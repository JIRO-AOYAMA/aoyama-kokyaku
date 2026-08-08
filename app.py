import base64
import calendar
import copy
import gzip
import hashlib
import hmac
import html
import json
import logging
import math
import mimetypes
import posixpath
import re
import secrets
import threading
import time
import urllib.parse
import unicodedata
import uuid
import zipfile
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from PIL import Image, ImageOps
except ImportError:
    Image = None
    ImageOps = None

try:
    from st_keyup import st_keyup
except ImportError:
    st_keyup = None

try:
    from streamlit_cookies_manager import EncryptedCookieManager
except ImportError:
    EncryptedCookieManager = None


# =========================
# 基本設定
# =========================
APP_TITLE = "取引先カルテ"

# Streamlitでは、st.set_page_config は他の st.* 呼び出しより先に実行する
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# 画像処理の診断記録は画面や保存データを変えず、Streamlitのサーバーログへだけ出す。
# アクセストークン・ダウンロードURL・画像本体は絶対に記録しない。
IMAGE_EVENT_LOGGER = logging.getLogger("aoyama.image")
if not any(getattr(handler, "_aoyama_image_handler", False) for handler in IMAGE_EVENT_LOGGER.handlers):
    image_log_handler = logging.StreamHandler()
    image_log_handler._aoyama_image_handler = True
    image_log_handler.setFormatter(logging.Formatter("%(message)s"))
    IMAGE_EVENT_LOGGER.addHandler(image_log_handler)
IMAGE_EVENT_LOGGER.setLevel(logging.INFO)
IMAGE_EVENT_LOGGER.propagate = False


def log_image_event(event, **details):
    """画像処理の成否・時間・容量だけを、安全なJSON形式でサーバーログへ残す。"""
    payload = {
        "event": str(event or "image_event")[:80],
        "recorded_at": datetime.now(timezone.utc).isoformat(),
    }
    for key, value in dict(details or {}).items():
        safe_key = re.sub(r"[^0-9A-Za-z_]", "_", str(key or "detail"))[:60]
        if isinstance(value, bool) or value is None:
            safe_value = value
        elif isinstance(value, (int, float)):
            safe_value = value
        else:
            safe_value = str(value or "")[:300]
        payload[safe_key] = safe_value
    try:
        IMAGE_EVENT_LOGGER.info(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    except Exception:
        # 記録機能の障害でアプリ本体を止めない。
        pass


EXCEL_FILE = "配車予定 次郎_修正版.xlsm"
SHEET_NAME = "Sheet1"

# secrets.toml に入れる設定
DROPBOX_APP_KEY = st.secrets.get("DROPBOX_APP_KEY", "")
DROPBOX_APP_SECRET = st.secrets.get("DROPBOX_APP_SECRET", "")
DROPBOX_REFRESH_TOKEN = st.secrets.get("DROPBOX_REFRESH_TOKEN", "")
# 移行期間用。Streamlit CloudではRefresh Token方式の3項目を使う。
DROPBOX_ACCESS_TOKEN = st.secrets.get("DROPBOX_ACCESS_TOKEN", "")
DROPBOX_DEFAULT_FILE_PATH = "/1共有　青山商店　本社/配車表-北海道-/配車予定 次郎.xlsm"
DROPBOX_FILE_PATH = st.secrets.get("DROPBOX_FILE_PATH", DROPBOX_DEFAULT_FILE_PATH)
DROPBOX_BACKUP_FOLDER = "/1共有　青山商店　本社/配車表-北海道-/Backups"
FULL_DATA_BACKUP_DROPBOX_FOLDER = st.secrets.get(
    "FULL_DATA_BACKUP_DROPBOX_FOLDER",
    str(DROPBOX_FILE_PATH).rsplit("/", 1)[0] + "/顧客カルテ全データ_Backups",
)
DROPBOX_FAST_CACHE_FILE = "/1共有　青山商店　本社/配車表-北海道-/顧客検索キャッシュ.json"
# Excelの列構成や読み込み処理を変更した時は、この番号を上げて古いJSONを無効化する。
DROPBOX_FAST_CACHE_VERSION = 6
DISPATCH_DROPBOX_DEFAULT_FILE_PATH = "/1共有　青山商店　本社/配車表-次郎-/配車表1.xlsm"
DISPATCH_DROPBOX_FILE_PATH = st.secrets.get(
    "DISPATCH_DROPBOX_FILE_PATH",
    DISPATCH_DROPBOX_DEFAULT_FILE_PATH,
)
DISPATCH_LOCAL_FILE = st.secrets.get(
    "DISPATCH_LOCAL_FILE",
    r"C:\Users\jiroa\Aoyama Dropbox\bulu jack\1共有　青山商店　本社\配車表-次郎-\配車表1.xlsm",
)
DISPATCH_MONTH_SHEETS = [f"{month}月" for month in range(1, 13)]
SOLUBLE_SHEET_NAME = "ソリュブル"
SOLUBLE_FILE_NAME = "aoベンチャーグレイン配車表.xlsx"
SOLUBLE_DROPBOX_DEFAULT_FILE_PATH = (
    str(DROPBOX_DEFAULT_FILE_PATH).rsplit("/", 1)[0] + "/" + SOLUBLE_FILE_NAME
)
SOLUBLE_DROPBOX_FILE_PATH = st.secrets.get(
    "SOLUBLE_DROPBOX_FILE_PATH",
    str(DROPBOX_FILE_PATH).rsplit("/", 1)[0] + "/" + SOLUBLE_FILE_NAME,
)
SOLUBLE_LOCAL_FILE = st.secrets.get(
    "SOLUBLE_LOCAL_FILE",
    r"C:\Users\jiroa\Aoyama Dropbox\bulu jack\1共有　青山商店　本社\配車表-北海道-\aoベンチャーグレイン配車表.xlsx",
)
SOLUBLE_BACKUP_FOLDER = str(SOLUBLE_DROPBOX_FILE_PATH).rsplit("/", 1)[0] + "/Backups"

# 仕入先・運送会社の基本情報を保存する別ブック。
# 配車予定 次郎.xlsm と同じDropboxフォルダに置く。
TRADE_PARTNER_FILE_NAME = "取引先カルテ.xlsx"
TRADE_PARTNER_DROPBOX_DEFAULT_FILE_PATH = (
    str(DROPBOX_DEFAULT_FILE_PATH).rsplit("/", 1)[0] + "/" + TRADE_PARTNER_FILE_NAME
)
TRADE_PARTNER_DROPBOX_FILE_PATH = st.secrets.get(
    "TRADE_PARTNER_DROPBOX_FILE_PATH",
    str(DROPBOX_FILE_PATH).rsplit("/", 1)[0] + "/" + TRADE_PARTNER_FILE_NAME,
)
TRADE_PARTNER_BACKUP_FOLDER = (
    str(TRADE_PARTNER_DROPBOX_FILE_PATH).rsplit("/", 1)[0] + "/取引先カルテ_Backups"
)
TRADE_PARTNER_MASTER_SHEET = "取引先マスター"
TRADE_PARTNER_CONTACT_SHEET = "担当者"
TRADE_PARTNER_PRODUCT_SHEET = "仕入商品"
TRADE_PARTNER_TRANSPORT_SHEET = "運送条件"
SOLUBLE_LOCATIONS = {
    "ノベルズ": {"usage": 3, "delivery": 4, "inventory": 5},
    "コスモアグリ": {"usage": 6, "delivery": 7, "inventory": 8},
}
# Excel内の識別名・既存ロジックは変えず、画面表示だけを分かりやすい名称にする。
SOLUBLE_LOCATION_DISPLAY_NAMES = {
    "ノベルズ": "ノベルズデイリー",
}
SOLUBLE_CUSTOMER_NAMES = ("三谷牧場", "熊林牧場")
SOLUBLE_CUSTOMER_COLUMNS = {
    "customer_name": 2,       # B列
    "delivery_date": 5,       # E列：配達日
    "delivery_quantity": 6,   # F列：配達数量
    "next_delivery": 7,       # G列：次回配達予定（数式・表示のみ）
    "usage": 8,               # H列：使用数量/日
}
DISPATCH_REQUIRED_COLUMNS = [
    "発注番号",
    "引取日",
    "引取先",
    "商品名",
    "数量",
    "運送会社",
    "納品先",
    "着日",
]
DELIVERY_SHEET_NAME = "次回配達日"
DELIVERY_HISTORY_SHEET_NAME = "配達履歴"
MANAGEMENT_SHEET_NAME = "管理"
SHEET1_CUSTOMER_COLUMN = 2   # B列：顧客名
SHEET1_HIRAGANA_COLUMN = 9   # I列：ひらがな
SHEET1_ADDRESS_COLUMN = 10   # J列：住所
SHEET1_MAP_COLUMN = 11       # K列：マップ位置
# v50までの共通パスワード設定はSecretsに残っていてもよいが、v51では認証に使用しない。
APP_PASSWORD = st.secrets.get("APP_PASSWORD", "")
LOGIN_TOKEN_SECRET = str(st.secrets.get("LOGIN_TOKEN_SECRET", "") or "").strip()
try:
    APP_AUTH_SETTINGS = st.secrets.get("app_auth", {})
except Exception:
    APP_AUTH_SETTINGS = {}
MICROSOFT_ALLOWED_SUB = str(
    APP_AUTH_SETTINGS.get("allowed_sub", "")
    or st.secrets.get("MICROSOFT_ALLOWED_SUB", "")
    or ""
).strip()
MICROSOFT_CLIENT_SECRET_EXPIRES_AT = str(
    APP_AUTH_SETTINGS.get("client_secret_expires_at", "") or ""
).strip()
MICROSOFT_CLIENT_SECRET_WARNING_DAYS = 60
MICROSOFT_CLIENT_SECRET_CRITICAL_DAYS = 30
MICROSOFT_CLIENT_SECRET_URGENT_DAYS = 7
MICROSOFT_ISSUER_PREFIX = "https://login.microsoftonline.com/"
MICROSOFT_AUTH_CLOCK_SKEW_SECONDS = 60
LOGIN_TOKEN_COOKIE_KEY = "signed_login_token"
LOGIN_TOKEN_COOKIE_PREFIX = "aoyama-kokyaku/login/"
LOGIN_TOKEN_TTL_SECONDS = 12 * 60 * 60
# 有効期限まで1時間を切った状態でアプリが開かれていれば、12時間へ更新する。
LOGIN_TOKEN_REFRESH_THRESHOLD_SECONDS = 60 * 60
LOGIN_TOKEN_VERSION = 1
LOGIN_TOKEN_AUDIENCE = "aoyama-kokyaku"
LOGIN_TOKEN_CLOCK_SKEW_SECONDS = 30

if EncryptedCookieManager is None:
    st.error(
        "ログイン用Cookie機能が読み込めません。requirements.txt に "
        "streamlit-cookies-manager-v2==0.3.1 を追加してください。"
    )
    st.stop()

LOGIN_COOKIES = EncryptedCookieManager(
    prefix=LOGIN_TOKEN_COOKIE_PREFIX,
    password=LOGIN_TOKEN_SECRET or "__LOGIN_TOKEN_SECRET_NOT_CONFIGURED__",
)
if not LOGIN_COOKIES.ready():
    st.stop()
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_SECRET_KEY = st.secrets.get("SUPABASE_SECRET_KEY", "")
SUPABASE_SERVICE_ROLE_KEY = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", "")
SUPABASE_NOTES_TABLE = st.secrets.get("SUPABASE_NOTES_TABLE", "notes")
SUPABASE_CUSTOMER_INFO_TABLE = st.secrets.get(
    "SUPABASE_CUSTOMER_INFO_TABLE",
    "customer_information",
)
SUPABASE_LOGIN_BROWSERS_TABLE = st.secrets.get(
    "SUPABASE_LOGIN_BROWSERS_TABLE",
    "app_login_browsers",
)
SUPABASE_LOGIN_EVENTS_TABLE = st.secrets.get(
    "SUPABASE_LOGIN_EVENTS_TABLE",
    "app_login_events",
)
LOGIN_BROWSER_COOKIE_KEY = "login_browser_id"
LOGIN_BROWSER_TOKEN_BYTES = 32
LOGIN_AUDIT_REQUEST_TIMEOUT = 15
LOGIN_HISTORY_PAGE_SIZE = 100
LINE_STATUS_NOTE_PREFIX = "line_status_"
LINE_STATUS_BODY = "__LINE_CONNECTED__"
HOME_TODO_CUSTOMER_NAME = "__HOME_TODO__"
HOME_TODO_ID_PREFIX = "home_todo_"
HOME_TODO_BODY_PREFIX = "__home_todo_v1__:"
HOME_TODO_LIST_HEIGHT = 420
VOICE_INPUT_HELP = "スマホではキーボードのマイクを押して音声入力できます。"
PAST_PRODUCT_NOTE_PREFIX = "__past_product_note__:"
ESTIMATE_PREFIX = "__estimate__:"
ESTIMATE_VERSION = 1
# 予想使用量は既存Excelを変更せず、customer_information内の内部レコードに保存する。
# 今日以降の在庫編集を基準として蓄積し、現在の「使用数量/日」は上書きしない。
INVENTORY_USAGE_SNAPSHOT_PREFIX = "__inventory_usage_snapshot__:"
INVENTORY_USAGE_SNAPSHOT_VERSION = 1
INVENTORY_USAGE_SNAPSHOT_CACHE_TTL_SECONDS = 5 * 60
CARRIER_FREIGHT_PREFIX = "__carrier_freight__:"
CARRIER_FREIGHT_VERSION = 1
CHANGE_HISTORY_CUSTOMER = "__CHANGE_HISTORY__"
CHANGE_HISTORY_VERSION = 1
CHANGE_HISTORY_PAGE_SIZE = 30
CHANGE_HISTORY_TARGETS = ("顧客", "仕入先", "運送会社")

REQUIRED_COLUMNS = [
    "ID",
    "顧客名",
    "地域",
    "商品名",
    "使用数量/日",
    "次回配達予定",
    "残数",
    "ひらがな",
]

REQUIRED_COLUMN_CANDIDATES = {
    "ID": ["ID", "id", "顧客ID", "顧客コード", "コード", "No", "NO", "番号"],
    "顧客名": ["顧客名", "牧場名", "取引先名", "得意先名", "お客様名", "名前", "名称"],
    "地域": ["地域", "地区", "エリア", "住所", "市町村"],
    "商品名": ["商品名", "商品", "品名", "製品名"],
    "使用数量/日": ["使用数量/日", "使用数量", "使用量/日", "一日使用量", "数量/日", "日量"],
    "次回配達予定": ["次回配達予定", "配達予定日", "配送予定日", "配達日", "配送日", "納品日", "予定日", "日付"],
    "残数": ["残数", "残量", "残", "在庫", "残り"],
    "ひらがな": ["ひらがな", "ふりがな", "フリガナ", "かな", "カナ", "よみがな", "読み仮名"],
}

ADDRESS_COLUMN_CANDIDATES = [
    "住所",
    "所在地",
    "配達先住所",
    "配送先住所",
    "納品先住所",
    "顧客住所",
    "牧場住所",
]

MAP_LOCATION_COLUMN_CANDIDATES = [
    "マップ位置",
    "地図位置",
    "Googleマップ",
    "GoogleマップURL",
    "Google Maps",
    "Google Map",
    "マップURL",
    "地図URL",
    "位置情報",
    "緯度経度",
    "緯度・経度",
    "座標",
]


# ホテル・宿泊先情報は、既存のcustomer_informationテーブル内に
# 通常の顧客情報と衝突しない内部レコードとして保存する。
# 新しいSupabaseテーブルやSecretsは不要。
HOTEL_INFORMATION_STORAGE_CUSTOMER = "__HOTEL_STAY_INFORMATION__"
HOTEL_INFORMATION_FIELD_PREFIX = "__hotel_stay_information__:"
HOTEL_INFORMATION_VERSION = 1
HOTEL_INFORMATION_REQUIRED_FIELDS = ("ホテル名",)
HOTEL_INFORMATION_FIXED_FIELDS = ("ホテル名", "地域", "住所", "Googleマップ")


# =========================
# WATER it接続（読み取り専用）
# =========================
# スマホでWATER itから手動ダウンロードしたCSVを選び、読み取り専用で表示する。
# 選択したCSVはStreamlitのセッション内だけに保持し、Excel・WATER it・Dropboxへは書き込まない。
# 未選択時は、従来どおり同じフォルダのdata.csvを参考表示できる。
WATER_IT_CSV_PATH = st.secrets.get("WATER_IT_CSV_PATH", "data.csv")
WATER_IT_CSV_URL = st.secrets.get("WATER_IT_CSV_URL", "")
WATER_IT_REQUEST_TIMEOUT = 20
WATER_IT_LOGIN_URL = "https://www.dms2.waterit.optex.net/WIA0101/Index01"
WATER_IT_UPLOAD_BYTES_KEY = "water_it_uploaded_csv_bytes"
WATER_IT_UPLOAD_NAME_KEY = "water_it_uploaded_csv_name"
WATER_IT_UPLOAD_HASH_KEY = "water_it_uploaded_csv_hash"
WATER_IT_UPLOAD_PERSISTED_KEY = "water_it_uploaded_csv_persisted"
# 既存のcustomer_informationテーブル内に、通常の顧客情報と衝突しない内部レコードとして保存する。
# 新しいSupabaseテーブルやSecretsは不要。Excel・WATER it・Dropboxには書き込まない。
WATER_IT_STORAGE_CUSTOMER = "__WATER_IT_STORAGE__"
WATER_IT_STORAGE_FIELD = "__water_it_csv_snapshot__"
WATER_IT_STORAGE_ID = str(
    uuid.uuid5(uuid.NAMESPACE_URL, "aoyama-water-it-csv-snapshot-v1")
)
WATER_IT_STORAGE_VERSION = 1
# 顧客詳細では、WATER it対象顧客の小さな照合索引だけを5分間再利用する。
# 元CSV・Supabase保存内容・顧客名の照合ルール自体は変更しない。
WATER_IT_CUSTOMER_INDEX_TTL_SECONDS = 5 * 60
WATER_IT_CUSTOMER_INDEX_SESSION_KEY = "water_it_customer_key_index"
WATER_IT_CUSTOMER_INDEX_HASH_SESSION_KEY = "water_it_customer_key_index_hash"
WATER_IT_REQUIRED_COLUMNS = [
    "測定日時",
    "測定項目",
    "測定値",
    "単位",
    "エリア",
    "ポイント",
]
WATER_IT_ALERT_COLUMNS = [
    "HOLD中",
    "メンテナンス時期",
    "校正時期",
    "消耗品交換時期",
    "オーバーホール時期",
    "ローバッテリ",
    "センサまたは変換器異常",
    "通信不良または断線",
    "状態",
]
# WATER it側の元名称は変更せず、画面表示と顧客照合だけを統一する。
WATER_IT_POINT_DISPLAY_NAMES = {
    "ノベルズデイリーファーム": "ノベルズデイリー",
}

# ソリュブル在庫とWATER itの対応。Excel内の識別名や既存列構成は変更しない。
SOLUBLE_WATER_IT_POINT_NAMES = {
    "ノベルズ": "ノベルズデイリー",
    "コスモアグリ": "コスモアグリ",
}
# 実績平均はアプリ上の参考表示だけに使い、Excelの使用量/日へは反映しない。
SOLUBLE_WATER_IT_USAGE_WINDOWS = (3, 7, 20, 30)


# =========================
# OneDrive接続（写真・資料）
# =========================
ONEDRIVE_AUTHORITY = "https://login.microsoftonline.com/consumers"
ONEDRIVE_AUTHORIZE_URL = ONEDRIVE_AUTHORITY + "/oauth2/v2.0/authorize"
ONEDRIVE_TOKEN_URL = ONEDRIVE_AUTHORITY + "/oauth2/v2.0/token"
ONEDRIVE_GRAPH_BASE = "https://graph.microsoft.com/v1.0"
ONEDRIVE_SCOPES = "openid profile offline_access User.Read Files.ReadWrite"
ONEDRIVE_REQUEST_TIMEOUT = 90
ONEDRIVE_AUTH_FLOW_TTL_SECONDS = 15 * 60
ONEDRIVE_PRODUCTION_REDIRECT_URI = "https://aoyama-kokyaku.streamlit.app"
ONEDRIVE_TEST_REDIRECT_HOST = "aoyama-onedrive-test.streamlit.app"
ONEDRIVE_ROOT_FOLDER = "取引先カルテ"
ONEDRIVE_CUSTOMER_FOLDER = "顧客"
ONEDRIVE_SUPPLIER_FOLDER = "仕入先"
ONEDRIVE_CARRIER_FOLDER = "運送会社"
ONEDRIVE_HOTEL_FOLDER = "ホテル"
ONEDRIVE_ATTACHMENT_PREFIX = "__onedrive_attachment__:"
ONEDRIVE_ATTACHMENT_VERSION = 2
ONEDRIVE_DISPLAY_IMAGE_MAX_EDGE = 1600
ONEDRIVE_DISPLAY_IMAGE_QUALITY = 82
ONEDRIVE_DISPLAY_IMAGE_METHOD = 4
ONEDRIVE_GALLERY_URL_WORKERS = 4
ONEDRIVE_FIXED_TAGS = ("設備", "名刺", "納品場所", "商品", "トラブル")
ONEDRIVE_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
ONEDRIVE_PDF_EXTENSIONS = {".pdf"}
ONEDRIVE_PAGE_SIZE = 12
ONEDRIVE_RECENT_CAMERA_DAYS = 14
ONEDRIVE_RECENT_CAMERA_MAX_ITEMS = 100
ONEDRIVE_RECENT_CAMERA_PAGE_SIZE = 12
# Secretsのrefresh_token由来かを、サーバー内の一時トークンだけで識別する。
# Microsoftのトークン本体やこの印はSupabase・OneDriveへ保存しない。
ONEDRIVE_CONFIGURED_TOKEN_SOURCE_KEY = "_aoyama_configured_refresh_source"


def is_mobile_browser():
    """現在のアクセスがスマホ・タブレット系ブラウザーかを判定する。"""
    try:
        user_agent = str(st.context.headers.get("User-Agent", "")).lower()
    except Exception:
        user_agent = ""
    return any(
        marker in user_agent
        for marker in ("android", "iphone", "ipad", "ipod", "mobile")
    )


def enable_mobile_camera_capture(uploader_label):
    """指定した画像アップローダーへ、スマホ背面カメラの起動指定を付ける。"""
    label_json = json.dumps(str(uploader_label), ensure_ascii=False)
    script = r'''<script>
(() => {
  const targetLabel = __TARGET_LABEL__;
  let stopped = false;
  let attempts = 0;

  const normalize = (value) => String(value || '').replace(/\s+/g, ' ').trim();

  const applyCapture = () => {
    if (stopped) return true;
    let parentDocument;
    try {
      parentDocument = window.parent.document;
    } catch (_) {
      return false;
    }
    if (!parentDocument) return false;

    const uploaders = Array.from(
      parentDocument.querySelectorAll('[data-testid="stFileUploader"]')
    );
    const target = uploaders.find((node) => {
      const label = node.querySelector('label');
      const text = normalize(label ? label.textContent : node.textContent);
      return text.includes(targetLabel);
    });
    if (!target) return false;

    const input = target.querySelector('input[type="file"]');
    if (!input) return false;

    input.setAttribute('accept', 'image/*');
    input.setAttribute('capture', 'environment');
    input.dataset.aoyamaCameraCapture = 'environment';
    return true;
  };

  const retry = () => {
    if (applyCapture() || attempts >= 60) return;
    attempts += 1;
    window.setTimeout(retry, 100);
  };

  retry();

  let observer;
  try {
    observer = new MutationObserver(() => applyCapture());
    observer.observe(window.parent.document.body, {childList: true, subtree: true});
    window.setTimeout(() => {
      stopped = true;
      observer.disconnect();
    }, 12000);
  } catch (_) {}
})();
</script>'''.replace('__TARGET_LABEL__', label_json)
    components.html(script, height=0, scrolling=False)


def enable_mobile_bulk_image_picker(uploader_label):
    """スマホ用の独立した複数画像ピッカーからStreamlitへ選択結果を渡す。"""
    label_json = json.dumps(str(uploader_label), ensure_ascii=False)
    component_html = r"""<style>
      html, body {
        margin: 0;
        padding: 0;
        background: transparent;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      }
      .aoyama-bulk-picker {
        display: flex;
        flex-direction: column;
        gap: 0.28rem;
      }
      .aoyama-bulk-picker-button {
        box-sizing: border-box;
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
        min-height: 2.75rem;
        padding: 0.65rem 0.9rem;
        border: 1px solid rgba(49, 51, 63, 0.28);
        border-radius: 0.5rem;
        background: white;
        color: rgb(49, 51, 63);
        font-size: 0.95rem;
        font-weight: 600;
        cursor: pointer;
        user-select: none;
      }
      .aoyama-bulk-picker-button:active {
        background: rgba(49, 51, 63, 0.06);
      }
      .aoyama-bulk-picker-button input {
        position: absolute;
        width: 1px;
        height: 1px;
        opacity: 0;
        pointer-events: none;
      }
      .aoyama-bulk-picker-status {
        min-height: 1.15rem;
        color: rgba(49, 51, 63, 0.68);
        font-size: 0.78rem;
        line-height: 1.35;
      }
    </style>
    <div class="aoyama-bulk-picker">
      <label class="aoyama-bulk-picker-button">
        🖼 写真をまとめて選ぶ
        <input id="aoyama-bulk-image-input" type="file" accept="image/*" multiple>
      </label>
      <div id="aoyama-bulk-image-status" class="aoyama-bulk-picker-status">
        3枚以上も一度に選択できます
      </div>
    </div>
    <script>
    (() => {
      const targetLabel = __TARGET_LABEL__;
      const picker = document.getElementById('aoyama-bulk-image-input');
      const status = document.getElementById('aoyama-bulk-image-status');
      if (!picker || !status) return;

      const normalize = (value) => String(value || '').replace(/\s+/g, ' ').trim();

      const findTargetUploader = () => {
        let parentDocument;
        try {
          parentDocument = window.parent.document;
        } catch (_) {
          return null;
        }
        const uploaders = Array.from(
          parentDocument.querySelectorAll('[data-testid="stFileUploader"]')
        );
        return uploaders.find((node) => {
          const label = node.querySelector('label');
          const text = normalize(label ? label.textContent : node.textContent);
          const visible = node.offsetParent !== null;
          return visible && text.includes(targetLabel);
        }) || uploaders.find((node) => {
          const label = node.querySelector('label');
          const text = normalize(label ? label.textContent : node.textContent);
          return text.includes(targetLabel);
        }) || null;
      };

      const prepareNativeUploader = () => {
        const target = findTargetUploader();
        if (!target) return null;
        const nativeInput = target.querySelector('input[type="file"]');
        if (!nativeInput) return null;
        nativeInput.setAttribute('multiple', '');

        const dropzone = target.querySelector('[data-testid="stFileUploaderDropzone"]');
        if (dropzone) {
          dropzone.style.display = 'none';
        }
        return nativeInput;
      };

      picker.addEventListener('change', () => {
        const selectedFiles = Array.from(picker.files || []);
        if (!selectedFiles.length) {
          status.textContent = '写真が選択されていません';
          return;
        }

        const nativeInput = prepareNativeUploader();
        if (!nativeInput) {
          status.textContent = '写真選択欄を確認できませんでした。画面を開き直してください。';
          return;
        }

        try {
          const parentWindow = window.parent;
          const transfer = new parentWindow.DataTransfer();
          selectedFiles.forEach((file) => transfer.items.add(file));

          const filesSetter = Object.getOwnPropertyDescriptor(
            parentWindow.HTMLInputElement.prototype,
            'files'
          );
          if (filesSetter && typeof filesSetter.set === 'function') {
            filesSetter.set.call(nativeInput, transfer.files);
          } else {
            nativeInput.files = transfer.files;
          }

          status.textContent = `${selectedFiles.length}枚をアプリへ渡しています…`;
          nativeInput.dispatchEvent(new parentWindow.Event('input', {bubbles: true}));
          nativeInput.dispatchEvent(new parentWindow.Event('change', {bubbles: true}));
        } catch (_) {
          status.textContent = '写真を渡せませんでした。もう一度選択してください。';
        }
      });

      let attempts = 0;
      const prepareWithRetry = () => {
        if (prepareNativeUploader() || attempts >= 80) return;
        attempts += 1;
        window.setTimeout(prepareWithRetry, 100);
      };
      prepareWithRetry();

      try {
        const observer = new MutationObserver(() => prepareNativeUploader());
        observer.observe(window.parent.document.body, {childList: true, subtree: true});
        window.setTimeout(() => observer.disconnect(), 15000);
      } catch (_) {}
    })();
    </script>""".replace('__TARGET_LABEL__', label_json)
    components.html(component_html, height=76, scrolling=False)


def read_onedrive_settings():
    """Streamlit SecretsからOneDrive接続設定を読む。"""
    try:
        settings = st.secrets["onedrive"]
        client_id = str(settings.get("client_id", "")).strip()
        client_secret = str(settings.get("client_secret", "")).strip()
        redirect_uri = str(settings.get("redirect_uri", "")).strip()
    except Exception:
        client_id = ""
        client_secret = ""
        redirect_uri = ""

    # テストアプリのURLが残っていても、本番カルテへ戻るよう安全側で補正する。
    if not redirect_uri or ONEDRIVE_TEST_REDIRECT_HOST in redirect_uri:
        redirect_uri = ONEDRIVE_PRODUCTION_REDIRECT_URI

    missing = []
    if not client_id:
        missing.append("client_id")
    if not client_secret or client_secret == "PASTE_SECRET_VALUE_HERE":
        missing.append("client_secret")
    if missing:
        raise RuntimeError(
            "StreamlitのSecretsにある[onedrive]へ"
            + "、".join(missing)
            + "を設定してください。"
        )
    return client_id, client_secret, redirect_uri


def read_onedrive_configured_refresh_token():
    """通常利用時の自動接続に使う更新トークンをSecretsから読む。"""
    try:
        value = str(st.secrets["onedrive"].get("refresh_token", "")).strip()
    except Exception:
        value = ""
    if value in {"PASTE_REFRESH_TOKEN_HERE", "Microsoftの更新トークン"}:
        return ""
    return value


@st.cache_resource(show_spinner=False)
def get_onedrive_pending_auth_store():
    """外部ログイン中だけ必要な認証情報を一時保持する。"""
    return {"lock": threading.RLock(), "flows": {}}


@st.cache_resource(show_spinner=False)
def get_onedrive_shared_token_store():
    """全利用者が同じOneDriveを使えるよう、サーバー内で認証結果を共有する。"""
    return {"lock": threading.RLock(), "token": None}


def get_onedrive_shared_token_result():
    store = get_onedrive_shared_token_store()
    with store["lock"]:
        token = store.get("token")
        return dict(token) if isinstance(token, dict) else None


def clear_onedrive_shared_token_result():
    store = get_onedrive_shared_token_store()
    with store["lock"]:
        store["token"] = None


def cleanup_onedrive_pending_auth_flows(store):
    now = time.time()
    expired = [
        state
        for state, entry in store["flows"].items()
        if now - float(entry.get("created_at", 0)) > ONEDRIVE_AUTH_FLOW_TTL_SECONDS
    ]
    for state in expired:
        store["flows"].pop(state, None)


def save_onedrive_pending_auth_flow(state, payload):
    store = get_onedrive_pending_auth_store()
    with store["lock"]:
        cleanup_onedrive_pending_auth_flows(store)
        store["flows"][state] = {
            "created_at": time.time(),
            **payload,
        }


def pop_onedrive_pending_auth_flow(state):
    store = get_onedrive_pending_auth_store()
    with store["lock"]:
        cleanup_onedrive_pending_auth_flows(store)
        return store["flows"].pop(state, None)


def get_raw_query_params():
    try:
        return {key: str(value) for key, value in st.query_params.items()}
    except Exception:
        legacy = st.experimental_get_query_params()
        return {
            key: str(value[0] if isinstance(value, list) and value else value)
            for key, value in legacy.items()
        }



def login_token_base64_encode(value):
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def login_token_base64_decode(value):
    text = str(value or "").strip()
    if not text:
        raise ValueError("empty base64 value")
    padding = "=" * (-len(text) % 4)
    return base64.urlsafe_b64decode((text + padding).encode("ascii"))


@st.cache_resource(show_spinner=False)
def get_revoked_login_token_store():
    """ログアウト済みトークンを有効期限までサーバー内で無効化する。"""
    return {"lock": threading.RLock(), "items": {}}


def cleanup_revoked_login_tokens(store, now=None):
    current_time = float(time.time() if now is None else now)
    expired = [
        token_id
        for token_id, expires_at in store["items"].items()
        if float(expires_at or 0) <= current_time
    ]
    for token_id in expired:
        store["items"].pop(token_id, None)


def is_login_token_revoked(token_id, now=None):
    token_id = str(token_id or "").strip()
    if not token_id:
        return True
    store = get_revoked_login_token_store()
    with store["lock"]:
        cleanup_revoked_login_tokens(store, now=now)
        return token_id in store["items"]


def create_login_token(now=None, token_id=None):
    """別Secretで署名した、発行時点から12時間有効なトークンを作る。"""
    if not LOGIN_TOKEN_SECRET:
        raise RuntimeError("LOGIN_TOKEN_SECRET が設定されていません。")
    issued_at = int(time.time() if now is None else now)
    session_token_id = str(token_id or "").strip() or uuid.uuid4().hex
    payload = {
        "aud": LOGIN_TOKEN_AUDIENCE,
        "exp": issued_at + LOGIN_TOKEN_TTL_SECONDS,
        "iat": issued_at,
        "jti": session_token_id,
        "ver": LOGIN_TOKEN_VERSION,
    }
    payload_bytes = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    payload_part = login_token_base64_encode(payload_bytes)
    signature = hmac.new(
        LOGIN_TOKEN_SECRET.encode("utf-8"),
        payload_part.encode("ascii"),
        hashlib.sha256,
    ).digest()
    return payload_part + "." + login_token_base64_encode(signature)


def validate_login_token(token, now=None, check_revocation=True):
    """署名・用途・発行時刻・12時間期限を確認し、有効時だけpayloadを返す。"""
    if not LOGIN_TOKEN_SECRET:
        return None
    token_text = str(token or "").strip()
    if token_text.count(".") != 1:
        return None
    payload_part, signature_part = token_text.split(".", 1)
    try:
        supplied_signature = login_token_base64_decode(signature_part)
        expected_signature = hmac.new(
            LOGIN_TOKEN_SECRET.encode("utf-8"),
            payload_part.encode("ascii"),
            hashlib.sha256,
        ).digest()
        if not hmac.compare_digest(supplied_signature, expected_signature):
            return None
        payload = json.loads(login_token_base64_decode(payload_part).decode("utf-8"))
        issued_at = int(payload.get("iat"))
        expires_at = int(payload.get("exp"))
        token_id = str(payload.get("jti") or "").strip()
        current_time = int(time.time() if now is None else now)
    except (ValueError, TypeError, UnicodeDecodeError, json.JSONDecodeError):
        return None

    if payload.get("aud") != LOGIN_TOKEN_AUDIENCE:
        return None
    if int(payload.get("ver") or 0) != LOGIN_TOKEN_VERSION:
        return None
    if not token_id:
        return None
    if issued_at > current_time + LOGIN_TOKEN_CLOCK_SKEW_SECONDS:
        return None
    if expires_at <= issued_at:
        return None
    if expires_at - issued_at != LOGIN_TOKEN_TTL_SECONDS:
        return None
    if current_time >= expires_at:
        return None
    if check_revocation and is_login_token_revoked(token_id, now=current_time):
        return None
    return payload


def refresh_login_token_if_needed(token, now=None):
    """有効期限が近いトークンを、アプリ利用中に12時間へ安全に更新する。"""
    payload = validate_login_token(token, now=now)
    if not payload:
        return ""

    current_time = int(time.time() if now is None else now)
    expires_at = int(payload.get("exp") or 0)
    if expires_at - current_time > LOGIN_TOKEN_REFRESH_THRESHOLD_SECONDS:
        return str(token or "").strip()

    # 同じログインセッションとして扱うため、監査用のjtiは引き継ぐ。
    refreshed_token = create_login_token(
        now=current_time,
        token_id=payload.get("jti"),
    )
    save_login_token_cookie(refreshed_token)
    st.session_state["login_token"] = refreshed_token
    return refreshed_token


def revoke_login_token(token):
    """ログアウトしたトークンを、残りの有効期限中は再利用できないようにする。"""
    payload = validate_login_token(token, check_revocation=False)
    if not payload:
        return
    token_id = str(payload.get("jti") or "").strip()
    expires_at = int(payload.get("exp") or 0)
    if not token_id or expires_at <= int(time.time()):
        return
    store = get_revoked_login_token_store()
    with store["lock"]:
        cleanup_revoked_login_tokens(store)
        store["items"][token_id] = expires_at


def get_login_token_from_cookie():
    """暗号化Cookieから署名付きログイントークンを読む。"""
    try:
        return str(LOGIN_COOKIES.get(LOGIN_TOKEN_COOKIE_KEY, "") or "").strip()
    except Exception:
        return ""


def save_login_token_cookie(token):
    """署名付きログイントークンをブラウザの永続Cookieへ保存する。"""
    token_text = str(token or "").strip()
    if not validate_login_token(token_text):
        raise RuntimeError("有効なログイントークンを保存できません。")
    LOGIN_COOKIES[LOGIN_TOKEN_COOKIE_KEY] = token_text
    LOGIN_COOKIES.save()


def clear_login_token_cookie():
    """ログインCookieを即時削除する。"""
    try:
        del LOGIN_COOKIES[LOGIN_TOKEN_COOKIE_KEY]
    except KeyError:
        pass
    except Exception:
        try:
            LOGIN_COOKIES[LOGIN_TOKEN_COOKIE_KEY] = ""
        except Exception:
            return
    try:
        LOGIN_COOKIES.save()
    except Exception:
        pass


def remove_obsolete_login_query_params():
    """旧 logged_in=1 と v49 の auth パラメータを認証に使わずURLから除く。"""
    try:
        for key in ("logged_in", "auth"):
            if key in st.query_params:
                del st.query_params[key]
    except Exception:
        pass


def get_active_login_token():
    session_token = str(st.session_state.get("login_token", "") or "").strip()
    if validate_login_token(session_token):
        return session_token

    cookie_token = get_login_token_from_cookie()
    if validate_login_token(cookie_token):
        st.session_state["login_token"] = cookie_token
        return cookie_token

    st.session_state.pop("login_token", None)
    return ""


def set_query_params_safely(params):
    clean_params = {
        str(key): str(value)
        for key, value in dict(params or {}).items()
        if value is not None and str(value) != ""
    }
    try:
        st.query_params.clear()
        for key, value in clean_params.items():
            st.query_params[key] = value
    except Exception:
        st.experimental_set_query_params(**clean_params)


def get_microsoft_user_claims():
    """Streamlitが検証したMicrosoft OIDCのユーザー情報を辞書で返す。"""
    try:
        if hasattr(st.user, "to_dict"):
            claims = st.user.to_dict()
        else:
            claims = dict(st.user)
    except Exception:
        return {}
    return dict(claims or {})


def is_private_test_login_bypass_enabled():
    """Privateなテストアプリで明示的に許可した時だけMicrosoftログインを省略する。"""
    try:
        value = st.secrets.get("TEST_PRIVATE_BYPASS_LOGIN", False)
    except Exception:
        return False
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def get_configured_auth_client_id():
    """[auth] に設定されたログイン専用MicrosoftアプリのClient IDを返す。"""
    try:
        auth_settings = st.secrets.get("auth", {})
        return str(auth_settings.get("client_id", "") or "").strip()
    except Exception:
        return ""


def get_microsoft_client_secret_expiry_info(today=None):
    """ログイン専用MicrosoftアプリのSecret期限と残り日数を返す。"""
    configured_value = str(MICROSOFT_CLIENT_SECRET_EXPIRES_AT or "").strip()
    if not configured_value:
        return {"configured": False, "valid": True, "expires_at": None, "days_remaining": None}

    try:
        expires_at = datetime.strptime(configured_value, "%Y-%m-%d").date()
    except ValueError:
        return {
            "configured": True,
            "valid": False,
            "expires_at": None,
            "days_remaining": None,
        }

    current_date = today or datetime.now(timezone(timedelta(hours=9))).date()
    return {
        "configured": True,
        "valid": True,
        "expires_at": expires_at,
        "days_remaining": (expires_at - current_date).days,
    }


def show_microsoft_client_secret_expiry_notice():
    """期限が近い場合だけ、ログイン後の全画面上部へ更新警告を表示する。"""
    info = get_microsoft_client_secret_expiry_info()
    if not info["configured"]:
        return

    if not info["valid"]:
        st.error(
            "Microsoftログイン用クライアントシークレットの期限日を確認できません。"
            " Secretsの [app_auth] client_secret_expires_at を YYYY-MM-DD 形式で確認してください。"
        )
        return

    days_remaining = int(info["days_remaining"])
    expires_text = info["expires_at"].strftime("%Y年%m月%d日")

    if days_remaining > MICROSOFT_CLIENT_SECRET_WARNING_DAYS:
        return
    if days_remaining < 0:
        st.error(
            f"Microsoftログイン用クライアントシークレットは {expires_text} に期限切れになっています。"
            " 新しいシークレットを作成し、Streamlit Secretsの [auth] client_secret を更新してください。"
        )
        return
    if days_remaining == 0:
        st.error(
            f"Microsoftログイン用クライアントシークレットは本日（{expires_text}）が期限です。"
            " 本日中に更新してください。"
        )
        return
    if days_remaining <= MICROSOFT_CLIENT_SECRET_URGENT_DAYS:
        st.error(
            f"Microsoftログイン用クライアントシークレットの期限まで残り {days_remaining} 日です"
            f"（{expires_text}）。早急に更新してください。"
        )
        return
    if days_remaining <= MICROSOFT_CLIENT_SECRET_CRITICAL_DAYS:
        st.warning(
            f"Microsoftログイン用クライアントシークレットの期限まで残り {days_remaining} 日です"
            f"（{expires_text}）。更新準備を進めてください。"
        )
        return

    st.warning(
        f"Microsoftログイン用クライアントシークレットの期限まで残り {days_remaining} 日です"
        f"（{expires_text}）。期限前に更新してください。"
    )


def validate_microsoft_identity(claims, require_allowed_sub=True):
    """Microsoftログイン情報を確認し、問題がある場合だけ理由コードを返す。"""
    data = dict(claims or {})
    if not bool(data.get("is_logged_in")):
        return "not_logged_in"

    subject = str(data.get("sub") or "").strip()
    issuer = str(data.get("iss") or "").strip()
    audience = str(data.get("aud") or "").strip()
    configured_client_id = get_configured_auth_client_id()

    if not subject:
        return "missing_sub"
    if not issuer.startswith(MICROSOFT_ISSUER_PREFIX):
        return "invalid_issuer"
    if not audience or not configured_client_id:
        return "missing_audience"
    if not hmac.compare_digest(audience, configured_client_id):
        return "invalid_audience"
    if require_allowed_sub and not MICROSOFT_ALLOWED_SUB:
        return "allowed_sub_not_configured"
    if MICROSOFT_ALLOWED_SUB and not hmac.compare_digest(subject, MICROSOFT_ALLOWED_SUB):
        return "account_not_allowed"
    return ""


def microsoft_auth_seconds_remaining(claims, now=None):
    """Microsoftログイン後、初回トークンを発行できる12時間の残り秒数を返す。"""
    try:
        issued_at = int(dict(claims or {}).get("iat"))
    except (TypeError, ValueError):
        return -1
    current_time = int(time.time() if now is None else now)
    if issued_at > current_time + MICROSOFT_AUTH_CLOCK_SKEW_SECONDS:
        return -1
    return issued_at + LOGIN_TOKEN_TTL_SECONDS - current_time


def is_microsoft_auth_current(claims, now=None):
    return microsoft_auth_seconds_remaining(claims, now=now) > 0


def clear_application_login_state(revoke_current=True):
    """独自の12時間トークンと画面内ログイン状態をまとめて消す。"""
    if revoke_current:
        revoke_login_token(get_active_login_token())
    clear_login_token_cookie()
    st.session_state.authenticated = False
    st.session_state.pop("login_token", None)


@st.fragment(run_every="10s")
def enforce_login_expiry():
    """Microsoft本人確認と独自トークンを確認し、利用中は期限を12時間へ更新する。"""
    if is_private_test_login_bypass_enabled():
        return
    claims = get_microsoft_user_claims()
    token = str(st.session_state.get("login_token", "") or "").strip()
    identity_error = validate_microsoft_identity(claims, require_allowed_sub=True)
    token_payload = validate_login_token(token)
    if not identity_error and token_payload:
        # 更新に失敗しても、現在のトークンが有効な間は画面を中断しない。
        try:
            refresh_login_token_if_needed(token)
        except Exception:
            pass
        return

    clear_application_login_state(revoke_current=True)
    st.session_state["microsoft_force_logout"] = True
    set_query_params_safely({"page": "home", "expired": "1"})
    st.rerun()



def get_login_audit_admin_key():
    """ログイン履歴専用に、ブラウザへ公開しないSupabase管理キーだけを返す。"""
    return str(SUPABASE_SECRET_KEY or SUPABASE_SERVICE_ROLE_KEY or "").strip()


def has_login_audit_config():
    return bool(str(SUPABASE_URL or "").strip() and get_login_audit_admin_key())


def validate_login_audit_table_name(value, default_name):
    table_name = str(value or default_name).strip()
    if not table_name.replace("_", "").isalnum():
        raise RuntimeError("ログイン履歴用Supabaseテーブル名が正しくありません。")
    return table_name


def get_login_audit_table_url(table_name):
    base_url = str(SUPABASE_URL or "").strip().rstrip("/")
    safe_table = urllib.parse.quote(str(table_name), safe="")
    return f"{base_url}/rest/v1/{safe_table}"


def get_login_browsers_table_name():
    return validate_login_audit_table_name(
        SUPABASE_LOGIN_BROWSERS_TABLE,
        "app_login_browsers",
    )


def get_login_events_table_name():
    return validate_login_audit_table_name(
        SUPABASE_LOGIN_EVENTS_TABLE,
        "app_login_events",
    )


def get_login_browsers_url():
    return get_login_audit_table_url(get_login_browsers_table_name())


def get_login_events_url():
    return get_login_audit_table_url(get_login_events_table_name())


def get_login_audit_headers(prefer=None):
    key = get_login_audit_admin_key()
    headers = {
        "apikey": key,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    if key and not key.startswith(("sb_secret_", "sb_publishable_")):
        headers["Authorization"] = f"Bearer {key}"
    if prefer:
        headers["Prefer"] = prefer
    return headers


def get_request_user_agent():
    try:
        return str(st.context.headers.get("User-Agent", "") or "").strip()
    except Exception:
        return ""


def get_request_ip_address():
    """IPは履歴表示の参考情報としてだけ保存し、本人判定には使わない。"""
    try:
        value = str(st.context.ip_address or "").strip()
        if value:
            return value[:200]
    except Exception:
        pass
    try:
        forwarded = str(st.context.headers.get("X-Forwarded-For", "") or "").strip()
        if forwarded:
            return forwarded.split(",", 1)[0].strip()[:200]
    except Exception:
        pass
    return ""


def parse_login_client_info(user_agent):
    ua = str(user_agent or "")
    lower = ua.lower()

    if "samsungbrowser/" in lower:
        browser_name = "Samsung Internet"
    elif "edg/" in lower or "edgios/" in lower or "edga/" in lower:
        browser_name = "Microsoft Edge"
    elif "crios/" in lower or "chrome/" in lower:
        browser_name = "Google Chrome"
    elif "firefox/" in lower or "fxios/" in lower:
        browser_name = "Firefox"
    elif "safari/" in lower and "chrome/" not in lower and "crios/" not in lower:
        browser_name = "Safari"
    else:
        browser_name = "不明なブラウザ"

    if "android" in lower:
        os_name = "Android"
    elif "iphone" in lower or "ipod" in lower:
        os_name = "iOS"
    elif "ipad" in lower:
        os_name = "iPadOS"
    elif "windows nt" in lower:
        os_name = "Windows"
    elif "mac os x" in lower or "macintosh" in lower:
        os_name = "macOS"
    elif "linux" in lower:
        os_name = "Linux"
    else:
        os_name = "不明なOS"

    if "ipad" in lower or "tablet" in lower:
        device_type = "タブレット"
    elif any(marker in lower for marker in ("android", "iphone", "ipod", "mobile")):
        device_type = "スマホ"
    else:
        device_type = "パソコン"

    return {
        "browser_name": browser_name,
        "os_name": os_name,
        "device_type": device_type,
    }


def get_login_browser_token(create_if_missing=True):
    try:
        token = str(LOGIN_COOKIES.get(LOGIN_BROWSER_COOKIE_KEY, "") or "").strip()
    except Exception:
        token = ""

    if token and re.fullmatch(r"[A-Za-z0-9_-]{40,200}", token):
        return token

    if not create_if_missing:
        return ""

    token = secrets.token_urlsafe(LOGIN_BROWSER_TOKEN_BYTES)
    LOGIN_COOKIES[LOGIN_BROWSER_COOKIE_KEY] = token
    LOGIN_COOKIES.save()
    return token


def hash_login_audit_value(value, purpose):
    secret_key = str(LOGIN_TOKEN_SECRET or "").encode("utf-8")
    if not secret_key:
        raise RuntimeError("LOGIN_TOKEN_SECRET が設定されていません。")
    message = f"{purpose}:{str(value or '')}".encode("utf-8")
    return hmac.new(secret_key, message, hashlib.sha256).hexdigest()


def get_current_browser_token_hash(create_if_missing=True):
    token = get_login_browser_token(create_if_missing=create_if_missing)
    if not token:
        return ""
    return hash_login_audit_value(token, "browser")


def get_microsoft_subject_hash(claims):
    subject = str(dict(claims or {}).get("sub") or "").strip()
    if not subject:
        return ""
    return hash_login_audit_value(subject, "microsoft-sub")


def get_microsoft_account_label(claims):
    data = dict(claims or {})
    return str(data.get("name") or "Microsoftアカウント").strip()[:200]


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()


def fetch_login_browser_by_hash(browser_token_hash):
    if not has_login_audit_config() or not browser_token_hash:
        return None
    response = requests.get(
        get_login_browsers_url(),
        headers=get_login_audit_headers(),
        params={
            "select": "id,browser_token_hash,first_seen_at,last_seen_at,is_active,is_acknowledged",
            "browser_token_hash": f"eq.{browser_token_hash}",
            "limit": "1",
        },
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"ログインブラウザ確認に失敗しました（{response.status_code}）。")
    rows = response.json() or []
    return rows[0] if rows else None


def save_login_browser_seen(browser_token_hash, user_agent, ip_address):
    """現在のブラウザを記録し、初見だった場合はTrueを返す。"""
    existing = fetch_login_browser_by_hash(browser_token_hash)
    info = parse_login_client_info(user_agent)
    now_iso = utc_now_iso()
    common_payload = {
        "last_seen_at": now_iso,
        "last_ip_address": str(ip_address or "")[:200],
        "last_user_agent": str(user_agent or "")[:1000],
        "browser_name": info["browser_name"],
        "os_name": info["os_name"],
        "device_type": info["device_type"],
        "updated_at": now_iso,
    }

    if existing:
        response = requests.patch(
            get_login_browsers_url(),
            headers=get_login_audit_headers(prefer="return=minimal"),
            params={"id": f"eq.{existing['id']}"},
            json=common_payload,
            timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
        )
        if response.status_code not in (200, 204):
            raise RuntimeError(f"ログインブラウザ更新に失敗しました（{response.status_code}）。")
        return False

    payload = {
        "browser_token_hash": browser_token_hash,
        "first_seen_at": now_iso,
        "first_ip_address": str(ip_address or "")[:200],
        "first_user_agent": str(user_agent or "")[:1000],
        "is_active": True,
        "is_acknowledged": False,
        "created_at": now_iso,
        **common_payload,
    }
    response = requests.post(
        get_login_browsers_url(),
        headers=get_login_audit_headers(prefer="return=minimal"),
        json=payload,
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code == 409:
        # 同時実行で別処理が先に登録した場合は既存扱いにする。
        return False
    if response.status_code not in (200, 201):
        raise RuntimeError(f"ログインブラウザ登録に失敗しました（{response.status_code}）。")
    return True


def record_login_audit_event(
    event_key,
    event_type,
    claims,
    browser_token_hash,
    is_new_browser=False,
    severity="info",
    details=None,
):
    if not has_login_audit_config():
        return False
    user_agent = get_request_user_agent()
    ip_address = get_request_ip_address()
    info = parse_login_client_info(user_agent)
    payload = {
        "event_key": str(event_key)[:300],
        "event_type": str(event_type)[:80],
        "occurred_at": utc_now_iso(),
        "browser_token_hash": str(browser_token_hash or "")[:128] or None,
        "microsoft_subject_hash": get_microsoft_subject_hash(claims) or None,
        "account_label": get_microsoft_account_label(claims),
        "is_new_browser": bool(is_new_browser),
        "severity": str(severity or "info")[:30],
        "ip_address": str(ip_address or "")[:200],
        "user_agent": str(user_agent or "")[:1000],
        "browser_name": info["browser_name"],
        "os_name": info["os_name"],
        "device_type": info["device_type"],
        "details": dict(details or {}),
    }
    response = requests.post(
        get_login_events_url(),
        headers=get_login_audit_headers(
            prefer="resolution=ignore-duplicates,return=minimal"
        ),
        params={"on_conflict": "event_key"},
        json=payload,
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code not in (200, 201, 204, 409):
        raise RuntimeError(f"ログイン履歴保存に失敗しました（{response.status_code}）。")
    return True


def ensure_login_audit_for_current_session(claims, login_payload):
    """ログイン成功を重複なく記録する。履歴障害でアプリ本体は止めない。"""
    if not has_login_audit_config():
        return "ログイン履歴用Supabase設定がありません。"

    event_id = str(dict(login_payload or {}).get("jti") or "").strip()
    if not event_id:
        return "ログイン履歴用のセッション識別子を確認できません。"

    session_key = f"login_audit_recorded_{event_id}"
    if st.session_state.get(session_key):
        return ""

    try:
        browser_hash = get_current_browser_token_hash(create_if_missing=True)
        user_agent = get_request_user_agent()
        ip_address = get_request_ip_address()
        is_new_browser = save_login_browser_seen(browser_hash, user_agent, ip_address)
        record_login_audit_event(
            event_key=f"login_success:{event_id}",
            event_type="login_success",
            claims=claims,
            browser_token_hash=browser_hash,
            is_new_browser=is_new_browser,
            severity="warning" if is_new_browser else "info",
            details={"authentication": "microsoft_oidc"},
        )
        st.session_state[session_key] = True
        clear_login_audit_caches()
        if is_new_browser:
            st.session_state["new_browser_login_notice"] = True
        return ""
    except Exception as exc:
        return str(exc)


def record_denied_microsoft_login(claims, reason):
    """許可外アカウントの到達を記録する。失敗しても拒否処理自体は継続する。"""
    if not has_login_audit_config():
        return
    try:
        browser_hash = get_current_browser_token_hash(create_if_missing=True)
        subject_hash = get_microsoft_subject_hash(claims)
        issued_at = str(dict(claims or {}).get("iat") or "")
        key_material = f"{subject_hash}:{issued_at}:{browser_hash}:{reason}"
        event_digest = hashlib.sha256(key_material.encode("utf-8")).hexdigest()
        record_login_audit_event(
            event_key=f"account_denied:{event_digest}",
            event_type="account_denied",
            claims=claims,
            browser_token_hash=browser_hash,
            is_new_browser=False,
            severity="danger",
            details={"reason": str(reason or "account_not_allowed")[:100]},
        )
        clear_login_audit_caches()
    except Exception:
        pass


def record_logout_event(claims, login_payload):
    if not has_login_audit_config():
        return
    try:
        event_id = str(dict(login_payload or {}).get("jti") or "").strip()
        if not event_id:
            return
        browser_hash = get_current_browser_token_hash(create_if_missing=False)
        record_login_audit_event(
            event_key=f"logout:{event_id}",
            event_type="logout",
            claims=claims,
            browser_token_hash=browser_hash,
            severity="info",
        )
        clear_login_audit_caches()
    except Exception:
        pass


def clear_login_audit_caches():
    for cached_function_name in (
        "load_login_browsers_from_supabase",
        "load_login_events_from_supabase",
    ):
        cached_function = globals().get(cached_function_name)
        if cached_function is not None and hasattr(cached_function, "clear"):
            try:
                cached_function.clear()
            except Exception:
                pass

def set_query_params_after_onedrive_auth(
    login_token,
    page="home",
    customer="",
    partner_id="",
    partner_type="",
):
    token_text = str(login_token or "").strip()
    if not validate_login_token(token_text):
        raise RuntimeError("ログイン期限が切れています。もう一度ログインしてください。")
    params = {
        "page": str(page or "home"),
    }
    if customer:
        params["customer"] = str(customer)
    if partner_id:
        params["partner_id"] = str(partner_id)
    if partner_type:
        params["partner_type"] = str(partner_type)
    set_query_params_safely(params)


def make_pkce_verifier():
    return uuid.uuid4().hex + uuid.uuid4().hex


def make_pkce_challenge(verifier):
    digest = hashlib.sha256(verifier.encode("ascii")).digest()
    return base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")


def build_onedrive_sign_in_url(
    return_page="home",
    customer_name="",
    partner_id="",
    partner_type="",
):
    client_id, _, redirect_uri = read_onedrive_settings()
    login_token = get_active_login_token()
    if not validate_login_token(login_token):
        raise RuntimeError("ログイン期限が切れています。もう一度ログインしてください。")
    state = uuid.uuid4().hex
    verifier = make_pkce_verifier()
    save_onedrive_pending_auth_flow(
        state,
        {
            "code_verifier": verifier,
            "redirect_uri": redirect_uri,
            "return_page": str(return_page or "home"),
            "customer_name": str(customer_name or ""),
            "partner_id": str(partner_id or ""),
            "partner_type": str(partner_type or ""),
            "login_token": login_token,
        },
    )
    params = {
        "client_id": client_id,
        "response_type": "code",
        "redirect_uri": redirect_uri,
        "response_mode": "query",
        "scope": ONEDRIVE_SCOPES,
        "state": state,
        "code_challenge": make_pkce_challenge(verifier),
        "code_challenge_method": "S256",
    }
    return ONEDRIVE_AUTHORIZE_URL + "?" + urllib.parse.urlencode(params)


def save_onedrive_token_result(result):
    token = dict(result or {})
    expires_in = int(token.get("expires_in") or 3600)
    token["expires_at"] = time.time() + max(expires_in, 60)
    st.session_state["onedrive_token_result"] = token
    store = get_onedrive_shared_token_store()
    with store["lock"]:
        store["token"] = dict(token)


def clear_onedrive_auth_state(clear_shared=False):
    st.session_state.pop("onedrive_token_result", None)
    for key in list(st.session_state.keys()):
        if str(key).startswith("onedrive_thumbnail_"):
            st.session_state.pop(key, None)
    if clear_shared:
        clear_onedrive_shared_token_result()


def refresh_onedrive_access_token(token=None):
    token_data = dict(token or {})
    configured_refresh_token = read_onedrive_configured_refresh_token()
    configured_source = bool(
        token_data.get(ONEDRIVE_CONFIGURED_TOKEN_SOURCE_KEY)
    )

    # Secretsにrefresh_tokenがある場合は、その接続先を常に正として扱う。
    # 一時的なMicrosoft再接続で別アカウントのトークンが混ざっても、保存先を変えない。
    if configured_refresh_token and not configured_source:
        refresh_token = configured_refresh_token
        configured_source = True
    else:
        refresh_token = str(token_data.get("refresh_token") or "").strip()
        if not refresh_token:
            refresh_token = configured_refresh_token
            configured_source = bool(configured_refresh_token)
    if not refresh_token:
        return None

    client_id, client_secret, redirect_uri = read_onedrive_settings()
    response = requests.post(
        ONEDRIVE_TOKEN_URL,
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "redirect_uri": redirect_uri,
            "scope": ONEDRIVE_SCOPES,
        },
        timeout=ONEDRIVE_REQUEST_TIMEOUT,
    )
    if response.status_code != 200:
        return None
    result = response.json()
    if not result.get("refresh_token"):
        result["refresh_token"] = refresh_token
    result[ONEDRIVE_CONFIGURED_TOKEN_SOURCE_KEY] = configured_source
    save_onedrive_token_result(result)
    return result


def get_onedrive_access_token():
    configured_refresh_token = read_onedrive_configured_refresh_token()

    # Secretsにrefresh_tokenがある時は、その値から取得したトークンだけを使う。
    # これにより、保存時と後日の自動接続時でOneDriveアカウントが変わるのを防ぐ。
    token = get_onedrive_shared_token_result()
    if not isinstance(token, dict):
        session_token = st.session_state.get("onedrive_token_result")
        token = dict(session_token) if isinstance(session_token, dict) else None
    if configured_refresh_token and not bool(
        (token or {}).get(ONEDRIVE_CONFIGURED_TOKEN_SOURCE_KEY)
    ):
        token = None

    access_token = str((token or {}).get("access_token") or "").strip()
    expires_at = float((token or {}).get("expires_at") or 0)
    if access_token and expires_at > time.time() + 60:
        return access_token

    # 同時アクセス時に更新処理が重ならないよう、共有ロック内で再確認する。
    store = get_onedrive_shared_token_store()
    with store["lock"]:
        current = store.get("token")
        if configured_refresh_token and not bool(
            (current or {}).get(ONEDRIVE_CONFIGURED_TOKEN_SOURCE_KEY)
        ):
            current = None
        if isinstance(current, dict):
            current_access_token = str(current.get("access_token") or "").strip()
            current_expires_at = float(current.get("expires_at") or 0)
            if current_access_token and current_expires_at > time.time() + 60:
                return current_access_token
            token = dict(current)

        refreshed = refresh_onedrive_access_token(token)
        if refreshed and refreshed.get("access_token"):
            return str(refreshed["access_token"])

    clear_onedrive_auth_state(clear_shared=True)
    return None


def process_onedrive_callback_if_present():
    """有効な12時間ログインを確認してから、Microsoft認証の戻りを処理する。"""
    params = get_raw_query_params()
    if not params.get("code") and not params.get("error"):
        return False

    state = str(params.get("state") or "")
    pending = pop_onedrive_pending_auth_flow(state) if state else None
    login_token = str((pending or {}).get("login_token") or "").strip()
    login_payload = validate_login_token(login_token)
    if not pending or not login_payload:
        st.session_state.authenticated = False
        st.session_state.pop("login_token", None)
        clear_login_token_cookie()
        set_query_params_safely({"page": "home", "auth_invalid": "1"})
        st.rerun()

    return_page = str(pending.get("return_page") or "home")
    customer_name = str(pending.get("customer_name") or "")
    partner_id = str(pending.get("partner_id") or "")
    partner_type = str(pending.get("partner_type") or "")

    st.session_state.authenticated = True
    st.session_state["login_token"] = login_token
    save_login_token_cookie(login_token)
    st.session_state["page"] = return_page
    st.session_state["selected_customer"] = customer_name if return_page == "detail" else None
    if return_page == "partner_detail":
        st.session_state["selected_partner_id"] = partner_id
        st.session_state["selected_partner_type"] = partner_type

    try:
        if params.get("error"):
            description = params.get("error_description") or params.get("error")
            raise RuntimeError(f"Microsoftへのサインインが完了しませんでした：{description}")

        client_id, client_secret, _ = read_onedrive_settings()
        response = requests.post(
            ONEDRIVE_TOKEN_URL,
            data={
                "client_id": client_id,
                "client_secret": client_secret,
                "grant_type": "authorization_code",
                "code": str(params.get("code") or ""),
                "redirect_uri": str(pending.get("redirect_uri") or ""),
                "code_verifier": str(pending.get("code_verifier") or ""),
                "scope": ONEDRIVE_SCOPES,
            },
            timeout=ONEDRIVE_REQUEST_TIMEOUT,
        )
        if response.status_code != 200:
            try:
                detail = response.json().get("error_description") or response.json().get("error")
            except Exception:
                detail = response.text
            raise RuntimeError(f"OneDriveの認証情報を取得できませんでした：{detail}")
        token_result = response.json()
        save_onedrive_token_result(token_result)
        if not read_onedrive_configured_refresh_token():
            setup_refresh_token = str(token_result.get("refresh_token") or "").strip()
            if setup_refresh_token:
                st.session_state["onedrive_refresh_token_setup_value"] = setup_refresh_token
        st.session_state["onedrive_auth_success"] = True
    except Exception as exc:
        st.session_state["onedrive_auth_error"] = str(exc)

    set_query_params_after_onedrive_auth(
        login_token,
        return_page,
        customer_name,
        partner_id,
        partner_type,
    )
    st.rerun()
    return True


def onedrive_graph_request(method, path, access_token, expected=(200,), **kwargs):
    headers = dict(kwargs.pop("headers", {}) or {})
    request_timeout = kwargs.pop("_request_timeout", ONEDRIVE_REQUEST_TIMEOUT)
    headers["Authorization"] = f"Bearer {access_token}"
    response = requests.request(
        method,
        ONEDRIVE_GRAPH_BASE + path,
        headers=headers,
        timeout=request_timeout,
        **kwargs,
    )
    if response.status_code not in expected:
        try:
            payload = response.json()
            message = str(payload.get("error", {}).get("message", "")).strip()
        except Exception:
            message = str(response.text or "").strip()
        if response.status_code == 401:
            clear_onedrive_auth_state(clear_shared=True)
            message = message or "OneDriveの認証が失効しています。管理者が再接続してください。"
        raise RuntimeError(
            f"Microsoft Graphでエラーが発生しました（{response.status_code}）"
            + (f"：{message}" if message else "")
        )
    return response


def get_onedrive_profile(access_token):
    return onedrive_graph_request(
        "GET",
        "/me?$select=displayName,mail,userPrincipalName",
        access_token,
    ).json()


def get_onedrive_path_item(access_token, path):
    encoded = urllib.parse.quote(str(path).strip("/"), safe="/")
    response = onedrive_graph_request(
        "GET",
        f"/me/drive/root:/{encoded}?$select=id,name,folder,webUrl",
        access_token,
        expected=(200, 404),
    )
    return None if response.status_code == 404 else response.json()


def ensure_onedrive_folder_path(access_token, path):
    segments = [segment for segment in str(path).replace("\\", "/").split("/") if segment]
    if not segments:
        raise RuntimeError("OneDriveの保存フォルダが空です。")
    current_path = ""
    parent_id = None
    item = None
    for segment in segments:
        current_path = f"{current_path}/{segment}".strip("/")
        item = get_onedrive_path_item(access_token, current_path)
        if item:
            parent_id = str(item.get("id") or "")
            continue
        target = "/me/drive/root/children" if not parent_id else f"/me/drive/items/{urllib.parse.quote(parent_id, safe='')}/children"
        response = onedrive_graph_request(
            "POST",
            target,
            access_token,
            expected=(200, 201),
            headers={"Content-Type": "application/json"},
            json={
                "name": segment,
                "folder": {},
                "@microsoft.graph.conflictBehavior": "fail",
            },
        )
        item = response.json()
        parent_id = str(item.get("id") or "")
    return item or {}


def upload_onedrive_file_to_existing_folder(
    access_token,
    folder_path,
    filename,
    content,
    content_type,
):
    """作成済みの保存先へファイル本体だけをアップロードする。"""
    clean_name = re.sub(r"[\\/:*?\"<>|]", "_", str(filename or "")).strip().rstrip(".")
    if not clean_name:
        raise RuntimeError("ファイル名が空です。")
    full_path = f"{str(folder_path).strip('/')}/{clean_name}"
    encoded = urllib.parse.quote(full_path, safe="/")
    response = onedrive_graph_request(
        "PUT",
        f"/me/drive/root:/{encoded}:/content",
        access_token,
        expected=(200, 201),
        headers={"Content-Type": content_type or "application/octet-stream"},
        data=content,
    )
    return response.json()


def upload_onedrive_file(access_token, folder_path, filename, content, content_type):
    ensure_onedrive_folder_path(access_token, folder_path)
    return upload_onedrive_file_to_existing_folder(
        access_token,
        folder_path,
        filename,
        content,
        content_type,
    )


def get_onedrive_camera_roll_folder(access_token):
    """OneDriveのカメラバックアップ用フォルダを取得する。"""
    response = onedrive_graph_request(
        "GET",
        "/me/drive/special/cameraroll",
        access_token,
        expected=(200, 404),
        params={"$select": "id,name,parentReference"},
    )
    if response.status_code == 404:
        return None
    item = response.json()
    return item if isinstance(item, dict) else None


def normalize_onedrive_item_path(value):
    """Graphが返すOneDrive内パスを比較用に正規化する。"""
    return urllib.parse.unquote(str(value or "")).replace("\\", "/").rstrip("/").casefold()


def normalize_image_content_for_comparison(content):
    """画像メタデータを除いた比較用バイト列を返す。解析できない形式は元データを返す。"""
    data = bytes(content or b"")
    if not data:
        return b""

    # JPEG: EXIF・XMP・ICC・コメント等のAPP/COMセグメントだけを除外する。
    if data.startswith(b"\xff\xd8"):
        output = bytearray(data[:2])
        position = 2
        try:
            while position < len(data):
                marker_start = position
                if data[position] != 0xFF:
                    return data
                while position < len(data) and data[position] == 0xFF:
                    position += 1
                if position >= len(data):
                    return data
                marker = data[position]
                position += 1

                # SOI/EOI/TEM/RSTは長さを持たない。
                if marker in {0xD8, 0xD9, 0x01} or 0xD0 <= marker <= 0xD7:
                    output.extend(data[marker_start:position])
                    if marker == 0xD9:
                        break
                    continue

                if position + 2 > len(data):
                    return data
                segment_length = int.from_bytes(data[position:position + 2], "big")
                segment_end = position + segment_length
                if segment_length < 2 or segment_end > len(data):
                    return data

                # SOS以降は圧縮画像本体なので、そのまま残す。
                if marker == 0xDA:
                    output.extend(data[marker_start:segment_end])
                    output.extend(data[segment_end:])
                    break

                is_metadata = 0xE0 <= marker <= 0xEF or marker == 0xFE
                if not is_metadata:
                    output.extend(data[marker_start:segment_end])
                position = segment_end
            return bytes(output)
        except Exception:
            return data

    # PNG: 表示に必須のcritical chunkだけを残す。
    png_signature = b"\x89PNG\r\n\x1a\n"
    if data.startswith(png_signature):
        output = bytearray(png_signature)
        position = len(png_signature)
        try:
            while position + 12 <= len(data):
                chunk_length = int.from_bytes(data[position:position + 4], "big")
                chunk_end = position + 12 + chunk_length
                if chunk_end > len(data):
                    return data
                chunk_type = data[position + 4:position + 8]
                # PNGでは先頭文字が大文字のchunkがcritical。
                if chunk_type and 65 <= chunk_type[0] <= 90:
                    output.extend(data[position:chunk_end])
                position = chunk_end
                if chunk_type == b"IEND":
                    break
            return bytes(output)
        except Exception:
            return data

    # WebP: EXIF/XMP/ICCだけを除き、画像データchunkを比較する。
    if len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        output = bytearray(b"WEBP")
        position = 12
        try:
            while position + 8 <= len(data):
                chunk_type = data[position:position + 4]
                chunk_length = int.from_bytes(data[position + 4:position + 8], "little")
                padded_length = chunk_length + (chunk_length % 2)
                chunk_end = position + 8 + padded_length
                if chunk_end > len(data):
                    return data
                if chunk_type not in {b"EXIF", b"XMP ", b"ICCP"}:
                    output.extend(chunk_type)
                    output.extend(data[position + 8:position + 8 + chunk_length])
                position = chunk_end
            return bytes(output)
        except Exception:
            return data

    return data


def build_image_visual_signature(content):
    """同じ写真の再圧縮・EXIF差を判定するための小さな画素署名を作る。"""
    if Image is None or not content:
        return None
    try:
        with Image.open(BytesIO(content)) as source_image:
            source_image.load()
            image = ImageOps.exif_transpose(source_image) if ImageOps is not None else source_image.copy()
            image = image.convert("RGB")
            width, height = image.size
            if width <= 0 or height <= 0:
                return None
            resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
            color_bytes = image.resize((32, 32), resampling).tobytes()
            grayscale_values = list(
                image.convert("L").resize((16, 16), resampling).tobytes()
            )
            average = sum(grayscale_values) / max(1, len(grayscale_values))
            average_hash = tuple(value >= average for value in grayscale_values)
            return {
                "size": (int(width), int(height)),
                "color": color_bytes,
                "average_hash": average_hash,
            }
    except Exception:
        return None


def image_visual_signatures_match(left, right):
    """縦横サイズと縮小画素が十分近い場合だけ、同じ写真と判定する。"""
    if not left or not right or left.get("size") != right.get("size"):
        return False
    left_color = left.get("color") or b""
    right_color = right.get("color") or b""
    if not left_color or len(left_color) != len(right_color):
        return False
    mean_absolute_difference = sum(
        abs(left_value - right_value)
        for left_value, right_value in zip(left_color, right_color)
    ) / len(left_color)
    left_hash = left.get("average_hash") or ()
    right_hash = right.get("average_hash") or ()
    if len(left_hash) != len(right_hash):
        return False
    hash_distance = sum(a != b for a, b in zip(left_hash, right_hash))
    # JPEG再圧縮やEXIF除去は許容するが、似た別写真を拾わないよう厳しめにする。
    return mean_absolute_difference <= 4.0 and hash_distance <= 8


def parse_onedrive_datetime(value):
    """Microsoft Graphの日付文字列をUTCのdatetimeへ変換する。"""
    text = str(value or "").strip()
    if not text:
        return None
    try:
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        return None


def onedrive_next_link_to_path(next_link):
    """Graphの@odata.nextLinkをonedrive_graph_request用の相対パスへ戻す。"""
    parsed = urllib.parse.urlparse(str(next_link or ""))
    path = parsed.path or ""
    marker = "/v1.0"
    if marker not in path:
        return ""
    relative = path.split(marker, 1)[1]
    if parsed.query:
        relative += "?" + parsed.query
    return relative


@st.cache_resource(show_spinner=False)
def get_onedrive_camera_folder_hint_store():
    """同じOneDriveで直前に見つけた写真フォルダーを、再検索用に短く保持する。"""
    return {"lock": threading.Lock(), "folders": {}}


def get_onedrive_camera_folder_hints(camera_roll_id):
    """セッション内とアプリ内キャッシュから、優先して調べるフォルダーIDを返す。"""
    clean_camera_roll_id = clean_value(camera_roll_id, blank_text="")
    if not clean_camera_roll_id:
        return []

    session_key = "_onedrive_camera_folder_hints_" + hashlib.sha256(
        clean_camera_roll_id.encode("utf-8")
    ).hexdigest()[:16]
    session_hints = st.session_state.get(session_key, [])
    if not isinstance(session_hints, list):
        session_hints = []

    shared_hints = []
    store = get_onedrive_camera_folder_hint_store()
    try:
        with store["lock"]:
            shared_hints = list(store["folders"].get(clean_camera_roll_id, []))
    except Exception:
        shared_hints = []

    result = []
    for folder_id in list(session_hints) + list(shared_hints):
        clean_folder_id = clean_value(folder_id, blank_text="")
        if clean_folder_id and clean_folder_id not in result:
            result.append(clean_folder_id)
    return result[:4]


def remember_onedrive_camera_folder_hint(camera_roll_id, folder_id):
    """一致した元写真の親フォルダーを、次回の高速検索用に記憶する。"""
    clean_camera_roll_id = clean_value(camera_roll_id, blank_text="")
    clean_folder_id = clean_value(folder_id, blank_text="")
    if not clean_camera_roll_id or not clean_folder_id:
        return

    session_key = "_onedrive_camera_folder_hints_" + hashlib.sha256(
        clean_camera_roll_id.encode("utf-8")
    ).hexdigest()[:16]
    current = st.session_state.get(session_key, [])
    if not isinstance(current, list):
        current = []
    session_hints = [clean_folder_id] + [
        value for value in current if clean_value(value, blank_text="") != clean_folder_id
    ]
    st.session_state[session_key] = session_hints[:4]

    store = get_onedrive_camera_folder_hint_store()
    try:
        with store["lock"]:
            shared = list(store["folders"].get(clean_camera_roll_id, []))
            shared = [clean_folder_id] + [
                value for value in shared if clean_value(value, blank_text="") != clean_folder_id
            ]
            store["folders"][clean_camera_roll_id] = shared[:4]
    except Exception:
        pass


def get_recent_camera_candidate_cache(camera_roll_id):
    """同じ保存操作中に取得済みの候補一覧を再利用する。"""
    cache = st.session_state.get("_onedrive_camera_candidate_cache")
    if not isinstance(cache, dict):
        return []
    if clean_value(cache.get("camera_roll_id"), blank_text="") != clean_value(
        camera_roll_id,
        blank_text="",
    ):
        return []
    try:
        cached_at = float(cache.get("cached_at") or 0)
    except Exception:
        cached_at = 0
    if time.time() - cached_at > 120:
        return []
    items = cache.get("items")
    return list(items) if isinstance(items, list) else []


def set_recent_camera_candidate_cache(camera_roll_id, items):
    """候補一覧を2分だけセッションへ保持し、複数枚保存時の再走査を避ける。"""
    st.session_state["_onedrive_camera_candidate_cache"] = {
        "camera_roll_id": clean_value(camera_roll_id, blank_text=""),
        "cached_at": time.time(),
        "items": list(items) if isinstance(items, list) else [],
    }


def get_camera_candidate_fingerprint(item_id):
    """直前に照合したOneDrive画像の小さな比較情報を再利用する。"""
    clean_item_id = clean_value(item_id, blank_text="")
    cache = st.session_state.get("_onedrive_camera_fingerprint_cache")
    if not clean_item_id or not isinstance(cache, dict):
        return None
    entry = cache.get(clean_item_id)
    if not isinstance(entry, dict):
        return None
    try:
        cached_at = float(entry.get("cached_at") or 0)
    except Exception:
        cached_at = 0
    if time.time() - cached_at > 120:
        return None
    return entry


def set_camera_candidate_fingerprint(item_id, content):
    """画像本体は保持せず、ハッシュと縮小画素だけを2分間保持する。"""
    clean_item_id = clean_value(item_id, blank_text="")
    if not clean_item_id or not content:
        return None
    entry = {
        "cached_at": time.time(),
        "size": len(content),
        "sha256": hashlib.sha256(content).digest(),
        "normalized_sha256": hashlib.sha256(
            normalize_image_content_for_comparison(content)
        ).digest(),
        "visual_signature": build_image_visual_signature(content),
    }
    cache = st.session_state.get("_onedrive_camera_fingerprint_cache")
    if not isinstance(cache, dict):
        cache = {}
    cache[clean_item_id] = entry
    # 写真本体は保持しないが、セッションが長時間続いても増え続けないよう上限を設ける。
    if len(cache) > 40:
        ordered = sorted(
            cache.items(),
            key=lambda pair: float((pair[1] or {}).get("cached_at") or 0),
            reverse=True,
        )[:40]
        cache = dict(ordered)
    st.session_state["_onedrive_camera_fingerprint_cache"] = cache
    return entry


def remove_recent_camera_candidate_cache_item(item_id):
    """移動済みの写真を候補キャッシュから外す。"""
    clean_item_id = clean_value(item_id, blank_text="")
    cache = st.session_state.get("_onedrive_camera_candidate_cache")
    if not clean_item_id or not isinstance(cache, dict):
        return
    items = cache.get("items")
    if not isinstance(items, list):
        return
    cache["items"] = [
        item
        for item in items
        if clean_value((item or {}).get("id"), blank_text="") != clean_item_id
    ]
    cache["cached_at"] = time.time()
    st.session_state["_onedrive_camera_candidate_cache"] = cache


def list_onedrive_folder_children(
    access_token,
    folder_id,
    max_pages=1,
    page_size=200,
    request_timeout=10,
):
    """検索インデックスを使わず、指定フォルダーの新しい項目を直接取得する。"""
    clean_folder_id = clean_value(folder_id, blank_text="")
    if not clean_folder_id:
        return []

    target = f"/me/drive/items/{urllib.parse.quote(clean_folder_id, safe='')}/children"
    params = {
        "$select": (
            "id,name,size,file,folder,parentReference,webUrl,image,photo,"
            "createdDateTime,lastModifiedDateTime"
        ),
        "$top": str(max(1, min(int(page_size or 200), 200))),
        "$orderby": "lastModifiedDateTime desc",
    }
    items = []
    page_count = 0

    while target and page_count < max(1, int(max_pages or 1)):
        if page_count == 0:
            response = onedrive_graph_request(
                "GET",
                target,
                access_token,
                params=params,
                _request_timeout=request_timeout,
            )
        else:
            response = onedrive_graph_request(
                "GET",
                target,
                access_token,
                _request_timeout=request_timeout,
            )
        payload = response.json()
        if not isinstance(payload, dict):
            break
        page_items = payload.get("value", [])
        if isinstance(page_items, list):
            items.extend(item for item in page_items if isinstance(item, dict))
        target = onedrive_next_link_to_path(payload.get("@odata.nextLink"))
        page_count += 1

    return items


def onedrive_camera_folder_priority(name):
    """Androidの写真バックアップで使われやすいフォルダー名を先に調べる。"""
    normalized = unicodedata.normalize("NFKC", str(name or "")).casefold()
    priority_tokens = (
        "camera",
        "camera roll",
        "dcim",
        "samsung",
        "gallery",
        "pictures",
        "photos",
        "カメラ",
        "写真",
        "画像",
    )
    return 0 if any(token in normalized for token in priority_tokens) else 1


def collect_onedrive_camera_backup_images(
    access_token,
    camera_roll,
    preferred_folder_ids=None,
    max_folders=30,
    max_items=800,
    max_depth=4,
    max_seconds=8,
):
    """候補フォルダーを新しい順・優先順で短時間だけたどる。"""
    camera_roll = camera_roll if isinstance(camera_roll, dict) else {}
    camera_roll_id = clean_value(camera_roll.get("id"), blank_text="")
    parent_reference = camera_roll.get("parentReference") or {}
    camera_parent_id = clean_value(parent_reference.get("id"), blank_text="")
    if not camera_roll_id:
        return []

    queue = []
    queued_folders = set()

    def enqueue(folder_id, depth, priority=False):
        clean_folder_id = clean_value(folder_id, blank_text="")
        if not clean_folder_id or clean_folder_id in queued_folders:
            return
        queued_folders.add(clean_folder_id)
        entry = (clean_folder_id, int(depth or 0))
        if priority:
            queue.insert(0, entry)
        else:
            queue.append(entry)

    for folder_id in preferred_folder_ids or []:
        enqueue(folder_id, 0, priority=True)
    enqueue(camera_roll_id, 0, priority=not bool(preferred_folder_ids))
    if camera_parent_id and camera_parent_id != camera_roll_id:
        enqueue(camera_parent_id, 0, priority=False)

    seen_folders = set()
    seen_items = set()
    images = []
    app_root_name = str(ONEDRIVE_ROOT_FOLDER).strip("/").casefold()
    app_root_token = "/" + app_root_name
    started_at = time.monotonic()

    while queue and len(seen_folders) < max_folders and len(seen_items) < max_items:
        if time.monotonic() - started_at >= max_seconds:
            break

        folder_id, depth = queue.pop(0)
        if not folder_id or folder_id in seen_folders:
            continue
        seen_folders.add(folder_id)

        remaining_seconds = max_seconds - (time.monotonic() - started_at)
        request_timeout = max(3, min(8, int(remaining_seconds) + 1))
        try:
            folder_items = list_onedrive_folder_children(
                access_token,
                folder_id,
                max_pages=1,
                page_size=200,
                request_timeout=request_timeout,
            )
        except Exception:
            continue

        priority_folders = []
        normal_folders = []
        for item in folder_items:
            item_id = clean_value(item.get("id"), blank_text="")
            if not item_id or item_id in seen_items:
                continue
            seen_items.add(item_id)
            if len(seen_items) > max_items:
                break

            parent_path = normalize_onedrive_item_path(
                (item.get("parentReference") or {}).get("path")
            )
            # すでにアプリの保存先へ移した写真は、移動元候補にしない。
            if parent_path.endswith(app_root_token) or app_root_token + "/" in parent_path:
                continue

            name = Path(str(item.get("name") or "")).name
            if item.get("folder"):
                if name.casefold() == app_root_name:
                    continue
                if depth < max_depth:
                    target = (item_id, depth + 1)
                    if onedrive_camera_folder_priority(name) == 0:
                        priority_folders.append(target)
                    else:
                        normal_folders.append(target)
                continue

            suffix = Path(name).suffix.casefold()
            mime_type = clean_value(
                (item.get("file") or {}).get("mimeType"),
                blank_text="",
            ).casefold()
            is_image = bool(
                item.get("image")
                or mime_type.startswith("image/")
                or suffix in {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
            )
            if is_image:
                images.append(item)

        # 優先フォルダーは次の周回で先に調べる。通常フォルダーは後ろへ回す。
        for child_id, child_depth in reversed(priority_folders):
            enqueue(child_id, child_depth, priority=True)
        for child_id, child_depth in normal_folders:
            enqueue(child_id, child_depth, priority=False)

        # 新しい画像を十分取得できたら、無制限にフォルダーを広げない。
        if len(images) >= 260 and len(seen_folders) >= 3:
            break

    images.sort(
        key=lambda item: str(
            item.get("lastModifiedDateTime")
            or item.get("createdDateTime")
            or ""
        ),
        reverse=True,
    )
    return images


def onedrive_image_dimensions(item):
    """Graphのimage facetから画像サイズを取得する。"""
    image_info = item.get("image") or {}
    try:
        width = int(image_info.get("width") or 0)
        height = int(image_info.get("height") or 0)
    except Exception:
        return None
    return (width, height) if width > 0 and height > 0 else None


def match_onedrive_photo_candidates(
    access_token,
    candidates,
    original_name,
    content,
    max_downloads=12,
    max_seconds=6,
):
    """取得済み候補から、完全一致または厳しい画素一致の写真を1枚だけ返す。"""
    uploaded_size = len(content)
    uploaded_sha1 = hashlib.sha1(content).hexdigest().upper()
    uploaded_sha256 = hashlib.sha256(content).digest()
    uploaded_normalized_sha256 = hashlib.sha256(
        normalize_image_content_for_comparison(content)
    ).digest()
    uploaded_visual_signature = build_image_visual_signature(content)
    uploaded_dimensions = (
        uploaded_visual_signature.get("size")
        if uploaded_visual_signature is not None
        else None
    )

    now_utc = datetime.now(timezone.utc)
    recent_cutoff = now_utc - timedelta(days=14)
    exact_name_candidates = []
    recent_dimension_candidates = []
    recent_unknown_dimension_candidates = []

    for candidate in candidates or []:
        if not isinstance(candidate, dict) or candidate.get("folder"):
            continue
        candidate_name = Path(str(candidate.get("name") or "")).name
        if candidate_name.casefold() == original_name.casefold():
            exact_name_candidates.append(candidate)
            continue

        modified_at = parse_onedrive_datetime(
            candidate.get("lastModifiedDateTime") or candidate.get("createdDateTime")
        )
        if modified_at is not None and modified_at < recent_cutoff:
            continue

        candidate_dimensions = onedrive_image_dimensions(candidate)
        if uploaded_dimensions and candidate_dimensions:
            if (
                candidate_dimensions == uploaded_dimensions
                or candidate_dimensions == uploaded_dimensions[::-1]
            ):
                recent_dimension_candidates.append(candidate)
        elif modified_at is not None:
            recent_unknown_dimension_candidates.append(candidate)

    selected_candidates = []
    seen_candidate_ids = set()
    for group, limit in (
        (exact_name_candidates, 20),
        (recent_dimension_candidates, 12),
        (recent_unknown_dimension_candidates, 4),
    ):
        for candidate in group[:limit]:
            candidate_id = clean_value(candidate.get("id"), blank_text="")
            if candidate_id and candidate_id not in seen_candidate_ids:
                seen_candidate_ids.add(candidate_id)
                selected_candidates.append(candidate)

    # Graphが返すSHA1で一致する場合は、画像をダウンロードせず判定できる。
    graph_hash_matches = []
    for candidate in selected_candidates:
        try:
            candidate_size = int(candidate.get("size") or 0)
        except Exception:
            candidate_size = 0
        hashes = ((candidate.get("file") or {}).get("hashes") or {})
        candidate_sha1 = clean_value(hashes.get("sha1Hash"), blank_text="").upper()
        if candidate_size == uploaded_size and candidate_sha1 == uploaded_sha1:
            graph_hash_matches.append(candidate)
    if len(graph_hash_matches) == 1:
        return graph_hash_matches[0]
    if len(graph_hash_matches) > 1:
        raise RuntimeError(
            "OneDriveの写真バックアップに同じ写真が複数見つかりました（完全一致）。"
            "誤った写真を移動しないよう保存を停止しました。"
        )

    raw_matches = []
    normalized_matches = []
    visual_matches = []
    started_at = time.monotonic()
    download_count = 0

    for candidate in selected_candidates:
        if time.monotonic() - started_at >= max_seconds:
            break
        candidate_id = clean_value(candidate.get("id"), blank_text="")
        if not candidate_id:
            continue

        fingerprint = get_camera_candidate_fingerprint(candidate_id)
        if fingerprint is None:
            if download_count >= max_downloads:
                break
            try:
                candidate_content = onedrive_graph_request(
                    "GET",
                    f"/me/drive/items/{urllib.parse.quote(candidate_id, safe='')}/content",
                    access_token,
                    expected=(200,),
                    _request_timeout=8,
                ).content
            except Exception:
                continue
            download_count += 1
            fingerprint = set_camera_candidate_fingerprint(
                candidate_id,
                candidate_content,
            )
        if not isinstance(fingerprint, dict):
            continue

        try:
            candidate_size = int(candidate.get("size") or fingerprint.get("size") or 0)
        except Exception:
            candidate_size = 0
        if (
            candidate_size == uploaded_size
            and fingerprint.get("sha256") == uploaded_sha256
        ):
            raw_matches.append(candidate)
            continue

        if fingerprint.get("normalized_sha256") == uploaded_normalized_sha256:
            normalized_matches.append(candidate)
            continue

        if uploaded_visual_signature is not None:
            candidate_visual_signature = fingerprint.get("visual_signature")
            if image_visual_signatures_match(
                uploaded_visual_signature,
                candidate_visual_signature,
            ):
                visual_matches.append(candidate)

    for match_level, matches in (
        ("完全一致", raw_matches),
        ("画像本体一致", normalized_matches),
        ("画素一致", visual_matches),
    ):
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise RuntimeError(
                f"OneDriveの写真バックアップに同じ写真が複数見つかりました（{match_level}）。"
                "誤った写真を移動しないよう保存を停止しました。"
            )
    return None


def find_matching_onedrive_camera_roll_file(
    access_token,
    uploaded_name,
    content,
    excluded_item_ids=None,
):
    """重複を作らず、短時間の優先検索でOneDrive上の元写真を特定する。"""
    original_name = Path(str(uploaded_name or "")).name
    if not original_name or not content:
        return None

    excluded_ids = {
        clean_value(item_id, blank_text="")
        for item_id in list(excluded_item_ids or [])
        if clean_value(item_id, blank_text="")
    }

    def available_candidates(items):
        return [
            item
            for item in list(items or [])
            if clean_value((item or {}).get("id"), blank_text="") not in excluded_ids
        ]

    camera_roll = get_onedrive_camera_roll_folder(access_token) or {}
    camera_roll_id = clean_value(camera_roll.get("id"), blank_text="")
    if not camera_roll_id:
        # カメラバックアップ用フォルダーがない場合は、
        # 呼び出し側で選択画像を新規アップロードする。
        return None

    # 同じ保存操作で取得済みの一覧があれば、再走査せず先に照合する。
    cached_candidates = get_recent_camera_candidate_cache(camera_roll_id)
    if cached_candidates:
        matched = match_onedrive_photo_candidates(
            access_token,
            available_candidates(cached_candidates),
            original_name,
            content,
            max_downloads=10,
            max_seconds=5,
        )
        if matched is not None:
            source_parent_id = clean_value(
                (matched.get("parentReference") or {}).get("id"),
                blank_text="",
            )
            remember_onedrive_camera_folder_hint(camera_roll_id, source_parent_id)
            remove_recent_camera_candidate_cache_item(matched.get("id"))
            return matched

    folder_hints = get_onedrive_camera_folder_hints(camera_roll_id)

    # 前回見つかったフォルダーがある場合は、まずそこだけを短時間で確認する。
    if folder_hints:
        fast_candidates = collect_onedrive_camera_backup_images(
            access_token,
            camera_roll,
            preferred_folder_ids=folder_hints,
            max_folders=5,
            max_items=350,
            max_depth=1,
            max_seconds=4,
        )
        matched = match_onedrive_photo_candidates(
            access_token,
            available_candidates(fast_candidates),
            original_name,
            content,
            max_downloads=8,
            max_seconds=4,
        )
        if matched is not None:
            source_parent_id = clean_value(
                (matched.get("parentReference") or {}).get("id"),
                blank_text="",
            )
            remember_onedrive_camera_folder_hint(camera_roll_id, source_parent_id)
            set_recent_camera_candidate_cache(camera_roll_id, fast_candidates)
            remove_recent_camera_candidate_cache_item(matched.get("id"))
            return matched

    # 初回だけは候補フォルダーを優先順で調べるが、時間・件数・ページ数に上限を設ける。
    broad_candidates = collect_onedrive_camera_backup_images(
        access_token,
        camera_roll,
        preferred_folder_ids=folder_hints,
        max_folders=30,
        max_items=800,
        max_depth=4,
        max_seconds=8,
    )
    set_recent_camera_candidate_cache(camera_roll_id, broad_candidates)
    matched = match_onedrive_photo_candidates(
        access_token,
        available_candidates(broad_candidates),
        original_name,
        content,
        max_downloads=12,
        max_seconds=6,
    )
    if matched is not None:
        source_parent_id = clean_value(
            (matched.get("parentReference") or {}).get("id"),
            blank_text="",
        )
        remember_onedrive_camera_folder_hint(camera_roll_id, source_parent_id)
        remove_recent_camera_candidate_cache_item(matched.get("id"))
        return matched

    # OneDrive内に同じ写真がない場合は、呼び出し側で新規アップロードする。
    return None


def load_recent_onedrive_camera_photos(
    access_token,
    days=ONEDRIVE_RECENT_CAMERA_DAYS,
    max_items=ONEDRIVE_RECENT_CAMERA_MAX_ITEMS,
):
    """OneDriveのカメラバックアップから、対応形式の最近の写真だけを返す。"""
    camera_roll = get_onedrive_camera_roll_folder(access_token) or {}
    camera_roll_id = clean_value(camera_roll.get("id"), blank_text="")
    if not camera_roll_id:
        return []

    folder_hints = get_onedrive_camera_folder_hints(camera_roll_id)
    candidates = collect_onedrive_camera_backup_images(
        access_token,
        camera_roll,
        preferred_folder_ids=folder_hints,
        max_folders=30,
        max_items=800,
        max_depth=4,
        max_seconds=8,
    )
    cutoff = datetime.now(timezone.utc) - timedelta(days=max(1, int(days or 1)))
    allowed_mime_types = {"image/jpeg", "image/png", "image/webp"}
    recent_items = []

    for item in candidates:
        if not isinstance(item, dict) or item.get("folder"):
            continue
        item_id = clean_value(item.get("id"), blank_text="")
        name = Path(str(item.get("name") or "")).name
        suffix = Path(name).suffix.casefold()
        mime_type = clean_value(
            (item.get("file") or {}).get("mimeType"),
            blank_text="",
        ).casefold()
        if not item_id or not name:
            continue
        # 既存アプリと同じ対応形式だけを表示し、HEIC等を新たに許可しない。
        if suffix not in ONEDRIVE_IMAGE_EXTENSIONS and mime_type not in allowed_mime_types:
            continue

        modified_at = parse_onedrive_datetime(
            item.get("lastModifiedDateTime") or item.get("createdDateTime")
        )
        if modified_at is None or modified_at < cutoff:
            continue
        recent_items.append(item)

    recent_items.sort(
        key=lambda item: str(
            item.get("lastModifiedDateTime")
            or item.get("createdDateTime")
            or ""
        ),
        reverse=True,
    )
    return recent_items[:max(1, int(max_items or 1))]


def move_onedrive_item(access_token, item_id, parent_id, filename):
    """同じOneDrive内でファイルを移動し、必要なら同時に名前を変更する。"""
    clean_name = re.sub(r'[\\/:*?"<>|]', "_", str(filename or "")).strip().rstrip(".")
    if not item_id or not parent_id or not clean_name:
        raise RuntimeError("OneDriveの移動先情報が不足しています。")
    response = onedrive_graph_request(
        "PATCH",
        f"/me/drive/items/{urllib.parse.quote(str(item_id), safe='')}",
        access_token,
        expected=(200,),
        headers={"Content-Type": "application/json"},
        json={
            "parentReference": {"id": str(parent_id)},
            "name": clean_name,
        },
    )
    return response.json()


def move_onedrive_file_to_folder(
    access_token,
    item_id,
    folder_path,
    filename,
):
    """保存先フォルダを確保してから、既存ファイルをそこへ移動する。"""
    target_folder = ensure_onedrive_folder_path(access_token, folder_path)
    target_folder_id = clean_value(target_folder.get("id"), blank_text="")
    if not target_folder_id:
        raise RuntimeError("OneDriveの移動先フォルダIDを取得できませんでした。")
    return move_onedrive_item(
        access_token,
        item_id,
        target_folder_id,
        filename,
    )


def delete_onedrive_file(access_token, item_id):
    onedrive_graph_request(
        "DELETE",
        f"/me/drive/items/{urllib.parse.quote(str(item_id), safe='')}",
        access_token,
        expected=(204,),
    )


def download_onedrive_file(access_token, item_id):
    response = onedrive_graph_request(
        "GET",
        f"/me/drive/items/{urllib.parse.quote(str(item_id), safe='')}/content",
        access_token,
        expected=(200,),
    )
    return response.content


def get_onedrive_item_download_url(access_token, item_id):
    """画像本文をアプリへ通さず、ブラウザーが直接読む短期URLを返す。"""
    clean_id = clean_value(item_id, blank_text="").strip()
    if not clean_id:
        return ""
    response = onedrive_graph_request(
        "GET",
        f"/me/drive/items/{urllib.parse.quote(clean_id, safe='')}",
        access_token,
        expected=(200,),
    )
    payload = response.json() if response.content else {}
    direct_url = str((payload or {}).get("@microsoft.graph.downloadUrl") or "").strip()
    if direct_url:
        return direct_url

    streamed = requests.get(
        ONEDRIVE_GRAPH_BASE
        + f"/me/drive/items/{urllib.parse.quote(clean_id, safe='')}/content",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=ONEDRIVE_REQUEST_TIMEOUT,
        allow_redirects=True,
        stream=True,
    )
    try:
        if streamed.status_code != 200:
            raise RuntimeError(
                f"Microsoft Graphでエラーが発生しました（{streamed.status_code}）"
            )
        final_url = str(streamed.url or "").strip()
        if not final_url or final_url.startswith(ONEDRIVE_GRAPH_BASE):
            return ""
        return final_url
    finally:
        streamed.close()


def prepare_onedrive_display_image(content, original_name, stored_name):
    """元画像を残し、通常表示用の長辺1600px WebPを作る。"""
    started_at = time.perf_counter()
    original_size = len(content or b"")
    if Image is None or not content:
        log_image_event(
            "display_webp_skipped",
            reason="pillow_unavailable" if Image is None else "empty_content",
            original_size=original_size,
        )
        return None
    try:
        with Image.open(BytesIO(content)) as source_image:
            source_image.load()
            image = (
                ImageOps.exif_transpose(source_image)
                if ImageOps is not None
                else source_image.copy()
            )
            if image.width <= 0 or image.height <= 0:
                log_image_event(
                    "display_webp_failed",
                    reason="invalid_dimensions",
                    original_size=original_size,
                    duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
                )
                return None
            resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
            if max(image.size) > ONEDRIVE_DISPLAY_IMAGE_MAX_EDGE:
                image.thumbnail(
                    (ONEDRIVE_DISPLAY_IMAGE_MAX_EDGE, ONEDRIVE_DISPLAY_IMAGE_MAX_EDGE),
                    resampling,
                )
            if image.mode in {"RGBA", "LA"} or (
                image.mode == "P" and "transparency" in image.info
            ):
                image = image.convert("RGBA")
            else:
                image = image.convert("RGB")
            output = BytesIO()
            image.save(
                output,
                format="WEBP",
                quality=ONEDRIVE_DISPLAY_IMAGE_QUALITY,
                method=ONEDRIVE_DISPLAY_IMAGE_METHOD,
                optimize=True,
            )
            display_content = output.getvalue()
            if not display_content:
                log_image_event(
                    "display_webp_failed",
                    reason="empty_output",
                    original_size=original_size,
                    duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
                )
                return None
            safe_stem = Path(str(stored_name or original_name or "image")).stem
            result = {
                "content": display_content,
                "stored_name": f"{safe_stem}__display.webp",
                "mime_type": "image/webp",
                "size": len(display_content),
                "width": int(image.width),
                "height": int(image.height),
            }
            log_image_event(
                "display_webp_created",
                original_size=original_size,
                display_size=len(display_content),
                width=int(image.width),
                height=int(image.height),
                compression_ratio=(
                    round(len(display_content) / original_size, 4)
                    if original_size > 0
                    else None
                ),
                duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
            )
            return result
    except Exception as exc:
        # 変換できない画像は、元画像の従来保存・従来表示へ戻す。
        log_image_event(
            "display_webp_failed",
            reason=type(exc).__name__,
            original_size=original_size,
            duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
        )
        return None


def get_onedrive_attachment_display_stored_name(attachment):
    """表示用WebPの保存名を返す。旧データは元画像の保存名から安全に補完する。"""
    if not isinstance(attachment, dict):
        return ""
    configured_name = clean_value(
        attachment.get("display_stored_name"), blank_text=""
    ).strip("/")
    if configured_name:
        return Path(configured_name).name
    original_stored_name = clean_value(
        attachment.get("stored_name"), blank_text=""
    ).strip("/")
    if not original_stored_name:
        return ""
    return f"{Path(original_stored_name).stem}__display.webp"


def update_onedrive_attachment_display_fields(
    attachment,
    display_item,
    requested_stored_name="",
    display_content=b"",
):
    """表示用WebP情報をメモリへ反映し、可能ならSupabaseにも保存する。"""
    if not isinstance(attachment, dict) or not isinstance(display_item, dict):
        return ""
    display_file_id = clean_value(display_item.get("id"), blank_text="").strip()
    if not display_file_id:
        return ""

    display_stored_name = (
        clean_value(display_item.get("name"), blank_text="").strip()
        or clean_value(requested_stored_name, blank_text="").strip()
    )
    updated = dict(attachment)
    updated["display_file_id"] = display_file_id
    updated["display_stored_name"] = display_stored_name
    updated["display_mime_type"] = "image/webp"
    updated["display_size"] = int(
        display_item.get("size")
        or len(display_content or b"")
        or attachment.get("display_size")
        or 0
    )
    updated["display_web_url"] = (
        clean_value(display_item.get("webUrl"), blank_text="")
        or clean_value(attachment.get("display_web_url"), blank_text="")
    )

    # 先に現在の画面へ反映する。Supabaseが一時的に更新できなくても、
    # OneDrive上の決定的な保存名から次回表示時に再取得できる。
    attachment.update(updated)
    metadata_id = clean_value(updated.get("id"), blank_text="")
    field_name = clean_value(updated.get("field_name"), blank_text="")
    if metadata_id and field_name:
        try:
            update_customer_information(
                metadata_id,
                field_name,
                serialize_onedrive_attachment(updated),
            )
        except Exception:
            # 表示用WebPの記録更新失敗で、従来の画像表示まで止めない。
            pass
    return display_file_id


def ensure_onedrive_attachment_display_image(access_token, attachment):
    """旧画像を初めて開く時だけ、通常表示用WebPを1枚生成して記録する。"""
    started_at = time.perf_counter()
    if not isinstance(attachment, dict) or attachment.get("file_type") != "image":
        return ""

    existing_display_id = clean_value(
        attachment.get("display_file_id"), blank_text=""
    ).strip()
    if existing_display_id:
        return existing_display_id

    folder_path = clean_value(
        attachment.get("onedrive_path"), blank_text=""
    ).strip("/")
    display_stored_name = get_onedrive_attachment_display_stored_name(attachment)
    if not folder_path or not display_stored_name:
        return ""

    try:
        # 前回のSupabase更新だけ失敗してWebP本体が残っている場合は、
        # 元画像を再ダウンロードせず、そのファイルIDだけを修復する。
        existing_item = get_onedrive_path_item(
            access_token,
            f"{folder_path}/{display_stored_name}",
        )
        if isinstance(existing_item, dict) and not existing_item.get("folder"):
            repaired_id = update_onedrive_attachment_display_fields(
                attachment,
                existing_item,
                requested_stored_name=display_stored_name,
            )
            log_image_event(
                "legacy_display_webp_reused",
                success=bool(repaired_id),
                duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
            )
            return repaired_id

        original_content = download_onedrive_attachment_file(
            access_token,
            attachment,
        )
        display = prepare_onedrive_display_image(
            original_content,
            attachment.get("original_name", ""),
            attachment.get("stored_name", ""),
        )
        if not display:
            return ""

        # 保存名は元画像のstored_nameから決まるため、同時アクセスでも同じ場所を使う。
        display["stored_name"] = display_stored_name
        display_item = upload_onedrive_file_to_existing_folder(
            access_token,
            folder_path,
            display_stored_name,
            display["content"],
            display["mime_type"],
        )
        display_file_id = update_onedrive_attachment_display_fields(
            attachment,
            display_item,
            requested_stored_name=display_stored_name,
            display_content=display["content"],
        )
        log_image_event(
            "legacy_display_webp_saved",
            success=bool(display_file_id),
            original_size=len(original_content or b""),
            display_size=len(display.get("content") or b""),
            duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
        )
        return display_file_id
    except Exception as exc:
        # 変換・保存・記録のどこかが失敗しても、呼び出し側は従来どおり元画像を表示する。
        log_image_event(
            "legacy_display_webp_failed",
            reason=type(exc).__name__,
            duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
        )
        return ""


def repair_onedrive_attachment_display_reference(access_token, attachment):
    """表示用WebPのIDだけが古い場合、保存名から取り直す。"""
    if not isinstance(attachment, dict):
        return ""
    folder_path = clean_value(attachment.get("onedrive_path"), blank_text="").strip("/")
    stored_name = get_onedrive_attachment_display_stored_name(attachment)
    if not folder_path or not stored_name:
        return ""
    item = get_onedrive_path_item(access_token, f"{folder_path}/{stored_name}")
    if not isinstance(item, dict) or item.get("folder"):
        return ""
    repaired_id = clean_value(item.get("id"), blank_text="")
    if not repaired_id:
        return ""
    old_id = clean_value(attachment.get("display_file_id"), blank_text="")
    updated = dict(attachment)
    updated["display_file_id"] = repaired_id
    updated["display_stored_name"] = (
        clean_value(item.get("name"), blank_text="") or stored_name
    )
    updated["display_web_url"] = clean_value(item.get("webUrl"), blank_text="")
    metadata_id = clean_value(updated.get("id"), blank_text="")
    field_name = clean_value(updated.get("field_name"), blank_text="")
    if metadata_id and field_name and repaired_id != old_id:
        update_customer_information(
            metadata_id,
            field_name,
            serialize_onedrive_attachment(updated),
        )
    attachment.update(updated)
    return repaired_id


def get_onedrive_attachment_image_download_url(access_token, attachment):
    """表示用WebPを優先し、旧画像は元画像へフォールバックする。"""
    display_id = clean_value(
        (attachment or {}).get("display_file_id"), blank_text=""
    ).strip()
    if display_id:
        try:
            direct_url = get_onedrive_item_download_url(access_token, display_id)
            if direct_url:
                return direct_url
        except Exception as exc:
            if not is_onedrive_not_found_error(exc):
                raise
            repaired_id = repair_onedrive_attachment_display_reference(
                access_token,
                attachment,
            )
            if repaired_id:
                direct_url = get_onedrive_item_download_url(access_token, repaired_id)
                if direct_url:
                    return direct_url

    item_id = clean_value((attachment or {}).get("file_id"), blank_text="").strip()
    if not item_id:
        item_id = repair_onedrive_attachment_reference(access_token, attachment)
    if not item_id:
        raise RuntimeError("OneDriveファイルIDを確認できませんでした。")
    try:
        return get_onedrive_item_download_url(access_token, item_id)
    except Exception as exc:
        if not is_onedrive_not_found_error(exc):
            raise
        repaired_id = repair_onedrive_attachment_reference(access_token, attachment)
        if not repaired_id or repaired_id == item_id:
            raise
        return get_onedrive_item_download_url(access_token, repaired_id)


def build_onedrive_image_gallery_items(access_token, attachments):
    """本文を読まず、短期URLだけを最大4件並行で取得する。"""
    image_attachments = [
        item
        for item in list(attachments or [])
        if isinstance(item, dict) and item.get("file_type") == "image"
    ]
    if not image_attachments:
        return [], [], []
    results = {}
    max_workers = min(ONEDRIVE_GALLERY_URL_WORKERS, len(image_attachments))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(
                get_onedrive_attachment_image_download_url,
                access_token,
                attachment,
            ): index
            for index, attachment in enumerate(image_attachments)
        }
        for future in as_completed(future_map):
            index = future_map[future]
            try:
                results[index] = (future.result(), "")
            except Exception as exc:
                results[index] = ("", str(exc))

    gallery_items = []
    missing_names = []
    other_errors = []
    for index, attachment in enumerate(image_attachments):
        filename = str(attachment.get("original_name") or "名称未設定")
        original_url = clean_value(
            attachment.get("web_url"), blank_text=""
        ).strip()
        direct_url, error_text = results.get(index, ("", ""))
        if direct_url:
            gallery_items.append(
                {
                    "url": direct_url,
                    "filename": filename,
                    "original_url": original_url,
                }
            )
        elif error_text and ("404" in error_text or "見つか" in error_text):
            missing_names.append(filename)
        elif error_text:
            other_errors.append((filename, error_text))
        else:
            missing_names.append(filename)
    return gallery_items, missing_names, other_errors


def open_onedrive_image_group_gallery(access_token, attachments):
    """最初の画像だけを読み、次の画像は移動した時に読み込ませる。"""
    started_at = time.perf_counter()
    image_attachments = [
        item
        for item in list(attachments or [])
        if isinstance(item, dict) and item.get("file_type") == "image"
    ]
    with st.spinner("画像を準備しています…"):
        # 旧画像は最初に表示する1枚だけを初回WebP化する。
        # グループ全件を変換すると元画像の一括ダウンロードへ戻るため行わない。
        if image_attachments:
            ensure_onedrive_attachment_display_image(
                access_token,
                image_attachments[0],
            )
        image_items, missing_names, other_errors = build_onedrive_image_gallery_items(
            access_token,
            image_attachments,
        )
    log_image_event(
        "gallery_prepared",
        requested_count=len(image_attachments),
        available_count=len(image_items),
        missing_count=len(missing_names),
        error_count=len(other_errors),
        duration_ms=round((time.perf_counter() - started_at) * 1000, 1),
    )
    if image_items:
        show_onedrive_image_gallery_dialog(image_items)
    if missing_names:
        st.warning(
            "OneDrive上に見つからない画像があります："
            + "、".join(missing_names)
        )
    if other_errors:
        st.error(
            "表示できない画像があります："
            + "、".join(name for name, _ in other_errors)
            + f"（{other_errors[0][1]}）"
        )


def repair_onedrive_attachment_reference(access_token, attachment):
    """保存済みパスからOneDriveの現在のfile_idを取り直し、記録を修復する。"""
    if not isinstance(attachment, dict):
        return ""
    folder_path = clean_value(attachment.get("onedrive_path"), blank_text="").strip("/")
    stored_name = clean_value(attachment.get("stored_name"), blank_text="").strip("/")
    if not folder_path or not stored_name:
        return ""

    item = get_onedrive_path_item(
        access_token,
        f"{folder_path}/{stored_name}",
    )
    if not isinstance(item, dict) or item.get("folder"):
        return ""
    repaired_id = clean_value(item.get("id"), blank_text="")
    if not repaired_id:
        return ""

    old_id = clean_value(attachment.get("file_id"), blank_text="")
    updated = dict(attachment)
    updated["file_id"] = repaired_id
    updated["stored_name"] = clean_value(item.get("name"), blank_text="") or stored_name
    updated["web_url"] = clean_value(item.get("webUrl"), blank_text="") or updated.get(
        "web_url",
        "",
    )

    metadata_id = clean_value(updated.get("id"), blank_text="")
    field_name = clean_value(updated.get("field_name"), blank_text="")
    if metadata_id and field_name and repaired_id != old_id:
        update_customer_information(
            metadata_id,
            field_name,
            serialize_onedrive_attachment(updated),
        )
    attachment.update(updated)
    if old_id and old_id != repaired_id:
        st.session_state.pop(f"onedrive_thumbnail_{old_id}", None)
    return repaired_id


def download_onedrive_attachment_file(access_token, attachment):
    """file_idが古い時だけ保存済みパスで修復してから画像本体を読む。"""
    item_id = clean_value((attachment or {}).get("file_id"), blank_text="")
    if not item_id:
        item_id = repair_onedrive_attachment_reference(access_token, attachment)
    if not item_id:
        raise RuntimeError("OneDriveファイルIDを確認できませんでした。")
    try:
        return download_onedrive_file(access_token, item_id)
    except Exception as exc:
        if not is_onedrive_not_found_error(exc):
            raise
        repaired_id = repair_onedrive_attachment_reference(access_token, attachment)
        if not repaired_id or repaired_id == item_id:
            raise
        return download_onedrive_file(access_token, repaired_id)


def download_onedrive_attachment_thumbnail(access_token, attachment):
    """file_idが古い時だけ保存済みパスで修復してからサムネイルを読む。"""
    item_id = clean_value((attachment or {}).get("file_id"), blank_text="")
    if not item_id:
        item_id = repair_onedrive_attachment_reference(access_token, attachment)
    if not item_id:
        return None
    try:
        return download_onedrive_thumbnail(access_token, item_id)
    except Exception as exc:
        if not is_onedrive_not_found_error(exc):
            raise
        repaired_id = repair_onedrive_attachment_reference(access_token, attachment)
        if not repaired_id or repaired_id == item_id:
            raise
        return download_onedrive_thumbnail(access_token, repaired_id)


def render_onedrive_pdf_inline(pdf_content, filename):
    """OneDriveへ移動せず、画面全体を使うPDF.jsビューアーで表示する。"""
    pdf_bytes = bytes(pdf_content or b"")
    if not pdf_bytes:
        st.error("PDFデータが空のため表示できません。")
        return

    encoded = base64.b64encode(pdf_bytes).decode("ascii")
    viewer_id = "onedrive_fullscreen_pdf_" + hashlib.sha256(
        (str(filename or "") + str(len(pdf_bytes))).encode("utf-8")
    ).hexdigest()[:16]

    # components.html の小さなiframe内ではなく、親画面へ全画面ビューアーを配置する。
    # スマホではピンチ・パン・ダブルタップ、PCではホイール・ドラッグで操作できる。
    viewer_html = r"""
    <script>
    (() => {
      const parentWindow = window.parent;
      const parentDocument = parentWindow.document;
      const viewerId = __VIEWER_ID_JSON__;
      const encodedPdf = __PDF_BASE64_JSON__;
      const filename = __FILENAME_JSON__;
      const viewerGlobalKey = "__aoyamaOneDriveFullscreenViewerCleanup";
      const pdfJsSources = [
        {
          script: "https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js",
          worker: "https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js",
        },
        {
          script: "https://cdn.jsdelivr.net/npm/pdfjs-dist@3.11.174/build/pdf.min.js",
          worker: "https://cdn.jsdelivr.net/npm/pdfjs-dist@3.11.174/build/pdf.worker.min.js",
        },
        {
          script: "https://unpkg.com/pdfjs-dist@3.11.174/build/pdf.min.js",
          worker: "https://unpkg.com/pdfjs-dist@3.11.174/build/pdf.worker.min.js",
        },
      ];

      if (typeof parentWindow[viewerGlobalKey] === "function") {
        try { parentWindow[viewerGlobalKey](); } catch (_) {}
      }

      const previousViewer = parentDocument.getElementById(viewerId);
      if (previousViewer) previousViewer.remove();
      const anyViewer = parentDocument.querySelector('[data-onedrive-fullscreen-viewer="1"]');
      if (anyViewer) anyViewer.remove();

      const oldOverflow = parentDocument.body.style.overflow;
      const overlay = parentDocument.createElement("div");
      overlay.id = viewerId;
      overlay.dataset.onedriveFullscreenViewer = "1";
      overlay.setAttribute("role", "dialog");
      overlay.setAttribute("aria-modal", "true");
      overlay.setAttribute("aria-label", "PDFを全画面表示");
      overlay.innerHTML = `
        <div class="odp-stage" aria-label="PDF表示領域">
          <div class="odp-page-anchor">
            <div class="odp-page-frame">
              <canvas class="odp-canvas" aria-label="PDFページ"></canvas>
            </div>
          </div>
          <div class="odp-loading">PDFを読み込んでいます…</div>
          <div class="odp-error" hidden>
            <strong>PDFを表示できませんでした。</strong><br>
            通信状態を確認して、いったん閉じてからもう一度開いてください。
          </div>
        </div>
        <div class="odp-toolbar">
          <div class="odp-filename"></div>
          <div class="odp-controls">
            <button class="odp-prev" type="button" aria-label="前のページ">‹</button>
            <span class="odp-page-status">読込中…</span>
            <button class="odp-next" type="button" aria-label="次のページ">›</button>
            <button class="odp-zoom-out" type="button" aria-label="縮小">−</button>
            <button class="odp-fit" type="button" aria-label="幅に合わせる">幅</button>
            <button class="odp-zoom-in" type="button" aria-label="拡大">＋</button>
            <button class="odp-fullscreen" type="button" aria-label="ブラウザの全画面表示">⛶</button>
            <button class="odp-close" type="button" aria-label="閉じる">×</button>
          </div>
        </div>
        <div class="odp-help">2本指で拡大・縮小　ドラッグで移動　ダブルタップで切替</div>
      `;

      const style = parentDocument.createElement("style");
      style.textContent = `
        #${viewerId} {
          position: fixed;
          inset: 0;
          z-index: 2147483647;
          width: 100vw;
          height: 100dvh;
          overflow: hidden;
          background: #151515;
          color: #fff;
          font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
          overscroll-behavior: none;
          -webkit-user-select: none;
          user-select: none;
        }
        #${viewerId}, #${viewerId} * { box-sizing: border-box; }
        #${viewerId} .odp-stage {
          position: absolute;
          inset: 0;
          overflow: hidden;
          background: #242424;
          touch-action: none;
          cursor: grab;
        }
        #${viewerId} .odp-stage.odp-dragging { cursor: grabbing; }
        #${viewerId} .odp-page-anchor {
          position: absolute;
          top: 62px;
          left: 50%;
          transform: translateX(-50%);
          will-change: transform;
        }
        #${viewerId} .odp-page-frame {
          transform-origin: center top;
          will-change: transform;
          background: #fff;
          box-shadow: 0 4px 24px rgba(0,0,0,.55);
        }
        #${viewerId} .odp-canvas {
          display: block;
          max-width: none;
          background: #fff;
          -webkit-user-drag: none;
          user-select: none;
        }
        #${viewerId} .odp-loading,
        #${viewerId} .odp-error {
          position: absolute;
          left: 50%;
          top: 50%;
          transform: translate(-50%, -50%);
          max-width: calc(100vw - 40px);
          padding: 16px 20px;
          border-radius: 12px;
          background: rgba(20,20,20,.82);
          color: #fff;
          font-size: 15px;
          line-height: 1.65;
          text-align: center;
          pointer-events: none;
        }
        #${viewerId} .odp-error { background: rgba(100,18,18,.90); }
        #${viewerId} .odp-toolbar {
          position: absolute;
          top: max(8px, env(safe-area-inset-top));
          left: 10px;
          right: 10px;
          z-index: 4;
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 12px;
          min-height: 46px;
          pointer-events: none;
        }
        #${viewerId} .odp-filename {
          min-width: 0;
          max-width: min(42vw, 520px);
          overflow: hidden;
          text-overflow: ellipsis;
          white-space: nowrap;
          padding: 8px 12px;
          border-radius: 18px;
          background: rgba(20,20,20,.72);
          color: #fff;
          font-size: 13px;
          line-height: 1.3;
          pointer-events: auto;
        }
        #${viewerId} .odp-controls {
          display: flex;
          align-items: center;
          justify-content: flex-end;
          gap: 6px;
          min-width: 0;
          pointer-events: auto;
        }
        #${viewerId} .odp-page-status {
          min-width: 62px;
          padding: 8px 7px;
          border-radius: 18px;
          background: rgba(20,20,20,.72);
          color: #fff;
          font-size: 13px;
          font-weight: 650;
          line-height: 1;
          text-align: center;
          white-space: nowrap;
        }
        #${viewerId} button {
          width: 42px;
          height: 42px;
          flex: 0 0 42px;
          display: inline-flex;
          align-items: center;
          justify-content: center;
          border: 0;
          border-radius: 21px;
          background: rgba(35,35,35,.82);
          color: #fff;
          font-size: 25px;
          font-weight: 600;
          line-height: 1;
          cursor: pointer;
          touch-action: manipulation;
        }
        #${viewerId} button:disabled { opacity: .34; cursor: default; }
        #${viewerId} .odp-fit { font-size: 14px; font-weight: 750; }
        #${viewerId} .odp-fullscreen { font-size: 21px; }
        #${viewerId} .odp-close { font-size: 28px; }
        #${viewerId} .odp-help {
          position: absolute;
          left: 50%;
          bottom: max(14px, env(safe-area-inset-bottom));
          z-index: 4;
          transform: translateX(-50%);
          max-width: calc(100vw - 28px);
          padding: 7px 12px;
          border-radius: 16px;
          background: rgba(20,20,20,.70);
          color: #fff;
          font-size: 13px;
          line-height: 1.35;
          text-align: center;
          white-space: nowrap;
          pointer-events: none;
          opacity: 1;
          transition: opacity .35s ease;
        }
        @media (max-width: 370px) {
          #${viewerId} .odp-zoom-out,
          #${viewerId} .odp-fit,
          #${viewerId} .odp-zoom-in { display: none; }
        }
        @media (max-width: 760px) {
          #${viewerId} .odp-toolbar {
            top: max(6px, env(safe-area-inset-top));
            left: 6px;
            right: 6px;
            gap: 6px;
          }
          #${viewerId} .odp-filename { display: none; }
          #${viewerId} .odp-controls { width: 100%; justify-content: center; gap: 4px; }
          #${viewerId} button {
            width: 38px;
            height: 38px;
            flex-basis: 38px;
            border-radius: 19px;
            font-size: 22px;
          }
          #${viewerId} .odp-fit { font-size: 13px; }
          #${viewerId} .odp-fullscreen { display: none; }
          #${viewerId} .odp-page-status { min-width: 54px; padding: 7px 5px; font-size: 12px; }
          #${viewerId} .odp-page-anchor { top: 54px; }
          #${viewerId} .odp-help { font-size: 12px; }
        }
      `;
      overlay.appendChild(style);
      parentDocument.body.appendChild(overlay);
      parentDocument.body.style.overflow = "hidden";

      const stage = overlay.querySelector(".odp-stage");
      const pageAnchor = overlay.querySelector(".odp-page-anchor");
      const pageFrame = overlay.querySelector(".odp-page-frame");
      const canvas = overlay.querySelector(".odp-canvas");
      const context = canvas.getContext("2d", { alpha: false });
      const loadingNode = overlay.querySelector(".odp-loading");
      const errorNode = overlay.querySelector(".odp-error");
      const filenameNode = overlay.querySelector(".odp-filename");
      const pageStatus = overlay.querySelector(".odp-page-status");
      const prevButton = overlay.querySelector(".odp-prev");
      const nextButton = overlay.querySelector(".odp-next");
      const zoomOutButton = overlay.querySelector(".odp-zoom-out");
      const fitButton = overlay.querySelector(".odp-fit");
      const zoomInButton = overlay.querySelector(".odp-zoom-in");
      const fullscreenButton = overlay.querySelector(".odp-fullscreen");
      const closeButton = overlay.querySelector(".odp-close");
      const helpNode = overlay.querySelector(".odp-help");
      filenameNode.textContent = filename;

      let pdfDocument = null;
      let currentPage = 1;
      let renderTask = null;
      let renderSerial = 0;
      let baseWidth = 1;
      let baseHeight = 1;
      let viewScale = 1;
      let translateX = 0;
      let translateY = 0;
      let resizeTimer = null;
      let closed = false;
      let helpTimer = parentWindow.setTimeout(() => {
        helpNode.style.opacity = "0";
      }, 2600);

      const pointers = new Map();
      let lastSinglePoint = null;
      let lastPinchDistance = 0;
      let lastPinchMidpoint = null;
      let pointerMoved = false;
      let lastTapAt = 0;
      let lastTapX = 0;
      let lastTapY = 0;

      const clamp = (value, min, max) => Math.min(max, Math.max(min, value));
      const topPadding = () => (parentWindow.innerWidth <= 760 ? 54 : 62);
      const bottomPadding = () => 12;
      const sidePadding = () => (parentWindow.innerWidth <= 760 ? 6 : 12);

      function decodeBase64(base64Text) {
        const binary = atob(base64Text);
        const bytes = new Uint8Array(binary.length);
        for (let index = 0; index < binary.length; index += 1) {
          bytes[index] = binary.charCodeAt(index);
        }
        return bytes;
      }

      function midpoint(a, b) {
        return { x: (a.x + b.x) / 2, y: (a.y + b.y) / 2 };
      }

      function distance(a, b) {
        return Math.hypot(a.x - b.x, a.y - b.y);
      }

      function clampTranslation() {
        const stageWidth = Math.max(1, stage.clientWidth);
        const stageHeight = Math.max(1, stage.clientHeight);
        const scaledWidth = baseWidth * viewScale;
        const scaledHeight = baseHeight * viewScale;
        const horizontalLimit = Math.max(0, (scaledWidth - stageWidth) / 2 + sidePadding());
        const minimumY = Math.min(0, stageHeight - topPadding() - bottomPadding() - scaledHeight);
        translateX = clamp(translateX, -horizontalLimit, horizontalLimit);
        translateY = clamp(translateY, minimumY, 0);
      }

      function applyTransform() {
        clampTranslation();
        pageFrame.style.transform = `translate3d(${translateX}px, ${translateY}px, 0) scale(${viewScale})`;
      }

      function resetView() {
        viewScale = 1;
        translateX = 0;
        translateY = 0;
        applyTransform();
      }

      function zoomAt(newScale, clientX, clientY) {
        const oldScale = viewScale;
        newScale = clamp(newScale, 1, 8);
        if (Math.abs(newScale - oldScale) < 0.001) return;
        const rect = stage.getBoundingClientRect();
        const stageX = clientX - rect.left - rect.width / 2;
        const stageY = clientY - rect.top - topPadding();
        const contentX = (stageX - translateX) / oldScale;
        const contentY = (stageY - translateY) / oldScale;
        viewScale = newScale;
        translateX = stageX - contentX * newScale;
        translateY = stageY - contentY * newScale;
        if (viewScale <= 1.001) resetView();
        else applyTransform();
      }

      function setControls() {
        const ready = Boolean(pdfDocument);
        prevButton.disabled = !ready || currentPage <= 1;
        nextButton.disabled = !ready || currentPage >= pdfDocument.numPages;
        zoomOutButton.disabled = !ready || viewScale <= 1.001;
        fitButton.disabled = !ready;
        zoomInButton.disabled = !ready || viewScale >= 7.999;
        pageStatus.textContent = ready ? `${currentPage} / ${pdfDocument.numPages}` : "読込中…";
      }

      function showError(error) {
        console.error("PDF viewer error", error);
        loadingNode.hidden = true;
        pageAnchor.style.display = "none";
        errorNode.hidden = false;
        pageStatus.textContent = "表示失敗";
      }

      async function loadPdfJs() {
        if (window.pdfjsLib) {
          return { library: window.pdfjsLib, worker: pdfJsSources[0].worker };
        }
        for (const source of pdfJsSources) {
          try {
            const library = await new Promise((resolve, reject) => {
              const script = document.createElement("script");
              script.src = source.script;
              script.async = true;
              script.onload = () => {
                if (window.pdfjsLib) resolve(window.pdfjsLib);
                else reject(new Error("PDF.jsを初期化できませんでした。"));
              };
              script.onerror = () => reject(new Error("PDF.jsを読み込めませんでした。"));
              document.head.appendChild(script);
            });
            return { library, worker: source.worker };
          } catch (error) {
            console.warn("PDF.js source failed", source.script, error);
          }
        }
        throw new Error("PDF.jsを読み込めませんでした。");
      }

      async function renderPage(resetPosition = true) {
        if (!pdfDocument || closed) return;
        const serial = ++renderSerial;
        if (renderTask && typeof renderTask.cancel === "function") {
          try { renderTask.cancel(); } catch (_) {}
        }
        setControls();
        try {
          const page = await pdfDocument.getPage(currentPage);
          if (serial !== renderSerial || closed) return;
          const unitViewport = page.getViewport({ scale: 1 });
          const availableWidth = Math.max(240, stage.clientWidth - sidePadding() * 2);
          const fitScale = availableWidth / unitViewport.width;
          const viewport = page.getViewport({ scale: fitScale });
          const pixelRatio = Math.min(3, Math.max(1.5, parentWindow.devicePixelRatio || 1));

          baseWidth = viewport.width;
          baseHeight = viewport.height;
          canvas.width = Math.max(1, Math.floor(baseWidth * pixelRatio));
          canvas.height = Math.max(1, Math.floor(baseHeight * pixelRatio));
          canvas.style.width = `${Math.floor(baseWidth)}px`;
          canvas.style.height = `${Math.floor(baseHeight)}px`;
          pageAnchor.style.width = `${Math.floor(baseWidth)}px`;
          pageAnchor.style.height = `${Math.floor(baseHeight)}px`;
          pageAnchor.style.display = "block";

          renderTask = page.render({
            canvasContext: context,
            viewport,
            transform: [pixelRatio, 0, 0, pixelRatio, 0, 0],
            background: "rgb(255,255,255)",
          });
          await renderTask.promise;
          if (serial !== renderSerial || closed) return;

          loadingNode.hidden = true;
          errorNode.hidden = true;
          if (resetPosition) resetView();
          else applyTransform();
          setControls();
        } catch (error) {
          if (error && error.name === "RenderingCancelledException") return;
          showError(error);
        } finally {
          renderTask = null;
        }
      }

      function changePage(delta) {
        if (!pdfDocument) return;
        const nextPage = clamp(currentPage + delta, 1, pdfDocument.numPages);
        if (nextPage === currentPage) return;
        currentPage = nextPage;
        renderPage(true);
      }

      function closeViewer() {
        if (closed) return;
        closed = true;
        parentWindow.clearTimeout(helpTimer);
        parentWindow.clearTimeout(resizeTimer);
        renderSerial += 1;
        if (renderTask && typeof renderTask.cancel === "function") {
          try { renderTask.cancel(); } catch (_) {}
        }
        parentDocument.removeEventListener("keydown", onKeyDown, true);
        parentWindow.removeEventListener("resize", onResize, true);
        if (parentDocument.fullscreenElement === overlay && parentDocument.exitFullscreen) {
          parentDocument.exitFullscreen().catch(() => {});
        }
        parentDocument.body.style.overflow = oldOverflow;
        if (overlay.isConnected) overlay.remove();
        if (parentWindow[viewerGlobalKey] === closeViewer) {
          delete parentWindow[viewerGlobalKey];
        }
      }

      parentWindow[viewerGlobalKey] = closeViewer;

      function onKeyDown(event) {
        if (event.key === "Escape") closeViewer();
        else if (event.key === "ArrowLeft" || event.key === "PageUp") changePage(-1);
        else if (event.key === "ArrowRight" || event.key === "PageDown") changePage(1);
        else if (event.key === "+" || event.key === "=") {
          const rect = stage.getBoundingClientRect();
          zoomAt(viewScale * 1.25, rect.left + rect.width / 2, rect.top + rect.height / 2);
          setControls();
        } else if (event.key === "-") {
          const rect = stage.getBoundingClientRect();
          zoomAt(viewScale / 1.25, rect.left + rect.width / 2, rect.top + rect.height / 2);
          setControls();
        }
      }

      function onResize() {
        if (!pdfDocument || closed) return;
        parentWindow.clearTimeout(resizeTimer);
        resizeTimer = parentWindow.setTimeout(() => renderPage(true), 180);
      }

      parentDocument.addEventListener("keydown", onKeyDown, true);
      parentWindow.addEventListener("resize", onResize, true);
      closeButton.addEventListener("click", closeViewer);
      prevButton.addEventListener("click", () => changePage(-1));
      nextButton.addEventListener("click", () => changePage(1));
      fitButton.addEventListener("click", () => {
        resetView();
        setControls();
      });
      zoomInButton.addEventListener("click", () => {
        const rect = stage.getBoundingClientRect();
        zoomAt(viewScale * 1.25, rect.left + rect.width / 2, rect.top + rect.height / 2);
        setControls();
      });
      zoomOutButton.addEventListener("click", () => {
        const rect = stage.getBoundingClientRect();
        zoomAt(viewScale / 1.25, rect.left + rect.width / 2, rect.top + rect.height / 2);
        setControls();
      });
      fullscreenButton.addEventListener("click", async () => {
        try {
          if (parentDocument.fullscreenElement) await parentDocument.exitFullscreen();
          else if (overlay.requestFullscreen) await overlay.requestFullscreen();
        } catch (_) {}
      });

      stage.addEventListener("pointerdown", (event) => {
        event.preventDefault();
        stage.setPointerCapture?.(event.pointerId);
        pointers.set(event.pointerId, { x: event.clientX, y: event.clientY });
        pointerMoved = false;
        if (pointers.size === 1) {
          lastSinglePoint = { x: event.clientX, y: event.clientY };
          stage.classList.add("odp-dragging");
        } else if (pointers.size === 2) {
          const values = Array.from(pointers.values());
          lastPinchDistance = distance(values[0], values[1]);
          lastPinchMidpoint = midpoint(values[0], values[1]);
        }
      }, { passive: false });

      stage.addEventListener("pointermove", (event) => {
        if (!pointers.has(event.pointerId)) return;
        event.preventDefault();
        const previous = pointers.get(event.pointerId);
        pointers.set(event.pointerId, { x: event.clientX, y: event.clientY });
        if (Math.hypot(event.clientX - previous.x, event.clientY - previous.y) > 2) {
          pointerMoved = true;
        }

        if (pointers.size >= 2) {
          const values = Array.from(pointers.values()).slice(0, 2);
          const currentDistance = distance(values[0], values[1]);
          const currentMidpoint = midpoint(values[0], values[1]);
          if (lastPinchDistance > 0 && lastPinchMidpoint) {
            const targetScale = viewScale * (currentDistance / lastPinchDistance);
            zoomAt(targetScale, currentMidpoint.x, currentMidpoint.y);
            translateX += currentMidpoint.x - lastPinchMidpoint.x;
            translateY += currentMidpoint.y - lastPinchMidpoint.y;
            applyTransform();
            setControls();
          }
          lastPinchDistance = currentDistance;
          lastPinchMidpoint = currentMidpoint;
          lastSinglePoint = null;
        } else if (pointers.size === 1 && lastSinglePoint) {
          translateX += event.clientX - lastSinglePoint.x;
          translateY += event.clientY - lastSinglePoint.y;
          lastSinglePoint = { x: event.clientX, y: event.clientY };
          applyTransform();
        }
      }, { passive: false });

      function finishPointer(event) {
        const point = pointers.get(event.pointerId) || { x: event.clientX, y: event.clientY };
        pointers.delete(event.pointerId);
        if (pointers.size === 1) {
          const remaining = Array.from(pointers.values())[0];
          lastSinglePoint = { x: remaining.x, y: remaining.y };
          lastPinchDistance = 0;
          lastPinchMidpoint = null;
        } else if (pointers.size === 0) {
          stage.classList.remove("odp-dragging");
          lastSinglePoint = null;
          lastPinchDistance = 0;
          lastPinchMidpoint = null;
          const now = Date.now();
          if (!pointerMoved && now - lastTapAt < 320 && Math.hypot(point.x - lastTapX, point.y - lastTapY) < 36) {
            if (viewScale > 1.05) resetView();
            else zoomAt(2.5, point.x, point.y);
            setControls();
            lastTapAt = 0;
          } else if (!pointerMoved) {
            lastTapAt = now;
            lastTapX = point.x;
            lastTapY = point.y;
          }
        }
      }

      stage.addEventListener("pointerup", finishPointer, { passive: false });
      stage.addEventListener("pointercancel", finishPointer, { passive: false });

      stage.addEventListener("wheel", (event) => {
        event.preventDefault();
        const factor = event.deltaY < 0 ? 1.18 : 1 / 1.18;
        zoomAt(viewScale * factor, event.clientX, event.clientY);
        setControls();
      }, { passive: false });

      (async () => {
        try {
          const pdfBytes = decodeBase64(encodedPdf);
          const loadedPdfJs = await loadPdfJs();
          if (closed) return;
          const pdfjsLib = loadedPdfJs.library;
          pdfjsLib.GlobalWorkerOptions.workerSrc = loadedPdfJs.worker;
          pdfDocument = await pdfjsLib.getDocument({ data: pdfBytes }).promise;
          if (!pdfDocument || pdfDocument.numPages < 1) {
            throw new Error("PDFに表示できるページがありません。");
          }
          currentPage = 1;
          await renderPage(true);
        } catch (error) {
          if (!closed) showError(error);
        }
      })();
    })();
    </script>
    """
    viewer_html = viewer_html.replace("__VIEWER_ID_JSON__", json.dumps(viewer_id))
    viewer_html = viewer_html.replace("__PDF_BASE64_JSON__", json.dumps(encoded))
    viewer_html = viewer_html.replace(
        "__FILENAME_JSON__",
        json.dumps(str(filename or "PDF"), ensure_ascii=False),
    )
    components.html(
        viewer_html,
        height=1,
        width=1,
        scrolling=False,
    )

def _download_onedrive_thumbnail_uncached(access_token, item_id):
    """Streamlit状態に触れず、並行取得用に1件の軽量サムネイルを読む。"""
    safe_item_id = urllib.parse.quote(str(item_id), safe="")
    response = requests.get(
        ONEDRIVE_GRAPH_BASE
        + f"/me/drive/items/{safe_item_id}/thumbnails?$select=small,medium",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=ONEDRIVE_REQUEST_TIMEOUT,
    )
    if response.status_code != 200:
        return None
    try:
        values = list(response.json().get("value", []))
    except Exception:
        return None
    if not values:
        return None
    thumbnail_set = values[0]
    thumbnail_info = thumbnail_set.get("small") or thumbnail_set.get("medium") or {}
    url = str(thumbnail_info.get("url") or "").strip()
    if not url:
        return None
    image_response = requests.get(url, timeout=ONEDRIVE_REQUEST_TIMEOUT)
    if image_response.status_code != 200:
        return None
    return image_response.content

@st.cache_data(ttl=6 * 60 * 60, max_entries=1200, show_spinner=False)
def download_onedrive_thumbnail(_access_token, item_id):
    """一覧用の軽量サムネイルをアプリ全体で共有キャッシュする。"""
    response = onedrive_graph_request(
        "GET",
        f"/me/drive/items/{urllib.parse.quote(str(item_id), safe='')}/thumbnails?$select=small,medium",
        _access_token,
        expected=(200,),
    )
    values = list(response.json().get("value", []))
    if not values:
        return None
    thumbnail_set = values[0]
    thumbnail_info = thumbnail_set.get("small") or thumbnail_set.get("medium") or {}
    url = str(thumbnail_info.get("url") or "").strip()
    if not url:
        return None
    image_response = requests.get(url, timeout=ONEDRIVE_REQUEST_TIMEOUT)
    if image_response.status_code != 200:
        return None
    return image_response.content

@st.cache_data(ttl=6 * 60 * 60, max_entries=300, show_spinner=False)
def download_onedrive_thumbnail_batch(_access_token, item_ids):
    """顧客カルテの初回表示用に、複数サムネイルを最大4件ずつ並行取得する。"""
    unique_ids = []
    seen = set()
    for item_id in item_ids:
        clean_id = str(item_id or "").strip()
        if not clean_id or clean_id in seen:
            continue
        seen.add(clean_id)
        unique_ids.append(clean_id)
    if not unique_ids:
        return {}

    results = {}
    max_workers = min(4, len(unique_ids))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(
                _download_onedrive_thumbnail_uncached,
                _access_token,
                item_id,
            ): item_id
            for item_id in unique_ids
        }
        for future in as_completed(futures):
            item_id = futures[future]
            try:
                content = future.result()
            except Exception:
                content = None
            results[item_id] = content if isinstance(content, bytes) else None
    return results


def render_clickable_onedrive_thumbnail(
    thumbnail_content,
    filename,
    trigger_label,
    compact_height=150,
    badge_text="",
):
    """小さなサムネイル自体をタップして、既存のStreamlit表示ボタンを実行する。"""
    thumbnail_bytes = bytes(thumbnail_content or b"")
    safe_filename = str(filename or "画像")
    safe_trigger_label = str(trigger_label or "")
    safe_badge_text = str(badge_text or "").strip()
    tile_id = "onedrive_thumbnail_tile_" + hashlib.sha256(
        (safe_filename + safe_trigger_label + str(len(thumbnail_bytes))).encode("utf-8")
    ).hexdigest()[:16]

    if thumbnail_bytes:
        guessed_type = mimetypes.guess_type(safe_filename)[0] or "image/jpeg"
        if not str(guessed_type).startswith("image/"):
            guessed_type = "image/jpeg"
        encoded = base64.b64encode(thumbnail_bytes).decode("ascii")
        image_url = f"data:{guessed_type};base64,{encoded}"
        image_markup = '<img class="od-thumb-image" alt="">'
    else:
        image_url = ""
        image_markup = '<div class="od-thumb-placeholder">🖼</div>'

    components.html(
        f"""
        <div id={json.dumps(tile_id)} class="od-thumb-tile" role="button" tabindex="0"
             aria-label={json.dumps(safe_filename + 'を表示')} title={json.dumps(safe_filename)}>
          {image_markup}
          <div class="od-thumb-count"></div>
          <div class="od-thumb-hint">タップして表示</div>
        </div>
        <style>
          html, body {{ margin:0; padding:0; background:transparent; overflow:hidden; }}
          .od-thumb-tile {{
            position:relative;
            width:100%;
            height:{int(compact_height)}px;
            overflow:hidden;
            border-radius:10px;
            background:#eef1f5;
            cursor:pointer;
            touch-action:manipulation;
            box-shadow:inset 0 0 0 1px rgba(49,51,63,.12);
          }}
          .od-thumb-image {{
            display:block;
            width:100%;
            height:100%;
            object-fit:cover;
            object-position:center;
            -webkit-user-drag:none;
            user-select:none;
          }}
          .od-thumb-placeholder {{
            width:100%;
            height:100%;
            display:flex;
            align-items:center;
            justify-content:center;
            font:42px/1 system-ui, sans-serif;
          }}
          .od-thumb-count {{
            position:absolute;
            top:7px;
            right:7px;
            display:none;
            padding:4px 8px;
            border-radius:12px;
            color:#fff;
            background:rgba(0,0,0,.68);
            font:700 12px/1.2 system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            pointer-events:none;
          }}
          .od-thumb-hint {{
            position:absolute;
            left:7px;
            bottom:7px;
            padding:3px 7px;
            border-radius:10px;
            color:#fff;
            background:rgba(0,0,0,.58);
            font:11px/1.35 system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            pointer-events:none;
          }}
          .od-thumb-tile:focus-visible {{ outline:3px solid #4c8bf5; outline-offset:-3px; }}
        </style>
        <script>
        (() => {{
          const triggerLabel = {json.dumps(safe_trigger_label)};
          const badgeText = {json.dumps(safe_badge_text)};
          const tile = document.getElementById({json.dumps(tile_id)});
          const badge = tile.querySelector('.od-thumb-count');
          if (badgeText) {{
            badge.textContent = badgeText;
            badge.style.display = 'block';
          }}
          const parentDocument = window.parent.document;
          const normalizeText = (value) => String(value || '').replace(/\\s+/g, '');
          const findTrigger = () => {{
            const frame = window.frameElement;
            const frameContainer = frame && frame.closest('[data-testid="stElementContainer"]');
            let previous = frameContainer ? frameContainer.previousElementSibling : null;
            while (previous) {{
              const nearbyButton = previous.querySelector && previous.querySelector('button');
              if (nearbyButton) return nearbyButton;
              previous = previous.previousElementSibling;
            }}
            return Array.from(parentDocument.querySelectorAll('button')).find(
              (button) => normalizeText(button.textContent).includes(normalizeText(triggerLabel))
            );
          }};
          const hideTrigger = () => {{
            const trigger = findTrigger();
            if (!trigger) return null;
            const wrapper = trigger.closest('[data-testid="stButton"]');
            if (wrapper) {{
              wrapper.style.setProperty('display', 'none', 'important');
              wrapper.style.setProperty('height', '0', 'important');
              wrapper.style.setProperty('margin', '0', 'important');
              wrapper.style.setProperty('padding', '0', 'important');
            }} else {{
              trigger.style.setProperty('display', 'none', 'important');
            }}
            return trigger;
          }};
          const activate = () => {{
            const trigger = hideTrigger();
            if (trigger) trigger.click();
          }};
          hideTrigger();
          window.setTimeout(hideTrigger, 0);
          window.setTimeout(hideTrigger, 40);
          window.setTimeout(hideTrigger, 160);
          window.setTimeout(hideTrigger, 500);
          const triggerObserver = new MutationObserver(hideTrigger);
          triggerObserver.observe(parentDocument.body, {{childList:true, subtree:true}});
          window.setTimeout(() => triggerObserver.disconnect(), 3000);
          tile.addEventListener('click', activate);
          tile.addEventListener('keydown', (event) => {{
            if (event.key === 'Enter' || event.key === ' ') {{
              event.preventDefault();
              activate();
            }}
          }});
          const image = tile.querySelector('.od-thumb-image');
          if (image) image.src = {json.dumps(image_url)};
        }})();
        </script>
        """,
        height=int(compact_height) + 2,
        scrolling=False,
    )


def render_clickable_onedrive_pdf_tile(filename, trigger_label, compact_height=150):
    """PDFを写真と同じ小型カードで表示し、カード全体をタップ可能にする。"""
    safe_filename = str(filename or "PDF")
    safe_trigger_label = str(trigger_label or "")
    tile_id = "onedrive_pdf_tile_" + hashlib.sha256(
        (safe_filename + safe_trigger_label).encode("utf-8")
    ).hexdigest()[:16]
    components.html(
        f"""
        <div id={json.dumps(tile_id)} class="od-pdf-tile" role="button" tabindex="0"
             aria-label={json.dumps(safe_filename + 'を表示')} title={json.dumps(safe_filename)}>
          <div class="od-pdf-icon">PDF</div>
          <div class="od-pdf-name"></div>
          <div class="od-pdf-hint">タップして表示</div>
        </div>
        <style>
          html, body {{ margin:0; padding:0; background:transparent; overflow:hidden; }}
          .od-pdf-tile {{
            position:relative;
            width:100%;
            height:{int(compact_height)}px;
            box-sizing:border-box;
            display:flex;
            flex-direction:column;
            align-items:center;
            justify-content:center;
            gap:8px;
            padding:14px 10px 25px;
            overflow:hidden;
            border-radius:10px;
            background:#f4f5f7;
            cursor:pointer;
            touch-action:manipulation;
            box-shadow:inset 0 0 0 1px rgba(49,51,63,.12);
            font-family:system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
          }}
          .od-pdf-icon {{
            display:flex;
            align-items:center;
            justify-content:center;
            width:58px;
            height:70px;
            border-radius:6px;
            background:#fff;
            box-shadow:0 1px 5px rgba(0,0,0,.15);
            color:#b42318;
            font-weight:800;
            font-size:19px;
          }}
          .od-pdf-name {{
            width:100%;
            overflow:hidden;
            text-overflow:ellipsis;
            white-space:nowrap;
            text-align:center;
            color:#31333f;
            font-size:12px;
          }}
          .od-pdf-hint {{
            position:absolute;
            left:7px;
            bottom:7px;
            padding:3px 7px;
            border-radius:10px;
            color:#fff;
            background:rgba(0,0,0,.58);
            font-size:11px;
            line-height:1.35;
            pointer-events:none;
          }}
          .od-pdf-tile:focus-visible {{ outline:3px solid #4c8bf5; outline-offset:-3px; }}
        </style>
        <script>
        (() => {{
          const triggerLabel = {json.dumps(safe_trigger_label)};
          const tile = document.getElementById({json.dumps(tile_id)});
          tile.querySelector('.od-pdf-name').textContent = {json.dumps(safe_filename)};
          const parentDocument = window.parent.document;
          const normalizeText = (value) => String(value || '').replace(/\\s+/g, '');
          const findTrigger = () => {{
            const frame = window.frameElement;
            const frameContainer = frame && frame.closest('[data-testid="stElementContainer"]');
            let previous = frameContainer ? frameContainer.previousElementSibling : null;
            while (previous) {{
              const nearbyButton = previous.querySelector && previous.querySelector('button');
              if (nearbyButton) return nearbyButton;
              previous = previous.previousElementSibling;
            }}
            return Array.from(parentDocument.querySelectorAll('button')).find(
              (button) => normalizeText(button.textContent).includes(normalizeText(triggerLabel))
            );
          }};
          const hideTrigger = () => {{
            const trigger = findTrigger();
            if (!trigger) return null;
            const wrapper = trigger.closest('[data-testid="stButton"]');
            if (wrapper) {{
              wrapper.style.setProperty('display', 'none', 'important');
              wrapper.style.setProperty('height', '0', 'important');
              wrapper.style.setProperty('margin', '0', 'important');
              wrapper.style.setProperty('padding', '0', 'important');
            }} else {{
              trigger.style.setProperty('display', 'none', 'important');
            }}
            return trigger;
          }};
          const activate = () => {{
            const trigger = hideTrigger();
            if (trigger) trigger.click();
          }};
          hideTrigger();
          window.setTimeout(hideTrigger, 0);
          window.setTimeout(hideTrigger, 40);
          window.setTimeout(hideTrigger, 160);
          window.setTimeout(hideTrigger, 500);
          const triggerObserver = new MutationObserver(hideTrigger);
          triggerObserver.observe(parentDocument.body, {{childList:true, subtree:true}});
          window.setTimeout(() => triggerObserver.disconnect(), 3000);
          tile.addEventListener('click', activate);
          tile.addEventListener('keydown', (event) => {{
            if (event.key === 'Enter' || event.key === ' ') {{
              event.preventDefault();
              activate();
            }}
          }});
        }})();
        </script>
        """,
        height=int(compact_height) + 2,
        scrolling=False,
    )


def show_onedrive_image_gallery_dialog(image_items):
    """複数画像を、ピンチ・パン・前後移動対応の全画面ビューアーで表示する。"""
    prepared = []
    for item in list(image_items or []):
        if isinstance(item, dict):
            image_content = item.get("content")
            filename = item.get("filename")
            direct_url = str(item.get("url") or "").strip()
            original_url = str(item.get("original_url") or "").strip()
        else:
            try:
                image_content, filename = item
            except Exception:
                continue
            direct_url = ""
            original_url = ""
        safe_filename = str(filename or "画像")
        if direct_url:
            prepared.append(
                {
                    "filename": safe_filename,
                    "url": direct_url,
                    "original_url": original_url,
                    "size": 0,
                }
            )
            continue
        image_bytes = bytes(image_content or b"")
        if not image_bytes:
            continue
        guessed_type = mimetypes.guess_type(safe_filename)[0] or "image/jpeg"
        if not str(guessed_type).startswith("image/"):
            guessed_type = "image/jpeg"
        encoded = base64.b64encode(image_bytes).decode("ascii")
        prepared.append(
            {
                "filename": safe_filename,
                "url": f"data:{guessed_type};base64,{encoded}",
                "original_url": original_url,
                "size": len(image_bytes),
            }
        )

    if not prepared:
        st.error("画像データが空のため表示できません。")
        return

    viewer_signature = "|".join(
        f"{item['filename']}:{item['size']}" for item in prepared
    )
    viewer_id = "onedrive_fullscreen_image_" + hashlib.sha256(
        viewer_signature.encode("utf-8")
    ).hexdigest()[:16]
    browser_images = [
        {
            "filename": item["filename"],
            "url": item["url"],
            "originalUrl": item.get("original_url", ""),
        }
        for item in prepared
    ]

    components.html(
        f"""
        <script>
        (() => {{
          const parentWindow = window.parent;
          const parentDocument = parentWindow.document;
          const viewerId = {json.dumps(viewer_id)};
          const images = {json.dumps(browser_images, ensure_ascii=False)};

          const previousViewer = parentDocument.getElementById(viewerId);
          if (previousViewer) previousViewer.remove();
          const anyViewer = parentDocument.querySelector('[data-onedrive-fullscreen-viewer="1"]');
          if (anyViewer) anyViewer.remove();

          const oldOverflow = parentDocument.body.style.overflow;
          const overlay = parentDocument.createElement('div');
          overlay.id = viewerId;
          overlay.dataset.onedriveFullscreenViewer = '1';
          overlay.setAttribute('role', 'dialog');
          overlay.setAttribute('aria-modal', 'true');
          overlay.setAttribute('aria-label', '画像を全画面表示');
          overlay.innerHTML = `
            <div class="odv-stage">
              <img class="odv-image" alt="">
            </div>
            <div class="odv-toolbar">
              <div class="odv-filename"></div>
              <div class="odv-count"></div>
              <button class="odv-fullscreen" type="button" aria-label="全画面表示">⛶</button>
              <button class="odv-close" type="button" aria-label="閉じる">×</button>
            </div>
            <button class="odv-prev odv-nav" type="button" aria-label="前の画像">‹</button>
            <button class="odv-next odv-nav" type="button" aria-label="次の画像">›</button>
            <div class="odv-help">2本指で拡大・縮小　左右スワイプで画像を切替</div>
          `;

          const style = parentDocument.createElement('style');
          style.textContent = `
            #${{viewerId}} {{
              position: fixed;
              inset: 0;
              z-index: 2147483647;
              width: 100vw;
              height: 100dvh;
              background: #000;
              overflow: hidden;
              overscroll-behavior: none;
              touch-action: none;
              -webkit-user-select: none;
              user-select: none;
            }}
            #${{viewerId}} .odv-stage {{
              position: absolute;
              inset: 0;
              display: flex;
              align-items: center;
              justify-content: center;
              overflow: hidden;
              touch-action: none;
            }}
            #${{viewerId}} .odv-image {{
              display: block;
              max-width: 100vw;
              max-height: 100dvh;
              width: auto;
              height: auto;
              object-fit: contain;
              transform-origin: center center;
              will-change: transform;
              pointer-events: none;
              -webkit-user-drag: none;
            }}
            #${{viewerId}} .odv-toolbar {{
              position: absolute;
              top: 0;
              left: 0;
              right: 0;
              min-height: 58px;
              display: flex;
              align-items: center;
              gap: 8px;
              padding: max(8px, env(safe-area-inset-top)) 10px 8px 14px;
              box-sizing: border-box;
              color: #fff;
              background: linear-gradient(rgba(0,0,0,.78), rgba(0,0,0,0));
              font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            }}
            #${{viewerId}} .odv-filename {{
              flex: 1;
              min-width: 0;
              overflow: hidden;
              text-overflow: ellipsis;
              white-space: nowrap;
              font-size: 14px;
              text-shadow: 0 1px 2px #000;
            }}
            #${{viewerId}} .odv-count {{
              flex: 0 0 auto;
              min-width: 46px;
              text-align: center;
              font: 700 13px/1.2 system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
              text-shadow: 0 1px 2px #000;
            }}
            #${{viewerId}} button {{
              width: 44px;
              height: 44px;
              flex: 0 0 44px;
              border: 0;
              border-radius: 22px;
              background: rgba(35,35,35,.78);
              color: #fff;
              font-size: 28px;
              line-height: 1;
              cursor: pointer;
              touch-action: manipulation;
            }}
            #${{viewerId}} .odv-fullscreen {{ font-size: 22px; }}
            #${{viewerId}} .odv-nav {{
              position:absolute;
              top:50%;
              transform:translateY(-50%);
              z-index:2;
              background:rgba(20,20,20,.58);
              font-size:38px;
            }}
            #${{viewerId}} .odv-prev {{ left:10px; }}
            #${{viewerId}} .odv-next {{ right:10px; }}
            #${{viewerId}} .odv-help {{
              position: absolute;
              left: 50%;
              bottom: max(16px, env(safe-area-inset-bottom));
              transform: translateX(-50%);
              max-width: calc(100vw - 32px);
              padding: 7px 12px;
              border-radius: 16px;
              color: #fff;
              background: rgba(20,20,20,.68);
              font: 13px/1.4 system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
              white-space: nowrap;
              pointer-events: none;
              opacity: 1;
              transition: opacity .35s ease;
            }}
            @media (max-width: 640px) {{
              #${{viewerId}} .odv-nav {{
                top:auto;
                bottom:max(68px, calc(env(safe-area-inset-bottom) + 54px));
                transform:none;
              }}
              #${{viewerId}} .odv-help {{ font-size:12px; }}
            }}
          `;
          overlay.appendChild(style);
          parentDocument.body.appendChild(overlay);
          parentDocument.body.style.overflow = 'hidden';

          const stage = overlay.querySelector('.odv-stage');
          const image = overlay.querySelector('.odv-image');
          const closeButton = overlay.querySelector('.odv-close');
          const fullscreenButton = overlay.querySelector('.odv-fullscreen');
          const filenameNode = overlay.querySelector('.odv-filename');
          const countNode = overlay.querySelector('.odv-count');
          const prevButton = overlay.querySelector('.odv-prev');
          const nextButton = overlay.querySelector('.odv-next');
          const helpNode = overlay.querySelector('.odv-help');

          let currentIndex = 0;
          let scale = 1;
          let translateX = 0;
          let translateY = 0;
          let lastTapAt = 0;
          let lastTapX = 0;
          let lastTapY = 0;
          let swipeStart = null;
          let helpTimer = parentWindow.setTimeout(() => {{
            helpNode.style.opacity = '0';
          }}, 2600);
          const pointers = new Map();
          let lastSinglePoint = null;
          let lastPinchDistance = 0;
          let lastPinchMidpoint = null;
          let nextImagePreloader = null;

          const clamp = (value, min, max) => Math.min(max, Math.max(min, value));
          const applyTransform = () => {{
            image.style.transform = `translate3d(${{translateX}}px, ${{translateY}}px, 0) scale(${{scale}})`;
          }};
          const resetView = () => {{
            scale = 1;
            translateX = 0;
            translateY = 0;
            applyTransform();
          }};
          const showImage = (index) => {{
            currentIndex = (index + images.length) % images.length;
            resetView();
            const current = images[currentIndex];
            image.alt = current.filename;
            image.src = current.url;
            filenameNode.textContent = current.filename;
            countNode.textContent = images.length > 1
              ? `${{currentIndex + 1}} / ${{images.length}}`
              : '';
            const navDisplay = images.length > 1 ? '' : 'none';
            prevButton.style.display = navDisplay;
            nextButton.style.display = navDisplay;
          }};
          const preloadNextImage = () => {{
            if (nextImagePreloader) {{
              nextImagePreloader.onload = null;
              nextImagePreloader.onerror = null;
              nextImagePreloader.src = '';
              nextImagePreloader = null;
            }}
            if (images.length <= 1) return;

            const nextIndex = (currentIndex + 1) % images.length;
            const nextImage = images[nextIndex];
            if (!nextImage || !nextImage.url) return;

            const preloader = new parentWindow.Image();
            nextImagePreloader = preloader;
            const releasePreloader = () => {{
              if (nextImagePreloader === preloader) {{
                nextImagePreloader = null;
              }}
              preloader.onload = null;
              preloader.onerror = null;
            }};
            preloader.onload = releasePreloader;
            preloader.onerror = releasePreloader;
            preloader.src = nextImage.url;
          }};
          const zoomAt = (newScale, clientX, clientY) => {{
            newScale = clamp(newScale, 1, 8);
            const rect = stage.getBoundingClientRect();
            const centerX = rect.left + rect.width / 2;
            const centerY = rect.top + rect.height / 2;
            const ratio = newScale / scale;
            translateX = (clientX - centerX) - ((clientX - centerX) - translateX) * ratio;
            translateY = (clientY - centerY) - ((clientY - centerY) - translateY) * ratio;
            scale = newScale;
            if (scale <= 1.001) resetView();
            else applyTransform();
          }};
          const midpoint = (a, b) => ({{
            x: (a.x + b.x) / 2,
            y: (a.y + b.y) / 2
          }});
          const distance = (a, b) => Math.hypot(a.x - b.x, a.y - b.y);

          const closeViewer = () => {{
            parentWindow.clearTimeout(helpTimer);
            if (nextImagePreloader) {{
              nextImagePreloader.onload = null;
              nextImagePreloader.onerror = null;
              nextImagePreloader.src = '';
              nextImagePreloader = null;
            }}
            parentDocument.removeEventListener('keydown', onKeyDown, true);
            if (parentDocument.fullscreenElement === overlay && parentDocument.exitFullscreen) {{
              parentDocument.exitFullscreen().catch(() => {{}});
            }}
            parentDocument.body.style.overflow = oldOverflow;
            overlay.remove();
          }};
          const onKeyDown = (event) => {{
            if (event.key === 'Escape') closeViewer();
            if (event.key === 'ArrowLeft' && images.length > 1) showImage(currentIndex - 1);
            if (event.key === 'ArrowRight' && images.length > 1) showImage(currentIndex + 1);
          }};
          parentDocument.addEventListener('keydown', onKeyDown, true);
          closeButton.addEventListener('click', closeViewer);
          prevButton.addEventListener('click', () => showImage(currentIndex - 1));
          nextButton.addEventListener('click', () => showImage(currentIndex + 1));
          fullscreenButton.addEventListener('click', async () => {{
            try {{
              if (parentDocument.fullscreenElement) {{
                await parentDocument.exitFullscreen();
              }} else if (overlay.requestFullscreen) {{
                await overlay.requestFullscreen();
              }}
            }} catch (_) {{}}
          }});

          stage.addEventListener('pointerdown', (event) => {{
            event.preventDefault();
            stage.setPointerCapture?.(event.pointerId);
            pointers.set(event.pointerId, {{x: event.clientX, y: event.clientY}});
            if (pointers.size === 1) {{
              lastSinglePoint = {{x: event.clientX, y: event.clientY}};
              swipeStart = {{x: event.clientX, y: event.clientY, at: Date.now()}};
            }} else if (pointers.size === 2) {{
              swipeStart = null;
              const values = Array.from(pointers.values());
              lastPinchDistance = distance(values[0], values[1]);
              lastPinchMidpoint = midpoint(values[0], values[1]);
            }}
          }}, {{passive: false}});

          stage.addEventListener('pointermove', (event) => {{
            if (!pointers.has(event.pointerId)) return;
            event.preventDefault();
            pointers.set(event.pointerId, {{x: event.clientX, y: event.clientY}});
            if (pointers.size >= 2) {{
              const values = Array.from(pointers.values()).slice(0, 2);
              const currentDistance = distance(values[0], values[1]);
              const currentMidpoint = midpoint(values[0], values[1]);
              if (lastPinchDistance > 0 && lastPinchMidpoint) {{
                const targetScale = scale * (currentDistance / lastPinchDistance);
                zoomAt(targetScale, currentMidpoint.x, currentMidpoint.y);
                translateX += currentMidpoint.x - lastPinchMidpoint.x;
                translateY += currentMidpoint.y - lastPinchMidpoint.y;
                applyTransform();
              }}
              lastPinchDistance = currentDistance;
              lastPinchMidpoint = currentMidpoint;
              lastSinglePoint = null;
            }} else if (pointers.size === 1 && lastSinglePoint && scale > 1) {{
              translateX += event.clientX - lastSinglePoint.x;
              translateY += event.clientY - lastSinglePoint.y;
              lastSinglePoint = {{x: event.clientX, y: event.clientY}};
              applyTransform();
            }}
          }}, {{passive: false}});

          const finishPointer = (event) => {{
            const point = pointers.get(event.pointerId) || {{x: event.clientX, y: event.clientY}};
            pointers.delete(event.pointerId);
            if (pointers.size === 1) {{
              const remaining = Array.from(pointers.values())[0];
              lastSinglePoint = {{x: remaining.x, y: remaining.y}};
              lastPinchDistance = 0;
              lastPinchMidpoint = null;
            }} else if (pointers.size === 0) {{
              lastSinglePoint = null;
              lastPinchDistance = 0;
              lastPinchMidpoint = null;
              const dx = swipeStart ? point.x - swipeStart.x : 0;
              const dy = swipeStart ? point.y - swipeStart.y : 0;
              const elapsed = swipeStart ? Date.now() - swipeStart.at : 9999;
              if (
                scale <= 1.001 && images.length > 1 && elapsed < 900 &&
                Math.abs(dx) > 60 && Math.abs(dx) > Math.abs(dy) * 1.2
              ) {{
                showImage(currentIndex + (dx < 0 ? 1 : -1));
                lastTapAt = 0;
              }} else {{
                const now = Date.now();
                if (now - lastTapAt < 320 && Math.hypot(point.x - lastTapX, point.y - lastTapY) < 36) {{
                  if (scale > 1.05) resetView();
                  else zoomAt(2.5, point.x, point.y);
                  lastTapAt = 0;
                }} else {{
                  lastTapAt = now;
                  lastTapX = point.x;
                  lastTapY = point.y;
                }}
              }}
              swipeStart = null;
            }}
          }};
          stage.addEventListener('pointerup', finishPointer, {{passive: false}});
          stage.addEventListener('pointercancel', finishPointer, {{passive: false}});

          stage.addEventListener('wheel', (event) => {{
            event.preventDefault();
            const factor = event.deltaY < 0 ? 1.18 : 1 / 1.18;
            zoomAt(scale * factor, event.clientX, event.clientY);
          }}, {{passive: false}});

          image.addEventListener('load', () => {{
            resetView();
            preloadNextImage();
          }});
          showImage(0);
        }})();
        </script>
        """,
        height=1,
        width=1,
        scrolling=False,
    )


def show_onedrive_image_dialog(image_content, filename):
    """従来の単独画像表示を、共通の全画面ギャラリーで開く。"""
    show_onedrive_image_gallery_dialog(
        [{"content": image_content, "filename": filename}]
    )

def show_onedrive_pdf_dialog(pdf_content, filename, mime_type, metadata_id):
    """現在の画面位置を保ったまま、PDFを全画面ビューアーで表示する。"""
    render_onedrive_pdf_inline(pdf_content, filename)



GLOBAL_DELETE_SCROLL_RESTORE_KEY = "restore_global_delete_scroll_position"


def render_global_delete_scroll_keeper(restore=False):
    """画面上の削除操作前後で、現在のスクロール位置を共通保持する。"""
    storage_key = "aoyama_global_delete_scroll"
    listener_key = "__aoyama_global_delete_scroll_listener"
    restore_script = ""
    if restore:
        restore_script = f"""
          const savedY = Number(parentWindow.sessionStorage.getItem({json.dumps(storage_key)}));
          if (Number.isFinite(savedY)) {{
            const restorePosition = () => parentWindow.scrollTo({{top: savedY, left: 0, behavior: 'auto'}});
            parentWindow.requestAnimationFrame(() => {{
              restorePosition();
              parentWindow.setTimeout(restorePosition, 80);
              parentWindow.setTimeout(restorePosition, 240);
              parentWindow.setTimeout(restorePosition, 520);
            }});
            parentWindow.sessionStorage.removeItem({json.dumps(storage_key)});
          }}
        """

    components.html(
        f"""
        <script>
        (() => {{
          const parentWindow = window.parent;
          const parentDocument = parentWindow.document;
          const listenerKey = {json.dumps(listener_key)};
          const storageKey = {json.dumps(storage_key)};

          if (!parentWindow[listenerKey]) {{
            const rememberScroll = (event) => {{
              const button = event.target && event.target.closest
                ? event.target.closest('button')
                : null;
              if (!button) return;
              const label = (button.innerText || button.textContent || '')
                .replace(/\\s+/g, ' ')
                .trim();
              if (label === '削除' || label === '削除する' || label === '本当に削除') {{
                parentWindow.sessionStorage.setItem(storageKey, String(parentWindow.scrollY || 0));
              }}
            }};
            parentDocument.addEventListener('click', rememberScroll, true);
            parentWindow[listenerKey] = true;
          }}

          {restore_script}
        }})();
        </script>
        """,
        height=1,
        width=1,
        scrolling=False,
    )


def render_onedrive_attachment_scroll_keeper(suffix, restore=False):
    """写真・資料の削除前後で、現在の画面位置をブラウザー側に保持する。"""
    safe_suffix = re.sub(r"[^0-9A-Za-z_-]", "", str(suffix or ""))
    storage_key = f"onedrive_attachment_scroll_{safe_suffix}"
    listener_key = f"__onedrive_attachment_scroll_listener_{safe_suffix}"
    restore_script = ""
    if restore:
        restore_script = f"""
          const savedY = Number(parentWindow.sessionStorage.getItem({json.dumps(storage_key)}));
          if (Number.isFinite(savedY)) {{
            const restorePosition = () => parentWindow.scrollTo({{top: savedY, left: 0, behavior: 'auto'}});
            parentWindow.requestAnimationFrame(() => {{
              restorePosition();
              parentWindow.setTimeout(restorePosition, 80);
              parentWindow.setTimeout(restorePosition, 240);
            }});
            parentWindow.sessionStorage.removeItem({json.dumps(storage_key)});
          }}
        """

    components.html(
        f"""
        <script>
        (() => {{
          const parentWindow = window.parent;
          const parentDocument = parentWindow.document;
          const listenerKey = {json.dumps(listener_key)};
          const storageKey = {json.dumps(storage_key)};

          if (!parentWindow[listenerKey]) {{
            const rememberScroll = (event) => {{
              const button = event.target && event.target.closest
                ? event.target.closest('button')
                : null;
              if (!button) return;
              const label = (button.innerText || button.textContent || '')
                .replace(/\\s+/g, ' ')
                .trim();
              if (label === '削除' || label === '削除する') {{
                parentWindow.sessionStorage.setItem(storageKey, String(parentWindow.scrollY || 0));
              }}
            }};
            parentDocument.addEventListener('click', rememberScroll, true);
            parentWindow[listenerKey] = true;
          }}

          {restore_script}
        }})();
        </script>
        """,
        height=1,
        width=1,
        scrolling=False,
    )


def is_onedrive_not_found_error(exc):
    text = str(exc or "")
    return "（404）" in text or "(404)" in text or "resource could not be found" in text.lower()


def confirm_onedrive_item_deleted(access_token, item_id):
    """削除後にOneDriveのIDが存在しないことを確認する。"""
    clean_id = str(item_id or "").strip()
    if not clean_id:
        return True
    response = onedrive_graph_request(
        "GET",
        f"/me/drive/items/{urllib.parse.quote(clean_id, safe='')}?$select=id",
        access_token,
        expected=(200, 404),
    )
    return response.status_code == 404


def confirm_customer_information_deleted(item_id):
    """削除後にSupabaseの添付情報が存在しないことを確認する。"""
    clean_id = str(item_id or "").strip()
    if not clean_id:
        return True
    try:
        response = requests.get(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(),
            params={"select": "id", "id": f"eq.{clean_id}", "limit": "1"},
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("削除後のSupabase確認中に接続できませんでした。") from exc
    check_customer_information_response("削除後の確認", response, (200,))
    rows = response.json()
    if not isinstance(rows, list):
        raise RuntimeError("削除後のSupabase確認結果が正しくありません。")
    return not rows


@st.dialog("写真・資料を削除")
def confirm_onedrive_attachment_delete_dialog(
    access_token,
    attachment,
    entity_type,
    entity_id,
    entity_name,
    success_key,
    open_key,
    restore_key,
    error_key="",
):
    """OneDrive実体がない場合も含め、単独またはグループの写真・資料を削除する。"""
    attachments = list(attachment) if isinstance(attachment, (list, tuple)) else [attachment]
    attachments = [item for item in attachments if isinstance(item, dict)]
    if not attachments:
        st.error("削除対象が見つかりません。")
        return

    representative = attachments[0]
    metadata_id = representative.get("id", "")
    image_count = sum(1 for item in attachments if item.get("file_type") == "image")
    filenames = [
        str(item.get("original_name") or "名称未設定")
        for item in attachments
    ]
    if len(attachments) > 1 and image_count == len(attachments):
        target_label = f"写真{len(attachments)}枚"
    elif len(attachments) > 1:
        target_label = f"写真・資料{len(attachments)}件"
    else:
        target_label = f"「{filenames[0]}」"

    st.warning(f"{target_label}をOneDriveから削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"onedrive_attachment_delete_dialog_yes_{metadata_id}",
            type="primary",
            use_container_width=True,
        ):
            if not access_token:
                st.error("削除するにはOneDriveへ接続してください。")
            else:
                deleted_items = []
                failed_items = []
                with st.spinner("削除しています…"):
                    for item in attachments:
                        item_metadata_id = str(item.get("id") or "")
                        item_id = str(item.get("file_id") or "")
                        item_name = str(item.get("original_name") or "名称未設定")
                        item_started_at = time.perf_counter()
                        display_item_id = str(item.get("display_file_id") or "")
                        log_image_event(
                            "attachment_delete_started",
                            file_type=str(item.get("file_type") or ""),
                            has_display_file=bool(display_item_id),
                            has_original_file=bool(item_id),
                            has_metadata=bool(item_metadata_id),
                        )
                        try:
                            if display_item_id:
                                try:
                                    delete_onedrive_file(access_token, display_item_id)
                                except Exception as exc:
                                    if not is_onedrive_not_found_error(exc):
                                        raise
                                if not confirm_onedrive_item_deleted(
                                    access_token,
                                    display_item_id,
                                ):
                                    raise RuntimeError("表示用WebPがOneDriveに残っています。")
                            if item_id:
                                try:
                                    delete_onedrive_file(access_token, item_id)
                                except Exception as exc:
                                    # OneDrive上で既に消えている場合は、孤立したSupabase記録だけ削除する。
                                    if not is_onedrive_not_found_error(exc):
                                        raise
                                if not confirm_onedrive_item_deleted(access_token, item_id):
                                    raise RuntimeError("元画像・資料がOneDriveに残っています。")
                            if item_metadata_id:
                                delete_customer_information(item_metadata_id)
                                if not confirm_customer_information_deleted(item_metadata_id):
                                    raise RuntimeError("写真・資料の登録情報がSupabaseに残っています。")
                            st.session_state.pop(f"onedrive_thumbnail_{item_id}", None)
                            deleted_items.append(item)
                            log_image_event(
                                "attachment_delete_completed",
                                success=True,
                                duration_ms=round(
                                    (time.perf_counter() - item_started_at) * 1000,
                                    1,
                                ),
                            )
                        except Exception as exc:
                            log_image_event(
                                "attachment_delete_failed",
                                success=False,
                                reason=type(exc).__name__,
                                message=str(exc),
                                duration_ms=round(
                                    (time.perf_counter() - item_started_at) * 1000,
                                    1,
                                ),
                            )
                            failed_items.append((item_name, str(exc)))

                if deleted_items:
                    deleted_names = "、".join(
                        str(item.get("original_name") or "名称未設定")
                        for item in deleted_items
                    )
                    remember_change_history_warning(
                        record_change_history_safely(
                            attachment_entity_label(entity_type),
                            entity_id or "",
                            entity_name,
                            "削除",
                            {"ファイル": (deleted_names, "")},
                            section="写真・資料",
                        )
                    )
                    if len(deleted_items) > 1:
                        st.session_state[success_key] = f"写真を{len(deleted_items)}枚削除しました。"
                    else:
                        st.session_state[success_key] = "写真・資料を削除しました。"

                if failed_items:
                    failed_names = "、".join(name for name, _ in failed_items)
                    first_error = failed_items[0][1]
                    message = (
                        f"{len(failed_items)}件削除できませんでした：{failed_names}"
                        + (f"（{first_error}）" if first_error else "")
                    )
                    if error_key:
                        st.session_state[error_key] = message
                    else:
                        st.error(message)

                if deleted_items or failed_items:
                    st.session_state[open_key] = True
                    st.session_state[restore_key] = True
                    st.rerun()

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"onedrive_attachment_delete_dialog_no_{metadata_id}",
            use_container_width=True,
        ):
            st.session_state[open_key] = True
            st.session_state[restore_key] = True
            st.rerun()


# 旧URL認証情報は無視して削除する。
remove_obsolete_login_query_params()


# =========================
# Microsoftログイン認証（v53：履歴・新規ブラウザ通知・Secret期限警告）
# =========================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

private_test_login_bypass = is_private_test_login_bypass_enabled()
if private_test_login_bypass:
    microsoft_claims = {
        "is_logged_in": True,
        "sub": str(MICROSOFT_ALLOWED_SUB or "private-test-user"),
        "iss": MICROSOFT_ISSUER_PREFIX + "private-test",
        "aud": get_configured_auth_client_id(),
        "iat": int(time.time()),
        "preferred_username": "Private test app",
        "name": "Private test app",
    }
else:
    microsoft_claims = get_microsoft_user_claims()
microsoft_is_logged_in = bool(microsoft_claims.get("is_logged_in"))

# 期限切れを検知したフラグがある場合は、OIDC Cookieも削除してログイン画面へ戻す。
if st.session_state.pop("microsoft_force_logout", False):
    clear_application_login_state(revoke_current=True)
    st.logout()
    st.stop()

if not microsoft_is_logged_in:
    # v50のパスワード認証Cookieだけが残っていても、Microsoft認証なしでは入れない。
    clear_application_login_state(revoke_current=False)
    st.title("🔒 取引先カルテ")
    st.write("Microsoftアカウントで本人確認してから開きます。")
    st.caption("この画面では顧客情報・取引先情報・Dropbox・WATER itのデータを読み込みません。")

    if not LOGIN_TOKEN_SECRET:
        st.error("LOGIN_TOKEN_SECRET が設定されていません。既存のSecrets設定を確認してください。")
        st.stop()

    if st.button("Microsoftでログイン", type="primary", use_container_width=True):
        try:
            st.login()
        except Exception as exc:
            st.error("Microsoftログインを開始できませんでした。[auth]設定とrequirements.txtを確認してください。")
            st.exception(exc)
    st.stop()

identity_error = validate_microsoft_identity(microsoft_claims, require_allowed_sub=False)
if identity_error:
    clear_application_login_state(revoke_current=False)
    st.title("🔒 取引先カルテ")
    st.error("Microsoftのログイン情報を安全に確認できなかったため、アプリを表示しません。")
    st.caption(f"確認コード: {identity_error}")
    if st.button("Microsoftからログアウト", use_container_width=True):
        st.logout()
    st.stop()

microsoft_sub = str(microsoft_claims.get("sub") or "").strip()
microsoft_account_name = str(
    microsoft_claims.get("preferred_username")
    or microsoft_claims.get("email")
    or microsoft_claims.get("name")
    or "Microsoftアカウント"
).strip()

# 初回だけ、本人のsubを確認してSecretsへ登録する。登録されるまではアプリデータを表示しない。
if not MICROSOFT_ALLOWED_SUB and not private_test_login_bypass:
    clear_application_login_state(revoke_current=False)
    st.title("🔐 Microsoftログイン 初回確認")
    st.success(f"Microsoftへのログインに成功しました：{microsoft_account_name}")
    st.warning("まだ許可アカウントが登録されていないため、取引先カルテ本体は表示していません。")
    st.write("下の2行をStreamlit CloudのSecretsの一番下へ追加してください。")
    st.code(f'[app_auth]\nallowed_sub = "{microsoft_sub}"', language="toml")
    st.caption("この識別子はパスワードやトークンではありません。このMicrosoftログイン専用アプリ内で、本人のアカウントを限定するために使います。")
    if st.button("Microsoftからログアウト", use_container_width=True):
        st.logout()
    st.stop()

if not private_test_login_bypass and not hmac.compare_digest(microsoft_sub, MICROSOFT_ALLOWED_SUB):
    record_denied_microsoft_login(microsoft_claims, "account_not_allowed")
    clear_application_login_state(revoke_current=False)
    st.title("⛔ ログインできません")
    st.error("このMicrosoftアカウントは、取引先カルテの許可アカウントではありません。")
    st.caption(f"ログイン中のアカウント：{microsoft_account_name}")
    if st.button("別のMicrosoftアカウントでやり直す", type="primary", use_container_width=True):
        st.logout()
    st.stop()

# OneDrive認証から戻った場合も、許可されたMicrosoftアカウントを確認した後で処理する。
process_onedrive_callback_if_present()

cookie_login_token = get_login_token_from_cookie()
login_token_payload = validate_login_token(cookie_login_token)

if cookie_login_token and not login_token_payload:
    # 旧3時間版・期限切れ・改変済みのアプリ用Cookieだけを削除する。
    # Microsoft本人確認はすでに成功しているため、ここでMicrosoftからはログアウトしない。
    clear_login_token_cookie()
    st.session_state.pop("login_token", None)
    cookie_login_token = ""

if not login_token_payload:
    # 初回発行はMicrosoftログインから12時間以内に限定する。
    # 有効な独自トークンがある間は、Microsoftのiatが12時間を超えても利用を継続できる。
    if not is_microsoft_auth_current(microsoft_claims):
        clear_application_login_state(revoke_current=False)
        st.session_state["microsoft_force_logout"] = True
        set_query_params_safely({"page": "home", "expired": "1"})
        st.rerun()
    cookie_login_token = create_login_token()
    save_login_token_cookie(cookie_login_token)
    login_token_payload = validate_login_token(cookie_login_token)

if login_token_payload:
    # 残り1時間を切っていれば、操作中に切れないよう12時間へ更新する。
    try:
        refreshed_login_token = refresh_login_token_if_needed(cookie_login_token)
        if refreshed_login_token:
            cookie_login_token = refreshed_login_token
            login_token_payload = validate_login_token(cookie_login_token)
    except Exception:
        # 更新に失敗しても、現在のトークンが有効な間は従来どおり利用を続ける。
        pass
    st.session_state.authenticated = True
    st.session_state["login_token"] = cookie_login_token
else:
    clear_application_login_state(revoke_current=False)
    st.error("ログイン状態を安全に保存できませんでした。")
    st.logout()
    st.stop()

active_login_token = get_active_login_token()
active_login_payload = validate_login_token(active_login_token)
if not active_login_payload:
    clear_application_login_state(revoke_current=True)
    st.session_state["microsoft_force_logout"] = True
    set_query_params_safely({"page": "home", "auth_invalid": "1"})
    st.rerun()

if private_test_login_bypass:
    login_audit_warning = ""
else:
    login_audit_warning = ensure_login_audit_for_current_session(
        microsoft_claims,
        active_login_payload,
    )
if login_audit_warning:
    st.session_state["login_audit_warning"] = login_audit_warning

enforce_login_expiry()


# =========================
# 共通CSS
# =========================
st.markdown(
    """
    <style>
    :root {
        --aoyama-bg: #f6f7fb;
        --aoyama-card: rgba(255, 255, 255, 0.92);
        --aoyama-line: rgba(15, 23, 42, 0.10);
        --aoyama-text: #172033;
        --aoyama-muted: #667085;
        --aoyama-blue: #2563eb;
        --aoyama-green: #0f766e;
        --aoyama-shadow: 0 12px 32px rgba(15, 23, 42, 0.08);
    }

    .stApp {
        background:
            radial-gradient(circle at top left, rgba(37, 99, 235, 0.13), transparent 28rem),
            radial-gradient(circle at top right, rgba(15, 118, 110, 0.11), transparent 24rem),
            linear-gradient(180deg, #f8fafc 0%, #eef2f7 100%);
        color: var(--aoyama-text);
    }

    [data-testid="stHeader"] {
        background: rgba(255, 255, 255, 0.72);
        backdrop-filter: blur(12px);
        border-bottom: 1px solid rgba(15, 23, 42, 0.06);
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
    }
    [data-testid="stSidebar"] * {
        color: #f8fafc !important;
    }
    [data-testid="stSidebar"] hr {
        border-color: rgba(255, 255, 255, 0.15);
    }

    .block-container {
        padding-top: 2.2rem;
        padding-bottom: 3rem;
        max-width: 1120px;
    }

    h1, h2, h3 {
        letter-spacing: -0.03em;
    }

    .app-nav-link {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 100%;
        min-height: 3.4rem;
        box-sizing: border-box;
        text-align: center;
        text-decoration: none !important;
        padding: 0.75rem 0.9rem;
        margin: 0.32rem 0;
        border: 1px solid rgba(15, 23, 42, 0.08);
        border-radius: 16px;
        color: #172033 !important;
        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
        font-weight: 800;
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.07);
        transition: transform 0.14s ease, box-shadow 0.14s ease, border-color 0.14s ease;
    }
    .app-nav-link:hover {
        transform: translateY(-1px);
        border-color: rgba(37, 99, 235, 0.32);
        box-shadow: 0 14px 32px rgba(37, 99, 235, 0.13);
        background: linear-gradient(180deg, #ffffff 0%, #eff6ff 100%);
    }


    /* ホーム画面のホテル・宿泊先情報だけを、横幅いっぱいの長いカードにする。 */
    .hotel-home-card-link {
        min-height: 4.5rem;
        justify-content: flex-start;
        padding-left: 1.25rem;
        font-size: 1.08rem;
        background: linear-gradient(135deg, #ffffff 0%, #fff7ed 100%);
        border-color: rgba(234, 88, 12, 0.18);
    }
    .hotel-home-card-link:hover {
        border-color: rgba(234, 88, 12, 0.36);
        background: linear-gradient(135deg, #ffffff 0%, #ffedd5 100%);
        box-shadow: 0 14px 32px rgba(234, 88, 12, 0.12);
    }

    [data-testid="stSidebar"] .app-nav-link {
        justify-content: flex-start;
        min-height: 2.9rem;
        color: #f8fafc !important;
        background: rgba(255, 255, 255, 0.08);
        border-color: rgba(255, 255, 255, 0.12);
        box-shadow: none;
    }
    [data-testid="stSidebar"] .app-nav-link:hover {
        background: rgba(255, 255, 255, 0.16);
        border-color: rgba(255, 255, 255, 0.28);
        transform: none;
    }

    .stButton > button {
        border-radius: 14px !important;
        border: 1px solid rgba(15, 23, 42, 0.10) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%) !important;
        box-shadow: 0 6px 18px rgba(15, 23, 42, 0.07);
        font-weight: 700 !important;
    }
    .stButton > button,
    .stButton > button *,
    .stFormSubmitButton > button,
    .stFormSubmitButton > button * {
        color: #172033 !important;
        -webkit-text-fill-color: #172033 !important;
        opacity: 1 !important;
    }
    .stButton > button:hover {
        border-color: rgba(37, 99, 235, 0.35) !important;
        box-shadow: 0 10px 24px rgba(37, 99, 235, 0.12);
    }

    .customer-information-row {
        display: grid;
        grid-template-columns: minmax(6.5rem, 1.4fr) minmax(0, 3fr);
        align-items: start;
        column-gap: 0.9rem;
        padding: 0.35rem 0;
        color: #172033;
        font-size: 1rem;
        line-height: 1.55;
    }
    .customer-information-label,
    .customer-information-content {
        margin: 0;
        padding: 0;
        overflow-wrap: anywhere;
    }
    .customer-information-label {
        font-weight: 800;
    }

    .customer-directory {
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 0.65rem;
        margin-top: 0.75rem;
    }
    .customer-directory-item {
        display: flex;
        flex-direction: column;
        justify-content: center;
        min-height: 4.3rem;
        box-sizing: border-box;
        padding: 0.72rem 0.9rem;
        border: 1px solid rgba(15, 23, 42, 0.10);
        border-radius: 14px;
        background: rgba(255, 255, 255, 0.92);
        color: #172033 !important;
        text-decoration: none !important;
        box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
    }
    .customer-directory-item:hover {
        border-color: rgba(37, 99, 235, 0.35);
        background: #eff6ff;
    }
    .customer-directory-name {
        color: #172033;
        font-size: 1rem;
        font-weight: 800;
        line-height: 1.4;
        overflow-wrap: anywhere;
    }
    .customer-directory-meta {
        margin-top: 0.22rem;
        color: #667085;
        font-size: 0.82rem;
        line-height: 1.35;
        overflow-wrap: anywhere;
    }


    /* 顧客・仕入先・運送会社の詳細へ移動する名前リンクだけをカード表示にする。 */
    a.entity-select-card-link:not(.app-nav-link):not(.customer-directory-item) {
        display: inline-flex;
        align-items: center;
        max-width: 100%;
        box-sizing: border-box;
        padding: 0.3rem 0.58rem;
        margin: 0.08rem 0;
        border: 1px solid rgba(15, 23, 42, 0.12);
        border-radius: 10px;
        background: rgba(255, 255, 255, 0.96);
        color: #2563eb !important;
        box-shadow: 0 3px 10px rgba(15, 23, 42, 0.06);
        line-height: 1.35;
        overflow-wrap: anywhere;
        white-space: normal;
        text-decoration: none !important;
        transition: border-color 0.14s ease, background 0.14s ease, box-shadow 0.14s ease;
    }
    a.entity-select-card-link:not(.app-nav-link):not(.customer-directory-item):hover {
        border-color: rgba(37, 99, 235, 0.38);
        background: #eff6ff;
        box-shadow: 0 6px 16px rgba(37, 99, 235, 0.10);
        text-decoration: none !important;
    }
    .dispatch-name a.entity-select-card-link {
        width: 100%;
    }

    .stTextInput input {
        border-radius: 14px !important;
        border: 1px solid rgba(15, 23, 42, 0.13) !important;
        background: rgba(255,255,255,0.92) !important;
        padding: 0.72rem 0.9rem !important;
    }

    [data-testid="stMetricValue"], .stCaptionContainer {
        color: var(--aoyama-muted);
    }

    hr {
        margin: 1.4rem 0;
        border-color: rgba(15, 23, 42, 0.08);
    }

    .note-card {
        border: 1px solid rgba(15, 23, 42, 0.10);
        border-radius: 16px;
        background: rgba(255, 255, 255, 0.92);
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
        padding: 0.9rem 1rem;
        margin: 0.65rem 0;
    }
    .note-meta {
        color: #667085;
        font-size: 0.86rem;
        margin-bottom: 0.35rem;
    }
    .note-body {
        color: #172033;
        font-size: 1rem;
        line-height: 1.65;
        white-space: pre-wrap;
        word-break: break-word;
    }

    .customer-name-row {
        display: flex;
        align-items: center;
        gap: 0.45rem;
        margin: 0.1rem 0 0.45rem;
        color: #172033;
        font-size: 1.25rem;
        font-weight: 700;
    }
    .line-status {
        font-size: 0.85rem;
        font-weight: 600;
        line-height: 1;
        white-space: nowrap;
    }
    .line-status-connected {
        color: #4f8f68;
    }
    .line-status-disconnected {
        color: #98a2b3;
    }
    .customer-detail-name-row {
        font-size: 1.65rem;
        margin-top: 0.15rem;
    }
    .line-detail-static {
        color: #4f8f68;
        font-size: 0.85rem;
        font-weight: 600;
        margin-top: 0.8rem;
        white-space: nowrap;
    }
    [data-testid="stPopover"] button {
        min-height: 0 !important;
        margin-top: 0.35rem !important;
        padding: 0.3rem 0.55rem !important;
        color: #667085 !important;
        font-size: 0.85rem !important;
        box-shadow: none !important;
        white-space: nowrap !important;
    }
    [data-testid="stPopover"] button p {
        white-space: nowrap !important;
    }


    @media (max-width: 640px) {
        .block-container {
            padding-left: 0.8rem;
            padding-right: 0.8rem;
            padding-top: 1.4rem;
        }

        /* Streamlit 1.38で列が縦に崩れないようにする */
        [data-testid="stHorizontalBlock"] {
            flex-wrap: nowrap !important;
            align-items: center !important;
            gap: 0.6rem !important;
        }
        [data-testid="stHorizontalBlock"] > div {
            min-width: 0 !important;
        }

        .app-nav-link {
            min-height: 3.1rem;
            border-radius: 14px;
            font-size: 0.95rem;
        }
        h1 {
            font-size: 1.55rem !important;
        }
        h2 {
            font-size: 1.25rem !important;
        }
        h3 {
            font-size: 1.08rem !important;
        }

        .customer-information-row {
            grid-template-columns: minmax(5.5rem, 1.35fr) minmax(0, 2.65fr);
            column-gap: 0.7rem;
            align-items: baseline;
        }

        .customer-directory {
            grid-template-columns: 1fr;
            gap: 0.55rem;
        }
        .customer-directory-item {
            min-height: 3.9rem;
        }
    }
    
    /* サイドバーの更新ボタン文字を白背景でも見える色に固定 */
    [data-testid="stSidebar"] .stButton > button,
    [data-testid="stSidebar"] .stButton > button *,
    [data-testid="stSidebar"] button[kind] *,
    [data-testid="stSidebar"] button[kind] {
        color: #1f2937 !important;
        -webkit-text-fill-color: #1f2937 !important;
        opacity: 1 !important;
    }

</style>
    """,
    unsafe_allow_html=True,
)

# =========================
# 表示用の整形
# =========================
def clean_value(value, blank_text="未設定"):
    if value is None:
        return blank_text

    if isinstance(value, float) and math.isnan(value):
        return blank_text

    text = str(value).strip()

    if text == "" or text.lower() == "nan":
        return blank_text

    if text.startswith("#"):
        return blank_text

    return text


def render_customer_name_with_line(customer_name, connected, detail=False):
    """顧客名の横に控えめなLINE ○／×を表示する。"""
    line_mark = "○" if connected else "×"
    line_class = "line-status-connected" if connected else "line-status-disconnected"
    detail_class = " customer-detail-name-row" if detail else ""
    st.markdown(
        f'<div class="customer-name-row{detail_class}">'
        f'<span>👤 {html.escape(clean_value(customer_name))}</span>'
        f'<span class="line-status {line_class}">LINE {line_mark}</span>'
        "</div>",
        unsafe_allow_html=True,
    )


def format_date(value):
    if value is None:
        return "未設定"

    if isinstance(value, float) and math.isnan(value):
        return "未設定"

    text = str(value).strip()

    if text == "" or text.lower() == "nan" or text.startswith("#"):
        return "未設定"

    try:
        dt = pd.to_datetime(value)
        return f"{dt.year}/{dt.month}/{dt.day}"
    except Exception:
        return text


def to_date(value):
    if value is None:
        return None

    if isinstance(value, float) and math.isnan(value):
        return None

    text = str(value).strip()

    if text == "" or text.lower() == "nan" or text.startswith("#"):
        return None

    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def format_number(value, decimal=1, blank_text="未設定"):
    if value is None:
        return blank_text

    if isinstance(value, float) and math.isnan(value):
        return blank_text

    text = str(value).strip()

    if text == "" or text.lower() == "nan":
        return blank_text

    if text.startswith("#"):
        return "計算不可"

    try:
        num = float(value)
        if num.is_integer():
            return str(int(num))
        return f"{num:.{decimal}f}"
    except Exception:
        return text


def is_blank_or_zero(value):
    """空白・NaN・0ならTrue。使用数量/日を非表示にする判定用。

    Excelから「0」「0.0」「０」「０．０」「0 kg」のような文字列で来ても
    0として扱えるように少し広めに判定する。
    """
    if value is None:
        return True

    if isinstance(value, float) and math.isnan(value):
        return True

    text = str(value).strip()

    if text == "" or text.lower() == "nan" or text.startswith("#"):
        return True

    # 全角数字・全角小数点・カンマを整理
    normalized = text.translate(str.maketrans({
        "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
        "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
        "．": ".", "，": ",",
    })).replace(",", "")

    # 単位などが付いていても、先頭の数値だけ取り出して判定
    import re
    match = re.match(r"^[-+]?\d+(?:\.\d+)?", normalized)
    if not match:
        return False

    try:
        return float(match.group(0)) == 0
    except Exception:
        return False


def find_existing_column(df, candidates):
    """候補名の中から、Excelに存在する列名を1つ探す"""
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return None


def find_required_column_mapping(column_names):
    """Excelの列名候補を、アプリ内で使う標準列名へ対応させる"""
    normalized_columns = [str(col).strip() for col in column_names]
    mapping = {}

    for required_column, candidates in REQUIRED_COLUMN_CANDIDATES.items():
        for candidate in candidates:
            if candidate in normalized_columns:
                mapping[required_column] = candidate
                break

    return mapping


def get_first_nonblank_column_value(df, column_name):
    """指定列から最初の空でない値を取り出す"""
    if not column_name or column_name not in df.columns:
        return ""

    for value in df[column_name].tolist():
        text = clean_value(value, blank_text="")
        if text:
            return text

    return ""


def parse_lat_lng(value):
    """「緯度,経度」形式なら緯度経度を返す"""
    import re

    try:
        text = clean_value(value, blank_text="")
        if not text:
            return None

        normalized = text.translate(str.maketrans({
            "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
            "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
            "．": ".", "，": ",",
        }))
        normalized = normalized.replace("、", ",")

        match = re.match(
            r"^\s*(?:緯度\s*[:：]?\s*)?([-+]?\d+(?:\.\d+)?)\s*[, ]\s*(?:経度\s*[:：]?\s*)?([-+]?\d+(?:\.\d+)?)\s*$",
            normalized,
        )
        if not match:
            return None

        lat = float(match.group(1))
        lng = float(match.group(2))
    except Exception:
        return None

    if -90 <= lat <= 90 and -180 <= lng <= 180:
        return lat, lng

    return None


def build_google_maps_url(value):
    """住所・緯度経度・URLからGoogleマップで開くURLを作る"""
    try:
        text = clean_value(value, blank_text="")
        if not text:
            return ""

        parsed = urllib.parse.urlparse(text)
        if parsed.scheme in ("http", "https") and parsed.netloc:
            return text

        lat_lng = parse_lat_lng(text)
        if lat_lng:
            lat, lng = lat_lng
            return f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"

        query = urllib.parse.quote(text)
        return f"https://www.google.com/maps/search/?api=1&query={query}"
    except Exception:
        return ""


def get_customer_map_info(detail):
    """顧客詳細で使う住所・地図情報を取り出す"""
    try:
        map_column = find_existing_column(detail, MAP_LOCATION_COLUMN_CANDIDATES)
        address_column = find_existing_column(detail, ADDRESS_COLUMN_CANDIDATES)
        map_value = get_first_nonblank_column_value(detail, map_column)
        address_value = get_first_nonblank_column_value(detail, address_column)
        target_value = map_value or address_value

        if not target_value:
            return None

        display_value = address_value or map_value
        display_label = "住所" if address_value else "マップ位置"
        target_column = map_column if map_value else address_column

        return {
            "display_label": display_label,
            "display_value": display_value,
            "target_column": target_column,
            "map_url": build_google_maps_url(target_value),
        }
    except Exception:
        return None


def show_google_maps_button(url):
    """Googleマップを開くボタンを表示する"""
    safe_url = clean_value(url, blank_text="")
    if not safe_url:
        return

    try:
        st.link_button("📍 Googleマップ", safe_url)
    except Exception:
        st.markdown(f"[📍 Googleマップ]({safe_url})")


def find_date_column(df):
    """配車カレンダーで使う日付列を探す"""
    candidates = [
        "配車日",
        "配車予定日",
        "配達日",
        "配達予定日",
        "配送日",
        "配送予定日",
        "納品日",
        "予定日",
        "日付",
        "次回配達予定",
    ]

    found = find_existing_column(df, candidates)
    if found:
        return found

    keywords = ["配車", "配達", "配送", "納品", "予定", "日付"]

    for col in df.columns:
        col_text = str(col)
        if "数量" in col_text:
            continue

        if any(keyword in col_text for keyword in keywords):
            parsed = pd.to_datetime(df[col], errors="coerce")
            if parsed.notna().any():
                return col

    return None


# =========================
# Dropbox API
# =========================
def make_dropbox_api_arg(path_or_id):
    """
    Dropbox-API-Argヘッダー用の文字列を作る。
    日本語パスでも送れるようにASCII化してからlatin1で渡す。
    """
    return json.dumps({"path": path_or_id}, ensure_ascii=True).encode("utf-8").decode("latin1")


@st.cache_data(ttl=3300, show_spinner=False)
def request_dropbox_access_token(app_key, app_secret, refresh_token):
    """Refresh Tokenから短期アクセストークンを取得する"""
    url = "https://api.dropboxapi.com/oauth2/token"
    data = {
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
    }

    try:
        response = requests.post(
            url,
            data=data,
            auth=(app_key, app_secret),
            timeout=30,
        )
    except Exception as e:
        return None, None, str(e)

    if response.status_code != 200:
        return None, response.status_code, response.text

    return response.json().get("access_token"), None, None


def has_dropbox_auth_config():
    return bool(
        DROPBOX_APP_KEY
        or DROPBOX_APP_SECRET
        or DROPBOX_REFRESH_TOKEN
        or DROPBOX_ACCESS_TOKEN
    )


def get_dropbox_access_token():
    """
    Streamlit CloudではRefresh Token方式で短期アクセストークンを取得する。
    DROPBOX_ACCESS_TOKENは既存環境をすぐ壊さないための移行用。
    """
    if DROPBOX_APP_KEY or DROPBOX_APP_SECRET or DROPBOX_REFRESH_TOKEN:
        missing = []
        if not DROPBOX_APP_KEY:
            missing.append("DROPBOX_APP_KEY")
        if not DROPBOX_APP_SECRET:
            missing.append("DROPBOX_APP_SECRET")
        if not DROPBOX_REFRESH_TOKEN:
            missing.append("DROPBOX_REFRESH_TOKEN")

        if missing:
            st.error("Dropbox API設定が不足しています。")
            st.write("Streamlit Cloud の Secrets に以下を追加してください。")
            st.code("\n".join(missing))
            st.stop()

        access_token, status_code, error_text = request_dropbox_access_token(
            DROPBOX_APP_KEY,
            DROPBOX_APP_SECRET,
            DROPBOX_REFRESH_TOKEN,
        )

        if not access_token:
            st.error("Dropboxのアクセストークン更新に失敗しました。")
            if status_code:
                st.write(f"Dropboxからの応答コード：{status_code}")
            st.code(error_text or "access_token が返りませんでした。")
            st.stop()

        return access_token

    if DROPBOX_ACCESS_TOKEN:
        return DROPBOX_ACCESS_TOKEN

    st.error("Dropbox API設定が不足しています。")
    st.write("secrets.toml に DROPBOX_APP_KEY / DROPBOX_APP_SECRET / DROPBOX_REFRESH_TOKEN を設定してください。")
    st.stop()


def download_dropbox_file(path_or_id, access_token):
    """Dropbox APIでExcelをダウンロードする"""
    url = "https://content.dropboxapi.com/2/files/download"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Dropbox-API-Arg": make_dropbox_api_arg(path_or_id),
    }

    try:
        response = requests.post(url, headers=headers, timeout=30)
    except Exception as e:
        st.error("Dropbox APIへの接続に失敗しました。")
        st.exception(e)
        st.stop()

    if response.status_code != 200:
        return None, response

    return response.content, response


@st.cache_data(ttl=3600, show_spinner=False)
def get_dropbox_root_info(access_token):
    """チームDropboxを含むルート名前空間と、メンバーフォルダのパスを取得する。"""
    response = requests.post(
        "https://api.dropboxapi.com/2/users/get_current_account",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
        data="null",
        timeout=30,
    )
    if response.status_code != 200:
        return None, response

    root_info = response.json().get("root_info", {})
    if not root_info.get("root_namespace_id"):
        return None, response
    return root_info, response


def download_dropbox_team_file(path_or_id, access_token):
    """チームルートを明示し、メンバーフォルダ内のファイルを取得する。"""
    root_info, response = get_dropbox_root_info(access_token)
    if root_info is None:
        return None, response

    rooted_path = str(path_or_id or "").strip()
    home_path = str(root_info.get("home_path") or "").rstrip("/")
    if home_path and rooted_path.startswith("/") and not rooted_path.startswith(home_path + "/"):
        rooted_path = home_path + rooted_path

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Dropbox-API-Path-Root": json.dumps(
            {".tag": "root", "root": root_info["root_namespace_id"]},
            ensure_ascii=False,
        ),
        "Dropbox-API-Arg": make_dropbox_api_arg(rooted_path),
    }
    try:
        response = requests.post(
            "https://content.dropboxapi.com/2/files/download",
            headers=headers,
            timeout=60,
        )
    except Exception as error:
        raise RuntimeError("Dropboxのチームルートへ接続できませんでした。") from error

    if response.status_code != 200:
        return None, response
    return response.content, response


def dropbox_error_text(response):
    """Dropboxのエラーを画面表示できる文字列にする。"""
    if response is None:
        return "Dropboxから応答がありませんでした。"
    try:
        body = json.dumps(response.json(), ensure_ascii=False, indent=2)
    except Exception:
        body = response.text
    return f"HTTP {response.status_code}\n{body}"


def get_download_revision(response):
    """download応答に含まれるファイルrevを取り出す。"""
    try:
        metadata = json.loads(response.headers.get("Dropbox-API-Result", "{}"))
        return metadata.get("rev", "")
    except Exception:
        return ""


def upload_dropbox_file(path, content, access_token, mode="add", rev=""):
    """競合ファイルを作らずDropboxへファイルをアップロードする。"""
    mode_arg = {".tag": "update", "update": rev} if mode == "update" else mode
    api_arg = {
        "path": path,
        "mode": mode_arg,
        "autorename": False,
        "mute": False,
        "strict_conflict": True,
    }
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/octet-stream",
        "Dropbox-API-Arg": json.dumps(api_arg, ensure_ascii=True).encode("utf-8").decode("latin1"),
    }
    try:
        return requests.post(
            "https://content.dropboxapi.com/2/files/upload",
            headers=headers,
            data=content,
            timeout=120,
        )
    except Exception as exc:
        raise RuntimeError(f"Dropboxへのアップロードに失敗しました: {exc}") from exc


def call_dropbox_rpc(endpoint, payload, access_token):
    """DropboxのメタデータAPIを呼び出す。"""
    try:
        return requests.post(
            f"https://api.dropboxapi.com/2/{endpoint}",
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError(f"Dropbox APIへの接続に失敗しました: {exc}") from exc


def get_dropbox_response_metadata(response):
    """Return FileMetadata from an upload/copy response."""
    try:
        payload = response.json()
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    metadata = payload.get("metadata")
    return metadata if isinstance(metadata, dict) else payload


def calculate_dropbox_content_hash(content):
    """Calculate Dropbox content_hash without downloading the remote file again."""
    block_size = 4 * 1024 * 1024
    combined = hashlib.sha256()
    for offset in range(0, len(content), block_size):
        combined.update(hashlib.sha256(content[offset:offset + block_size]).digest())
    return combined.hexdigest()


def verify_dropbox_file_metadata(metadata, expected_content, previous_revision=""):
    """Verify size, content hash, and revision from Dropbox FileMetadata."""
    if not isinstance(metadata, dict):
        raise RuntimeError("Dropbox\u306e\u4fdd\u5b58\u7d50\u679c\u3092\u78ba\u8a8d\u3067\u304d\u307e\u305b\u3093\u3067\u3057\u305f\u3002")

    expected_size = len(expected_content)
    expected_hash = calculate_dropbox_content_hash(expected_content)
    remote_size = metadata.get("size")
    remote_hash = str(metadata.get("content_hash") or "")
    remote_revision = str(metadata.get("rev") or "")

    try:
        size_matches = int(remote_size) == expected_size
    except Exception:
        size_matches = False
    if not size_matches:
        raise RuntimeError("Dropbox\u4fdd\u5b58\u5f8c\u306e\u30d5\u30a1\u30a4\u30eb\u30b5\u30a4\u30ba\u304c\u4e00\u81f4\u3057\u307e\u305b\u3093\u3002")
    if not remote_hash or remote_hash != expected_hash:
        raise RuntimeError("Dropbox\u4fdd\u5b58\u5f8c\u306e\u30d5\u30a1\u30a4\u30eb\u5185\u5bb9\u304c\u4e00\u81f4\u3057\u307e\u305b\u3093\u3002")
    if not remote_revision:
        raise RuntimeError("Dropbox\u4fdd\u5b58\u5f8c\u306erev\u3092\u78ba\u8a8d\u3067\u304d\u307e\u305b\u3093\u3002")
    if previous_revision and remote_revision == previous_revision:
        raise RuntimeError("Dropbox\u306e\u66f4\u65b0\u756a\u53f7\u304c\u5909\u308f\u3063\u3066\u3044\u306a\u3044\u305f\u3081\u3001\u4fdd\u5b58\u3092\u5b8c\u4e86\u3067\u304d\u307e\u305b\u3093\u3002")
    return remote_revision


def get_dropbox_file_metadata(path, access_token):
    """Fetch only metadata; this does not download the Excel bytes."""
    response = call_dropbox_rpc("files/get_metadata", {"path": path}, access_token)
    if response.status_code != 200:
        raise RuntimeError(
            "Dropbox\u306e\u30d5\u30a1\u30a4\u30eb\u60c5\u5831\u3092\u53d6\u5f97\u3067\u304d\u307e\u305b\u3093\u3067\u3057\u305f\u3002\n"
            + dropbox_error_text(response)
        )
    metadata = get_dropbox_response_metadata(response)
    if not metadata:
        raise RuntimeError("Dropbox\u306e\u30d5\u30a1\u30a4\u30eb\u60c5\u5831\u3092\u8aad\u307f\u53d6\u308c\u307e\u305b\u3093\u3067\u3057\u305f\u3002")
    return metadata


def copy_dropbox_file(from_path, to_path, access_token):
    """Create a server-side Dropbox copy without re-uploading the Excel bytes."""
    return call_dropbox_rpc(
        "files/copy_v2",
        {
            "from_path": from_path,
            "to_path": to_path,
            "allow_shared_folder": False,
            "autorename": False,
            "allow_ownership_transfer": False,
        },
        access_token,
    )


def get_dropbox_revision(path, access_token):
    """Dropboxファイルの現在のrevだけを軽量に取得する。"""
    response = call_dropbox_rpc("files/get_metadata", {"path": path}, access_token)
    if response.status_code != 200:
        raise RuntimeError("Dropboxのファイル情報を取得できませんでした。\n" + dropbox_error_text(response))
    return str(response.json().get("rev", ""))


def ensure_dropbox_backup_folder(access_token):
    """Backupsフォルダがなければ作成する。既に存在する場合は成功扱いにする。"""
    response = call_dropbox_rpc(
        "files/create_folder_v2",
        {"path": DROPBOX_BACKUP_FOLDER, "autorename": False},
        access_token,
    )
    if response.status_code == 200:
        return
    if response.status_code == 409:
        try:
            error_data = response.json()
            summary = str(error_data.get("error_summary", ""))
            # path/conflict/folder/ は「同名フォルダが既にある」という正常状態。
            if "conflict" in summary and "folder" in summary:
                return
        except Exception:
            pass
    raise RuntimeError(
        "Dropboxにバックアップフォルダを作成できませんでした。\n"
        + dropbox_error_text(response)
    )


def create_dropbox_backup(target_path, backup_path, original_content, access_token):
    """Create and verify the pre-save backup, preferring a fast server-side copy."""
    ensure_dropbox_backup_folder(access_token)
    copy_response = copy_dropbox_file(target_path, backup_path, access_token)

    if copy_response.status_code == 200:
        metadata = get_dropbox_response_metadata(copy_response)
        if not metadata.get("content_hash") or metadata.get("size") is None:
            metadata = get_dropbox_file_metadata(backup_path, access_token)
        try:
            verify_dropbox_file_metadata(metadata, original_content)
            return
        except Exception:
            call_dropbox_rpc("files/delete_v2", {"path": backup_path}, access_token)
            raise RuntimeError(
                "PC\u307e\u305f\u306f\u5225\u7aef\u672b\u3067Excel\u304c\u66f4\u65b0\u3055\u308c\u305f\u53ef\u80fd\u6027\u304c\u3042\u308a\u307e\u3059\u3002"
                "\u518d\u8aad\u307f\u8fbc\u307f\u3057\u3066\u304b\u3089\u3084\u308a\u76f4\u3057\u3066\u304f\u3060\u3055\u3044\u3002"
            )

    # Fallback keeps the previous behavior if server-side copy is unavailable.
    backup_response = upload_dropbox_file(
        backup_path,
        original_content,
        access_token,
        mode="add",
    )
    if backup_response.status_code != 200:
        raise RuntimeError(
            "\u30d0\u30c3\u30af\u30a2\u30c3\u30d7\u3092\u4f5c\u6210\u3067\u304d\u306a\u3044\u305f\u3081\u3001\u672c\u756a\u30d5\u30a1\u30a4\u30eb\u306f\u66f4\u65b0\u3057\u307e\u305b\u3093\u3002\n"
            + dropbox_error_text(backup_response)
        )
    metadata = get_dropbox_response_metadata(backup_response)
    if not metadata.get("content_hash") or metadata.get("size") is None:
        metadata = get_dropbox_file_metadata(backup_path, access_token)
    verify_dropbox_file_metadata(metadata, original_content)


def trim_old_dropbox_backups(access_token, keep=30):
    """対象ブックのバックアップを新しい順にkeep件だけ残す。"""
    entries = []
    response = call_dropbox_rpc(
        "files/list_folder",
        {"path": DROPBOX_BACKUP_FOLDER, "recursive": False, "include_deleted": False},
        access_token,
    )
    if response.status_code != 200:
        return f"バックアップ一覧を取得できませんでした。\n{dropbox_error_text(response)}"

    data = response.json()
    entries.extend(data.get("entries", []))
    while data.get("has_more"):
        response = call_dropbox_rpc("files/list_folder/continue", {"cursor": data["cursor"]}, access_token)
        if response.status_code != 200:
            return f"バックアップ一覧の続きが取得できませんでした。\n{dropbox_error_text(response)}"
        data = response.json()
        entries.extend(data.get("entries", []))

    pattern = re.compile(r"^配車予定 次郎_\d{8}_\d{6}(?:_\d+)?\.xlsm$")
    backups = [item for item in entries if item.get(".tag") == "file" and pattern.match(item.get("name", ""))]
    backups.sort(key=lambda item: (item.get("server_modified", ""), item.get("name", "")), reverse=True)
    warnings = []
    for item in backups[keep:]:
        delete_response = call_dropbox_rpc("files/delete_v2", {"path": item["path_lower"]}, access_token)
        if delete_response.status_code != 200:
            warnings.append(item.get("name", "不明なファイル"))
    return "削除できなかった古いバックアップ: " + ", ".join(warnings) if warnings else ""


def normalize_match_value(value):
    return clean_value(value, blank_text="").strip()


def find_sheet1_customer_rows(workbook, customer_name):
    """
    Sheet1のB列は顧客名の値ではなく、例: =次回配達日!B7 の数式。
    数式の参照先（次回配達日シートB列）をたどって対象顧客の行を返す。
    値が直接入っている場合にも対応する。
    """
    if SHEET_NAME not in workbook.sheetnames or DELIVERY_SHEET_NAME not in workbook.sheetnames:
        return []

    sheet1 = workbook[SHEET_NAME]
    delivery_ws = workbook[DELIVERY_SHEET_NAME]
    target = normalize_match_value(customer_name)
    rows = []
    formula_pattern = re.compile(
        r"^=\s*(?:'次回配達日'|次回配達日)!\$?B\$?(\d+)\s*$",
        re.IGNORECASE,
    )

    for row in range(2, sheet1.max_row + 1):
        value = sheet1.cell(row, SHEET1_CUSTOMER_COLUMN).value

        # 顧客名が直接入っている場合
        if normalize_match_value(value) == target:
            rows.append(row)
            continue

        # =次回配達日!B7 のような数式の場合
        if isinstance(value, str):
            match = formula_pattern.match(value.strip())
            if match:
                source_row = int(match.group(1))
                source_customer = normalize_match_value(
                    delivery_ws.cell(source_row, 2).value
                )
                if source_customer == target:
                    rows.append(row)

    return rows


def find_header_column_in_worksheet(ws, candidates, max_rows=50):
    """見出し行を走査し、候補名に完全一致する列番号を返す。"""
    candidate_set = {str(item).strip() for item in candidates}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows)):
        for cell in row:
            if normalize_match_value(cell.value) in candidate_set:
                return cell.column
    return None


def find_product_rows_by_usage(ws, customer_name, product_name):
    """同一顧客・同一商品の行を、使用中行と過去行に分けて返す。"""
    matched_rows = [
        row for row in range(1, ws.max_row + 1)
        if normalize_match_value(ws.cell(row, 2).value) == customer_name
        and normalize_match_value(ws.cell(row, 5).value) == product_name
    ]
    active_rows = [
        row for row in matched_rows
        if not is_blank_or_zero(ws.cell(row, 7).value)
    ]
    inactive_rows = [row for row in matched_rows if row not in active_rows]
    return matched_rows, active_rows, inactive_rows


@st.cache_data(ttl=60, show_spinner=False)
def read_edit_values_from_bytes(content, customer_name, product_name):
    """最新ブックから編集欄の現在値を取得する。"""
    workbook = load_workbook(BytesIO(content), keep_vba=True, data_only=False, read_only=False)
    try:
        if DELIVERY_SHEET_NAME not in workbook.sheetnames or SHEET_NAME not in workbook.sheetnames:
            raise ValueError("必要なシート（次回配達日 または Sheet1）が見つかりません。")
        delivery_ws = workbook[DELIVERY_SHEET_NAME]
        matches, active_rows, _ = find_product_rows_by_usage(
            delivery_ws, customer_name, product_name
        )
        product_values = {}
        if len(active_rows) == 1:
            row = active_rows[0]
            product_values = {
                "メーカー": delivery_ws.cell(row, 6).value,
                "在庫本数": delivery_ws.cell(row, 8).value,
                "本数": delivery_ws.cell(row, 9).value,
                "kg/本": delivery_ws.cell(row, 10).value,
                "配達日": delivery_ws.cell(row, 11).value,
            }

        customer_ws = workbook[SHEET_NAME]
        customer_rows = find_sheet1_customer_rows(workbook, customer_name)
        first_row = customer_rows[0] if customer_rows else None
        return {
            **product_values,
            "住所": customer_ws.cell(first_row, SHEET1_ADDRESS_COLUMN).value if first_row else None,
            "マップ位置": customer_ws.cell(first_row, SHEET1_MAP_COLUMN).value if first_row else None,
            "商品一致件数": len(active_rows),
            "商品全行件数": len(matches),
            "顧客一致件数": len(customer_rows),
        }
    finally:
        workbook.close()


@st.cache_data(ttl=60, show_spinner=False)
def read_customer_edit_bundle_from_bytes(content, customer_name):
    """顧客詳細に必要な地図と全商品の編集値を、Excelを1回開いてまとめて読む。"""
    workbook = load_workbook(BytesIO(content), keep_vba=True, data_only=False, read_only=False)
    try:
        if DELIVERY_SHEET_NAME not in workbook.sheetnames or SHEET_NAME not in workbook.sheetnames:
            raise ValueError("必要なシート（次回配達日 または Sheet1）が見つかりません。")

        delivery_ws = workbook[DELIVERY_SHEET_NAME]
        target = normalize_match_value(customer_name)
        product_rows = {}
        for row in range(1, delivery_ws.max_row + 1):
            if normalize_match_value(delivery_ws.cell(row, 2).value) != target:
                continue
            product = normalize_match_value(delivery_ws.cell(row, 5).value)
            if product:
                product_rows.setdefault(product, []).append(row)

        products = {}
        for product, rows in product_rows.items():
            active_rows = [
                row for row in rows
                if not is_blank_or_zero(delivery_ws.cell(row, 7).value)
            ]
            selected_row = active_rows[0] if len(active_rows) == 1 else None
            products[product] = {
                "メーカー": delivery_ws.cell(selected_row, 6).value if selected_row else None,
                "在庫本数": delivery_ws.cell(selected_row, 8).value if selected_row else None,
                "本数": delivery_ws.cell(selected_row, 9).value if selected_row else None,
                "kg/本": delivery_ws.cell(selected_row, 10).value if selected_row else None,
                "配達日": delivery_ws.cell(selected_row, 11).value if selected_row else None,
                "商品一致件数": len(active_rows),
                "商品全行件数": len(rows),
            }

        customer_rows = find_sheet1_customer_rows(workbook, customer_name)
        first_customer_row = customer_rows[0] if customer_rows else None
        ws = workbook[SHEET_NAME]
        map_values = {
            "住所": ws.cell(first_customer_row, SHEET1_ADDRESS_COLUMN).value if first_customer_row else None,
            "マップ位置": ws.cell(first_customer_row, SHEET1_MAP_COLUMN).value if first_customer_row else None,
            "顧客一致件数": len(customer_rows),
        }
        for values in products.values():
            values.update(map_values)
        return {"map": map_values, "products": products}
    finally:
        workbook.close()


def delivery_record_fingerprint(source, row_number, values):
    """履歴修正中の別端末更新を検知するため、A:Nの現在値から指紋を作る。"""
    normalized = [change_history_value(value) for value in list(values or [])[:14]]
    payload = json.dumps(
        {
            "source": str(source or ""),
            "row_number": int(row_number),
            "values": normalized,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


@st.cache_data(ttl=60, show_spinner=False)
def read_product_delivery_history_from_bytes(
    content,
    customer_name,
    product_name,
    limit=20,
):
    """
    配達履歴と次回配達日シートの現在行を、修正用の一覧として返す。

    画面に必要な新しい順の件数だけを保持し、古い履歴すべての辞書化・指紋計算・
    並べ替えは行わない。履歴自体は削除せず、表示順と修正ルールも従来どおり。
    """
    try:
        requested_limit = max(1, int(limit))
    except (TypeError, ValueError):
        requested_limit = 20

    # 読み取り専用なのでVBA本体を展開しない。元ファイルやマクロには一切変更を加えない。
    workbook = load_workbook(
        BytesIO(content),
        keep_vba=False,
        data_only=False,
        read_only=True,
    )
    try:
        if DELIVERY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("次回配達日シートが見つかりません。")
        if DELIVERY_HISTORY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("配達履歴シートが見つかりません。")

        target_customer = normalize_match_value(customer_name)
        target_product = normalize_match_value(product_name)
        delivery_ws = workbook[DELIVERY_SHEET_NAME]
        active_rows = []
        for row_number, values in enumerate(
            delivery_ws.iter_rows(min_row=1, max_col=14, values_only=True),
            start=1,
        ):
            values = tuple(values or ())
            row_customer = normalize_match_value(values[1] if len(values) >= 2 else None)
            row_product = normalize_match_value(values[4] if len(values) >= 5 else None)
            if row_customer != target_customer or row_product != target_product:
                continue
            usage = values[6] if len(values) >= 7 else None
            if not is_blank_or_zero(usage):
                active_rows.append((row_number, values))

        if len(active_rows) > 1:
            raise ValueError("同じ顧客名・商品名の使用中行が複数見つかりました。確認してください。")
        if not active_rows:
            raise ValueError("使用中の商品行が見つからないため、納品履歴を表示できません。")

        current_row, current_values = active_rows[0]
        target_identity = delivery_history_identity(current_values)
        if target_identity is None:
            raise ValueError("顧客・商品を履歴と結び付ける情報が見つかりません。")

        def candidate_sort_key(source, row_number, values):
            event_date = to_date(values[10] if len(values) >= 11 else None) or date.min
            return (
                1 if source == "current" else 0,
                event_date,
                int(row_number),
            )

        # 候補は画面に必要な件数だけ保持する。全履歴件数は「さらに表示」のため数える。
        candidates = [
            {
                "source": "current",
                "row_number": current_row,
                "values": current_values,
                "sort_key": candidate_sort_key("current", current_row, current_values),
            }
        ]
        total_count = 1

        history_ws = workbook[DELIVERY_HISTORY_SHEET_NAME]
        for row_number, values in enumerate(
            history_ws.iter_rows(min_row=2, max_col=14, values_only=True),
            start=2,
        ):
            values = tuple(values or ())
            if delivery_history_identity(values) != target_identity:
                continue

            total_count += 1
            candidate = {
                "source": "history",
                "row_number": row_number,
                "values": values,
                "sort_key": candidate_sort_key("history", row_number, values),
            }
            if len(candidates) < requested_limit:
                candidates.append(candidate)
                continue

            oldest_index = min(
                range(len(candidates)),
                key=lambda index: candidates[index]["sort_key"],
            )
            if candidate["sort_key"] > candidates[oldest_index]["sort_key"]:
                candidates[oldest_index] = candidate

        candidates.sort(key=lambda item: item["sort_key"], reverse=True)

        records = []
        for candidate in candidates:
            source = candidate["source"]
            row_number = candidate["row_number"]
            values = candidate["values"]
            next_delivery = (
                calculate_delivery_values(values)[0]
                if source == "current"
                else (values[12] if len(values) >= 13 else None)
            )
            records.append(
                {
                    "record_key": f"{source}:{row_number}",
                    "source": source,
                    "row_number": row_number,
                    "identity": target_identity,
                    "fingerprint": delivery_record_fingerprint(
                        source,
                        row_number,
                        values,
                    ),
                    "メーカー": values[5] if len(values) >= 6 else None,
                    "在庫本数": values[7] if len(values) >= 8 else None,
                    "本数": values[8] if len(values) >= 9 else None,
                    "kg/本": values[9] if len(values) >= 10 else None,
                    "配達日": values[10] if len(values) >= 11 else None,
                    "配達数量": values[11] if len(values) >= 12 else None,
                    "次回配達予定": next_delivery,
                }
            )

        return {
            "records": records,
            "total_count": total_count,
        }
    finally:
        workbook.close()


def parse_optional_nonnegative_number(text, integer=False):
    value = str(text).strip().translate(str.maketrans("０１２３４５６７８９．，", "0123456789.,"))
    # 音声入力で付きやすい単位を許可する。
    value = re.sub(r"\s*(?:本|kg|KG|ｋｇ|キロ|キログラム)\s*$", "", value, flags=re.IGNORECASE)
    value = value.replace(",", "")
    if value == "":
        return None
    try:
        number = float(value)
    except Exception as exc:
        raise ValueError("数値で入力してください。") from exc
    if not math.isfinite(number) or number < 0:
        raise ValueError("0以上の数値で入力してください。")
    if integer and not number.is_integer():
        raise ValueError("整数で入力してください。")
    return int(number) if integer else number


def validate_map_location(value):
    text = str(value).strip()
    if not text:
        return ""
    parsed = urllib.parse.urlparse(text)
    if parsed.scheme or "://" in text:
        if parsed.scheme not in ("http", "https") or not parsed.netloc:
            raise ValueError("マップ位置のURLが正しくありません。")
    # 数字と区切りだけで座標らしく見える入力に限って、緯度経度として検証する。
    # 通常の住所・施設名にカンマが含まれていても許可する。
    coordinate_like = bool(re.match(r"^[\s+\-\d０-９.,，．、:：緯度経]+$", text))
    if coordinate_like and not parse_lat_lng(text):
        raise ValueError("緯度,経度は例のように入力してください（43.123456, 143.123456）。")
    return text


def is_blank_excel_value(value):
    """Excel由来のNone・空文字・pandasのNaN/NaTを同じ空欄として扱う。"""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def normalize_existing_excel_value(value):
    """編集欄が未入力のとき、pandasのNaN/NaTをExcelへ書き戻さない。"""
    return None if is_blank_excel_value(value) else value


def same_excel_value(old, new):
    if is_blank_excel_value(old) and is_blank_excel_value(new):
        return True
    if isinstance(old, (datetime, date)) and isinstance(new, (datetime, date)):
        # datetime と date の組み合わせでも、date 側へ .date() を呼ばない。
        # 例：Excelの datetime(8/6) をアプリの date(8/3)へ直す場合も安全に比較する。
        old_date = old.date() if isinstance(old, datetime) else old
        new_date = new.date() if isinstance(new, datetime) else new
        return old_date == new_date
    return old == new


def enable_excel_recalculation(workbook):
    """Excelで開いたときに残数・次回配達予定などの数式を必ず再計算させる。"""
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        # openpyxlの版によって属性名が異なる場合でも、セル保存は続行する。
        pass


def verify_remote_workbook_changes(content, changed_cells):
    """Dropboxから再取得したブックに、変更値が実在することを検証する。"""
    workbook = load_workbook(BytesIO(content), keep_vba=True, data_only=False, read_only=False)
    try:
        for sheet, row, column, expected in changed_cells:
            if sheet not in workbook.sheetnames:
                raise RuntimeError(f"Dropbox保存後の確認でシート「{sheet}」が見つかりません。")
            cell = workbook[sheet].cell(row, column)
            if not same_excel_value(cell.value, expected):
                raise RuntimeError(
                    f"Dropbox保存後の確認で {sheet}!{cell.coordinate} が更新されていません。"
                )
    finally:
        workbook.close()


def confirm_dropbox_upload(target_path, access_token, changed_cells):
    """アップロード後にDropbox本体を読み直し、保存完了を保証する。"""
    uploaded_content, response = download_dropbox_file(target_path, access_token)
    if uploaded_content is None:
        raise RuntimeError(
            "Dropboxへ送信後、更新済みExcelを再取得できませんでした。\n"
            + dropbox_error_text(response)
        )
    verify_remote_workbook_changes(uploaded_content, changed_cells)
    return uploaded_content, get_download_revision(response)


def _xlsx_column_number_from_reference(cell_reference):
    """A1形式のセル参照から列番号だけを返す。"""
    match = re.match(r"^([A-Za-z]+)", str(cell_reference or ""))
    if not match:
        return 0
    number = 0
    for char in match.group(1).upper():
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _xlsx_cell_reference(row_number, column_number):
    """行・列番号をA1形式へ変換する。"""
    column_number = int(column_number)
    letters = []
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters)) + str(int(row_number))


def _xlsx_workbook_info(archive):
    """xlsm ZIP内のシート順・XMLパス・日付基準を取得する。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rels_root.findall(f"{{{package_rel_ns}}}Relationship")
    }

    sheet_names = []
    sheet_paths = {}
    sheets = workbook_root.find(f"{{{main_ns}}}sheets")
    if sheets is not None:
        for sheet in sheets:
            name = str(sheet.attrib.get("name") or "")
            relation_id = str(sheet.attrib.get(f"{{{office_rel_ns}}}id") or "")
            target = str(rel_targets.get(relation_id) or "").replace("\\", "/")
            if not name or not target:
                continue
            if target.startswith("/"):
                path = target.lstrip("/")
            elif target.startswith("xl/"):
                path = posixpath.normpath(target)
            else:
                path = posixpath.normpath(posixpath.join("xl", target))
            sheet_names.append(name)
            sheet_paths[name] = path

    workbook_pr = workbook_root.find(f"{{{main_ns}}}workbookPr")
    date1904 = False
    if workbook_pr is not None:
        date1904 = str(workbook_pr.attrib.get("date1904") or "").strip().lower() in {
            "1",
            "true",
        }
    return sheet_names, sheet_paths, date1904


def _xlsx_shared_strings(archive):
    """sharedStrings.xmlがあるブックだけ共有文字列を読む。openpyxl保存後は通常空。"""
    path = "xl/sharedStrings.xml"
    if path not in archive.namelist():
        return []
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(archive.read(path))
    values = []
    for item in root.findall(f"{{{main_ns}}}si"):
        values.append("".join(node.text or "" for node in item.iter(f"{{{main_ns}}}t")))
    return values


def _xlsx_excel_datetime(number, date1904=False):
    """Excelシリアル値をdatetimeへ変換する。"""
    epoch = datetime(1904, 1, 1) if date1904 else datetime(1899, 12, 30)
    return epoch + timedelta(days=float(number))


def _xlsx_cell_value(cell, shared_strings, date1904=False, force_date=False):
    """openpyxlを起動せず、worksheet XMLの1セルをPython値へ戻す。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    formula = cell.find(f"{{{main_ns}}}f")
    if formula is not None:
        return "=" + str(formula.text or "")

    cell_type = str(cell.attrib.get("t") or "")
    if cell_type == "inlineStr":
        inline = cell.find(f"{{{main_ns}}}is")
        if inline is None:
            return ""
        return "".join(node.text or "" for node in inline.iter(f"{{{main_ns}}}t"))

    value_node = cell.find(f"{{{main_ns}}}v")
    value_text = value_node.text if value_node is not None else None
    if value_text is None:
        return None

    if cell_type == "s":
        try:
            return shared_strings[int(value_text)]
        except Exception:
            return None
    if cell_type == "b":
        return str(value_text).strip() == "1"
    if cell_type in {"str", "e"}:
        return str(value_text)
    if cell_type == "d":
        try:
            return datetime.fromisoformat(str(value_text).replace("Z", "+00:00"))
        except Exception:
            return str(value_text)

    try:
        number = float(value_text)
    except Exception:
        return str(value_text)
    if force_date:
        try:
            return _xlsx_excel_datetime(number, date1904=date1904)
        except Exception:
            return number
    return int(number) if number.is_integer() else number


def _xlsx_iter_sheet_rows(
    archive,
    sheet_path,
    shared_strings,
    max_col,
    date1904=False,
    date_columns=None,
):
    """worksheet XMLを1回だけ流し読みし、必要列だけタプルで返す。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    row_tag = f"{{{main_ns}}}row"
    cell_tag = f"{{{main_ns}}}c"
    date_columns = set(date_columns or ())

    with archive.open(sheet_path) as stream:
        for _, element in ET.iterparse(stream, events=("end",)):
            if element.tag != row_tag:
                continue
            try:
                row_number = int(element.attrib.get("r") or 0)
            except Exception:
                row_number = 0
            values = [None] * int(max_col)
            for cell in element:
                if cell.tag != cell_tag:
                    continue
                column_number = _xlsx_column_number_from_reference(cell.attrib.get("r"))
                if column_number < 1 or column_number > max_col:
                    continue
                values[column_number - 1] = _xlsx_cell_value(
                    cell,
                    shared_strings,
                    date1904=date1904,
                    force_date=column_number in date_columns,
                )
            element.clear()
            if row_number > 0:
                yield row_number, tuple(values)


def _xlsx_verify_changed_cells(
    archive,
    sheet_paths,
    shared_strings,
    changed_cells,
    date1904=False,
):
    """変更セルだけをworksheet XMLから確認する。空欄セルはXMLに無くても正常扱い。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    cell_tag = f"{{{main_ns}}}c"
    targets_by_sheet = {}
    for sheet, row, column, expected in changed_cells:
        targets_by_sheet.setdefault(sheet, {})[_xlsx_cell_reference(row, column)] = expected

    for sheet, targets in targets_by_sheet.items():
        sheet_path = sheet_paths.get(sheet)
        if not sheet_path or sheet_path not in archive.namelist():
            raise ValueError(f"保存後の検証で{sheet}シートが見つかりません。")

        remaining = dict(targets)
        with archive.open(sheet_path) as stream:
            for _, element in ET.iterparse(stream, events=("end",)):
                if element.tag != cell_tag:
                    continue
                reference = str(element.attrib.get("r") or "")
                if reference not in remaining:
                    element.clear()
                    continue
                expected = remaining.pop(reference)
                force_date = isinstance(expected, (datetime, date))
                actual = _xlsx_cell_value(
                    element,
                    shared_strings,
                    date1904=date1904,
                    force_date=force_date,
                )
                element.clear()
                if not same_excel_value(actual, expected):
                    raise ValueError(
                        f"保存後の検証で{sheet}!{reference}の値が一致しません。"
                    )
                if not remaining:
                    break

        for reference, expected in remaining.items():
            # openpyxlも未作成セルをNoneとして読むため、空欄期待なら一致扱いにする。
            if not is_blank_excel_value(expected):
                raise ValueError(
                    f"保存後の検証で{sheet}!{reference}の値を確認できません。"
                )


def _rebuild_changed_customer_product_record_full_xml(
    content,
    customer_name,
    product_name,
    base_record,
):
    """従来どおり保存済みxlsmを走査して1商品を作り直す、安全側フォールバック。"""
    with zipfile.ZipFile(BytesIO(content), "r") as archive:
        _, sheet_paths, date1904 = _xlsx_workbook_info(archive)
        if DELIVERY_SHEET_NAME not in sheet_paths:
            raise ValueError("次回配達日シートが見つかりません。")

        shared_strings = _xlsx_shared_strings(archive)
        target_customer = normalize_match_value(customer_name)
        target_product = normalize_match_value(product_name)
        active_rows = []
        for row_number, values in _xlsx_iter_sheet_rows(
            archive,
            sheet_paths[DELIVERY_SHEET_NAME],
            shared_strings,
            max_col=16,
            date1904=date1904,
            date_columns={11, 13},
        ):
            if normalize_match_value(values[1] if len(values) >= 2 else None) != target_customer:
                continue
            if normalize_match_value(values[4] if len(values) >= 5 else None) != target_product:
                continue
            if is_blank_or_zero(values[6] if len(values) >= 7 else None):
                continue
            active_rows.append((row_number, values))

        if len(active_rows) != 1:
            raise ValueError("変更対象の商品行を1件に確定できません。")

        _, delivery_values = active_rows[0]
        history_key = delivery_history_identity(delivery_values)
        predicted_usage = None
        if history_key is not None:
            states = []
            history_path = sheet_paths.get(DELIVERY_HISTORY_SHEET_NAME)
            if history_path and history_path in archive.namelist():
                for _, history_values in _xlsx_iter_sheet_rows(
                    archive,
                    history_path,
                    shared_strings,
                    max_col=14,
                    date1904=date1904,
                    date_columns={11, 13},
                ):
                    if delivery_history_identity(history_values) == history_key:
                        states.append(history_values)
            states.append(delivery_values)
            predicted_usage = calculate_predicted_daily_usage_from_states(states)

        next_delivery, remaining = calculate_delivery_values(delivery_values)
        refreshed_record = dict(base_record or {})
        refreshed_record.update(
            {
                "ID": delivery_values[0] if len(delivery_values) >= 1 else None,
                "顧客名": delivery_values[1] if len(delivery_values) >= 2 else None,
                "地域": delivery_values[2] if len(delivery_values) >= 3 else None,
                "コンサル": delivery_values[3] if len(delivery_values) >= 4 else None,
                "商品名": delivery_values[4] if len(delivery_values) >= 5 else None,
                "使用数量/日": delivery_values[6] if len(delivery_values) >= 7 else None,
                "予想使用量/日": predicted_usage,
                "次回配達予定": next_delivery,
                "残数": remaining,
                "メーカー": delivery_values[5] if len(delivery_values) >= 6 else None,
                "在庫本数": delivery_values[7] if len(delivery_values) >= 8 else None,
                "本数": delivery_values[8] if len(delivery_values) >= 9 else None,
                "kg/本": delivery_values[9] if len(delivery_values) >= 10 else None,
                "配達日": delivery_values[10] if len(delivery_values) >= 11 else None,
                "_配達数量": delivery_values[11] if len(delivery_values) >= 12 else None,
            }
        )
        return refreshed_record


def _display_delivery_values_from_base_record(base_record, changed_cells):
    """直前revの表示1行へ、実際に保存した次回配達日シートの変更セルだけを反映する。"""
    if not isinstance(base_record, dict) or not base_record:
        return None

    values = [None] * 14
    values[0] = base_record.get("ID")
    values[1] = base_record.get("顧客名")
    values[2] = base_record.get("地域")
    values[3] = base_record.get("コンサル")
    values[4] = base_record.get("商品名")
    values[5] = base_record.get("メーカー")
    values[6] = base_record.get("使用数量/日")
    values[7] = base_record.get("在庫本数")
    values[8] = base_record.get("本数")
    values[9] = base_record.get("kg/本")
    values[10] = base_record.get("配達日")
    values[11] = base_record.get("_配達数量")
    values[12] = base_record.get("次回配達予定")

    delivery_rows = set()
    for sheet, row, column, expected in changed_cells or ():
        if sheet != DELIVERY_SHEET_NAME:
            continue
        try:
            row_number = int(row)
            column_number = int(column)
        except Exception:
            return None
        delivery_rows.add(row_number)
        if 1 <= column_number <= len(values):
            values[column_number - 1] = expected

    # 1回の保存で別々の現在行を同時更新する既存ルールはない。
    # 万一その状態になった場合は高速経路を使わず従来処理へ戻す。
    if len(delivery_rows) > 1:
        return None
    return tuple(values)


def _xlsx_matching_history_states_lightweight(
    archive,
    sheet_path,
    shared_strings,
    history_key,
    date1904=False,
):
    """履歴XMLを流し読みし、対象ID/顧客・商品に必要な列だけPython値へ戻す。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    row_tag = f"{{{main_ns}}}row"
    cell_tag = f"{{{main_ns}}}c"
    identity_columns = {1, 2, 5}
    calculation_columns = {8, 9, 10, 11}
    states = []

    with archive.open(sheet_path) as stream:
        for _, element in ET.iterparse(stream, events=("end",)):
            if element.tag != row_tag:
                continue

            identity_values = [None] * 14
            for cell in element:
                if cell.tag != cell_tag:
                    continue
                column_number = _xlsx_column_number_from_reference(cell.attrib.get("r"))
                if column_number not in identity_columns:
                    continue
                identity_values[column_number - 1] = _xlsx_cell_value(
                    cell,
                    shared_strings,
                    date1904=date1904,
                    force_date=False,
                )

            if delivery_history_identity(identity_values) == history_key:
                values = list(identity_values)
                for cell in element:
                    if cell.tag != cell_tag:
                        continue
                    column_number = _xlsx_column_number_from_reference(cell.attrib.get("r"))
                    if column_number not in calculation_columns:
                        continue
                    values[column_number - 1] = _xlsx_cell_value(
                        cell,
                        shared_strings,
                        date1904=date1904,
                        force_date=column_number == 11,
                    )
                states.append(tuple(values))

            element.clear()

    return states


def rebuild_changed_customer_product_record(
    content,
    customer_name,
    product_name,
    base_record,
    changed_cells=None,
):
    """直前revの1行を土台に履歴だけ読み、無理な場合は従来の全XML走査へ戻す。"""
    try:
        delivery_values = _display_delivery_values_from_base_record(
            base_record,
            changed_cells,
        )
        if delivery_values is None:
            raise ValueError("表示用の現在行を安全に組み立てられません。")

        if normalize_match_value(delivery_values[1]) != normalize_match_value(customer_name):
            raise ValueError("表示用の顧客が一致しません。")
        if normalize_match_value(delivery_values[4]) != normalize_match_value(product_name):
            raise ValueError("表示用の商品が一致しません。")
        if is_blank_or_zero(delivery_values[6]):
            raise ValueError("表示用の使用中商品を確認できません。")

        history_key = delivery_history_identity(delivery_values)
        predicted_usage = None
        if history_key is not None:
            with zipfile.ZipFile(BytesIO(content), "r") as archive:
                _, sheet_paths, date1904 = _xlsx_workbook_info(archive)
                history_path = sheet_paths.get(DELIVERY_HISTORY_SHEET_NAME)
                if not history_path or history_path not in archive.namelist():
                    raise ValueError("配達履歴シートが見つかりません。")
                shared_strings = _xlsx_shared_strings(archive)
                states = _xlsx_matching_history_states_lightweight(
                    archive,
                    history_path,
                    shared_strings,
                    history_key,
                    date1904=date1904,
                )
            states.append(delivery_values)
            predicted_usage = calculate_predicted_daily_usage_from_states(states)

        next_delivery, remaining = calculate_delivery_values(delivery_values)
        refreshed_record = dict(base_record)
        refreshed_record.update(
            {
                "ID": delivery_values[0],
                "顧客名": delivery_values[1],
                "地域": delivery_values[2],
                "コンサル": delivery_values[3],
                "商品名": delivery_values[4],
                "使用数量/日": delivery_values[6],
                "予想使用量/日": predicted_usage,
                "次回配達予定": next_delivery,
                "残数": remaining,
                "メーカー": delivery_values[5],
                "在庫本数": delivery_values[7],
                "本数": delivery_values[8],
                "kg/本": delivery_values[9],
                "配達日": delivery_values[10],
                "_配達数量": delivery_values[11],
            }
        )
        return refreshed_record
    except Exception:
        # 表示用の高速化だけの失敗で本体保存を危険にしない。
        # 従来と同じ保存済みxlsm全走査へ戻して正確さを優先する。
        return _rebuild_changed_customer_product_record_full_xml(
            content,
            customer_name,
            product_name,
            base_record,
        )



def try_refresh_fast_dropbox_cache_for_changed_product(
    content,
    previous_revision,
    access_token,
    customer_name,
    product_name,
    changed_cells=None,
    diagnostic_timings=None,
):
    """直前revと一致する表示用JSONがある時だけ、変更した1商品を差し替える。"""
    diagnostic_step_started = time.perf_counter()
    cache_content, _ = download_dropbox_file(
        DROPBOX_FAST_CACHE_FILE,
        access_token,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　表示① 既存JSON取得"] = time.perf_counter() - diagnostic_step_started
    if cache_content is None:
        return None

    diagnostic_step_started = time.perf_counter()
    payload = json.loads(cache_content.decode("utf-8"))
    if payload.get("cache_version") != DROPBOX_FAST_CACHE_VERSION:
        if diagnostic_timings is not None:
            diagnostic_timings["　表示② JSON解析・全体再計算"] = time.perf_counter() - diagnostic_step_started
        return None
    if str(payload.get("excel_revision") or "") != str(previous_revision or ""):
        if diagnostic_timings is not None:
            diagnostic_timings["　表示② JSON解析・全体再計算"] = time.perf_counter() - diagnostic_step_started
        return None

    records = payload.get("records", [])
    if not isinstance(records, list) or not records:
        if diagnostic_timings is not None:
            diagnostic_timings["　表示② JSON解析・全体再計算"] = time.perf_counter() - diagnostic_step_started
        return None

    base_df = recalculate_customer_inventory_for_today(pd.DataFrame(records))
    # 既存JSONの配達日は文字列型だが、変更対象の1行はExcelからdate値で戻る。
    # 値や保存形式は変えず、date値を安全に差し替えられる型へ広げるだけにする。
    if "配達日" in base_df.columns:
        base_df["配達日"] = base_df["配達日"].astype("object")

    target_customer = normalize_match_value(customer_name)
    target_product = normalize_match_value(product_name)
    matching_indexes = [
        index
        for index, row in base_df.iterrows()
        if normalize_match_value(row.get("顧客名")) == target_customer
        and normalize_match_value(row.get("商品名")) == target_product
    ]
    if diagnostic_timings is not None:
        diagnostic_timings["　表示② JSON解析・全体再計算"] = time.perf_counter() - diagnostic_step_started
    if len(matching_indexes) != 1:
        return None

    diagnostic_step_started = time.perf_counter()
    target_index = matching_indexes[0]
    refreshed_record = rebuild_changed_customer_product_record(
        content,
        customer_name,
        product_name,
        base_df.loc[target_index].to_dict(),
        changed_cells=changed_cells,
    )
    for column, value in refreshed_record.items():
        if column not in base_df.columns:
            base_df[column] = None
        base_df.at[target_index, column] = value
    if diagnostic_timings is not None:
        diagnostic_timings["　表示③ 対象商品差替え"] = time.perf_counter() - diagnostic_step_started
    return base_df


def refresh_fast_dropbox_cache_after_save(
    content,
    excel_revision,
    access_token,
    previous_revision="",
    customer_name="",
    product_name="",
    changed_cells=None,
    diagnostic_timings=None,
):
    """保存直後の表示用JSONを更新する。安全に差分更新できない時は従来の全体再生成へ戻す。"""
    try:
        refreshed_df = None

        # 保存直前のExcel revと既存JSONのrevが完全一致する場合だけ、
        # 変更した顧客・商品1行の差分更新を使う。
        # 条件が1つでも合わない場合は、従来どおりExcel全体から作り直す。
        if previous_revision and customer_name and product_name:
            try:
                refreshed_df = try_refresh_fast_dropbox_cache_for_changed_product(
                    content,
                    previous_revision,
                    access_token,
                    customer_name,
                    product_name,
                    changed_cells=changed_cells,
                    diagnostic_timings=diagnostic_timings,
                )
            except Exception as exc:
                if diagnostic_timings is not None:
                    diagnostic_timings["表示高速経路エラー"] = f"{type(exc).__name__}: {exc}"
                refreshed_df = None

        if not isinstance(refreshed_df, pd.DataFrame) or refreshed_df.empty:
            diagnostic_step_started = time.perf_counter()
            refreshed_df = rebuild_sheet1_from_formula_references(BytesIO(content))
            if diagnostic_timings is not None:
                diagnostic_timings["　表示④ 全体再生成（フォールバック）"] = time.perf_counter() - diagnostic_step_started
        if refreshed_df.empty:
            return "保存は完了しましたが、表示用キャッシュを更新できませんでした。更新ボタンを押してください。"

        # Dropbox側の更新番号やJSONの反映待ちに左右されず、保存直後の1回目の
        # 再表示では、今保存したExcelから作った最新データをそのまま使用する。
        # 次の画面実行で1度だけ取り出し、その後は従来どおりDropboxキャッシュを使う。
        diagnostic_step_started = time.perf_counter()
        st.session_state["customer_excel_immediate_df"] = refreshed_df.copy()
        if diagnostic_timings is not None:
            diagnostic_timings["　表示⑤ 即時表示データ保持"] = time.perf_counter() - diagnostic_step_started

        diagnostic_step_started = time.perf_counter()
        records = json.loads(
            refreshed_df.to_json(
                orient="records",
                date_format="iso",
                force_ascii=False,
            )
        )
        cache_payload = json.dumps(
            {
                "cache_version": DROPBOX_FAST_CACHE_VERSION,
                "excel_revision": excel_revision,
                "records": records,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
        if diagnostic_timings is not None:
            diagnostic_timings["　表示⑥ JSON変換"] = time.perf_counter() - diagnostic_step_started

        diagnostic_step_started = time.perf_counter()
        cache_response = upload_dropbox_file(
            DROPBOX_FAST_CACHE_FILE,
            cache_payload,
            access_token,
            mode="overwrite",
        )
        if diagnostic_timings is not None:
            diagnostic_timings["　表示⑦ JSON保存"] = time.perf_counter() - diagnostic_step_started
        if cache_response.status_code != 200:
            return "保存は完了しましたが、表示用キャッシュを更新できませんでした。更新ボタンを押してください。"
        return ""
    except Exception:
        # 本番Excelの保存と検証は完了しているため、キャッシュ更新失敗だけで保存失敗にはしない。
        return "保存は完了しましたが、表示用キャッシュを更新できませんでした。更新ボタンを押してください。"


def clear_customer_excel_caches_after_save():
    """顧客Excel更新で内容が変わるキャッシュだけを無効化する。

    写真・メモ・OneDrive・配車表・取引先など、在庫変更と無関係な
    キャッシュは残し、保存直後の画面再表示を重くしない。
    """
    cache_function_names = (
        "get_cached_dropbox_excel_content",
        "load_fast_dropbox_data",
        "load_data",
        "load_product_search_index",
    )
    for function_name in cache_function_names:
        cached_function = globals().get(function_name)
        if cached_function is not None and hasattr(cached_function, "clear"):
            cached_function.clear()


def find_next_delivery_history_row(ws):
    """配達履歴A列の最終データ行の次を返す。書式だけの空行は数えない。"""
    row = max(int(ws.max_row or 1), 1)
    while row > 1 and is_blank_excel_value(ws.cell(row, 1).value):
        row -= 1
    return row + 1


def read_cached_delivery_row_values_lightweight(original_content, product_row):
    """計算済み値だけを読み取り専用で取得する。元Excelやマクロは変更しない。"""
    cached_workbook = load_workbook(
        BytesIO(original_content),
        keep_vba=False,
        data_only=True,
        read_only=True,
    )
    try:
        if DELIVERY_SHEET_NAME not in cached_workbook.sheetnames:
            return None
        cached_delivery_ws = cached_workbook[DELIVERY_SHEET_NAME]
        row_values = next(
            cached_delivery_ws.iter_rows(
                min_row=product_row,
                max_row=product_row,
                min_col=1,
                max_col=14,
                values_only=True,
            ),
            None,
        )
        return tuple(row_values) if row_values is not None else None
    finally:
        cached_workbook.close()


def verify_saved_workbook_lightweight(
    saved_content,
    original_sheets,
    required_sheets,
    changed_cells,
):
    """openpyxlで再オープンせず、xlsm内XMLから従来と同じ確認項目を検証する。"""
    try:
        with zipfile.ZipFile(BytesIO(saved_content), "r") as archive:
            saved_sheets, sheet_paths, date1904 = _xlsx_workbook_info(archive)
            if list(saved_sheets) != list(original_sheets):
                raise ValueError("保存後にシート構成が変わったため、更新を中止しました。")
            if not set(required_sheets).issubset(set(saved_sheets)):
                raise ValueError("保存後の検証で必要なシートが見つかりません。")
            if "xl/vbaProject.bin" not in archive.namelist():
                raise ValueError("保存後の検証でVBAプロジェクトを確認できません。")

            shared_strings = _xlsx_shared_strings(archive)
            _xlsx_verify_changed_cells(
                archive,
                sheet_paths,
                shared_strings,
                changed_cells,
                date1904=date1904,
            )
    except zipfile.BadZipFile as exc:
        raise ValueError("保存後の検証でExcelファイルを開けません。") from exc



def copy_previous_delivery_state_to_history(
    workbook,
    cached_row_values,
    delivery_ws,
    product_row,
    changed_cells,
):
    """Excelの入力マクロと同じく、更新前のA:Nを配達履歴へ1行追加する。"""
    if DELIVERY_HISTORY_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("配達履歴シートが見つからないため、在庫・配達情報を安全に保存できません。")
    if MANAGEMENT_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("管理シートが見つからないため、在庫・配達情報を安全に保存できません。")

    history_ws = workbook[DELIVERY_HISTORY_SHEET_NAME]
    management_ws = workbook[MANAGEMENT_SHEET_NAME]
    history_row = find_next_delivery_history_row(history_ws)

    old_values = [delivery_ws.cell(product_row, column).value for column in range(1, 17)]
    calculated_next_delivery, _ = calculate_delivery_values(old_values)

    for column in range(1, 15):
        value = delivery_ws.cell(product_row, column).value
        if isinstance(value, str) and value.startswith("="):
            cached_value = (
                cached_row_values[column - 1]
                if cached_row_values is not None and len(cached_row_values) >= column
                else None
            )
            if cached_value is not None:
                value = cached_value
            elif column == 13:
                value = calculated_next_delivery
            else:
                value = None
        history_ws.cell(history_row, column).value = value
        changed_cells.append((DELIVERY_HISTORY_SHEET_NAME, history_row, column, value))

    history_count = history_row - 1
    management_ws["B8"] = history_count
    changed_cells.append((MANAGEMENT_SHEET_NAME, 8, 2, history_count))


def is_test_xml_fast_save_enabled():
    """Privateテスト環境で明示的に有効化した時だけXML直接保存を使う。"""
    try:
        value = st.secrets.get("TEST_XML_FAST_SAVE", False)
    except Exception:
        return False
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


class _XmlFastSaveUnavailable(Exception):
    """安全にXML直接保存できない条件では従来方式へ戻すための内部例外。"""


def _xml_fast_root_start_tag(xml_bytes):
    """XML宣言の直後にあるルート要素の開始タグを返す。"""
    match = re.search(br"<([A-Za-z_][\\w:.-]*)(?:\\s[^>]*)?>", bytes(xml_bytes or b""))
    return match


def _xml_fast_register_original_namespaces(original_xml):
    """元XMLのnamespace prefixをElementTreeへ登録し、既存prefixを極力維持する。"""
    match = _xml_fast_root_start_tag(original_xml)
    if not match:
        return
    start_tag = match.group(0)
    namespace_pattern = re.compile(
        br"\\s+xmlns(?::([A-Za-z_][\\w.-]*))?=(?:\"([^\"]*)\"|'([^']*)')"
    )
    for namespace_match in namespace_pattern.finditer(start_tag):
        prefix = (namespace_match.group(1) or b"").decode("utf-8", errors="ignore")
        uri_bytes = namespace_match.group(2) or namespace_match.group(3) or b""
        uri = uri_bytes.decode("utf-8", errors="ignore")
        if not uri:
            continue
        try:
            ET.register_namespace(prefix, uri)
        except Exception:
            pass


def _xml_fast_serialize(root, original_xml):
    """未知の拡張要素を保持しつつ、元XMLのnamespace宣言も残してシリアライズする。"""
    original_xml = bytes(original_xml or b"")
    _xml_fast_register_original_namespaces(original_xml)
    serialized = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    original_match = _xml_fast_root_start_tag(original_xml)
    serialized_match = _xml_fast_root_start_tag(serialized)
    if not original_match or not serialized_match:
        return serialized

    original_start = original_match.group(0)
    serialized_start = serialized_match.group(0)
    declaration_pattern = re.compile(
        br"\\s+xmlns(?::[A-Za-z_][\\w.-]*)?=(?:\"[^\"]*\"|'[^']*')"
    )
    additions = []
    for declaration in declaration_pattern.findall(original_start):
        name = declaration.strip().split(b"=", 1)[0]
        if name + b"=" not in serialized_start:
            additions.append(declaration)
    if not additions:
        return serialized

    new_start = serialized_start[:-1] + b"".join(additions) + serialized_start[-1:]
    return (
        serialized[: serialized_match.start()]
        + new_start
        + serialized[serialized_match.end() :]
    )


def _xml_fast_rebuild_zip(original_content, replacements):
    """指定XMLだけ差し替え、その他のxlsm内部ファイルは元バイトをそのままコピーする。"""
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(original_content), "r") as source_archive:
        with zipfile.ZipFile(output, "w") as target_archive:
            for info in source_archive.infolist():
                payload = replacements.get(info.filename)
                if payload is None:
                    payload = source_archive.read(info.filename)
                target_archive.writestr(info, payload)
    return output.getvalue()


def _xml_fast_find_row(sheet_root, row_number):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet_data = sheet_root.find(f"{{{main_ns}}}sheetData")
    if sheet_data is None:
        return None
    target = int(row_number)
    for row in sheet_data.findall(f"{{{main_ns}}}row"):
        try:
            current = int(row.attrib.get("r") or 0)
        except Exception:
            current = 0
        if current == target:
            return row
    return None


def _xml_fast_cell_in_row(row, column_number):
    if row is None:
        return None
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    for cell in row.findall(f"{{{main_ns}}}c"):
        if _xlsx_column_number_from_reference(cell.attrib.get("r")) == int(column_number):
            return cell
    return None


def _xml_fast_cell_value(
    cell,
    shared_strings,
    date1904=False,
    force_date=False,
    formula_cache=False,
):
    """通常値または数式の保存済みキャッシュ値をXMLからPython値へ戻す。"""
    if cell is None:
        return None
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    formula = cell.find(f"{{{main_ns}}}f")
    if formula is not None and not formula_cache:
        return "=" + str(formula.text or "")

    cell_type = str(cell.attrib.get("t") or "")
    if cell_type == "inlineStr":
        inline = cell.find(f"{{{main_ns}}}is")
        if inline is None:
            return ""
        return "".join(node.text or "" for node in inline.iter(f"{{{main_ns}}}t"))

    value_node = cell.find(f"{{{main_ns}}}v")
    value_text = value_node.text if value_node is not None else None
    if value_text is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(value_text)]
        except Exception:
            return None
    if cell_type == "b":
        return str(value_text).strip() == "1"
    if cell_type in {"str", "e"}:
        return str(value_text)
    if cell_type == "d":
        try:
            return datetime.fromisoformat(str(value_text).replace("Z", "+00:00"))
        except Exception:
            return str(value_text)

    try:
        number = float(value_text)
    except Exception:
        return str(value_text)
    if force_date:
        try:
            return _xlsx_excel_datetime(number, date1904=date1904)
        except Exception:
            return number
    return int(number) if number.is_integer() else number


def _xml_fast_row_values(
    row,
    shared_strings,
    max_col=16,
    date1904=False,
    date_columns=None,
    formula_cache=False,
):
    date_columns = set(date_columns or ())
    values = [None] * int(max_col)
    if row is None:
        return tuple(values)
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    for cell in row.findall(f"{{{main_ns}}}c"):
        column_number = _xlsx_column_number_from_reference(cell.attrib.get("r"))
        if column_number < 1 or column_number > max_col:
            continue
        values[column_number - 1] = _xml_fast_cell_value(
            cell,
            shared_strings,
            date1904=date1904,
            force_date=column_number in date_columns,
            formula_cache=formula_cache,
        )
    return tuple(values)


def _xml_fast_excel_serial(value, date1904=False):
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime.combine(value, datetime.min.time())
    if not isinstance(value, datetime):
        raise TypeError("Excel日付へ変換できない値です。")
    if value.tzinfo is not None:
        value = value.replace(tzinfo=None)
    epoch = datetime(1904, 1, 1) if date1904 else datetime(1899, 12, 30)
    delta = value - epoch
    return delta.days + (delta.seconds + delta.microseconds / 1_000_000) / 86400


def _xml_fast_insert_cell_sorted(row, cell, column_number):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    for index, other in enumerate(list(row)):
        if other.tag != f"{{{main_ns}}}c":
            continue
        other_column = _xlsx_column_number_from_reference(other.attrib.get("r"))
        if other_column > int(column_number):
            row.insert(index, cell)
            return
    row.append(cell)


def _xml_fast_set_cell_value(
    row,
    row_number,
    column_number,
    value,
    date1904=False,
    date_style=None,
):
    """既存styleを保持し、値だけをExcel XMLへ書く。新規日付は履歴の既存日付styleを継承する。"""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    cell = _xml_fast_cell_in_row(row, column_number)
    if cell is None:
        cell = ET.Element(
            f"{{{main_ns}}}c",
            {"r": _xlsx_cell_reference(row_number, column_number)},
        )
        _xml_fast_insert_cell_sorted(row, cell, column_number)

    if isinstance(value, (date, datetime)) and date_style and not cell.attrib.get("s"):
        cell.attrib["s"] = str(date_style)

    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)

    if value is None:
        return cell
    if isinstance(value, bool):
        cell.attrib["t"] = "b"
        ET.SubElement(cell, f"{{{main_ns}}}v").text = "1" if value else "0"
        return cell
    if isinstance(value, (date, datetime)):
        serial = _xml_fast_excel_serial(value, date1904=date1904)
        cell.attrib["t"] = "n"
        ET.SubElement(cell, f"{{{main_ns}}}v").text = (
            str(int(serial)) if float(serial).is_integer() else str(serial)
        )
        return cell
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and not math.isfinite(value):
            raise ValueError("有限でない数値はExcelへ保存できません。")
        cell.attrib["t"] = "n"
        number_text = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
        ET.SubElement(cell, f"{{{main_ns}}}v").text = number_text
        return cell

    text = str(value)
    if text.startswith("="):
        formula = ET.SubElement(cell, f"{{{main_ns}}}f")
        formula.text = text[1:]
        ET.SubElement(cell, f"{{{main_ns}}}v")
        return cell

    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{main_ns}}}is")
    text_node = ET.SubElement(inline, f"{{{main_ns}}}t")
    if text[:1].isspace() or text[-1:].isspace():
        text_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    text_node.text = text
    return cell


def _xml_fast_clear_formula_caches(row):
    """変更行の数式は式を残し、古い計算済み値だけ空にしてExcel再計算へ任せる。"""
    if row is None:
        return
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    for cell in row.findall(f"{{{main_ns}}}c"):
        if cell.find(f"{{{main_ns}}}f") is None:
            continue
        value_node = cell.find(f"{{{main_ns}}}v")
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{main_ns}}}v")
        value_node.text = None


def _xml_fast_update_dimension(sheet_root, last_row):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    dimension = sheet_root.find(f"{{{main_ns}}}dimension")
    if dimension is None:
        return
    current = str(dimension.attrib.get("ref") or "")
    if ":" not in current:
        return
    first, last = current.split(":", 1)
    match = re.match(r"^([A-Za-z]+)", last)
    last_column = match.group(1) if match else "A"
    dimension.attrib["ref"] = f"{first}:{last_column}{int(last_row)}"


def _xml_fast_find_recent_date_style(sheet_data, column_number):
    if sheet_data is None:
        return None
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rows = sheet_data.findall(f"{{{main_ns}}}row")
    for row in reversed(rows):
        cell = _xml_fast_cell_in_row(row, column_number)
        if cell is not None and cell.attrib.get("s"):
            return cell.attrib.get("s")
    return None


def _xml_fast_patch_recalculation(workbook_root):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    calculation = workbook_root.find(f"{{{main_ns}}}calcPr")
    if calculation is None:
        calculation = ET.SubElement(workbook_root, f"{{{main_ns}}}calcPr")
    calculation.attrib["calcMode"] = "auto"
    calculation.attrib["fullCalcOnLoad"] = "1"
    calculation.attrib["forceFullCalc"] = "1"


def _xml_fast_validate_vba_unchanged(original_vba, saved_content):
    try:
        with zipfile.ZipFile(BytesIO(saved_content), "r") as archive:
            saved_vba = archive.read("xl/vbaProject.bin")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ValueError("保存後の検証でVBAプロジェクトを確認できません。") from exc
    if saved_vba != original_vba:
        raise ValueError("保存後にVBAプロジェクトが変化したため、更新を中止しました。")


def _xml_fast_find_active_product_row(
    delivery_root,
    shared_strings,
    customer_name,
    product_name,
    date1904=False,
):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet_data = delivery_root.find(f"{{{main_ns}}}sheetData")
    if sheet_data is None:
        raise _XmlFastSaveUnavailable("次回配達日シートのデータを確認できません。")

    matched_rows = []
    active_rows = []
    for row in sheet_data.findall(f"{{{main_ns}}}row"):
        values = _xml_fast_row_values(
            row,
            shared_strings,
            max_col=7,
            date1904=date1904,
            date_columns=(),
        )
        if (
            normalize_match_value(values[1]) == customer_name
            and normalize_match_value(values[4]) == product_name
        ):
            row_number = int(row.attrib.get("r") or 0)
            matched_rows.append(row_number)
            if not is_blank_or_zero(values[6]):
                active_rows.append(row_number)

    if not matched_rows:
        raise ValueError("顧客名・商品名が一致する行が見つかりません。")
    if len(active_rows) > 1:
        raise ValueError("同じ顧客名・商品名の行が複数見つかりました。確認してください。")
    if not active_rows:
        raise ValueError("使用数量/日に値が入っている行が見つからないため編集できません。")
    return active_rows[0]


def _xml_fast_next_history_row(history_root, shared_strings, date1904=False):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet_data = history_root.find(f"{{{main_ns}}}sheetData")
    if sheet_data is None:
        raise _XmlFastSaveUnavailable("配達履歴シートのデータを確認できません。")
    last_nonblank_row = 1
    for row in sheet_data.findall(f"{{{main_ns}}}row"):
        first_cell = _xml_fast_cell_in_row(row, 1)
        value = _xml_fast_cell_value(
            first_cell,
            shared_strings,
            date1904=date1904,
        )
        if not is_blank_excel_value(value):
            last_nonblank_row = max(last_nonblank_row, int(row.attrib.get("r") or 0))
    return last_nonblank_row + 1, sheet_data


def _update_workbook_bytes_xml_fast(
    original_content,
    customer_name,
    product_name,
    proposed,
    diagnostic_timings=None,
):
    """Privateテスト用。既存xlsmの必要XMLだけを書き換え、VBA等は元バイトのまま保持する。"""
    initial_started = time.perf_counter()
    replacements = {}
    changed_cells = []

    with zipfile.ZipFile(BytesIO(original_content), "r") as archive:
        original_sheets, sheet_paths, date1904 = _xlsx_workbook_info(archive)
        required_sheets = {
            DELIVERY_SHEET_NAME,
            SHEET_NAME,
            DELIVERY_HISTORY_SHEET_NAME,
            MANAGEMENT_SHEET_NAME,
        }
        if not required_sheets.issubset(set(original_sheets)):
            raise _XmlFastSaveUnavailable("必要なシートを確認できません。")
        if "xl/vbaProject.bin" not in archive.namelist():
            raise _XmlFastSaveUnavailable("元ExcelのVBAプロジェクトを確認できません。")
        original_vba = archive.read("xl/vbaProject.bin")
        shared_strings = _xlsx_shared_strings(archive)

        delivery_path = sheet_paths[DELIVERY_SHEET_NAME]
        delivery_xml = archive.read(delivery_path)
        delivery_root = ET.fromstring(delivery_xml)
        product_row = _xml_fast_find_active_product_row(
            delivery_root,
            shared_strings,
            customer_name,
            product_name,
            date1904=date1904,
        )
        delivery_row = _xml_fast_find_row(delivery_root, product_row)
        if delivery_row is None:
            raise _XmlFastSaveUnavailable("編集対象行をXMLから確認できません。")
        old_values = list(
            _xml_fast_row_values(
                delivery_row,
                shared_strings,
                max_col=16,
                date1904=date1904,
                date_columns={11, 13},
            )
        )
        cached_values = list(
            _xml_fast_row_values(
                delivery_row,
                shared_strings,
                max_col=14,
                date1904=date1904,
                date_columns={11, 13},
                formula_cache=True,
            )
        )
        if diagnostic_timings is not None:
            diagnostic_timings["　Excel① マクロ付きExcelを開く"] = (
                time.perf_counter() - initial_started
            )
            diagnostic_timings["　Excel② 計算値用Excelを開く"] = 0.0

        edit_started = time.perf_counter()
        event_changed = any(
            not same_excel_value(old_values[column - 1], proposed.get(label))
            for label, column in {
                "在庫本数": 8,
                "本数": 9,
                "配達日": 11,
            }.items()
        )

        if event_changed:
            history_path = sheet_paths[DELIVERY_HISTORY_SHEET_NAME]
            history_xml = archive.read(history_path)
            history_root = ET.fromstring(history_xml)
            history_row_number, history_sheet_data = _xml_fast_next_history_row(
                history_root,
                shared_strings,
                date1904=date1904,
            )
            new_history_row = ET.Element(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row",
                {"r": str(history_row_number)},
            )
            calculated_next_delivery, _ = calculate_delivery_values(old_values)
            for column in range(1, 15):
                value = old_values[column - 1]
                source_cell = _xml_fast_cell_in_row(delivery_row, column)
                if source_cell is not None and source_cell.find(
                    "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f"
                ) is not None:
                    cached_value = cached_values[column - 1]
                    if cached_value is not None:
                        value = cached_value
                    elif column == 13:
                        value = calculated_next_delivery
                    else:
                        value = None

                if value is not None:
                    date_style = None
                    if isinstance(value, (date, datetime)):
                        date_style = _xml_fast_find_recent_date_style(
                            history_sheet_data,
                            column,
                        )
                        if not date_style and source_cell is not None:
                            date_style = source_cell.attrib.get("s")
                    _xml_fast_set_cell_value(
                        new_history_row,
                        history_row_number,
                        column,
                        value,
                        date1904=date1904,
                        date_style=date_style,
                    )
                changed_cells.append(
                    (DELIVERY_HISTORY_SHEET_NAME, history_row_number, column, value)
                )

            history_sheet_data.append(new_history_row)
            _xml_fast_update_dimension(history_root, history_row_number)
            replacements[history_path] = _xml_fast_serialize(history_root, history_xml)

            management_path = sheet_paths[MANAGEMENT_SHEET_NAME]
            management_xml = archive.read(management_path)
            management_root = ET.fromstring(management_xml)
            management_row = _xml_fast_find_row(management_root, 8)
            if management_row is None:
                raise _XmlFastSaveUnavailable("管理シートB8を確認できません。")
            history_count = history_row_number - 1
            _xml_fast_set_cell_value(
                management_row,
                8,
                2,
                history_count,
                date1904=date1904,
            )
            changed_cells.append((MANAGEMENT_SHEET_NAME, 8, 2, history_count))
            replacements[management_path] = _xml_fast_serialize(
                management_root,
                management_xml,
            )

        column_mapping = {
            "メーカー": 6,
            "在庫本数": 8,
            "本数": 9,
            "kg/本": 10,
            "配達日": 11,
        }
        for label, column in column_mapping.items():
            new_value = proposed.get(label)
            if not same_excel_value(old_values[column - 1], new_value):
                _xml_fast_set_cell_value(
                    delivery_row,
                    product_row,
                    column,
                    new_value,
                    date1904=date1904,
                )
                changed_cells.append(
                    (DELIVERY_SHEET_NAME, product_row, column, new_value)
                )

        inventory_count = inventory_usage_number(proposed.get("在庫本数"))
        delivery_count = inventory_usage_number(proposed.get("本数"))
        kg_per_bottle = inventory_usage_number(proposed.get("kg/本"))
        delivery_quantity = None
        if inventory_count is not None or delivery_count is not None:
            if kg_per_bottle is None or kg_per_bottle <= 0:
                raise ValueError(
                    "在庫本数または本数を入力する場合は、kg/本に0より大きい数値が必要です。"
                )
            delivery_quantity = (inventory_count or 0) + (delivery_count or 0)
            delivery_quantity *= kg_per_bottle
            if not math.isfinite(delivery_quantity):
                raise ValueError("配達数量を正しく計算できませんでした。")
            if delivery_quantity.is_integer():
                delivery_quantity = int(delivery_quantity)

        if not same_excel_value(old_values[11], delivery_quantity):
            _xml_fast_set_cell_value(
                delivery_row,
                product_row,
                12,
                delivery_quantity,
                date1904=date1904,
            )
            changed_cells.append(
                (DELIVERY_SHEET_NAME, product_row, 12, delivery_quantity)
            )

        if not changed_cells:
            raise ValueError("変更された項目がありません。")

        _xml_fast_clear_formula_caches(delivery_row)
        replacements[delivery_path] = _xml_fast_serialize(delivery_root, delivery_xml)

        workbook_xml = archive.read("xl/workbook.xml")
        workbook_root = ET.fromstring(workbook_xml)
        _xml_fast_patch_recalculation(workbook_root)
        replacements["xl/workbook.xml"] = _xml_fast_serialize(
            workbook_root,
            workbook_xml,
        )

    if diagnostic_timings is not None:
        diagnostic_timings["　Excel③ 対象検索・書換え"] = (
            time.perf_counter() - edit_started
        )

    save_started = time.perf_counter()
    saved_content = _xml_fast_rebuild_zip(original_content, replacements)
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel④ xlsm保存"] = time.perf_counter() - save_started

    vba_started = time.perf_counter()
    _xml_fast_validate_vba_unchanged(original_vba, saved_content)
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑤ 保存後Excelを開く"] = time.perf_counter() - vba_started

    verify_started = time.perf_counter()
    verify_saved_workbook_lightweight(
        saved_content,
        original_sheets,
        required_sheets,
        changed_cells,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑥ 保存後内容検証"] = time.perf_counter() - verify_started
    return saved_content, changed_cells


def _update_delivery_history_record_bytes_xml_fast(
    original_content,
    customer_name,
    product_name,
    record_ref,
    proposed,
    diagnostic_timings=None,
):
    """Privateテスト用。現在行または過去履歴の指定1行だけをXML直接更新する。"""
    source = str((record_ref or {}).get("source") or "").strip()
    if source not in {"current", "history"}:
        raise _XmlFastSaveUnavailable("修正対象の種類を確認できないため従来方式を使用します。")

    initial_started = time.perf_counter()
    replacements = {}
    changed_cells = []
    row_number = int((record_ref or {}).get("row_number") or 0)
    expected_fingerprint = str((record_ref or {}).get("fingerprint") or "").strip()
    expected_identity = tuple((record_ref or {}).get("identity") or ())

    with zipfile.ZipFile(BytesIO(original_content), "r") as archive:
        original_sheets, sheet_paths, date1904 = _xlsx_workbook_info(archive)
        required_sheets = {
            DELIVERY_SHEET_NAME,
            DELIVERY_HISTORY_SHEET_NAME,
            MANAGEMENT_SHEET_NAME,
            SHEET_NAME,
        }
        if not required_sheets.issubset(set(original_sheets)):
            raise _XmlFastSaveUnavailable("必要なシートを確認できません。")
        if "xl/vbaProject.bin" not in archive.namelist():
            raise _XmlFastSaveUnavailable("元ExcelのVBAプロジェクトを確認できません。")
        original_vba = archive.read("xl/vbaProject.bin")
        shared_strings = _xlsx_shared_strings(archive)

        sheet_name = (
            DELIVERY_SHEET_NAME
            if source == "current"
            else DELIVERY_HISTORY_SHEET_NAME
        )
        sheet_path = sheet_paths[sheet_name]
        sheet_xml = archive.read(sheet_path)
        sheet_root = ET.fromstring(sheet_xml)
        row = _xml_fast_find_row(sheet_root, row_number)
        if row is None:
            raise ValueError("修正対象の行が見つかりません。再読み込みしてください。")
        row_values = list(
            _xml_fast_row_values(
                row,
                shared_strings,
                max_col=14,
                date1904=date1904,
                date_columns={11, 13},
            )
        )

        # 現在行には共有数式が含まれる場合があり、XML単体では数式文字列を
        # openpyxlと完全に同じ形へ復元できないことがある。競合検知の指紋は
        # 従来と同じA:Nの値で比較するため、現在行だけ読み取り専用で1行取得する。
        # 編集・保存自体はXML直接更新のままなので、通常の重いopenpyxl保存には戻さない。
        if source == "current":
            validation_workbook = load_workbook(
                BytesIO(original_content),
                keep_vba=False,
                data_only=False,
                read_only=True,
            )
            try:
                validation_ws = validation_workbook[DELIVERY_SHEET_NAME]
                validation_rows = list(
                    validation_ws.iter_rows(
                        min_row=row_number,
                        max_row=row_number,
                        max_col=14,
                        values_only=True,
                    )
                )
                if not validation_rows:
                    raise ValueError("修正対象の行が見つかりません。再読み込みしてください。")
                row_values = list(validation_rows[0])
            finally:
                validation_workbook.close()

        if diagnostic_timings is not None:
            diagnostic_timings["　Excel① マクロ付きExcelを開く"] = (
                time.perf_counter() - initial_started
            )

        actual_identity = delivery_history_identity(row_values)
        if expected_identity and actual_identity != expected_identity:
            raise ValueError("修正対象の顧客・商品が変わっています。再読み込みしてください。")
        if normalize_match_value(row_values[1]) != normalize_match_value(customer_name):
            raise ValueError("修正対象の顧客が一致しません。再読み込みしてください。")
        if normalize_match_value(row_values[4]) != normalize_match_value(product_name):
            raise ValueError("修正対象の商品が一致しません。再読み込みしてください。")
        actual_fingerprint = delivery_record_fingerprint(
            source,
            row_number,
            row_values,
        )
        if expected_fingerprint and actual_fingerprint != expected_fingerprint:
            raise ValueError(
                "この納品履歴はPCまたは別端末で変更されています。"
                "納品履歴を開き直してから修正してください。"
            )

        edit_started = time.perf_counter()
        column_mapping = {
            "メーカー": 6,
            "在庫本数": 8,
            "本数": 9,
            "kg/本": 10,
            "配達日": 11,
        }
        old_values = {
            label: row_values[column - 1]
            for label, column in column_mapping.items()
        }
        for label, column in column_mapping.items():
            new_value = proposed.get(label)
            if not same_excel_value(row_values[column - 1], new_value):
                _xml_fast_set_cell_value(
                    row,
                    row_number,
                    column,
                    new_value,
                    date1904=date1904,
                )
                changed_cells.append((sheet_name, row_number, column, new_value))

        inventory_changed = not same_excel_value(
            old_values.get("在庫本数"),
            proposed.get("在庫本数"),
        )
        delivery_changed = not same_excel_value(
            old_values.get("本数"),
            proposed.get("本数"),
        )
        kg_changed = not same_excel_value(
            old_values.get("kg/本"),
            proposed.get("kg/本"),
        )
        date_changed = not same_excel_value(
            old_values.get("配達日"),
            proposed.get("配達日"),
        )

        quantity_changed = False
        should_recalculate_quantity = inventory_changed or delivery_changed
        if kg_changed:
            should_recalculate_quantity = should_recalculate_quantity or (
                inventory_usage_number(proposed.get("在庫本数")) is not None
                or inventory_usage_number(proposed.get("本数")) is not None
            )

        if should_recalculate_quantity:
            inventory_count = inventory_usage_number(proposed.get("在庫本数"))
            delivery_count = inventory_usage_number(proposed.get("本数"))
            kg_per_bottle = inventory_usage_number(proposed.get("kg/本"))
            corrected_quantity = None
            if inventory_count is not None or delivery_count is not None:
                if kg_per_bottle is None or kg_per_bottle <= 0:
                    raise ValueError(
                        "在庫本数または本数を入れる場合は、kg/本に0より大きい数値が必要です。"
                    )
                corrected_quantity = (
                    (inventory_count or 0) + (delivery_count or 0)
                ) * kg_per_bottle
                if not math.isfinite(corrected_quantity):
                    raise ValueError("配達数量を正しく計算できませんでした。")
                if corrected_quantity.is_integer():
                    corrected_quantity = int(corrected_quantity)

            if not same_excel_value(row_values[11], corrected_quantity):
                _xml_fast_set_cell_value(
                    row,
                    row_number,
                    12,
                    corrected_quantity,
                    date1904=date1904,
                )
                changed_cells.append(
                    (sheet_name, row_number, 12, corrected_quantity)
                )
                quantity_changed = True

        # 配達履歴シートは数式ではないため、訂正後の次回配達予定も同じ行で再計算する。
        # 次回配達日シート（現在の登録）は従来どおりM列数式そのものを維持し、
        # Excel再計算に任せる。ここではM列の式や計算ルールを変更しない。
        if source == "history" and (date_changed or quantity_changed):
            corrected_values = list(
                _xml_fast_row_values(
                    row,
                    shared_strings,
                    max_col=14,
                    date1904=date1904,
                    date_columns={11, 13},
                )
            )
            corrected_next_delivery, _ = calculate_delivery_values(corrected_values)
            if not same_excel_value(row_values[12], corrected_next_delivery):
                _xml_fast_set_cell_value(
                    row,
                    row_number,
                    13,
                    corrected_next_delivery,
                    date1904=date1904,
                )
                changed_cells.append(
                    (sheet_name, row_number, 13, corrected_next_delivery)
                )

        if not changed_cells:
            raise ValueError("変更された項目がありません。")

        if source == "current":
            # 現在行のM列など既存数式は残したまま、古い計算キャッシュだけを消す。
            # 通常のXML高速保存と同じ扱いにしてExcel再計算へ任せる。
            _xml_fast_clear_formula_caches(row)

        replacements[sheet_path] = _xml_fast_serialize(sheet_root, sheet_xml)
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_root = ET.fromstring(workbook_xml)
        _xml_fast_patch_recalculation(workbook_root)
        replacements["xl/workbook.xml"] = _xml_fast_serialize(
            workbook_root,
            workbook_xml,
        )

    if diagnostic_timings is not None:
        diagnostic_timings["　Excel③ 対象検索・書換え"] = (
            time.perf_counter() - edit_started
        )

    save_started = time.perf_counter()
    saved_content = _xml_fast_rebuild_zip(original_content, replacements)
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel④ xlsm保存"] = time.perf_counter() - save_started

    vba_started = time.perf_counter()
    _xml_fast_validate_vba_unchanged(original_vba, saved_content)
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑤ 保存後Excelを開く"] = time.perf_counter() - vba_started

    verify_started = time.perf_counter()
    verify_saved_workbook_lightweight(
        saved_content,
        original_sheets,
        required_sheets,
        changed_cells,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑥ 保存後内容検証"] = time.perf_counter() - verify_started
    return saved_content, changed_cells

def update_workbook_bytes(
    original_content,
    customer_name,
    product_name,
    proposed,
    diagnostic_timings=None,
):
    """H列の在庫本数・I列の配達本数を含む指定項目とL列配達数量を安全に更新する。"""
    if is_test_xml_fast_save_enabled():
        try:
            return _update_workbook_bytes_xml_fast(
                original_content,
                customer_name,
                product_name,
                proposed,
                diagnostic_timings=diagnostic_timings,
            )
        except Exception as exc:
            # PrivateテストのXML直接保存で少しでも問題があれば、
            # 本番で実績のある従来openpyxl方式へその場で戻す。
            # 診断版では、フォールバック理由だけを保存時間診断へ残す。
            if diagnostic_timings is not None:
                diagnostic_timings["XML直接保存エラー"] = f"{type(exc).__name__}: {exc}"
            pass
    diagnostic_step_started = time.perf_counter()
    workbook = load_workbook(BytesIO(original_content), keep_vba=True, data_only=False, read_only=False)
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel① マクロ付きExcelを開く"] = time.perf_counter() - diagnostic_step_started

    cached_row_values = None
    cached_values_seconds = 0.0

    original_sheets = list(workbook.sheetnames)
    changed_cells = []
    diagnostic_step_started = time.perf_counter()
    try:
        if DELIVERY_SHEET_NAME not in workbook.sheetnames or SHEET_NAME not in workbook.sheetnames:
            raise ValueError("必要なシート（次回配達日 または Sheet1）が見つかりません。")
        delivery_ws = workbook[DELIVERY_SHEET_NAME]
        product_rows, active_rows, _ = find_product_rows_by_usage(
            delivery_ws, customer_name, product_name
        )
        if not product_rows:
            raise ValueError("顧客名・商品名が一致する行が見つかりません。")
        if len(active_rows) > 1:
            raise ValueError("同じ顧客名・商品名の行が複数見つかりました。確認してください。")
        if not active_rows:
            raise ValueError("使用数量/日に値が入っている行が見つからないため編集できません。")

        product_row = active_rows[0]
        column_mapping = {
            "メーカー": 6,
            "在庫本数": 8,
            "本数": 9,
            "kg/本": 10,
            "配達日": 11,
        }

        event_changed = any(
            not same_excel_value(delivery_ws.cell(product_row, column).value, proposed.get(label))
            for label, column in {
                "在庫本数": 8,
                "本数": 9,
                "配達日": 11,
            }.items()
        )
        if event_changed:
            cached_values_started = time.perf_counter()
            try:
                cached_row_values = read_cached_delivery_row_values_lightweight(
                    original_content,
                    product_row,
                )
            except Exception:
                cached_row_values = None
            cached_values_seconds = time.perf_counter() - cached_values_started
            copy_previous_delivery_state_to_history(
                workbook,
                cached_row_values,
                delivery_ws,
                product_row,
                changed_cells,
            )

        for label, column in column_mapping.items():
            cell = delivery_ws.cell(product_row, column)
            new_value = proposed.get(label)
            if not same_excel_value(cell.value, new_value):
                cell.value = new_value
                changed_cells.append((DELIVERY_SHEET_NAME, product_row, column, new_value))

        # L列「配達数量」は、H列在庫本数＋I列本数（今回配達本数）の合計×kg/本。
        inventory_count = inventory_usage_number(proposed.get("在庫本数"))
        delivery_count = inventory_usage_number(proposed.get("本数"))
        kg_per_bottle = inventory_usage_number(proposed.get("kg/本"))
        delivery_quantity = None
        if inventory_count is not None or delivery_count is not None:
            if kg_per_bottle is None or kg_per_bottle <= 0:
                raise ValueError("在庫本数または本数を入力する場合は、kg/本に0より大きい数値が必要です。")
            delivery_quantity = (inventory_count or 0) + (delivery_count or 0)
            delivery_quantity *= kg_per_bottle
            if not math.isfinite(delivery_quantity):
                raise ValueError("配達数量を正しく計算できませんでした。")
            if delivery_quantity.is_integer():
                delivery_quantity = int(delivery_quantity)

        quantity_cell = delivery_ws.cell(product_row, 12)
        if not same_excel_value(quantity_cell.value, delivery_quantity):
            quantity_cell.value = delivery_quantity
            changed_cells.append(
                (DELIVERY_SHEET_NAME, product_row, 12, delivery_quantity)
            )

        # 商品・在庫の保存では住所とマップ位置を変更しない。
        # 住所・マップ位置は専用の編集処理だけで保存する。

        if not changed_cells:
            raise ValueError("変更された項目がありません。")
        enable_excel_recalculation(workbook)
        if diagnostic_timings is not None:
            diagnostic_timings["　Excel② 計算値用Excelを開く"] = cached_values_seconds
            diagnostic_timings["　Excel③ 対象検索・書換え"] = (
                time.perf_counter() - diagnostic_step_started - cached_values_seconds
            )

        output = BytesIO()
        diagnostic_step_started = time.perf_counter()
        workbook.save(output)
        if diagnostic_timings is not None:
            diagnostic_timings["　Excel④ xlsm保存"] = time.perf_counter() - diagnostic_step_started
    finally:
        workbook.close()

    saved_content = output.getvalue()
    diagnostic_step_started = time.perf_counter()
    try:
        with zipfile.ZipFile(BytesIO(saved_content), "r") as archive:
            if "xl/vbaProject.bin" not in archive.namelist():
                raise ValueError("保存後の検証でVBAプロジェクトを確認できません。")
    except zipfile.BadZipFile as exc:
        raise ValueError("保存後の検証でExcelファイルを開けません。") from exc
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑤ 保存後Excelを開く"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    verify_saved_workbook_lightweight(
        saved_content,
        original_sheets,
        {
            DELIVERY_SHEET_NAME,
            SHEET_NAME,
            DELIVERY_HISTORY_SHEET_NAME,
            MANAGEMENT_SHEET_NAME,
        },
        changed_cells,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑥ 保存後内容検証"] = time.perf_counter() - diagnostic_step_started
    return saved_content, changed_cells


def save_customer_excel_changes(customer_name, product_name, proposed):
    """Fast save path with backup, local validation, rev conflict protection, and hash verification."""
    diagnostic_total_started = time.perf_counter()
    diagnostic_timings = {}

    diagnostic_step_started = time.perf_counter()
    access_token = get_dropbox_access_token()
    target_path = get_dropbox_file_path()
    diagnostic_timings["Dropbox接続準備"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    original_content, download_response = download_dropbox_file(target_path, access_token)
    diagnostic_timings["Excel取得"] = time.perf_counter() - diagnostic_step_started
    if original_content is None:
        raise RuntimeError("\u6700\u65b0\u306eExcel\u3092\u53d6\u5f97\u3067\u304d\u307e\u305b\u3093\u3067\u3057\u305f\u3002\n" + dropbox_error_text(download_response))
    revision = get_download_revision(download_response)
    if not revision:
        raise RuntimeError("Dropbox\u306erev\u3092\u53d6\u5f97\u3067\u304d\u306a\u3044\u305f\u3081\u3001\u5b89\u5168\u306e\u305f\u3081\u66f4\u65b0\u3092\u4e2d\u6b62\u3057\u307e\u3057\u305f\u3002")

    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = f"{DROPBOX_BACKUP_FOLDER}/\u914d\u8eca\u4e88\u5b9a \u6b21\u90ce_{timestamp}.xlsm"

    # Dropbox-internal copy avoids uploading the same 1.8 MB file a second time.
    # The copied backup is hash-checked before the production file is touched.
    diagnostic_step_started = time.perf_counter()
    create_dropbox_backup(
        target_path,
        backup_path,
        original_content,
        access_token,
    )
    diagnostic_timings["バックアップ作成"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    saved_content, changed_cells = update_workbook_bytes(
        original_content,
        customer_name,
        product_name,
        proposed,
        diagnostic_timings=diagnostic_timings,
    )
    diagnostic_timings["Excel編集・保存・検証"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    upload_response = upload_dropbox_file(
        target_path,
        saved_content,
        access_token,
        mode="update",
        rev=revision,
    )
    diagnostic_timings["Dropbox本番保存"] = time.perf_counter() - diagnostic_step_started
    if upload_response.status_code == 409:
        raise RuntimeError("PC\u307e\u305f\u306f\u5225\u7aef\u672b\u3067Excel\u304c\u66f4\u65b0\u3055\u308c\u3066\u3044\u307e\u3059\u3002\u518d\u8aad\u307f\u8fbc\u307f\u3057\u3066\u304b\u3089\u3084\u308a\u76f4\u3057\u3066\u304f\u3060\u3055\u3044")
    if upload_response.status_code != 200:
        raise RuntimeError("\u672c\u756aExcel\u3092\u66f4\u65b0\u3067\u304d\u307e\u305b\u3093\u3067\u3057\u305f\u3002\u5fc5\u8981\u306aDropbox\u6a29\u9650\u306f files.content.write \u3067\u3059\u3002\n" + dropbox_error_text(upload_response))

    # The upload response contains size, rev, and content_hash. Verifying those
    # guarantees the bytes without downloading the full workbook again.
    diagnostic_step_started = time.perf_counter()
    upload_metadata = get_dropbox_response_metadata(upload_response)
    if not upload_metadata.get("content_hash") or upload_metadata.get("size") is None:
        upload_metadata = get_dropbox_file_metadata(target_path, access_token)
    confirmed_revision = verify_dropbox_file_metadata(
        upload_metadata,
        saved_content,
        previous_revision=revision,
    )
    diagnostic_timings["保存結果確認"] = time.perf_counter() - diagnostic_step_started

    # Use the already verified local bytes to rebuild the immediate-display JSON.
    diagnostic_step_started = time.perf_counter()
    cache_warning = refresh_fast_dropbox_cache_after_save(
        saved_content,
        confirmed_revision,
        access_token,
        previous_revision=revision,
        customer_name=customer_name,
        product_name=product_name,
        changed_cells=changed_cells,
        diagnostic_timings=diagnostic_timings,
    )
    diagnostic_timings["表示用データ更新"] = time.perf_counter() - diagnostic_step_started

    # Keep the existing exact rule: retain the newest 30 backups.
    diagnostic_step_started = time.perf_counter()
    cleanup_warning = trim_old_dropbox_backups(access_token, keep=30)
    diagnostic_timings["バックアップ整理"] = time.perf_counter() - diagnostic_step_started
    warnings = [warning for warning in (cleanup_warning, cache_warning) if warning]
    # 在庫保存では顧客Excelに関係するキャッシュだけを更新する。
    # 他機能のキャッシュを残し、保存直後の画面再表示を軽くする。
    diagnostic_step_started = time.perf_counter()
    clear_customer_excel_caches_after_save()
    diagnostic_timings["関連キャッシュ更新"] = time.perf_counter() - diagnostic_step_started
    diagnostic_save_seconds = time.perf_counter() - diagnostic_total_started
    return {
        "backup_path": backup_path,
        "updated_at": get_jst_now(),
        "changed_cells": changed_cells,
        "cleanup_warning": "\n".join(warnings),
        "diagnostic_timings": diagnostic_timings,
        "diagnostic_save_seconds": diagnostic_save_seconds,
    }



def update_delivery_history_record_bytes(
    original_content,
    customer_name,
    product_name,
    record_ref,
    proposed,
    diagnostic_timings=None,
):
    """納品履歴の指定1行を訂正する。新しい履歴行は追加しない。"""
    if is_test_xml_fast_save_enabled():
        try:
            return _update_delivery_history_record_bytes_xml_fast(
                original_content,
                customer_name,
                product_name,
                record_ref,
                proposed,
                diagnostic_timings=diagnostic_timings,
            )
        except Exception as exc:
            # 過去履歴以外・XML条件不一致・検証失敗は従来方式へ戻す。
            # 診断版では、フォールバック理由だけを保存時間診断へ残す。
            if diagnostic_timings is not None:
                diagnostic_timings["XML直接保存エラー"] = f"{type(exc).__name__}: {exc}"
            pass
    diagnostic_step_started = time.perf_counter()
    workbook = load_workbook(
        BytesIO(original_content),
        keep_vba=True,
        data_only=False,
        read_only=False,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel① マクロ付きExcelを開く"] = time.perf_counter() - diagnostic_step_started

    original_sheets = list(workbook.sheetnames)
    changed_cells = []
    diagnostic_step_started = time.perf_counter()
    try:
        source = str((record_ref or {}).get("source") or "").strip()
        row_number = int((record_ref or {}).get("row_number") or 0)
        expected_fingerprint = str(
            (record_ref or {}).get("fingerprint") or ""
        ).strip()
        expected_identity = tuple((record_ref or {}).get("identity") or ())

        if source == "current":
            sheet_name = DELIVERY_SHEET_NAME
        elif source == "history":
            sheet_name = DELIVERY_HISTORY_SHEET_NAME
        else:
            raise ValueError("修正する納品履歴を特定できません。")

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{sheet_name}シートが見つかりません。")
        if DELIVERY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("次回配達日シートが見つかりません。")
        if DELIVERY_HISTORY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("配達履歴シートが見つかりません。")
        if MANAGEMENT_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("管理シートが見つかりません。")

        ws = workbook[sheet_name]
        if row_number < 1 or row_number > ws.max_row:
            raise ValueError("修正対象の行が見つかりません。再読み込みしてください。")

        row_values = tuple(ws.cell(row_number, column).value for column in range(1, 15))
        actual_identity = delivery_history_identity(row_values)
        if expected_identity and actual_identity != expected_identity:
            raise ValueError("修正対象の顧客・商品が変わっています。再読み込みしてください。")
        if normalize_match_value(row_values[1] if len(row_values) >= 2 else None) != normalize_match_value(customer_name):
            raise ValueError("修正対象の顧客が一致しません。再読み込みしてください。")
        if normalize_match_value(row_values[4] if len(row_values) >= 5 else None) != normalize_match_value(product_name):
            raise ValueError("修正対象の商品が一致しません。再読み込みしてください。")

        actual_fingerprint = delivery_record_fingerprint(
            source,
            row_number,
            row_values,
        )
        if expected_fingerprint and actual_fingerprint != expected_fingerprint:
            raise ValueError(
                "この納品履歴はPCまたは別端末で変更されています。"
                "納品履歴を開き直してから修正してください。"
            )

        column_mapping = {
            "メーカー": 6,
            "在庫本数": 8,
            "本数": 9,
            "kg/本": 10,
            "配達日": 11,
        }
        old_values = {
            label: ws.cell(row_number, column).value
            for label, column in column_mapping.items()
        }

        for label, column in column_mapping.items():
            cell = ws.cell(row_number, column)
            new_value = proposed.get(label)
            if not same_excel_value(cell.value, new_value):
                cell.value = new_value
                changed_cells.append((sheet_name, row_number, column, new_value))

        inventory_changed = not same_excel_value(
            old_values.get("在庫本数"),
            proposed.get("在庫本数"),
        )
        delivery_changed = not same_excel_value(
            old_values.get("本数"),
            proposed.get("本数"),
        )
        kg_changed = not same_excel_value(
            old_values.get("kg/本"),
            proposed.get("kg/本"),
        )
        date_changed = not same_excel_value(
            old_values.get("配達日"),
            proposed.get("配達日"),
        )

        quantity_changed = False
        should_recalculate_quantity = inventory_changed or delivery_changed
        if kg_changed:
            # H/Iがある新しい形式の行だけ、kg/本の訂正を配達数量へ反映する。
            # H/Iが空欄の古い履歴では、既存の配達数量を消さずに保持する。
            should_recalculate_quantity = should_recalculate_quantity or (
                inventory_usage_number(proposed.get("在庫本数")) is not None
                or inventory_usage_number(proposed.get("本数")) is not None
            )

        if should_recalculate_quantity:
            inventory_count = inventory_usage_number(proposed.get("在庫本数"))
            delivery_count = inventory_usage_number(proposed.get("本数"))
            kg_per_bottle = inventory_usage_number(proposed.get("kg/本"))
            corrected_quantity = None
            if inventory_count is not None or delivery_count is not None:
                if kg_per_bottle is None or kg_per_bottle <= 0:
                    raise ValueError(
                        "在庫本数または本数を入れる場合は、kg/本に0より大きい数値が必要です。"
                    )
                corrected_quantity = (
                    (inventory_count or 0) + (delivery_count or 0)
                ) * kg_per_bottle
                if not math.isfinite(corrected_quantity):
                    raise ValueError("配達数量を正しく計算できませんでした。")
                if corrected_quantity.is_integer():
                    corrected_quantity = int(corrected_quantity)

            quantity_cell = ws.cell(row_number, 12)
            if not same_excel_value(quantity_cell.value, corrected_quantity):
                quantity_cell.value = corrected_quantity
                changed_cells.append(
                    (sheet_name, row_number, 12, corrected_quantity)
                )
                quantity_changed = True

        # 配達履歴シートは数式ではないため、訂正後の次回配達予定も同じ行で再計算する。
        # 次回配達日シートは従来のM列数式をそのまま維持し、Excel再計算に任せる。
        if source == "history" and (date_changed or quantity_changed):
            corrected_values = [
                ws.cell(row_number, column).value
                for column in range(1, 15)
            ]
            corrected_next_delivery, _ = calculate_delivery_values(corrected_values)
            next_cell = ws.cell(row_number, 13)
            if not same_excel_value(next_cell.value, corrected_next_delivery):
                next_cell.value = corrected_next_delivery
                changed_cells.append(
                    (sheet_name, row_number, 13, corrected_next_delivery)
                )

        if not changed_cells:
            raise ValueError("変更された項目がありません。")

        enable_excel_recalculation(workbook)
        if diagnostic_timings is not None:
            diagnostic_timings["　Excel③ 対象検索・書換え"] = time.perf_counter() - diagnostic_step_started

        output = BytesIO()
        diagnostic_step_started = time.perf_counter()
        workbook.save(output)
        if diagnostic_timings is not None:
            diagnostic_timings["　Excel④ xlsm保存"] = time.perf_counter() - diagnostic_step_started
    finally:
        workbook.close()

    saved_content = output.getvalue()
    diagnostic_step_started = time.perf_counter()
    try:
        with zipfile.ZipFile(BytesIO(saved_content), "r") as archive:
            if "xl/vbaProject.bin" not in archive.namelist():
                raise ValueError("保存後の検証でVBAプロジェクトを確認できません。")
    except zipfile.BadZipFile as exc:
        raise ValueError("保存後の検証でExcelファイルを開けません。") from exc
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑤ 保存後Excelを開く"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    verify_saved_workbook_lightweight(
        saved_content,
        original_sheets,
        {
            DELIVERY_SHEET_NAME,
            DELIVERY_HISTORY_SHEET_NAME,
            MANAGEMENT_SHEET_NAME,
            SHEET_NAME,
        },
        changed_cells,
    )
    if diagnostic_timings is not None:
        diagnostic_timings["　Excel⑥ 保存後内容検証"] = time.perf_counter() - diagnostic_step_started
    return saved_content, changed_cells


def save_customer_delivery_history_correction(
    customer_name,
    product_name,
    record_ref,
    proposed,
):
    """バックアップ・競合防止・ハッシュ確認付きで納品履歴の1件を訂正する。"""
    diagnostic_total_started = time.perf_counter()
    diagnostic_timings = {}

    diagnostic_step_started = time.perf_counter()
    access_token = get_dropbox_access_token()
    target_path = get_dropbox_file_path()
    diagnostic_timings["Dropbox接続準備"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    original_content, download_response = download_dropbox_file(
        target_path,
        access_token,
    )
    diagnostic_timings["Excel取得"] = time.perf_counter() - diagnostic_step_started
    if original_content is None:
        raise RuntimeError(
            "最新のExcelを取得できませんでした。\n"
            + dropbox_error_text(download_response)
        )
    revision = get_download_revision(download_response)
    if not revision:
        raise RuntimeError(
            "Dropboxのrevを取得できないため、安全のため更新を中止しました。"
        )

    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = f"{DROPBOX_BACKUP_FOLDER}/配車予定 次郎_{timestamp}.xlsm"
    diagnostic_step_started = time.perf_counter()
    create_dropbox_backup(
        target_path,
        backup_path,
        original_content,
        access_token,
    )
    diagnostic_timings["バックアップ作成"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    saved_content, changed_cells = update_delivery_history_record_bytes(
        original_content,
        customer_name,
        product_name,
        record_ref,
        proposed,
        diagnostic_timings=diagnostic_timings,
    )
    diagnostic_timings["Excel編集・保存・検証"] = time.perf_counter() - diagnostic_step_started

    diagnostic_step_started = time.perf_counter()
    upload_response = upload_dropbox_file(
        target_path,
        saved_content,
        access_token,
        mode="update",
        rev=revision,
    )
    diagnostic_timings["Dropbox本番保存"] = time.perf_counter() - diagnostic_step_started
    if upload_response.status_code == 409:
        raise RuntimeError(
            "PCまたは別端末でExcelが更新されています。"
            "再読み込みしてからやり直してください"
        )
    if upload_response.status_code != 200:
        raise RuntimeError(
            "本番Excelを更新できませんでした。\n"
            + dropbox_error_text(upload_response)
        )

    diagnostic_step_started = time.perf_counter()
    upload_metadata = get_dropbox_response_metadata(upload_response)
    if not upload_metadata.get("content_hash") or upload_metadata.get("size") is None:
        upload_metadata = get_dropbox_file_metadata(target_path, access_token)
    confirmed_revision = verify_dropbox_file_metadata(
        upload_metadata,
        saved_content,
        previous_revision=revision,
    )
    diagnostic_timings["保存結果確認"] = time.perf_counter() - diagnostic_step_started

    # 過去履歴の訂正でも予想使用量が変わるため、表示用JSONを必ず作り直す。
    diagnostic_step_started = time.perf_counter()
    cache_warning = refresh_fast_dropbox_cache_after_save(
        saved_content,
        confirmed_revision,
        access_token,
        previous_revision=revision,
        customer_name=customer_name,
        product_name=product_name,
        changed_cells=changed_cells,
        diagnostic_timings=diagnostic_timings,
    )
    diagnostic_timings["表示用データ更新"] = time.perf_counter() - diagnostic_step_started
    diagnostic_step_started = time.perf_counter()
    cleanup_warning = trim_old_dropbox_backups(access_token, keep=30)
    diagnostic_timings["バックアップ整理"] = time.perf_counter() - diagnostic_step_started
    warnings = [warning for warning in (cleanup_warning, cache_warning) if warning]
    # 納品履歴の訂正でも、顧客Excelに関係するキャッシュだけを更新する。
    # 写真・メモ・OneDrive・配車表・取引先など、納品履歴修正と無関係な
    # キャッシュは残し、保存直後の画面再表示を重くしない。
    diagnostic_step_started = time.perf_counter()
    clear_customer_excel_caches_after_save()
    diagnostic_timings["関連キャッシュ更新"] = time.perf_counter() - diagnostic_step_started
    diagnostic_save_seconds = time.perf_counter() - diagnostic_total_started
    return {
        "backup_path": backup_path,
        "updated_at": get_jst_now(),
        "changed_cells": changed_cells,
        "cleanup_warning": "\n".join(warnings),
        "diagnostic_timings": diagnostic_timings,
        "diagnostic_save_seconds": diagnostic_save_seconds,
    }



@st.cache_data(ttl=60, show_spinner=False)
def read_customer_map_values_from_bytes(content, customer_name):
    """Sheet1の表示値で顧客を探し、J列住所・K列マップ位置を返す。"""
    workbook = load_workbook(BytesIO(content), keep_vba=True, data_only=False, read_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError("Sheet1が見つかりません。")
        ws = workbook[SHEET_NAME]
        rows = find_sheet1_customer_rows(workbook, customer_name)
        if not rows:
            raise ValueError("Sheet1のB列に表示されている顧客名と一致する行が見つかりません。")
        first_row = rows[0]
        return {
            "住所": ws.cell(first_row, SHEET1_ADDRESS_COLUMN).value,
            "マップ位置": ws.cell(first_row, SHEET1_MAP_COLUMN).value,
            "顧客一致件数": len(rows),
        }
    finally:
        workbook.close()


def update_customer_map_workbook_bytes(original_content, customer_name, address, map_location):
    """Sheet1のJ列住所・K列マップ位置だけを更新する。"""
    workbook = load_workbook(BytesIO(original_content), keep_vba=True, data_only=False, read_only=False)
    original_sheets = list(workbook.sheetnames)
    changed_cells = []
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError("Sheet1が見つかりません。")
        ws = workbook[SHEET_NAME]
        rows = find_sheet1_customer_rows(workbook, customer_name)
        if not rows:
            raise ValueError("Sheet1のB列に表示されている顧客名と一致する行が見つかりません。")

        if not normalize_match_value(ws.cell(1, SHEET1_ADDRESS_COLUMN).value):
            ws.cell(1, SHEET1_ADDRESS_COLUMN).value = "住所"
            changed_cells.append((SHEET_NAME, 1, SHEET1_ADDRESS_COLUMN, "住所"))
        if not normalize_match_value(ws.cell(1, SHEET1_MAP_COLUMN).value):
            ws.cell(1, SHEET1_MAP_COLUMN).value = "マップ位置"
            changed_cells.append((SHEET_NAME, 1, SHEET1_MAP_COLUMN, "マップ位置"))

        for row in rows:
            for label, column, new_value in (
                ("住所", SHEET1_ADDRESS_COLUMN, address),
                ("マップ位置", SHEET1_MAP_COLUMN, map_location),
            ):
                cell = ws.cell(row, column)
                if not same_excel_value(cell.value, new_value):
                    cell.value = new_value
                    changed_cells.append((SHEET_NAME, row, column, new_value))

        if not changed_cells:
            raise ValueError("変更された項目がありません。")

        enable_excel_recalculation(workbook)
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    saved_content = output.getvalue()
    verified = load_workbook(BytesIO(saved_content), keep_vba=True, data_only=False, read_only=False)
    try:
        if list(verified.sheetnames) != original_sheets:
            raise ValueError("保存後にシート構成が変わったため、更新を中止しました。")
        if verified.vba_archive is None:
            raise ValueError("保存後の検証でVBAプロジェクトを確認できません。")
        for sheet, row, column, expected in changed_cells:
            actual = verified[sheet].cell(row, column).value
            if not same_excel_value(actual, expected):
                coordinate = verified[sheet].cell(row, column).coordinate
                raise ValueError(f"保存後の検証で{sheet}!{coordinate}の値が一致しません。")
    finally:
        verified.close()

    return saved_content, changed_cells


def save_customer_map_changes(customer_name, address, map_location):
    """更新前バックアップ、rev競合防止付きで住所・マップ位置を保存する。"""
    access_token = get_dropbox_access_token()
    target_path = get_dropbox_file_path()
    original_content, download_response = download_dropbox_file(target_path, access_token)
    if original_content is None:
        raise RuntimeError("最新のExcelを取得できませんでした。\n" + dropbox_error_text(download_response))

    revision = get_download_revision(download_response)
    if not revision:
        raise RuntimeError("Dropboxのrevを取得できないため、安全のため更新を中止しました。")

    ensure_dropbox_backup_folder(access_token)
    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = f"{DROPBOX_BACKUP_FOLDER}/配車予定 次郎_{timestamp}.xlsm"
    backup_response = upload_dropbox_file(
        backup_path,
        original_content,
        access_token,
        mode="add",
    )
    if backup_response.status_code != 200:
        raise RuntimeError(
            "バックアップを作成できないため、本番ファイルは更新しません。\n"
            + dropbox_error_text(backup_response)
        )

    saved_content, changed_cells = update_customer_map_workbook_bytes(
        original_content,
        customer_name,
        address,
        map_location,
    )
    upload_response = upload_dropbox_file(
        target_path,
        saved_content,
        access_token,
        mode="update",
        rev=revision,
    )
    if upload_response.status_code == 409:
        raise RuntimeError(
            "PCまたは別端末でExcelが更新されています。再読み込みしてからやり直してください"
        )
    if upload_response.status_code != 200:
        raise RuntimeError(
            "本番Excelを更新できませんでした。\n"
            + dropbox_error_text(upload_response)
        )

    # Dropbox上のファイルを読み直し、住所・地図が正しいセルへ入ったことを確認する。
    confirm_dropbox_upload(target_path, access_token, changed_cells)

    cleanup_warning = trim_old_dropbox_backups(access_token, keep=30)
    st.cache_data.clear()
    return {
        "backup_path": backup_path,
        "updated_at": get_jst_now(),
        "changed_cells": changed_cells,
        "cleanup_warning": cleanup_warning,
    }


def get_dropbox_file_path():
    """Dropboxから直接取得するファイルパスを返す"""
    path = str(DROPBOX_FILE_PATH or "").strip()

    if not path:
        return DROPBOX_DEFAULT_FILE_PATH

    return path


def show_dropbox_download_error(path, response):
    st.error("DropboxからExcelファイルを直接取得できませんでした。")
    st.write("Dropboxの指定パスを確認してください。")

    st.write("現在アプリが取得しようとしたパス：")
    st.code(path)
    st.write("Secretsに設定すべき正しい値：")
    st.code(f'DROPBOX_FILE_PATH = "{DROPBOX_DEFAULT_FILE_PATH}"')

    if response is not None:
        st.write(f"Dropboxからの応答コード：{response.status_code}")
        try:
            error_body = json.dumps(response.json(), ensure_ascii=False, indent=2)
        except Exception:
            error_body = response.text
        st.code(error_body)

    st.write("よくある原因：")
    st.write("- パスの先頭に不要なフォルダ名が入っている")
    st.write("- 全角スペースが半角スペースに変わっている")
    st.write("- Dropbox上でファイル名またはフォルダ名が変更された")
    st.stop()


@st.cache_data(ttl=60, show_spinner=False)
def get_cached_dropbox_excel_content():
    """同じExcelを画面操作ごとに再取得せず、1分間だけ共有する。"""
    access_token = get_dropbox_access_token()
    dropbox_file_path = get_dropbox_file_path()
    content, response = download_dropbox_file(dropbox_file_path, access_token)
    if content is None:
        raise RuntimeError(dropbox_error_text(response))
    return content


def read_excel_from_dropbox_api():
    """Dropbox APIでExcelをダウンロードして読み込む"""
    if not has_dropbox_auth_config():
        st.error("Dropbox API設定が不足しています。")
        st.write("secrets.toml に DROPBOX_APP_KEY / DROPBOX_APP_SECRET / DROPBOX_REFRESH_TOKEN を設定してください。")
        st.stop()

    dropbox_file_path = get_dropbox_file_path()
    try:
        content = get_cached_dropbox_excel_content()
    except Exception:
        # 従来の詳しいエラー表示を維持するため、失敗時だけ直接取得する。
        access_token = get_dropbox_access_token()
        content, response = download_dropbox_file(dropbox_file_path, access_token)
        if content is None:
            show_dropbox_download_error(dropbox_file_path, response)

    return BytesIO(content)


def read_excel_local():
    """同じフォルダにあるローカルExcelを読み込む"""
    excel_path = Path(EXCEL_FILE)

    if not excel_path.exists():
        st.error(f"Excelファイルが見つかりません：{EXCEL_FILE}")
        st.stop()

    return excel_path

# =========================
# メモ帳（Supabase保存）
# =========================
def get_jst_now():
    """日本時間の現在日時を返す"""
    return datetime.now(timezone(timedelta(hours=9)))


def format_note_datetime(value):
    """ISO形式の日時を見やすく表示する"""
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone(timedelta(hours=9)))
        return dt.strftime("%Y/%m/%d %H:%M")
    except Exception:
        return clean_value(value, blank_text="")


def get_supabase_key():
    """Supabaseへ接続するキーを返す。StreamlitではSecrets内だけに置く。"""
    return str(SUPABASE_SECRET_KEY or SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY or "").strip()


def has_supabase_config():
    return bool(str(SUPABASE_URL or "").strip() and get_supabase_key())


def get_supabase_notes_table():
    table_name = str(SUPABASE_NOTES_TABLE or "notes").strip()

    if not table_name.replace("_", "").isalnum():
        st.error("Supabaseのメモ帳テーブル名が正しくありません。")
        st.write("SUPABASE_NOTES_TABLE は英数字とアンダースコアだけで指定してください。")
        st.stop()

    return table_name


def get_supabase_notes_url():
    base_url = str(SUPABASE_URL or "").strip().rstrip("/")
    table_name = urllib.parse.quote(get_supabase_notes_table(), safe="")
    return f"{base_url}/rest/v1/{table_name}"


def get_supabase_headers(prefer=None):
    key = get_supabase_key()
    headers = {
        "apikey": key,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    # 新しいsb_secret_/sb_publishable_キーはJWTではないため、
    # Authorization: Bearerには入れず、apikeyヘッダーだけで送る。
    # 旧service_role/anonのJWTキーは従来どおりBearerにも設定する。
    if not key.startswith(("sb_secret_", "sb_publishable_")):
        headers["Authorization"] = f"Bearer {key}"
    if prefer:
        headers["Prefer"] = prefer
    return headers


def make_line_status_id(customer_name):
    """顧客名から、LINE状態保存用の重複しないIDを作る。"""
    customer = clean_value(customer_name, blank_text="")
    digest = hashlib.sha256(customer.encode("utf-8")).hexdigest()
    return f"{LINE_STATUS_NOTE_PREFIX}{digest}"


@st.cache_data(ttl=30, show_spinner=False)
def load_line_statuses_from_supabase():
    """LINE接続中の顧客名だけを小さなデータとして読み込む。"""
    if not has_supabase_config():
        return {}

    try:
        response = requests.get(
            get_supabase_notes_url(),
            headers=get_supabase_headers(),
            params={
                "select": "customer_name",
                "id": f"like.{LINE_STATUS_NOTE_PREFIX}*",
                "limit": "5000",
            },
            timeout=15,
        )
    except Exception:
        return {}

    if response.status_code != 200:
        return {}

    try:
        rows = response.json()
    except Exception:
        return {}

    if not isinstance(rows, list):
        return {}

    return {
        clean_value(row.get("customer_name"), blank_text=""): True
        for row in rows
        if clean_value(row.get("customer_name"), blank_text="")
    }


def get_line_connected(customer_name):
    customer = clean_value(customer_name, blank_text="")
    return bool(load_line_statuses_from_supabase().get(customer, False))


def save_line_connected(customer_name, connected):
    """Excelを変更せず、LINE状態だけをSupabaseへ保存する。"""
    previous_connected = get_line_connected(customer_name)
    if not has_supabase_config():
        st.error("LINE状態を保存するための接続設定がありません。")
        return False

    customer = clean_value(customer_name, blank_text="")
    status_id = make_line_status_id(customer)

    try:
        if connected:
            response = requests.post(
                get_supabase_notes_url(),
                headers=get_supabase_headers(
                    prefer="resolution=merge-duplicates,return=minimal"
                ),
                json={
                    "id": status_id,
                    "customer_name": customer,
                    "body": LINE_STATUS_BODY,
                    "created_at": get_jst_now().isoformat(),
                },
                timeout=15,
            )
            success = response.status_code in (200, 201)
        else:
            response = requests.delete(
                get_supabase_notes_url(),
                headers=get_supabase_headers(prefer="return=minimal"),
                params={"id": f"eq.{status_id}"},
                timeout=15,
            )
            success = response.status_code in (200, 204)
    except Exception as exc:
        st.error(f"LINE状態を保存できませんでした：{exc}")
        return False

    if not success:
        st.error("LINE状態を保存できませんでした。")
        return False

    load_line_statuses_from_supabase.clear()
    warning = record_change_history_safely(
        "顧客",
        "",
        customer,
        "変更",
        {"LINE状態": ("○" if previous_connected else "×", "○" if connected else "×")},
        section="LINE状態",
    )
    remember_change_history_warning(warning)
    return True


def show_supabase_config_error():
    st.error("メモ帳を使うにはSupabase設定が必要です。")
    st.write("Streamlit Cloud の Secrets に以下を追加してください。")
    st.code(
        '\n'.join(
            [
                'SUPABASE_URL = "https://xxxx.supabase.co"',
                'SUPABASE_SECRET_KEY = "SupabaseのSecret key"',
                'SUPABASE_NOTES_TABLE = "notes"',
            ]
        )
    )
    st.stop()


def show_supabase_response_error(action, response):
    st.error(f"メモ帳をSupabaseに{action}できませんでした。")
    if response is not None:
        st.write(f"Supabaseからの応答コード：{response.status_code}")
        try:
            st.code(json.dumps(response.json(), ensure_ascii=False, indent=2))
        except Exception:
            st.code(response.text)
    st.stop()


@st.cache_data(ttl=30, show_spinner=False)
def load_notes_from_supabase(customer_name=None, limit=500):
    """Supabaseのnotesテーブルからメモを新しい順で読み込む"""
    if not has_supabase_config():
        show_supabase_config_error()

    params = {
        "select": "id,customer_name,body,created_at",
        "order": "created_at.desc",
        "limit": str(limit),
        "id": f"not.like.{LINE_STATUS_NOTE_PREFIX}*",
    }

    if customer_name is not None:
        target = clean_value(customer_name, blank_text="")
        params["customer_name"] = f"eq.{target}"

    try:
        response = requests.get(
            get_supabase_notes_url(),
            headers=get_supabase_headers(),
            params=params,
            timeout=30,
        )
    except Exception as e:
        st.error("メモ帳の読み込み中にSupabaseへの接続に失敗しました。")
        st.exception(e)
        st.stop()

    if response.status_code != 200:
        show_supabase_response_error("読み込み", response)

    try:
        notes = response.json()
    except Exception:
        st.error("Supabaseから返ったメモ帳データの形式が正しくありません。")
        st.stop()

    if not isinstance(notes, list):
        return []

    # ホーム専用の「やることメモ」は、既存の顧客メモ・取引先メモには混ぜない。
    return [
        note
        for note in notes
        if clean_value(note.get("customer_name"), blank_text="") != HOME_TODO_CUSTOMER_NAME
    ]


@st.cache_data(ttl=300, show_spinner=False)
def load_note_presence_index():
    """顧客メモが存在する顧客名だけを、小さい索引としてSupabaseから読む。"""
    if not has_supabase_config():
        return tuple()

    customer_names = set()
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_name",
            "order": "id.asc",
            "limit": str(page_size),
            "offset": str(offset),
            "id": f"not.like.{LINE_STATUS_NOTE_PREFIX}*",
        }
        try:
            response = requests.get(
                get_supabase_notes_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("メモの索引読み込み中にSupabaseへ接続できませんでした。") from exc
        if response.status_code != 200:
            detail = str(response.text or "").strip()[:500]
            message = f"メモの索引を読み込めませんでした（{response.status_code}）。"
            if detail:
                message += f" {detail}"
            raise RuntimeError(message)
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返ったメモ索引の形式が正しくありません。")

        for note in page:
            if not isinstance(note, dict):
                continue
            customer_name = clean_value(note.get("customer_name"), blank_text="").strip()
            if customer_name and customer_name != HOME_TODO_CUSTOMER_NAME:
                customer_names.add(customer_name)

        if len(page) < page_size:
            break
        offset += page_size

    return tuple(sorted(customer_names))

def customer_has_notes(customer_name):
    """存在索引だけで、その顧客にメモがあるか確認する。"""
    target = clean_value(customer_name, blank_text="").strip()
    return target in set(load_note_presence_index())

def insert_note_to_supabase(note):
    """Supabaseのnotesテーブルへメモを1件追加する"""
    if not has_supabase_config():
        show_supabase_config_error()

    try:
        response = requests.post(
            get_supabase_notes_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            json=note,
            timeout=30,
        )
    except Exception as e:
        st.error("メモ帳の保存中にSupabaseへの接続に失敗しました。")
        st.exception(e)
        st.stop()

    if response.status_code not in (200, 201):
        show_supabase_response_error("保存", response)
    load_notes_from_supabase.clear()
    try:
        load_note_presence_index.clear()
    except Exception:
        pass


def delete_note_from_supabase(note_id):
    """Supabaseのnotesテーブルからメモを1件削除する"""
    target_id = clean_value(note_id, blank_text="")
    if not target_id:
        st.warning("削除するメモが見つかりません。")
        return False

    if not has_supabase_config():
        show_supabase_config_error()

    try:
        response = requests.delete(
            get_supabase_notes_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            params={"id": f"eq.{target_id}"},
            timeout=30,
        )
    except Exception as e:
        st.error("メモ帳の削除中にSupabaseへの接続に失敗しました。")
        st.exception(e)
        st.stop()

    if response.status_code not in (200, 204):
        show_supabase_response_error("削除", response)

    load_notes_from_supabase.clear()
    try:
        load_note_presence_index.clear()
    except Exception:
        pass
    return True


def make_note_id():
    return get_jst_now().strftime("%Y%m%d%H%M%S%f")


def add_note(customer_name, body):
    """顧客名に紐づくメモを1件追加する"""
    note_text = str(body or "").strip()
    if not note_text:
        st.warning("メモ本文を入力してください。")
        return False

    now = get_jst_now().isoformat()
    note = {
        "id": make_note_id(),
        "customer_name": clean_value(customer_name, blank_text=""),
        "body": note_text,
        "created_at": now,
    }

    insert_note_to_supabase(note)
    return True


def get_notes_for_customer(customer_name):
    return load_notes_from_supabase(customer_name=customer_name)


def encode_home_todo_body(text, completed=False):
    payload = {
        "text": str(text or "").strip(),
        "completed": bool(completed),
    }
    return HOME_TODO_BODY_PREFIX + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def decode_home_todo_body(body):
    raw = str(body or "")
    if not raw.startswith(HOME_TODO_BODY_PREFIX):
        return {"text": raw.strip(), "completed": False}
    try:
        payload = json.loads(raw[len(HOME_TODO_BODY_PREFIX):])
    except Exception:
        return {"text": raw[len(HOME_TODO_BODY_PREFIX):].strip(), "completed": False}
    return {
        "text": str(payload.get("text") or "").strip(),
        "completed": bool(payload.get("completed", False)),
    }


@st.cache_data(ttl=30, show_spinner=False)
def load_home_todos_from_supabase(limit=500):
    """ホーム画面専用のやることメモを読み込む。"""
    if not has_supabase_config():
        return []
    try:
        response = requests.get(
            get_supabase_notes_url(),
            headers=get_supabase_headers(),
            params={
                "select": "id,customer_name,body,created_at",
                "customer_name": f"eq.{HOME_TODO_CUSTOMER_NAME}",
                "order": "created_at.desc",
                "limit": str(limit),
            },
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError(f"やることメモを読み込めませんでした：{exc}") from exc
    if response.status_code != 200:
        raise RuntimeError(f"やることメモを読み込めませんでした（{response.status_code}）。")
    try:
        rows = response.json()
    except Exception as exc:
        raise RuntimeError("やることメモのデータ形式が正しくありません。") from exc
    if not isinstance(rows, list):
        return []
    todos = []
    for row in rows:
        parsed = decode_home_todo_body(row.get("body"))
        if not parsed["text"]:
            continue
        todos.append({**row, **parsed})
    # 未完了を先にし、同じ状態では新しい順を保つ。
    todos.sort(key=lambda item: bool(item.get("completed")))
    return todos


def save_home_todo(text, existing=None):
    todo_text = str(text or "").strip()
    if not todo_text:
        raise ValueError("メモを入力してください。")
    existing = existing or {}
    note_id = clean_value(existing.get("id"), blank_text="") or (
        HOME_TODO_ID_PREFIX + make_note_id()
    )
    completed = bool(existing.get("completed", False))
    if existing:
        try:
            response = requests.patch(
                get_supabase_notes_url(),
                headers=get_supabase_headers(prefer="return=minimal"),
                params={"id": f"eq.{note_id}"},
                json={"body": encode_home_todo_body(todo_text, completed)},
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError(f"やることメモを更新できませんでした：{exc}") from exc
        if response.status_code not in (200, 204):
            raise RuntimeError(f"やることメモを更新できませんでした（{response.status_code}）。")
    else:
        insert_note_to_supabase(
            {
                "id": note_id,
                "customer_name": HOME_TODO_CUSTOMER_NAME,
                "body": encode_home_todo_body(todo_text, False),
                "created_at": get_jst_now().isoformat(),
            }
        )
    load_home_todos_from_supabase.clear()
    load_notes_from_supabase.clear()


def set_home_todo_completed(todo, completed):
    note_id = clean_value(todo.get("id"), blank_text="")
    if not note_id:
        raise ValueError("対象のメモが見つかりません。")
    try:
        response = requests.patch(
            get_supabase_notes_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            params={"id": f"eq.{note_id}"},
            json={
                "body": encode_home_todo_body(
                    todo.get("text", ""),
                    completed,
                )
            },
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError(f"完了状態を更新できませんでした：{exc}") from exc
    if response.status_code not in (200, 204):
        raise RuntimeError(f"完了状態を更新できませんでした（{response.status_code}）。")
    load_home_todos_from_supabase.clear()
    load_notes_from_supabase.clear()


def delete_home_todo(todo):
    note_id = clean_value(todo.get("id"), blank_text="")
    if not note_id:
        raise ValueError("対象のメモが見つかりません。")
    if not delete_note_from_supabase(note_id):
        raise RuntimeError("やることメモを削除できませんでした。")
    load_home_todos_from_supabase.clear()


@st.dialog("やることメモを追加")
def show_home_todo_add_dialog():
    text = st.text_area(
        "メモ",
        key="home_todo_add_text",
        height=140,
        placeholder="例：7月20日の週で青雲、醤油粕",
        help=VOICE_INPUT_HELP,
    )
    save_col, cancel_col = st.columns(2)
    with save_col:
        if st.button("追加する", type="primary", use_container_width=True):
            try:
                save_home_todo(text)
                st.session_state.pop("home_todo_add_text", None)
                st.session_state["home_todo_message"] = "メモを追加しました。"
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    with cancel_col:
        if st.button("キャンセル", use_container_width=True):
            st.rerun()


@st.dialog("やることメモを編集")
def show_home_todo_edit_dialog(todo):
    note_id = clean_value(todo.get("id"), blank_text="")
    text = st.text_area(
        "メモ",
        value=str(todo.get("text") or ""),
        key=f"home_todo_edit_text_{note_id}",
        height=140,
        help=VOICE_INPUT_HELP,
    )
    save_col, cancel_col = st.columns(2)
    with save_col:
        if st.button("保存する", type="primary", use_container_width=True):
            try:
                save_home_todo(text, existing=todo)
                st.session_state["home_todo_message"] = "メモを更新しました。"
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    with cancel_col:
        if st.button("キャンセル", use_container_width=True):
            st.rerun()


@st.dialog("やることメモを削除")
def show_home_todo_delete_dialog(todo):
    preview = str(todo.get("text") or "").strip()
    st.warning("このメモを削除します。")
    if preview:
        st.caption(preview if len(preview) <= 100 else preview[:97] + "…")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)
    with delete_col:
        if st.button("削除する", type="primary", use_container_width=True):
            try:
                delete_home_todo(todo)
                st.session_state["home_todo_message"] = "メモを削除しました。"
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    with cancel_col:
        if st.button("キャンセル", use_container_width=True):
            st.rerun()


def render_home_todo_section():
    """今のホーム画面を変えず、その最下部にやることメモだけを追加する。"""
    st.markdown("---")
    title_col, add_col = st.columns([4, 1])
    with title_col:
        st.subheader("☑️ やることメモ")
        st.caption("未完了を上に表示します。メモが多い場合は枠内をスクロールできます。")
    with add_col:
        if st.button("＋ 追加", key="home_todo_add_button", use_container_width=True):
            show_home_todo_add_dialog()

    message = st.session_state.pop("home_todo_message", None)
    if message:
        st.success(message)

    if not has_supabase_config():
        st.info("やることメモを使うには、現在のSupabase設定が必要です。")
        return

    try:
        todos = load_home_todos_from_supabase()
    except Exception as exc:
        st.warning(str(exc))
        return

    if not todos:
        st.info("やることメモはまだありません。右上の「＋ 追加」から登録できます。")
        return

    st.markdown(
        """
        <style>
        /* やることメモだけを対象にし、他のボタンやカードには影響させない。 */
        div[class*="st-key-home_todo_card_"] {
            background: #ffffff !important;
            border: 1px solid rgba(49, 51, 63, 0.18) !important;
            border-radius: 14px !important;
            padding: 0.55rem 0.7rem !important;
            margin-bottom: 0.65rem !important;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06) !important;
        }
        div[class*="st-key-home_todo_check_"] button {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            padding: 0 !important;
            min-height: 1.875rem !important;
            font-size: 1.875rem !important;
            line-height: 1 !important;
            color: #d8d0f2 !important;
        }
        div[class*="st-key-home_todo_check_"] button p {
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 1.875rem !important;
            height: 1.875rem !important;
            margin: 0 !important;
            border: 2.25px solid #d8d0f2 !important;
            border-radius: 0.4125rem !important;
            background: #d8d0f2 !important;
            box-shadow: none !important;
            font-size: 1.35rem !important;
            font-weight: 700 !important;
            line-height: 1 !important;
            color: #ffffff !important;
            text-shadow: none !important;
        }
        div[class*="st-key-home_todo_check_"] button:hover,
        div[class*="st-key-home_todo_check_"] button:focus,
        div[class*="st-key-home_todo_check_"] button:active {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    with st.container(height=HOME_TODO_LIST_HEIGHT, border=True):
        for index, todo in enumerate(todos):
            note_id = clean_value(todo.get("id"), blank_text=str(index))
            completed = bool(todo.get("completed"))
            todo_text = str(todo.get("text") or "").strip()
            button_text = todo_text.replace("\r", " ").replace("\n", " ") or "（内容なし）"

            # 1件ずつ白いカードにし、本文タップで編集を開く。
            with st.container(border=False, key=f"home_todo_card_{note_id}"):
                check_col, text_col, delete_col = st.columns([0.9, 5.7, 1.35], vertical_alignment="center")
                with check_col:
                    check_label = "✓" if completed else "\u00a0"
                    if st.button(
                        check_label,
                        key=f"home_todo_check_{note_id}",
                        help="未完了に戻す" if completed else "完了にする",
                        use_container_width=True,
                    ):
                        try:
                            set_home_todo_completed(todo, not completed)
                            st.rerun()
                        except Exception as exc:
                            st.error(str(exc))
                with text_col:
                    if st.button(
                        button_text,
                        key=f"home_todo_edit_{note_id}",
                        help="タップして編集",
                        use_container_width=True,
                        type="tertiary",
                    ):
                        show_home_todo_edit_dialog(todo)
                with delete_col:
                    if st.button(
                        "削除",
                        key=f"home_todo_delete_{note_id}",
                        use_container_width=True,
                    ):
                        show_home_todo_delete_dialog(todo)


def render_note_card(note, show_customer=True):
    customer_name = clean_value(note.get("customer_name"), blank_text="未設定")
    created_at = format_note_datetime(note.get("created_at", ""))
    body = html.escape(clean_value(note.get("body"), blank_text="")).replace("\n", "<br>")

    if show_customer:
        customer_link = build_customer_detail_link(customer_name, class_name="dispatch-month-link")
        meta = f"{html.escape(created_at)}　{customer_link}"
    else:
        meta = html.escape(created_at)

    st.markdown(
        f"""
        <div class="note-card">
            <div class="note-meta">{meta}</div>
            <div class="note-body">{body}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


NOTE_DELETE_SUCCESS_KEY = "note_delete_success_message"


def show_note_delete_success_message():
    message = st.session_state.pop(NOTE_DELETE_SUCCESS_KEY, None)
    if message:
        st.success(message)


@st.dialog("メモを削除")
def confirm_note_delete_dialog(note):
    """画面位置を保ったまま、メモ削除を確認して実行する。"""
    note_id = clean_value(note.get("id"), blank_text="")
    body = clean_value(note.get("body"), blank_text="").strip()
    preview = body if len(body) <= 80 else body[:77] + "…"

    st.warning("このメモを削除します。")
    if preview:
        st.caption(preview)
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"delete_note_dialog_yes_{note_id}",
            type="primary",
            use_container_width=True,
        ):
            if delete_note_from_supabase(note_id):
                st.session_state[NOTE_DELETE_SUCCESS_KEY] = "メモを削除しました。"
                st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
                st.rerun()

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"delete_note_dialog_no_{note_id}",
            use_container_width=True,
        ):
            st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
            st.rerun()


def render_note_delete_controls(note):
    note_id = clean_value(note.get("id"), blank_text="")
    if not note_id:
        return

    if st.button("削除", key=f"delete_note_start_{note_id}"):
        confirm_note_delete_dialog(note)


def show_customer_notes(customer_name):
    st.markdown("---")
    st.subheader("📝 この顧客のメモ")
    show_note_delete_success_message()
    st.caption(f"🎤 {VOICE_INPUT_HELP}")

    note_key = f"customer_note_input_{customer_name}"
    clear_note_key = f"clear_{note_key}"

    # Streamlitでは生成済みウィジェットの値を同じ実行中に変更できないため、
    # 保存後の次回実行で入力欄を空にする。
    if st.session_state.pop(clear_note_key, False):
        st.session_state[note_key] = ""

    note_text = st.text_area(
        "メモ本文",
        key=note_key,
        height=120,
        placeholder="例：次回は午前中希望。サンプル持参。など",
        help=VOICE_INPUT_HELP,
    )

    if st.button("メモを保存", key=f"save_customer_note_{customer_name}"):
        if add_note(customer_name, note_text):
            st.session_state[clear_note_key] = True
            st.session_state["note_save_success"] = customer_name
            st.rerun()

    if st.session_state.pop("note_save_success", None) == customer_name:
        st.success("メモを保存しました。")

    try:
        has_notes = customer_has_notes(customer_name)
    except Exception:
        # 索引確認に失敗した場合は、表示欠落を避けるため従来どおり詳細取得へ戻す。
        has_notes = None

    if has_notes is False:
        customer_notes = []
    else:
        customer_notes = get_notes_for_customer(customer_name)

    if not customer_notes:
        st.info("この顧客のメモはまだありません。")
        return

    st.markdown("#### メモ履歴")
    for note in customer_notes:
        render_note_card(note, show_customer=False)
        render_note_delete_controls(note)


def show_notes_page(df):
    show_back_home_button("notes_back_home")

    st.markdown("---")
    st.header("📝 メモ帳")
    show_note_delete_success_message()
    st.caption("全顧客のメモを新しい順で表示します。")

    notes = load_notes_from_supabase()

    if not notes:
        st.info("メモはまだありません。顧客詳細画面から保存できます。")
        return

    for note in notes:
        render_note_card(note, show_customer=True)
        render_note_delete_controls(note)


# =========================
# 顧客情報（Supabase保存）
# =========================
def get_supabase_customer_information_table():
    table_name = str(SUPABASE_CUSTOMER_INFO_TABLE or "customer_information").strip()
    if not table_name.replace("_", "").isalnum():
        raise RuntimeError("Supabaseの顧客情報テーブル名が正しくありません。")
    return table_name


def get_supabase_customer_information_url():
    base_url = str(SUPABASE_URL or "").strip().rstrip("/")
    table_name = urllib.parse.quote(get_supabase_customer_information_table(), safe="")
    return f"{base_url}/rest/v1/{table_name}"


def change_history_value(value):
    """変更前後の値を、JSONと画面表示で安定して扱える文字列へ変換する。"""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def normalize_change_history_changes(changes):
    """辞書または変更明細リストを、共通の変更明細リストへ整形する。"""
    normalized = []
    if isinstance(changes, dict):
        iterable = []
        for field_name, values in changes.items():
            if isinstance(values, dict):
                before = values.get("before")
                after = values.get("after")
            elif isinstance(values, (list, tuple)) and len(values) >= 2:
                before, after = values[0], values[1]
            else:
                before, after = "", values
            iterable.append(
                {
                    "field": field_name,
                    "before": before,
                    "after": after,
                }
            )
    elif isinstance(changes, list):
        iterable = changes
    else:
        iterable = []

    for item in iterable:
        if not isinstance(item, dict):
            continue
        field_name = clean_value(
            item.get("field") or item.get("項目"),
            blank_text="",
        )
        if not field_name:
            continue
        before = change_history_value(item.get("before"))
        after = change_history_value(item.get("after"))
        if before == after:
            continue
        normalized.append(
            {
                "field": field_name,
                "before": before,
                "after": after,
            }
        )
    return normalized


def clear_change_history_cache():
    try:
        load_change_history_page.clear()
    except Exception:
        pass


def record_change_history(
    target_type,
    target_id,
    target_name,
    action,
    changes,
    section="",
):
    """既存のcustomer_informationテーブルへ、通常顧客情報と分離して履歴を保存する。"""
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がないため変更履歴を保存できません。")

    target_type = clean_value(target_type, blank_text="")
    target_name = clean_value(target_name, blank_text="")
    action = clean_value(action, blank_text="")
    normalized_changes = normalize_change_history_changes(changes)
    if not target_type or not target_name or not action or not normalized_changes:
        return

    now = get_jst_now().isoformat()
    history_id = str(uuid.uuid4())
    content = json.dumps(
        {
            "version": CHANGE_HISTORY_VERSION,
            "target_type": target_type,
            "target_id": clean_value(target_id, blank_text=""),
            "target_name": target_name,
            "action": action,
            "section": clean_value(section, blank_text=""),
            "changes": normalized_changes,
            "source": "app",
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    payload = {
        "id": history_id,
        "customer_key": None,
        "customer_name": CHANGE_HISTORY_CUSTOMER,
        "field_name": target_type,
        "content": content,
        "sort_order": 0,
        "created_at": now,
        "updated_at": now,
    }
    try:
        response = requests.post(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            json=payload,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("変更履歴の保存中にSupabaseへ接続できませんでした。") from exc
    if response.status_code not in (200, 201):
        detail = str(response.text or "").strip()[:500]
        message = f"変更履歴を保存できませんでした（{response.status_code}）。"
        if detail:
            message += f" {detail}"
        raise RuntimeError(message)
    clear_change_history_cache()


def record_change_history_safely(*args, **kwargs):
    """本体の保存成功を取り消さず、履歴保存だけを警告として返す。"""
    try:
        record_change_history(*args, **kwargs)
        return ""
    except Exception as exc:
        return f"本体の保存は完了しましたが、変更履歴を保存できませんでした：{exc}"


def remember_change_history_warning(warning):
    if warning:
        st.session_state["change_history_warning"] = str(warning)


@st.cache_data(ttl=15, show_spinner=False)
def load_change_history_page(target_type="", start_iso="", offset=0, limit=CHANGE_HISTORY_PAGE_SIZE):
    """変更履歴を新しい順に必要件数だけ取得する。"""
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がありません。")
    params = {
        "select": "id,customer_name,field_name,content,created_at,updated_at",
        "customer_name": f"eq.{CHANGE_HISTORY_CUSTOMER}",
        "order": "created_at.desc,id.desc",
        "limit": str(int(limit) + 1),
        "offset": str(max(0, int(offset))),
    }
    if target_type:
        params["field_name"] = f"eq.{target_type}"
    if start_iso:
        params["created_at"] = f"gte.{start_iso}"
    try:
        response = requests.get(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(),
            params=params,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("変更履歴の読み込み中にSupabaseへ接続できませんでした。") from exc
    if response.status_code != 200:
        raise RuntimeError(
            f"変更履歴を読み込めませんでした（{response.status_code}）。"
        )
    rows = response.json()
    if not isinstance(rows, list):
        raise RuntimeError("Supabaseから返った変更履歴の形式が正しくありません。")
    return rows


def parse_change_history_row(row):
    payload = {}
    try:
        parsed = json.loads(str(row.get("content") or "{}"))
        if isinstance(parsed, dict):
            payload = parsed
    except Exception:
        payload = {}
    return {
        "id": clean_value(row.get("id"), blank_text=""),
        "created_at": clean_value(row.get("created_at"), blank_text=""),
        "target_type": clean_value(
            payload.get("target_type") or row.get("field_name"),
            blank_text="",
        ),
        "target_id": clean_value(payload.get("target_id"), blank_text=""),
        "target_name": clean_value(payload.get("target_name"), blank_text=""),
        "action": clean_value(payload.get("action"), blank_text=""),
        "section": clean_value(payload.get("section"), blank_text=""),
        "changes": normalize_change_history_changes(payload.get("changes", [])),
        "raw_content": clean_value(row.get("content"), blank_text=""),
    }


def change_history_rows_to_dataframe(rows):
    """変更履歴を、1変更項目につき1行のCSV向け表へ変換する。"""
    records = []
    for row in rows:
        parsed = parse_change_history_row(row)
        changes = parsed["changes"] or [
            {"field": "解析できない履歴", "before": "", "after": parsed["raw_content"]}
        ]
        for change in changes:
            records.append(
                {
                    "変更日時": parsed["created_at"],
                    "対象区分": parsed["target_type"],
                    "対象ID": parsed["target_id"],
                    "対象名": parsed["target_name"],
                    "操作": parsed["action"],
                    "変更箇所": parsed["section"],
                    "項目": change.get("field", ""),
                    "変更前": change.get("before", ""),
                    "変更後": change.get("after", ""),
                    "履歴ID": parsed["id"],
                }
            )
    return backup_dataframe(
        records,
        [
            "変更日時", "対象区分", "対象ID", "対象名", "操作",
            "変更箇所", "項目", "変更前", "変更後", "履歴ID",
        ],
    )


def display_change_history_value(value):
    """変更確認画面では、ISO形式の日時を日付だけで表示する。"""
    if value is None or value == "":
        return "（空欄）"
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y/%m/%d")

    text = str(value).strip()
    iso_date_match = re.fullmatch(
        r"(\d{4})-(\d{2})-(\d{2})(?:[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:?\d{2})?)?",
        text,
    )
    if iso_date_match:
        year, month, day = iso_date_match.groups()
        return f"{year}/{month}/{day}"
    return text


def change_history_target_url(parsed):
    """変更履歴カードから、既存の対象詳細画面へ移動するURLを返す。"""
    target_type = clean_value(parsed.get("target_type"), blank_text="").strip()
    target_name = clean_value(parsed.get("target_name"), blank_text="").strip()
    target_id = clean_value(parsed.get("target_id"), blank_text="").strip()

    if target_type == "顧客" and target_name:
        return make_app_url(page="detail", customer=target_name)
    if target_type == "仕入先" and target_id:
        return make_app_url(
            page="partner_detail",
            partner_id=target_id,
            partner_type="supplier",
        )
    if target_type == "運送会社" and target_id:
        return make_app_url(
            page="partner_detail",
            partner_id=target_id,
            partner_type="carrier",
        )
    return ""


def render_change_history_card(parsed):
    """変更履歴1件を、対象詳細へ移動できる1枚のカードとして表示する。"""
    title = "　".join(
        part for part in (
            parsed["target_type"],
            parsed["target_name"],
        ) if part
    ) or "変更履歴"
    meta = " ｜ ".join(
        part for part in (
            format_note_datetime(parsed["created_at"]),
            parsed["action"],
        ) if part
    )

    # クリック可能なカード内はspanだけで構成する。
    # st.markdownがリンク内のブロック要素を分割して、各項目が別カードに見えるのを防ぐ。
    parts = [
        '<span class="change-history-card-title">',
        html.escape(title),
        '</span>',
    ]
    if meta:
        parts.extend([
            '<span class="change-history-card-meta">',
            html.escape(meta),
            '</span>',
        ])
    if parsed["section"]:
        parts.extend([
            '<span class="change-history-card-section">変更箇所：',
            html.escape(parsed["section"]),
            '</span>',
        ])

    for change in parsed["changes"]:
        before = html.escape(
            display_change_history_value(change.get("before", ""))
        )
        after = html.escape(
            display_change_history_value(change.get("after", ""))
        )
        field_name = html.escape(
            clean_value(change.get("field"), blank_text="変更内容")
        )
        parts.extend([
            '<span class="change-history-card-change">',
            field_name,
            '：',
            before,
            ' → ',
            after,
            '</span>',
        ])

    body = "".join(parts)
    target_url = change_history_target_url(parsed)
    if target_url:
        safe_url = html.escape(target_url, quote=True)
        st.markdown(
            f'<a class="change-history-card" href="{safe_url}" target="_self">{body}</a>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="change-history-card change-history-card-static">{body}</div>',
            unsafe_allow_html=True,
        )


def show_change_history_page():
    st.header("🕘 変更確認")
    st.caption("アプリから正常に保存された変更を新しい順に表示します。メモ帳は対象外です。")
    st.markdown(
        """
        <style>
        .change-history-card {
            display: block;
            width: 100%;
            box-sizing: border-box;
            margin: 0 0 0.8rem 0;
            padding: 1rem 1.15rem;
            border: 1px solid rgba(15, 23, 42, 0.16);
            border-radius: 12px;
            background: rgba(255, 255, 255, 0.50);
            color: #172033 !important;
            text-decoration: none !important;
            box-shadow: none;
            overflow-wrap: anywhere;
        }
        a.change-history-card:hover {
            border-color: rgba(37, 99, 235, 0.36);
            background: rgba(239, 246, 255, 0.82);
            text-decoration: none !important;
        }
        .change-history-card-title {
            display: block;
            color: #172033;
            font-weight: 800;
            font-size: 1rem;
            line-height: 1.45;
        }
        .change-history-card-meta {
            display: block;
            margin-top: 0.7rem;
            color: #667085;
            font-size: 0.86rem;
            line-height: 1.4;
        }
        .change-history-card-section {
            display: block;
            margin-top: 0.85rem;
            color: #172033;
            font-weight: 800;
            line-height: 1.5;
        }
        .change-history-card-change {
            display: block;
            margin-top: 0.72rem;
            color: #172033;
            line-height: 1.65;
            white-space: pre-wrap;
        }
        .change-history-card-static {
            cursor: default;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    target_label = st.selectbox(
        "対象",
        ["すべて", *CHANGE_HISTORY_TARGETS],
        key="change_history_target_filter",
    )
    period_label = st.selectbox(
        "期間",
        ["今日", "7日間", "30日間", "すべて"],
        index=2,
        key="change_history_period_filter",
    )
    target_type = "" if target_label == "すべて" else target_label
    now = get_jst_now()
    if period_label == "今日":
        start_iso = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif period_label == "7日間":
        start_iso = (now - timedelta(days=7)).replace(
            hour=0, minute=0, second=0, microsecond=0
        ).isoformat()
    elif period_label == "30日間":
        start_iso = (now - timedelta(days=30)).replace(
            hour=0, minute=0, second=0, microsecond=0
        ).isoformat()
    else:
        start_iso = ""

    signature = f"{target_type}|{start_iso}"
    if st.session_state.get("change_history_filter_signature") != signature:
        st.session_state["change_history_filter_signature"] = signature
        st.session_state["change_history_offset"] = 0
    offset = max(0, int(st.session_state.get("change_history_offset", 0)))

    try:
        rows = load_change_history_page(
            target_type=target_type,
            start_iso=start_iso,
            offset=offset,
            limit=CHANGE_HISTORY_PAGE_SIZE,
        )
    except Exception as exc:
        st.error(str(exc))
        return

    has_next = len(rows) > CHANGE_HISTORY_PAGE_SIZE
    visible_rows = rows[:CHANGE_HISTORY_PAGE_SIZE]
    if not visible_rows:
        st.info("該当する変更履歴はありません。")
    else:
        st.caption(
            f"{offset + 1}件目～{offset + len(visible_rows)}件目を表示"
        )
        for row in visible_rows:
            parsed = parse_change_history_row(row)
            render_change_history_card(parsed)

    previous_col, page_col, next_col = st.columns([1, 1, 1])
    with previous_col:
        if st.button(
            "← 前の30件",
            key="change_history_previous",
            disabled=offset == 0,
            use_container_width=True,
        ):
            st.session_state["change_history_offset"] = max(
                0, offset - CHANGE_HISTORY_PAGE_SIZE
            )
            st.rerun()
    with page_col:
        st.caption(f"ページ {offset // CHANGE_HISTORY_PAGE_SIZE + 1}")
    with next_col:
        if st.button(
            "次の30件 →",
            key="change_history_next",
            disabled=not has_next,
            use_container_width=True,
        ):
            st.session_state["change_history_offset"] = offset + CHANGE_HISTORY_PAGE_SIZE
            st.rerun()


def get_stable_customer_key(detail):
    """同一顧客の全行でIDが1種類だけの場合に限り、安定キーとして使う。"""
    if "ID" not in detail.columns:
        return None
    customer_ids = {
        clean_value(value, blank_text="")
        for value in detail["ID"].tolist()
        if clean_value(value, blank_text="")
    }
    return next(iter(customer_ids)) if len(customer_ids) == 1 else None


def customer_information_query(customer_name, customer_key=None):
    params = {
        "select": "id,customer_key,customer_name,field_name,content,sort_order,created_at,updated_at",
        "order": "sort_order.asc,created_at.asc,id.asc",
    }
    if customer_key:
        params["customer_key"] = f"eq.{customer_key}"
    else:
        params["customer_key"] = "is.null"
        params["customer_name"] = f"eq.{customer_name}"
    return params


def check_customer_information_response(action, response, success_codes):
    if response.status_code in success_codes:
        return
    detail = str(response.text or "").strip()[:500]
    message = f"顧客情報を{action}できませんでした（{response.status_code}）。"
    if detail:
        message += f" {detail}"
    raise RuntimeError(message)


def is_inventory_usage_snapshot_item(item):
    """予想使用量の内部レコードかどうかを返す。通常の顧客情報には表示しない。"""
    field_name = clean_value((item or {}).get("field_name"), blank_text="")
    return field_name.startswith(INVENTORY_USAGE_SNAPSHOT_PREFIX)


def is_regular_customer_information_item(item):
    """顧客情報カードへ表示する通常項目かどうかを、従来と同じ除外規則で判定する。"""
    return not (
        is_past_product_note_item(item)
        or is_estimate_item(item)
        or is_carrier_freight_item(item)
        or is_onedrive_attachment_item(item)
        or is_inventory_usage_snapshot_item(item)
    )

@st.cache_data(ttl=300, show_spinner=False)
def load_customer_information_presence_index():
    """通常の顧客情報が存在する顧客だけを、小さい索引としてSupabaseから読む。"""
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がありません。")

    key_values = set()
    name_values = set()
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name",
            "order": "id.asc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("顧客情報の索引読み込み中にSupabaseへ接続できませんでした。") from exc
        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った顧客情報索引の形式が正しくありません。")

        for item in page:
            if not isinstance(item, dict) or not is_regular_customer_information_item(item):
                continue
            item_key = clean_value(item.get("customer_key"), blank_text="").strip()
            item_name = clean_value(item.get("customer_name"), blank_text="").strip()
            if item_key:
                key_values.add(item_key)
            elif item_name:
                name_values.add(item_name)

        if len(page) < page_size:
            break
        offset += page_size

    return {
        "customer_keys": tuple(sorted(key_values)),
        "customer_names": tuple(sorted(name_values)),
    }

def customer_has_regular_information(customer_name, customer_key=None):
    """存在索引だけで、顧客情報カード用の通常項目があるか確認する。"""
    index = load_customer_information_presence_index()
    if customer_key:
        target = clean_value(customer_key, blank_text="").strip()
        return target in set(index.get("customer_keys", ()))
    target = clean_value(customer_name, blank_text="").strip()
    return target in set(index.get("customer_names", ()))

@st.cache_data(ttl=30, show_spinner=False)
def load_customer_information(customer_name, customer_key=None):
    """件数上限を設けず、Supabaseからページ単位で顧客情報を読む。"""
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がありません。")

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = customer_information_query(customer_name, customer_key)
        params.update({"limit": str(page_size), "offset": str(offset)})
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("顧客情報の読み込み中にSupabaseへ接続できませんでした。") from exc
        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った顧客情報の形式が正しくありません。")
        rows.extend(page)
        if len(page) < page_size:
            break
        offset += page_size
    return rows


def clear_customer_information_cache():
    load_customer_information.clear()
    try:
        load_customer_information_presence_index.clear()
    except Exception:
        pass
    global_attachment_loader = globals().get("load_all_onedrive_attachments_from_supabase")
    if global_attachment_loader is not None:
        try:
            global_attachment_loader.clear()
        except Exception:
            pass


def insert_customer_information(customer_name, customer_key, field_name, content, sort_order):
    now = get_jst_now().isoformat()
    payload = {
        "id": str(uuid.uuid4()),
        "customer_key": customer_key or None,
        "customer_name": clean_value(customer_name, blank_text=""),
        "field_name": str(field_name).strip(),
        "content": str(content or ""),
        "sort_order": int(sort_order),
        "created_at": now,
        "updated_at": now,
    }
    try:
        response = requests.post(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            json=payload,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("顧客情報の保存中にSupabaseへ接続できませんでした。") from exc
    check_customer_information_response("保存", response, (200, 201))
    clear_customer_information_cache()


def update_customer_information(item_id, field_name, content):
    try:
        response = requests.patch(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            params={"id": f"eq.{item_id}"},
            json={
                "field_name": str(field_name).strip(),
                "content": str(content or ""),
            },
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("顧客情報の更新中にSupabaseへ接続できませんでした。") from exc
    check_customer_information_response("更新", response, (200, 204))
    clear_customer_information_cache()


def delete_customer_information(item_id):
    try:
        response = requests.delete(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(prefer="return=minimal"),
            params={"id": f"eq.{item_id}"},
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("顧客情報の削除中にSupabaseへ接続できませんでした。") from exc
    check_customer_information_response("削除", response, (200, 204))
    clear_customer_information_cache()


# =========================
# 予想使用量（Supabase内部保存・Excelの使用数量/日は変更しない）
# =========================
def make_inventory_usage_snapshot_field_name(product_name):
    """商品名から、顧客内で安定する内部レコード名を作る。"""
    normalized_product = normalize_match_value(product_name)
    digest = hashlib.sha256(normalized_product.encode("utf-8")).hexdigest()[:32]
    return f"{INVENTORY_USAGE_SNAPSHOT_PREFIX}{digest}"


def inventory_usage_number(value):
    """在庫履歴用の値を有限の数値へ変換し、変換できない場合はNoneを返す。"""
    if is_blank_excel_value(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def parse_inventory_usage_snapshot_item(item):
    """Supabase内部レコードを、予想使用量表示・次回計算用の辞書へ変換する。"""
    if not is_inventory_usage_snapshot_item(item):
        return None
    try:
        payload = json.loads(str((item or {}).get("content") or "{}"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    product_name = clean_value(payload.get("product_name"), blank_text="").strip()
    if not product_name:
        return None
    return {
        "id": clean_value((item or {}).get("id"), blank_text=""),
        "field_name": clean_value((item or {}).get("field_name"), blank_text=""),
        "product_name": product_name,
        "baseline_date": clean_value(payload.get("baseline_date"), blank_text=""),
        "baseline_bottles": inventory_usage_number(payload.get("baseline_bottles")),
        "baseline_kg_per_bottle": inventory_usage_number(payload.get("baseline_kg_per_bottle")),
        "baseline_inventory_kg": inventory_usage_number(payload.get("baseline_inventory_kg")),
        "comparison_date": clean_value(payload.get("comparison_date"), blank_text=""),
        "comparison_inventory_kg": inventory_usage_number(payload.get("comparison_inventory_kg")),
        "predicted_daily_usage": inventory_usage_number(payload.get("predicted_daily_usage")),
        "created_at": clean_value((item or {}).get("created_at"), blank_text=""),
        "updated_at": clean_value((item or {}).get("updated_at"), blank_text=""),
    }


@st.cache_data(ttl=INVENTORY_USAGE_SNAPSHOT_CACHE_TTL_SECONDS, show_spinner=False)
def load_customer_inventory_usage_snapshots(customer_name, customer_key=None):
    """対象顧客の予想使用量内部レコードだけを、小さい問い合わせで取得する。"""
    if not has_supabase_config():
        return []

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = customer_information_query(customer_name, customer_key)
        params.update(
            {
                "field_name": f"like.{INVENTORY_USAGE_SNAPSHOT_PREFIX}*",
                "order": "updated_at.desc,created_at.desc,id.desc",
                "limit": str(page_size),
                "offset": str(offset),
            }
        )
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("予想使用量を読み込めませんでした。") from exc
        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った予想使用量の形式が正しくありません。")
        rows.extend(page)
        if len(page) < page_size:
            break
        offset += page_size
    return rows


@st.cache_data(ttl=INVENTORY_USAGE_SNAPSHOT_CACHE_TTL_SECONDS, show_spinner=False)
def load_inventory_usage_snapshot_presence_index():
    """予想使用量を持つ顧客だけを、小さい索引として5分間再利用する。"""
    if not has_supabase_config():
        return {"customer_keys": tuple(), "customer_names": tuple()}

    key_values = set()
    name_values = set()
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "customer_key,customer_name,field_name",
            "field_name": f"like.{INVENTORY_USAGE_SNAPSHOT_PREFIX}*",
            "order": "id.asc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("予想使用量の索引を読み込めませんでした。") from exc
        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った予想使用量索引の形式が正しくありません。")
        for item in page:
            item_key = clean_value((item or {}).get("customer_key"), blank_text="").strip()
            item_name = clean_value((item or {}).get("customer_name"), blank_text="").strip()
            if item_key:
                key_values.add(item_key)
            elif item_name:
                name_values.add(item_name)
        if len(page) < page_size:
            break
        offset += page_size

    return {
        "customer_keys": tuple(sorted(key_values)),
        "customer_names": tuple(sorted(name_values)),
    }


def customer_has_inventory_usage_snapshot(customer_name, customer_key=None):
    index = load_inventory_usage_snapshot_presence_index()
    if customer_key:
        target = clean_value(customer_key, blank_text="").strip()
        return target in set(index.get("customer_keys", ()))
    target = clean_value(customer_name, blank_text="").strip()
    return target in set(index.get("customer_names", ()))


def clear_inventory_usage_snapshot_cache():
    try:
        load_customer_inventory_usage_snapshots.clear()
    except Exception:
        pass
    try:
        load_inventory_usage_snapshot_presence_index.clear()
    except Exception:
        pass


def get_customer_inventory_usage_snapshot_map(customer_name, customer_key=None):
    """商品名をキーに、最新の予想使用量内部レコードを返す。"""
    try:
        has_snapshot = customer_has_inventory_usage_snapshot(customer_name, customer_key)
    except Exception:
        # 索引取得に失敗した場合は表示欠落を避け、従来どおり詳細取得へ進む。
        has_snapshot = None
    if has_snapshot is False:
        return {}

    snapshots = {}
    for item in load_customer_inventory_usage_snapshots(customer_name, customer_key):
        parsed = parse_inventory_usage_snapshot_item(item)
        if not parsed:
            continue
        product_key = normalize_match_value(parsed.get("product_name"))
        if product_key and product_key not in snapshots:
            snapshots[product_key] = parsed
    return snapshots


def build_inventory_usage_snapshot(existing, product_name, delivery_date, bottles, kg_per_bottle):
    """
    今日以降に保存した基準値だけで予想使用量を作る。
    Excelの残数・現在の使用数量/日・過去データからの逆算は行わない。
    """
    baseline_date = to_date(delivery_date)
    baseline_bottles = inventory_usage_number(bottles)
    baseline_kg_per_bottle = inventory_usage_number(kg_per_bottle)
    if baseline_date is None or baseline_bottles is None or baseline_kg_per_bottle is None:
        raise ValueError("予想使用量の記録には、配達日・本数・kg/本が必要です。")
    if baseline_bottles < 0 or baseline_kg_per_bottle <= 0:
        raise ValueError("予想使用量の記録には、0以上の本数と0より大きいkg/本が必要です。")

    baseline_inventory_kg = baseline_bottles * baseline_kg_per_bottle
    comparison_date = None
    comparison_inventory_kg = None
    predicted_daily_usage = None

    existing = dict(existing or {})
    previous_baseline_date = to_date(existing.get("baseline_date"))
    previous_baseline_inventory_kg = inventory_usage_number(
        existing.get("baseline_inventory_kg")
    )
    existing_comparison_date = to_date(existing.get("comparison_date"))
    existing_comparison_inventory_kg = inventory_usage_number(
        existing.get("comparison_inventory_kg")
    )

    if previous_baseline_date is not None and previous_baseline_inventory_kg is not None:
        if baseline_date < previous_baseline_date:
            # 最初の基準を訂正する場合だけ、前の日付への変更を許可する。
            if existing_comparison_date is not None:
                raise ValueError(
                    "配達日が前回の基準日より前のため、予想使用量の履歴は更新しませんでした。"
                )
        elif baseline_date == previous_baseline_date:
            # 同じ日の入力訂正では、直前の比較開始点を保って再計算する。
            comparison_date = existing_comparison_date
            comparison_inventory_kg = existing_comparison_inventory_kg
        else:
            comparison_date = previous_baseline_date
            comparison_inventory_kg = previous_baseline_inventory_kg

    if comparison_date is not None and comparison_inventory_kg is not None:
        elapsed_days = (baseline_date - comparison_date).days
        inventory_decrease_kg = comparison_inventory_kg - baseline_inventory_kg
        if elapsed_days > 0 and inventory_decrease_kg >= 0:
            predicted_daily_usage = inventory_decrease_kg / elapsed_days
        # 在庫が増えた場合は納品・補充とみなし、負の使用量を出さず新しい基準だけ保存する。

    return {
        "version": INVENTORY_USAGE_SNAPSHOT_VERSION,
        "product_name": normalize_match_value(product_name),
        "baseline_date": baseline_date.isoformat(),
        "baseline_bottles": baseline_bottles,
        "baseline_kg_per_bottle": baseline_kg_per_bottle,
        "baseline_inventory_kg": baseline_inventory_kg,
        "comparison_date": comparison_date.isoformat() if comparison_date else "",
        "comparison_inventory_kg": comparison_inventory_kg,
        "predicted_daily_usage": predicted_daily_usage,
        "recorded_at": get_jst_now().isoformat(),
    }


def save_customer_inventory_usage_snapshot(
    customer_name,
    customer_key,
    product_name,
    proposed,
    changes,
):
    """在庫基準が変わった時だけ、予想使用量の基準を保存する。"""
    relevant_fields = {"本数", "kg/本", "配達日"}
    if not relevant_fields.intersection(set((changes or {}).keys())):
        return
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がないため、予想使用量の基準を保存できません。")

    snapshots = get_customer_inventory_usage_snapshot_map(customer_name, customer_key)
    product_key = normalize_match_value(product_name)
    existing = snapshots.get(product_key)
    payload = build_inventory_usage_snapshot(
        existing,
        product_name,
        (proposed or {}).get("配達日"),
        (proposed or {}).get("本数"),
        (proposed or {}).get("kg/本"),
    )
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    field_name = make_inventory_usage_snapshot_field_name(product_name)

    if existing and existing.get("id"):
        update_customer_information(existing["id"], field_name, content)
    else:
        insert_customer_information(
            customer_name,
            customer_key,
            field_name,
            content,
            0,
        )
    clear_inventory_usage_snapshot_cache()


def save_customer_inventory_usage_snapshot_safely(*args, **kwargs):
    """Excel保存を取り消さず、予想使用量の記録失敗だけを警告として返す。"""
    try:
        save_customer_inventory_usage_snapshot(*args, **kwargs)
        return ""
    except Exception as exc:
        return f"在庫の保存は完了しましたが、予想使用量を記録できませんでした：{exc}"


# =========================
# ホテル・宿泊先情報（Supabase保存）
# =========================
def make_hotel_information_field_name():
    return HOTEL_INFORMATION_FIELD_PREFIX + uuid.uuid4().hex


def is_hotel_information_item(item):
    return clean_value(item.get("field_name"), blank_text="").startswith(
        HOTEL_INFORMATION_FIELD_PREFIX
    )


def normalize_hotel_custom_fields(values):
    """自由項目を、項目名と内容を持つ安定したリストへ整形する。"""
    result = []
    if not isinstance(values, list):
        return result
    seen_names = set()
    for value in values:
        if not isinstance(value, dict):
            continue
        field_name = clean_value(value.get("name"), blank_text="").strip()
        content = clean_value(value.get("value"), blank_text="").strip()
        if not field_name or not content:
            continue
        normalized_name = field_name.casefold()
        if normalized_name in seen_names:
            continue
        if field_name in HOTEL_INFORMATION_FIXED_FIELDS:
            continue
        seen_names.add(normalized_name)
        result.append({"name": field_name, "value": content})
    return result


def parse_hotel_information_item(item):
    """customer_informationの内部レコードをホテル表示用へ変換する。"""
    if not isinstance(item, dict) or not is_hotel_information_item(item):
        return None
    try:
        payload = json.loads(str(item.get("content") or "{}"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    hotel_name = clean_value(payload.get("hotel_name"), blank_text="").strip()
    region = clean_value(payload.get("region"), blank_text="").strip()
    raw_custom_fields = payload.get("custom_fields")
    # 地域が自由項目だった旧データは、表示を失わず新しい地域欄へ引き継ぐ。
    if not region and isinstance(raw_custom_fields, list):
        for field in raw_custom_fields:
            if not isinstance(field, dict):
                continue
            if clean_value(field.get("name"), blank_text="").strip().casefold() != "地域".casefold():
                continue
            region = clean_value(field.get("value"), blank_text="").strip()
            if region:
                break
    address = clean_value(payload.get("address"), blank_text="").strip()
    google_map = clean_value(payload.get("google_map"), blank_text="").strip()
    if not hotel_name:
        return None
    return {
        "id": clean_value(item.get("id"), blank_text=""),
        "field_name": clean_value(item.get("field_name"), blank_text=""),
        "hotel_name": hotel_name,
        "region": region,
        "address": address,
        "google_map": google_map,
        "custom_fields": normalize_hotel_custom_fields(raw_custom_fields),
        "created_at": clean_value(item.get("created_at"), blank_text=""),
        "updated_at": clean_value(item.get("updated_at"), blank_text=""),
    }


def load_hotel_information_records():
    """ホテル情報だけを新しい順に読み込む。"""
    items = load_customer_information(HOTEL_INFORMATION_STORAGE_CUSTOMER)
    records = []
    for item in items:
        parsed = parse_hotel_information_item(item)
        if parsed is not None:
            records.append(parsed)
    records.sort(
        key=lambda record: str(record.get("updated_at") or record.get("created_at") or ""),
        reverse=True,
    )
    return records


def encode_hotel_information_content(hotel_name, region, address, google_map, custom_fields):
    return json.dumps(
        {
            "version": HOTEL_INFORMATION_VERSION,
            "hotel_name": str(hotel_name).strip(),
            "region": str(region or "").strip(),
            "address": str(address).strip(),
            "google_map": str(google_map).strip(),
            "custom_fields": normalize_hotel_custom_fields(custom_fields),
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )


def save_hotel_information_record(record, hotel_name, region, address, google_map, custom_fields):
    content = encode_hotel_information_content(
        hotel_name,
        region,
        address,
        google_map,
        custom_fields,
    )
    record = record if isinstance(record, dict) else {}
    item_id = clean_value(record.get("id"), blank_text="")
    field_name = clean_value(record.get("field_name"), blank_text="")
    if item_id and field_name:
        update_customer_information(item_id, field_name, content)
        return "ホテル情報を更新しました。"
    insert_customer_information(
        HOTEL_INFORMATION_STORAGE_CUSTOMER,
        None,
        make_hotel_information_field_name(),
        content,
        0,
    )
    return "ホテル情報を追加しました。"


def clear_hotel_information_editor_state(state_prefix):
    keys = [
        key for key in list(st.session_state.keys())
        if str(key).startswith(state_prefix)
    ]
    for key in keys:
        st.session_state.pop(key, None)


@st.dialog("ホテル・宿泊先情報")
def show_hotel_information_editor_dialog(record=None):
    record = record if isinstance(record, dict) else {}
    item_id = clean_value(record.get("id"), blank_text="")
    state_suffix = item_id or "new"
    state_prefix = f"hotel_information_editor_{state_suffix}_"
    initialized_key = state_prefix + "initialized"
    hotel_name_key = state_prefix + "hotel_name"
    region_key = state_prefix + "region"
    address_key = state_prefix + "address"
    google_map_key = state_prefix + "google_map"
    count_key = state_prefix + "custom_count"

    if not st.session_state.get(initialized_key):
        custom_fields = normalize_hotel_custom_fields(record.get("custom_fields"))
        st.session_state[hotel_name_key] = clean_value(
            record.get("hotel_name"), blank_text=""
        )
        st.session_state[region_key] = clean_value(record.get("region"), blank_text="")
        st.session_state[address_key] = clean_value(record.get("address"), blank_text="")
        st.session_state[google_map_key] = clean_value(
            record.get("google_map"), blank_text=""
        )
        st.session_state[count_key] = max(1, len(custom_fields))
        for index, field in enumerate(custom_fields):
            st.session_state[state_prefix + f"custom_name_{index}"] = field["name"]
            st.session_state[state_prefix + f"custom_value_{index}"] = field["value"]
        st.session_state[initialized_key] = True

    st.caption("必須項目はホテル名だけです。住所・Googleマップやその他の項目は、あとから追加できます。")
    st.text_input("ホテル名（必須）", key=hotel_name_key, autocomplete="off")
    st.text_input(
        "地域（任意）",
        key=region_key,
        placeholder="例：札幌、帯広、名古屋",
        autocomplete="off",
    )
    st.text_area("住所（任意）", key=address_key, height=90)
    st.text_input(
        "Googleマップ（任意）",
        key=google_map_key,
        placeholder="共有URL・緯度経度・施設名のいずれでも入力できます",
        autocomplete="off",
    )

    st.markdown("#### 自由項目")
    if st.button(
        "＋ 自由項目を追加",
        key=state_prefix + "add_custom_field",
        use_container_width=True,
    ):
        current_count = max(1, int(st.session_state.get(count_key, 1)))
        st.session_state[state_prefix + f"custom_name_{current_count}"] = ""
        st.session_state[state_prefix + f"custom_value_{current_count}"] = ""
        st.session_state[count_key] = current_count + 1

    custom_count = max(1, int(st.session_state.get(count_key, 1)))
    for index in range(custom_count):
        name_key = state_prefix + f"custom_name_{index}"
        value_key = state_prefix + f"custom_value_{index}"
        if name_key not in st.session_state:
            st.session_state[name_key] = ""
        if value_key not in st.session_state:
            st.session_state[value_key] = ""
        with st.container(border=True):
            st.text_input(
                f"自由項目 {index + 1}：項目名",
                key=name_key,
                placeholder="例：バイク駐車場",
                autocomplete="off",
            )
            st.text_area(
                f"自由項目 {index + 1}：内容",
                key=value_key,
                placeholder="例：屋根付き・大型可・1泊500円・事前連絡必要",
                height=90,
            )
    st.caption("自由項目を削除するときは、その項目名と内容を両方空欄にして保存してください。")

    save_col, cancel_col = st.columns(2)
    with save_col:
        save_clicked = st.button(
            "保存",
            key=state_prefix + "save",
            type="primary",
            use_container_width=True,
        )
    with cancel_col:
        cancel_clicked = st.button(
            "キャンセル",
            key=state_prefix + "cancel",
            use_container_width=True,
        )

    if cancel_clicked:
        clear_hotel_information_editor_state(state_prefix)
        st.rerun()

    if not save_clicked:
        return

    hotel_name = str(st.session_state.get(hotel_name_key, "")).strip()
    region = str(st.session_state.get(region_key, "")).strip()
    address = str(st.session_state.get(address_key, "")).strip()
    google_map = str(st.session_state.get(google_map_key, "")).strip()
    missing = [
        label for label, value in (
            ("ホテル名", hotel_name),
        ) if not value
    ]
    if missing:
        st.error("必須項目を入力してください：" + "、".join(missing))
        return

    custom_fields = []
    custom_names = set()
    for index in range(custom_count):
        field_name = str(
            st.session_state.get(state_prefix + f"custom_name_{index}", "")
        ).strip()
        value = str(
            st.session_state.get(state_prefix + f"custom_value_{index}", "")
        ).strip()
        if not field_name and not value:
            continue
        if not field_name or not value:
            st.error(f"自由項目 {index + 1}は、項目名と内容の両方を入力してください。")
            return
        if field_name in HOTEL_INFORMATION_FIXED_FIELDS:
            st.error(f"「{field_name}」は上の固定項目にあるため、自由項目には登録できません。")
            return
        normalized_name = field_name.casefold()
        if normalized_name in custom_names:
            st.error(f"自由項目「{field_name}」が重複しています。")
            return
        custom_names.add(normalized_name)
        custom_fields.append({"name": field_name, "value": value})

    try:
        message = save_hotel_information_record(
            record,
            hotel_name,
            region,
            address,
            google_map,
            custom_fields,
        )
        clear_hotel_information_editor_state(state_prefix)
        st.session_state["hotel_information_message"] = message
        st.rerun()
    except Exception as exc:
        st.error(str(exc))


@st.dialog("ホテル情報を削除")
def confirm_hotel_information_delete_dialog(record):
    hotel_name = clean_value(record.get("hotel_name"), blank_text="ホテル")
    item_id = clean_value(record.get("id"), blank_text="")
    st.warning(f"「{hotel_name}」を削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)
    with delete_col:
        if st.button(
            "削除する",
            key=f"hotel_information_delete_yes_{item_id}",
            type="primary",
            use_container_width=True,
        ):
            try:
                delete_customer_information(item_id)
                st.session_state["hotel_information_message"] = "ホテル情報を削除しました。"
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"hotel_information_delete_no_{item_id}",
            use_container_width=True,
        ):
            st.rerun()


def render_hotel_information_field(field_name, content):
    safe_name = html.escape(clean_value(field_name, blank_text=""))
    safe_content = html.escape(clean_value(content, blank_text="")).replace("\n", "<br>")
    st.markdown(
        (
            '<div class="customer-information-row">'
            f'<div class="customer-information-label">{safe_name}</div>'
            f'<div class="customer-information-content">{safe_content}</div>'
            '</div>'
        ),
        unsafe_allow_html=True,
    )


def show_hotel_information_page():
    st.header("🏨 ホテル・宿泊先情報")
    st.caption("ホテル名だけ必須です。住所・Googleマップやその他の項目は、あとから追加できます。")

    if not has_supabase_config():
        st.warning("ホテル・宿泊先情報を使うには、現在のSupabase設定が必要です。")
        return

    if st.button(
        "＋ ホテルを追加",
        key="hotel_information_add_button",
        type="primary",
        use_container_width=True,
    ):
        clear_hotel_information_editor_state("hotel_information_editor_new_")
        show_hotel_information_editor_dialog()

    message = st.session_state.pop("hotel_information_message", None)
    if message:
        st.success(message)

    try:
        records = load_hotel_information_records()
    except Exception as exc:
        st.error(str(exc))
        return

    if not records:
        st.info("ホテル・宿泊先情報はまだありません。")
        return

    query_param = clean_value(
        get_query_value("hotel_search", ""),
        blank_text="",
    ).strip()
    query_key = "hotel_information_search"
    query_source_key = "hotel_information_search_source"
    if query_param and st.session_state.get(query_source_key) != query_param:
        st.session_state[query_key] = query_param
        st.session_state[query_source_key] = query_param
    query = st.text_input(
        "ホテル名・地域・住所・自由項目を検索",
        key=query_key,
        placeholder="ホテル名・地域・住所・項目名・内容の一部を入力",
        autocomplete="off",
    )
    normalized_query = clean_value(query, blank_text="").strip().casefold()
    query_terms = [
        term for term in re.split(r"[\s　]+", normalized_query) if term
    ]
    filtered_records = []
    for record in records:
        custom_field_search_values = []
        for field in record.get("custom_fields", []):
            if not isinstance(field, dict):
                continue
            custom_field_search_values.extend(
                [
                    clean_value(field.get("name"), blank_text=""),
                    clean_value(field.get("value"), blank_text=""),
                ]
            )
        searchable = " ".join(
            [
                clean_value(record.get("hotel_name"), blank_text=""),
                clean_value(record.get("region"), blank_text=""),
                clean_value(record.get("address"), blank_text=""),
                *custom_field_search_values,
            ]
        ).casefold()
        if query_terms and not all(term in searchable for term in query_terms):
            continue
        filtered_records.append(record)

    st.caption(f"該当：{len(filtered_records)}件")
    if not filtered_records:
        st.info("検索条件に一致する宿泊先はありません。")
        return

    for record in filtered_records:
        record_id = clean_value(record.get("id"), blank_text="")
        card_suffix = hashlib.sha256(record_id.encode("utf-8")).hexdigest()[:16]
        with st.container(border=True, key=f"hotel_information_card_{card_suffix}"):
            st.subheader(f"🏨 {record['hotel_name']}")
            if clean_value(record.get("region"), blank_text="").strip():
                render_hotel_information_field("地域", record["region"])
            render_hotel_information_field("住所", record["address"])
            show_google_maps_button(
                build_google_maps_url(record.get("google_map") or record.get("address"))
            )
            for field in record.get("custom_fields", []):
                render_hotel_information_field(field["name"], field["value"])

            # 顧客・仕入先・運送会社と同じOneDrive/Supabaseのルールで管理する。
            render_customer_attachments_section(
                record["hotel_name"],
                record_id,
                entity_type="hotel",
            )

            edit_col, delete_col = st.columns(2)
            with edit_col:
                if st.button(
                    "編集",
                    key=f"hotel_information_edit_{record_id}",
                    use_container_width=True,
                ):
                    clear_hotel_information_editor_state(
                        f"hotel_information_editor_{record_id}_"
                    )
                    show_hotel_information_editor_dialog(record)
            with delete_col:
                if st.button(
                    "削除",
                    key=f"hotel_information_delete_{record_id}",
                    use_container_width=True,
                ):
                    confirm_hotel_information_delete_dialog(record)


def make_onedrive_attachment_field_name():
    return ONEDRIVE_ATTACHMENT_PREFIX + uuid.uuid4().hex


def is_onedrive_attachment_item(item):
    return clean_value(item.get("field_name"), blank_text="").startswith(
        ONEDRIVE_ATTACHMENT_PREFIX
    )


def normalize_attachment_tags(values):
    if isinstance(values, str):
        candidates = re.split(r"[,、\n]+", values)
    elif isinstance(values, (list, tuple, set)):
        candidates = list(values)
    else:
        candidates = []
    result = []
    seen = set()
    for value in candidates:
        tag = clean_value(value, blank_text="").strip().lstrip("#").strip()
        if not tag or tag in seen:
            continue
        seen.add(tag)
        result.append(tag)
    return result


def normalize_attachment_entity_type(value):
    text = clean_value(value, blank_text="").strip().lower()
    return text if text in {"customer", "supplier", "carrier", "hotel"} else "customer"


def attachment_entity_label(entity_type):
    return {
        "customer": "顧客",
        "supplier": "仕入先",
        "carrier": "運送会社",
        "hotel": "ホテル",
    }[normalize_attachment_entity_type(entity_type)]


def attachment_entity_folder(entity_type):
    return {
        "customer": ONEDRIVE_CUSTOMER_FOLDER,
        "supplier": ONEDRIVE_SUPPLIER_FOLDER,
        "carrier": ONEDRIVE_CARRIER_FOLDER,
        "hotel": ONEDRIVE_HOTEL_FOLDER,
    }[normalize_attachment_entity_type(entity_type)]


def attachment_storage_key(entity_type, entity_id, entity_name):
    entity_type = normalize_attachment_entity_type(entity_type)
    entity_id = clean_value(entity_id, blank_text="").strip()
    if entity_type == "customer":
        return entity_id or None
    if not entity_id:
        entity_id = hashlib.sha256(str(entity_name).encode("utf-8")).hexdigest()[:16]
    return f"onedrive_attachment:{entity_type}:{entity_id}"


def parse_onedrive_attachment_item(item):
    if not is_onedrive_attachment_item(item):
        return None
    try:
        payload = json.loads(str(item.get("content") or "{}"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None

    entity_type = normalize_attachment_entity_type(payload.get("entity_type"))
    entity_id = clean_value(payload.get("entity_id"), blank_text="").strip()
    entity_name = clean_value(payload.get("entity_name"), blank_text="").strip()
    stored_customer_key = clean_value(item.get("customer_key"), blank_text="")
    stored_customer_name = clean_value(item.get("customer_name"), blank_text="")

    # v41以前の顧客データには取引先種別がないため、従来どおり顧客として扱う。
    if not entity_id:
        entity_id = stored_customer_key if entity_type == "customer" else ""
    if not entity_name:
        entity_name = stored_customer_name

    return {
        "id": str(item.get("id") or ""),
        "field_name": str(item.get("field_name") or ""),
        "customer_key": stored_customer_key,
        "customer_name": stored_customer_name,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity_name": entity_name,
        "file_id": clean_value(payload.get("file_id"), blank_text=""),
        "original_name": clean_value(payload.get("original_name"), blank_text=""),
        "stored_name": clean_value(payload.get("stored_name"), blank_text=""),
        "file_type": clean_value(payload.get("file_type"), blank_text=""),
        "mime_type": clean_value(payload.get("mime_type"), blank_text=""),
        "size": payload.get("size") or 0,
        "onedrive_path": clean_value(payload.get("onedrive_path"), blank_text=""),
        "web_url": clean_value(payload.get("web_url"), blank_text=""),
        "display_file_id": clean_value(payload.get("display_file_id"), blank_text=""),
        "display_stored_name": clean_value(payload.get("display_stored_name"), blank_text=""),
        "display_mime_type": clean_value(payload.get("display_mime_type"), blank_text=""),
        "display_size": payload.get("display_size") or 0,
        "display_web_url": clean_value(payload.get("display_web_url"), blank_text=""),
        "tags": normalize_attachment_tags(payload.get("tags") or []),
        "remarks": clean_value(payload.get("remarks"), blank_text=""),
        "uploaded_by": clean_value(payload.get("uploaded_by"), blank_text=""),
        "created_at": clean_value(payload.get("created_at"), blank_text="")
        or clean_value(item.get("created_at"), blank_text=""),
        "updated_at": clean_value(item.get("updated_at"), blank_text=""),
        "group_id": clean_value(payload.get("group_id"), blank_text=""),
        "group_index": int(payload.get("group_index") or 0),
        "group_size": max(1, int(payload.get("group_size") or 1)),
        "version": payload.get("version") or 1,
    }

def serialize_onedrive_attachment(attachment):
    payload = {
        "version": ONEDRIVE_ATTACHMENT_VERSION,
        "entity_type": normalize_attachment_entity_type(
            attachment.get("entity_type")
        ),
        "entity_id": attachment.get("entity_id", ""),
        "entity_name": attachment.get("entity_name", ""),
        "file_id": attachment.get("file_id", ""),
        "original_name": attachment.get("original_name", ""),
        "stored_name": attachment.get("stored_name", ""),
        "file_type": attachment.get("file_type", ""),
        "mime_type": attachment.get("mime_type", ""),
        "size": int(attachment.get("size") or 0),
        "onedrive_path": attachment.get("onedrive_path", ""),
        "web_url": attachment.get("web_url", ""),
        "display_file_id": attachment.get("display_file_id", ""),
        "display_stored_name": attachment.get("display_stored_name", ""),
        "display_mime_type": attachment.get("display_mime_type", ""),
        "display_size": int(attachment.get("display_size") or 0),
        "display_web_url": attachment.get("display_web_url", ""),
        "tags": normalize_attachment_tags(attachment.get("tags") or []),
        "remarks": attachment.get("remarks", ""),
        "uploaded_by": attachment.get("uploaded_by", ""),
        "created_at": attachment.get("created_at", ""),
        "group_id": attachment.get("group_id", ""),
        "group_index": int(attachment.get("group_index") or 0),
        "group_size": max(1, int(attachment.get("group_size") or 1)),
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

def attachment_group_index(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def group_onedrive_attachments_for_display(attachments):
    """同時選択で保存した画像を1カードへまとめ、従来データは単独のまま返す。"""
    groups = []
    by_key = {}
    for position, attachment in enumerate(list(attachments or [])):
        group_id = clean_value(attachment.get("group_id"), blank_text="").strip()
        if attachment.get("file_type") == "image" and group_id:
            key = "group:" + ":".join(
                [
                    normalize_attachment_entity_type(attachment.get("entity_type")),
                    clean_value(attachment.get("entity_id"), blank_text=""),
                    group_id,
                ]
            )
        else:
            stable_id = (
                clean_value(attachment.get("id"), blank_text="")
                or clean_value(attachment.get("file_id"), blank_text="")
                or f"position_{position}"
            )
            key = f"single:{stable_id}"

        if key not in by_key:
            group = {"key": key, "items": []}
            by_key[key] = group
            groups.append(group)
        by_key[key]["items"].append(attachment)

    for group in groups:
        items = group["items"]
        items.sort(
            key=lambda item: (
                attachment_group_index(item.get("group_index")),
                str(item.get("created_at") or item.get("updated_at") or ""),
                str(item.get("id") or ""),
            )
        )
        representative = items[0]
        group["representative"] = representative
        group["count"] = len(items)
        group["ui_id"] = hashlib.sha256(group["key"].encode("utf-8")).hexdigest()[:20]
    return groups


def update_onedrive_attachment_group_metadata(attachments, tags, remarks):
    """グループ内の全画像へ同じタグ・備考を保存する。"""
    updated = []
    for attachment in list(attachments or []):
        updated.append(
            update_customer_onedrive_attachment_metadata(
                attachment,
                tags,
                remarks,
            )
        )
    return updated


def get_customer_onedrive_folder_key(customer_key, customer_name):
    if customer_key:
        raw = f"顧客ID_{customer_key}"
    else:
        digest = hashlib.sha256(str(customer_name).encode("utf-8")).hexdigest()[:16]
        raw = f"顧客仮ID_{digest}"
    safe = re.sub(r"[\\/:*?\"<>|]", "_", raw).strip().rstrip(".")
    return safe or "顧客仮ID_未設定"


def get_attachment_onedrive_folder_key(entity_type, entity_id, entity_name):
    entity_type = normalize_attachment_entity_type(entity_type)
    if entity_type == "customer":
        return get_customer_onedrive_folder_key(entity_id, entity_name)

    label = attachment_entity_label(entity_type)
    entity_id = clean_value(entity_id, blank_text="").strip()
    if entity_id:
        raw = f"{label}ID_{entity_id}"
    else:
        digest = hashlib.sha256(str(entity_name).encode("utf-8")).hexdigest()[:16]
        raw = f"{label}仮ID_{digest}"
    safe = re.sub(r"[\\/:*?\"<>|]", "_", raw).strip().rstrip(".")
    return safe or f"{label}仮ID_未設定"


def get_entity_attachments(entity_type, entity_name, entity_id=None):
    entity_type = normalize_attachment_entity_type(entity_type)
    storage_key = attachment_storage_key(entity_type, entity_id, entity_name)
    items = load_customer_information(entity_name, storage_key)
    attachments = []
    for item in items:
        parsed = parse_onedrive_attachment_item(item)
        if not parsed:
            continue
        if normalize_attachment_entity_type(parsed.get("entity_type")) != entity_type:
            continue
        attachments.append(parsed)
    attachments.sort(
        key=lambda row: str(row.get("created_at") or row.get("updated_at") or ""),
        reverse=True,
    )
    return attachments


def get_customer_attachments(customer_name, customer_key):
    return get_entity_attachments("customer", customer_name, customer_key)


@st.cache_data(ttl=30, show_spinner=False)
def load_all_onedrive_attachments_from_supabase():
    """全取引先の写真・資料メタデータだけをSupabaseから読み込む。"""
    if not has_supabase_config():
        return []

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name,content,sort_order,created_at,updated_at",
            "field_name": f"like.{ONEDRIVE_ATTACHMENT_PREFIX}*",
            "order": "created_at.desc,id.desc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("写真・資料の読み込み中にSupabaseへ接続できませんでした。") from exc

        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った写真・資料の形式が正しくありません。")
        rows.extend(page)
        if len(page) < page_size:
            break
        offset += page_size

    attachments = []
    for row in rows:
        parsed = parse_onedrive_attachment_item(row)
        if parsed:
            attachments.append(parsed)
    attachments.sort(
        key=lambda item: str(item.get("created_at") or item.get("updated_at") or ""),
        reverse=True,
    )
    return attachments


def attachment_tag_history_options(attachments):
    """使われたタグを、最近使われた順で重複なく返す。"""
    sorted_attachments = sorted(
        list(attachments or []),
        key=lambda item: str(item.get("updated_at") or item.get("created_at") or ""),
        reverse=True,
    )
    options = []
    seen = set()
    for attachment in sorted_attachments:
        for tag in normalize_attachment_tags(attachment.get("tags") or []):
            if tag in seen:
                continue
            seen.add(tag)
            options.append(tag)
    return options


def get_attachment_tag_history_options(fallback_attachments=None):
    """全体履歴を優先し、読み込めない場合だけ現在画面のタグを使う。"""
    try:
        return attachment_tag_history_options(load_all_onedrive_attachments_from_supabase())
    except Exception:
        return attachment_tag_history_options(fallback_attachments or [])


def attachment_file_kind(filename, mime_type=""):
    suffix = Path(str(filename or "")).suffix.lower()
    mime = str(mime_type or "").lower()
    if mime.startswith("image/") or suffix in ONEDRIVE_IMAGE_EXTENSIONS:
        return "image"
    if mime == "application/pdf" or suffix in ONEDRIVE_PDF_EXTENSIONS:
        return "pdf"
    return ""


def get_attachment_uploaded_by(access_token):
    """同じ保存操作中に再利用できるOneDriveの登録者名を返す。"""
    profile = {}
    try:
        profile = get_onedrive_profile(access_token)
    except Exception:
        profile = {}
    return (
        clean_value(profile.get("displayName"), blank_text="")
        or clean_value(profile.get("mail"), blank_text="")
        or clean_value(profile.get("userPrincipalName"), blank_text="")
    )


def build_entity_onedrive_attachment_record(
    entity_type,
    entity_name,
    entity_id,
    original_name,
    content,
    mime_type,
    tags,
    remarks,
    uploaded_item,
    requested_stored_name,
    folder_path,
    uploaded_by,
    group_id="",
    group_index=0,
    group_size=1,
    display_item=None,
    requested_display_name="",
    display_content=b"",
):
    """従来の元画像情報に、任意の通常表示用WebP情報だけを加える。"""
    file_kind = attachment_file_kind(original_name, mime_type)
    file_id = clean_value((uploaded_item or {}).get("id"), blank_text="")
    if not file_id:
        raise RuntimeError("OneDriveから保存済みファイルIDを取得できませんでした。")
    stored_name = (
        clean_value((uploaded_item or {}).get("name"), blank_text="")
        or clean_value(requested_stored_name, blank_text="")
    )
    display_file_id = clean_value((display_item or {}).get("id"), blank_text="")
    display_stored_name = (
        clean_value((display_item or {}).get("name"), blank_text="")
        or clean_value(requested_display_name, blank_text="")
    )
    return {
        "entity_type": normalize_attachment_entity_type(entity_type),
        "entity_id": clean_value(entity_id, blank_text="").strip(),
        "entity_name": clean_value(entity_name, blank_text="").strip(),
        "file_id": file_id,
        "original_name": Path(str(original_name or "file")).name,
        "stored_name": stored_name,
        "file_type": file_kind,
        "mime_type": mime_type
        or mimetypes.guess_type(str(original_name or ""))[0]
        or "application/octet-stream",
        "size": int((uploaded_item or {}).get("size") or len(content)),
        "onedrive_path": folder_path,
        "web_url": clean_value((uploaded_item or {}).get("webUrl"), blank_text=""),
        "display_file_id": display_file_id,
        "display_stored_name": display_stored_name if display_file_id else "",
        "display_mime_type": "image/webp" if display_file_id else "",
        "display_size": (
            int((display_item or {}).get("size") or len(display_content or b""))
            if display_file_id
            else 0
        ),
        "display_web_url": clean_value((display_item or {}).get("webUrl"), blank_text=""),
        "tags": normalize_attachment_tags(tags),
        "remarks": str(remarks or "").strip(),
        "uploaded_by": str(uploaded_by or ""),
        "created_at": get_jst_now().isoformat(),
        "group_id": str(group_id or "").strip(),
        "group_index": int(group_index or 0),
        "group_size": max(1, int(group_size or 1)),
    }

def register_entity_onedrive_attachment(
    attachment,
    access_token,
    moved_source=None,
):
    """Supabase登録に失敗した時は、元画像と表示用画像を従来どおり戻す。"""
    entity_type = normalize_attachment_entity_type(attachment.get("entity_type"))
    entity_id = clean_value(attachment.get("entity_id"), blank_text="").strip()
    entity_name = clean_value(attachment.get("entity_name"), blank_text="").strip()
    file_id = clean_value(attachment.get("file_id"), blank_text="")
    display_file_id = clean_value(attachment.get("display_file_id"), blank_text="")
    stored_name = clean_value(attachment.get("stored_name"), blank_text="")
    storage_key = attachment_storage_key(entity_type, entity_id, entity_name)
    try:
        insert_customer_information(
            entity_name,
            storage_key,
            make_onedrive_attachment_field_name(),
            serialize_onedrive_attachment(attachment),
            int(time.time()),
        )
    except Exception:
        if display_file_id:
            try:
                delete_onedrive_file(access_token, display_file_id)
            except Exception:
                pass
        if moved_source:
            try:
                move_onedrive_item(
                    access_token,
                    moved_source["item_id"],
                    moved_source["parent_id"],
                    moved_source["name"],
                )
            except Exception as rollback_exc:
                raise RuntimeError(
                    "写真はOneDrive内で移動されましたが、アプリへの登録に失敗し、"
                    "元のカメラロールへ戻せませんでした。"
                    f"OneDriveで「{stored_name}」を確認してください。"
                ) from rollback_exc
        else:
            try:
                delete_onedrive_file(access_token, file_id)
            except Exception:
                pass
        raise
    return attachment

def save_multiple_entity_onedrive_photos_parallel(
    entity_type,
    entity_name,
    entity_id,
    photo_uploads,
    tags,
    remarks,
    access_token,
    group_id,
    max_workers=3,
    move_from_camera_roll=False,
):
    """元画像の保存ルールを保ち、通常表示用WebPを加えて最大3件ずつ保存する。"""
    entity_type = normalize_attachment_entity_type(entity_type)
    entity_name = clean_value(entity_name, blank_text="").strip()
    entity_id = clean_value(entity_id, blank_text="").strip()
    uploads = list(photo_uploads or [])
    if len(uploads) < 2:
        raise ValueError("複数写真の保存処理には2枚以上の写真が必要です。")

    folder_key = get_attachment_onedrive_folder_key(
        entity_type,
        entity_id,
        entity_name,
    )
    folder_path = "/".join(
        [
            ONEDRIVE_ROOT_FOLDER,
            attachment_entity_folder(entity_type),
            folder_key,
            "写真",
        ]
    )
    target_folder = ensure_onedrive_folder_path(access_token, folder_path)
    target_folder_id = clean_value(target_folder.get("id"), blank_text="")
    if not target_folder_id:
        raise RuntimeError("OneDriveの移動先フォルダIDを取得できませんでした。")
    uploaded_by = get_attachment_uploaded_by(access_token)

    prepared_jobs = []
    failed_by_index = {}
    reserved_source_ids = set()
    group_size = len(uploads)

    # カメラバックアップ照合は従来どおり直列。表示用WebPは通信前に作る。
    for upload_index, upload in enumerate(uploads):
        original_name = Path(str(upload.get("name") or "file")).name
        content = bytes(upload.get("content") or b"")
        mime_type = str(upload.get("mime_type") or "application/octet-stream")
        try:
            if attachment_file_kind(original_name, mime_type) != "image":
                raise ValueError("画像（JPG・JPEG・PNG・WEBP）を選んでください。")
            if not content:
                raise ValueError("選択したファイルが空です。")
            source_item = None
            if move_from_camera_roll:
                source_item = find_matching_onedrive_camera_roll_file(
                    access_token,
                    original_name,
                    content,
                    excluded_item_ids=reserved_source_ids,
                )
                source_item_id = clean_value((source_item or {}).get("id"), blank_text="")
                if source_item_id:
                    reserved_source_ids.add(source_item_id)
            timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S")
            stored_name = f"{timestamp}_{uuid.uuid4().hex[:8]}_{original_name}"
            display = prepare_onedrive_display_image(
                content,
                original_name,
                stored_name,
            )
            prepared_jobs.append(
                {
                    "index": upload_index,
                    "original_name": original_name,
                    "content": content,
                    "mime_type": mime_type,
                    "stored_name": stored_name,
                    "source_item": source_item,
                    "display": display,
                }
            )
        except Exception as exc:
            failed_by_index[upload_index] = (original_name, str(exc))

    def save_to_onedrive(job):
        source_item = job.get("source_item") or {}
        display = job.get("display") or {}
        moved_source = None
        uploaded_item = None
        display_item = None
        try:
            if source_item:
                source_parent_reference = source_item.get("parentReference") or {}
                source_item_id = clean_value(source_item.get("id"), blank_text="")
                source_parent_id = clean_value(source_parent_reference.get("id"), blank_text="")
                source_name = (
                    clean_value(source_item.get("name"), blank_text="")
                    or job["original_name"]
                )
                if not source_item_id or not source_parent_id:
                    raise RuntimeError(
                        "OneDriveの元写真の保存場所を確認できませんでした。"
                        "誤移動を防ぐため保存を停止しました。"
                    )
                uploaded_item = move_onedrive_item(
                    access_token,
                    source_item_id,
                    target_folder_id,
                    job["stored_name"],
                )
                moved_source = {
                    "item_id": source_item_id,
                    "parent_id": source_parent_id,
                    "name": source_name,
                }
            else:
                uploaded_item = upload_onedrive_file_to_existing_folder(
                    access_token,
                    folder_path,
                    job["stored_name"],
                    job["content"],
                    job["mime_type"],
                )

            if display:
                display_item = upload_onedrive_file_to_existing_folder(
                    access_token,
                    folder_path,
                    display["stored_name"],
                    display["content"],
                    display["mime_type"],
                )
            return job, uploaded_item, display_item, moved_source
        except Exception:
            if display_item and display_item.get("id"):
                try:
                    delete_onedrive_file(access_token, display_item["id"])
                except Exception:
                    pass
            if moved_source:
                try:
                    move_onedrive_item(
                        access_token,
                        moved_source["item_id"],
                        moved_source["parent_id"],
                        moved_source["name"],
                    )
                except Exception:
                    pass
            elif uploaded_item and uploaded_item.get("id"):
                try:
                    delete_onedrive_file(access_token, uploaded_item["id"])
                except Exception:
                    pass
            raise

    onedrive_results = {}
    worker_count = max(1, min(int(max_workers or 1), 3, len(prepared_jobs)))
    if prepared_jobs:
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            future_map = {
                executor.submit(save_to_onedrive, job): job
                for job in prepared_jobs
            }
            for future in as_completed(future_map):
                job = future_map[future]
                try:
                    onedrive_results[job["index"]] = future.result()
                except Exception as exc:
                    failed_by_index[job["index"]] = (
                        job["original_name"],
                        str(exc),
                    )

    saved_items = []
    # Supabase登録と失敗時ロールバックは、従来どおり結果順に直列で行う。
    for upload_index in range(group_size):
        result = onedrive_results.get(upload_index)
        if result is None:
            continue
        job, uploaded_item, display_item, moved_source = result
        display = job.get("display") or {}
        try:
            attachment = build_entity_onedrive_attachment_record(
                entity_type,
                entity_name,
                entity_id,
                job["original_name"],
                job["content"],
                job["mime_type"],
                tags,
                remarks,
                uploaded_item,
                job["stored_name"],
                folder_path,
                uploaded_by,
                group_id=group_id,
                group_index=upload_index,
                group_size=group_size,
                display_item=display_item,
                requested_display_name=display.get("stored_name", ""),
                display_content=display.get("content", b""),
            )
            saved_items.append(
                register_entity_onedrive_attachment(
                    attachment,
                    access_token,
                    moved_source=moved_source,
                )
            )
        except Exception as exc:
            failed_by_index[upload_index] = (
                job["original_name"],
                str(exc),
            )

    failed_items = [
        failed_by_index[index]
        for index in sorted(failed_by_index)
    ]
    return saved_items, failed_items

def save_selected_onedrive_camera_photos_parallel(
    entity_type,
    entity_name,
    entity_id,
    selected_items,
    tags,
    remarks,
    access_token,
    group_id="",
    max_workers=3,
):
    """選択済みの正確なOneDriveファイルIDを使い、照合せず最大3件ずつ保存する。"""
    entity_type = normalize_attachment_entity_type(entity_type)
    entity_name = clean_value(entity_name, blank_text="").strip()
    entity_id = clean_value(entity_id, blank_text="").strip()
    items = [item for item in list(selected_items or []) if isinstance(item, dict)]
    if not items:
        raise ValueError("OneDriveの写真を選んでください。")

    folder_key = get_attachment_onedrive_folder_key(
        entity_type,
        entity_id,
        entity_name,
    )
    folder_path = "/".join(
        [
            ONEDRIVE_ROOT_FOLDER,
            attachment_entity_folder(entity_type),
            folder_key,
            "写真",
        ]
    )
    target_folder = ensure_onedrive_folder_path(access_token, folder_path)
    target_folder_id = clean_value(target_folder.get("id"), blank_text="")
    if not target_folder_id:
        raise RuntimeError("OneDriveの移動先フォルダIDを取得できませんでした。")
    uploaded_by = get_attachment_uploaded_by(access_token)
    group_size = len(items)
    allowed_mime_types = {"image/jpeg", "image/png", "image/webp"}

    def rollback_saved_files(display_item, moved_source, stored_name):
        if display_item and display_item.get("id"):
            try:
                delete_onedrive_file(access_token, display_item["id"])
            except Exception:
                pass
        if moved_source:
            try:
                move_onedrive_item(
                    access_token,
                    moved_source["item_id"],
                    moved_source["parent_id"],
                    moved_source["name"],
                )
            except Exception as rollback_exc:
                raise RuntimeError(
                    "写真はOneDrive内で移動されましたが、保存処理に失敗し、"
                    "元のカメラロールへ戻せませんでした。"
                    f"OneDriveで「{stored_name}」を確認してください。"
                ) from rollback_exc

    def save_to_onedrive(index, source_item):
        source_item_id = clean_value(source_item.get("id"), blank_text="")
        source_parent_reference = source_item.get("parentReference") or {}
        source_parent_id = clean_value(source_parent_reference.get("id"), blank_text="")
        original_name = Path(str(source_item.get("name") or "file")).name
        mime_type = clean_value(
            (source_item.get("file") or {}).get("mimeType"),
            blank_text="",
        ) or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
        suffix = Path(original_name).suffix.casefold()
        if not source_item_id or not source_parent_id:
            raise RuntimeError("OneDriveの元写真の保存場所を確認できませんでした。")
        if suffix not in ONEDRIVE_IMAGE_EXTENSIONS and mime_type.casefold() not in allowed_mime_types:
            raise ValueError("JPG・JPEG・PNG・WEBPの写真だけ保存できます。")

        content = download_onedrive_file(access_token, source_item_id)
        if not content:
            raise RuntimeError("OneDriveの元写真を読み込めませんでした。")
        timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S")
        stored_name = f"{timestamp}_{uuid.uuid4().hex[:8]}_{original_name}"
        display = prepare_onedrive_display_image(content, original_name, stored_name)
        moved_source = {
            "item_id": source_item_id,
            "parent_id": source_parent_id,
            "name": original_name,
        }
        moved_item = None
        display_item = None
        try:
            moved_item = move_onedrive_item(
                access_token,
                source_item_id,
                target_folder_id,
                stored_name,
            )
            if display:
                display_item = upload_onedrive_file_to_existing_folder(
                    access_token,
                    folder_path,
                    display["stored_name"],
                    display["content"],
                    display["mime_type"],
                )
            return {
                "index": index,
                "original_name": original_name,
                "content": content,
                "mime_type": mime_type,
                "stored_name": stored_name,
                "display": display or {},
                "moved_item": moved_item,
                "display_item": display_item,
                "moved_source": moved_source,
            }
        except Exception:
            rollback_saved_files(display_item, moved_source if moved_item else None, stored_name)
            raise

    prepared_results = {}
    failed_by_index = {}
    worker_count = max(1, min(int(max_workers or 1), 3, len(items)))
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        future_map = {
            executor.submit(save_to_onedrive, index, item): index
            for index, item in enumerate(items)
        }
        for future in as_completed(future_map):
            index = future_map[future]
            try:
                prepared_results[index] = future.result()
            except Exception as exc:
                original_name = Path(str(items[index].get("name") or "名称未設定")).name
                failed_by_index[index] = (original_name, str(exc))

    saved_items = []
    for index in range(group_size):
        result = prepared_results.get(index)
        if result is None:
            continue
        registration_started = False
        try:
            display = result.get("display") or {}
            attachment = build_entity_onedrive_attachment_record(
                entity_type,
                entity_name,
                entity_id,
                result["original_name"],
                result["content"],
                result["mime_type"],
                tags,
                remarks,
                result["moved_item"],
                result["stored_name"],
                folder_path,
                uploaded_by,
                group_id=group_id,
                group_index=index if group_id else 0,
                group_size=group_size if group_id else 1,
                display_item=result.get("display_item"),
                requested_display_name=display.get("stored_name", ""),
                display_content=display.get("content", b""),
            )
            registration_started = True
            saved_items.append(
                register_entity_onedrive_attachment(
                    attachment,
                    access_token,
                    moved_source=result.get("moved_source"),
                )
            )
        except Exception as exc:
            if not registration_started:
                try:
                    rollback_saved_files(
                        result.get("display_item"),
                        result.get("moved_source"),
                        result.get("stored_name", ""),
                    )
                except Exception as rollback_exc:
                    exc = rollback_exc
            failed_by_index[index] = (result.get("original_name", "名称未設定"), str(exc))

    failed_items = [failed_by_index[index] for index in sorted(failed_by_index)]
    return saved_items, failed_items


def save_entity_onedrive_attachment(
    entity_type,
    entity_name,
    entity_id,
    uploaded_name,
    content,
    mime_type,
    tags,
    remarks,
    access_token,
    group_id="",
    group_index=0,
    group_size=1,
    move_from_camera_roll=False,
):
    entity_type = normalize_attachment_entity_type(entity_type)
    entity_name = clean_value(entity_name, blank_text="").strip()
    entity_id = clean_value(entity_id, blank_text="").strip()
    file_kind = attachment_file_kind(uploaded_name, mime_type)
    if not file_kind:
        raise ValueError("画像（JPG・JPEG・PNG・WEBP）またはPDFを選んでください。")
    if not content:
        raise ValueError("選択したファイルが空です。")

    folder_key = get_attachment_onedrive_folder_key(
        entity_type,
        entity_id,
        entity_name,
    )
    category_folder = "写真" if file_kind == "image" else "資料"
    folder_path = "/".join(
        [
            ONEDRIVE_ROOT_FOLDER,
            attachment_entity_folder(entity_type),
            folder_key,
            category_folder,
        ]
    )
    original_name = Path(str(uploaded_name or "file")).name
    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S")
    stored_name = f"{timestamp}_{uuid.uuid4().hex[:8]}_{original_name}"
    display = (
        prepare_onedrive_display_image(content, original_name, stored_name)
        if file_kind == "image"
        else None
    )

    moved_source = None
    uploaded_item = None
    display_item = None
    try:
        if file_kind == "image" and move_from_camera_roll:
            source_item = find_matching_onedrive_camera_roll_file(
                access_token,
                original_name,
                content,
            )
            if source_item:
                source_parent_reference = source_item.get("parentReference") or {}
                source_item_id = clean_value(source_item.get("id"), blank_text="")
                source_parent_id = clean_value(source_parent_reference.get("id"), blank_text="")
                source_name = clean_value(source_item.get("name"), blank_text="") or original_name
                if not source_item_id or not source_parent_id:
                    raise RuntimeError(
                        "OneDriveの元写真の保存場所を確認できませんでした。"
                        "誤移動を防ぐため保存を停止しました。"
                    )
                uploaded_item = move_onedrive_file_to_folder(
                    access_token,
                    source_item_id,
                    folder_path,
                    stored_name,
                )
                moved_source = {
                    "item_id": source_item_id,
                    "parent_id": source_parent_id,
                    "name": source_name,
                }
            else:
                uploaded_item = upload_onedrive_file(
                    access_token,
                    folder_path,
                    stored_name,
                    content,
                    mime_type,
                )
        else:
            uploaded_item = upload_onedrive_file(
                access_token,
                folder_path,
                stored_name,
                content,
                mime_type,
            )

        if display:
            display_item = upload_onedrive_file_to_existing_folder(
                access_token,
                folder_path,
                display["stored_name"],
                display["content"],
                display["mime_type"],
            )
    except Exception:
        if display_item and display_item.get("id"):
            try:
                delete_onedrive_file(access_token, display_item["id"])
            except Exception:
                pass
        if moved_source:
            try:
                move_onedrive_item(
                    access_token,
                    moved_source["item_id"],
                    moved_source["parent_id"],
                    moved_source["name"],
                )
            except Exception:
                pass
        elif uploaded_item and uploaded_item.get("id"):
            try:
                delete_onedrive_file(access_token, uploaded_item["id"])
            except Exception:
                pass
        raise

    uploaded_by = get_attachment_uploaded_by(access_token)
    attachment = build_entity_onedrive_attachment_record(
        entity_type,
        entity_name,
        entity_id,
        original_name,
        content,
        mime_type,
        tags,
        remarks,
        uploaded_item,
        stored_name,
        folder_path,
        uploaded_by,
        group_id=group_id,
        group_index=group_index,
        group_size=group_size,
        display_item=display_item,
        requested_display_name=(display or {}).get("stored_name", ""),
        display_content=(display or {}).get("content", b""),
    )
    return register_entity_onedrive_attachment(
        attachment,
        access_token,
        moved_source=moved_source,
    )

def save_customer_onedrive_attachment(
    customer_name,
    customer_key,
    uploaded_name,
    content,
    mime_type,
    tags,
    remarks,
    access_token,
):
    return save_entity_onedrive_attachment(
        "customer",
        customer_name,
        customer_key,
        uploaded_name,
        content,
        mime_type,
        tags,
        remarks,
        access_token,
    )


def update_customer_onedrive_attachment_metadata(attachment, tags, remarks):
    updated = dict(attachment)
    updated["tags"] = normalize_attachment_tags(tags)
    updated["remarks"] = str(remarks or "").strip()
    update_customer_information(
        attachment["id"],
        attachment["field_name"],
        serialize_onedrive_attachment(updated),
    )
    return updated


def format_attachment_size(value):
    try:
        size = float(value or 0)
    except Exception:
        size = 0
    units = ["B", "KB", "MB", "GB"]
    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.0f} {unit}" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024
    return "0 B"


def format_attachment_datetime(value):
    text = clean_value(value, blank_text="")
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.astimezone(timezone(timedelta(hours=9))).strftime("%Y/%m/%d %H:%M")
    except Exception:
        return text


def render_collapsible_attachment_remarks(value):
    """備考は通常2行まで表示し、長い場合だけ全文を折りたたんで表示する。"""
    remarks = clean_value(value, blank_text="").strip()
    if not remarks:
        return

    normalized = remarks.replace("\r\n", "\n").replace("\r", "\n")
    visible_character_count = len(re.sub(r"\s+", "", normalized))
    needs_collapse = len(normalized.split("\n")) > 2 or visible_character_count > 34
    if not needs_collapse:
        st.caption(normalized)
        return

    safe_remarks = html.escape(normalized).replace("\n", "<br>")
    st.markdown(
        f"""
        <details class="aoyama-attachment-remarks">
          <summary>
            <span class="aoyama-attachment-remarks-preview">{safe_remarks}</span>
            <span class="aoyama-attachment-remarks-toggle"></span>
          </summary>
          <div class="aoyama-attachment-remarks-full">{safe_remarks}</div>
        </details>
        <style>
          details.aoyama-attachment-remarks {{
            margin: 0.25rem 0 0.4rem 0;
            color: rgba(49, 51, 63, 0.68);
            font-size: 0.875rem;
            line-height: 1.6;
          }}
          details.aoyama-attachment-remarks summary {{
            cursor: pointer;
            list-style: none;
          }}
          details.aoyama-attachment-remarks summary::-webkit-details-marker {{
            display: none;
          }}
          .aoyama-attachment-remarks-preview {{
            display: -webkit-box;
            -webkit-box-orient: vertical;
            -webkit-line-clamp: 2;
            overflow: hidden;
            overflow-wrap: anywhere;
          }}
          .aoyama-attachment-remarks-toggle {{
            display: inline-block;
            margin-top: 0.12rem;
            color: rgb(0, 104, 201);
            font-size: 0.82rem;
            font-weight: 600;
          }}
          .aoyama-attachment-remarks-toggle::before {{
            content: "続きを読む";
          }}
          details.aoyama-attachment-remarks[open] .aoyama-attachment-remarks-preview {{
            display: none;
          }}
          details.aoyama-attachment-remarks[open] .aoyama-attachment-remarks-toggle::before {{
            content: "閉じる";
          }}
          .aoyama-attachment-remarks-full {{
            margin-top: 0.22rem;
            overflow-wrap: anywhere;
            white-space: normal;
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_collapsible_record_remarks(value):
    """商品・運賃などの備考を全幅で2行まで表示し、長い場合だけ折りたたむ。"""
    remarks = clean_value(value, blank_text="").strip()
    if not remarks:
        return

    normalized = remarks.replace("\r\n", "\n").replace("\r", "\n")
    visible_character_count = len(re.sub(r"\s+", "", normalized))
    needs_collapse = len(normalized.split("\n")) > 2 or visible_character_count > 36
    safe_remarks = html.escape(normalized).replace("\n", "<br>")

    if not needs_collapse:
        st.markdown(
            f'<div class="aoyama-record-remarks-text">{safe_remarks}</div>'
            '<style>'
            '.aoyama-record-remarks-text {'
            'font-weight: 700; line-height: 1.65; overflow-wrap: anywhere; '
            'word-break: break-word; margin: 0.05rem 0 0.45rem 0;'
            '}'
            '</style>',
            unsafe_allow_html=True,
        )
        return

    st.markdown(
        f"""
        <details class="aoyama-record-remarks">
          <summary>
            <span class="aoyama-record-remarks-preview">{safe_remarks}</span>
            <span class="aoyama-record-remarks-toggle"></span>
          </summary>
          <div class="aoyama-record-remarks-full">{safe_remarks}</div>
        </details>
        <style>
          details.aoyama-record-remarks {{
            width: 100%;
            margin: 0.05rem 0 0.45rem 0;
            font-weight: 700;
            line-height: 1.65;
            overflow-wrap: anywhere;
            word-break: break-word;
          }}
          details.aoyama-record-remarks summary {{
            cursor: pointer;
            list-style: none;
          }}
          details.aoyama-record-remarks summary::-webkit-details-marker {{
            display: none;
          }}
          .aoyama-record-remarks-preview {{
            display: -webkit-box;
            -webkit-box-orient: vertical;
            -webkit-line-clamp: 2;
            overflow: hidden;
            overflow-wrap: anywhere;
            word-break: break-word;
          }}
          .aoyama-record-remarks-toggle {{
            display: inline-block;
            margin-top: 0.12rem;
            color: rgb(0, 104, 201);
            font-size: 0.86rem;
            font-weight: 600;
          }}
          .aoyama-record-remarks-toggle::before {{
            content: "続きを読む";
          }}
          details.aoyama-record-remarks[open] .aoyama-record-remarks-preview {{
            display: none;
          }}
          details.aoyama-record-remarks[open] .aoyama-record-remarks-toggle::before {{
            content: "閉じる";
          }}
          .aoyama-record-remarks-full {{
            margin-top: 0.22rem;
            overflow-wrap: anywhere;
            word-break: break-word;
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )



def render_attachment_card_layout_styles():
    """写真・資料の一覧カードだけを、行内で上揃え・同じ高さに整える。"""
    st.markdown(
        """
        <style>
          .aoyama-attachment-card-marker {
            display: none !important;
          }

          /* 既存のスマホ向け中央揃えを、このカード一覧だけ上書きする。 */
          [data-testid="stHorizontalBlock"]:has(.aoyama-attachment-card-marker) {
            align-items: stretch !important;
          }
          [data-testid="stHorizontalBlock"]:has(.aoyama-attachment-card-marker) > div {
            align-self: stretch !important;
            display: flex !important;
            min-width: 0 !important;
          }
          [data-testid="stHorizontalBlock"]:has(.aoyama-attachment-card-marker)
            > div > [data-testid="stVerticalBlock"] {
            display: flex !important;
            flex: 1 1 auto !important;
            flex-direction: column !important;
            width: 100% !important;
            min-width: 0 !important;
          }
          [data-testid="stVerticalBlockBorderWrapper"]:has(.aoyama-attachment-card-marker) {
            display: flex !important;
            flex: 1 1 auto !important;
            width: 100% !important;
          }
          [data-testid="stVerticalBlockBorderWrapper"]:has(.aoyama-attachment-card-marker)
            > [data-testid="stVerticalBlock"] {
            width: 100% !important;
          }

          .aoyama-attachment-card-entity {
            min-height: 3.55rem;
            margin: 0.08rem 0 0.2rem 0;
            line-height: 1.45;
            overflow: hidden;
          }
          .aoyama-attachment-card-entity a {
            display: -webkit-box;
            overflow: hidden;
            color: rgb(0, 104, 201);
            font-weight: 700;
            text-decoration: none;
            overflow-wrap: anywhere;
            -webkit-box-orient: vertical;
            -webkit-line-clamp: 2;
          }
          .aoyama-attachment-card-date {
            min-height: 1.55rem;
            color: rgba(49, 51, 63, 0.62);
            font-size: 0.875rem;
            line-height: 1.45;
          }
          .aoyama-attachment-card-tags {
            min-height: 2.7rem;
            margin: 0.08rem 0 0.16rem 0;
            line-height: 1.75;
            overflow-wrap: anywhere;
          }
          .aoyama-attachment-card-tag {
            display: inline-block;
            margin: 0 0.16rem 0.14rem 0;
            padding: 0.04rem 0.28rem;
            border-radius: 0.24rem;
            background: rgba(22, 163, 74, 0.06);
            color: rgb(20, 128, 59);
            font-size: 0.82rem;
            line-height: 1.45;
            text-decoration: none;
            cursor: pointer;
          }
          .aoyama-attachment-card-tag:hover,
          .aoyama-attachment-card-tag:focus-visible {
            background: rgba(22, 163, 74, 0.14);
            color: rgb(16, 104, 48);
            text-decoration: none;
          }
        </style>
        """,
        unsafe_allow_html=True,
    )

def onedrive_attachment_rows_to_dataframe(rows):
    records = []
    for row in rows:
        attachment = parse_onedrive_attachment_item(row)
        if not attachment:
            continue
        entity_type = normalize_attachment_entity_type(
            attachment.get("entity_type")
        )
        records.append(
            {
                "顧客ID": attachment.get("entity_id", "") if entity_type == "customer" else "",
                "顧客名": attachment.get("entity_name", "") if entity_type == "customer" else "",
                "取引先種別": attachment_entity_label(entity_type),
                "取引先ID": attachment.get("entity_id", ""),
                "取引先名": attachment.get("entity_name", ""),
                "種類": "写真" if attachment.get("file_type") == "image" else "PDF",
                "元ファイル名": attachment.get("original_name", ""),
                "OneDrive保存名": attachment.get("stored_name", ""),
                "OneDrive保存先": attachment.get("onedrive_path", ""),
                "タグ": " ".join(f"#{tag}" for tag in attachment.get("tags", [])),
                "備考": attachment.get("remarks", ""),
                "サイズ": attachment.get("size", ""),
                "登録者": attachment.get("uploaded_by", ""),
                "登録日時": attachment.get("created_at", ""),
                "OneDriveファイルID": attachment.get("file_id", ""),
                "保存ID": attachment.get("id", ""),
            }
        )
    records.sort(key=lambda record: str(record.get("登録日時") or ""), reverse=True)
    return backup_dataframe(
        records,
        [
            "顧客ID", "顧客名", "取引先種別", "取引先ID", "取引先名",
            "種類", "元ファイル名", "OneDrive保存名", "OneDrive保存先",
            "タグ", "備考", "サイズ", "登録者", "登録日時",
            "OneDriveファイルID", "保存ID",
        ],
    )


def build_attachment_onedrive_sign_in_url(entity_type, entity_id, entity_name):
    entity_type = normalize_attachment_entity_type(entity_type)
    if entity_type == "customer":
        return build_onedrive_sign_in_url("detail", entity_name)
    if entity_type == "hotel":
        return build_onedrive_sign_in_url("hotel_information")
    return build_onedrive_sign_in_url(
        "partner_detail",
        partner_id=entity_id,
        partner_type=entity_type,
    )


def render_customer_attachments_section(
    customer_name,
    customer_key=None,
    entity_type="customer",
):
    entity_type = normalize_attachment_entity_type(entity_type)
    entity_name = clean_value(customer_name, blank_text="").strip()
    entity_id = clean_value(customer_key, blank_text="").strip()
    identity = f"{entity_type}:{entity_id or entity_name}"
    suffix = hashlib.sha256(str(identity).encode("utf-8")).hexdigest()[:16]
    add_form_version_key = f"onedrive_attachment_add_form_version_{suffix}"
    camera_form_version_key = f"onedrive_attachment_camera_form_version_{suffix}"
    recent_form_version_key = f"onedrive_attachment_recent_form_version_{suffix}"
    add_form_version = int(st.session_state.get(add_form_version_key, 0) or 0)
    camera_form_version = int(st.session_state.get(camera_form_version_key, 0) or 0)
    recent_form_version = int(st.session_state.get(recent_form_version_key, 0) or 0)
    add_widget_suffix = f"{suffix}_{add_form_version}"
    camera_widget_suffix = f"{suffix}_{camera_form_version}"
    recent_widget_suffix = f"{suffix}_{recent_form_version}"
    recent_items_key = f"onedrive_attachment_recent_items_{suffix}"
    recent_limit_key = f"onedrive_attachment_recent_limit_{suffix}"
    recent_loaded_key = f"onedrive_attachment_recent_loaded_{suffix}"
    recent_error_key = f"onedrive_attachment_recent_error_{suffix}"
    success_key = f"onedrive_attachment_success_{suffix}"
    upload_error_key = f"onedrive_attachment_upload_error_{suffix}"
    edit_key = f"onedrive_attachment_edit_{suffix}"
    limit_key = f"onedrive_attachment_limit_{suffix}"
    open_key = f"onedrive_attachment_force_open_{suffix}"
    restore_key = f"onedrive_attachment_restore_scroll_{suffix}"

    if not has_supabase_config():
        with st.expander("📎 写真・資料"):
            st.warning("写真・資料の管理にはSupabase設定が必要です。")
        return

    try:
        attachments = get_entity_attachments(entity_type, entity_name, entity_id)
    except Exception as exc:
        with st.expander("📎 写真・資料"):
            st.warning(f"写真・資料の一覧を読み込めませんでした：{exc}")
        return

    attachment_groups = group_onedrive_attachments_for_display(attachments)
    force_open = bool(st.session_state.pop(open_key, False))
    restore_scroll = bool(st.session_state.pop(restore_key, False))
    with st.expander(f"📎 写真・資料　{len(attachment_groups)}件", expanded=force_open):
        render_onedrive_attachment_scroll_keeper(suffix, restore=restore_scroll)
        success_message = st.session_state.pop(success_key, None)
        if success_message:
            st.success(success_message)
        upload_error_message = st.session_state.pop(upload_error_key, None)
        if upload_error_message:
            st.error(upload_error_message)
        auth_success = st.session_state.pop("onedrive_auth_success", None)
        if auth_success:
            st.success("OneDriveの初回接続が完了しました。")
        auth_error = st.session_state.pop("onedrive_auth_error", None)
        if auth_error:
            st.error(auth_error)

        try:
            read_onedrive_settings()
        except Exception as exc:
            st.warning(str(exc))
            st.code(
                "[onedrive]\n"
                'client_id = "MicrosoftのクライアントID"\n'
                'client_secret = "Microsoftのシークレットの値"\n'
                'redirect_uri = "https://aoyama-kokyaku.streamlit.app"\n'
                'refresh_token = "初回接続後に表示される値"'
            )
            return

        configured_refresh_token = read_onedrive_configured_refresh_token()
        setup_refresh_token = str(
            st.session_state.get("onedrive_refresh_token_setup_value") or ""
        ).strip()
        if setup_refresh_token and not configured_refresh_token:
            st.warning(
                "次回から自動接続するため、下の1行を顧客カルテの"
                "Streamlit Secretsにある[onedrive]の中へ追加してください。"
            )
            st.code(f'refresh_token = "{setup_refresh_token}"', language="toml")
            st.caption("追加して保存するとアプリが再起動し、通常画面から接続ボタンが消えます。")

        access_token = get_onedrive_access_token()

        # Streamlitのタブは、選択していないタブの中身も実行される。
        # そのため全体タグ履歴はアルバム閲覧だけでは読み込まず、
        # 実際に追加するファイルを選んだ時、写真を撮った時、または編集開始時だけ取得する。
        tag_history_options_cache = None

        def load_tag_history_options_for_action():
            nonlocal tag_history_options_cache
            if tag_history_options_cache is None:
                tag_history_options_cache = get_attachment_tag_history_options(attachments)
            return list(tag_history_options_cache or [])

        mobile_browser = is_mobile_browser()
        if mobile_browser:
            album_tab, add_tab, camera_tab = st.tabs(
                ["🖼 アルバム", "➕ 追加", "📷 写真撮影"]
            )
        else:
            album_tab, add_tab = st.tabs(["🖼 アルバム", "➕ 追加"])
            camera_tab = None

        # 追加タブは保存済み画像とPDF専用。タブの初期表示は先頭のアルバム。
        with add_tab:
            if not access_token:
                if configured_refresh_token:
                    st.warning(
                        "OneDriveへ自動接続できませんでした。Microsoft側で認証が失効した可能性があります。"
                    )
                    connect_label = "OneDriveを再接続（管理者用）"
                else:
                    st.info(
                        "最初の1回だけ管理者がOneDriveへ接続し、表示された更新トークンをSecretsへ追加してください。"
                    )
                    connect_label = "OneDrive初回設定（管理者用）"
                try:
                    st.link_button(
                        connect_label,
                        build_attachment_onedrive_sign_in_url(entity_type, entity_id, entity_name),
                        use_container_width=True,
                    )
                except Exception as exc:
                    st.error(f"OneDriveへの接続を開始できませんでした：{exc}")
            else:
                st.markdown("#### 追加")
                st.markdown("##### ☁️ OneDriveの最近の写真から選ぶ")
                st.caption(
                    "OneDriveのカメラバックアップにある直近14日以内の写真を、"
                    "スマホから再送信せずに追加します。写真は最後の保存ボタンを押すまで移動しません。"
                )
                recent_items = st.session_state.get(recent_items_key, [])
                if not isinstance(recent_items, list):
                    recent_items = []
                load_recent_label = "一覧を更新" if recent_items else "最近の写真を読み込む"
                if st.button(
                    load_recent_label,
                    key=f"onedrive_attachment_recent_load_{suffix}",
                    use_container_width=True,
                ):
                    try:
                        with st.spinner("OneDriveの最近の写真を確認しています…"):
                            recent_items = load_recent_onedrive_camera_photos(access_token)
                        st.session_state[recent_items_key] = recent_items
                        st.session_state[recent_limit_key] = ONEDRIVE_RECENT_CAMERA_PAGE_SIZE
                        st.session_state[recent_loaded_key] = True
                        st.session_state.pop(recent_error_key, None)
                        st.session_state[recent_form_version_key] = recent_form_version + 1
                        st.rerun()
                    except Exception as exc:
                        st.session_state[recent_error_key] = str(exc)
                        st.rerun()

                recent_error = st.session_state.pop(recent_error_key, None)
                if recent_error:
                    st.warning(f"最近の写真を読み込めませんでした：{recent_error}")

                recent_items = st.session_state.get(recent_items_key, [])
                if isinstance(recent_items, list) and recent_items:
                    recent_limit = int(
                        st.session_state.get(
                            recent_limit_key,
                            ONEDRIVE_RECENT_CAMERA_PAGE_SIZE,
                        )
                        or ONEDRIVE_RECENT_CAMERA_PAGE_SIZE
                    )
                    visible_recent_items = recent_items[:recent_limit]
                    recent_ids = [
                        clean_value(item.get("id"), blank_text="")
                        for item in visible_recent_items
                        if clean_value(item.get("id"), blank_text="")
                    ]
                    recent_thumbnails = download_onedrive_thumbnail_batch(
                        access_token,
                        tuple(recent_ids),
                    ) if recent_ids else {}
                    selected_recent_items = []
                    recent_grid_count = 2 if mobile_browser else 3
                    recent_columns = []
                    for recent_index, recent_item in enumerate(visible_recent_items):
                        if recent_index % recent_grid_count == 0:
                            recent_columns = st.columns(recent_grid_count, gap="small")
                        item_id = clean_value(recent_item.get("id"), blank_text="")
                        original_name = Path(
                            str(recent_item.get("name") or "名称未設定")
                        ).name
                        item_key = hashlib.sha256(item_id.encode("utf-8")).hexdigest()[:16]
                        modified_at = parse_onedrive_datetime(
                            recent_item.get("lastModifiedDateTime")
                            or recent_item.get("createdDateTime")
                        )
                        modified_text = ""
                        if modified_at is not None:
                            modified_text = modified_at.astimezone(
                                timezone(timedelta(hours=9))
                            ).strftime("%m/%d %H:%M")
                        with recent_columns[recent_index % recent_grid_count]:
                            with st.container(border=True):
                                thumbnail = recent_thumbnails.get(item_id)
                                if isinstance(thumbnail, bytes):
                                    st.image(thumbnail, use_container_width=True)
                                else:
                                    st.caption("サムネイルを表示できません")
                                selected = st.checkbox(
                                    "この写真を選ぶ",
                                    key=(
                                        "onedrive_attachment_recent_select_"
                                        f"{recent_widget_suffix}_{item_key}"
                                    ),
                                )
                                st.caption(
                                    original_name
                                    + (f"　{modified_text}" if modified_text else "")
                                )
                                if selected:
                                    selected_recent_items.append(recent_item)

                    st.caption(
                        f"{len(recent_items)}件中 {len(visible_recent_items)}件を表示　｜　"
                        f"{len(selected_recent_items)}件を選択中"
                    )
                    if len(recent_items) > recent_limit:
                        if st.button(
                            "さらに表示",
                            key=f"onedrive_attachment_recent_more_{suffix}",
                            use_container_width=True,
                        ):
                            st.session_state[recent_limit_key] = (
                                recent_limit + ONEDRIVE_RECENT_CAMERA_PAGE_SIZE
                            )
                            st.rerun()

                    recent_tag_options = (
                        load_tag_history_options_for_action()
                        if selected_recent_items
                        else []
                    )
                    recent_history_tags = st.multiselect(
                        "選択した写真のタグ",
                        recent_tag_options,
                        key=f"onedrive_attachment_recent_history_tags_{recent_widget_suffix}",
                    )
                    recent_new_tags_text = st.text_input(
                        "選択した写真の新しいタグ",
                        placeholder="例：北海道、タンク、要確認",
                        key=f"onedrive_attachment_recent_new_tags_{recent_widget_suffix}",
                        autocomplete="off",
                    )
                    recent_remarks = st.text_area(
                        "選択した写真の備考",
                        placeholder="写真について残したい内容",
                        height=90,
                        key=f"onedrive_attachment_recent_remarks_{recent_widget_suffix}",
                    )
                    if st.button(
                        "選択した写真を保存",
                        type="primary",
                        use_container_width=True,
                        key=f"onedrive_attachment_recent_save_{recent_widget_suffix}",
                    ):
                        if not selected_recent_items:
                            st.warning("保存する写真を選んでください。")
                        else:
                            recent_tags = list(recent_history_tags) + normalize_attachment_tags(
                                recent_new_tags_text
                            )
                            recent_group_id = (
                                uuid.uuid4().hex if len(selected_recent_items) > 1 else ""
                            )
                            with st.spinner("選択した写真をOneDrive内で保存しています…"):
                                try:
                                    recent_saved_items, recent_failed_items = (
                                        save_selected_onedrive_camera_photos_parallel(
                                            entity_type,
                                            entity_name,
                                            entity_id,
                                            selected_recent_items,
                                            recent_tags,
                                            recent_remarks,
                                            access_token,
                                            group_id=recent_group_id,
                                            max_workers=3,
                                        )
                                    )
                                except Exception as exc:
                                    recent_saved_items = []
                                    recent_failed_items = [
                                        (
                                            Path(
                                                str(item.get("name") or "名称未設定")
                                            ).name,
                                            str(exc),
                                        )
                                        for item in selected_recent_items
                                    ]

                            for saved in recent_saved_items:
                                remember_change_history_warning(
                                    record_change_history_safely(
                                        attachment_entity_label(entity_type),
                                        entity_id or "",
                                        entity_name,
                                        "追加",
                                        {
                                            "ファイル": (
                                                "",
                                                saved.get("original_name", ""),
                                            ),
                                            "タグ": (
                                                "",
                                                " ".join(
                                                    f"#{tag}"
                                                    for tag in saved.get("tags", [])
                                                ),
                                            ),
                                        },
                                        section="写真・資料",
                                    )
                                )

                            if recent_saved_items:
                                saved_ids = {
                                    clean_value(saved.get("file_id"), blank_text="")
                                    for saved in recent_saved_items
                                }
                                st.session_state[recent_items_key] = [
                                    item
                                    for item in recent_items
                                    if clean_value(item.get("id"), blank_text="")
                                    not in saved_ids
                                ]
                                st.session_state[success_key] = (
                                    "写真・資料を保存しました。"
                                    if len(recent_saved_items) == 1
                                    else f"写真・資料を{len(recent_saved_items)}件保存しました。"
                                )
                                st.session_state[limit_key] = ONEDRIVE_PAGE_SIZE
                                st.session_state[recent_limit_key] = (
                                    ONEDRIVE_RECENT_CAMERA_PAGE_SIZE
                                )
                                st.session_state[recent_form_version_key] = (
                                    recent_form_version + 1
                                )
                            if recent_failed_items:
                                failed_names = "、".join(
                                    name for name, _ in recent_failed_items
                                )
                                first_error = recent_failed_items[0][1]
                                st.session_state[upload_error_key] = (
                                    f"{len(recent_failed_items)}件保存できませんでした："
                                    f"{failed_names}"
                                    + (f"（{first_error}）" if first_error else "")
                                )
                            if recent_saved_items or recent_failed_items:
                                st.rerun()
                elif bool(st.session_state.get(recent_loaded_key, False)):
                    st.info("直近14日以内のJPG・JPEG・PNG・WEBP写真は見つかりませんでした。")
                elif isinstance(recent_items, list):
                    st.caption(
                        "直近14日以内の対応画像がまだ読み込まれていません。"
                        "上のボタンで確認してください。"
                    )

                st.markdown("---")
                st.markdown("##### 端末から画像・PDFを追加")
                # Androidなどで3枚以上をまとめて選ぶと、拡張子フィルター付きの
                # 写真ピッカーがStreamlitへ結果を返さない場合がある。
                # スマホだけ汎用ファイル選択に切り替え、受け取り後に画像形式を厳密に確認する。
                image_uploader_types = None if mobile_browser else ["jpg", "jpeg", "png", "webp"]
                image_uploader_label = (
                    "🖼 保存済み画像を選ぶ" if mobile_browser else "🖼 画像を選ぶ"
                )
                selected_image_files = st.file_uploader(
                    image_uploader_label,
                    type=image_uploader_types,
                    accept_multiple_files=True,
                    key=f"onedrive_attachment_photo_uploader_{add_widget_suffix}",
                )
                if mobile_browser:
                    enable_mobile_bulk_image_picker(image_uploader_label)
                    st.caption(
                        "OneDriveのカメラバックアップに同じ写真がある場合は、"
                        "その写真をこの取引先の写真フォルダへ移動します。"
                        "同じ写真が見つからない場合は、選んだ画像を新規アップロードします。"
                    )
                photo_files = list(selected_image_files or [])
                supported_image_mime_types = {"image/jpeg", "image/png", "image/webp"}
                invalid_photo_files = [
                    uploaded
                    for uploaded in photo_files
                    if (
                        Path(str(getattr(uploaded, "name", "") or "")).suffix.lower()
                        not in ONEDRIVE_IMAGE_EXTENSIONS
                        and str(getattr(uploaded, "type", "") or "").lower()
                        not in supported_image_mime_types
                    )
                ]
                if photo_files:
                    st.caption(f"画像を{len(photo_files)}枚選択中")
                if invalid_photo_files:
                    invalid_names = "、".join(
                        Path(str(getattr(uploaded, "name", "") or "名称未設定")).name
                        for uploaded in invalid_photo_files
                    )
                    st.warning(
                        "画像はJPG・JPEG・PNG・WEBPだけ選べます：" + invalid_names
                    )
                pdf_file = st.file_uploader(
                    "📄 PDFを選ぶ",
                    type=["pdf"],
                    accept_multiple_files=False,
                    key=f"onedrive_attachment_pdf_uploader_{add_widget_suffix}",
                )

                pending_add_files = bool(photo_files) or pdf_file is not None
                add_tag_history_options = (
                    load_tag_history_options_for_action()
                    if pending_add_files
                    else []
                )
                selected_history_tags = st.multiselect(
                    "タグ（入力すると過去の候補を絞り込み）",
                    add_tag_history_options,
                    key=f"onedrive_attachment_history_tags_{add_widget_suffix}",
                )
                new_tags_text = st.text_input(
                    "新しいタグ（候補にない場合）",
                    placeholder="例：北海道、タンク、要確認",
                    key=f"onedrive_attachment_new_tags_{add_widget_suffix}",
                    autocomplete="off",
                )
                if add_tag_history_options:
                    st.caption(
                        "最近使ったタグ："
                        + "　".join(f"#{tag}" for tag in add_tag_history_options[:8])
                    )
                remarks = st.text_area(
                    "備考",
                    placeholder="写真や資料について残したい内容",
                    height=90,
                    key=f"onedrive_attachment_remarks_{add_widget_suffix}",
                )
                if st.button(
                    "OneDriveへ保存",
                    type="primary",
                    use_container_width=True,
                    key=f"onedrive_attachment_upload_{add_widget_suffix}",
                ):
                    uploaded_files = photo_files if photo_files else ([pdf_file] if pdf_file is not None else [])
                    if invalid_photo_files:
                        st.warning("対応していないファイルを外してから保存してください。")
                    elif not uploaded_files:
                        st.warning("画像またはPDFを選んでください。")
                    else:
                        tags = list(selected_history_tags) + normalize_attachment_tags(new_tags_text)
                        saved_items = []
                        failed_items = []
                        image_group_id = uuid.uuid4().hex if len(photo_files) > 1 else ""
                        image_group_size = len(photo_files) if photo_files else 1
                        with st.spinner("OneDriveへ保存しています…"):
                            if len(photo_files) > 1:
                                photo_uploads = [
                                    {
                                        "name": uploaded.name,
                                        "content": uploaded.getvalue(),
                                        "mime_type": uploaded.type
                                        or mimetypes.guess_type(uploaded.name)[0]
                                        or "application/octet-stream",
                                    }
                                    for uploaded in photo_files
                                ]
                                try:
                                    saved_items, failed_items = save_multiple_entity_onedrive_photos_parallel(
                                        entity_type,
                                        entity_name,
                                        entity_id,
                                        photo_uploads,
                                        tags,
                                        remarks,
                                        access_token,
                                        image_group_id,
                                        max_workers=3,
                                        move_from_camera_roll=bool(mobile_browser),
                                    )
                                except Exception as exc:
                                    saved_items = []
                                    failed_items = [
                                        (uploaded.name, str(exc))
                                        for uploaded in photo_files
                                    ]
                                for saved in saved_items:
                                    remember_change_history_warning(
                                        record_change_history_safely(
                                            attachment_entity_label(entity_type),
                                            entity_id or "",
                                            entity_name,
                                            "追加",
                                            {
                                                "ファイル": ("", saved.get("original_name", "")),
                                                "タグ": ("", " ".join(f"#{tag}" for tag in saved.get("tags", []))),
                                            },
                                            section="写真・資料",
                                        )
                                    )
                            else:
                                for upload_index, uploaded in enumerate(uploaded_files):
                                    try:
                                        saved = save_entity_onedrive_attachment(
                                            entity_type,
                                            entity_name,
                                            entity_id,
                                            uploaded.name,
                                            uploaded.getvalue(),
                                            uploaded.type or mimetypes.guess_type(uploaded.name)[0] or "application/octet-stream",
                                            tags,
                                            remarks,
                                            access_token,
                                            group_id=image_group_id,
                                            group_index=upload_index if image_group_id else 0,
                                            group_size=image_group_size if image_group_id else 1,
                                            move_from_camera_roll=bool(mobile_browser and photo_files),
                                        )
                                        saved_items.append(saved)
                                        remember_change_history_warning(
                                            record_change_history_safely(
                                                attachment_entity_label(entity_type),
                                                entity_id or "",
                                                entity_name,
                                                "追加",
                                                {
                                                    "ファイル": ("", saved.get("original_name", "")),
                                                    "タグ": ("", " ".join(f"#{tag}" for tag in saved.get("tags", []))),
                                                },
                                                section="写真・資料",
                                            )
                                        )
                                    except Exception as exc:
                                        failed_items.append((uploaded.name, str(exc)))

                        if saved_items:
                            if len(saved_items) == 1:
                                st.session_state[success_key] = "写真・資料を保存しました。"
                            else:
                                st.session_state[success_key] = f"写真・資料を{len(saved_items)}件保存しました。"
                            st.session_state[limit_key] = ONEDRIVE_PAGE_SIZE
                            # 全件正常保存できた時だけ、次の追加を空の入力状態で始める。
                            # 一部失敗時は、再確認できるよう現在の入力を残す。
                            if not failed_items:
                                st.session_state[add_form_version_key] = add_form_version + 1
                        if failed_items:
                            failed_names = "、".join(name for name, _ in failed_items)
                            first_error = failed_items[0][1]
                            st.session_state[upload_error_key] = (
                                f"{len(failed_items)}件保存できませんでした：{failed_names}"
                                + (f"（{first_error}）" if first_error else "")
                            )
                        if saved_items or failed_items:
                            st.rerun()

            # 自動接続設定後は利用者ごとのMicrosoft接続操作を表示しない。
            if access_token and not configured_refresh_token:
                st.markdown("---")
                if st.button(
                    "初回設定中の一時接続を解除",
                    key=f"onedrive_attachment_signout_{suffix}",
                    use_container_width=True,
                ):
                    clear_onedrive_auth_state(clear_shared=True)
                    st.session_state.pop("onedrive_refresh_token_setup_value", None)
                    st.rerun()

        # スマホだけ、写真撮影を追加タブから分離する。
        if camera_tab is not None:
            with camera_tab:
                if not access_token:
                    st.info("先に「＋追加」タブでOneDriveへ接続してください。")
                else:
                    st.markdown("#### 写真撮影")
                    camera_label = "📷 写真を撮る"
                    camera_file = st.file_uploader(
                        camera_label,
                        type=["jpg", "jpeg", "png", "webp"],
                        accept_multiple_files=False,
                        key=f"onedrive_attachment_camera_uploader_{camera_widget_suffix}",
                        help="スマホの背面カメラを起動して撮影します。",
                    )
                    enable_mobile_camera_capture(camera_label)

                    camera_tag_history_options = (
                        load_tag_history_options_for_action()
                        if camera_file is not None
                        else []
                    )
                    camera_history_tags = st.multiselect(
                        "タグ（入力すると過去の候補を絞り込み）",
                        camera_tag_history_options,
                        key=f"onedrive_attachment_camera_history_tags_{camera_widget_suffix}",
                    )
                    camera_new_tags_text = st.text_input(
                        "新しいタグ（候補にない場合）",
                        placeholder="例：北海道、タンク、要確認",
                        key=f"onedrive_attachment_camera_new_tags_{camera_widget_suffix}",
                        autocomplete="off",
                    )
                    if camera_tag_history_options:
                        st.caption(
                            "最近使ったタグ："
                            + "　".join(f"#{tag}" for tag in camera_tag_history_options[:8])
                        )
                    camera_remarks = st.text_area(
                        "備考",
                        placeholder="写真について残したい内容",
                        height=90,
                        key=f"onedrive_attachment_camera_remarks_{camera_widget_suffix}",
                    )
                    if st.button(
                        "OneDriveへ保存",
                        type="primary",
                        use_container_width=True,
                        key=f"onedrive_attachment_camera_upload_{camera_widget_suffix}",
                    ):
                        if camera_file is None:
                            st.warning("写真を撮ってください。")
                        else:
                            try:
                                tags = list(camera_history_tags) + normalize_attachment_tags(
                                    camera_new_tags_text
                                )
                                with st.spinner("OneDriveへ保存しています…"):
                                    saved = save_entity_onedrive_attachment(
                                        entity_type,
                                        entity_name,
                                        entity_id,
                                        camera_file.name,
                                        camera_file.getvalue(),
                                        camera_file.type
                                        or mimetypes.guess_type(camera_file.name)[0]
                                        or "application/octet-stream",
                                        tags,
                                        camera_remarks,
                                        access_token,
                                    )
                                remember_change_history_warning(
                                    record_change_history_safely(
                                        attachment_entity_label(entity_type),
                                        entity_id or "",
                                        entity_name,
                                        "追加",
                                        {
                                            "ファイル": ("", saved.get("original_name", "")),
                                            "タグ": (
                                                "",
                                                " ".join(
                                                    f"#{tag}" for tag in saved.get("tags", [])
                                                ),
                                            ),
                                        },
                                        section="写真・資料",
                                    )
                                )
                                st.session_state[success_key] = "写真・資料を保存しました。"
                                st.session_state[limit_key] = ONEDRIVE_PAGE_SIZE
                                # 保存後は、次の撮影を空のタグ・備考・写真選択で始める。
                                st.session_state[camera_form_version_key] = camera_form_version + 1
                                st.rerun()
                            except Exception as exc:
                                st.error(f"保存できませんでした：{exc}")


        with album_tab:
            st.markdown("#### アルバム")
            if not attachments:
                st.info("保存されている写真・資料はありません。")
                return

            type_filter = st.selectbox(
                "種類",
                ["すべて", "写真", "PDF"],
                key=f"onedrive_attachment_type_filter_{suffix}",
            )
            all_tags = attachment_tag_history_options(attachments)
            tag_filter = st.multiselect(
                "タグで絞り込み（入力すると候補を絞り込み）",
                all_tags,
                key=f"onedrive_attachment_tag_filter_{suffix}",
            ) if all_tags else []

            filtered = []
            for attachment in attachments:
                if type_filter == "写真" and attachment.get("file_type") != "image":
                    continue
                if type_filter == "PDF" and attachment.get("file_type") != "pdf":
                    continue
                if tag_filter and not set(tag_filter).issubset(set(attachment.get("tags", []))):
                    continue
                filtered.append(attachment)

            limit = int(st.session_state.get(limit_key, ONEDRIVE_PAGE_SIZE))
            active_edit_id = st.session_state.get(edit_key)

            if not filtered:
                st.info("条件に一致する写真・資料はありません。")

            display_groups = group_onedrive_attachments_for_display(filtered)
            visible_groups = display_groups[:limit]
            grid_column_count = 2 if is_mobile_browser() else 3
            grid_columns = []

            # 顧客カルテの初回表示だけ、画面に出す画像の先頭サムネイルを並行取得する。
            # Streamlitのsession_state更新はメインスレッドで行い、既存の修復処理は
            # 取得できなかった場合の従来フォールバックとしてそのまま残す。
            if entity_type == "customer" and access_token:
                prefetch_ids = []
                for attachment_group in visible_groups:
                    representative = attachment_group.get("representative", {})
                    if representative.get("file_type") != "image":
                        continue
                    for thumbnail_item in attachment_group.get("items", []):
                        thumbnail_item_id = clean_value(
                            thumbnail_item.get("file_id"),
                            blank_text="",
                        ).strip()
                        if not thumbnail_item_id:
                            continue
                        thumb_key = f"onedrive_thumbnail_{thumbnail_item_id}"
                        if isinstance(st.session_state.get(thumb_key), bytes):
                            break
                        prefetch_ids.append(thumbnail_item_id)
                        break

                if prefetch_ids:
                    prefetched = download_onedrive_thumbnail_batch(
                        access_token,
                        tuple(prefetch_ids),
                    )
                    for thumbnail_item_id, thumbnail in prefetched.items():
                        if isinstance(thumbnail, bytes):
                            st.session_state[
                                f"onedrive_thumbnail_{thumbnail_item_id}"
                            ] = thumbnail

            for group_index, attachment_group in enumerate(visible_groups):
                if group_index % grid_column_count == 0:
                    grid_columns = st.columns(grid_column_count, gap="small")

                group_items = attachment_group["items"]
                attachment = attachment_group["representative"]
                group_count = attachment_group["count"]
                group_ui_id = attachment_group["ui_id"]
                item_id = attachment.get("file_id", "")
                metadata_id = attachment.get("id", "")
                filename = attachment.get("original_name", "名称未設定")
                with grid_columns[group_index % grid_column_count]:
                    with st.container(border=True):
                        preview_digest = hashlib.sha256(group_ui_id.encode("utf-8")).digest()[:8]
                        preview_bits = "".join(f"{byte:08b}" for byte in preview_digest)
                        preview_trigger_label = "⁣" + "".join(
                            "​" if bit == "0" else "‌" for bit in preview_bits
                        )
                        preview_clicked = st.button(
                            preview_trigger_label,
                            key=f"onedrive_attachment_preview_button_{group_ui_id}",
                        )

                        if attachment.get("file_type") == "image":
                            thumbnail_content = None
                            thumbnail_filename = filename
                            if access_token:
                                for thumbnail_item in group_items:
                                    thumbnail_item_id = thumbnail_item.get("file_id", "")
                                    if not thumbnail_item_id:
                                        continue
                                    thumb_key = f"onedrive_thumbnail_{thumbnail_item_id}"
                                    if not isinstance(st.session_state.get(thumb_key), bytes):
                                        try:
                                            thumbnail = download_onedrive_attachment_thumbnail(
                                                access_token,
                                                thumbnail_item,
                                            )
                                            if thumbnail:
                                                st.session_state[thumb_key] = thumbnail
                                        except Exception:
                                            pass
                                    if isinstance(st.session_state.get(thumb_key), bytes):
                                        thumbnail_content = st.session_state[thumb_key]
                                        thumbnail_filename = thumbnail_item.get(
                                            "original_name",
                                            filename,
                                        )
                                        break
                            render_clickable_onedrive_thumbnail(
                                thumbnail_content,
                                thumbnail_filename,
                                preview_trigger_label,
                                compact_height=142 if is_mobile_browser() else 150,
                                badge_text=f"{group_count}枚" if group_count > 1 else "",
                            )
                        else:
                            render_clickable_onedrive_pdf_tile(
                                filename,
                                preview_trigger_label,
                                compact_height=142 if is_mobile_browser() else 150,
                            )

                        if preview_clicked:
                            if not access_token or not item_id:
                                st.error("表示するにはOneDriveへ接続してください。")
                            else:
                                if attachment.get("file_type") == "image":
                                    open_onedrive_image_group_gallery(
                                        access_token,
                                        group_items,
                                    )
                                else:
                                    try:
                                        with st.spinner("ファイルを読み込んでいます…"):
                                            content = download_onedrive_file(access_token, item_id)
                                        show_onedrive_pdf_dialog(
                                            content,
                                            filename,
                                            attachment.get("mime_type") or "application/pdf",
                                            metadata_id,
                                        )
                                    except Exception as exc:
                                        st.error(f"表示できませんでした：{exc}")

                        attachment_date = format_attachment_datetime(
                            attachment.get("created_at")
                        ).split(" ", 1)[0]
                        attachment_tags = normalize_attachment_tags(
                            attachment.get("tags") or []
                        )
                        date_and_tags = []
                        if attachment_date:
                            date_and_tags.append(attachment_date)
                        if attachment_tags:
                            date_and_tags.append(
                                " ".join(f"`#{tag}`" for tag in attachment_tags)
                            )
                        if date_and_tags:
                            st.caption("　".join(date_and_tags))
                        render_collapsible_attachment_remarks(
                            attachment.get("remarks")
                        )

                        if active_edit_id == group_ui_id:
                            current_tags = normalize_attachment_tags(attachment.get("tags") or [])
                            edit_tag_options = load_tag_history_options_for_action()
                            for current_tag in reversed(current_tags):
                                if current_tag not in edit_tag_options:
                                    edit_tag_options.insert(0, current_tag)
                            edited_history_tags = st.multiselect(
                                "タグを編集（入力すると過去の候補を絞り込み）",
                                edit_tag_options,
                                default=current_tags,
                                key=f"onedrive_attachment_edit_history_{group_ui_id}",
                            )
                            edited_new_tags = st.text_input(
                                "新しいタグを追加",
                                value="",
                                placeholder="候補にないタグだけ入力",
                                key=f"onedrive_attachment_edit_new_{group_ui_id}",
                                autocomplete="off",
                            )
                            edited_remarks = st.text_area(
                                "備考を編集",
                                value=attachment.get("remarks", ""),
                                height=90,
                                key=f"onedrive_attachment_edit_remarks_{group_ui_id}",
                            )
                            save_col, cancel_col = st.columns(2)
                            with save_col:
                                if st.button(
                                    "保存",
                                    key=f"onedrive_attachment_edit_save_{group_ui_id}",
                                    type="primary",
                                    use_container_width=True,
                                ):
                                    try:
                                        old_tags = " ".join(f"#{tag}" for tag in attachment.get("tags", []))
                                        new_tags = list(edited_history_tags) + normalize_attachment_tags(edited_new_tags)
                                        update_onedrive_attachment_group_metadata(
                                            group_items,
                                            new_tags,
                                            edited_remarks,
                                        )
                                        changes = {}
                                        new_tags_text = " ".join(f"#{tag}" for tag in normalize_attachment_tags(new_tags))
                                        if old_tags != new_tags_text:
                                            changes["タグ"] = (old_tags, new_tags_text)
                                        if attachment.get("remarks", "") != str(edited_remarks or "").strip():
                                            changes["備考"] = (attachment.get("remarks", ""), str(edited_remarks or "").strip())
                                        remember_change_history_warning(
                                            record_change_history_safely(
                                                attachment_entity_label(entity_type),
                                                entity_id or "",
                                                entity_name,
                                                "変更",
                                                changes,
                                                section=(
                                                    f"写真・資料：画像{group_count}枚"
                                                    if group_count > 1
                                                    else f"写真・資料：{filename}"
                                                ),
                                            )
                                        )
                                        st.session_state.pop(edit_key, None)
                                        st.session_state[success_key] = "タグ・備考を更新しました。"
                                        st.rerun()
                                    except Exception as exc:
                                        st.error(f"更新できませんでした：{exc}")
                            with cancel_col:
                                if st.button(
                                    "キャンセル",
                                    key=f"onedrive_attachment_edit_cancel_{group_ui_id}",
                                    use_container_width=True,
                                ):
                                    st.session_state.pop(edit_key, None)
                                    st.rerun()
                            continue

                        action_col, delete_col = st.columns(2, gap="small")
                        with action_col:
                            if st.button(
                                "編集",
                                key=f"onedrive_attachment_edit_button_{group_ui_id}",
                                use_container_width=True,
                                help="タグ・備考を編集",
                            ):
                                st.session_state[edit_key] = group_ui_id
                                st.rerun()
                        with delete_col:
                            if st.button(
                                "削除",
                                key=f"onedrive_attachment_delete_button_{group_ui_id}",
                                use_container_width=True,
                            ):
                                st.session_state.pop(edit_key, None)
                                confirm_onedrive_attachment_delete_dialog(
                                    access_token,
                                    group_items,
                                    entity_type,
                                    entity_id,
                                    entity_name,
                                    success_key,
                                    open_key,
                                    restore_key,
                                    upload_error_key,
                                )

            if len(display_groups) > limit:
                if st.button(
                    "さらに表示",
                    key=f"onedrive_attachment_more_{suffix}",
                    use_container_width=True,
                ):
                    st.session_state[limit_key] = limit + ONEDRIVE_PAGE_SIZE
                    st.rerun()

def _open_customer_attachments_lazy_section(load_key, force_open_key):
    """顧客詳細の写真・PDFを、利用者が開いた時だけ読み込む。"""
    st.session_state[load_key] = True
    st.session_state[force_open_key] = True

def render_customer_attachments_lazy_section(customer_name, customer_key=None):
    """初期表示ではOneDrive/Supabaseへ接続せず、ボタン操作後だけ既存機能を表示する。"""
    entity_type = "customer"
    entity_name = clean_value(customer_name, blank_text="").strip()
    entity_id = clean_value(customer_key, blank_text="").strip()
    identity = f"{entity_type}:{entity_id or entity_name}"
    suffix = hashlib.sha256(str(identity).encode("utf-8")).hexdigest()[:16]
    load_key = f"customer_attachments_lazy_loaded_{suffix}"
    force_open_key = f"onedrive_attachment_force_open_{suffix}"

    if not st.session_state.get(load_key, False):
        st.button(
            "📎 写真・PDFを開く",
            key=f"customer_attachments_lazy_open_{suffix}",
            use_container_width=True,
            on_click=_open_customer_attachments_lazy_section,
            args=(load_key, force_open_key),
        )
        return

    render_customer_attachments_section(
        customer_name,
        customer_key,
        entity_type=entity_type,
    )






def show_attachment_search_page():
    """顧客・仕入先・運送会社・ホテルの写真・資料を横断検索する。"""
    st.header("🔎 写真・資料検索")
    st.caption("顧客・仕入先・運送会社・ホテルの写真・PDFを検索します。タグ欄は文字を入力すると過去の候補が絞り込まれます。")

    clicked_tag = clean_value(
        get_query_value("attachment_tag", ""),
        blank_text="",
    ).strip()
    clicked_tags = normalize_attachment_tags([clicked_tag]) if clicked_tag else []
    if clicked_tags:
        st.session_state["attachment_global_tag_filter"] = clicked_tags[:1]
        update_query_params(attachment_tag=None)
        st.rerun()

    if not has_supabase_config():
        st.warning("写真・資料検索にはSupabase設定が必要です。")
        return

    try:
        attachments = load_all_onedrive_attachments_from_supabase()
    except Exception as exc:
        st.error(f"写真・資料を読み込めませんでした：{exc}")
        return

    if not attachments:
        st.info("保存されている写真・資料はありません。")
        return

    query = st.text_input(
        "取引先名・タグ・備考を検索",
        placeholder="文字を入力",
        key="attachment_global_text_filter",
        autocomplete="off",
    )
    tag_options = attachment_tag_history_options(attachments)
    selected_tags = st.multiselect(
        "タグ（入力すると履歴から候補を表示）",
        tag_options,
        key="attachment_global_tag_filter",
    ) if tag_options else []
    if tag_options:
        st.caption(
            "最近使ったタグ："
            + "　".join(f"#{tag}" for tag in tag_options[:10])
        )
    entity_filter = st.selectbox(
        "取引先種別",
        ["すべて", "顧客", "仕入先", "運送会社", "ホテル"],
        key="attachment_global_entity_filter",
    )
    type_filter = st.selectbox(
        "種類",
        ["すべて", "写真", "PDF"],
        key="attachment_global_type_filter",
    )

    normalized_query = clean_value(query, blank_text="").strip().casefold()
    query_terms = [term for term in re.split(r"[\s　]+", normalized_query) if term]
    filtered = []
    for attachment in attachments:
        attachment_entity_type = normalize_attachment_entity_type(
            attachment.get("entity_type")
        )
        if (
            entity_filter != "すべて"
            and attachment_entity_label(attachment_entity_type) != entity_filter
        ):
            continue
        if type_filter == "写真" and attachment.get("file_type") != "image":
            continue
        if type_filter == "PDF" and attachment.get("file_type") != "pdf":
            continue
        attachment_tags = normalize_attachment_tags(attachment.get("tags") or [])
        if selected_tags and not set(selected_tags).issubset(set(attachment_tags)):
            continue
        searchable = " ".join(
            [
                clean_value(attachment.get("entity_name"), blank_text=""),
                " ".join(attachment_tags),
                clean_value(attachment.get("remarks"), blank_text=""),
                clean_value(attachment.get("original_name"), blank_text=""),
            ]
        ).casefold()
        if query_terms and not all(term in searchable for term in query_terms):
            continue
        filtered.append(attachment)

    signature = json.dumps(
        {
            "query": normalized_query,
            "tags": list(selected_tags),
            "entity": entity_filter,
            "type": type_filter,
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    signature_key = "attachment_global_filter_signature"
    limit_key = "attachment_global_result_limit"
    if st.session_state.get(signature_key) != signature:
        st.session_state[signature_key] = signature
        st.session_state[limit_key] = ONEDRIVE_PAGE_SIZE

    display_groups = group_onedrive_attachments_for_display(filtered)
    render_attachment_card_layout_styles()
    st.caption(f"該当：{len(display_groups)}件")
    if not display_groups:
        st.info("条件に一致する写真・資料はありません。")
        return

    access_token = get_onedrive_access_token()
    if not access_token:
        st.warning("画像やPDFを開くには、管理者によるOneDrive接続が必要です。")

    limit = int(st.session_state.get(limit_key, ONEDRIVE_PAGE_SIZE))
    visible_groups = display_groups[:limit]
    grid_column_count = 2 if is_mobile_browser() else 4
    grid_columns = []

    for group_index, attachment_group in enumerate(visible_groups):
        if group_index % grid_column_count == 0:
            grid_columns = st.columns(grid_column_count, gap="small")

        group_items = attachment_group["items"]
        attachment = attachment_group["representative"]
        group_count = attachment_group["count"]
        group_ui_id = attachment_group["ui_id"]
        item_id = attachment.get("file_id", "")
        metadata_id = attachment.get("id", "")
        filename = attachment.get("original_name", "名称未設定")
        entity_type = normalize_attachment_entity_type(attachment.get("entity_type"))
        entity_id = clean_value(attachment.get("entity_id"), blank_text="")
        entity_name = clean_value(attachment.get("entity_name"), blank_text="") or "名称未設定"
        entity_label = attachment_entity_label(entity_type)

        with grid_columns[group_index % grid_column_count]:
            with st.container(border=True):
                st.markdown(
                    '<span class="aoyama-attachment-card-marker" aria-hidden="true"></span>',
                    unsafe_allow_html=True,
                )
                preview_digest = hashlib.sha256(
                    f"global:{group_ui_id}".encode("utf-8")
                ).digest()[:8]
                preview_bits = "".join(f"{byte:08b}" for byte in preview_digest)
                preview_trigger_label = "⁣" + "".join(
                    "​" if bit == "0" else "‌" for bit in preview_bits
                )
                preview_clicked = st.button(
                    preview_trigger_label,
                    key=f"onedrive_attachment_global_preview_button_{group_ui_id}",
                )

                if attachment.get("file_type") == "image":
                    thumbnail_content = None
                    thumbnail_filename = filename
                    if access_token:
                        for thumbnail_item in group_items:
                            thumbnail_item_id = thumbnail_item.get("file_id", "")
                            if not thumbnail_item_id:
                                continue
                            thumb_key = f"onedrive_thumbnail_{thumbnail_item_id}"
                            if not isinstance(st.session_state.get(thumb_key), bytes):
                                try:
                                    thumbnail = download_onedrive_attachment_thumbnail(
                                        access_token,
                                        thumbnail_item,
                                    )
                                    if thumbnail:
                                        st.session_state[thumb_key] = thumbnail
                                except Exception:
                                    pass
                            if isinstance(st.session_state.get(thumb_key), bytes):
                                thumbnail_content = st.session_state[thumb_key]
                                thumbnail_filename = thumbnail_item.get(
                                    "original_name",
                                    filename,
                                )
                                break
                    render_clickable_onedrive_thumbnail(
                        thumbnail_content,
                        thumbnail_filename,
                        preview_trigger_label,
                        compact_height=132 if is_mobile_browser() else 145,
                        badge_text=f"{group_count}枚" if group_count > 1 else "",
                    )
                else:
                    render_clickable_onedrive_pdf_tile(
                        filename,
                        preview_trigger_label,
                        compact_height=132 if is_mobile_browser() else 145,
                    )

                if preview_clicked:
                    if not access_token or not item_id:
                        st.error("表示するにはOneDriveへ接続してください。")
                    else:
                        if attachment.get("file_type") == "image":
                            open_onedrive_image_group_gallery(
                                access_token,
                                group_items,
                            )
                        else:
                            try:
                                with st.spinner("ファイルを読み込んでいます…"):
                                    content = download_onedrive_file(access_token, item_id)
                                show_onedrive_pdf_dialog(
                                    content,
                                    filename,
                                    attachment.get("mime_type") or "application/pdf",
                                    f"global_{metadata_id}",
                                )
                            except Exception as exc:
                                st.error(f"表示できませんでした：{exc}")

                if entity_type == "customer":
                    entity_url = make_app_url(page="detail", customer=entity_name)
                elif entity_type == "hotel":
                    entity_url = make_app_url(
                        page="hotel_information",
                        hotel_search=entity_name,
                    )
                else:
                    entity_url = make_app_url(
                        page="partner_detail",
                        partner_id=entity_id,
                        partner_type=entity_type,
                    )
                safe_entity_url = html.escape(entity_url, quote=True)
                safe_entity_title = html.escape(entity_name)
                st.markdown(
                    f'<div class="aoyama-attachment-card-entity">'
                    f'<a href="{safe_entity_url}" target="_self" title="{safe_entity_title}">'
                    f'{safe_entity_title}</a></div>',
                    unsafe_allow_html=True,
                )
                attachment_date = format_attachment_datetime(
                    attachment.get("created_at")
                ).split(" ", 1)[0]
                st.markdown(
                    '<div class="aoyama-attachment-card-date">'
                    + (html.escape(attachment_date) if attachment_date else '&nbsp;')
                    + '</div>',
                    unsafe_allow_html=True,
                )
                attachment_tags = normalize_attachment_tags(attachment.get("tags") or [])
                tags_html = "".join(
                    f'<a class="aoyama-attachment-card-tag" '
                    f'href="{html.escape("?" + urllib.parse.urlencode({"page": "attachment_search", "attachment_tag": tag}), quote=True)}" '
                    f'target="_self" title="#{html.escape(tag, quote=True)}で絞り込む">'
                    f'#{html.escape(tag)}</a>'
                    for tag in attachment_tags
                )
                st.markdown(
                    '<div class="aoyama-attachment-card-tags">'
                    + (tags_html if tags_html else '&nbsp;')
                    + '</div>',
                    unsafe_allow_html=True,
                )
                render_collapsible_attachment_remarks(
                    attachment.get("remarks")
                )

    if len(display_groups) > limit:
        if st.button(
            "さらに表示",
            key="attachment_global_more",
            use_container_width=True,
        ):
            st.session_state[limit_key] = limit + ONEDRIVE_PAGE_SIZE
            st.rerun()


def reorder_customer_information(first_item, second_item):
    """隣接2行を1回のupsertで入れ替え、並び順をまとめて保存する。"""
    payload = []
    for item, new_order in (
        (first_item, second_item.get("sort_order", 0)),
        (second_item, first_item.get("sort_order", 0)),
    ):
        payload.append(
            {
                "id": item["id"],
                "customer_key": item.get("customer_key"),
                "customer_name": item["customer_name"],
                "field_name": item["field_name"],
                "content": item.get("content", ""),
                "sort_order": int(new_order),
                "created_at": item["created_at"],
            }
        )
    try:
        response = requests.post(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(
                prefer="resolution=merge-duplicates,return=minimal"
            ),
            params={"on_conflict": "id"},
            json=payload,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("顧客情報の並び替え中にSupabaseへ接続できませんでした。") from exc
    check_customer_information_response("並び替え", response, (200, 201))
    clear_customer_information_cache()



def make_past_product_note_field(product_name):
    """顧客情報テーブル内で、過去商品メモを通常項目と分けるための内部項目名を作る。"""
    product = clean_value(product_name, blank_text="").strip()
    return f"{PAST_PRODUCT_NOTE_PREFIX}{product}"


def extract_past_product_name(field_name):
    """内部項目名から商品名だけを取り出す。"""
    field = clean_value(field_name, blank_text="")
    if not field.startswith(PAST_PRODUCT_NOTE_PREFIX):
        return ""
    return field[len(PAST_PRODUCT_NOTE_PREFIX):].strip()


def is_past_product_note_item(item):
    """顧客情報テーブル上の商品メモ専用レコードか判定する。"""
    return clean_value(item.get("field_name"), blank_text="").startswith(PAST_PRODUCT_NOTE_PREFIX)


def get_past_product_names(detail, visible_detail):
    """使用数量/日が0または空白の商品を、過去に使用した商品として抽出する。"""
    active_products = {
        clean_value(value, blank_text="").strip()
        for value in visible_detail.get("商品名", []).tolist()
        if clean_value(value, blank_text="").strip()
    }

    past_products = []
    for _, row in detail.iterrows():
        product_name = clean_value(row.get("商品名"), blank_text="").strip()
        if not product_name:
            continue
        if product_name in active_products:
            continue
        if not is_blank_or_zero(row.get("使用数量/日")):
            continue
        if product_name not in past_products:
            past_products.append(product_name)

    return past_products


def get_past_product_note_items(customer_name, customer_key):
    """過去商品メモを商品名ごとの辞書で返す。"""
    items = load_customer_information(customer_name, customer_key)
    result = {}
    for item in items:
        if not is_past_product_note_item(item):
            continue
        product_name = extract_past_product_name(item.get("field_name"))
        if product_name and product_name not in result:
            result[product_name] = item
    return result


@st.cache_data(ttl=30, show_spinner=False)
def load_all_past_product_notes_from_supabase():
    """取引先メモ画面で使う過去商品メモを、全顧客分読み込む。"""
    if not has_supabase_config():
        return []

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name,content,sort_order,created_at,updated_at",
            "field_name": f"like.{PAST_PRODUCT_NOTE_PREFIX}*",
            "order": "updated_at.desc,created_at.desc,id.desc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("過去商品メモの読み込み中にSupabaseへ接続できませんでした。") from exc

        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った過去商品メモの形式が正しくありません。")

        rows.extend(item for item in page if is_past_product_note_item(item))
        if len(page) < page_size:
            break
        offset += page_size

    return rows


def save_past_product_note(customer_name, customer_key, product_name, content):
    """過去商品の商品別メモを、既存の顧客情報テーブルへ内部項目として保存する。"""
    field_name = make_past_product_note_field(product_name)
    items = load_customer_information(customer_name, customer_key)
    existing = next(
        (
            item for item in items
            if clean_value(item.get("field_name"), blank_text="") == field_name
        ),
        None,
    )

    if existing:
        update_customer_information(existing["id"], field_name, content)
    else:
        next_order = max(
            (int(item.get("sort_order", 0)) for item in items),
            default=0,
        ) + 10
        insert_customer_information(
            customer_name,
            customer_key,
            field_name,
            content,
            next_order,
        )

    load_all_past_product_notes_from_supabase.clear()


def delete_past_product_note(note_item):
    """過去商品の商品別メモを削除する。"""
    item_id = clean_value(note_item.get("id"), blank_text="")
    if not item_id:
        raise RuntimeError("削除する商品メモが見つかりません。")
    delete_customer_information(item_id)
    load_all_past_product_notes_from_supabase.clear()


@st.dialog("商品メモを削除")
def confirm_past_product_note_delete_dialog(
    note_item, product_name, delete_success_key, keep_open_key
):
    """画面位置を保ったまま、過去商品メモの削除を確認する。"""
    item_id = clean_value(note_item.get("id"), blank_text="")
    st.warning(f"「{product_name}」の商品メモを削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"past_product_note_delete_dialog_yes_{item_id}",
            type="primary",
            use_container_width=True,
        ):
            try:
                delete_past_product_note(note_item)
                st.session_state[delete_success_key] = True
                st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
                st.rerun()
            except Exception as exc:
                st.error(f"商品メモを削除できませんでした：{exc}")

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"past_product_note_delete_dialog_no_{item_id}",
            use_container_width=True,
        ):
            st.session_state[keep_open_key] = True
            st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
            st.rerun()


def render_past_products_section(customer_name, customer_key, detail, visible_detail):
    """顧客詳細の最下部に、過去に使用した商品と商品別メモを表示する。"""
    past_products = get_past_product_names(detail, visible_detail)
    if not past_products:
        return

    st.markdown("---")
    st.subheader("過去に使用した商品")

    if not has_supabase_config():
        st.warning("商品メモを使うにはSupabase設定が必要です。")
        return

    try:
        note_items = get_past_product_note_items(customer_name, customer_key)
    except Exception as exc:
        st.warning(f"商品メモを読み込めませんでした：{exc}")
        note_items = {}

    identity = customer_key or customer_name
    for product_name in past_products:
        note_item = note_items.get(product_name)
        current_content = clean_value(
            note_item.get("content") if note_item else "",
            blank_text="",
        )
        state_suffix = hashlib.sha256(
            f"past-product|{identity}|{product_name}".encode("utf-8")
        ).hexdigest()[:16]
        save_success_key = f"past_product_note_save_success_{state_suffix}"
        delete_success_key = f"past_product_note_delete_success_{state_suffix}"
        keep_open_key = f"past_product_note_keep_open_{state_suffix}"

        save_succeeded = bool(st.session_state.pop(save_success_key, False))
        delete_succeeded = bool(st.session_state.pop(delete_success_key, False))
        keep_open = bool(st.session_state.pop(keep_open_key, False))
        with st.expander(
            product_name,
            expanded=save_succeeded or delete_succeeded or keep_open,
        ):
            if save_succeeded:
                st.success("メモを保存しました。")
            if delete_succeeded:
                st.success("商品メモを削除しました。")

            memo = st.text_area(
                "メモ",
                value=current_content,
                key=f"past_product_note_{state_suffix}",
                height=110,
                placeholder="例：値上げのため中止、効果が薄かった、別商品へ変更 など",
            )

            save_col, delete_col = st.columns(2)
            with save_col:
                if st.button(
                    "保存",
                    key=f"past_product_note_save_{state_suffix}",
                    type="primary",
                    use_container_width=True,
                ):
                    try:
                        save_past_product_note(
                            customer_name,
                            customer_key,
                            product_name,
                            memo,
                        )
                        st.session_state[save_success_key] = True
                        st.rerun()
                    except Exception as exc:
                        st.error(f"商品メモを保存できませんでした：{exc}")

            with delete_col:
                if note_item and st.button(
                    "削除",
                    key=f"past_product_note_delete_{state_suffix}",
                    use_container_width=True,
                ):
                    confirm_past_product_note_delete_dialog(
                        note_item,
                        product_name,
                        delete_success_key,
                        keep_open_key,
                    )



def make_estimate_field_name():
    """顧客情報テーブル内で、見積りを通常項目と分ける内部項目名を作る。"""
    return f"{ESTIMATE_PREFIX}{uuid.uuid4()}"


def is_estimate_item(item):
    """顧客情報テーブル上の提案・見積り専用レコードか判定する。"""
    return clean_value(item.get("field_name"), blank_text="").startswith(ESTIMATE_PREFIX)


def estimate_date_text(value):
    """見積りの日付をYYYY-MM-DDへそろえる。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_value(value, blank_text="").strip()
    match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if not match:
        return text
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return text


def estimate_date_input_value(value):
    """保存済みの日付をst.date_inputで使えるdateへ変換する。"""
    text = estimate_date_text(value)
    try:
        return date.fromisoformat(text)
    except (TypeError, ValueError):
        return get_jst_now().date()


def format_estimate_date(value):
    text = estimate_date_text(value)
    try:
        return date.fromisoformat(text).strftime("%Y/%m/%d")
    except (TypeError, ValueError):
        return text or "未入力"


def serialize_estimate(proposal_date, product_name, manufacturer, unit_price, remarks):
    payload = {
        "version": ESTIMATE_VERSION,
        "proposal_date": estimate_date_text(proposal_date),
        "product_name": clean_value(product_name, blank_text="").strip(),
        "manufacturer": clean_value(manufacturer, blank_text="").strip(),
        "unit_price": clean_value(unit_price, blank_text="").strip(),
        "remarks": clean_value(remarks, blank_text="").strip(),
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def parse_estimate_item(item):
    """Supabase内部レコードを画面表示用の見積り辞書へ変換する。"""
    if not is_estimate_item(item):
        return None
    try:
        payload = json.loads(str(item.get("content") or "{}"))
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    return {
        "id": clean_value(item.get("id"), blank_text=""),
        "field_name": clean_value(item.get("field_name"), blank_text=""),
        "customer_key": clean_value(item.get("customer_key"), blank_text=""),
        "customer_name": clean_value(item.get("customer_name"), blank_text=""),
        "proposal_date": estimate_date_text(payload.get("proposal_date")),
        "product_name": clean_value(payload.get("product_name"), blank_text="").strip(),
        "manufacturer": clean_value(payload.get("manufacturer"), blank_text="").strip(),
        "unit_price": clean_value(payload.get("unit_price"), blank_text="").strip(),
        "remarks": clean_value(payload.get("remarks"), blank_text="").strip(),
        "created_at": item.get("created_at", ""),
        "updated_at": item.get("updated_at", ""),
        "sort_order": item.get("sort_order", 0),
    }


def estimate_sort_key(item):
    return (
        estimate_date_text(item.get("proposal_date")),
        str(item.get("updated_at") or item.get("created_at") or ""),
        str(item.get("id") or ""),
    )


@st.cache_data(ttl=300, show_spinner=False)
def load_estimate_presence_index():
    """提案・見積りが存在する顧客だけを、小さい索引としてSupabaseから読む。"""
    if not has_supabase_config():
        return {"customer_keys": tuple(), "customer_names": tuple()}

    key_values = set()
    name_values = set()
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name",
            "field_name": f"like.{ESTIMATE_PREFIX}*",
            "order": "id.asc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("見積りの索引読み込み中にSupabaseへ接続できませんでした。") from exc
        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った見積り索引の形式が正しくありません。")

        for item in page:
            if not isinstance(item, dict) or not is_estimate_item(item):
                continue
            item_key = clean_value(item.get("customer_key"), blank_text="").strip()
            item_name = clean_value(item.get("customer_name"), blank_text="").strip()
            if item_key:
                key_values.add(item_key)
            elif item_name:
                name_values.add(item_name)

        if len(page) < page_size:
            break
        offset += page_size

    return {
        "customer_keys": tuple(sorted(key_values)),
        "customer_names": tuple(sorted(name_values)),
    }

def customer_has_estimates(customer_name, customer_key=None):
    """存在索引だけで、その顧客に提案・見積りがあるか確認する。"""
    index = load_estimate_presence_index()
    if customer_key:
        target = clean_value(customer_key, blank_text="").strip()
        return target in set(index.get("customer_keys", ()))
    target = clean_value(customer_name, blank_text="").strip()
    return target in set(index.get("customer_names", ()))

def get_customer_estimates(customer_name, customer_key):
    try:
        has_estimates = customer_has_estimates(customer_name, customer_key)
    except Exception:
        # 索引確認に失敗した場合は、表示欠落を避けるため従来どおり詳細取得へ戻す。
        has_estimates = None

    if has_estimates is False:
        return []

    items = load_customer_information(customer_name, customer_key)
    estimates = []
    for item in items:
        parsed = parse_estimate_item(item)
        if parsed:
            estimates.append(parsed)
    estimates.sort(key=estimate_sort_key, reverse=True)
    return estimates


@st.cache_data(ttl=30, show_spinner=False)
def load_all_estimates_from_supabase():
    """ホームの見積り画面で使う全顧客分の見積りを読み込む。"""
    if not has_supabase_config():
        return []

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name,content,sort_order,created_at,updated_at",
            "field_name": f"like.{ESTIMATE_PREFIX}*",
            "order": "created_at.desc,id.desc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("見積りの読み込み中にSupabaseへ接続できませんでした。") from exc

        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った見積りの形式が正しくありません。")

        rows.extend(item for item in page if is_estimate_item(item))
        if len(page) < page_size:
            break
        offset += page_size

    return rows


def clear_estimate_cache():
    try:
        load_all_estimates_from_supabase.clear()
    except Exception:
        pass
    try:
        load_estimate_presence_index.clear()
    except Exception:
        pass


def estimate_values(item):
    return {
        "提案日": estimate_date_text(item.get("proposal_date")),
        "商品名": clean_value(item.get("product_name"), blank_text=""),
        "メーカー": clean_value(item.get("manufacturer"), blank_text=""),
        "単価": clean_value(item.get("unit_price"), blank_text=""),
        "備考": clean_value(item.get("remarks"), blank_text=""),
    }


def estimate_history_changes(before, after):
    before_values = estimate_values(before or {})
    after_values = estimate_values(after or {})
    return {
        field_name: (before_values.get(field_name, ""), after_values.get(field_name, ""))
        for field_name in after_values
        if before_values.get(field_name, "") != after_values.get(field_name, "")
    }


def save_customer_estimate(
    customer_name,
    customer_key,
    proposal_date,
    product_name,
    manufacturer,
    unit_price,
    remarks,
    existing=None,
):
    content = serialize_estimate(
        proposal_date,
        product_name,
        manufacturer,
        unit_price,
        remarks,
    )
    if existing:
        update_customer_information(existing["id"], existing["field_name"], content)
    else:
        items = load_customer_information(customer_name, customer_key)
        next_order = max(
            (int(item.get("sort_order", 0)) for item in items),
            default=0,
        ) + 10
        insert_customer_information(
            customer_name,
            customer_key,
            make_estimate_field_name(),
            content,
            next_order,
        )
    clear_estimate_cache()


def delete_customer_estimate(item):
    item_id = clean_value(item.get("id"), blank_text="")
    if not item_id:
        raise RuntimeError("削除する見積りが見つかりません。")
    delete_customer_information(item_id)
    clear_estimate_cache()


@st.dialog("見積りを削除")
def confirm_customer_estimate_delete_dialog(
    estimate,
    customer_name,
    customer_key,
    success_key,
    keep_open_key,
    add_key,
    edit_key,
    delete_key,
):
    """画面位置を保ったまま、見積り削除を確認して実行する。"""
    estimate_id = clean_value(estimate.get("id"), blank_text="")
    product_name = clean_value(estimate.get("product_name"), blank_text="商品名未入力")
    st.warning(f"「{product_name}」の見積りを削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"estimate_delete_dialog_yes_{estimate_id}",
            type="primary",
            use_container_width=True,
        ):
            try:
                delete_customer_estimate(estimate)
                remember_change_history_warning(
                    record_change_history_safely(
                        "顧客",
                        customer_key or "",
                        customer_name,
                        "削除",
                        estimate_history_changes(estimate, {}),
                        section="提案・見積り",
                    )
                )
                st.session_state.pop(add_key, None)
                st.session_state.pop(edit_key, None)
                st.session_state.pop(delete_key, None)
                st.session_state[success_key] = "見積りを削除しました。"
                st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
                st.rerun()
            except Exception as exc:
                st.error(f"見積りを削除できませんでした：{exc}")

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"estimate_delete_dialog_no_{estimate_id}",
            use_container_width=True,
        ):
            st.session_state[keep_open_key] = True
            st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
            st.rerun()


def estimate_price_label(item):
    price = clean_value(item.get("unit_price"), blank_text="").strip()
    return price or "未入力"


def render_customer_estimates_section(customer_name, customer_key=None):
    """顧客詳細に、折りたたみ式の提案・見積りを表示する。"""
    identity = customer_key or customer_name
    state_suffix = hashlib.sha256(
        f"estimate|{identity}".encode("utf-8")
    ).hexdigest()[:16]
    add_key = f"estimate_add_{state_suffix}"
    edit_key = f"estimate_edit_{state_suffix}"
    delete_key = f"estimate_delete_{state_suffix}"
    success_key = f"estimate_success_{state_suffix}"
    keep_open_key = f"estimate_keep_open_{state_suffix}"

    try:
        estimates = get_customer_estimates(customer_name, customer_key)
    except Exception as exc:
        estimates = []
        load_error = str(exc)
    else:
        load_error = ""

    success_message = st.session_state.pop(success_key, None)
    keep_open = bool(st.session_state.pop(keep_open_key, False))
    expanded = bool(
        success_message
        or keep_open
        or st.session_state.get(add_key)
        or st.session_state.get(edit_key)
    )

    with st.expander(f"📄 提案・見積り　{len(estimates)}件", expanded=expanded):
        if not has_supabase_config():
            st.warning("提案・見積りを使うにはSupabase設定が必要です。")
            return
        if load_error:
            st.warning(f"見積りを読み込めませんでした：{load_error}")
            return
        if success_message:
            st.success(success_message)

        if not st.session_state.get(add_key):
            if st.button(
                "＋ 見積りを追加",
                key=f"estimate_add_button_{state_suffix}",
                use_container_width=True,
            ):
                st.session_state[add_key] = True
                st.session_state.pop(edit_key, None)
                st.session_state.pop(delete_key, None)
                st.rerun()
        else:
            st.markdown("**新しい見積り**")
            with st.form(f"estimate_add_form_{state_suffix}"):
                proposal_date = st.date_input(
                    "提案日",
                    value=get_jst_now().date(),
                )
                product_name = st.text_input("商品名", autocomplete="off")
                manufacturer = st.text_input("メーカー", autocomplete="off")
                unit_price = st.text_input("単価", placeholder="例：85、3,500", autocomplete="off")
                remarks = st.text_area("備考", height=110)
                save_col, cancel_col = st.columns(2)
                with save_col:
                    save = st.form_submit_button(
                        "保存", type="primary", use_container_width=True
                    )
                with cancel_col:
                    cancel = st.form_submit_button(
                        "キャンセル", use_container_width=True
                    )

            if cancel:
                st.session_state.pop(add_key, None)
                st.rerun()
            if save:
                if not clean_value(product_name, blank_text="").strip():
                    st.warning("商品名を入力してください。")
                else:
                    after = {
                        "proposal_date": proposal_date,
                        "product_name": product_name,
                        "manufacturer": manufacturer,
                        "unit_price": unit_price,
                        "remarks": remarks,
                    }
                    try:
                        save_customer_estimate(
                            customer_name,
                            customer_key,
                            proposal_date,
                            product_name,
                            manufacturer,
                            unit_price,
                            remarks,
                        )
                        remember_change_history_warning(
                            record_change_history_safely(
                                "顧客",
                                customer_key or "",
                                customer_name,
                                "追加",
                                estimate_history_changes({}, after),
                                section="提案・見積り",
                            )
                        )
                        st.session_state.pop(add_key, None)
                        st.session_state[success_key] = "見積りを保存しました。"
                        st.rerun()
                    except Exception as exc:
                        st.error(f"見積りを保存できませんでした：{exc}")

        if not estimates:
            st.info("提案・見積りはまだありません。")
            return

        active_edit_id = st.session_state.get(edit_key)

        for estimate in estimates:
            estimate_id = estimate["id"]
            with st.container(border=True):
                if active_edit_id == estimate_id:
                    st.markdown("**見積りを編集**")
                    with st.form(f"estimate_edit_form_{estimate_id}"):
                        proposal_date = st.date_input(
                            "提案日",
                            value=estimate_date_input_value(estimate.get("proposal_date")),
                        )
                        product_name = st.text_input(
                            "商品名", value=estimate.get("product_name", ""),
                            autocomplete="off",
                        )
                        manufacturer = st.text_input(
                            "メーカー", value=estimate.get("manufacturer", ""),
                            autocomplete="off",
                        )
                        unit_price = st.text_input(
                            "単価", value=estimate.get("unit_price", ""),
                            autocomplete="off",
                        )
                        remarks = st.text_area(
                            "備考", value=estimate.get("remarks", ""), height=110
                        )
                        save_col, cancel_col = st.columns(2)
                        with save_col:
                            save = st.form_submit_button(
                                "保存", type="primary", use_container_width=True
                            )
                        with cancel_col:
                            cancel = st.form_submit_button(
                                "キャンセル", use_container_width=True
                            )

                    if cancel:
                        st.session_state.pop(edit_key, None)
                        st.rerun()
                    if save:
                        if not clean_value(product_name, blank_text="").strip():
                            st.warning("商品名を入力してください。")
                        else:
                            after = {
                                "proposal_date": proposal_date,
                                "product_name": product_name,
                                "manufacturer": manufacturer,
                                "unit_price": unit_price,
                                "remarks": remarks,
                            }
                            changes = estimate_history_changes(estimate, after)
                            if not changes:
                                st.warning("変更された項目がありません。")
                            else:
                                try:
                                    save_customer_estimate(
                                        customer_name,
                                        customer_key,
                                        proposal_date,
                                        product_name,
                                        manufacturer,
                                        unit_price,
                                        remarks,
                                        existing=estimate,
                                    )
                                    remember_change_history_warning(
                                        record_change_history_safely(
                                            "顧客",
                                            customer_key or "",
                                            customer_name,
                                            "変更",
                                            changes,
                                            section="提案・見積り",
                                        )
                                    )
                                    st.session_state.pop(edit_key, None)
                                    st.session_state[success_key] = "見積りを保存しました。"
                                    st.rerun()
                                except Exception as exc:
                                    st.error(f"見積りを保存できませんでした：{exc}")
                    continue

                st.markdown(
                    f"**{html.escape(estimate.get('product_name') or '商品名未入力')}**"
                )
                st.caption(f"提案日：{format_estimate_date(estimate.get('proposal_date'))}")
                info_col, price_col = st.columns(2)
                with info_col:
                    st.caption("メーカー")
                    st.write(estimate.get("manufacturer") or "未入力")
                with price_col:
                    st.caption("単価")
                    st.write(estimate_price_label(estimate))
                if estimate.get("remarks"):
                    st.caption("備考")
                    render_collapsible_record_remarks(estimate["remarks"])

                edit_col, delete_col = st.columns(2)
                with edit_col:
                    if st.button(
                        "編集",
                        key=f"estimate_edit_button_{estimate_id}",
                        use_container_width=True,
                    ):
                        st.session_state[edit_key] = estimate_id
                        st.session_state.pop(add_key, None)
                        st.session_state.pop(delete_key, None)
                        st.rerun()
                with delete_col:
                    if st.button(
                        "削除",
                        key=f"estimate_delete_button_{estimate_id}",
                        use_container_width=True,
                    ):
                        st.session_state.pop(add_key, None)
                        st.session_state.pop(edit_key, None)
                        st.session_state.pop(delete_key, None)
                        confirm_customer_estimate_delete_dialog(
                            estimate,
                            customer_name,
                            customer_key,
                            success_key,
                            keep_open_key,
                            add_key,
                            edit_key,
                            delete_key,
                        )


def estimate_rows_to_dataframe(rows):
    records = []
    for row in rows or []:
        estimate = parse_estimate_item(row)
        if not estimate:
            continue
        records.append(
            {
                "提案日": estimate.get("proposal_date", ""),
                "顧客ID": estimate.get("customer_key", ""),
                "顧客名": estimate.get("customer_name", ""),
                "商品名": estimate.get("product_name", ""),
                "メーカー": estimate.get("manufacturer", ""),
                "単価": estimate.get("unit_price", ""),
                "備考": estimate.get("remarks", ""),
                "保存ID": estimate.get("id", ""),
                "作成日時": estimate.get("created_at", ""),
                "更新日時": estimate.get("updated_at", ""),
            }
        )
    records.sort(
        key=lambda record: (
            estimate_date_text(record.get("提案日")),
            str(record.get("更新日時") or record.get("作成日時") or ""),
        ),
        reverse=True,
    )
    return backup_dataframe(
        records,
        [
            "提案日", "顧客ID", "顧客名", "商品名", "メーカー",
            "単価", "備考", "保存ID", "作成日時", "更新日時",
        ],
    )


def show_estimates_page():
    st.header("📄 提案・見積り")
    st.caption("全顧客の提案・見積りを、提案日の新しい順に表示します。")

    if not has_supabase_config():
        st.warning("提案・見積りを使うにはSupabase設定が必要です。")
        return

    try:
        rows = load_all_estimates_from_supabase()
    except Exception as exc:
        st.error(str(exc))
        return

    estimates = []
    for row in rows:
        parsed = parse_estimate_item(row)
        if parsed:
            estimates.append(parsed)
    estimates.sort(key=estimate_sort_key, reverse=True)

    if not estimates:
        st.info("提案・見積りはまだありません。")
        return

    st.write(f"見積り：{len(estimates)}件")
    for estimate in estimates:
        with st.container(border=True):
            customer_name = estimate.get("customer_name") or "顧客名未設定"
            customer_link = build_customer_detail_link(
                customer_name,
                class_name="dispatch-month-link",
            )
            st.markdown(customer_link, unsafe_allow_html=True)
            st.markdown(
                f"**{html.escape(estimate.get('product_name') or '商品名未入力')}**"
            )
            st.caption(f"提案日：{format_estimate_date(estimate.get('proposal_date'))}")
            info_col, price_col = st.columns(2)
            with info_col:
                st.caption("メーカー")
                st.write(estimate.get("manufacturer") or "未入力")
            with price_col:
                st.caption("単価")
                st.write(estimate_price_label(estimate))
            if estimate.get("remarks"):
                st.caption("備考")
                render_collapsible_record_remarks(estimate["remarks"])


# =========================
# 運送会社の運賃登録・比較（Supabase保存）
# =========================
def make_carrier_freight_field_name():
    """顧客情報テーブル内で、運送会社の運賃を通常項目と分ける内部項目名を作る。"""
    return f"{CARRIER_FREIGHT_PREFIX}{uuid.uuid4()}"


def carrier_freight_storage_key(carrier_id):
    return f"carrier_freight:{clean_value(carrier_id, blank_text='').strip()}"


def is_carrier_freight_item(item):
    """顧客情報テーブル上の運送会社運賃専用レコードか判定する。"""
    return clean_value(item.get("field_name"), blank_text="").startswith(CARRIER_FREIGHT_PREFIX)


def carrier_freight_date_text(value):
    """運賃の適用日をYYYY-MM-DDへそろえる。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_value(value, blank_text="").strip()
    match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if not match:
        return text
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return text


def carrier_freight_date_input_value(value):
    text = carrier_freight_date_text(value)
    try:
        return date.fromisoformat(text)
    except (TypeError, ValueError):
        return get_jst_now().date()


def format_carrier_freight_date(value):
    text = carrier_freight_date_text(value)
    try:
        return date.fromisoformat(text).strftime("%Y/%m/%d")
    except (TypeError, ValueError):
        return text or "未入力"


def parse_carrier_freight_number(value, label):
    """任意入力の正数をDecimalへ変換する。空欄はNone。"""
    text = unicodedata.normalize("NFKC", clean_value(value, blank_text="")).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("，", "").replace(" ", "")
    if not re.fullmatch(r"\d+(?:\.\d+)?", text):
        raise ValueError(f"{label}は数字で入力してください。")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{label}は数字で入力してください。") from exc
    if number <= 0:
        raise ValueError(f"{label}は0より大きい数字で入力してください。")
    return number


def carrier_freight_decimal_text(value):
    if value is None:
        return ""
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return "0" if text in {"", "-0"} else text


def normalize_carrier_freight_amounts(truck_freight, quantity_kg, kg_rate):
    """確定した計算ルールに従い、運賃・数量・kg単価を整合させる。"""
    truck = parse_carrier_freight_number(truck_freight, "1車運賃")
    quantity = parse_carrier_freight_number(quantity_kg, "数量")
    rate = parse_carrier_freight_number(kg_rate, "kg単価")
    calculation_source = ""

    if quantity is not None and truck is not None:
        # 3項目すべて入力された場合も、1車運賃と数量を正としてkg単価を計算する。
        rate = (truck / quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        calculation_source = "kg単価を自動計算"
    elif quantity is not None and rate is not None:
        truck = (quantity * rate).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        calculation_source = "1車運賃を自動計算"

    if truck is None and rate is None:
        raise ValueError("1車運賃またはkg単価のどちらかを入力してください。")

    return {
        "truck_freight": carrier_freight_decimal_text(truck),
        "quantity_kg": carrier_freight_decimal_text(quantity),
        "kg_rate": carrier_freight_decimal_text(rate),
        "calculation_source": calculation_source,
    }


def carrier_freight_route_key(value):
    """比較時の表記揺れを減らすため、全半角と空白をそろえる。"""
    text = unicodedata.normalize("NFKC", clean_value(value, blank_text="")).strip().lower()
    return re.sub(r"\s+", "", text)


def serialize_carrier_freight(record):
    payload = {
        "version": CARRIER_FREIGHT_VERSION,
        "carrier_id": clean_value(record.get("carrier_id"), blank_text="").strip(),
        "carrier_name": clean_value(record.get("carrier_name"), blank_text="").strip(),
        "effective_date": carrier_freight_date_text(record.get("effective_date")),
        "pickup_location": clean_value(record.get("pickup_location"), blank_text="").strip(),
        "delivery_destination": clean_value(record.get("delivery_destination"), blank_text="").strip(),
        "truck_freight": clean_value(record.get("truck_freight"), blank_text="").strip(),
        "quantity_kg": clean_value(record.get("quantity_kg"), blank_text="").strip(),
        "kg_rate": clean_value(record.get("kg_rate"), blank_text="").strip(),
        "calculation_source": clean_value(record.get("calculation_source"), blank_text="").strip(),
        "remarks": clean_value(record.get("remarks"), blank_text="").strip(),
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def parse_carrier_freight_item(item):
    """Supabase内部レコードを画面表示用の運賃辞書へ変換する。"""
    if not is_carrier_freight_item(item):
        return None
    try:
        payload = json.loads(str(item.get("content") or "{}"))
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    carrier_id = clean_value(payload.get("carrier_id"), blank_text="").strip()
    if not carrier_id:
        storage_key = clean_value(item.get("customer_key"), blank_text="")
        prefix = "carrier_freight:"
        if storage_key.startswith(prefix):
            carrier_id = storage_key[len(prefix):].strip()

    return {
        "id": clean_value(item.get("id"), blank_text=""),
        "field_name": clean_value(item.get("field_name"), blank_text=""),
        "carrier_id": carrier_id,
        "carrier_name": clean_value(
            payload.get("carrier_name") or item.get("customer_name"),
            blank_text="",
        ).strip(),
        "effective_date": carrier_freight_date_text(payload.get("effective_date")),
        "pickup_location": clean_value(payload.get("pickup_location"), blank_text="").strip(),
        "delivery_destination": clean_value(payload.get("delivery_destination"), blank_text="").strip(),
        "truck_freight": clean_value(payload.get("truck_freight"), blank_text="").strip(),
        "quantity_kg": clean_value(payload.get("quantity_kg"), blank_text="").strip(),
        "kg_rate": clean_value(payload.get("kg_rate"), blank_text="").strip(),
        "calculation_source": clean_value(payload.get("calculation_source"), blank_text="").strip(),
        "remarks": clean_value(payload.get("remarks"), blank_text="").strip(),
        "created_at": item.get("created_at", ""),
        "updated_at": item.get("updated_at", ""),
        "sort_order": item.get("sort_order", 0),
    }


def carrier_freight_sort_key(item):
    return (
        carrier_freight_date_text(item.get("effective_date")),
        str(item.get("updated_at") or item.get("created_at") or ""),
        str(item.get("id") or ""),
    )


@st.cache_data(ttl=30, show_spinner=False)
def load_carrier_freight_rows_from_supabase(carrier_id=""):
    """運送会社運賃をSupabaseからページ単位で読み込む。"""
    if not has_supabase_config():
        return []

    rows = []
    page_size = 1000
    offset = 0
    while True:
        params = {
            "select": "id,customer_key,customer_name,field_name,content,sort_order,created_at,updated_at",
            "field_name": f"like.{CARRIER_FREIGHT_PREFIX}*",
            "order": "created_at.desc,id.desc",
            "limit": str(page_size),
            "offset": str(offset),
        }
        if carrier_id:
            params["customer_key"] = f"eq.{carrier_freight_storage_key(carrier_id)}"
        try:
            response = requests.get(
                get_supabase_customer_information_url(),
                headers=get_supabase_headers(),
                params=params,
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError("運賃の読み込み中にSupabaseへ接続できませんでした。") from exc

        check_customer_information_response("読み込み", response, (200,))
        page = response.json()
        if not isinstance(page, list):
            raise RuntimeError("Supabaseから返った運賃の形式が正しくありません。")

        rows.extend(item for item in page if is_carrier_freight_item(item))
        if len(page) < page_size:
            break
        offset += page_size

    return rows


def get_carrier_freights(carrier_id):
    freights = []
    for item in load_carrier_freight_rows_from_supabase(carrier_id):
        parsed = parse_carrier_freight_item(item)
        if parsed:
            freights.append(parsed)
    freights.sort(key=carrier_freight_sort_key, reverse=True)
    return freights


def get_all_carrier_freights():
    freights = []
    for item in load_carrier_freight_rows_from_supabase(""):
        parsed = parse_carrier_freight_item(item)
        if parsed:
            freights.append(parsed)
    freights.sort(key=carrier_freight_sort_key, reverse=True)
    return freights


def clear_carrier_freight_cache():
    try:
        load_carrier_freight_rows_from_supabase.clear()
    except Exception:
        pass


def carrier_freight_values(item):
    return {
        "適用日": carrier_freight_date_text(item.get("effective_date")),
        "引取場所": clean_value(item.get("pickup_location"), blank_text=""),
        "納品先": clean_value(item.get("delivery_destination"), blank_text=""),
        "1車運賃": clean_value(item.get("truck_freight"), blank_text=""),
        "数量kg": clean_value(item.get("quantity_kg"), blank_text=""),
        "kg単価": clean_value(item.get("kg_rate"), blank_text=""),
        "備考": clean_value(item.get("remarks"), blank_text=""),
    }


def carrier_freight_history_changes(before, after):
    before_values = carrier_freight_values(before or {})
    after_values = carrier_freight_values(after or {})
    return {
        field_name: (before_values.get(field_name, ""), after_values.get(field_name, ""))
        for field_name in after_values
        if before_values.get(field_name, "") != after_values.get(field_name, "")
    }


def build_carrier_freight_record(
    carrier_id,
    carrier_name,
    effective_date,
    pickup_location,
    delivery_destination,
    truck_freight,
    quantity_kg,
    kg_rate,
    remarks,
):
    pickup = clean_value(pickup_location, blank_text="").strip()
    destination = clean_value(delivery_destination, blank_text="").strip()
    if not pickup:
        raise ValueError("引取場所を入力してください。")
    if not destination:
        raise ValueError("納品先を入力してください。")

    amounts = normalize_carrier_freight_amounts(truck_freight, quantity_kg, kg_rate)
    return {
        "carrier_id": clean_value(carrier_id, blank_text="").strip(),
        "carrier_name": clean_value(carrier_name, blank_text="").strip(),
        "effective_date": carrier_freight_date_text(effective_date),
        "pickup_location": pickup,
        "delivery_destination": destination,
        "truck_freight": amounts["truck_freight"],
        "quantity_kg": amounts["quantity_kg"],
        "kg_rate": amounts["kg_rate"],
        "calculation_source": amounts["calculation_source"],
        "remarks": clean_value(remarks, blank_text="").strip(),
    }


def save_carrier_freight(record, existing=None):
    content = serialize_carrier_freight(record)
    carrier_id = record["carrier_id"]
    if existing:
        update_customer_information(existing["id"], existing["field_name"], content)
    else:
        rows = load_carrier_freight_rows_from_supabase(carrier_id)
        next_order = max(
            (int(row.get("sort_order", 0)) for row in rows),
            default=0,
        ) + 10
        insert_customer_information(
            record["carrier_name"],
            carrier_freight_storage_key(carrier_id),
            make_carrier_freight_field_name(),
            content,
            next_order,
        )
    clear_carrier_freight_cache()


def delete_carrier_freight(item):
    item_id = clean_value(item.get("id"), blank_text="")
    if not item_id:
        raise RuntimeError("削除する運賃が見つかりません。")
    delete_customer_information(item_id)
    clear_carrier_freight_cache()


@st.dialog("運賃記録を削除")
def confirm_carrier_freight_delete_dialog(
    freight,
    carrier_id,
    company_name,
    success_key,
    keep_open_key,
    edit_key,
    delete_key,
):
    """画面位置を保ったまま、運賃記録の削除を確認して実行する。"""
    freight_id = clean_value(freight.get("id"), blank_text="")
    pickup = clean_value(freight.get("pickup_location"), blank_text="未設定")
    destination = clean_value(freight.get("delivery_destination"), blank_text="未設定")
    st.warning(f"「{pickup} → {destination}」の運賃記録を削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"carrier_freight_delete_dialog_yes_{freight_id}",
            type="primary",
            use_container_width=True,
        ):
            try:
                delete_carrier_freight(freight)
                remember_change_history_warning(
                    record_change_history_safely(
                        "運送会社",
                        carrier_id,
                        company_name,
                        "削除",
                        carrier_freight_history_changes(freight, {}),
                        section="運賃",
                    )
                )
                st.session_state.pop(edit_key, None)
                st.session_state.pop(delete_key, None)
                st.session_state[success_key] = "運賃を削除しました。"
                st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
                st.rerun()
            except Exception as exc:
                st.error(f"運賃を削除できませんでした：{exc}")

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"carrier_freight_delete_dialog_no_{freight_id}",
            use_container_width=True,
        ):
            st.session_state[keep_open_key] = True
            st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
            st.rerun()


def carrier_freight_decimal_display(value, maximum_decimals=2):
    text = clean_value(value, blank_text="").strip()
    if not text:
        return ""
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text
    if maximum_decimals == 0:
        return f"{number:,.0f}"
    formatted = f"{number:,.{maximum_decimals}f}"
    return formatted.rstrip("0").rstrip(".")


def carrier_freight_truck_label(item):
    value = carrier_freight_decimal_display(item.get("truck_freight"), 0)
    return f"{value}円" if value else "未入力"


def carrier_freight_quantity_label(item):
    value = carrier_freight_decimal_display(item.get("quantity_kg"), 2)
    return f"{value}kg" if value else "未入力"


def carrier_freight_rate_label(item):
    value = carrier_freight_decimal_display(item.get("kg_rate"), 4)
    return f"{value}円/kg" if value else "未入力"


def carrier_freight_success_message(action, record):
    message = f"運賃を{action}しました。"
    calculation = clean_value(record.get("calculation_source"), blank_text="")
    if calculation:
        message += f" {calculation}しました。"
    return message


def render_carrier_freight_display(item):
    route = (
        f"{clean_value(item.get('pickup_location'), blank_text='未入力')}"
        f" → {clean_value(item.get('delivery_destination'), blank_text='未入力')}"
    )
    st.markdown(f"**{html.escape(route)}**")
    st.caption(f"適用日：{format_carrier_freight_date(item.get('effective_date'))}")
    truck_col, quantity_col, rate_col = st.columns(3)
    with truck_col:
        st.caption("1車運賃")
        st.write(carrier_freight_truck_label(item))
    with quantity_col:
        st.caption("数量")
        st.write(carrier_freight_quantity_label(item))
    with rate_col:
        st.caption("kg単価")
        st.write(carrier_freight_rate_label(item))
    if item.get("remarks"):
        st.caption("備考")
        render_collapsible_record_remarks(item["remarks"])


def render_carrier_freight_form(form_key, existing=None):
    existing = existing or {}
    with st.form(form_key):
        effective_date = st.date_input(
            "適用日",
            value=carrier_freight_date_input_value(existing.get("effective_date")),
            key=f"{form_key}_effective_date",
        )
        pickup_location = st.text_input(
            "引取場所",
            value=existing.get("pickup_location", ""),
            placeholder="例：○○工場",
            key=f"{form_key}_pickup_location",
            autocomplete="off",
        )
        delivery_destination = st.text_input(
            "納品先",
            value=existing.get("delivery_destination", ""),
            placeholder="例：△△牧場",
            key=f"{form_key}_delivery_destination",
            autocomplete="off",
        )
        truck_col, quantity_col, rate_col = st.columns(3)
        with truck_col:
            truck_freight = st.text_input(
                "1車運賃（円）",
                value=existing.get("truck_freight", ""),
                placeholder="例：200000",
                key=f"{form_key}_truck_freight",
                autocomplete="off",
            )
        with quantity_col:
            quantity_kg = st.text_input(
                "数量（kg）",
                value=existing.get("quantity_kg", ""),
                placeholder="例：20000",
                key=f"{form_key}_quantity_kg",
                autocomplete="off",
            )
        with rate_col:
            kg_rate = st.text_input(
                "kg単価（円）",
                value=existing.get("kg_rate", ""),
                placeholder="例：10",
                key=f"{form_key}_kg_rate",
                autocomplete="off",
            )
        st.caption(
            "数量と1車運賃がある場合はkg単価を自動計算します。"
            "数量とkg単価があり1車運賃が空欄の場合は、1車運賃を自動計算します。"
        )
        remarks = st.text_area(
            "備考",
            value=existing.get("remarks", ""),
            height=100,
            placeholder="例：高速代込み、冬季料金 など",
            key=f"{form_key}_remarks",
        )
        save_col, cancel_col = st.columns(2)
        with save_col:
            submitted = st.form_submit_button(
                "自動計算して保存",
                type="primary",
                use_container_width=True,
            )
        with cancel_col:
            cancelled = st.form_submit_button("キャンセル", use_container_width=True)
    return submitted, cancelled, {
        "effective_date": effective_date,
        "pickup_location": pickup_location,
        "delivery_destination": delivery_destination,
        "truck_freight": truck_freight,
        "quantity_kg": quantity_kg,
        "kg_rate": kg_rate,
        "remarks": remarks,
    }


def render_carrier_freight_section(carrier_id, company_name):
    """運送会社詳細に、折りたたみ式の運賃登録・履歴を表示する。"""
    state_suffix = hashlib.sha256(
        f"carrier-freight|{carrier_id}".encode("utf-8")
    ).hexdigest()[:16]
    add_key = f"carrier_freight_add_{state_suffix}"
    edit_key = f"carrier_freight_edit_{state_suffix}"
    delete_key = f"carrier_freight_delete_{state_suffix}"
    success_key = f"carrier_freight_success_{state_suffix}"
    keep_open_key = f"carrier_freight_keep_open_{state_suffix}"

    try:
        freights = get_carrier_freights(carrier_id)
    except Exception as exc:
        freights = []
        load_error = str(exc)
    else:
        load_error = ""

    success_message = st.session_state.pop(success_key, None)
    keep_open = bool(st.session_state.pop(keep_open_key, False))
    expanded = bool(
        success_message
        or keep_open
        or st.session_state.get(add_key)
        or st.session_state.get(edit_key)
    )

    st.markdown("---")
    with st.expander(f"💰 運賃登録・履歴　{len(freights)}件", expanded=expanded):
        if success_message:
            st.success(success_message)
        if load_error:
            st.warning(load_error)
            return
        if not has_supabase_config():
            st.warning("運賃登録を使うにはSupabase設定が必要です。")
            return

        st.caption(
            "新しい運賃は既存記録を上書きせず追加し、過去分を履歴として残します。"
            "既存記録の編集・削除は入力ミスの訂正用です。"
        )

        if not st.session_state.get(add_key):
            if st.button(
                "＋ 運賃を追加",
                key=f"carrier_freight_add_button_{state_suffix}",
                use_container_width=True,
            ):
                st.session_state[add_key] = True
                st.session_state.pop(edit_key, None)
                st.session_state.pop(delete_key, None)
                st.rerun()
        else:
            st.markdown("#### 新しい運賃")
            submitted, cancelled, values = render_carrier_freight_form(
                f"carrier_freight_add_form_{state_suffix}"
            )
            if cancelled:
                st.session_state.pop(add_key, None)
                st.rerun()
            if submitted:
                try:
                    record = build_carrier_freight_record(
                        carrier_id,
                        company_name,
                        **values,
                    )
                    save_carrier_freight(record)
                    remember_change_history_warning(
                        record_change_history_safely(
                            "運送会社",
                            carrier_id,
                            company_name,
                            "追加",
                            carrier_freight_history_changes({}, record),
                            section="運賃",
                        )
                    )
                    st.session_state.pop(add_key, None)
                    st.session_state[success_key] = carrier_freight_success_message("追加", record)
                    st.rerun()
                except Exception as exc:
                    st.error(f"運賃を保存できませんでした：{exc}")

        if not freights:
            st.info("登録されている運賃はありません。")
            return

        st.markdown("#### 運賃履歴")
        active_edit_id = st.session_state.get(edit_key)

        for freight in freights:
            freight_id = freight["id"]
            with st.container(border=True):
                if active_edit_id == freight_id:
                    st.markdown("**入力ミスを訂正**")
                    submitted, cancelled, values = render_carrier_freight_form(
                        f"carrier_freight_edit_form_{freight_id}",
                        existing=freight,
                    )
                    if cancelled:
                        st.session_state.pop(edit_key, None)
                        st.rerun()
                    if submitted:
                        try:
                            record = build_carrier_freight_record(
                                carrier_id,
                                company_name,
                                **values,
                            )
                            changes = carrier_freight_history_changes(freight, record)
                            if changes:
                                save_carrier_freight(record, existing=freight)
                                remember_change_history_warning(
                                    record_change_history_safely(
                                        "運送会社",
                                        carrier_id,
                                        company_name,
                                        "変更",
                                        changes,
                                        section="運賃",
                                    )
                                )
                            st.session_state.pop(edit_key, None)
                            st.session_state[success_key] = carrier_freight_success_message("保存", record)
                            st.rerun()
                        except Exception as exc:
                            st.error(f"運賃を保存できませんでした：{exc}")
                    continue

                render_carrier_freight_display(freight)
                edit_col, delete_col = st.columns(2)
                with edit_col:
                    if st.button(
                        "編集",
                        key=f"carrier_freight_edit_button_{freight_id}",
                        use_container_width=True,
                    ):
                        st.session_state[edit_key] = freight_id
                        st.session_state.pop(add_key, None)
                        st.session_state.pop(delete_key, None)
                        st.rerun()
                with delete_col:
                    if st.button(
                        "削除",
                        key=f"carrier_freight_delete_button_{freight_id}",
                        use_container_width=True,
                    ):
                        st.session_state.pop(edit_key, None)
                        st.session_state.pop(delete_key, None)
                        confirm_carrier_freight_delete_dialog(
                            freight,
                            carrier_id,
                            company_name,
                            success_key,
                            keep_open_key,
                            edit_key,
                            delete_key,
                        )


def carrier_freight_numeric_value(item, field_name):
    text = clean_value(item.get(field_name), blank_text="").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def render_carrier_freight_ranking(title, records, field_name, current_names):
    available = [
        record for record in records
        if carrier_freight_numeric_value(record, field_name) is not None
    ]
    available.sort(key=lambda record: carrier_freight_numeric_value(record, field_name))

    st.subheader(title)
    if not available:
        st.info("比較できる運賃はありません。")
        return

    for index, record in enumerate(available):
        with st.container(border=True):
            carrier_id = record.get("carrier_id", "")
            company_name = (
                current_names.get(carrier_id)
                or record.get("carrier_name")
                or "運送会社名未設定"
            )
            company_link = render_page_link(
                company_name,
                page="partner_detail",
                partner_id=carrier_id,
                partner_type="carrier",
                class_name="dispatch-month-link",
            )
            if index == 0:
                st.markdown(f"**最安**　{company_link}", unsafe_allow_html=True)
            else:
                st.markdown(company_link, unsafe_allow_html=True)

            if field_name == "truck_freight":
                st.markdown(f"### {carrier_freight_truck_label(record)}")
            else:
                st.markdown(f"### {carrier_freight_rate_label(record)}")

            st.caption(f"適用日：{format_carrier_freight_date(record.get('effective_date'))}")
            detail_col1, detail_col2 = st.columns(2)
            with detail_col1:
                st.caption("1車運賃")
                st.write(carrier_freight_truck_label(record))
            with detail_col2:
                st.caption("数量・kg単価")
                st.write(
                    f"{carrier_freight_quantity_label(record)} ／ "
                    f"{carrier_freight_rate_label(record)}"
                )
            if record.get("remarks"):
                st.caption("備考")
                render_collapsible_record_remarks(record["remarks"])


def show_carrier_freight_compare():
    """同じ引取場所・納品先の最新運賃を運送会社別に比較する。"""
    show_trade_partner_home_link("carrier")
    st.header("💰 運賃比較")
    st.caption(
        "同じ引取場所と納品先について、各運送会社の適用日が最も新しい記録を比較します。"
    )

    if not has_supabase_config():
        st.warning("運賃比較を使うにはSupabase設定が必要です。")
        return

    try:
        freights = get_all_carrier_freights()
    except Exception as exc:
        st.error(str(exc))
        return

    if not freights:
        st.info("比較できる運賃はまだ登録されていません。")
        return

    # 新しい記録の表記を候補名として優先する。
    pickup_names = {}
    for record in freights:
        pickup_key = carrier_freight_route_key(record.get("pickup_location"))
        if pickup_key:
            pickup_names.setdefault(pickup_key, record.get("pickup_location", ""))

    pickup_keys = sorted(pickup_names, key=lambda key: pickup_names[key])
    selected_pickup_key = st.selectbox(
        "引取場所",
        pickup_keys,
        format_func=lambda key: pickup_names.get(key, key),
        key="carrier_freight_compare_pickup",
    )

    destination_names = {}
    for record in freights:
        if carrier_freight_route_key(record.get("pickup_location")) != selected_pickup_key:
            continue
        destination_key = carrier_freight_route_key(record.get("delivery_destination"))
        if destination_key:
            destination_names.setdefault(
                destination_key,
                record.get("delivery_destination", ""),
            )

    destination_keys = sorted(destination_names, key=lambda key: destination_names[key])
    selected_destination_key = st.selectbox(
        "納品先",
        destination_keys,
        format_func=lambda key: destination_names.get(key, key),
        key="carrier_freight_compare_destination",
    )

    route_records = [
        record for record in freights
        if carrier_freight_route_key(record.get("pickup_location")) == selected_pickup_key
        and carrier_freight_route_key(record.get("delivery_destination")) == selected_destination_key
    ]
    route_records.sort(key=carrier_freight_sort_key, reverse=True)

    latest_by_carrier = {}
    for record in route_records:
        carrier_id = clean_value(record.get("carrier_id"), blank_text="")
        if carrier_id and carrier_id not in latest_by_carrier:
            latest_by_carrier[carrier_id] = record
    latest_records = list(latest_by_carrier.values())

    try:
        partner_data = load_trade_partner_data()
        current_names = {
            trade_partner_text(row.get("取引先ID")): trade_partner_text(row.get("会社名"))
            for row in get_trade_partner_master_rows(partner_data, "carrier")
        }
    except Exception:
        current_names = {}

    st.markdown("---")
    st.markdown(
        f"**{html.escape(pickup_names[selected_pickup_key])}"
        f" → {html.escape(destination_names[selected_destination_key])}**"
    )
    st.caption(f"比較対象：{len(latest_records)}社（各社の最新記録）")

    render_carrier_freight_ranking(
        "1車運賃が安い順",
        latest_records,
        "truck_freight",
        current_names,
    )
    st.markdown("---")
    render_carrier_freight_ranking(
        "kg単価が安い順",
        latest_records,
        "kg_rate",
        current_names,
    )


def carrier_freight_rows_to_dataframe(rows):
    records = []
    for row in rows or []:
        freight = parse_carrier_freight_item(row)
        if not freight:
            continue
        records.append(
            {
                "運送会社ID": freight.get("carrier_id", ""),
                "運送会社": freight.get("carrier_name", ""),
                "適用日": freight.get("effective_date", ""),
                "引取場所": freight.get("pickup_location", ""),
                "納品先": freight.get("delivery_destination", ""),
                "1車運賃": freight.get("truck_freight", ""),
                "数量kg": freight.get("quantity_kg", ""),
                "kg単価": freight.get("kg_rate", ""),
                "計算方法": freight.get("calculation_source", ""),
                "備考": freight.get("remarks", ""),
                "保存ID": freight.get("id", ""),
                "作成日時": freight.get("created_at", ""),
                "更新日時": freight.get("updated_at", ""),
            }
        )
    records.sort(
        key=lambda record: (
            carrier_freight_date_text(record.get("適用日")),
            str(record.get("更新日時") or record.get("作成日時") or ""),
        ),
        reverse=True,
    )
    return backup_dataframe(
        records,
        [
            "運送会社ID", "運送会社", "適用日", "引取場所", "納品先",
            "1車運賃", "数量kg", "kg単価", "計算方法", "備考",
            "保存ID", "作成日時", "更新日時",
        ],
    )


@st.dialog("顧客情報を削除")
def confirm_customer_information_delete_dialog(
    item,
    customer_name,
    customer_key,
    success_key,
    editing_item_key,
    deleting_item_key,
):
    """画面位置を保ったまま、顧客情報項目の削除を確認して実行する。"""
    item_id = clean_value(item.get("id"), blank_text="")
    field_name = clean_value(item.get("field_name"), blank_text="項目名未設定")
    content = clean_value(item.get("content"), blank_text="")
    st.warning(f"「{field_name}」を削除します。")
    st.caption("この操作は元に戻せません。")
    delete_col, cancel_col = st.columns(2)

    with delete_col:
        if st.button(
            "削除する",
            key=f"customer_information_delete_dialog_yes_{item_id}",
            type="primary",
            use_container_width=True,
        ):
            try:
                delete_customer_information(item_id)
                remember_change_history_warning(
                    record_change_history_safely(
                        "顧客",
                        customer_key or "",
                        customer_name,
                        "削除",
                        {field_name: (content, "")},
                        section="顧客情報",
                    )
                )
                st.session_state.pop(editing_item_key, None)
                st.session_state.pop(deleting_item_key, None)
                st.session_state[success_key] = "項目を削除しました。"
                st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
                st.rerun()
            except RuntimeError as exc:
                st.error(str(exc))

    with cancel_col:
        if st.button(
            "キャンセル",
            key=f"customer_information_delete_dialog_no_{item_id}",
            use_container_width=True,
        ):
            st.session_state[GLOBAL_DELETE_SCROLL_RESTORE_KEY] = True
            st.rerun()


def render_customer_information_form(customer_name, customer_key, items, state_suffix):
    add_key = f"customer_information_add_{state_suffix}"
    if not st.session_state.get(add_key):
        if st.button("＋ 項目を追加", key=f"customer_information_add_button_{state_suffix}"):
            st.session_state[add_key] = True
            st.rerun()
        return

    st.markdown("**新しい項目**")
    with st.form(f"customer_information_add_form_{state_suffix}"):
        field_name = st.text_input("項目名", placeholder="例：担当者", autocomplete="off")
        content = st.text_area(
            "内容",
            placeholder="内容を入力（複数行可）",
            height=120,
        )
        save_col, cancel_col = st.columns(2)
        with save_col:
            save = st.form_submit_button(
                "保存", type="primary", use_container_width=True
            )
        with cancel_col:
            cancel = st.form_submit_button("キャンセル", use_container_width=True)

    if cancel:
        st.session_state.pop(add_key, None)
        st.rerun()
    if save:
        if not str(field_name).strip():
            st.warning("項目名を入力してください。")
            return
        next_order = max(
            (int(item.get("sort_order", 0)) for item in items),
            default=0,
        ) + 10
        try:
            insert_customer_information(
                customer_name,
                customer_key,
                field_name,
                content,
                next_order,
            )
            remember_change_history_warning(
                record_change_history_safely(
                    "顧客",
                    customer_key or "",
                    customer_name,
                    "追加",
                    {str(field_name).strip(): ("", content)},
                    section="顧客情報",
                )
            )
            st.session_state.pop(add_key, None)
            st.session_state[f"customer_information_success_{state_suffix}"] = "項目を追加しました。"
            st.rerun()
        except RuntimeError as exc:
            st.error(str(exc))


def render_customer_information_card(customer_name, customer_key=None):
    identity = customer_key or customer_name
    state_suffix = hashlib.sha256(str(identity).encode("utf-8")).hexdigest()[:16]
    edit_mode_key = f"customer_information_edit_mode_{state_suffix}"
    editing_item_key = f"customer_information_editing_item_{state_suffix}"
    deleting_item_key = f"customer_information_deleting_item_{state_suffix}"

    with st.container(border=True):
        title_col, action_col = st.columns([4, 1])
        with title_col:
            st.subheader("顧客情報")
        with action_col:
            edit_mode = bool(st.session_state.get(edit_mode_key))
            if st.button(
                "完了" if edit_mode else "編集",
                key=f"customer_information_mode_button_{state_suffix}",
                use_container_width=True,
            ):
                st.session_state[edit_mode_key] = not edit_mode
                st.session_state.pop(editing_item_key, None)
                st.session_state.pop(deleting_item_key, None)
                st.rerun()

        if not has_supabase_config():
            st.warning("顧客情報を使うにはSupabase設定が必要です。")
            return

        try:
            has_regular_items = customer_has_regular_information(
                customer_name,
                customer_key,
            )
        except Exception:
            # 索引確認に失敗した場合は、表示欠落を避けるため従来どおり詳細取得へ戻す。
            has_regular_items = None

        if has_regular_items is False:
            items = []
        else:
            try:
                items = load_customer_information(customer_name, customer_key)
                items = [
                    item for item in items
                    if is_regular_customer_information_item(item)
                ]
            except Exception as exc:
                st.warning(str(exc))
                return

        success_key = f"customer_information_success_{state_suffix}"
        success_message = st.session_state.pop(success_key, None)
        if success_message:
            st.success(success_message)

        if not items:
            st.info("登録されている情報はありません。")

        edit_mode = bool(st.session_state.get(edit_mode_key))
        active_edit_id = st.session_state.get(editing_item_key)

        for index, item in enumerate(items):
            item_id = str(item.get("id", ""))
            field_name = clean_value(item.get("field_name"), blank_text="")
            content = clean_value(item.get("content"), blank_text="")

            if edit_mode and active_edit_id == item_id:
                with st.form(f"customer_information_edit_form_{item_id}"):
                    edited_name = st.text_input("項目名", value=field_name, autocomplete="off")
                    edited_content = st.text_area(
                        "内容", value=content, height=120
                    )
                    save_col, cancel_col = st.columns(2)
                    with save_col:
                        save = st.form_submit_button(
                            "保存", type="primary", use_container_width=True
                        )
                    with cancel_col:
                        cancel = st.form_submit_button(
                            "キャンセル", use_container_width=True
                        )
                if cancel:
                    st.session_state.pop(editing_item_key, None)
                    st.rerun()
                if save:
                    if not str(edited_name).strip():
                        st.warning("項目名を入力してください。")
                    else:
                        try:
                            update_customer_information(
                                item_id, edited_name, edited_content
                            )
                            history_changes = {}
                            if str(edited_name).strip() != field_name:
                                history_changes["項目名"] = (field_name, str(edited_name).strip())
                            if str(edited_content) != content:
                                history_changes[str(edited_name).strip() or field_name] = (
                                    content,
                                    edited_content,
                                )
                            remember_change_history_warning(
                                record_change_history_safely(
                                    "顧客",
                                    customer_key or "",
                                    customer_name,
                                    "変更",
                                    history_changes,
                                    section="顧客情報",
                                )
                            )
                            st.session_state.pop(editing_item_key, None)
                            st.session_state[success_key] = "項目を更新しました。"
                            st.rerun()
                        except RuntimeError as exc:
                            st.error(str(exc))
                continue

            if edit_mode:
                up_col, down_col, name_col, content_col, edit_col, delete_col = st.columns(
                    [0.55, 0.55, 1.5, 3, 0.8, 0.8]
                )
                with up_col:
                    if st.button(
                        "↑",
                        key=f"customer_information_up_{item_id}",
                        disabled=index == 0,
                    ):
                        try:
                            other_item = items[index - 1]
                            reorder_customer_information(item, other_item)
                            remember_change_history_warning(
                                record_change_history_safely(
                                    "顧客",
                                    customer_key or "",
                                    customer_name,
                                    "並び替え",
                                    {
                                        "表示順": (
                                            f"{clean_value(other_item.get('field_name'), blank_text='')} → {field_name}",
                                            f"{field_name} → {clean_value(other_item.get('field_name'), blank_text='')}",
                                        )
                                    },
                                    section="顧客情報",
                                )
                            )
                            st.rerun()
                        except RuntimeError as exc:
                            st.error(str(exc))
                with down_col:
                    if st.button(
                        "↓",
                        key=f"customer_information_down_{item_id}",
                        disabled=index == len(items) - 1,
                    ):
                        try:
                            other_item = items[index + 1]
                            reorder_customer_information(item, other_item)
                            remember_change_history_warning(
                                record_change_history_safely(
                                    "顧客",
                                    customer_key or "",
                                    customer_name,
                                    "並び替え",
                                    {
                                        "表示順": (
                                            f"{field_name} → {clean_value(other_item.get('field_name'), blank_text='')}",
                                            f"{clean_value(other_item.get('field_name'), blank_text='')} → {field_name}",
                                        )
                                    },
                                    section="顧客情報",
                                )
                            )
                            st.rerun()
                        except RuntimeError as exc:
                            st.error(str(exc))
                with name_col:
                    st.markdown(f"**{html.escape(field_name)}**", unsafe_allow_html=True)
                with content_col:
                    safe_content = html.escape(content).replace("\n", "<br>")
                    st.markdown(
                        f'<div style="overflow-wrap:anywhere">{safe_content}</div>',
                        unsafe_allow_html=True,
                    )
                with edit_col:
                    if st.button("編集", key=f"customer_information_edit_{item_id}"):
                        st.session_state[editing_item_key] = item_id
                        st.session_state.pop(deleting_item_key, None)
                        st.rerun()
                with delete_col:
                    if st.button("削除", key=f"customer_information_delete_{item_id}"):
                        st.session_state.pop(editing_item_key, None)
                        st.session_state.pop(deleting_item_key, None)
                        confirm_customer_information_delete_dialog(
                            item,
                            customer_name,
                            customer_key,
                            success_key,
                            editing_item_key,
                            deleting_item_key,
                        )
            else:
                safe_name = html.escape(field_name)
                safe_content = html.escape(content).replace("\n", "<br>")
                st.markdown(
                    (
                        '<div class="customer-information-row">'
                        f'<div class="customer-information-label">{safe_name}</div>'
                        f'<div class="customer-information-content">{safe_content}</div>'
                        '</div>'
                    ),
                    unsafe_allow_html=True,
                )

        if edit_mode:
            st.markdown("---")
            render_customer_information_form(
                customer_name,
                customer_key,
                items,
                state_suffix,
            )



# =========================
# Excel読み込み・整形
# =========================
def delivery_history_identity(values):
    """IDを優先し、IDがない場合だけ顧客名で履歴と現在行を結ぶ。"""
    values = list(values or [])
    customer_id = clean_value(values[0] if len(values) >= 1 else None, blank_text="").strip()
    customer_name = normalize_match_value(values[1] if len(values) >= 2 else None)
    product_name = normalize_match_value(values[4] if len(values) >= 5 else None)
    if not product_name:
        return None
    if customer_id:
        return "id", customer_id, product_name
    if customer_name:
        return "name", customer_name, product_name
    return None


def calculate_predicted_daily_usage_from_states(states):
    """
    H列の実在庫確認を基準に、間にあるI列の配達本数を加味してkg/日の予想使用量を返す。

    HとIが同時入力された行では、Hは配達前の実在庫、Iはその日に配達した本数。
    次回の基準在庫はH+Iとし、今回の予想使用量には同日のIを含めない。
    """
    last_observation_date = None
    last_post_delivery_inventory_kg = None
    deliveries_since_observation_kg = 0.0
    # 新しいH列が使われる前は、従来のI列「本数」が次回配達計算の基準在庫だった。
    # 最初の実在庫確認だけは、直前の従来基準から予想使用量を計算できるようにする。
    legacy_baseline_date = None
    legacy_baseline_inventory_kg = None
    latest_prediction = None

    # 履歴修正で日付の順番が変わる場合もあるため、計算時だけ配達日順へ並べる。
    # 同じ日付の行は、Excelの元の並び順を維持する。
    indexed_states = list(enumerate(states or []))

    def state_sort_key(item):
        original_index, raw_values = item
        values = list(raw_values or [])
        event_date = to_date(values[10] if len(values) >= 11 else None)
        return event_date or date.max, original_index

    indexed_states.sort(key=state_sort_key)

    for _, values in indexed_states:
        values = list(values or [])
        inventory_count = inventory_usage_number(values[7] if len(values) >= 8 else None)
        delivery_count = inventory_usage_number(values[8] if len(values) >= 9 else None)
        kg_per_bottle = inventory_usage_number(values[9] if len(values) >= 10 else None)
        event_date = to_date(values[10] if len(values) >= 11 else None)

        if event_date is None or kg_per_bottle is None or kg_per_bottle <= 0:
            continue
        if inventory_count is not None and inventory_count < 0:
            continue
        if delivery_count is not None and delivery_count < 0:
            continue

        delivery_kg = (delivery_count or 0) * kg_per_bottle

        if inventory_count is None:
            if last_observation_date is not None:
                if delivery_count is not None:
                    deliveries_since_observation_kg += delivery_kg
            elif delivery_count is not None:
                # H列導入前の行は、I列本数をその時点の運用上の基準在庫として扱う。
                # 古い行を合算せず、実在庫確認の直前にある最新基準だけを使う。
                legacy_baseline_date = event_date
                legacy_baseline_inventory_kg = delivery_kg
            continue

        current_inventory_kg = inventory_count * kg_per_bottle
        if last_observation_date is not None and last_post_delivery_inventory_kg is not None:
            elapsed_days = (event_date - last_observation_date).days
            used_kg = (
                last_post_delivery_inventory_kg
                + deliveries_since_observation_kg
                - current_inventory_kg
            )
            if elapsed_days > 0 and used_kg >= 0 and math.isfinite(used_kg):
                latest_prediction = used_kg / elapsed_days
        elif legacy_baseline_date is not None and legacy_baseline_inventory_kg is not None:
            elapsed_days = (event_date - legacy_baseline_date).days
            used_kg = legacy_baseline_inventory_kg - current_inventory_kg
            if elapsed_days > 0 and used_kg >= 0 and math.isfinite(used_kg):
                latest_prediction = used_kg / elapsed_days

        # 同じ日の配達は今回の使用量計算には含めず、次回期間の開始在庫へ加える。
        last_observation_date = event_date
        last_post_delivery_inventory_kg = current_inventory_kg + delivery_kg
        deliveries_since_observation_kg = 0.0

    return latest_prediction


def build_predicted_usage_map(workbook, current_delivery_rows):
    """配達履歴と現在行を1回ずつ読み、各ID・顧客・商品の最新予想使用量を返す。"""
    target_keys = {
        key
        for values in current_delivery_rows.values()
        if (key := delivery_history_identity(values)) is not None
    }
    if not target_keys:
        return {}

    grouped_states = {key: [] for key in target_keys}
    if DELIVERY_HISTORY_SHEET_NAME in workbook.sheetnames:
        history_ws = workbook[DELIVERY_HISTORY_SHEET_NAME]
        for values in history_ws.iter_rows(min_row=2, max_col=14, values_only=True):
            key = delivery_history_identity(values)
            if key in grouped_states:
                grouped_states[key].append(values)

    # 配達履歴には更新前の状態が入り、現在の最新状態は次回配達日シートに残る。
    for row_number in sorted(current_delivery_rows):
        values = current_delivery_rows[row_number]
        key = delivery_history_identity(values)
        if key in grouped_states:
            grouped_states[key].append(values)

    return {
        key: calculate_predicted_daily_usage_from_states(states)
        for key, states in grouped_states.items()
    }


def calculate_delivery_values(delivery_row_values):
    """
    Excelの最新入力値から、次回配達予定と今日時点の残数をアプリ側で計算する。

    残数はExcelの残数セルや数式結果を参照せず、日本時間の今日を基準に、
    配達数量－（配達日から今日までの経過日数×使用数量/日）をkg/本で割る。
    計算結果が0未満の場合も、在庫不足が分かるようマイナスのまま返す。
    """
    def column_value(column_number):
        index = column_number - 1
        return delivery_row_values[index] if index < len(delivery_row_values) else None

    usage = column_value(7)
    kg_per_bottle = column_value(10)
    delivery_date = column_value(11)
    stored_delivery_quantity = column_value(12)

    delivery_day = to_date(delivery_date)
    next_delivery = None
    remaining = None

    try:
        delivery_quantity = float(stored_delivery_quantity)
        daily_usage = float(usage)
        if not all(math.isfinite(value) for value in (delivery_quantity, daily_usage)):
            raise ValueError("次回配達予定の計算に使用する数値が有限値ではありません。")
        if daily_usage == 0:
            raise ValueError("使用数量/日は0以外である必要があります。")
    except Exception:
        return None, None

    # ExcelのM列と同じく、L列「配達数量」を使って次回配達予定を計算する。
    if delivery_day is not None:
        next_delivery = delivery_day + timedelta(
            days=math.floor(delivery_quantity / daily_usage)
        )

        try:
            bottle_weight = float(kg_per_bottle)
            if not math.isfinite(bottle_weight) or bottle_weight == 0:
                raise ValueError("kg/本は0以外の有限値である必要があります。")
            today = get_jst_now().date()
            elapsed_days = (today - delivery_day).days
            remaining_kg = delivery_quantity - (elapsed_days * daily_usage)
            remaining = remaining_kg / bottle_weight
        except Exception:
            remaining = None

    return next_delivery, remaining


def rebuild_sheet1_from_formula_references(excel_source):
    """数式参照元を1回だけ走査し、顧客検索用のSheet1相当データを復元する。"""
    if isinstance(excel_source, BytesIO):
        content = excel_source.getvalue()
    else:
        content = Path(excel_source).read_bytes()

    # 顧客検索時に同じExcelを二重に開かない。read_onlyシートはcell()で
    # 飛び飛びに読むと非常に遅いため、iter_rows()で各シートを1回だけ走査する。
    workbook = load_workbook(
        BytesIO(content),
        # 読み取り専用処理ではVBA本体を展開しない。元のExcel・マクロは変更しない。
        keep_vba=False,
        data_only=False,
        read_only=True,
    )
    try:
        if SHEET_NAME not in workbook.sheetnames or DELIVERY_SHEET_NAME not in workbook.sheetnames:
            return pd.DataFrame()

        sheet1 = workbook[SHEET_NAME]
        delivery = workbook[DELIVERY_SHEET_NAME]

        sheet1_records = []
        sheet1_max_column = max(
            2,
            SHEET1_HIRAGANA_COLUMN,
            SHEET1_ADDRESS_COLUMN,
            SHEET1_MAP_COLUMN,
        )
        for values in sheet1.iter_rows(
            min_row=2,
            max_col=sheet1_max_column,
            values_only=True,
        ):
            source_row = None
            for formula in values[:2]:
                if isinstance(formula, str) and formula.startswith("="):
                    match = re.search(r"(\d+)\s*$", formula.strip())
                    if match:
                        source_row = int(match.group(1))
                        break
            if source_row is None:
                continue

            sheet1_records.append(
                {
                    "source_row": source_row,
                    "ひらがな": values[SHEET1_HIRAGANA_COLUMN - 1]
                    if len(values) >= SHEET1_HIRAGANA_COLUMN
                    else None,
                    "住所": values[SHEET1_ADDRESS_COLUMN - 1]
                    if len(values) >= SHEET1_ADDRESS_COLUMN
                    else None,
                    "マップ位置": values[SHEET1_MAP_COLUMN - 1]
                    if len(values) >= SHEET1_MAP_COLUMN
                    else None,
                }
            )

        if not sheet1_records:
            return pd.DataFrame()

        required_source_rows = {record["source_row"] for record in sheet1_records}
        delivery_rows = {}
        for row_number, values in enumerate(
            delivery.iter_rows(min_row=1, max_col=16, values_only=True),
            start=1,
        ):
            if row_number in required_source_rows:
                delivery_rows[row_number] = values
                if len(delivery_rows) == len(required_source_rows):
                    break

        predicted_usage_map = build_predicted_usage_map(workbook, delivery_rows)

        rows = []
        for sheet1_record in sheet1_records:
            source_row = sheet1_record["source_row"]
            delivery_values = delivery_rows.get(source_row)
            if delivery_values is None:
                continue

            customer_name = delivery_values[1] if len(delivery_values) >= 2 else None
            product_name = delivery_values[4] if len(delivery_values) >= 5 else None
            # 見積り段階など商品名が未登録でも、顧客名があれば顧客として残す。
            # 空の商品名は顧客詳細側で商品カードにしない。
            if not normalize_match_value(customer_name):
                continue

            next_delivery, remaining = calculate_delivery_values(delivery_values)
            history_key = delivery_history_identity(delivery_values)
            rows.append(
                {
                    "ID": delivery_values[0] if len(delivery_values) >= 1 else None,
                    "顧客名": customer_name,
                    "地域": delivery_values[2] if len(delivery_values) >= 3 else None,
                    "コンサル": delivery_values[3] if len(delivery_values) >= 4 else None,
                    "商品名": product_name,
                    "使用数量/日": delivery_values[6] if len(delivery_values) >= 7 else None,
                    "予想使用量/日": predicted_usage_map.get(history_key),
                    "次回配達予定": next_delivery,
                    "残数": remaining,
                    "ひらがな": sheet1_record["ひらがな"],
                    "住所": sheet1_record["住所"],
                    "マップ位置": sheet1_record["マップ位置"],
                    "メーカー": delivery_values[5] if len(delivery_values) >= 6 else None,
                    "在庫本数": delivery_values[7] if len(delivery_values) >= 8 else None,
                    "本数": delivery_values[8] if len(delivery_values) >= 9 else None,
                    "kg/本": delivery_values[9] if len(delivery_values) >= 10 else None,
                    "配達日": delivery_values[10] if len(delivery_values) >= 11 else None,
                    "_配達数量": delivery_values[11] if len(delivery_values) >= 12 else None,
                }
            )
        return pd.DataFrame(rows)
    finally:
        workbook.close()

def normalize_excel_table(excel_source):
    """
    ExcelのSheet1から、顧客一覧表を取り出す。

    対応できる形：
    1) 1行目が見出し
       ID / 顧客名 / 地域 / 商品名 / 使用数量/日 / 次回配達予定 / 残数 / ひらがな

    2) 上部に大きな表示があり、途中の行に見出しがある形
       9行目などに ID / 顧客名 / 地域 / 商品名 ... がある
    """
    # 現在のブック構造ではこちらが最短経路。数式キャッシュの有無にも影響されない。
    try:
        rebuilt = rebuild_sheet1_from_formula_references(excel_source)
        if not rebuilt.empty:
            return rebuilt
    except Exception:
        pass

    try:
        raw = pd.read_excel(
            excel_source,
            sheet_name=SHEET_NAME,
            header=None,
            engine="openpyxl",
        )
    except Exception as e:
        st.error("Excelファイルを読み込めませんでした。")
        st.exception(e)
        st.stop()

    header_row_index = None

    for idx, row in raw.iterrows():
        values = [str(v).strip() for v in row.tolist() if not pd.isna(v)]
        column_mapping = find_required_column_mapping(values)
        score = len(column_mapping)

        # IDだけは上部表示にも出るので、顧客名・ひらがなに相当する列がある行を重視
        if "顧客名" in column_mapping and "ひらがな" in column_mapping and score >= 5:
            header_row_index = idx
            break

    if header_row_index is None:
        # openpyxlでxlsmを保存すると数式セルの前回計算結果が消える。
        # その場合はSheet1の数式参照先である「次回配達日」から一覧を復元する。
        rebuilt = rebuild_sheet1_from_formula_references(excel_source)
        if rebuilt.empty:
            st.error("必要な見出し行が見つかりません。")
            st.write("必要な列：", REQUIRED_COLUMNS)
            st.stop()
        return rebuilt

    header = raw.iloc[header_row_index].tolist()
    df = raw.iloc[header_row_index + 1:].copy()
    df.columns = header

    # 列名の空白除去
    df.columns = [str(c).strip() for c in df.columns]

    column_mapping = find_required_column_mapping(df.columns)
    missing = [col for col in REQUIRED_COLUMNS if col not in column_mapping]
    if missing:
        st.error("必要な列が見つかりません。")
        st.write("見つからない列：", missing)
        st.write("使用できる列名候補：")
        for col in missing:
            st.write(f"- {col}: {', '.join(REQUIRED_COLUMN_CANDIDATES[col])}")
        st.write("Excelから読み取れた列：", list(df.columns))
        st.stop()

    rename_mapping = {
        actual_column: required_column
        for required_column, actual_column in column_mapping.items()
        if actual_column != required_column
    }
    df = df.rename(columns=rename_mapping)

    # 既存機能に必要な列を先頭に置きつつ、備考などの追加列も残す
    ordered_columns = REQUIRED_COLUMNS.copy()
    for col in df.columns:
        if col not in ordered_columns:
            ordered_columns.append(col)

    df = df[ordered_columns].copy()

    # 検索に必要な行だけ残す
    df = df.dropna(subset=["顧客名", "ひらがな"])

    df["顧客名"] = df["顧客名"].astype(str).str.strip()
    df["ひらがな"] = df["ひらがな"].astype(str).str.strip()

    df = df[(df["顧客名"] != "") & (df["ひらがな"] != "")]
    return df


def recalculate_customer_inventory_for_today(df):
    """表示用データの次回配達予定と残数を、日本時間の今日で再計算する。"""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    required_columns = {"使用数量/日", "kg/本", "配達日", "_配達数量"}
    if not required_columns.issubset(df.columns):
        return df

    recalculated = df.copy()
    # JSONから戻した日付列は文字列型になるため、再計算したPythonのdate値を
    # そのまま保持できるよう受け皿だけobject型にする。計算値そのものは変更しない。
    for result_column in ("次回配達予定", "残数"):
        if result_column in recalculated.columns:
            recalculated[result_column] = recalculated[result_column].astype("object")

    for index, row in recalculated.iterrows():
        delivery_values = [None] * 12
        delivery_values[6] = row.get("使用数量/日")
        delivery_values[9] = row.get("kg/本")
        delivery_values[10] = row.get("配達日")
        delivery_values[11] = row.get("_配達数量")
        next_delivery, remaining = calculate_delivery_values(delivery_values)
        recalculated.at[index, "次回配達予定"] = next_delivery
        recalculated.at[index, "残数"] = remaining
    return recalculated


@st.cache_data(ttl=300, show_spinner=False)
def load_fast_dropbox_data(jst_date_key):
    """同じ日本日付ではDropbox確認を5分間再利用し、Excel変更時だけ再生成する。"""
    access_token = get_dropbox_access_token()
    excel_path = get_dropbox_file_path()
    excel_revision = get_dropbox_revision(excel_path, access_token)

    cache_content, cache_response = download_dropbox_file(
        DROPBOX_FAST_CACHE_FILE,
        access_token,
    )
    if cache_content is not None:
        try:
            payload = json.loads(cache_content.decode("utf-8"))
            if (
                payload.get("cache_version") == DROPBOX_FAST_CACHE_VERSION
                and payload.get("excel_revision") == excel_revision
            ):
                records = payload.get("records", [])
                if isinstance(records, list) and records:
                    return recalculate_customer_inventory_for_today(pd.DataFrame(records))
        except Exception:
            pass

    # Excelが更新された時だけ、1回だけ重い解析を行ってJSONを作り直す。
    excel_content, response = download_dropbox_file(excel_path, access_token)
    if excel_content is None:
        raise RuntimeError("Excelを取得できませんでした。\n" + dropbox_error_text(response))
    df = normalize_excel_table(BytesIO(excel_content))
    records = json.loads(df.to_json(orient="records", date_format="iso", force_ascii=False))
    payload = json.dumps(
        {
            "cache_version": DROPBOX_FAST_CACHE_VERSION,
            "excel_revision": excel_revision,
            "records": records,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    upload_response = upload_dropbox_file(
        DROPBOX_FAST_CACHE_FILE,
        payload,
        access_token,
        mode="overwrite",
    )
    if upload_response.status_code != 200:
        # キャッシュ作成失敗でも、取得済みデータで画面表示は続ける。
        return df
    return df


@st.cache_data(ttl=60)
def load_data():
    """
    Dropbox API設定があればDropbox上のExcelを読む。
    設定がなければ同じフォルダのローカルExcelを読む。
    """
    if has_dropbox_auth_config():
        # 日本時間の日付をキャッシュキーに含め、日付が変われば自動的に別キャッシュを使う。
        # 残数と次回配達予定の再計算場所・計算方法は従来のまま変更しない。
        return load_fast_dropbox_data(get_jst_now().date().isoformat())

    return normalize_excel_table(read_excel_local())

@st.cache_data(ttl=300, show_spinner=False)
def load_dispatch_customer_names():
    """配車表の納品先リンク判定用に、顧客名だけを5分間再利用する。"""
    customer_df = load_data()
    if "顧客名" not in customer_df.columns:
        return set()
    return {
        clean_value(value, blank_text="").strip()
        for value in customer_df["顧客名"].tolist()
        if clean_value(value, blank_text="").strip()
    }


# =========================
# 画面遷移
# =========================
def get_query_value(key, default=""):
    """URLパラメータを安全に1つ取り出す"""
    try:
        value = st.query_params.get(key, default)
    except Exception:
        return default

    if isinstance(value, list):
        return value[0] if value else default

    return value if value is not None else default


def update_query_params(**params):
    """ブラウザの戻るボタンを維持しつつ、画面遷移用パラメータだけを更新する。"""
    try:
        for key, value in params.items():
            if value is None or value == "":
                if key in st.query_params:
                    del st.query_params[key]
            else:
                st.query_params[key] = str(value)
    except Exception:
        pass



def make_app_url(
    page="home",
    customer=None,
    customer_search=None,
    region_search=None,
    product_search=None,
    partner_id=None,
    partner_type=None,
    partner_search=None,
    hotel_search=None,
):
    """ブラウザの戻るボタンで戻れるように、通常リンク用URLを作る。"""
    params = {"page": page}
    if customer:
        params["customer"] = str(customer)
    if customer_search:
        params["customer_search"] = str(customer_search)
    if region_search:
        params["region_search"] = str(region_search)
    if product_search:
        params["product_search"] = str(product_search)
    if partner_id:
        params["partner_id"] = str(partner_id)
    if partner_type:
        params["partner_type"] = str(partner_type)
    if partner_search:
        params["partner_search"] = str(partner_search)
    if hotel_search:
        params["hotel_search"] = str(hotel_search)
    return "?" + urllib.parse.urlencode(params)


def render_page_link(
    label,
    page="home",
    customer=None,
    customer_search=None,
    region_search=None,
    product_search=None,
    partner_id=None,
    partner_type=None,
    partner_search=None,
    hotel_search=None,
    class_name="app-nav-link",
):
    """st.buttonではなくHTMLリンクで画面遷移する。これによりブラウザ戻るが効く。"""
    url = make_app_url(
        page=page,
        customer=customer,
        customer_search=customer_search,
        region_search=region_search,
        product_search=product_search,
        partner_id=partner_id,
        partner_type=partner_type,
        partner_search=partner_search,
        hotel_search=hotel_search,
    )
    if page in {"detail", "partner_detail"}:
        class_name = f"{class_name} entity-select-card-link".strip()
    return f'<a class="{class_name}" href="{url}" target="_self">{html.escape(str(label))}</a>'

def sync_page_from_query_params():
    """URLの画面情報を読み、ブラウザ戻る・進むに追従する。"""
    page = str(get_query_value("page", "home")).strip() or "home"
    customer = str(get_query_value("customer", "")).strip()
    partner_id = str(get_query_value("partner_id", "")).strip()
    partner_type = str(get_query_value("partner_type", "")).strip()

    valid_pages = {
        "home",
        "customer_home",
        "customer_list",
        "customer",
        "region",
        "product",
        "calendar",
        "dispatch_table",
        "soluble_inventory",
        "water_it_test",
        "notes",
        "trade_notes",
        "hotel_information",
        "detail",
        "supplier_home",
        "supplier_list",
        "supplier_search",
        "supplier_product",
        "supplier_register",
        "carrier_home",
        "carrier_list",
        "carrier_search",
        "carrier_freight_compare",
        "carrier_register",
        "partner_detail",
        "change_history",
        "attachment_search",
        "estimates",
        "data_backup",
        "login_history",
    }

    raw_page = str(get_query_value("page", "")).strip()
    if page not in valid_pages:
        page = "home"
    if customer and not raw_page:
        page = "detail"

    st.session_state["page"] = page

    if page == "detail" and customer:
        st.session_state["selected_customer"] = customer
    elif page != "detail":
        st.session_state["selected_customer"] = None

    if page == "partner_detail" and partner_id:
        st.session_state["selected_partner_id"] = partner_id
        st.session_state["selected_partner_type"] = partner_type
    elif page != "partner_detail":
        st.session_state["selected_partner_id"] = None
        st.session_state["selected_partner_type"] = None
        clear_trade_partner_immediate_data()


def set_page(page_name, rerun=False):
    st.session_state["page"] = page_name

    if page_name != "detail":
        st.session_state["selected_customer"] = None
    if page_name != "partner_detail":
        st.session_state["selected_partner_id"] = None
        st.session_state["selected_partner_type"] = None
        clear_trade_partner_immediate_data()

    update_query_params(
        page=page_name,
        customer=None,
        partner_id=None,
        partner_type=None,
    )

    if rerun:
        st.rerun()


def select_customer(customer_name, page_name="detail"):
    st.session_state["selected_customer"] = customer_name
    st.session_state["page"] = page_name
    update_query_params(page=page_name, customer=customer_name)


def show_back_home_button(key):
    """既存の顧客画面から顧客メニューへ戻る共通リンク。"""
    st.markdown(
        render_page_link("← 顧客メニューへ戻る", page="customer_home"),
        unsafe_allow_html=True,
    )


def show_detail_search_shortcuts():
    """顧客詳細から、次の検索をすぐ始めるためのショートカット。"""
    col_customer, col_region = st.columns(2)
    with col_customer:
        st.markdown(
            render_page_link("🔍 顧客名で検索", page="customer"),
            unsafe_allow_html=True,
        )
    with col_region:
        st.markdown(
            render_page_link("📍 地域名で検索", page="region"),
            unsafe_allow_html=True,
        )


# =========================
# 顧客詳細
# =========================
def value_for_input(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def date_for_input(value):
    parsed = to_date(value)
    return parsed.strftime("%Y/%m/%d") if parsed else ""


def parse_optional_date(text):
    value = str(text).strip()
    if not value:
        return None
    try:
        value = value.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        # 「2026年7月15日」のような音声入力結果も受け付ける。
        value = value.replace("年", "/").replace("月", "/").replace("日", "")
        # 「7/14」「7月14日」のように年がなければ、現在の年を自動補完する。
        if re.fullmatch(r"\d{1,2}\s*[/\-]\s*\d{1,2}", value):
            value = f"{date.today().year}/{value}"
        parsed = pd.to_datetime(value, errors="raise")
        return parsed.to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
    except Exception as exc:
        raise ValueError("配達日は 2026/07/15 のように入力してください。") from exc


def display_change_value(value):
    if value is None or value == "":
        return "（空欄）"
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y/%m/%d")
    return str(value)


def render_delivery_history_scroll_to_top_once(key_suffix):
    """納品履歴を開いた直後だけ、履歴欄の先頭が見える位置へ戻す。"""
    safe_suffix = re.sub(r"[^0-9A-Za-z_-]", "", str(key_suffix or ""))
    anchor_id = f"aoyama-delivery-history-top-{safe_suffix}"
    scroll_key = f"delivery_history_scroll_to_top_{key_suffix}"

    st.markdown(
        f'<div id="{html.escape(anchor_id, quote=True)}"></div>',
        unsafe_allow_html=True,
    )
    if not st.session_state.pop(scroll_key, False):
        return

    components.html(
        f"""
        <script>
        (() => {{
          const parentWindow = window.parent;
          const parentDocument = parentWindow.document;
          const anchorId = {json.dumps(anchor_id)};
          const move = () => {{
            const target = parentDocument.getElementById(anchorId);
            if (!target) return;
            const rect = target.getBoundingClientRect();
            parentWindow.scrollTo({{
              top: Math.max(0, parentWindow.scrollY + rect.top - 78),
              left: parentWindow.scrollX || 0,
              behavior: 'auto',
            }});
          }};
          parentWindow.requestAnimationFrame(move);
          [80, 220, 520].forEach((delay) => parentWindow.setTimeout(move, delay));
        }})();
        </script>
        """,
        height=0,
        scrolling=False,
    )


def render_customer_delivery_history_editor(
    customer_name,
    product_name,
    key_suffix,
    history_open_key,
    history_edit_key,
):
    """編集画面の中だけで、現在の登録と過去の納品履歴を修正できるようにする。"""
    history_limit_key = f"delivery_history_limit_{key_suffix}"
    render_delivery_history_scroll_to_top_once(key_suffix)
    st.markdown("#### 納品履歴")
    st.caption(
        "Excelの納品ボタンまたはアプリから登録した履歴です。"
        "ここでの修正は新しい納品を追加せず、選んだ記録だけを訂正します。"
    )

    if st.button(
        "編集画面に戻る",
        key=f"delivery_history_back_{key_suffix}",
        use_container_width=True,
    ):
        st.session_state.pop(history_open_key, None)
        st.session_state.pop(history_edit_key, None)
        st.session_state.pop(history_limit_key, None)
        st.rerun()

    try:
        history_limit = max(20, int(st.session_state.get(history_limit_key, 20)))
    except (TypeError, ValueError):
        history_limit = 20

    try:
        content = get_cached_dropbox_excel_content()
        history_result = read_product_delivery_history_from_bytes(
            content,
            customer_name,
            product_name,
            history_limit,
        )
        records = list(history_result.get("records") or [])
        total_count = int(history_result.get("total_count") or 0)
    except Exception as exc:
        st.error(f"納品履歴を読み込めませんでした：{exc}")
        return

    if not records:
        st.info("納品履歴はありません。")
        return

    active_record_key = st.session_state.get(history_edit_key)

    for record in records:
        record_key = record["record_key"]
        with st.container(border=True):
            title = format_date(record.get("配達日"))
            if record.get("source") == "current":
                st.markdown(f"**{title}　現在の登録**")
            else:
                st.markdown(f"**{title}**")

            col1, col2 = st.columns(2)
            with col1:
                st.caption("メーカー")
                st.markdown(f"**{clean_value(record.get('メーカー'))}**")
                st.caption("在庫本数")
                st.markdown(f"**{format_number(record.get('在庫本数'))}**")
                st.caption("本数")
                st.markdown(f"**{format_number(record.get('本数'))}**")
            with col2:
                st.caption("kg/本")
                st.markdown(f"**{format_number(record.get('kg/本'))}**")
                st.caption("配達数量")
                st.markdown(f"**{format_number(record.get('配達数量'))}**")
                st.caption("次回配達予定")
                st.markdown(f"**{format_date(record.get('次回配達予定'))}**")

            if active_record_key != record_key:
                if st.button(
                    "修正",
                    key=f"delivery_history_edit_{key_suffix}_{record_key}",
                    use_container_width=True,
                ):
                    st.session_state[history_edit_key] = record_key
                    st.rerun()
                continue

            st.markdown("**この記録を修正**")
            with st.form(f"delivery_history_form_{key_suffix}_{record_key}"):
                maker = st.text_input(
                    "メーカー",
                    value=value_for_input(record.get("メーカー")),
                    help=VOICE_INPUT_HELP,
                    autocomplete="off",
                )
                inventory_bottles = st.text_input(
                    "在庫本数",
                    value=value_for_input(record.get("在庫本数")),
                    help=VOICE_INPUT_HELP,
                    autocomplete="off",
                )
                bottles = st.text_input(
                    "本数",
                    value=value_for_input(record.get("本数")),
                    help=VOICE_INPUT_HELP,
                    autocomplete="off",
                )
                kg_per_bottle = st.text_input(
                    "kg/本",
                    value=value_for_input(record.get("kg/本")),
                    help=VOICE_INPUT_HELP,
                    autocomplete="off",
                )
                delivery_date = st.text_input(
                    "配達日",
                    value=date_for_input(record.get("配達日")),
                    placeholder="例：2026年7月15日",
                    help=VOICE_INPUT_HELP,
                    autocomplete="off",
                )
                save_col, cancel_col = st.columns(2)
                with save_col:
                    save_correction = st.form_submit_button(
                        "修正を保存",
                        type="primary",
                        use_container_width=True,
                    )
                with cancel_col:
                    cancel_correction = st.form_submit_button(
                        "キャンセル",
                        use_container_width=True,
                    )

            if cancel_correction:
                st.session_state.pop(history_edit_key, None)
                st.rerun()

            if save_correction:
                try:
                    proposed = {
                        "メーカー": str(maker).strip() or None,
                        "在庫本数": (
                            parse_optional_nonnegative_number(
                                inventory_bottles,
                                integer=True,
                            )
                            if str(inventory_bottles).strip()
                            else None
                        ),
                        "本数": (
                            parse_optional_nonnegative_number(
                                bottles,
                                integer=True,
                            )
                            if str(bottles).strip()
                            else None
                        ),
                        "kg/本": (
                            parse_optional_nonnegative_number(
                                kg_per_bottle,
                                integer=False,
                            )
                            if str(kg_per_bottle).strip()
                            else None
                        ),
                        "配達日": (
                            parse_optional_date(delivery_date)
                            if str(delivery_date).strip()
                            else None
                        ),
                    }
                    changes = {
                        label: (record.get(label), proposed[label])
                        for label in proposed
                        if not same_excel_value(record.get(label), proposed[label])
                    }
                    if not changes:
                        st.warning("変更された項目がありません。")
                        continue

                    diagnostic_started_at = time.time()
                    with st.spinner("バックアップを作成して履歴を修正しています…"):
                        result = save_customer_delivery_history_correction(
                            customer_name,
                            product_name,
                            record,
                            proposed,
                        )
                        result["usage_warning"] = ""
                        diagnostic_history_started = time.perf_counter()
                        result["history_warning"] = record_change_history_safely(
                            "顧客",
                            "",
                            customer_name,
                            "変更",
                            changes,
                            section=f"商品：{product_name}／納品履歴",
                        )
                        result.setdefault("diagnostic_timings", {})["変更履歴保存"] = (
                            time.perf_counter() - diagnostic_history_started
                        )

                    result["diagnostic_started_at"] = diagnostic_started_at
                    result["diagnostic_before_rerun_seconds"] = time.time() - diagnostic_started_at
                    # 納品履歴の修正が終わったら、履歴画面と通常編集画面をすべて閉じる。
                    st.session_state.pop(history_edit_key, None)
                    st.session_state.pop(history_open_key, None)
                    st.session_state.pop(history_limit_key, None)
                    st.session_state.pop(f"excel_edit_{key_suffix}", None)
                    st.session_state.pop(f"excel_confirm_{key_suffix}", None)
                    st.session_state["excel_save_success"] = {
                        **result,
                        "customer_name": customer_name,
                        "product_name": f"{product_name}（納品履歴修正）",
                    }
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))
                except Exception as exc:
                    st.error(f"納品履歴を修正できませんでした：{exc}")

    if total_count > len(records):
        remaining_count = total_count - len(records)
        if st.button(
            f"さらに表示（残り{remaining_count}件）",
            key=f"delivery_history_more_{key_suffix}_{history_limit}",
            use_container_width=True,
        ):
            st.session_state[history_limit_key] = history_limit + 20
            st.rerun()


def render_customer_excel_editor(customer_name, customer_key, product_name, current):
    """商品カード内に、確認画面付きのExcel編集欄を追加する。"""
    identity = f"{customer_name}|{product_name}"
    key_suffix = str(abs(hash(identity)))
    edit_key = f"excel_edit_{key_suffix}"
    confirm_key = f"excel_confirm_{key_suffix}"
    history_open_key = f"delivery_history_open_{key_suffix}"
    history_edit_key = f"delivery_history_active_edit_{key_suffix}"
    history_limit_key = f"delivery_history_limit_{key_suffix}"

    if current.get("商品一致件数") == 0:
        st.error("顧客名・商品名が一致する行が見つからないため編集できません。")
        return
    if current.get("商品一致件数", 0) > 1:
        st.error("同じ顧客名・商品名の行が複数見つかりました。確認してください。")
        return

    st.caption("メーカー")
    st.markdown(f"**{clean_value(current.get('メーカー'))}**")
    col_a, col_b = st.columns(2)
    with col_a:
        st.caption("在庫本数")
        st.markdown(f"**{format_number(current.get('在庫本数'))}**")
    with col_b:
        st.caption("本数")
        st.markdown(f"**{format_number(current.get('本数'))}**")
    col_c, col_d = st.columns(2)
    with col_c:
        st.caption("kg/本")
        st.markdown(f"**{format_number(current.get('kg/本'))}**")
    with col_d:
        st.caption("配達日")
        st.markdown(f"**{format_date(current.get('配達日'))}**")

    if st.session_state.get(history_open_key):
        render_customer_delivery_history_editor(
            customer_name,
            product_name,
            key_suffix,
            history_open_key,
            history_edit_key,
        )
        return

    if not st.session_state.get(edit_key) and not st.session_state.get(confirm_key):
        if st.button("編集", key=f"edit_button_{key_suffix}"):
            st.session_state[edit_key] = True
            st.rerun()
        return

    if st.session_state.get(confirm_key):
        pending = st.session_state[confirm_key]
        st.markdown("**保存前の確認**")
        for label, values in pending["changes"].items():
            st.write(f"{label}：{display_change_value(values[0])} → {display_change_value(values[1])}")
        save_col, cancel_col = st.columns(2)
        with save_col:
            if st.button("保存", key=f"save_confirm_{key_suffix}", type="primary", use_container_width=True):
                try:
                    diagnostic_started_at = time.time()
                    with st.spinner("バックアップを作成して保存しています…"):
                        result = save_customer_excel_changes(customer_name, product_name, pending["proposed"])
                        result["usage_warning"] = ""
                        diagnostic_history_started = time.perf_counter()
                        result["history_warning"] = record_change_history_safely(
                            "顧客",
                            "",
                            customer_name,
                            "変更",
                            pending["changes"],
                            section=f"商品：{product_name}",
                        )
                        result.setdefault("diagnostic_timings", {})["変更履歴保存"] = (
                            time.perf_counter() - diagnostic_history_started
                        )
                    result["diagnostic_started_at"] = diagnostic_started_at
                    result["diagnostic_before_rerun_seconds"] = time.time() - diagnostic_started_at
                    st.session_state.pop(confirm_key, None)
                    st.session_state.pop(edit_key, None)
                    st.session_state.pop(history_open_key, None)
                    st.session_state.pop(history_edit_key, None)
                    st.session_state.pop(history_limit_key, None)
                    st.session_state["excel_save_success"] = {
                        **result,
                        "customer_name": customer_name,
                        "product_name": product_name,
                    }
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))
        with cancel_col:
            if st.button("キャンセル", key=f"cancel_confirm_{key_suffix}", use_container_width=True):
                st.session_state.pop(confirm_key, None)
                st.session_state.pop(edit_key, None)
                st.rerun()
        return

    with st.form(f"excel_edit_form_{key_suffix}"):
        st.caption(f"🎤 {VOICE_INPUT_HELP} 入力欄は毎回空白から始まります。")
        maker = st.text_input(
            "メーカー",
            value="",
            placeholder="メーカー名を入力",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        st.caption(
            "在庫本数＝配達前に確認した本数、本数＝今回配達する本数です。"
            "両方入力した場合は合計で次回配達予定を計算します。"
        )
        inventory_bottles = st.text_input(
            "在庫本数",
            value="",
            placeholder="確認できた場合：例 3本",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        bottles = st.text_input(
            "本数",
            value="",
            placeholder="今回配達：例 10本",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        kg_per_bottle = st.text_input(
            "kg/本",
            value="",
            placeholder="例：450キロ",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        delivery_date = st.text_input(
            "配達日",
            value="",
            placeholder="例：2026年7月15日",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        save_col, cancel_col = st.columns(2)
        with save_col:
            proceed = st.form_submit_button("保存", type="primary", use_container_width=True)
        with cancel_col:
            cancel = st.form_submit_button("キャンセル", use_container_width=True)

    if st.button(
        "納品履歴",
        key=f"delivery_history_open_button_{key_suffix}",
        use_container_width=True,
    ):
        st.session_state[history_open_key] = True
        st.session_state[history_limit_key] = 20
        st.session_state[f"delivery_history_scroll_to_top_{key_suffix}"] = True
        st.session_state.pop(history_edit_key, None)
        st.rerun()

    if cancel:
        st.session_state.pop(edit_key, None)
        st.session_state.pop(history_open_key, None)
        st.session_state.pop(history_edit_key, None)
        st.session_state.pop(history_limit_key, None)
        st.rerun()
    if proceed:
        try:
            inventory_text = str(inventory_bottles).strip()
            bottles_text = str(bottles).strip()
            delivery_date_text = str(delivery_date).strip()
            event_input = bool(inventory_text or bottles_text or delivery_date_text)

            if event_input and not delivery_date_text:
                raise ValueError("在庫本数または本数を入力する場合は、配達日も入力してください。")
            if delivery_date_text and not (inventory_text or bottles_text):
                raise ValueError("配達日を入力する場合は、在庫本数または本数も入力してください。")

            proposed = {
                "メーカー": (
                    str(maker).strip()
                    or normalize_existing_excel_value(current.get("メーカー"))
                ),
                # 新しい在庫・配達を登録するときは、空欄側をそのまま空欄として保存する。
                # メーカーやkg/本だけを直す場合は、現在の在庫・配達値を維持する。
                "在庫本数": (
                    parse_optional_nonnegative_number(inventory_bottles, integer=True)
                    if inventory_text
                    else (None if event_input else normalize_existing_excel_value(current.get("在庫本数")))
                ),
                "本数": (
                    parse_optional_nonnegative_number(bottles, integer=True)
                    if bottles_text
                    else (None if event_input else normalize_existing_excel_value(current.get("本数")))
                ),
                "kg/本": (
                    parse_optional_nonnegative_number(kg_per_bottle, integer=False)
                    if str(kg_per_bottle).strip()
                    else normalize_existing_excel_value(current.get("kg/本"))
                ),
                "配達日": (
                    parse_optional_date(delivery_date)
                    if delivery_date_text
                    else normalize_existing_excel_value(current.get("配達日"))
                ),
            }
            changes = {
                label: (current.get(label), proposed[label])
                for label in proposed
                if not same_excel_value(current.get(label), proposed[label])
            }
            if not changes:
                st.warning("変更された項目がありません。")
            else:
                diagnostic_started_at = time.time()
                with st.spinner("DropboxのExcelへ保存しています…"):
                    result = save_customer_excel_changes(
                        customer_name,
                        product_name,
                        proposed,
                    )
                    result["usage_warning"] = ""
                    diagnostic_history_started = time.perf_counter()
                    result["history_warning"] = record_change_history_safely(
                        "顧客",
                        "",
                        customer_name,
                        "変更",
                        changes,
                        section=f"商品：{product_name}",
                    )
                    result.setdefault("diagnostic_timings", {})["変更履歴保存"] = (
                        time.perf_counter() - diagnostic_history_started
                    )
                result["diagnostic_started_at"] = diagnostic_started_at
                result["diagnostic_before_rerun_seconds"] = time.time() - diagnostic_started_at
                st.session_state.pop(confirm_key, None)
                st.session_state.pop(edit_key, None)
                st.session_state.pop(history_open_key, None)
                st.session_state.pop(history_edit_key, None)
                st.session_state.pop(history_limit_key, None)
                st.session_state["excel_save_success"] = {
                    **result,
                    "customer_name": customer_name,
                    "product_name": product_name,
                }
                st.rerun()
        except ValueError as exc:
            st.error(str(exc))
        except Exception as exc:
            st.error(f"保存できませんでした：{exc}")



def render_customer_map_editor(customer_name, current):
    """顧客単位の住所・マップ位置専用編集欄。"""
    key_suffix = str(abs(hash(f"map|{customer_name}")))
    edit_key = f"map_edit_{key_suffix}"
    confirm_key = f"map_confirm_{key_suffix}"

    st.markdown("### 📍 住所・マップ位置")
    st.write(f"**住所：** {clean_value(current.get('住所'))}")
    st.write(f"**マップ位置：** {clean_value(current.get('マップ位置'))}")

    if not st.session_state.get(edit_key) and not st.session_state.get(confirm_key):
        if st.button("住所・マップ位置を編集", key=f"map_edit_button_{key_suffix}"):
            st.session_state[edit_key] = True
            st.rerun()
        return

    if st.session_state.get(confirm_key):
        pending = st.session_state[confirm_key]
        st.markdown("**保存前の確認**")
        for label, values in pending["changes"].items():
            st.write(
                f"{label}：{display_change_value(values[0])} → "
                f"{display_change_value(values[1])}"
            )

        save_col, cancel_col = st.columns(2)
        with save_col:
            if st.button(
                "保存",
                key=f"map_save_confirm_{key_suffix}",
                type="primary",
                use_container_width=True,
            ):
                try:
                    with st.spinner("バックアップを作成して保存しています…"):
                        result = save_customer_map_changes(
                            customer_name,
                            pending["住所"],
                            pending["マップ位置"],
                        )
                        result["history_warning"] = record_change_history_safely(
                            "顧客",
                            "",
                            customer_name,
                            "変更",
                            pending["changes"],
                            section="住所・マップ位置",
                        )
                    st.session_state.pop(confirm_key, None)
                    st.session_state.pop(edit_key, None)
                    st.session_state["excel_save_success"] = {
                        **result,
                        "customer_name": customer_name,
                        "product_name": "住所・マップ位置",
                    }
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

        with cancel_col:
            if st.button(
                "キャンセル",
                key=f"map_cancel_confirm_{key_suffix}",
                use_container_width=True,
            ):
                st.session_state.pop(confirm_key, None)
                st.session_state.pop(edit_key, None)
                st.rerun()
        return

    with st.form(f"map_edit_form_{key_suffix}"):
        st.caption(f"🎤 {VOICE_INPUT_HELP}")
        address = st.text_input(
            "住所",
            value=value_for_input(current.get("住所")),
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        map_location = st.text_input(
            "マップ位置",
            value=value_for_input(current.get("マップ位置")),
            help=f"緯度,経度／Googleマップ共有URL／文字列を入力できます。{VOICE_INPUT_HELP}",
            autocomplete="off",
        )
        save_col, cancel_col = st.columns(2)
        with save_col:
            proceed = st.form_submit_button(
                "保存",
                type="primary",
                use_container_width=True,
            )
        with cancel_col:
            cancel = st.form_submit_button(
                "キャンセル",
                use_container_width=True,
            )

    if cancel:
        st.session_state.pop(edit_key, None)
        st.rerun()

    if proceed:
        try:
            proposed_address = str(address)
            proposed_map = validate_map_location(map_location)
            changes = {}
            if not same_excel_value(current.get("住所"), proposed_address):
                changes["住所"] = (current.get("住所"), proposed_address)
            if not same_excel_value(current.get("マップ位置"), proposed_map):
                changes["マップ位置"] = (
                    current.get("マップ位置"),
                    proposed_map,
                )

            if not changes:
                st.warning("変更された項目がありません。")
            else:
                with st.spinner("DropboxのExcelへ保存しています…"):
                    result = save_customer_map_changes(
                        customer_name,
                        proposed_address,
                        proposed_map,
                    )
                    result["history_warning"] = record_change_history_safely(
                        "顧客",
                        "",
                        customer_name,
                        "変更",
                        changes,
                        section="住所・マップ位置",
                    )
                st.session_state.pop(confirm_key, None)
                st.session_state.pop(edit_key, None)
                st.session_state["excel_save_success"] = {
                    **result,
                    "customer_name": customer_name,
                    "product_name": "住所・マップ位置",
                }
                st.rerun()
        except ValueError as exc:
            st.error(str(exc))
        except Exception as exc:
            st.error(f"保存できませんでした：{exc}")


def show_customer_detail(df, customer_name):
    detail = df[df["顧客名"] == customer_name].copy()

    if detail.empty:
        st.warning("選択した顧客の情報が見つかりません。")
        return

    show_back_home_button("detail_back_home")
    show_detail_search_shortcuts()

    # 商品名が空欄の行は顧客情報として残すが、空の商品カードにはしない。
    # 商品名がある行については、従来どおり使用数量/日が0・空白・NaNなら表示しない。
    named_product_mask = detail["商品名"].apply(
        lambda value: bool(normalize_match_value(value))
    )
    visible_detail = detail[
        named_product_mask & (~detail["使用数量/日"].apply(is_blank_or_zero))
    ].copy()
    has_named_product = bool(named_product_mask.any())

    region = clean_value(detail.iloc[0]["地域"])
    consultant = clean_value(detail.iloc[0].get("コンサル"), blank_text="")

    st.markdown("---")
    line_connected = get_line_connected(customer_name)

    name_col, line_col, _ = st.columns([6, 3, 3])
    with name_col:
        st.markdown(
            '<div class="customer-name-row customer-detail-name-row">'
            f'<span>👤 {html.escape(clean_value(customer_name))}</span>'
            "</div>",
            unsafe_allow_html=True,
        )
    with line_col:
        if line_connected:
            st.markdown(
                '<div class="line-detail-static">LINE ○</div>',
                unsafe_allow_html=True,
            )
        else:
            with st.popover("LINE ×"):
                st.caption("LINEを○にしますか？")
                if st.button(
                    "○にする",
                    key=f"line_status_{make_line_status_id(customer_name)}",
                ):
                    with st.spinner("保存しています…"):
                        if save_line_connected(customer_name, True):
                            st.toast("LINEを○にしました。")
                            st.rerun()

    st.write(f"**地域：** {region}　　**コンサル：** {consultant}")
    st.write(f"**商品数：** {len(visible_detail)}件")

    success = st.session_state.pop("excel_save_success", None)
    if success:
        st.success("保存しました")
        st.success("バックアップを作成しました")
        st.write(f"**更新日時：** {success['updated_at'].strftime('%Y/%m/%d %H:%M:%S')}")
        st.write(f"**更新した顧客名：** {success['customer_name']}")
        st.write(f"**更新した商品名：** {success['product_name']}")
        if success.get("cleanup_warning"):
            st.warning(success["cleanup_warning"])
        if success.get("usage_warning"):
            st.warning(success["usage_warning"])
        if success.get("history_warning"):
            st.warning(success["history_warning"])

        diagnostic_timings = success.get("diagnostic_timings") or {}
        if diagnostic_timings:
            try:
                diagnostic_started_at = float(success.get("diagnostic_started_at") or 0)
            except (TypeError, ValueError):
                diagnostic_started_at = 0.0
            try:
                diagnostic_before_rerun = float(
                    success.get("diagnostic_before_rerun_seconds") or 0
                )
            except (TypeError, ValueError):
                diagnostic_before_rerun = 0.0

            diagnostic_total_to_display = (
                max(0.0, time.time() - diagnostic_started_at)
                if diagnostic_started_at > 0
                else diagnostic_before_rerun
            )
            diagnostic_rerun_seconds = max(
                0.0,
                diagnostic_total_to_display - diagnostic_before_rerun,
            )

            st.markdown("**⏱ 保存時間診断**")
            diagnostic_order = (
                "Dropbox接続準備",
                "Excel取得",
                "バックアップ作成",
                "Excel編集・保存・検証",
                "　Excel① マクロ付きExcelを開く",
                "　Excel② 計算値用Excelを開く",
                "　Excel③ 対象検索・書換え",
                "　Excel④ xlsm保存",
                "　Excel⑤ 保存後Excelを開く",
                "　Excel⑥ 保存後内容検証",
                "Dropbox本番保存",
                "保存結果確認",
                "表示用データ更新",
                "　表示① 既存JSON取得",
                "　表示② JSON解析・全体再計算",
                "　表示③ 対象商品差替え",
                "　表示④ 全体再生成（フォールバック）",
                "　表示⑤ 即時表示データ保持",
                "　表示⑥ JSON変換",
                "　表示⑦ JSON保存",
                "バックアップ整理",
                "関連キャッシュ更新",
                "変更履歴保存",
            )
            for diagnostic_label in diagnostic_order:
                if diagnostic_label not in diagnostic_timings:
                    continue
                try:
                    diagnostic_seconds = float(diagnostic_timings[diagnostic_label])
                except (TypeError, ValueError):
                    continue
                st.write(f"{diagnostic_label}：{diagnostic_seconds:.2f}秒")
            diagnostic_fast_error = diagnostic_timings.get("表示高速経路エラー")
            if diagnostic_fast_error:
                st.write(f"**表示高速経路エラー：{diagnostic_fast_error}**")
            diagnostic_xml_error = diagnostic_timings.get("XML直接保存エラー")
            if diagnostic_xml_error:
                st.write(f"**XML直接保存エラー：{diagnostic_xml_error}**")
            try:
                diagnostic_save_seconds = float(success.get("diagnostic_save_seconds") or 0)
            except (TypeError, ValueError):
                diagnostic_save_seconds = 0.0
            if diagnostic_save_seconds > 0:
                st.write(f"**保存処理小計：{diagnostic_save_seconds:.2f}秒**")
            st.write(f"画面再表示：{diagnostic_rerun_seconds:.2f}秒")
            st.write(f"**保存開始から表示まで：{diagnostic_total_to_display:.2f}秒**")

    try:
        map_info = get_customer_map_info(detail)
        if map_info and map_info["map_url"]:
            show_google_maps_button(map_info["map_url"])
    except Exception:
        pass

    # 詳細表示では重いExcelを開かず、高速JSON内の現在値を使う。
    first_detail = detail.iloc[0]
    current_map_values = {
        "住所": first_detail.get("住所"),
        "マップ位置": first_detail.get("マップ位置"),
        "顧客一致件数": len(detail),
    }
    render_customer_map_editor(customer_name, current_map_values)

    customer_key = get_stable_customer_key(detail)

    # WATER itのポイント名と顧客名が一致する場合だけ、最新値を読み取り専用で表示する。
    render_customer_water_it_card(customer_name)

    if visible_detail.empty:
        if not has_named_product:
            st.info("登録されている商品はありません。")
        else:
            st.info("表示対象の商品はありません。使用数量/日が0または空白の商品は非表示にしています。")

    # 同じ商品に使用中行が複数あっても、商品カードは1つだけ表示する。
    # 複数の使用中行がある場合はカード内で警告し、編集を停止する。
    visible_products = []
    seen_products = set()
    for _, candidate_row in visible_detail.iterrows():
        candidate_product = clean_value(candidate_row["商品名"], blank_text="").strip()
        if not candidate_product or candidate_product in seen_products:
            continue
        seen_products.add(candidate_product)
        visible_products.append(candidate_row)


    for row in visible_products:
        product_name = clean_value(row["商品名"])
        customer_id = clean_value(row["ID"])
        usage = format_number(row["使用数量/日"])
        predicted_usage = format_number(
            row.get("予想使用量/日"),
            blank_text="",
        )
        next_date = format_date(row["次回配達予定"])
        remaining = format_number(row["残数"])

        with st.container(border=True):
            st.subheader(f"📦 {product_name}")

            col1, col2 = st.columns(2)

            with col1:
                st.caption("ID")
                st.markdown(f"**{customer_id}**")

                st.caption("使用数量/日")
                st.markdown(f"**{usage}**")

                st.caption("予想使用量/日")
                if predicted_usage:
                    st.markdown(f"**{predicted_usage}**")
                else:
                    # 比較できる前回基準がない場合だけ空白表示。
                    st.markdown("&nbsp;", unsafe_allow_html=True)

            with col2:
                st.caption("次回配達予定")
                st.markdown(f"**{next_date}**")

                st.caption("残数")
                st.markdown(f"**{remaining}**")

            product_match_count = int(
                (
                    (detail["商品名"].astype(str).str.strip() == product_name)
                    & (~detail["使用数量/日"].apply(is_blank_or_zero))
                ).sum()
            )
            current_edit_values = {
                "メーカー": row.get("メーカー"),
                "在庫本数": row.get("在庫本数"),
                "本数": row.get("本数"),
                "kg/本": row.get("kg/本"),
                "配達日": row.get("配達日"),
                "住所": current_map_values["住所"],
                "マップ位置": current_map_values["マップ位置"],
                "商品一致件数": product_match_count,
                "顧客一致件数": len(detail),
            }
            render_customer_excel_editor(customer_name, customer_key, product_name, current_edit_values)


    if normalize_soluble_customer_name(customer_name) in {
        normalize_soluble_customer_name(name) for name in SOLUBLE_CUSTOMER_NAMES
    }:
        try:
            soluble_content, _ = load_soluble_workbook_content()
            soluble_customer_summary = get_soluble_customer_summary(
                soluble_content,
                customer_name,
            )
            if soluble_customer_summary is not None:
                render_soluble_customer_product_card(
                    customer_name,
                    soluble_customer_summary,
                    key_scope="customer_detail",
                )
            else:
                st.warning(f"ソリュブルシートに「{customer_name}」が見つかりません。")
        except Exception as exc:
            st.warning(f"ソリュブル情報を読み込めませんでした：{exc}")

    render_customer_estimates_section(customer_name, customer_key)
    render_customer_attachments_lazy_section(customer_name, customer_key)

    render_customer_information_card(customer_name, customer_key)
    show_customer_notes(customer_name)
    render_past_products_section(customer_name, customer_key, detail, visible_detail)

# =========================
# 顧客名一覧
# =========================
CUSTOMER_DIRECTORY_GROUPS = {
    "あ行": set("あいうえおぁぃぅぇぉ"),
    "か行": set("かきくけこがぎぐげご"),
    "さ行": set("さしすせそざじずぜぞ"),
    "た行": set("たちつてとだぢづでどっ"),
    "な行": set("なにぬねの"),
    "は行": set("はひふへほばびぶべぼぱぴぷぺぽ"),
    "ま行": set("まみむめも"),
    "や行": set("やゆよゃゅょ"),
    "ら行": set("らりるれろ"),
    "わ行": set("わをんゎ"),
}


def normalize_directory_kana(value):
    """顧客名一覧の並び替え用に、カタカナをひらがなへ寄せる。"""
    text = clean_value(value, blank_text="").strip()
    return "".join(
        chr(ord(char) - 0x60) if "ァ" <= char <= "ヶ" else char
        for char in text
    )


def get_customer_directory_group(value):
    kana = normalize_directory_kana(value)
    if not kana:
        return "その他"
    first = kana[0]
    for group_name, characters in CUSTOMER_DIRECTORY_GROUPS.items():
        if first in characters:
            return group_name
    return "その他"


def show_customer_directory(df=None):
    st.subheader("👥 顧客名一覧")
    show_back_home_button("customer_directory_back_home")
    st.caption("Sheet1の顧客を五十音順で表示します。顧客名を押すと詳細を開きます。")

    if df is None:
        with st.spinner("顧客データを読み込んでいます…"):
            df = load_data()

    directory = df[["顧客名", "地域", "商品名", "ひらがな"]].copy()
    for column in directory.columns:
        directory[column] = directory[column].fillna("").astype(str).str.strip()
    directory = directory[directory["顧客名"] != ""]

    if directory.empty:
        st.info("表示できる顧客がありません。")
        return

    customers = (
        directory.groupby("顧客名", as_index=False)
        .agg(
            地域=("地域", "first"),
            ひらがな=("ひらがな", "first"),
            商品数=("商品名", lambda values: values[values != ""].nunique()),
        )
    )
    customers["並び順"] = customers.apply(
        lambda row: normalize_directory_kana(row["ひらがな"] or row["顧客名"]),
        axis=1,
    )
    customers["五十音"] = customers["並び順"].map(get_customer_directory_group)

    kana_filter = st.selectbox(
        "五十音で絞り込み",
        ["すべて", *CUSTOMER_DIRECTORY_GROUPS.keys(), "その他"],
        key="customer_directory_kana_filter",
    )
    keyword = st.text_input(
        "一覧を絞り込み",
        placeholder="顧客名・ひらがな・地域",
        key="customer_directory_keyword",
        autocomplete="off",
    ).strip()

    filtered = customers
    if kana_filter != "すべて":
        filtered = filtered[filtered["五十音"] == kana_filter]
    if keyword:
        keyword_kana = normalize_directory_kana(keyword)
        name_text = filtered["顧客名"].astype(str)
        region_text = filtered["地域"].astype(str)
        kana_text = filtered["並び順"].astype(str)
        filtered = filtered[
            name_text.str.contains(keyword, case=False, na=False, regex=False)
            | region_text.str.contains(keyword, case=False, na=False, regex=False)
            | kana_text.str.contains(keyword_kana, case=False, na=False, regex=False)
        ]

    filtered = filtered.sort_values(["並び順", "顧客名"]).reset_index(drop=True)
    st.write(f"顧客：{len(filtered)}件")

    if filtered.empty:
        st.info("条件に一致する顧客がありません。")
        return

    parts = ['<div class="customer-directory">']
    for _, row in filtered.iterrows():
        name = clean_value(row["顧客名"])
        region = clean_value(row["地域"], blank_text="未設定")
        product_count = int(row["商品数"])
        url = html.escape(make_app_url(page="detail", customer=name), quote=True)
        parts.append(
            (
                f'<a class="customer-directory-item" href="{url}" target="_self">'
                f'<span class="customer-directory-name">{html.escape(name)}</span>'
                f'<span class="customer-directory-meta">地域：{html.escape(region)}　商品：{product_count}件</span>'
                '</a>'
            )
        )
    parts.append("</div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


# =========================
# 全ページ共通の顧客検索
# =========================
GLOBAL_CUSTOMER_SEARCH_RESULT_LIMIT = 10
GLOBAL_CUSTOMER_SEARCH_ROUTE_KEY = "_global_customer_search_route"
GLOBAL_CUSTOMER_SEARCH_LIVE_KEY = "global_customer_search_live"
GLOBAL_CUSTOMER_SEARCH_INPUT_KEY = "global_customer_search_input"


def normalize_customer_search_terms(keyword):
    """顧客検索文字を、全角・半角スペース区切りのAND検索語へ整形する。"""
    normalized = clean_value(keyword, blank_text="").strip().casefold()
    return [term for term in re.split(r"[\s　]+", normalized) if term]


def find_customer_search_candidates(df, keyword):
    """顧客名・ひらがな・地域を対象に、全検索語を含む顧客を返す。"""
    columns = ["顧客名", "ひらがな", "地域"]
    if not isinstance(df, pd.DataFrame) or not set(columns).issubset(df.columns):
        return pd.DataFrame(columns=columns)

    terms = normalize_customer_search_terms(keyword)
    if not terms:
        return pd.DataFrame(columns=columns)

    customers = df[columns].copy()
    for column in columns:
        customers[column] = customers[column].apply(
            lambda value: clean_value(value, blank_text="").strip()
        )
    customers = customers[customers["顧客名"] != ""]
    searchable = customers.apply(
        lambda row: " ".join(
            [row["顧客名"], row["ひらがな"], row["地域"]]
        ).casefold(),
        axis=1,
    )
    matched = customers[
        searchable.apply(lambda value: all(term in value for term in terms))
    ].drop_duplicates(
        subset=["顧客名"],
        keep="first",
    )
    return matched.reset_index(drop=True)


@st.cache_data(ttl=300, show_spinner=False)
def load_global_customer_search_index():
    """共通検索用に、顧客名・ひらがな・地域だけを5分間再利用する。"""
    columns = ["顧客名", "ひらがな", "地域"]
    customer_df = load_data()
    if not isinstance(customer_df, pd.DataFrame) or not set(columns).issubset(customer_df.columns):
        return pd.DataFrame(columns=columns + ["_検索文字"])

    customers = customer_df[columns].copy()
    for column in columns:
        customers[column] = customers[column].apply(
            lambda value: clean_value(value, blank_text="").strip()
        )
    customers = customers[customers["顧客名"] != ""].drop_duplicates(
        subset=["顧客名"],
        keep="first",
    )
    customers["_検索文字"] = customers.apply(
        lambda row: " ".join(
            [row["顧客名"], row["ひらがな"], row["地域"]]
        ).casefold(),
        axis=1,
    )
    return customers.reset_index(drop=True)

def find_global_customer_search_candidates(search_index, keyword):
    """共通検索用の小さな索引から、全検索語を含む顧客を返す。"""
    columns = ["顧客名", "ひらがな", "地域"]
    if (
        not isinstance(search_index, pd.DataFrame)
        or not set(columns + ["_検索文字"]).issubset(search_index.columns)
    ):
        return pd.DataFrame(columns=columns)

    terms = normalize_customer_search_terms(keyword)
    if not terms:
        return pd.DataFrame(columns=columns)

    matched = pd.Series(True, index=search_index.index)
    searchable = search_index["_検索文字"].fillna("").astype(str)
    for term in terms:
        matched &= searchable.str.contains(term, regex=False, na=False)

    return search_index.loc[matched, columns].reset_index(drop=True)

def render_customer_search_candidate_cards(customers, limit=None):
    """顧客検索候補を、顧客詳細へ直接移動できるカードで表示する。"""
    visible = customers if limit is None else customers.head(int(limit))
    parts = ['<div class="customer-directory">']
    for _, row in visible.iterrows():
        name = clean_value(row.get("顧客名"), blank_text="").strip()
        if not name:
            continue
        region = clean_value(row.get("地域"), blank_text="未設定").strip() or "未設定"
        url = html.escape(make_app_url(page="detail", customer=name), quote=True)
        parts.append(
            (
                f'<a class="customer-directory-item" href="{url}" target="_self">'
                f'<span class="customer-directory-name">{html.escape(name)}</span>'
                f'<span class="customer-directory-meta">地域：{html.escape(region)}</span>'
                '</a>'
            )
        )
    parts.append("</div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


def clear_global_customer_search_when_route_changes():
    """別画面・別顧客へ移動した後は、共通検索欄を空に戻す。"""
    route = (
        st.session_state.get("page", "home"),
        st.session_state.get("selected_customer"),
        st.session_state.get("selected_partner_id"),
        st.session_state.get("selected_partner_type"),
    )
    previous_route = st.session_state.get(GLOBAL_CUSTOMER_SEARCH_ROUTE_KEY)
    if previous_route is not None and previous_route != route:
        st.session_state.pop(GLOBAL_CUSTOMER_SEARCH_LIVE_KEY, None)
        st.session_state.pop(GLOBAL_CUSTOMER_SEARCH_INPUT_KEY, None)
    st.session_state[GLOBAL_CUSTOMER_SEARCH_ROUTE_KEY] = route


def render_global_customer_search():
    """どの機能ページからでも使える、入力欄常設型の顧客検索。"""
    clear_global_customer_search_when_route_changes()

    if st_keyup is not None:
        keyword = str(
            st_keyup(
                "🔍 顧客名・ひらがな・地域で検索",
                value="",
                placeholder="例：三谷、みたに、帯広、帯広 牧場",
                debounce=250,
                key=GLOBAL_CUSTOMER_SEARCH_LIVE_KEY,
            )
            or ""
        ).strip()
    else:
        keyword = st.text_input(
            "🔍 顧客名・ひらがな・地域で検索",
            value="",
            placeholder="例：三谷、みたに、帯広、帯広 牧場",
            key=GLOBAL_CUSTOMER_SEARCH_INPUT_KEY,
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        ).strip()

    if not keyword:
        return

    try:
        with st.spinner("顧客を検索しています…"):
            customers = find_global_customer_search_candidates(
                load_global_customer_search_index(),
                keyword,
            )
    except Exception as exc:
        st.warning(f"顧客検索を読み込めませんでした：{exc}")
        return

    if customers.empty:
        st.warning("該当する顧客がありません。")
        return

    total = len(customers)
    if total > GLOBAL_CUSTOMER_SEARCH_RESULT_LIMIT:
        st.caption(
            f"候補：{total}件（先頭{GLOBAL_CUSTOMER_SEARCH_RESULT_LIMIT}件を表示）"
        )
    else:
        st.caption(f"候補：{total}件")
    render_customer_search_candidate_cards(
        customers,
        limit=GLOBAL_CUSTOMER_SEARCH_RESULT_LIMIT,
    )


# =========================
# 顧客検索
# =========================
def show_customer_search(df=None, show_home_link=False):
    st.subheader("🔍 顧客検索")
    if show_home_link:
        show_back_home_button("customer_back_home")
    st.caption(f"🎤 {VOICE_INPUT_HELP} 顧客名・ひらがな・地域を、スペース区切りで絞り込めます。")

    page_name = "customer" if show_home_link else "customer_home"

    default_keyword = str(get_query_value("customer_search", "")).strip()
    if st_keyup is not None:
        keyword = str(
            st_keyup(
                "顧客名・ひらがな・地域で検索",
                value=default_keyword,
                placeholder="例：三谷、みたに、帯広、帯広 牧場",
                debounce=250,
                key="customer_search_live",
            )
            or ""
        ).strip()
    else:
        # 追加部品がまだインストールされていない環境でもアプリを止めない。
        keyword = st.text_input(
            "顧客名・ひらがな・地域で検索",
            value=default_keyword,
            placeholder="例：三谷、みたに、帯広、帯広 牧場",
            key="customer_search_input",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        ).strip()

    if keyword:
        update_query_params(page=page_name, customer_search=keyword)
    else:
        update_query_params(page=page_name, customer_search=None)

    if not keyword:
        st.info("顧客名・ひらがな・地域を入力してください。")
        return

    # ログイン直後はExcelを読まず、実際に検索が始まった時だけ取得する。
    # 通常は共通検索と同じ小さな索引を再利用し、全行の整形を繰り返さない。
    if df is None:
        with st.spinner("顧客データを読み込んでいます…"):
            customers = find_global_customer_search_candidates(
                load_global_customer_search_index(),
                keyword,
            )
    else:
        customers = find_customer_search_candidates(df, keyword)

    if customers.empty:
        st.warning("該当する顧客がありません。")
        return
    line_by_customer = load_line_statuses_from_supabase()

    st.write(f"候補：{len(customers)}件")

    for i, row in customers.iterrows():
        name = clean_value(row["顧客名"])
        region = clean_value(row["地域"])
        line_connected = line_by_customer.get(row["顧客名"], False)

        with st.container(border=True):
            render_customer_name_with_line(name, line_connected)
            st.write(f"地域：{region}")

            st.markdown(
                render_page_link("この顧客を見る", page="detail", customer=name, customer_search=keyword),
                unsafe_allow_html=True,
            )


# =========================
# 商品検索
# =========================
def get_product_search_rows(df):
    """商品名・顧客名がある全行を、現在・過去を問わず商品検索用に返す。"""
    required_columns = {
        "顧客名",
        "地域",
        "商品名",
        "使用数量/日",
    }
    if not required_columns.issubset(df.columns):
        return pd.DataFrame(columns=list(required_columns))

    rows = df.copy()
    rows["_商品名検索"] = rows["商品名"].apply(
        lambda value: clean_value(value, blank_text="").strip()
    )
    if "メーカー" in rows.columns:
        rows["_メーカー検索"] = rows["メーカー"].apply(
            lambda value: clean_value(value, blank_text="").strip()
        )
    else:
        rows["_メーカー検索"] = ""
    rows["_顧客名検索"] = rows["顧客名"].apply(
        lambda value: clean_value(value, blank_text="").strip()
    )
    rows = rows[
        (rows["_商品名検索"] != "")
        & (rows["_顧客名検索"] != "")
    ].copy()
    return rows


@st.cache_data(ttl=300, show_spinner=False)
def load_product_search_index():
    """商品検索用に必要な行だけを整形し、5分間再利用する。"""
    return get_product_search_rows(load_data())


def get_product_search_candidates(product_rows, keyword):
    """商品名またはメーカー名に入力文字を含む商品を候補タブ用に返す。"""
    keyword = str(keyword or "").strip()
    if product_rows.empty or not keyword:
        return []

    product_matches = product_rows["_商品名検索"].str.contains(
        keyword,
        case=False,
        na=False,
        regex=False,
    )
    manufacturer_matches = (
        product_rows["_メーカー検索"].str.contains(
            keyword,
            case=False,
            na=False,
            regex=False,
        )
        & (~product_rows["使用数量/日"].apply(is_blank_or_zero))
    )
    matches = product_rows[product_matches | manufacturer_matches]
    candidates = matches["_商品名検索"].drop_duplicates().tolist()
    keyword_folded = keyword.casefold()
    return sorted(
        candidates,
        key=lambda product_name: (
            product_name.casefold() != keyword_folded,
            not product_name.casefold().startswith(keyword_folded),
            product_name.casefold().find(keyword_folded),
            len(product_name),
            product_name,
        ),
    )


def build_exact_product_search_results(product_rows, product_name, keyword=""):
    """商品名検索は従来どおり、メーカー検索は現在使用中の一致行だけを返す。"""
    exact = product_rows[product_rows["_商品名検索"] == product_name].copy()

    keyword_text = str(keyword or "").strip()
    product_name_matches_keyword = (
        bool(keyword_text)
        and keyword_text.casefold() in str(product_name or "").casefold()
    )
    if keyword_text and not product_name_matches_keyword:
        exact = exact[
            exact["_メーカー検索"].str.contains(
                keyword_text,
                case=False,
                na=False,
                regex=False,
            )
            & (~exact["使用数量/日"].apply(is_blank_or_zero))
        ].copy()

    if exact.empty:
        return [], []

    current_results = []
    past_results = []

    for customer_name, group in exact.groupby("_顧客名検索", sort=False):
        group = group.copy()
        active_group = group[
            ~group["使用数量/日"].apply(is_blank_or_zero)
        ].copy()

        region_values = [
            clean_value(value, blank_text="").strip()
            for value in group["地域"].tolist()
        ]
        region = next((value for value in region_values if value), "未設定")

        if not active_group.empty:
            duplicate_count = len(active_group)
            if duplicate_count == 1:
                active_row = active_group.iloc[0]
                usage_text = format_number(active_row["使用数量/日"])
                manufacturer = clean_value(
                    active_row.get("メーカー"),
                    blank_text="未設定",
                )
                next_delivery = format_date(active_row.get("次回配達予定"))
            else:
                usage_text = "複数行（確認が必要）"
                manufacturer = "複数行（確認が必要）"
                next_delivery = "複数行（確認が必要）"

            current_results.append(
                {
                    "顧客名": customer_name,
                    "地域": region,
                    "メーカー": manufacturer,
                    "使用数量/日": usage_text,
                    "次回配達予定": next_delivery,
                    "重複件数": duplicate_count,
                }
            )
        else:
            past_row = group.iloc[0]
            past_results.append(
                {
                    "顧客名": customer_name,
                    "地域": region,
                    "メーカー": clean_value(
                        past_row.get("メーカー"),
                        blank_text="未設定",
                    ),
                    "次回配達予定": format_date(
                        past_row.get("次回配達予定")
                    ),
                }
            )

    # 商品名・メーカー名のどちらで検索した場合も、現在使用中は
    # 次回配達予定の早い順に表示する。日付なし・判定不能は最後、同日は顧客名順。
    current_results.sort(
        key=lambda item: (
            to_date(item.get("次回配達予定")) is None,
            to_date(item.get("次回配達予定")) or date.max,
            item["顧客名"],
        )
    )
    past_results.sort(key=lambda item: item["顧客名"])
    return current_results, past_results


def render_product_search_customer(item, keyword, current):
    """商品検索の顧客カードを、現在使用中・過去使用の共通形式で表示する。"""
    customer_name = item["顧客名"]
    with st.container(border=True):
        st.markdown(
            render_page_link(
                f"👤 {customer_name}",
                page="detail",
                customer=customer_name,
                product_search=keyword,
            ),
            unsafe_allow_html=True,
        )

        st.caption("地域")
        st.markdown(f"**{html.escape(item['地域'])}**")

        st.caption("メーカー")
        st.markdown(f"**{html.escape(item['メーカー'])}**")

        if current:
            st.caption("使用数量/日")
            st.markdown(f"**{html.escape(item['使用数量/日'])}**")

        st.caption("次回配達予定")
        st.markdown(f"**{html.escape(item['次回配達予定'])}**")

        if current:
            if item["重複件数"] > 1:
                st.warning(
                    "同じ顧客名・商品名の使用中行が複数見つかりました。"
                    "顧客詳細で確認してください。"
                )
        else:
            st.caption("この商品を過去に使用")


def render_product_search_customer_grid(items, keyword, current):
    """商品検索の顧客カードを、1段につき2件ずつ表示する。"""
    for start in range(0, len(items), 2):
        columns = st.columns(2)
        for column, item in zip(columns, items[start:start + 2]):
            with column:
                render_product_search_customer(item, keyword, current)


def render_product_search_results(product_rows, product_name, keyword):
    """選択した商品の顧客を、現在使用中と過去使用に分けて表示する。"""
    current_results, past_results = build_exact_product_search_results(
        product_rows,
        product_name,
        keyword,
    )

    st.markdown(
        f"**現在使用中 {len(current_results)}件 ／ "
        f"過去に使用 {len(past_results)}件**"
    )

    st.markdown(f"### 🟢 現在使用中　{len(current_results)}件")
    if not current_results:
        st.info("この商品を現在使用している顧客はいません。")
    else:
        render_product_search_customer_grid(
            current_results,
            keyword,
            current=True,
        )

    with st.expander(
        f"⚪ 過去に使用　{len(past_results)}件",
        expanded=False,
    ):
        if not past_results:
            st.info("この商品を過去に使用した顧客はいません。")
        else:
            render_product_search_customer_grid(
                past_results,
                keyword,
                current=False,
            )


def show_product_search(df=None):
    st.subheader("🔎 商品検索")
    show_back_home_button("product_back_home")
    st.caption(
        "商品名またはメーカー名の一部を入力し、候補タブを選んでください。"
        "次回配達日や残数には関係なく、現在使用中の顧客と過去に使用した顧客を分けて表示します。"
    )

    default_keyword = str(get_query_value("product_search", "")).strip()
    if st_keyup is not None:
        keyword = str(
            st_keyup(
                "商品名・メーカーで検索",
                value=default_keyword,
                placeholder="例：酒、日清丸紅",
                debounce=250,
                key="product_search_live",
            )
            or ""
        ).strip()
    else:
        keyword = st.text_input(
            "商品名・メーカーで検索",
            value=default_keyword,
            placeholder="例：酒、日清丸紅",
            key="product_search_input",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        ).strip()

    if keyword:
        update_query_params(page="product", product_search=keyword)
    else:
        update_query_params(page="product", product_search=None)

    if not keyword:
        st.info("商品名またはメーカー名を入力してください。")
        return

    if df is None:
        with st.spinner("商品データを読み込んでいます…"):
            product_rows = load_product_search_index()
    else:
        product_rows = get_product_search_rows(df)
    candidates = get_product_search_candidates(product_rows, keyword)

    if not candidates:
        st.warning("該当する商品がありません。")
        return

    st.write(f"商品候補：{len(candidates)}件")
    tabs = st.tabs([f"📦 {product_name}" for product_name in candidates])
    for product_name, tab in zip(candidates, tabs):
        with tab:
            st.markdown(f"#### {html.escape(product_name)}")
            render_product_search_results(
                product_rows,
                product_name,
                keyword,
            )


# =========================
# 地域検索
# =========================
@st.cache_data(ttl=300, show_spinner=False)
def load_region_search_index():
    """地域検索に必要な顧客名・地域だけを整形し、5分間再利用する。"""
    df = load_data()
    required = ["顧客名", "地域"]
    if not isinstance(df, pd.DataFrame) or not set(required).issubset(df.columns):
        return pd.DataFrame(columns=required + ["_地域検索"])
    index = df[required].copy()
    index["顧客名"] = index["顧客名"].apply(
        lambda value: clean_value(value, blank_text="").strip()
    )
    index["地域"] = index["地域"].apply(
        lambda value: clean_value(value, blank_text="").strip()
    )
    index = index[index["顧客名"] != ""].drop_duplicates()
    index["_地域検索"] = index["地域"].str.casefold()
    return index.reset_index(drop=True)


def show_region_search(df=None):
    st.subheader("📍 地域検索")
    show_back_home_button("region_back_home")
    st.caption(f"🎤 {VOICE_INPUT_HELP}")

    default_keyword = str(get_query_value("region_search", "")).strip()
    keyword = st.text_input(
        "地域名で検索",
        value=default_keyword,
        placeholder="例：帯広、芽室、釧路",
        key="region_search_input",
        help=VOICE_INPUT_HELP,
        autocomplete="off",
    ).strip()

    if keyword:
        update_query_params(page="region", region_search=keyword)
    else:
        update_query_params(page="region", region_search=None)

    if not keyword:
        st.info("地域名を入力してください。")
        return

    if df is None:
        search_index = load_region_search_index()
    else:
        required = ["顧客名", "地域"]
        if not isinstance(df, pd.DataFrame) or not set(required).issubset(df.columns):
            search_index = pd.DataFrame(columns=required + ["_地域検索"])
        else:
            search_index = df[required].copy()
            search_index["顧客名"] = search_index["顧客名"].apply(
                lambda value: clean_value(value, blank_text="").strip()
            )
            search_index["地域"] = search_index["地域"].apply(
                lambda value: clean_value(value, blank_text="").strip()
            )
            search_index = search_index[search_index["顧客名"] != ""].drop_duplicates()
            search_index["_地域検索"] = search_index["地域"].str.casefold()

    target = keyword.casefold()
    hit = search_index[
        search_index["_地域検索"].str.contains(target, na=False, regex=False)
    ]

    if hit.empty:
        st.warning("該当する地域の顧客が見つかりません。")
        return

    customers = (
        hit[["顧客名", "地域"]]
        .drop_duplicates()
        .sort_values(["地域", "顧客名"])
        .reset_index(drop=True)
    )

    st.write(f"候補：{len(customers)}件")

    for i, row in customers.iterrows():
        name = clean_value(row["顧客名"])
        region = clean_value(row["地域"])

        with st.container(border=True):
            st.markdown(f"### 👤 {name}")
            st.write(f"地域：{region}")

            st.markdown(
                render_page_link("この顧客を見る", page="detail", customer=name, region_search=keyword),
                unsafe_allow_html=True,
            )


# =========================
# 配車カレンダー
# =========================
DISPATCH_COLUMN_CANDIDATES = {
    "date": ["次回配達予定", "配達予定日", "配送予定日", "配達日", "配送日", "納品日", "予定日", "日付"],
    "customer": ["顧客名", "牧場名", "取引先名", "得意先名", "お客様名", "名前", "名称"],
    "region": ["地域", "地区", "エリア", "住所", "市町村"],
    "product": ["商品名", "商品", "品名", "製品名"],
    "maker": ["メーカー", "製造元", "製造メーカー"],
}

DISPATCH_REQUIRED_LABELS = {
    "date": "日付（例：次回配達予定）",
    "customer": "顧客名・牧場名",
    "region": "地域",
    "product": "商品名",
}

WEEKDAYS_JA = ["月", "火", "水", "木", "金", "土", "日"]


def find_dispatch_columns(df):
    """配車カレンダーに必要な列を、候補名から探す"""
    return {
        key: find_existing_column(df, candidates)
        for key, candidates in DISPATCH_COLUMN_CANDIDATES.items()
    }


def show_missing_dispatch_columns_error(df, dispatch_columns):
    missing = [
        label
        for key, label in DISPATCH_REQUIRED_LABELS.items()
        if not dispatch_columns.get(key)
    ]

    if not missing:
        return False

    st.error("配車カレンダーに必要な列が見つかりません。")
    st.write("見つからない項目：", missing)
    st.write("次のような列名が使えます。")
    st.code(
        "日付: 次回配達予定 / 配達予定日 / 日付\n"
        "顧客: 顧客名 / 牧場名 / 取引先名\n"
        "地域: 地域 / 地区 / エリア\n"
        "商品: 商品名 / 商品 / 品名"
    )
    st.write("Excelから読み取れた列：")
    st.write(list(df.columns))
    return True


def get_default_dispatch_date(df, date_column):
    """予定がある日付のうち、今日以降で一番近い日を初期表示にする"""
    parsed = pd.to_datetime(df[date_column], errors="coerce").dropna()
    if parsed.empty:
        return date.today()

    available_dates = sorted(set(parsed.dt.date))
    today = date.today()

    for target_date in available_dates:
        if target_date >= today:
            return target_date

    return available_dates[-1]


def get_calendar_month_start(df, date_column):
    """表示中の月をsession_stateで保持する"""
    if "dispatch_calendar_year" not in st.session_state or "dispatch_calendar_month" not in st.session_state:
        default_date = get_default_dispatch_date(df, date_column)
        st.session_state["dispatch_calendar_year"] = default_date.year
        st.session_state["dispatch_calendar_month"] = default_date.month

    return date(
        st.session_state["dispatch_calendar_year"],
        st.session_state["dispatch_calendar_month"],
        1,
    )


def change_dispatch_month(delta, scroll_target=None):
    current = date(
        st.session_state["dispatch_calendar_year"],
        st.session_state["dispatch_calendar_month"],
        1,
    )
    month = current.month + delta
    year = current.year

    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    st.session_state["dispatch_calendar_year"] = year
    st.session_state["dispatch_calendar_month"] = month
    if scroll_target in {"top", "bottom"}:
        st.session_state["dispatch_calendar_month_scroll_target"] = scroll_target


def render_month_navigation_anchor(anchor_id):
    """月移動後のスクロール先として使う非表示アンカーを置く。"""
    safe_anchor_id = re.sub(r"[^0-9A-Za-z_-]", "", str(anchor_id or ""))
    if not safe_anchor_id:
        return
    st.markdown(
        f'<div id="{safe_anchor_id}" style="height:0;scroll-margin-top:88px;"></div>',
        unsafe_allow_html=True,
    )


def restore_month_navigation_scroll(
    scroll_target,
    top_target_selector,
    bottom_target_selector,
    fallback_anchor_id=None,
):
    """下側の月移動後に、実際の先頭行または最終行を画面へ位置合わせする。"""
    if scroll_target not in {"top", "bottom"}:
        return

    safe_fallback_anchor_id = re.sub(
        r"[^0-9A-Za-z_-]",
        "",
        str(fallback_anchor_id or ""),
    )
    target_json = json.dumps(scroll_target)
    top_selector_json = json.dumps(str(top_target_selector or ""))
    bottom_selector_json = json.dumps(str(bottom_target_selector or ""))
    fallback_anchor_json = json.dumps(safe_fallback_anchor_id)
    script = f'''<script>
    (() => {{
      const scrollTarget = {target_json};
      const topSelector = {top_selector_json};
      const bottomSelector = {bottom_selector_json};
      const fallbackAnchorId = {fallback_anchor_json};
      let parentWindow;
      let parentDocument;
      try {{
        parentWindow = window.parent;
        parentDocument = parentWindow.document;
      }} catch (_) {{
        return;
      }}

      let cancelled = false;
      const cancelAutomaticMove = () => {{ cancelled = true; }};
      ['wheel', 'touchstart', 'pointerdown', 'keydown'].forEach((eventName) => {{
        parentWindow.addEventListener(eventName, cancelAutomaticMove, {{once: true, passive: true}});
      }});

      const visibleTarget = (selector) => {{
        if (!selector) return null;
        const candidates = Array.from(parentDocument.querySelectorAll(selector));
        for (const candidate of candidates) {{
          const style = parentWindow.getComputedStyle(candidate);
          if (
            style.display !== 'none' &&
            style.visibility !== 'hidden' &&
            candidate.getClientRects().length > 0
          ) {{
            return candidate;
          }}
        }}
        return null;
      }};

      const verticalScrollParent = (element) => {{
        let current = element ? element.parentElement : null;
        while (current && current !== parentDocument.body && current !== parentDocument.documentElement) {{
          const style = parentWindow.getComputedStyle(current);
          const overflowY = style.overflowY;
          if (
            (overflowY === 'auto' || overflowY === 'scroll') &&
            current.scrollHeight > current.clientHeight + 2
          ) {{
            return current;
          }}
          current = current.parentElement;
        }}
        return null;
      }};

      const move = () => {{
        if (cancelled) return;

        const selector = scrollTarget === 'top' ? topSelector : bottomSelector;
        let target = visibleTarget(selector);
        if (!target && fallbackAnchorId) {{
          target = parentDocument.getElementById(fallbackAnchorId);
        }}
        if (!target) return;

        const innerScroller = verticalScrollParent(target);
        if (innerScroller) {{
          if (scrollTarget === 'top') {{
            innerScroller.scrollTop = 0;
          }} else {{
            innerScroller.scrollTop = Math.max(
              0,
              target.offsetTop + target.offsetHeight - innerScroller.clientHeight,
            );
          }}
        }}

        parentWindow.requestAnimationFrame(() => {{
          if (cancelled) return;
          const rect = target.getBoundingClientRect();
          const topOffset = 82;
          const bottomOffset = 12;
          const destination = scrollTarget === 'top'
            ? parentWindow.scrollY + rect.top - topOffset
            : parentWindow.scrollY + rect.bottom - parentWindow.innerHeight + bottomOffset;
          parentWindow.scrollTo({{
            top: Math.max(0, destination),
            left: parentWindow.scrollX || 0,
            behavior: 'auto',
          }});
        }});
      }};

      [60, 220, 520, 950, 1500, 2300, 3200].forEach((delay) => {{
        window.setTimeout(move, delay);
      }});
    }})();
    </script>'''
    components.html(script, height=0, scrolling=False)


def clean_dispatch_maker(value):
    """カレンダーに表示するメーカー名を整える。空白と数値の0は表示しない。"""
    maker = clean_value(value, blank_text="").strip()
    if not maker:
        return ""

    try:
        if float(maker.replace(",", "")) == 0:
            return ""
    except ValueError:
        pass

    return maker


def format_month_day(target_day):
    weekday = WEEKDAYS_JA[target_day.weekday()]
    return f"{target_day.month}/{target_day.day}（{weekday}）"


def make_dispatch_items_by_day(df, month_start, dispatch_columns):
    last_day = calendar.monthrange(month_start.year, month_start.month)[1]
    start_day = date(month_start.year, month_start.month, 1)
    end_day = date(month_start.year, month_start.month, last_day)

    rows_by_day = {
        date(month_start.year, month_start.month, day_num): []
        for day_num in range(1, last_day + 1)
    }

    date_column = dispatch_columns["date"]
    customer_column = dispatch_columns["customer"]
    region_column = dispatch_columns["region"]
    product_column = dispatch_columns["product"]
    maker_column = dispatch_columns.get("maker")
    parsed_dates = pd.to_datetime(df[date_column], errors="coerce").dt.date

    for idx, row in df.iterrows():
        delivery_date = parsed_dates.loc[idx]

        if pd.isna(delivery_date) or not (start_day <= delivery_date <= end_day):
            continue

        item = {
            "顧客名": clean_value(row[customer_column]),
            "地域": clean_value(row[region_column]),
            "商品名": clean_value(row[product_column]),
            "メーカー": clean_dispatch_maker(row[maker_column]) if maker_column else "",
        }
        rows_by_day[delivery_date].append(item)

    for delivery_date, items in rows_by_day.items():
        rows_by_day[delivery_date] = sorted(
            items,
            key=lambda item: (item["顧客名"], item["地域"], item["商品名"]),
        )

    return rows_by_day


def inject_dispatch_calendar_css():
    st.markdown(
        """
        <style>
        .dispatch-month-title {
            text-align: center;
            font-size: 1.2rem;
            font-weight: 700;
            padding-top: 0.35rem;
        }
        .dispatch-two-day-row {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
            gap: 0.75rem;
            margin: 0.75rem 0 1.25rem;
        }
        .dispatch-day-panel {
            border: 1px solid rgba(49, 51, 63, 0.18);
            border-radius: 16px;
            color: #111827 !important;
            overflow: visible;
            padding: 0.8rem;
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            box-shadow: 0 8px 22px rgba(15, 23, 42, 0.07);
            min-width: 0;
        }
        .dispatch-day-title {
            display: block;
            color: #111827 !important;
            background: linear-gradient(135deg, #eff6ff 0%, #ecfdf5 100%);
            border-radius: 12px;
            font-size: 0.95rem;
            font-weight: 700;
            line-height: 1.35;
            margin-bottom: 0.55rem;
            overflow-wrap: anywhere;
            padding: 0.25rem 0.4rem;
            text-align: center;
            white-space: normal;
        }
        .dispatch-item {
            border-top: 1px solid rgba(49, 51, 63, 0.12);
            display: block;
            overflow: visible;
            padding: 0.55rem 0;
        }
        .dispatch-item:first-of-type {
            border-top: 0;
            padding-top: 0;
        }
        .dispatch-name {
            color: #111827 !important;
            display: block;
            font-size: 0.95rem;
            font-weight: 700;
            line-height: 1.35;
            overflow-wrap: anywhere;
            white-space: normal;
            word-break: normal;
        }
        .dispatch-name a,
        .dispatch-month-link {
            color: #2563eb !important;
            font-weight: 700;
            text-decoration: none;
        }
        .dispatch-name a:hover,
        .dispatch-month-link:hover {
            text-decoration: underline;
        }
        .dispatch-month-product {
            color: #374151 !important;
            font-size: 0.82rem;
            white-space: normal;
        }
        .dispatch-line,
        .dispatch-empty {
            color: #374151 !important;
            display: block;
            font-size: 0.9rem;
            line-height: 1.45;
            overflow-wrap: anywhere;
            white-space: normal;
            word-break: normal;
        }
        .dispatch-month-scroll {
            width: 100%;
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
            border: 1px solid rgba(15, 23, 42, 0.10);
            border-radius: 16px;
            background: #ffffff;
            box-shadow: 0 10px 26px rgba(15, 23, 42, 0.08);
        }
        .dispatch-month-table {
            border-collapse: collapse;
            min-width: 760px;
            width: max-content;
            color: #111827 !important;
            table-layout: auto;
        }
        .dispatch-month-table th,
        .dispatch-month-table td {
            border-bottom: 1px solid rgba(49, 51, 63, 0.12);
            border-right: 1px solid rgba(49, 51, 63, 0.08);
            padding: 0.45rem 0.6rem;
            text-align: left;
            vertical-align: top;
            white-space: nowrap;
            min-width: 130px;
            font-size: 0.9rem;
            position: static !important;
        }
        .dispatch-month-table th:first-child,
        .dispatch-month-table td:first-child {
            min-width: 86px;
            position: sticky !important;
            left: 0;
            z-index: 3;
            background: #ffffff;
        }
        .dispatch-month-table th:first-child {
            z-index: 4;
            background: #f3f4f6;
        }
        .dispatch-month-table th {
            background: #eff6ff;
            font-weight: 800;
        }
        @media (max-width: 420px) {
            .dispatch-two-day-row {
                gap: 0.45rem;
            }
            .dispatch-day-panel {
                padding: 0.6rem;
            }
            .dispatch-name,
            .dispatch-line,
            .dispatch-empty {
                font-size: 0.85rem;
            }
            .dispatch-day-title {
                font-size: 0.82rem;
                padding-left: 0.25rem;
                padding-right: 0.25rem;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def escape_html(value):
    return html.escape(clean_value(value), quote=True)


def build_customer_detail_link(customer_name, label=None, class_name="dispatch-month-link"):
    """配車カレンダーから顧客詳細へ移動するリンクを作る"""
    customer = clean_value(customer_name, blank_text="").strip()

    if not customer:
        return escape_html(label or customer_name)

    link_label = label or customer
    url = make_app_url(page="detail", customer=customer)
    class_name = f"{class_name} entity-select-card-link".strip()
    return f'<a class="{class_name}" href="{url}" target="_self">{escape_html(link_label)}</a>'


def handle_customer_query_param():
    """旧リンク互換用。URLの顧客名を消さず、ブラウザ戻るに使えるよう保持する。"""
    sync_page_from_query_params()


def build_two_day_panel_html(target_day, items):
    parts = [
        '<div class="dispatch-day-panel">',
        f'<div class="dispatch-day-title">{html.escape(format_month_day(target_day))}</div>',
    ]

    if not items:
        parts.append('<div class="dispatch-empty">予定なし</div>')
    else:
        for item in items:
            customer_link = build_customer_detail_link(item.get("顧客名"), class_name="dispatch-month-link")
            parts.append('<div class="dispatch-item">')
            parts.append(f'<div class="dispatch-name">👤 {customer_link}</div>')
            parts.append(f'<div class="dispatch-line">地域：{escape_html(item.get("地域"))}</div>')
            parts.append(f'<div class="dispatch-line">商品：{escape_html(item.get("商品名"))}</div>')
            maker = clean_dispatch_maker(item.get("メーカー"))
            if maker:
                parts.append(f'<div class="dispatch-line">メーカー：{escape_html(maker)}</div>')
            parts.append('</div>')

    parts.append('</div>')
    return "".join(parts)


def show_dispatch_month_switcher(month_start, key_suffix="top"):
    col_prev, col_month, col_next = st.columns([1, 2, 1])
    is_bottom = key_suffix == "bottom"
    prev_args = (-1, "bottom") if is_bottom else (-1,)
    next_args = (1, "top") if is_bottom else (1,)

    with col_prev:
        st.button(
            "◀",
            key=f"dispatch_prev_month_{key_suffix}",
            use_container_width=True,
            on_click=change_dispatch_month,
            args=prev_args,
        )

    with col_month:
        st.markdown(
            f'<div class="dispatch-month-title">{month_start.year}年{month_start.month}月</div>',
            unsafe_allow_html=True,
        )

    with col_next:
        st.button(
            "▶",
            key=f"dispatch_next_month_{key_suffix}",
            use_container_width=True,
            on_click=change_dispatch_month,
            args=next_args,
        )


def show_two_day_dispatch_calendar(rows_by_day, month_start):
    st.subheader("📱 2日表示")
    st.caption("スマホでも2日分を横並びで表示します。")

    last_day = calendar.monthrange(month_start.year, month_start.month)[1]

    for day_num in range(1, last_day + 1, 2):
        day1 = date(month_start.year, month_start.month, day_num)
        day2 = date(month_start.year, month_start.month, day_num + 1) if day_num + 1 <= last_day else None

        left_panel = build_two_day_panel_html(day1, rows_by_day.get(day1, []))
        right_panel = build_two_day_panel_html(day2, rows_by_day.get(day2, [])) if day2 else '<div></div>'
        row_classes = ["dispatch-two-day-row"]
        if day_num == 1:
            row_classes.append("dispatch-calendar-scroll-top")
        if day_num + 1 >= last_day:
            row_classes.append("dispatch-calendar-scroll-bottom")

        st.markdown(
            f'<div class="{" ".join(row_classes)}">{left_panel}{right_panel}</div>',
            unsafe_allow_html=True,
        )

def format_month_cell_item(item):
    customer_name = clean_value(item.get("顧客名"))
    product_name = clean_value(item.get("商品名"), blank_text="").strip()
    maker = clean_dispatch_maker(item.get("メーカー"))
    customer_link = build_customer_detail_link(customer_name, class_name="dispatch-month-link")

    if not product_name and not maker:
        return customer_link

    product_label = f"{product_name}/{maker}" if product_name and maker else product_name or maker
    return f'{customer_link}<br><span class="dispatch-month-product">{escape_html(product_label)}</span>'


def make_month_dispatch_table(rows_by_day, month_start):
    last_day = calendar.monthrange(month_start.year, month_start.month)[1]
    max_count = max((len(items) for items in rows_by_day.values()), default=0)
    farm_column_count = max(5, max_count)

    table_rows = []

    for day_num in range(1, last_day + 1):
        target_day = date(month_start.year, month_start.month, day_num)
        items = rows_by_day.get(target_day, [])

        row_data = {"月/日": format_month_day(target_day)}

        for item_index in range(farm_column_count):
            column_name = f"牧場名{item_index + 1}"
            row_data[column_name] = format_month_cell_item(items[item_index]) if item_index < len(items) else ""

        table_rows.append(row_data)

    return pd.DataFrame(table_rows)


def show_month_dispatch_calendar(rows_by_day, month_start):
    st.subheader("🗓 月表示")
    st.caption("横スクロールで1か月分を確認できます。日付列は固定表示します。")

    month_df = make_month_dispatch_table(rows_by_day, month_start)

    header_cells = "".join(
        f'<th>{html.escape(str(column))}</th>'
        for column in month_df.columns
    )

    body_rows = []
    last_row_index = len(month_df) - 1
    for row_index, (_, row) in enumerate(month_df.iterrows()):
        row_cells = []
        for column in month_df.columns:
            value = row[column]
            if str(value) == "nan":
                cell_value = ""
            elif column == "月/日":
                cell_value = html.escape(str(value))
            else:
                cell_value = str(value)
            row_cells.append(f"<td>{cell_value}</td>")
        cells = "".join(row_cells)
        row_classes = []
        if row_index == 0:
            row_classes.append("dispatch-calendar-scroll-top")
        if row_index == last_row_index:
            row_classes.append("dispatch-calendar-scroll-bottom")
        class_attribute = f' class="{" ".join(row_classes)}"' if row_classes else ""
        body_rows.append(f"<tr{class_attribute}>{cells}</tr>")

    table_html = f"""
    <div class="dispatch-month-scroll">
      <table class="dispatch-month-table">
        <thead><tr>{header_cells}</tr></thead>
        <tbody>{''.join(body_rows)}</tbody>
      </table>
    </div>
    """

    st.markdown(table_html, unsafe_allow_html=True)

def show_dispatch_calendar(df):
    st.markdown("---")
    st.header("🗓 配車カレンダー")
    show_back_home_button("calendar_back_home")
    month_scroll_target = st.session_state.pop(
        "dispatch_calendar_month_scroll_target",
        None,
    )

    inject_dispatch_calendar_css()

    if df.empty:
        st.warning("Excelから読み込めるデータがありません。")
        return

    dispatch_columns = find_dispatch_columns(df)

    if show_missing_dispatch_columns_error(df, dispatch_columns):
        return

    month_start = get_calendar_month_start(df, dispatch_columns["date"])
    show_dispatch_month_switcher(month_start, key_suffix="top")
    rows_by_day = make_dispatch_items_by_day(df, month_start, dispatch_columns)

    st.caption(
        f"使用している日付列：{dispatch_columns['date']} / "
        f"顧客名：{dispatch_columns['customer']} / "
        f"地域：{dispatch_columns['region']} / "
        f"商品名：{dispatch_columns['product']}"
    )

    total_count = sum(len(items) for items in rows_by_day.values())
    st.write(f"{month_start.month}月の予定：{total_count}件")

    view = st.radio(
        "表示切替",
        ["📱 2日表示", "🗓 月表示"],
        horizontal=True,
        key="dispatch_calendar_view",
    )

    render_month_navigation_anchor("dispatch_calendar_month_content_top")
    if view == "📱 2日表示":
        show_two_day_dispatch_calendar(rows_by_day, month_start)
    else:
        show_month_dispatch_calendar(rows_by_day, month_start)

    st.markdown("---")
    show_dispatch_month_switcher(month_start, key_suffix="bottom")
    restore_month_navigation_scroll(
        month_scroll_target,
        ".dispatch-calendar-scroll-top",
        ".dispatch-calendar-scroll-bottom",
        fallback_anchor_id="dispatch_calendar_month_content_top",
    )


# =========================
# 配車表（配車表1.xlsm・1月～12月）
# =========================

# 配車表Excelの読み込み・整形は、分割した処理を使用する。
from app_modules.dispatch_excel import (
    normalize_dispatch_text,
    read_dispatch_month_sheets,
)


def read_dispatch_selected_month_sheet(excel_source, selected_month):
    """
    既存の配車表読み込みルールをそのまま使い、表示対象月の明細行だけを読む。

    12か月すべてのシート存在確認とA～H列の見出し確認は従来どおり行う。
    違いは、2行目以降の明細走査を selected_month の1シートだけに限定する点だけ。
    """
    selected_month = str(selected_month or "").strip()
    if selected_month not in DISPATCH_MONTH_SHEETS:
        raise ValueError(f"表示月が正しくありません：{selected_month}")

    if isinstance(excel_source, BytesIO):
        source = BytesIO(excel_source.getvalue())
    else:
        source = excel_source

    workbook = load_workbook(source, read_only=True, data_only=True)
    rows = []
    try:
        # 従来どおり、1月～12月がすべて存在することを確認する。
        missing_sheets = [
            name for name in DISPATCH_MONTH_SHEETS
            if name not in workbook.sheetnames
        ]
        if missing_sheets:
            raise ValueError("月別シートが見つかりません：" + "、".join(missing_sheets))

        # 従来どおり、全12シートのA～H列見出しを確認する。
        # 明細行はここでは読まないため、この確認は軽い。
        for sheet_name in DISPATCH_MONTH_SHEETS:
            ws = workbook[sheet_name]
            headers = [
                normalize_dispatch_text(ws.cell(1, column).value)
                for column in range(1, 9)
            ]
            if headers != DISPATCH_REQUIRED_COLUMNS:
                raise ValueError(
                    f"{sheet_name}のA～H列の見出しが想定と異なります。\n"
                    f"読み取った見出し：{' / '.join(headers)}"
                )

        # 明細行だけ、現在表示している月に限定して読む。
        ws = workbook[selected_month]
        for values in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            if not any(
                value is not None and normalize_dispatch_text(value)
                for value in values
            ):
                continue

            record = dict(zip(DISPATCH_REQUIRED_COLUMNS, values))
            record["参照シート"] = selected_month
            rows.append(record)
    finally:
        workbook.close()

    df = pd.DataFrame(
        rows,
        columns=DISPATCH_REQUIRED_COLUMNS + ["参照シート"],
    )
    if df.empty:
        return df

    # 以下の整形は従来の read_dispatch_month_sheets と同一。
    for column in ["引取先", "商品名", "数量", "運送会社", "納品先"]:
        df[column] = df[column].map(normalize_dispatch_text)

    pickup_dates = pd.to_datetime(df["引取日"], errors="coerce")
    arrival_dates = pd.to_datetime(df["着日"], errors="coerce")
    df["_引取日"] = pickup_dates.map(
        lambda value: value.date() if pd.notna(value) else None
    )
    df["_着日"] = arrival_dates.map(
        lambda value: value.date() if pd.notna(value) else None
    )
    return df


@st.cache_data(ttl=60, show_spinner=False)
def get_cached_dispatch_dropbox_content():
    access_token = get_dropbox_access_token()
    content, response = download_dropbox_team_file(
        str(DISPATCH_DROPBOX_FILE_PATH or DISPATCH_DROPBOX_DEFAULT_FILE_PATH).strip(),
        access_token,
    )
    if content is None:
        raise RuntimeError("配車表1.xlsmをDropboxから取得できませんでした。\n" + dropbox_error_text(response))
    return content


@st.cache_data(ttl=300, show_spinner=False)
def load_dispatch_board_data():
    """本番はDropbox、設定がない場合は指定されたローカルファイルから読む。"""
    dropbox_error = None
    if has_dropbox_auth_config():
        try:
            return read_dispatch_month_sheets(BytesIO(get_cached_dispatch_dropbox_content()))
        except Exception as error:
            # Dropbox側に配車表フォルダがまだ共有されていないPCでは、同期済みローカル版を使う。
            dropbox_error = error

    local_path = Path(str(DISPATCH_LOCAL_FILE or "").strip())
    if not local_path.exists():
        message = f"配車表1.xlsmが見つかりません：{local_path}"
        if dropbox_error is not None:
            message += f"\nDropbox取得エラー：{dropbox_error}"
        raise FileNotFoundError(message)
    return read_dispatch_month_sheets(local_path)


@st.cache_data(ttl=300, show_spinner=False)
def load_dispatch_board_month_data(selected_month):
    """
    配車表画面専用。表示中の月だけを読み、失敗した場合は従来の全月読み込みへ戻す。
    """
    selected_month = str(selected_month or "").strip()
    if selected_month not in DISPATCH_MONTH_SHEETS:
        return load_dispatch_board_data()

    if has_dropbox_auth_config():
        try:
            return read_dispatch_selected_month_sheet(
                BytesIO(get_cached_dispatch_dropbox_content()),
                selected_month,
            )
        except Exception:
            # 高速経路だけの問題で配車表を止めない。従来処理へ安全に切り戻す。
            full_df = load_dispatch_board_data()
            return full_df[full_df["参照シート"] == selected_month].copy()

    local_path = Path(str(DISPATCH_LOCAL_FILE or "").strip())
    if not local_path.exists():
        # エラー文やDropboxフォールバックを含め、従来関数へ任せる。
        full_df = load_dispatch_board_data()
        return full_df[full_df["参照シート"] == selected_month].copy()

    try:
        return read_dispatch_selected_month_sheet(local_path, selected_month)
    except Exception:
        full_df = load_dispatch_board_data()
        return full_df[full_df["参照シート"] == selected_month].copy()


def dispatch_date_label(value):
    target = to_date(value)
    if target is None:
        return "未入力"
    weekdays = "月火水木金土日"
    return f"{target.month}/{target.day}（{weekdays[target.weekday()]}）"


# 配車表の絞り込み処理は、分割した純粋関数を使用する。
from app_modules.dispatch_filters import (
    apply_dispatch_choice_filter,
    apply_dispatch_date_filter,
    dispatch_filter_options,
)


def dispatch_mobile_text_filter(df, column, keyword):
    """スマホ配車表の文字入力を、指定列の部分一致で絞り込む。"""
    keyword_text = normalize_dispatch_text(keyword).strip().casefold()
    if not keyword_text:
        return df
    target_text = df[column].map(normalize_dispatch_text).str.casefold()
    return df[target_text.str.contains(keyword_text, regex=False, na=False)]


def dispatch_text_filter(df, column, keyword):
    """配車表の文字項目を部分一致で絞り込む。"""
    keyword_text = normalize_dispatch_text(keyword).strip().casefold()
    if not keyword_text:
        return df
    target_text = df[column].map(normalize_dispatch_text).str.casefold()
    return df[target_text.str.contains(keyword_text, regex=False, na=False)]


def parse_dispatch_date_keyword(keyword):
    """8/1・08/01・8月1日・2026/8/1 などを日付として解釈する。"""
    text = normalize_dispatch_text(keyword).strip()
    if not text:
        return None

    normalized = (
        text.replace('年', '/')
            .replace('月', '/')
            .replace('日', '')
            .replace('-', '/')
            .replace('.', '/')
    )
    normalized = re.sub(r'/+', '/', normalized).strip('/')
    parts = [part.strip() for part in normalized.split('/') if part.strip()]

    try:
        if len(parts) == 2:
            month, day = map(int, parts)
            return ('month_day', month, day)
        if len(parts) == 3:
            year, month, day = map(int, parts)
            return ('full_date', date(year, month, day))
    except (TypeError, ValueError):
        return None
    return None


def dispatch_exact_date_filter(df, column, keyword):
    """日付は文字列の部分一致ではなく、入力された日付と完全一致で絞り込む。"""
    parsed = parse_dispatch_date_keyword(keyword)
    if parsed is None:
        if not normalize_dispatch_text(keyword).strip():
            return df
        return df.iloc[0:0]

    if parsed[0] == 'full_date':
        target_date = parsed[1]
        return df[df[column] == target_date]

    _, month, day = parsed
    return df[
        df[column].map(
            lambda value: isinstance(value, date) and value.month == month and value.day == day
        )
    ]


def show_dispatch_filters(df):
    """Excelフィルターに近いAND条件の絞り込みを表示する。"""
    st.markdown(
        """
        <style>
        div[class*="st-key-dispatch_filter_white_panel"] [data-testid="stExpander"] details,
        div[class*="st-key-dispatch_filter_white_panel"] [data-testid="stExpander"] summary,
        div[class*="st-key-dispatch_filter_white_panel"] [data-testid="stExpanderDetails"] {
            background: #ffffff !important;
            background-color: #ffffff !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    with st.container(key="dispatch_filter_white_panel"):
        with st.expander("🔎 絞り込み", expanded=False):
            pickup_date_keyword = st.text_input(
                "引取日",
                key="dispatch_filter_pickup_date_text",
                placeholder="例：8/1、8月1日、2026/8/1",
                autocomplete="off",
            ).strip()
            arrival_date_keyword = st.text_input(
                "着日",
                key="dispatch_filter_arrival_date_text",
                placeholder="例：8/1、8月1日、2026/8/1",
                autocomplete="off",
            ).strip()
            pickup_place_keyword = st.text_input(
                "引取先",
                key="dispatch_filter_pickup_place_text",
                placeholder="引取先の一部を入力",
                autocomplete="off",
            ).strip()
            product_keyword = st.text_input(
                "商品名",
                key="dispatch_filter_product_text",
                placeholder="商品名の一部を入力",
                autocomplete="off",
            ).strip()
            quantity_keyword = st.text_input(
                "数量",
                key="dispatch_filter_quantity",
                placeholder="例：450㎏、44本",
                autocomplete="off",
            ).strip()
            carrier_keyword = st.text_input(
                "運送会社",
                key="dispatch_filter_carrier_text",
                placeholder="運送会社の一部を入力",
                autocomplete="off",
            ).strip()
            destination_keyword = st.text_input(
                "納品先",
                key="dispatch_filter_destination_text",
                placeholder="納品先の一部を入力",
                autocomplete="off",
            ).strip()

            if st.button("条件をすべて解除", use_container_width=True, key="dispatch_filter_clear"):
                for key in list(st.session_state.keys()):
                    if key.startswith("dispatch_filter_"):
                        del st.session_state[key]
                st.rerun()

    filtered = dispatch_exact_date_filter(df, "_引取日", pickup_date_keyword)
    filtered = dispatch_exact_date_filter(filtered, "_着日", arrival_date_keyword)
    filtered = dispatch_text_filter(filtered, "引取先", pickup_place_keyword)
    filtered = dispatch_text_filter(filtered, "商品名", product_keyword)
    filtered = dispatch_text_filter(filtered, "運送会社", carrier_keyword)
    filtered = dispatch_text_filter(filtered, "納品先", destination_keyword)

    if quantity_keyword:
        quantity_text = filtered["数量"].map(normalize_dispatch_text)
        filtered = filtered[quantity_text.str.contains(quantity_keyword, regex=False, na=False)]
    return filtered


def render_dispatch_board_card(row):
    pickup_date = dispatch_date_label(row.get("_引取日"))
    arrival_date = dispatch_date_label(row.get("_着日"))
    pickup_place = normalize_dispatch_text(row.get("引取先")) or "未入力"
    destination = normalize_dispatch_text(row.get("納品先")) or "未入力"
    product = normalize_dispatch_text(row.get("商品名")) or "未入力"
    quantity = normalize_dispatch_text(row.get("数量")) or "未入力"
    carrier = normalize_dispatch_text(row.get("運送会社")) or "未入力"
    order_number = normalize_dispatch_text(row.get("発注番号"))

    with st.container(border=True):
        st.markdown(f"### {pickup_date} 引取 → {arrival_date} 着")
        st.write(f"**引取先：** {pickup_place}")
        st.write(f"**納品先：** {destination}")
        st.write(f"**商品名：** {product}")
        st.write(f"**数量：** {quantity}")
        st.write(f"**運送会社：** {carrier}")
        if order_number:
            st.caption(f"発注番号：{order_number}")


def show_dispatch_day_cards(df, basis_column, selected_day):
    day_rows = df[df[basis_column] == selected_day].copy()
    if day_rows.empty:
        st.info("この日の配車はありません。")
        return

    day_rows = day_rows.sort_values(
        ["_引取日", "_着日", "引取先", "納品先"],
        na_position="last",
    )
    st.subheader(f"{dispatch_date_label(selected_day)}：{len(day_rows)}件")
    for _, row in day_rows.iterrows():
        render_dispatch_board_card(row)


def show_dispatch_month_calendar(df):
    basis_label = st.radio(
        "カレンダー基準",
        ["引取日", "着日"],
        horizontal=True,
        key="dispatch_board_basis",
    )
    basis_column = "_引取日" if basis_label == "引取日" else "_着日"
    available_dates = sorted({value for value in df[basis_column] if pd.notna(value)})
    if not available_dates:
        st.info(f"{basis_label}が入力された配車はありません。")
        return

    periods = sorted({(value.year, value.month) for value in available_dates})
    today_period = (date.today().year, date.today().month)
    default_index = periods.index(today_period) if today_period in periods else len(periods) - 1
    selected_period = st.selectbox(
        "表示月",
        periods,
        index=default_index,
        format_func=lambda value: f"{value[0]}年{value[1]}月",
        key=f"dispatch_board_month_{basis_column}",
    )
    year, month = selected_period
    month_rows = df[
        df[basis_column].map(
            lambda value: pd.notna(value) and value.year == year and value.month == month
        )
    ]
    counts = month_rows.groupby(basis_column).size().to_dict()

    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
    header_columns = st.columns(7)
    for column, label in zip(header_columns, weekday_names):
        column.markdown(f"**{label}**")

    selected_key = f"dispatch_selected_day_{basis_column}"
    for week in calendar.Calendar(firstweekday=0).monthdayscalendar(year, month):
        columns = st.columns(7)
        for column, day_number in zip(columns, week):
            if day_number == 0:
                column.write("")
                continue
            target_day = date(year, month, day_number)
            count = int(counts.get(target_day, 0))
            label = f"{day_number}\n{count}件" if count else str(day_number)
            if column.button(
                label,
                key=f"dispatch_day_{basis_column}_{target_day.isoformat()}",
                use_container_width=True,
                disabled=count == 0,
            ):
                st.session_state[selected_key] = target_day

    selected_day = st.session_state.get(selected_key)
    if not isinstance(selected_day, date) or (selected_day.year, selected_day.month) != selected_period:
        selected_day = date.today() if date.today() in counts else min(counts)
        st.session_state[selected_key] = selected_day
    show_dispatch_day_cards(df, basis_column, selected_day)


def show_dispatch_filtered_list(df):
    st.subheader(f"絞り込み結果：{len(df)}件")
    if df.empty:
        st.info("条件に一致する配車はありません。")
        return

    sorted_df = df.sort_values(["_引取日", "_着日", "引取先"], na_position="last")
    for _, row in sorted_df.iterrows():
        render_dispatch_board_card(row)


def render_dispatch_responsive_list(display_df, customer_names=None):
    """PCはExcel風一覧、スマホは横スクロール不要の縦型カードで表示する。"""
    customer_names = set(customer_names or [])
    st.markdown(
        """
        <style>
        .dispatch-desktop-view {
            display: block;
            max-height: 760px;
            overflow-y: auto;
            overflow-x: hidden;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            background: #ffffff;
        }
        .dispatch-excel-table {
            width: 100%;
            min-width: 0;
            table-layout: fixed;
            border-collapse: separate;
            border-spacing: 0;
            font-size: 13px;
            color: #172033;
        }
        .dispatch-excel-table th:nth-child(1),
        .dispatch-excel-table td:nth-child(1) { width: 9%; }
        .dispatch-excel-table th:nth-child(2),
        .dispatch-excel-table td:nth-child(2) { width: 20%; }
        .dispatch-excel-table th:nth-child(3),
        .dispatch-excel-table td:nth-child(3) { width: 18%; }
        .dispatch-excel-table th:nth-child(4),
        .dispatch-excel-table td:nth-child(4) { width: 10%; }
        .dispatch-excel-table th:nth-child(5),
        .dispatch-excel-table td:nth-child(5) { width: 15%; }
        .dispatch-excel-table th:nth-child(6),
        .dispatch-excel-table td:nth-child(6) { width: 19%; }
        .dispatch-excel-table th:nth-child(7),
        .dispatch-excel-table td:nth-child(7) { width: 9%; }
        .dispatch-excel-table th {
            position: sticky;
            top: 0;
            z-index: 2;
            padding: 9px 6px;
            background: #dbeaf7;
            border-right: 1px solid #94a3b8;
            border-bottom: 2px solid #64748b;
            text-align: center;
            white-space: nowrap;
        }
        .dispatch-excel-table td {
            padding: 7px 6px;
            border-right: 1px solid #cbd5e1;
            border-bottom: 1px solid #cbd5e1;
            vertical-align: middle;
            background: #ffffff;
            overflow-wrap: anywhere;
            word-break: break-word;
        }
        .dispatch-excel-table tr:nth-child(even) td { background: #f8fafc; }
        .dispatch-excel-table .date-cell,
        .dispatch-excel-table .quantity-cell { text-align: center; white-space: nowrap; }
        .dispatch-excel-table .dispatch-teshikaga-text,
        .dispatch-excel-table .dispatch-teshikaga-text a { color: #dc2626 !important; }
        .dispatch-mobile-view { display: none; }

        @media (max-width: 768px) {
            .dispatch-desktop-view { display: none; }
            .dispatch-mobile-view { display: block; }
            .dispatch-day-group { margin: 0 0 18px 0; }
            .dispatch-day-heading {
                position: sticky;
                top: 0;
                z-index: 2;
                margin: 0 0 7px 0;
                padding: 8px 10px;
                border-left: 5px solid #2563eb;
                border-radius: 7px;
                background: #eaf2ff;
                color: #172033;
                font-size: 16px;
                font-weight: 800;
            }
            .dispatch-mobile-card {
                margin: 0 0 8px 0;
                padding: 10px 11px;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                background: #ffffff;
                box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08);
                color: #172033;
            }
            .dispatch-date-line {
                display: grid;
                grid-template-columns: minmax(0, 1fr) 20px minmax(0, 1fr);
                align-items: center;
                gap: 4px;
                margin-bottom: 8px;
            }
            .dispatch-date-box {
                padding: 6px 7px;
                border-radius: 7px;
                text-align: center;
                line-height: 1.25;
            }
            .dispatch-pickup-date { background: #e8f1ff; color: #174ea6; }
            .dispatch-arrival-date { background: #e8f8ef; color: #176b3a; }
            .dispatch-date-label { display: block; font-size: 10px; font-weight: 700; }
            .dispatch-date-value { display: block; margin-top: 2px; font-size: 14px; font-weight: 800; }
            .dispatch-date-arrow { text-align: center; color: #64748b; font-weight: 800; }
            .dispatch-route {
                display: grid;
                grid-template-columns: minmax(0, 1fr) 20px minmax(0, 1fr);
                align-items: stretch;
                gap: 4px;
                margin-bottom: 8px;
            }
            .dispatch-route-box {
                min-width: 0;
                padding: 6px 7px;
                border: 1px solid #e2e8f0;
                border-radius: 7px;
                background: #fafcff;
            }
            .dispatch-route-box.dispatch-teshikaga-destination { background: #fee2e2; }
            .dispatch-route-label,
            .dispatch-detail-label { display: block; color: #64748b; font-size: 10px; font-weight: 700; }
            .dispatch-route-value { display: block; margin-top: 2px; font-size: 13px; font-weight: 800; overflow-wrap: anywhere; }
            .dispatch-route-arrow { align-self: center; text-align: center; color: #64748b; font-weight: 800; }
            .dispatch-details {
                display: grid;
                grid-template-columns: minmax(0, 1.4fr) minmax(0, 0.8fr);
                gap: 6px;
                margin-bottom: 6px;
            }
            .dispatch-detail-box {
                min-width: 0;
                padding: 5px 7px;
                border-radius: 6px;
                background: #f8fafc;
            }
            .dispatch-detail-value { display: block; margin-top: 1px; font-size: 13px; font-weight: 700; overflow-wrap: anywhere; }
            .dispatch-carrier {
                padding: 6px 7px;
                border-radius: 6px;
                background: #fff7e6;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    def safe_value(value):
        text = normalize_dispatch_text(value) or "未入力"
        return html.escape(text)

    def is_teshikaga_destination(value):
        return normalize_dispatch_text(value) == "弟子屈"

    def destination_value(value, highlight_text=False):
        destination = normalize_dispatch_text(value) or "未入力"
        if destination in customer_names:
            rendered = build_customer_detail_link(
                destination,
                class_name="dispatch-month-link",
            )
        else:
            rendered = html.escape(destination)
        if highlight_text and destination == "弟子屈":
            return f'<span class="dispatch-teshikaga-text">{rendered}</span>'
        return rendered

    columns = ["引取日", "引取先", "商品名", "数量", "運送会社", "納品先", "着日"]
    desktop_parts = [
        '<div class="dispatch-desktop-view">',
        '<table class="dispatch-excel-table"><thead><tr>',
    ]
    desktop_parts.extend(f"<th>{html.escape(column)}</th>" for column in columns)
    desktop_parts.append("</tr></thead><tbody>")
    last_display_row_index = len(display_df) - 1
    for row_index, (_, row) in enumerate(display_df.iterrows()):
        row_classes = []
        if row_index == 0:
            row_classes.append("dispatch-table-scroll-top")
        if row_index == last_display_row_index:
            row_classes.append("dispatch-table-scroll-bottom")
        class_attribute = f' class="{" ".join(row_classes)}"' if row_classes else ""
        desktop_parts.append(f"<tr{class_attribute}>")
        for column in columns:
            css_class = "date-cell" if column in ["引取日", "着日"] else "quantity-cell" if column == "数量" else ""
            cell_value = (
                destination_value(row.get(column), highlight_text=True)
                if column == "納品先"
                else safe_value(row.get(column))
            )
            desktop_parts.append(f'<td class="{css_class}">{cell_value}</td>')
        desktop_parts.append("</tr>")
    desktop_parts.append("</tbody></table></div>")

    mobile_parts = ['<div class="dispatch-mobile-view">']
    mobile_row_index = 0
    for pickup_date, day_rows in display_df.groupby("引取日", sort=False, dropna=False):
        pickup_label = safe_value(pickup_date)
        mobile_parts.append('<section class="dispatch-day-group">')
        mobile_parts.append(
            f'<div class="dispatch-day-heading">{pickup_label}　引取 {len(day_rows)}件</div>'
        )
        for _, row in day_rows.iterrows():
            card_classes = ["dispatch-mobile-card"]
            if mobile_row_index == 0:
                card_classes.append("dispatch-table-scroll-top")
            if mobile_row_index == last_display_row_index:
                card_classes.append("dispatch-table-scroll-bottom")
            mobile_parts.extend(
                [
                    f'<article class="{" ".join(card_classes)}">',
                    '<div class="dispatch-date-line">',
                    f'<div class="dispatch-date-box dispatch-pickup-date"><span class="dispatch-date-label">引取日</span><span class="dispatch-date-value">{safe_value(row.get("引取日"))}</span></div>',
                    '<div class="dispatch-date-arrow">→</div>',
                    f'<div class="dispatch-date-box dispatch-arrival-date"><span class="dispatch-date-label">着日</span><span class="dispatch-date-value">{safe_value(row.get("着日"))}</span></div>',
                    '</div>',
                    '<div class="dispatch-route">',
                    f'<div class="dispatch-route-box"><span class="dispatch-route-label">引取先</span><span class="dispatch-route-value">{safe_value(row.get("引取先"))}</span></div>',
                    '<div class="dispatch-route-arrow">→</div>',
                    f'<div class="dispatch-route-box{" dispatch-teshikaga-destination" if is_teshikaga_destination(row.get("納品先")) else ""}"><span class="dispatch-route-label">納品先</span><span class="dispatch-route-value">{destination_value(row.get("納品先"))}</span></div>',
                    '</div>',
                    '<div class="dispatch-details">',
                    f'<div class="dispatch-detail-box"><span class="dispatch-detail-label">商品名</span><span class="dispatch-detail-value">{safe_value(row.get("商品名"))}</span></div>',
                    f'<div class="dispatch-detail-box"><span class="dispatch-detail-label">数量</span><span class="dispatch-detail-value">{safe_value(row.get("数量"))}</span></div>',
                    '</div>',
                    f'<div class="dispatch-carrier"><span class="dispatch-detail-label">運送会社</span><span class="dispatch-detail-value">{safe_value(row.get("運送会社"))}</span></div>',
                    '</article>',
                ]
            )
            mobile_row_index += 1
        mobile_parts.append("</section>")
    mobile_parts.append("</div>")

    st.markdown("".join(desktop_parts + mobile_parts), unsafe_allow_html=True)


def get_dispatch_table_month_name():
    """配車表で表示する月をsession_stateに保持する。"""
    selected = str(st.session_state.get("dispatch_table_month") or "").strip()
    if selected not in DISPATCH_MONTH_SHEETS:
        current_month_name = f"{date.today().month}月"
        selected = current_month_name if current_month_name in DISPATCH_MONTH_SHEETS else DISPATCH_MONTH_SHEETS[0]
        st.session_state["dispatch_table_month"] = selected
    return selected


def change_dispatch_table_month(delta, scroll_target=None):
    """配車表の表示月を前月・翌月へ移動し、月別の絞り込みだけを解除する。"""
    current = get_dispatch_table_month_name()
    current_index = DISPATCH_MONTH_SHEETS.index(current)
    next_index = (current_index + int(delta)) % len(DISPATCH_MONTH_SHEETS)
    st.session_state["dispatch_table_month"] = DISPATCH_MONTH_SHEETS[next_index]
    if scroll_target in {"top", "bottom"}:
        st.session_state["dispatch_table_month_scroll_target"] = scroll_target

    for key in list(st.session_state.keys()):
        if key.startswith("dispatch_filter_"):
            del st.session_state[key]


def dispatch_table_month_title(df, selected_month):
    """配車表の月見出しに使う年を、該当月の実データから決める。"""
    try:
        month_number = int(str(selected_month).replace("月", ""))
    except ValueError:
        month_number = date.today().month

    month_rows = df[df["参照シート"] == selected_month]
    years = []
    for column in ("_引取日", "_着日"):
        if column not in month_rows.columns:
            continue
        for value in month_rows[column]:
            if isinstance(value, date) and value.month == month_number:
                years.append(value.year)

    year = max(set(years), key=years.count) if years else date.today().year
    return f"{year}年{month_number}月"


def show_dispatch_table_month_switcher(df, selected_month, key_suffix="top"):
    """配車カレンダーと同じ形で、配車表の表示月を切り替える。"""
    col_prev, col_month, col_next = st.columns([1, 2, 1])
    is_bottom = key_suffix == "bottom"
    prev_args = (-1, "bottom") if is_bottom else (-1,)
    next_args = (1, "top") if is_bottom else (1,)

    with col_prev:
        st.button(
            "◀",
            key=f"dispatch_table_prev_month_{key_suffix}",
            use_container_width=True,
            on_click=change_dispatch_table_month,
            args=prev_args,
        )

    with col_month:
        st.markdown(
            '<div style="text-align:center;font-size:1.2rem;font-weight:700;padding-top:0.35rem;">'
            f'{dispatch_table_month_title(df, selected_month)}</div>',
            unsafe_allow_html=True,
        )

    with col_next:
        st.button(
            "▶",
            key=f"dispatch_table_next_month_{key_suffix}",
            use_container_width=True,
            on_click=change_dispatch_table_month,
            args=next_args,
        )


def show_dispatch_board():
    st.markdown("---")
    st.header("🚚 配車表")
    show_back_home_button("dispatch_board_back_home")
    month_scroll_target = st.session_state.pop(
        "dispatch_table_month_scroll_target",
        None,
    )
    st.caption("配車表1.xlsmの月別シートを、元のExcelに近い一覧で表示します。")

    selected_month = get_dispatch_table_month_name()
    with st.spinner("配車表を読み込んでいます…"):
        df = load_dispatch_board_month_data(selected_month)

    if df.empty:
        # 従来は全12か月を読んだ結果が空の時だけこの警告を出していた。
        # 表示月だけが空の場合との区別を保つため、その場合だけ従来処理で確認する。
        full_df = load_dispatch_board_data()
        if full_df.empty:
            st.warning("1月～12月シートに表示できるデータがありません。")
            return
        df = full_df[full_df["参照シート"] == selected_month].copy()

    show_dispatch_table_month_switcher(df, selected_month, key_suffix="top")

    month_df = df[df["参照シート"] == selected_month].copy()
    filtered = show_dispatch_filters(month_df)

    st.markdown(
        f"**参照：{selected_month}シート　｜　全 {len(month_df)}件　｜　条件一致 {len(filtered)}件**"
    )
    st.caption("※ 1件は、元のExcelの月別シートにある明細1行です。")
    render_month_navigation_anchor("dispatch_table_month_content_top")

    if filtered.empty:
        st.info("条件に一致する配車はありません。")
        st.markdown("---")
        show_dispatch_table_month_switcher(df, selected_month, key_suffix="bottom")
        restore_month_navigation_scroll(
            month_scroll_target,
            ".dispatch-table-scroll-top",
            ".dispatch-table-scroll-bottom",
            fallback_anchor_id="dispatch_table_month_content_top",
        )
        return

    display_df = filtered.sort_values(
        ["_引取日", "_着日", "発注番号"],
        na_position="last",
    )[["引取日", "引取先", "商品名", "数量", "運送会社", "納品先", "着日"]].copy()

    display_df["引取日"] = filtered.loc[display_df.index, "_引取日"].map(dispatch_date_label)
    display_df["着日"] = filtered.loc[display_df.index, "_着日"].map(dispatch_date_label)
    for column in ["引取先", "商品名", "数量", "運送会社", "納品先"]:
        display_df[column] = display_df[column].map(normalize_dispatch_text)

    try:
        customer_names = load_dispatch_customer_names()
    except Exception:
        # 顧客名一覧を確認できない場合も、配車表は従来どおり文字表示で続行する。
        customer_names = set()

    render_dispatch_responsive_list(display_df, customer_names)

    st.markdown("---")
    show_dispatch_table_month_switcher(df, selected_month, key_suffix="bottom")
    restore_month_navigation_scroll(
        month_scroll_target,
        ".dispatch-table-scroll-top",
        ".dispatch-table-scroll-bottom",
        fallback_anchor_id="dispatch_table_month_content_top",
    )


# =========================
# ソリュブル在庫（aoベンチャーグレイン配車表.xlsx）
# =========================
def soluble_cell_is_manual(cell):
    """Excelで黄色に塗られたセルを手入力値として扱う。"""
    if cell.fill.fill_type != "solid":
        return False
    color = cell.fill.fgColor
    if color.type == "rgb":
        return str(color.rgb or "").upper().endswith("FFFF00")
    return False


def soluble_formula_value(formula_ws, value_ws, row, column, memo=None, visiting=None):
    """保存後にExcelの計算キャッシュが空でも、対象表の単純な加減式を表示できるようにする。"""
    memo = memo if memo is not None else {}
    visiting = visiting if visiting is not None else set()
    key = (row, column)
    if key in memo:
        return memo[key]
    if key in visiting:
        return None

    raw = formula_ws.cell(row, column).value
    cached = value_ws.cell(row, column).value
    if not (isinstance(raw, str) and raw.startswith("=")):
        memo[key] = raw
        return raw
    if cached is not None:
        memo[key] = cached
        return cached

    expression = raw[1:].replace(" ", "").replace("$", "").upper()
    tokens = re.findall(r"[A-Z]+\d+|\d+(?:\.\d+)?|[+-]", expression)
    if not tokens or "".join(tokens) != expression:
        return None

    visiting.add(key)
    try:
        def token_value(token):
            match = re.fullmatch(r"([A-Z]+)(\d+)", token)
            if match:
                letters, target_row = match.groups()
                target_column = 0
                for letter in letters:
                    target_column = target_column * 26 + (ord(letter) - 64)
                return soluble_formula_value(
                    formula_ws,
                    value_ws,
                    int(target_row),
                    target_column,
                    memo,
                    visiting,
                )
            return float(token)

        result = token_value(tokens[0])
        index = 1
        while index < len(tokens):
            operator = tokens[index]
            right = token_value(tokens[index + 1])
            if result is None:
                result = 0
            if right is None:
                right = 0
            if isinstance(result, (date, datetime)) and isinstance(right, (int, float)):
                result = result + timedelta(days=right if operator == "+" else -right)
            else:
                result = result + right if operator == "+" else result - right
            index += 2
        memo[key] = result
        return result
    except Exception:
        return None
    finally:
        visiting.discard(key)


def soluble_date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def soluble_number_label(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "—"
    if isinstance(value, (int, float)):
        number = float(value)
        if number.is_integer():
            return f"{int(number):,}"
        return f"{number:,.2f}".rstrip("0").rstrip(".")
    return str(value)


def soluble_input_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value)


def parse_soluble_number(text, label):
    cleaned = str(text or "").strip().replace(",", "").replace("，", "")
    if not cleaned:
        return None
    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"{label}は数字で入力してください。") from exc
    return int(number) if number.is_integer() else number


def same_soluble_value(left, right):
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=0, abs_tol=1e-9)
    return left == right


def normalize_soluble_customer_name(value):
    """ソリュブル上段の顧客名照合用。半角・全角空白の違いだけを吸収する。"""
    return re.sub(r"[\s　]+", "", clean_value(value, blank_text=""))


def find_soluble_customer_row(ws, customer_name):
    """ソリュブルシート上段から顧客名で対象行を探す。"""
    target = normalize_soluble_customer_name(customer_name)
    if not target:
        return None

    customer_column = SOLUBLE_CUSTOMER_COLUMNS["customer_name"]
    # 上段の顧客一覧は2行目の見出しから、日別表が始まる10行目より前にある。
    for row_number in range(3, min(ws.max_row, 10) + 1):
        if normalize_soluble_customer_name(ws.cell(row_number, customer_column).value) == target:
            return row_number
    return None


def calculate_soluble_customer_next_delivery(delivery_date_value, delivery_quantity, usage):
    """G列の「配達数量÷使用数量/日＋配達日」と同じ表示日を計算する。"""
    delivery_day = soluble_date_value(delivery_date_value)
    if delivery_day is None:
        return None
    try:
        quantity = float(delivery_quantity)
        daily_usage = float(usage)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(quantity) or not math.isfinite(daily_usage) or daily_usage <= 0:
        return None
    calculated = datetime.combine(delivery_day, datetime.min.time()) + timedelta(
        days=quantity / daily_usage
    )
    return calculated.date()


@st.cache_data(ttl=60, show_spinner=False)
def read_soluble_customer_summaries(content):
    """三谷牧場・熊林牧場の上段4項目をまとめて読む。"""
    formula_wb = load_workbook(BytesIO(content), data_only=False, read_only=False)
    value_wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    try:
        if SOLUBLE_SHEET_NAME not in formula_wb.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        formula_ws = formula_wb[SOLUBLE_SHEET_NAME]
        value_ws = value_wb[SOLUBLE_SHEET_NAME]
        result = {}

        for customer_name in SOLUBLE_CUSTOMER_NAMES:
            row_number = find_soluble_customer_row(formula_ws, customer_name)
            if row_number is None:
                continue

            delivery_date_value = formula_ws.cell(
                row_number, SOLUBLE_CUSTOMER_COLUMNS["delivery_date"]
            ).value
            delivery_quantity = formula_ws.cell(
                row_number, SOLUBLE_CUSTOMER_COLUMNS["delivery_quantity"]
            ).value
            usage = formula_ws.cell(
                row_number, SOLUBLE_CUSTOMER_COLUMNS["usage"]
            ).value
            next_delivery_value = value_ws.cell(
                row_number, SOLUBLE_CUSTOMER_COLUMNS["next_delivery"]
            ).value
            next_delivery = soluble_date_value(next_delivery_value)
            if next_delivery is None:
                next_delivery = calculate_soluble_customer_next_delivery(
                    delivery_date_value,
                    delivery_quantity,
                    usage,
                )

            result[customer_name] = {
                "row": row_number,
                "顧客名": customer_name,
                "配達日": soluble_date_value(delivery_date_value),
                "配達数量": delivery_quantity,
                "次回配達予定": next_delivery,
                "使用数量/日": usage,
            }
        return result
    finally:
        formula_wb.close()
        value_wb.close()


def get_soluble_customer_summary(content, customer_name):
    """選択中の顧客に対応するソリュブル上段情報を返す。"""
    target = normalize_soluble_customer_name(customer_name)
    for name, summary in read_soluble_customer_summaries(content).items():
        if normalize_soluble_customer_name(name) == target:
            return summary
    return None


@st.cache_data(ttl=60, show_spinner=False)
def load_soluble_workbook_content():
    """Dropboxを優先し、開発用PCでは同期済みローカルファイルも利用する。"""
    target_path = str(SOLUBLE_DROPBOX_FILE_PATH or SOLUBLE_DROPBOX_DEFAULT_FILE_PATH).strip()
    if has_dropbox_auth_config():
        access_token = get_dropbox_access_token()
        content, response = download_dropbox_file(target_path, access_token)
        if content is not None:
            return content, "Dropbox"
        local_path = Path(str(SOLUBLE_LOCAL_FILE))
        if not local_path.exists():
            raise RuntimeError(
                "aoベンチャーグレイン配車表.xlsxをDropboxから取得できませんでした。\n"
                + dropbox_error_text(response)
            )
        return local_path.read_bytes(), "同期済みローカルファイル"

    local_path = Path(str(SOLUBLE_LOCAL_FILE))
    if not local_path.exists():
        raise FileNotFoundError(f"対象ファイルが見つかりません：{local_path}")
    return local_path.read_bytes(), "同期済みローカルファイル"


def read_soluble_rows(content):
    formula_wb = load_workbook(BytesIO(content), data_only=False, read_only=False)
    value_wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    try:
        if SOLUBLE_SHEET_NAME not in formula_wb.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        formula_ws = formula_wb[SOLUBLE_SHEET_NAME]
        value_ws = value_wb[SOLUBLE_SHEET_NAME]
        memo = {}
        rows = []
        for row_number in range(11, formula_ws.max_row + 1):
            day_value = soluble_formula_value(formula_ws, value_ws, row_number, 2, memo)
            day = soluble_date_value(day_value)
            if day is None:
                continue
            record = {"row": row_number, "date": day}
            for location, columns in SOLUBLE_LOCATIONS.items():
                for field, column in columns.items():
                    record[f"{location}_{field}"] = soluble_formula_value(
                        formula_ws, value_ws, row_number, column, memo
                    )
                    record[f"{location}_{field}_manual"] = soluble_cell_is_manual(
                        formula_ws.cell(row_number, column)
                    )
                    record[f"{location}_{field}_formula"] = (
                        formula_ws.cell(row_number, column).value
                        if isinstance(formula_ws.cell(row_number, column).value, str)
                        and formula_ws.cell(row_number, column).value.startswith("=")
                        else ""
                    )
            rows.append(record)
        return rows
    finally:
        formula_wb.close()
        value_wb.close()


def _disabled_unsafe_xml_builder(content, row_number, location, updates):
    """XLSX全体を再生成せず、対象セルのXMLだけを変更して既存の計算結果を保つ。"""
    if location not in SOLUBLE_LOCATIONS:
        raise ValueError("対象の会社が正しくありません。")
    if row_number < 11:
        raise ValueError("更新する行が正しくありません。")
    if not updates:
        raise ValueError("変更された項目がありません。")

    columns = SOLUBLE_LOCATIONS[location]
    for field in updates:
        if field not in columns:
            raise ValueError("更新項目が正しくありません。")

    original_rows = read_soluble_rows(content)
    current_row = next((row for row in original_rows if row["row"] == row_number), None)
    previous_row = next((row for row in original_rows if row["row"] == row_number - 1), None)
    if current_row is None:
        raise ValueError("更新する日付行が見つかりません。")

    # openpyxlは共有数式を各セルの通常の式へ展開して読めるため、表示値の再計算に利用する。
    formula_book = load_workbook(BytesIO(content), data_only=False, read_only=False)
    try:
        if SOLUBLE_SHEET_NAME not in formula_book.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        formula_sheet = formula_book[SOLUBLE_SHEET_NAME]
        expanded_formulas = {
            cell.coordinate: cell.value[1:]
            for row in formula_sheet.iter_rows(min_row=11, min_col=2, max_col=8)
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        }
    finally:
        formula_book.close()

    resolved_updates = {}
    cached_values = {}
    for field, requested_value in updates.items():
        if requested_value == "__AUTO_INVENTORY__":
            if field != "inventory" or previous_row is None:
                raise ValueError("この日は在庫を自動計算にできません。")
            inventory_column = columns["inventory"]
            usage_column = columns["usage"]
            delivery_column = columns["delivery"]
            inventory_letter = chr(64 + inventory_column)
            usage_letter = chr(64 + usage_column)
            delivery_letter = chr(64 + delivery_column)
            formula = f"={inventory_letter}{row_number - 1}-{usage_letter}{row_number}+{delivery_letter}{row_number}"
            previous_inventory = previous_row.get(f"{location}_inventory") or 0
            current_usage = updates.get("usage", current_row.get(f"{location}_usage")) or 0
            current_delivery = updates.get("delivery", current_row.get(f"{location}_delivery")) or 0
            if not all(isinstance(value, (int, float)) for value in (previous_inventory, current_usage, current_delivery)):
                raise ValueError("在庫の自動計算に使う値が数字ではありません。")
            resolved_updates[field] = formula
            cached_values[field] = previous_inventory - current_usage + current_delivery
        else:
            resolved_updates[field] = requested_value
            cached_values[field] = requested_value

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", main_ns)

    with zipfile.ZipFile(BytesIO(content), "r") as source_zip:
        workbook_root = ET.fromstring(source_zip.read("xl/workbook.xml"))
        relationship_id = ""
        for sheet_node in workbook_root.findall(f".//{{{main_ns}}}sheet"):
            if sheet_node.get("name") == SOLUBLE_SHEET_NAME:
                relationship_id = sheet_node.get(f"{{{office_rel_ns}}}id", "")
                break
        if not relationship_id:
            raise ValueError("ソリュブルシートが見つかりません。")

        relationships_root = ET.fromstring(source_zip.read("xl/_rels/workbook.xml.rels"))
        sheet_target = ""
        for relationship in relationships_root.findall(f"{{{package_rel_ns}}}Relationship"):
            if relationship.get("Id") == relationship_id:
                sheet_target = relationship.get("Target", "")
                break
        if not sheet_target:
            raise ValueError("ソリュブルシートの保存先を確認できません。")
        sheet_part = (
            sheet_target.lstrip("/")
            if sheet_target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", sheet_target))
        )

        sheet_root = ET.fromstring(source_zip.read(sheet_part))
        styles_root = ET.fromstring(source_zip.read("xl/styles.xml"))
        fills = styles_root.find(f"{{{main_ns}}}fills")
        cell_xfs = styles_root.find(f"{{{main_ns}}}cellXfs")
        if fills is None or cell_xfs is None:
            raise ValueError("Excelの表示形式を確認できません。")

        yellow_fill_id = None
        for fill_index, fill in enumerate(list(fills)):
            foreground = fill.find(f".//{{{main_ns}}}fgColor")
            if foreground is not None and str(foreground.get("rgb", "")).upper().endswith("FFFF00"):
                yellow_fill_id = fill_index
                break
        if yellow_fill_id is None:
            fill = ET.Element(f"{{{main_ns}}}fill")
            pattern = ET.SubElement(fill, f"{{{main_ns}}}patternFill", {"patternType": "solid"})
            ET.SubElement(pattern, f"{{{main_ns}}}fgColor", {"rgb": "FFFFFF00"})
            ET.SubElement(pattern, f"{{{main_ns}}}bgColor", {"indexed": "64"})
            fills.append(fill)
            yellow_fill_id = len(fills) - 1
            fills.set("count", str(len(fills)))

        style_cache = {}

        def style_with_manual_fill(style_id, manual):
            style_id = int(style_id or 0)
            if style_id >= len(cell_xfs):
                style_id = 0
            original_xf = cell_xfs[style_id]
            current_fill_id = int(original_xf.get("fillId", "0"))
            is_yellow = current_fill_id == yellow_fill_id
            if is_yellow == manual:
                return style_id
            cache_key = (style_id, manual)
            if cache_key in style_cache:
                return style_cache[cache_key]
            new_xf = copy.deepcopy(original_xf)
            new_xf.set("fillId", str(yellow_fill_id if manual else 0))
            new_xf.set("applyFill", "1" if manual else "0")
            cell_xfs.append(new_xf)
            new_style_id = len(cell_xfs) - 1
            cell_xfs.set("count", str(len(cell_xfs)))
            style_cache[cache_key] = new_style_id
            return new_style_id

        sheet_data = sheet_root.find(f"{{{main_ns}}}sheetData")
        if sheet_data is None:
            raise ValueError("ソリュブルシートのセルを確認できません。")
        row_node = sheet_data.find(f"{{{main_ns}}}row[@r='{row_number}']")
        if row_node is None:
            raise ValueError("更新する日付行が見つかりません。")

        changed = []
        for field, new_value in resolved_updates.items():
            coordinate = f"{chr(64 + columns[field])}{row_number}"
            cell_node = row_node.find(f"{{{main_ns}}}c[@r='{coordinate}']")
            if cell_node is None:
                cell_node = ET.SubElement(row_node, f"{{{main_ns}}}c", {"r": coordinate})
            is_formula = isinstance(new_value, str) and new_value.startswith("=")
            should_be_manual = not is_formula
            cell_node.set("s", str(style_with_manual_fill(cell_node.get("s", "0"), should_be_manual)))
            cell_node.attrib.pop("t", None)
            for child in list(cell_node):
                if child.tag in {f"{{{main_ns}}}f", f"{{{main_ns}}}v", f"{{{main_ns}}}is"}:
                    cell_node.remove(child)
            if is_formula:
                formula_node = ET.SubElement(cell_node, f"{{{main_ns}}}f")
                formula_node.text = new_value[1:]
            cached_value = cached_values[field]
            if cached_value is not None:
                value_node = ET.SubElement(cell_node, f"{{{main_ns}}}v")
                value_node.text = str(int(cached_value)) if isinstance(cached_value, float) and cached_value.is_integer() else str(cached_value)
            changed.append((coordinate, new_value, should_be_manual))

        # 変更後の数値を使って、ソリュブル表内の単純な加減式の表示値も更新する。
        # これによりExcelを開く前でも、アプリとExcelプレビューで最新在庫を確認できる。
        cell_nodes = {
            cell.get("r", ""): cell
            for cell in sheet_root.findall(f".//{{{main_ns}}}c")
            if cell.get("r")
        }
        calculated = {}
        calculating = set()

        def calculate_xml_cell(coordinate):
            if coordinate in calculated:
                return calculated[coordinate]
            if coordinate in calculating:
                return 0
            node = cell_nodes.get(coordinate)
            if node is None:
                return 0
            formula_node = node.find(f"{{{main_ns}}}f")
            value_node = node.find(f"{{{main_ns}}}v")
            formula_text = (
                formula_node.text
                if formula_node is not None and formula_node.text
                else expanded_formulas.get(coordinate, "") if formula_node is not None else ""
            )
            if not formula_text:
                try:
                    value = float(value_node.text) if value_node is not None and value_node.text else 0
                except ValueError:
                    value = 0
                calculated[coordinate] = value
                return value

            expression = formula_text.replace(" ", "").replace("$", "").upper()
            tokens = re.findall(r"[A-Z]+\d+|\d+(?:\.\d+)?|[+-]", expression)
            if not tokens or "".join(tokens) != expression:
                return 0
            calculating.add(coordinate)
            try:
                def token_number(token):
                    return calculate_xml_cell(token) if re.fullmatch(r"[A-Z]+\d+", token) else float(token)

                result = token_number(tokens[0])
                index = 1
                while index < len(tokens):
                    right = token_number(tokens[index + 1])
                    result = result + right if tokens[index] == "+" else result - right
                    index += 2
                calculated[coordinate] = result
                return result
            finally:
                calculating.discard(coordinate)

        for coordinate, node in cell_nodes.items():
            match = re.fullmatch(r"([B-H])(\d+)", coordinate)
            formula_node = node.find(f"{{{main_ns}}}f")
            if not match or int(match.group(2)) < 11 or formula_node is None:
                continue
            result = calculate_xml_cell(coordinate)
            value_node = node.find(f"{{{main_ns}}}v")
            if value_node is None:
                value_node = ET.SubElement(node, f"{{{main_ns}}}v")
            value_node.text = str(int(result)) if float(result).is_integer() else str(result)

        calculation_properties = workbook_root.find(f"{{{main_ns}}}calcPr")
        if calculation_properties is not None:
            calculation_properties.set("calcMode", "auto")
            calculation_properties.set("fullCalcOnLoad", "1")
            calculation_properties.set("forceFullCalc", "1")

        replacement_parts = {
            sheet_part: ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True),
            "xl/styles.xml": ET.tostring(styles_root, encoding="utf-8", xml_declaration=True),
            "xl/workbook.xml": ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True),
        }
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as target_zip:
            for item in source_zip.infolist():
                target_zip.writestr(item, replacement_parts.get(item.filename, source_zip.read(item.filename)))

    saved_content = output.getvalue()
    formula_wb = load_workbook(BytesIO(saved_content), data_only=False, read_only=False)
    value_wb = load_workbook(BytesIO(saved_content), data_only=True, read_only=False)
    try:
        ws = formula_wb[SOLUBLE_SHEET_NAME]
        value_ws = value_wb[SOLUBLE_SHEET_NAME]
        for coordinate, expected, expected_manual in changed:
            cell = ws[coordinate]
            if cell.value != expected or soluble_cell_is_manual(cell) != expected_manual:
                raise ValueError(f"保存確認で{SOLUBLE_SHEET_NAME}!{coordinate}が一致しません。")
            if expected is not None and value_ws[coordinate].value is None:
                raise ValueError(f"保存確認で{SOLUBLE_SHEET_NAME}!{coordinate}の表示値がありません。")
    finally:
        formula_wb.close()
        value_wb.close()
    return saved_content, changed


def build_soluble_updated_workbook(content, row_number, location, updates):
    """openpyxlの標準保存を使い、Excel本体で開ける形式のまま対象セルを更新する。"""
    if location not in SOLUBLE_LOCATIONS:
        raise ValueError("対象の会社が正しくありません。")
    if row_number < 11:
        raise ValueError("更新する行が正しくありません。")
    if not updates:
        raise ValueError("変更された項目がありません。")

    workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    original_sheets = list(workbook.sheetnames)
    changed = []
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    clear_fill = PatternFill(fill_type=None)
    try:
        if SOLUBLE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        ws = workbook[SOLUBLE_SHEET_NAME]
        if row_number > ws.max_row:
            raise ValueError("更新する日付行が見つかりません。")
        columns = SOLUBLE_LOCATIONS[location]

        for field, requested_value in updates.items():
            if field not in columns:
                raise ValueError("更新項目が正しくありません。")
            if requested_value == "__AUTO_INVENTORY__":
                if field != "inventory" or row_number <= 11:
                    raise ValueError("この日は在庫を自動計算にできません。")
                inventory_letter = ws.cell(row_number, columns["inventory"]).column_letter
                usage_letter = ws.cell(row_number, columns["usage"]).column_letter
                delivery_letter = ws.cell(row_number, columns["delivery"]).column_letter
                new_value = (
                    f"={inventory_letter}{row_number - 1}-{usage_letter}{row_number}+{delivery_letter}{row_number}"
                )
                manual = False
            else:
                new_value = requested_value
                manual = True

            cell = ws.cell(row_number, columns[field])
            cell.value = new_value
            cell.fill = yellow_fill if manual else clear_fill
            changed.append((cell.coordinate, new_value, manual))

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    saved_content = output.getvalue()
    verified = load_workbook(BytesIO(saved_content), data_only=False, read_only=False)
    try:
        if list(verified.sheetnames) != original_sheets:
            raise ValueError("保存後にシート構成が変わったため、更新を中止しました。")
        ws = verified[SOLUBLE_SHEET_NAME]
        for coordinate, expected, expected_manual in changed:
            cell = ws[coordinate]
            if cell.value != expected or soluble_cell_is_manual(cell) != expected_manual:
                raise ValueError(f"保存確認で{SOLUBLE_SHEET_NAME}!{coordinate}が一致しません。")
    finally:
        verified.close()
    return saved_content, changed


# 分割したExcel処理をここから使用する。
# 上に残る同名処理は移行中の比較用で、実行時には以下の関数へ切り替わる。
from app_modules.soluble_excel import (
    build_soluble_updated_workbook,
    parse_soluble_number,
    read_soluble_rows,
    same_soluble_value,
    soluble_cell_is_manual,
    soluble_date_value,
    soluble_formula_value,
    soluble_input_value,
    soluble_number_label,
)


def ensure_soluble_backup_folder(access_token):
    response = call_dropbox_rpc(
        "files/create_folder_v2",
        {"path": SOLUBLE_BACKUP_FOLDER, "autorename": False},
        access_token,
    )
    if response.status_code == 200:
        return
    if response.status_code == 409 and "conflict" in str(response.text).lower():
        return
    raise RuntimeError("Dropboxにバックアップフォルダを作成できませんでした。\n" + dropbox_error_text(response))


def save_soluble_changes(row_number, location, updates):
    target_path = str(SOLUBLE_DROPBOX_FILE_PATH or SOLUBLE_DROPBOX_DEFAULT_FILE_PATH).strip()
    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")

    if has_dropbox_auth_config():
        access_token = get_dropbox_access_token()
        original_content, response = download_dropbox_file(target_path, access_token)
        if original_content is None:
            raise RuntimeError("最新の対象Excelを取得できませんでした。\n" + dropbox_error_text(response))
        revision = get_download_revision(response)
        if not revision:
            raise RuntimeError("Dropboxの更新番号を取得できないため、保存を中止しました。")
        saved_content, changed = build_soluble_updated_workbook(
            original_content, row_number, location, updates
        )
        ensure_soluble_backup_folder(access_token)
        backup_path = f"{SOLUBLE_BACKUP_FOLDER}/aoベンチャーグレイン配車表_{timestamp}.xlsx"
        backup_response = upload_dropbox_file(backup_path, original_content, access_token, mode="add")
        if backup_response.status_code != 200:
            raise RuntimeError("バックアップを作成できないため、本番ファイルは更新しません。\n" + dropbox_error_text(backup_response))
        upload_response = upload_dropbox_file(
            target_path, saved_content, access_token, mode="update", rev=revision
        )
        if upload_response.status_code == 409:
            raise RuntimeError("保存中にPCなどでExcelが更新されました。再読み込みしてからやり直してください。")
        if upload_response.status_code != 200:
            raise RuntimeError("対象Excelを更新できませんでした。\n" + dropbox_error_text(upload_response))
    else:
        local_path = Path(str(SOLUBLE_LOCAL_FILE))
        if not local_path.exists():
            raise FileNotFoundError(f"対象ファイルが見つかりません：{local_path}")
        original_content = local_path.read_bytes()
        saved_content, changed = build_soluble_updated_workbook(
            original_content, row_number, location, updates
        )
        backup_dir = local_path.parent / "Backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"aoベンチャーグレイン配車表_{timestamp}.xlsx"
        backup_path.write_bytes(original_content)
        local_path.write_bytes(saved_content)

    st.cache_data.clear()
    return changed


def build_soluble_customer_updated_workbook(content, customer_name, updates):
    """上段顧客のE/F/H列だけを更新し、G列の数式は保持する。"""
    if normalize_soluble_customer_name(customer_name) not in {
        normalize_soluble_customer_name(name) for name in SOLUBLE_CUSTOMER_NAMES
    }:
        raise ValueError("編集対象の顧客が正しくありません。")
    if not updates:
        raise ValueError("変更された項目がありません。")

    allowed_columns = {
        "delivery_date": SOLUBLE_CUSTOMER_COLUMNS["delivery_date"],
        "delivery_quantity": SOLUBLE_CUSTOMER_COLUMNS["delivery_quantity"],
        "usage": SOLUBLE_CUSTOMER_COLUMNS["usage"],
    }
    if any(field not in allowed_columns for field in updates):
        raise ValueError("更新項目が正しくありません。")

    workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    original_sheets = list(workbook.sheetnames)
    changed = []
    try:
        if SOLUBLE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        ws = workbook[SOLUBLE_SHEET_NAME]
        row_number = find_soluble_customer_row(ws, customer_name)
        if row_number is None:
            raise ValueError(f"ソリュブルシートに「{customer_name}」が見つかりません。")

        next_delivery_cell = ws.cell(
            row_number,
            SOLUBLE_CUSTOMER_COLUMNS["next_delivery"],
        )
        original_next_delivery_formula = next_delivery_cell.value
        if not (
            isinstance(original_next_delivery_formula, str)
            and original_next_delivery_formula.startswith("=")
        ):
            raise ValueError("次回配達予定の数式が見つからないため、更新を中止しました。")

        for field, new_value in updates.items():
            column_number = allowed_columns[field]
            cell = ws.cell(row_number, column_number)
            if not same_excel_value(cell.value, new_value):
                cell.value = new_value
                changed.append((cell.coordinate, new_value))

        if not changed:
            raise ValueError("変更された項目がありません。")

        # 次回配達予定は編集せず、既存のG列数式に任せる。
        if next_delivery_cell.value != original_next_delivery_formula:
            raise ValueError("次回配達予定の数式が変更されたため、保存を中止しました。")
        enable_excel_recalculation(workbook)
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    saved_content = output.getvalue()
    verified = load_workbook(BytesIO(saved_content), data_only=False, read_only=False)
    try:
        if list(verified.sheetnames) != original_sheets:
            raise ValueError("保存後にシート構成が変わったため、更新を中止しました。")
        ws = verified[SOLUBLE_SHEET_NAME]
        verified_row = find_soluble_customer_row(ws, customer_name)
        if verified_row is None:
            raise ValueError("保存後に対象顧客の行を確認できません。")
        if ws.cell(
            verified_row,
            SOLUBLE_CUSTOMER_COLUMNS["next_delivery"],
        ).value != original_next_delivery_formula:
            raise ValueError("保存後に次回配達予定の数式が変わっています。")
        for coordinate, expected in changed:
            if not same_excel_value(ws[coordinate].value, expected):
                raise ValueError(f"保存確認で{SOLUBLE_SHEET_NAME}!{coordinate}が一致しません。")
    finally:
        verified.close()
    return saved_content, changed


def verify_soluble_customer_saved_content(content, customer_name, changed):
    """Dropbox保存後に、変更セルとG列数式が残っていることを確認する。"""
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    try:
        if SOLUBLE_SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError("保存後の確認でソリュブルシートが見つかりません。")
        ws = workbook[SOLUBLE_SHEET_NAME]
        row_number = find_soluble_customer_row(ws, customer_name)
        if row_number is None:
            raise RuntimeError("保存後の確認で対象顧客が見つかりません。")
        formula = ws.cell(row_number, SOLUBLE_CUSTOMER_COLUMNS["next_delivery"]).value
        if not (isinstance(formula, str) and formula.startswith("=")):
            raise RuntimeError("保存後の確認で次回配達予定の数式が見つかりません。")
        for coordinate, expected in changed:
            if not same_excel_value(ws[coordinate].value, expected):
                raise RuntimeError(
                    f"Dropbox保存後の確認で{SOLUBLE_SHEET_NAME}!{coordinate}が更新されていません。"
                )
    finally:
        workbook.close()


def save_soluble_customer_changes(customer_name, updates):
    """上段顧客の変更を、既存ソリュブル保存と同じバックアップ方式で保存する。"""
    target_path = str(SOLUBLE_DROPBOX_FILE_PATH or SOLUBLE_DROPBOX_DEFAULT_FILE_PATH).strip()
    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")

    if has_dropbox_auth_config():
        access_token = get_dropbox_access_token()
        original_content, response = download_dropbox_file(target_path, access_token)
        if original_content is None:
            raise RuntimeError(
                "最新の対象Excelを取得できませんでした。\n" + dropbox_error_text(response)
            )
        revision = get_download_revision(response)
        if not revision:
            raise RuntimeError("Dropboxの更新番号を取得できないため、保存を中止しました。")

        saved_content, changed = build_soluble_customer_updated_workbook(
            original_content,
            customer_name,
            updates,
        )
        ensure_soluble_backup_folder(access_token)
        backup_path = (
            f"{SOLUBLE_BACKUP_FOLDER}/"
            f"aoベンチャーグレイン配車表_{timestamp}.xlsx"
        )
        backup_response = upload_dropbox_file(
            backup_path,
            original_content,
            access_token,
            mode="add",
        )
        if backup_response.status_code != 200:
            raise RuntimeError(
                "バックアップを作成できないため、本番ファイルは更新しません。\n"
                + dropbox_error_text(backup_response)
            )
        upload_response = upload_dropbox_file(
            target_path,
            saved_content,
            access_token,
            mode="update",
            rev=revision,
        )
        if upload_response.status_code == 409:
            raise RuntimeError(
                "保存中にPCなどでExcelが更新されました。再読み込みしてからやり直してください。"
            )
        if upload_response.status_code != 200:
            raise RuntimeError(
                "対象Excelを更新できませんでした。\n"
                + dropbox_error_text(upload_response)
            )

        confirmed_content, confirmed_response = download_dropbox_file(target_path, access_token)
        if confirmed_content is None:
            raise RuntimeError(
                "保存後のExcelを再取得できませんでした。\n"
                + dropbox_error_text(confirmed_response)
            )
        verify_soluble_customer_saved_content(
            confirmed_content,
            customer_name,
            changed,
        )
    else:
        local_path = Path(str(SOLUBLE_LOCAL_FILE))
        if not local_path.exists():
            raise FileNotFoundError(f"対象ファイルが見つかりません：{local_path}")
        original_content = local_path.read_bytes()
        saved_content, changed = build_soluble_customer_updated_workbook(
            original_content,
            customer_name,
            updates,
        )
        backup_dir = local_path.parent / "Backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"aoベンチャーグレイン配車表_{timestamp}.xlsx"
        backup_path.write_bytes(original_content)
        local_path.write_bytes(saved_content)
        verify_soluble_customer_saved_content(
            local_path.read_bytes(),
            customer_name,
            changed,
        )

    st.cache_data.clear()
    return changed


def render_soluble_customer_editor(customer_name, current, key_scope):
    """既存の商品カードと同じ操作で、配達日・配達数量・使用数量/日だけ編集する。"""
    identity = f"{key_scope}|{customer_name}|soluble"
    key_suffix = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16]
    edit_key = f"soluble_customer_edit_{key_suffix}"

    if not st.session_state.get(edit_key):
        if st.button("編集", key=f"soluble_customer_edit_button_{key_suffix}"):
            st.session_state[edit_key] = True
            st.rerun()
        return

    with st.form(f"soluble_customer_edit_form_{key_suffix}"):
        st.caption(f"🎤 {VOICE_INPUT_HELP} 入力欄は毎回空白から始まります。")
        delivery_date_text = st.text_input(
            "配達日",
            value="",
            placeholder="例：2026年7月15日",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        delivery_quantity_text = st.text_input(
            "配達数量",
            value="",
            placeholder="例：15000",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        usage_text = st.text_input(
            "使用数量/日",
            value="",
            placeholder="例：1000",
            help=VOICE_INPUT_HELP,
            autocomplete="off",
        )
        st.caption("次回配達予定はExcelの数式で計算されるため、直接編集しません。")
        save_col, cancel_col = st.columns(2)
        with save_col:
            save = st.form_submit_button("保存", type="primary", use_container_width=True)
        with cancel_col:
            cancel = st.form_submit_button("キャンセル", use_container_width=True)

    if cancel:
        st.session_state.pop(edit_key, None)
        st.rerun()

    if save:
        try:
            updates = {}
            if str(delivery_date_text).strip():
                new_delivery_date = parse_optional_date(delivery_date_text)
                if not same_excel_value(new_delivery_date, current.get("配達日")):
                    updates["delivery_date"] = new_delivery_date
            if str(delivery_quantity_text).strip():
                new_delivery_quantity = parse_optional_nonnegative_number(
                    delivery_quantity_text,
                    integer=False,
                )
                if not same_soluble_value(
                    new_delivery_quantity,
                    current.get("配達数量"),
                ):
                    updates["delivery_quantity"] = new_delivery_quantity
            if str(usage_text).strip():
                new_usage = parse_optional_nonnegative_number(
                    usage_text,
                    integer=False,
                )
                if not same_soluble_value(new_usage, current.get("使用数量/日")):
                    updates["usage"] = new_usage

            if not updates:
                st.warning("変更された項目がありません。")
                return

            with st.spinner("元ファイルをバックアップして保存しています…"):
                changed = save_soluble_customer_changes(customer_name, updates)
                field_map = {
                    "delivery_date": ("配達日", current.get("配達日")),
                    "delivery_quantity": ("配達数量", current.get("配達数量")),
                    "usage": ("使用数量/日", current.get("使用数量/日")),
                }
                history_changes = {
                    field_map[field][0]: (field_map[field][1], value)
                    for field, value in updates.items()
                    if field in field_map
                }
                remember_change_history_warning(
                    record_change_history_safely(
                        "顧客",
                        "",
                        customer_name,
                        "変更",
                        history_changes,
                        section="ソリュブル",
                    )
                )
            st.session_state.pop(edit_key, None)
            st.session_state["soluble_customer_save_success"] = {
                "customer_name": customer_name,
                "changed_count": len(changed),
            }
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))
        except Exception as exc:
            st.error(f"保存できませんでした：{exc}")


def render_soluble_customer_product_card(customer_name, current, key_scope):
    """三谷牧場・熊林牧場のソリュブル情報を既存商品カードと同じ形で表示する。"""
    with st.container(border=True):
        st.subheader("📦 ソリュブル")

        success = st.session_state.get("soluble_customer_save_success")
        if success and normalize_soluble_customer_name(
            success.get("customer_name")
        ) == normalize_soluble_customer_name(customer_name):
            st.success(f"保存しました（{success.get('changed_count', 0)}セル更新）。")
            st.session_state.pop("soluble_customer_save_success", None)

        col1, col2 = st.columns(2)
        with col1:
            st.caption("配達日")
            st.markdown(f"**{format_date(current.get('配達日'))}**")
            st.caption("配達数量")
            st.markdown(f"**{soluble_number_label(current.get('配達数量'))}**")
        with col2:
            st.caption("次回配達予定")
            st.markdown(f"**{format_date(current.get('次回配達予定'))}**")
            st.caption("使用数量/日")
            st.markdown(f"**{soluble_number_label(current.get('使用数量/日'))}**")

        render_soluble_customer_editor(customer_name, current, key_scope)



def get_soluble_water_it_history(dataframe, location):
    """選択会社に対応するWATER itのタンク履歴を、kgの数値行だけ返す。"""
    point_name = SOLUBLE_WATER_IT_POINT_NAMES.get(location)
    if not point_name or dataframe is None or dataframe.empty:
        return pd.DataFrame()

    history = get_water_it_customer_rows(dataframe, point_name)
    if history.empty:
        return history

    # ノベルズ・コスモは保管タンク1基。別の測定項目を誤って混ぜない。
    tank_rows = history[
        history["測定項目"].astype(str).str.contains("保管タンク", regex=False, na=False)
    ].copy()
    tank_rows = tank_rows[
        tank_rows["測定値_数値"].notna()
        & tank_rows["単位_表示"].map(normalize_water_it_unit).eq("kg")
    ].copy()
    if tank_rows.empty:
        return tank_rows

    # 測定項目が複数ある場合は、最新時刻を持つ1項目だけに限定する。
    latest_by_item = tank_rows.groupby("測定項目")["測定日時_解析"].max()
    selected_item = latest_by_item.idxmax()
    tank_rows = tank_rows[tank_rows["測定項目"] == selected_item].copy()
    tank_rows.sort_values("測定日時_解析", inplace=True)
    tank_rows.reset_index(drop=True, inplace=True)
    return tank_rows


def get_soluble_water_it_daily_actuals(history):
    """WATER it履歴から、各日の9:00実測値を日付ごとに返す。

    CSVを数日取り込まなかった場合でも、次回CSVに含まれる過去日の9:00値を
    アプリ表示へまとめて反映する。Excel本体は変更しない。
    """
    if history is None or history.empty:
        return {}

    daily = history.dropna(subset=["測定日時_解析", "測定値_数値"]).copy()
    if daily.empty:
        return {}

    measured_at = pd.to_datetime(daily["測定日時_解析"], errors="coerce")
    daily = daily[
        measured_at.notna()
        & measured_at.dt.hour.eq(9)
        & measured_at.dt.minute.eq(0)
    ].copy()
    if daily.empty:
        return {}

    daily["_water_it_measured_at"] = pd.to_datetime(
        daily["測定日時_解析"],
        errors="coerce",
    )
    daily["_water_it_date"] = daily["_water_it_measured_at"].dt.date
    daily.sort_values("_water_it_measured_at", inplace=True)

    # 同じ日の9:00行が重複していても、CSV内で最後の1件だけを採用する。
    daily = daily.groupby("_water_it_date", sort=True, as_index=False).tail(1)
    result = {}
    for _, row in daily.iterrows():
        value = float(row["測定値_数値"])
        if not math.isfinite(value):
            continue
        if value.is_integer():
            value = int(value)
        result[row["_water_it_date"]] = {
            "value": value,
            "measured_at": row["_water_it_measured_at"],
            "source": "09:00",
        }
    return result


def estimate_soluble_water_it_daily_usage(history, days):
    """WATER it履歴から1日平均使用量を参考値として推定する。

    1時間ごとの中央値で短時間の揺れをならし、大きな上昇は納品として区切る。
    各区間の開始値と終了値の減少分だけを合計するため、Excelの使用量/日や
    将来予測へは一切反映しない。
    """
    result = {
        "days": int(days),
        "average": None,
        "available_days": 0.0,
        "enough_data": False,
    }
    if history is None or history.empty:
        return result

    series = (
        history.dropna(subset=["測定日時_解析", "測定値_数値"])
        .sort_values("測定日時_解析")
        .set_index("測定日時_解析")["測定値_数値"]
        .astype(float)
    )
    series = series[~series.index.duplicated(keep="last")]
    if len(series) < 2:
        return result

    latest_time = series.index.max()
    oldest_time = series.index.min()
    available_days = max(
        0.0,
        (latest_time - oldest_time).total_seconds() / 86400.0,
    )
    result["available_days"] = available_days

    # 10分程度の端数は許容するが、期間が足りない時は無理に期間平均を出さない。
    if available_days < float(days) - 0.1:
        return result

    cutoff = latest_time - pd.Timedelta(days=int(days))
    hourly = series.resample("1h").median().interpolate(limit=2)
    hourly = hourly[(hourly.index >= cutoff) & (hourly.index <= latest_time)].dropna()
    if len(hourly) < 2:
        return result

    median_level = float(hourly.median())
    # 小さな測定揺れは納品扱いにしない。実タンクでは納品上昇が数千kg単位になる。
    delivery_jump = max(1000.0, abs(median_level) * 0.04)
    differences = hourly.diff()
    split_positions = [
        hourly.index.get_loc(timestamp)
        for timestamp in hourly.index[differences > delivery_jump]
    ]

    starts = [0] + split_positions
    ends = [position - 1 for position in split_positions] + [len(hourly) - 1]
    total_decrease = 0.0
    for start_position, end_position in zip(starts, ends):
        if end_position <= start_position:
            continue
        decrease = float(hourly.iloc[start_position] - hourly.iloc[end_position])
        if math.isfinite(decrease) and decrease > 0:
            total_decrease += decrease

    elapsed_days = (
        hourly.index[-1] - hourly.index[0]
    ).total_seconds() / 86400.0
    if elapsed_days <= 0:
        return result

    average = total_decrease / elapsed_days
    if not math.isfinite(average) or average < 0:
        return result

    result["average"] = average
    result["enough_data"] = True
    return result


def get_soluble_water_it_context(location, rows):
    """ソリュブル画面用の実測値・差額・参考平均をまとめる。失敗時はNone。"""
    if location not in SOLUBLE_WATER_IT_POINT_NAMES:
        return None
    try:
        dataframe, source = get_active_water_it_data()
        history = get_soluble_water_it_history(dataframe, location)
    except Exception:
        return None
    if history.empty:
        return None

    latest = history.iloc[-1]
    measured_at = latest["測定日時_解析"]
    if pd.isna(measured_at):
        return None
    actual_value = float(latest["測定値_数値"])
    if not math.isfinite(actual_value):
        return None
    if actual_value.is_integer():
        actual_value = int(actual_value)

    measured_date = measured_at.date()
    excel_row = next((row for row in rows if row.get("date") == measured_date), None)
    excel_value = (
        excel_row.get(f"{location}_inventory")
        if excel_row is not None
        else None
    )
    excel_usage = (
        excel_row.get(f"{location}_usage")
        if excel_row is not None
        else None
    )
    difference = None
    if isinstance(excel_value, (int, float)):
        difference = float(actual_value) - float(excel_value)
        if difference.is_integer():
            difference = int(difference)

    usage_averages = {
        days: estimate_soluble_water_it_daily_usage(history, days)
        for days in SOLUBLE_WATER_IT_USAGE_WINDOWS
    }

    # 日付ごとの在庫欄は、最新値ではなく各日の9:00実測だけを使う。
    # 上部の「現在の実測在庫」は従来どおりCSV内の最新値を表示する。
    daily_actuals = get_soluble_water_it_daily_actuals(history)

    today = get_jst_now().date()
    today_actual = daily_actuals.get(today)
    today_excel_row = next((row for row in rows if row.get("date") == today), None)
    today_excel_value = (
        today_excel_row.get(f"{location}_inventory")
        if today_excel_row is not None
        else None
    )

    return {
        "source": source,
        "history": history,
        "actual_value": actual_value,
        "measured_at": measured_at,
        "measured_date": measured_date,
        "unit": "kg",
        "excel_row": excel_row,
        "excel_value": excel_value,
        "excel_usage": excel_usage,
        "difference": difference,
        "usage_averages": usage_averages,
        "daily_actuals": daily_actuals,
        "today_9am_actual": today_actual,
        "today_9am_excel_row": today_excel_row,
        "today_9am_excel_value": today_excel_value,
    }


def apply_soluble_water_it_forecast(rows, location, context):
    """黄色い手入力を最優先し、その次にWATER it実測・予測を表示する。

    表示優先順位は、黄色い手入力値、WATER it実測値、WATER it予測値、
    Excel自動計算値の順とする。Excel本体・元のrows・使用量/日・納品は変更しない。
    """
    display_rows = [dict(row) for row in rows]
    if not context:
        return display_rows

    daily_actuals = context.get("daily_actuals") or {}
    if not daily_actuals:
        return display_rows

    previous_inventory = None
    started = False
    for row in sorted(display_rows, key=lambda item: item["date"]):
        row_date = row["date"]
        display_key = f"{location}_inventory_display"
        source_key = f"{location}_inventory_display_source"

        # 黄色い手入力値は、その日のWATER it実測値より優先する。
        # 以後の予測も、この手入力値を新しい基準として続ける。
        if row.get(f"{location}_inventory_manual"):
            manual_value = row.get(f"{location}_inventory")
            if isinstance(manual_value, (int, float)):
                previous_inventory = manual_value
                row[display_key] = manual_value
                row[source_key] = "excel_manual_baseline"
                started = True
                continue

        # 黄色い手入力値がない日は、その日の9:00実測値を優先する。
        actual = daily_actuals.get(row_date)
        if actual is not None:
            actual_value = actual.get("value")
            if isinstance(actual_value, (int, float)) and math.isfinite(float(actual_value)):
                previous_inventory = actual_value
                row[display_key] = actual_value
                row[source_key] = "water_it_actual"
                row[f"{location}_inventory_measured_at"] = actual.get("measured_at")
                started = True
                continue

        if not started:
            row[display_key] = row.get(f"{location}_inventory")
            row[source_key] = "excel"
            continue

        usage = row.get(f"{location}_usage")
        delivery = row.get(f"{location}_delivery")
        usage = usage if isinstance(usage, (int, float)) else 0
        delivery = delivery if isinstance(delivery, (int, float)) else 0
        if isinstance(previous_inventory, (int, float)):
            previous_inventory = previous_inventory - usage + delivery
            row[display_key] = previous_inventory
            row[source_key] = "water_it_forecast"
        else:
            row[display_key] = row.get(f"{location}_inventory")
            row[source_key] = "excel"

    return display_rows


def find_soluble_row_by_date(content, target_date):
    matches = [row for row in read_soluble_rows(content) if row.get("date") == target_date]
    if len(matches) != 1:
        if not matches:
            raise RuntimeError(
                f"Excelのソリュブルシートに{target_date.strftime('%Y/%m/%d')}の行がありません。"
            )
        raise RuntimeError("同じ日付の行が複数あるため、安全のため更新を中止しました。")
    return matches[0]


def verify_soluble_water_it_baseline(content, location, target_date, expected_value, next_formula):
    """Dropbox保存後に実測値・黄色・翌日の式が保たれていることを確認する。"""
    row = find_soluble_row_by_date(content, target_date)
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    try:
        if SOLUBLE_SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError("保存後のExcelにソリュブルシートがありません。")
        ws = workbook[SOLUBLE_SHEET_NAME]
        inventory_column = SOLUBLE_LOCATIONS[location]["inventory"]
        cell = ws.cell(row["row"], inventory_column)
        if not same_soluble_value(cell.value, expected_value):
            raise RuntimeError("保存後のExcelで実測値が一致しません。")
        if not soluble_cell_is_manual(cell):
            raise RuntimeError("保存後のExcelで実測値のセルが黄色になっていません。")
        if next_formula is not None and row["row"] < ws.max_row:
            actual_next_formula = ws.cell(row["row"] + 1, inventory_column).value
            if actual_next_formula != next_formula:
                raise RuntimeError("保存後のExcelで翌日の在庫計算式が変わっています。")
    finally:
        workbook.close()


def save_soluble_water_it_baseline(location, context):
    """選択会社の今日の実測値だけを、バックアップ付きでExcelの黄色い基準値にする。"""
    if location not in SOLUBLE_WATER_IT_POINT_NAMES:
        raise RuntimeError("WATER it実測値を反映できる会社ではありません。")
    if not context:
        raise RuntimeError("WATER itの実測値を確認できません。")

    today_actual = context.get("today_9am_actual") or {}
    measured_at = today_actual.get("measured_at")
    target_date = get_jst_now().date()
    actual_value = today_actual.get("value")
    if measured_at is None or not isinstance(actual_value, (int, float)):
        raise RuntimeError("今日9:00の実測値がないため、Excelへの反映を中止しました。")
    if measured_at.date() != target_date or measured_at.hour != 9 or measured_at.minute != 0:
        raise RuntimeError("今日9:00の実測値を確認できないため、Excelへの反映を中止しました。")
    if normalize_water_it_unit(context.get("unit")) != "kg":
        raise RuntimeError("kg以外の実測値はExcelへ反映しません。")
    if actual_value < 0:
        raise RuntimeError("実測値がマイナスのため、Excelへの反映を中止しました。")

    target_path = str(SOLUBLE_DROPBOX_FILE_PATH or SOLUBLE_DROPBOX_DEFAULT_FILE_PATH).strip()
    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")

    if has_dropbox_auth_config():
        access_token = get_dropbox_access_token()
        original_content, response = download_dropbox_file(target_path, access_token)
        if original_content is None:
            raise RuntimeError(
                "最新の対象Excelを取得できませんでした。\n" + dropbox_error_text(response)
            )
        revision = get_download_revision(response)
        if not revision:
            raise RuntimeError("Dropboxの更新番号を取得できないため、保存を中止しました。")

        latest_row = find_soluble_row_by_date(original_content, target_date)
        inventory_column = SOLUBLE_LOCATIONS[location]["inventory"]
        formula_book = load_workbook(BytesIO(original_content), data_only=False, read_only=False)
        try:
            ws = formula_book[SOLUBLE_SHEET_NAME]
            next_formula = (
                ws.cell(latest_row["row"] + 1, inventory_column).value
                if latest_row["row"] < ws.max_row
                else None
            )
        finally:
            formula_book.close()

        saved_content, changed = build_soluble_updated_workbook(
            original_content,
            latest_row["row"],
            location,
            {"inventory": actual_value},
        )
        ensure_soluble_backup_folder(access_token)
        backup_path = (
            f"{SOLUBLE_BACKUP_FOLDER}/"
            f"aoベンチャーグレイン配車表_{timestamp}.xlsx"
        )
        backup_response = upload_dropbox_file(
            backup_path,
            original_content,
            access_token,
            mode="add",
        )
        if backup_response.status_code != 200:
            raise RuntimeError(
                "バックアップを作成できないため、本番ファイルは更新しません。\n"
                + dropbox_error_text(backup_response)
            )
        upload_response = upload_dropbox_file(
            target_path,
            saved_content,
            access_token,
            mode="update",
            rev=revision,
        )
        if upload_response.status_code == 409:
            raise RuntimeError(
                "保存中にPCなどでExcelが更新されました。再読み込みしてからやり直してください。"
            )
        if upload_response.status_code != 200:
            raise RuntimeError(
                "対象Excelを更新できませんでした。\n" + dropbox_error_text(upload_response)
            )

        confirmed_content, confirmed_response = download_dropbox_file(target_path, access_token)
        if confirmed_content is None:
            raise RuntimeError(
                "保存後のExcelを再取得できませんでした。\n"
                + dropbox_error_text(confirmed_response)
            )
        verify_soluble_water_it_baseline(
            confirmed_content,
            location,
            target_date,
            actual_value,
            next_formula,
        )
    else:
        local_path = Path(str(SOLUBLE_LOCAL_FILE))
        if not local_path.exists():
            raise FileNotFoundError(f"対象ファイルが見つかりません：{local_path}")
        original_content = local_path.read_bytes()
        latest_row = find_soluble_row_by_date(original_content, target_date)
        inventory_column = SOLUBLE_LOCATIONS[location]["inventory"]
        formula_book = load_workbook(BytesIO(original_content), data_only=False, read_only=False)
        try:
            ws = formula_book[SOLUBLE_SHEET_NAME]
            next_formula = (
                ws.cell(latest_row["row"] + 1, inventory_column).value
                if latest_row["row"] < ws.max_row
                else None
            )
        finally:
            formula_book.close()
        saved_content, changed = build_soluble_updated_workbook(
            original_content,
            latest_row["row"],
            location,
            {"inventory": actual_value},
        )
        backup_dir = local_path.parent / "Backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"aoベンチャーグレイン配車表_{timestamp}.xlsx"
        backup_path.write_bytes(original_content)
        local_path.write_bytes(saved_content)
        verify_soluble_water_it_baseline(
            local_path.read_bytes(),
            location,
            target_date,
            actual_value,
            next_formula,
        )

    st.cache_data.clear()
    return changed


def render_soluble_water_it_summary(location, context):
    """ソリュブル画面に実測値、Excelとの差、参考平均、任意反映ボタンを表示する。"""
    if not context:
        return

    success = st.session_state.pop("soluble_water_it_excel_success", None)
    if success and success.get("location") == location:
        st.success(
            f"{success['date']}の実測値 {soluble_number_label(success['value'])} kg を"
            "Excelの基準値として保存しました。対象セルは黄色です。"
        )

    display_name = SOLUBLE_LOCATION_DISPLAY_NAMES.get(location, location)
    actual_value = context["actual_value"]
    measured_at = context["measured_at"]
    excel_value = context.get("excel_value")
    excel_usage = context.get("excel_usage")
    difference = context.get("difference")

    with st.container(border=True):
        st.subheader(f"💧 {display_name}のWATER it実測")
        st.caption(
            f"最終受信：{measured_at.strftime('%Y/%m/%d %H:%M')}　｜　参照：{context['source']}"
        )
        def summary_card(label, value, tone=""):
            tone_class = f" {tone}" if tone else ""
            return (
                f'<div class="soluble-waterit-stat{tone_class}">'
                f'<span class="soluble-waterit-stat-label">{html.escape(str(label))}</span>'
                f'<span class="soluble-waterit-stat-value">{html.escape(str(value))}</span>'
                '</div>'
            )

        st.markdown(
            """
            <style>
            .soluble-waterit-summary-grid {
                display: grid;
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 0.75rem;
                margin: 0.4rem 0 1.1rem;
            }
            .soluble-waterit-usage-grid {
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 0.75rem;
                margin: 0.45rem 0 0.8rem;
            }
            .soluble-waterit-stat {
                min-width: 0;
                box-sizing: border-box;
                padding: 0.9rem 1rem;
                border: 1px solid rgba(15, 23, 42, 0.12);
                border-radius: 14px;
                background: rgba(255, 255, 255, 0.9);
            }
            .soluble-waterit-stat.actual {
                background: #dcfce7;
                border-color: #4ade80;
            }
            .soluble-waterit-stat.excel {
                background: #f8fafc;
                border-color: #cbd5e1;
            }
            .soluble-waterit-stat.difference {
                background: #eff6ff;
                border-color: #93c5fd;
            }
            .soluble-waterit-stat.average {
                background: #f0fdfa;
                border-color: #5eead4;
            }
            .soluble-waterit-stat-label {
                display: block;
                color: #667085;
                font-size: 0.88rem;
                font-weight: 700;
                line-height: 1.35;
                margin-bottom: 0.35rem;
                overflow-wrap: anywhere;
            }
            .soluble-waterit-stat-value {
                display: block;
                color: #172033;
                font-size: clamp(1.55rem, 3.2vw, 2.2rem);
                font-weight: 800;
                line-height: 1.15;
                letter-spacing: -0.02em;
                white-space: normal;
                overflow: visible;
                text-overflow: clip;
                overflow-wrap: anywhere;
            }
            @media (max-width: 640px) {
                .soluble-waterit-summary-grid,
                .soluble-waterit-usage-grid {
                    grid-template-columns: 1fr;
                    gap: 0.55rem;
                }
                .soluble-waterit-stat {
                    padding: 0.78rem 0.85rem;
                }
                .soluble-waterit-stat-label {
                    font-size: 0.82rem;
                }
                .soluble-waterit-stat-value {
                    font-size: 1.45rem;
                    white-space: nowrap;
                    overflow-wrap: normal;
                }
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        actual_label = f"{soluble_number_label(actual_value)} kg"
        excel_label = (
            f"{soluble_number_label(excel_value)} kg"
            if excel_value is not None
            else "—"
        )
        difference_label = "—" if difference is None else f"{difference:+,.0f} kg"
        st.markdown(
            '<div class="soluble-waterit-summary-grid">'
            + summary_card(f"現在の実測在庫/{measured_at.strftime('%H:%M')}", actual_label, "actual")
            + summary_card("同日のExcel計算在庫", excel_label, "excel")
            + summary_card("実測 − Excel", difference_label, "difference")
            + '</div>',
            unsafe_allow_html=True,
        )

        st.markdown("#### 実績から見た1日平均使用量（参考）")
        excel_usage_label = (
            f"{soluble_number_label(excel_usage)} kg/日"
            if isinstance(excel_usage, (int, float))
            else "—"
        )
        st.markdown(
            '<div class="soluble-waterit-usage-grid">'
            + summary_card("Excel設定使用量", excel_usage_label, "excel")
            + '</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "WATER it履歴を1時間単位でならし、大きな在庫増加は納品として区切った推定値です。"
            "Excelの使用量/日や予測計算へは自動反映しません。"
        )

        # st.tabsは先頭タブが初期表示になるため、7日を先頭にして標準表示にする。
        usage_tab_specs = (
            (7, "7日（標準）"),
            (3, "3日"),
            (20, "20日"),
            (30, "30日"),
        )
        usage_tabs = st.tabs([label for _, label in usage_tab_specs])
        for tab, (days, _) in zip(usage_tabs, usage_tab_specs):
            estimate = context["usage_averages"][days]
            with tab:
                if estimate.get("enough_data") and estimate.get("average") is not None:
                    average_label = f"{estimate['average']:,.0f} kg/日"
                    detail_label = f"直近{days}日の実績平均"
                else:
                    average_label = "データ不足"
                    detail_label = (
                        f"直近{days}日分には不足しています（現在 約"
                        f"{estimate.get('available_days', 0):.1f}日分）"
                    )
                st.markdown(
                    '<div class="soluble-waterit-usage-grid">'
                    + summary_card(detail_label, average_label, "average")
                    + '</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("#### Excelへ反映（任意）")
        st.caption(
            "アプリは各日の9:00実測値を在庫に使います。Excelは自動変更せず、ここで確認して押した場合だけ、"
            "今日9:00の在庫を基準値として黄色セルへ保存します。保存前バックアップと保存後確認を行います。"
        )

        today = get_jst_now().date()
        today_actual = context.get("today_9am_actual") or {}
        baseline_value = today_actual.get("value")
        baseline_measured_at = today_actual.get("measured_at")
        excel_row = context.get("today_9am_excel_row")
        baseline_excel_value = context.get("today_9am_excel_value")
        has_today_9am = (
            isinstance(baseline_value, (int, float))
            and baseline_measured_at is not None
            and baseline_measured_at.date() == today
            and baseline_measured_at.hour == 9
            and baseline_measured_at.minute == 0
        )
        same_value = (
            has_today_9am
            and baseline_excel_value is not None
            and same_soluble_value(baseline_excel_value, baseline_value)
        )
        if not has_today_9am:
            st.warning("今日9:00の実測値がCSVにないため、Excelへの反映ボタンは使えません。")
            return
        if excel_row is None:
            st.warning("今日の行がExcelにないため、反映できません。")
            return
        if same_value:
            st.info("今日のExcel在庫は、すでに9:00実測値と同じです。")
            return

        confirm_key = f"soluble_water_it_confirm_{location}_{today.isoformat()}"
        confirmed = st.checkbox(
            f"{today.strftime('%Y/%m/%d')}のExcel在庫を "
            f"9:00実測の {soluble_number_label(baseline_value)} kg に変更する",
            key=confirm_key,
        )
        if st.button(
            "今日の実測値をExcelの基準値にする",
            key=f"soluble_water_it_save_{location}_{today.isoformat()}",
            type="primary",
            use_container_width=True,
            disabled=not confirmed,
        ):
            try:
                with st.spinner("元ファイルをバックアップし、9:00実測値を保存・確認しています…"):
                    changed = save_soluble_water_it_baseline(location, context)
                    remember_change_history_warning(
                        record_change_history_safely(
                            "顧客",
                            "",
                            SOLUBLE_LOCATION_DISPLAY_NAMES.get(location, location),
                            "変更",
                            {"在庫基準値": (baseline_excel_value, baseline_value)},
                            section="ソリュブル在庫（WATER it実測反映）",
                        )
                    )
                st.session_state["soluble_water_it_excel_success"] = {
                    "location": location,
                    "date": today.strftime("%Y/%m/%d"),
                    "value": baseline_value,
                    "changed_count": len(changed),
                }
                st.rerun()
            except Exception as exc:
                st.error(f"Excelへ反映できませんでした：{exc}")

def show_soluble_inventory_page():
    st.markdown("---")
    st.header("🧪 ソリュブル在庫")
    show_back_home_button("soluble_back_home")
    st.caption("aoベンチャーグレイン配車表.xlsx の「ソリュブル」シートを表示します。")

    with st.spinner("ソリュブル在庫を読み込んでいます…"):
        content, source = load_soluble_workbook_content()
        rows = read_soluble_rows(content)
        customer_summaries = read_soluble_customer_summaries(content)
    if not rows and not customer_summaries:
        st.warning("ソリュブルシートに表示できるデータがありません。")
        return

    location = st.radio(
        "表示する会社",
        list(SOLUBLE_LOCATIONS.keys()) + list(SOLUBLE_CUSTOMER_NAMES),
        horizontal=True,
        key="soluble_location",
        format_func=lambda name: SOLUBLE_LOCATION_DISPLAY_NAMES.get(name, name),
    )

    if location in SOLUBLE_CUSTOMER_NAMES:
        current = customer_summaries.get(location)
        if current is None:
            st.warning(f"{location}の行がソリュブルシートに見つかりません。")
            return
        st.caption(f"参照：{source}")
        render_soluble_customer_product_card(
            location,
            current,
            key_scope="soluble_inventory_page",
        )
        return

    if not rows:
        st.warning("ソリュブルシートに表示できる日付がありません。")
        return

    # 数値がまだ空の日も、ここから新しく入力できるように日付行はすべて表示対象にする。
    water_it_context = get_soluble_water_it_context(location, rows)
    if water_it_context is not None:
        render_soluble_water_it_summary(location, water_it_context)
        active_rows = apply_soluble_water_it_forecast(rows, location, water_it_context)
    else:
        # WATER itを読めない時も、既存のExcel表示・編集ルールは従来どおり動かす。
        active_rows = list(rows)
        if location in SOLUBLE_WATER_IT_POINT_NAMES:
            st.info("WATER itの保存済みデータを確認できないため、Excelの値だけを表示しています。")
    if not active_rows:
        st.info(f"{location}の表示データはありません。")
        return

    month_keys = sorted({(row["date"].year, row["date"].month) for row in active_rows})
    month_labels = [f"{year}年{month}月" for year, month in month_keys]
    # Streamlit CloudはUTCで動くため、日本時間の「今日」を使う。
    today = get_jst_now().date()
    today_key = (today.year, today.month)
    today_month_label = f"{today.year}年{today.month}月"
    default_month = month_keys.index(today_key) if today_key in month_keys else len(month_keys) - 1

    # 日付が変わった最初の表示だけ、表示月と開始日を今日へ戻す。
    # 同じ日のうちは、ユーザーが選んだ別の日付・表示月をそのまま維持する。
    month_widget_key = f"soluble_month_{location}"
    daily_reset_key = f"soluble_daily_default_{location}"
    if st.session_state.get(daily_reset_key) != today.isoformat():
        if today_month_label in month_labels:
            st.session_state[month_widget_key] = today_month_label
        st.session_state[daily_reset_key] = today.isoformat()

    selected_month_label = st.selectbox(
        "表示月",
        month_labels,
        index=default_month,
        key=month_widget_key,
    )
    selected_month_key = month_keys[month_labels.index(selected_month_label)]
    month_rows = [
        row for row in active_rows
        if (row["date"].year, row["date"].month) == selected_month_key
    ]

    day_options = [row["date"] for row in month_rows]
    default_day = day_options.index(today) if today in day_options else 0
    start_widget_key = f"soluble_start_{location}_{selected_month_label}"
    start_default_key = f"{start_widget_key}_default_{today.isoformat()}"
    if today in day_options and not st.session_state.get(start_default_key):
        st.session_state[start_widget_key] = today
        st.session_state[start_default_key] = True

    control_left, control_right = st.columns(2)
    with control_left:
        start_day = st.selectbox(
            "開始日",
            day_options,
            index=default_day,
            format_func=lambda day: f"{day.month}/{day.day}（{'月火水木金土日'[day.weekday()]}）",
            key=start_widget_key,
        )
    with control_right:
        period_widget_key = f"soluble_period_{location}"
        period_options = ["7日間", "14日間", "1か月"]
        period_default_key = f"{period_widget_key}_default_one_month_v2"
        if (
            not st.session_state.get(period_default_key)
            or st.session_state.get(period_widget_key) not in period_options
        ):
            st.session_state[period_widget_key] = "1か月"
            st.session_state[period_default_key] = True
        period = st.selectbox(
            "表示期間",
            period_options,
            index=2,
            key=period_widget_key,
        )
    manual_only = st.checkbox("黄色の手入力だけ表示", key=f"soluble_manual_only_{location}")

    if period == "1か月":
        next_month = 1 if start_day.month == 12 else start_day.month + 1
        next_year = start_day.year + 1 if start_day.month == 12 else start_day.year
        one_month_later = date(
            next_year,
            next_month,
            min(start_day.day, calendar.monthrange(next_year, next_month)[1]),
        )
        visible_rows = [
            row for row in active_rows
            if start_day <= row["date"] < one_month_later
        ]
    else:
        visible_rows = [
            row for row in month_rows
            if start_day <= row["date"] < start_day + timedelta(
                days=7 if period == "7日間" else 14
            )
        ]
    if manual_only:
        visible_rows = [
            row for row in visible_rows
            if any(row.get(f"{location}_{field}_manual") for field in ("usage", "delivery", "inventory"))
        ]

    st.markdown(
        """
        <style>
        .soluble-legend {display:flex; gap:.7rem; align-items:center; margin:.35rem 0 1rem; color:#596273;}
        .soluble-yellow-chip {display:inline-block; width:1.25rem; height:1.25rem; background:#fff59d; border:1px solid #e3cb42; border-radius:.3rem;}
        .soluble-card {background:rgba(255,255,255,.78); border:1px solid #cbd5e1; border-radius:16px; padding:14px 16px; margin:.65rem 0 .25rem; box-shadow:0 6px 16px rgba(30,41,59,.05);}
        .soluble-card-date {font-size:1.12rem; font-weight:800; margin-bottom:10px;}
        .soluble-values {display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:8px;}
        .soluble-value {background:#f8fafc; border-radius:12px; padding:9px 10px; min-width:0;}
        .soluble-value.manual {background:#fff59d; border:1px solid #e3cb42;}
        .soluble-value.waterit-actual {background:#dcfce7; border:1px solid #4ade80;}
        .soluble-value.waterit-forecast {background:#e0f2fe; border:1px solid #7dd3fc;}
        .soluble-value.negative {background:#fee2e2; border:1px solid #f87171;}
        .soluble-label {display:block; color:#697386; font-size:.78rem; margin-bottom:3px;}
        .soluble-number {display:block; color:#182033; font-size:1.04rem; font-weight:800; overflow-wrap:anywhere;}
        @media (max-width: 640px) {
          .soluble-card {padding:13px 12px; border-radius:14px;}
          .soluble-values {gap:6px;}
          .soluble-value {padding:9px 7px;}
          .soluble-label {font-size:.72rem;}
          .soluble-number {font-size:.96rem;}
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div class="soluble-legend"><span class="soluble-yellow-chip"></span><span>黄色はExcelの手入力　｜　緑はWATER it実測　｜　水色は実測起点のアプリ予測　｜　参照：{html.escape(source)}</span></div>',
        unsafe_allow_html=True,
    )

    if not visible_rows:
        st.info("この条件で表示するデータはありません。")
        return

    weekday = "月火水木金土日"
    for row in visible_rows:
        usage = row.get(f"{location}_usage")
        delivery = row.get(f"{location}_delivery")
        inventory = row.get(f"{location}_inventory")
        inventory_display = row.get(f"{location}_inventory_display", inventory)
        inventory_display_source = row.get(f"{location}_inventory_display_source", "excel")
        inventory_label = {
            "water_it_actual": "在庫（実測）",
            "water_it_forecast": "在庫（実測起点予測）",
            "excel_manual_baseline": "在庫（Excel手入力基準）",
        }.get(inventory_display_source, "在庫")
        cells = []
        for label, field, value in (
            ("使用量/日", "usage", usage),
            ("納品", "delivery", delivery),
            (inventory_label, "inventory", inventory_display),
        ):
            classes = ["soluble-value"]
            if field == "inventory" and inventory_display_source == "water_it_actual":
                classes.append("waterit-actual")
            elif field == "inventory" and inventory_display_source == "water_it_forecast":
                classes.append("waterit-forecast")
            elif row.get(f"{location}_{field}_manual"):
                classes.append("manual")
            if field == "inventory" and isinstance(value, (int, float)) and value < 0:
                classes.append("negative")
            cells.append(
                f'<div class="{" ".join(classes)}"><span class="soluble-label">{label}</span>'
                f'<span class="soluble-number">{html.escape(soluble_number_label(value))}</span></div>'
            )
        day = row["date"]
        st.markdown(
            f'<section class="soluble-card"><div class="soluble-card-date">{day.month}/{day.day}（{weekday[day.weekday()]}）</div>'
            f'<div class="soluble-values">{"".join(cells)}</div></section>',
            unsafe_allow_html=True,
        )

        with st.expander(f"✏️ {day.month}/{day.day}を入力・修正"):
            form_key = f"soluble_form_{location}_{row['row']}"
            with st.form(form_key):
                usage_text = st.text_input(
                    "使用量/日",
                    value=soluble_input_value(usage),
                    key=f"{form_key}_usage",
                    autocomplete="off",
                )
                delivery_text = st.text_input(
                    "納品",
                    value=soluble_input_value(delivery),
                    key=f"{form_key}_delivery",
                    autocomplete="off",
                )
                current_formula = bool(row.get(f"{location}_inventory_formula")) and not bool(
                    row.get(f"{location}_inventory_manual")
                )
                auto_inventory = st.checkbox(
                    "在庫は「前日在庫 − 使用量 + 納品」で自動計算する",
                    value=current_formula,
                    key=f"{form_key}_auto",
                )
                inventory_text = st.text_input(
                    "在庫（自動計算を外した場合に使用）",
                    value=soluble_input_value(inventory),
                    key=f"{form_key}_inventory",
                    autocomplete="off",
                )
                submitted = st.form_submit_button("バックアップして保存", use_container_width=True)

            if submitted:
                try:
                    new_usage = parse_soluble_number(usage_text, "使用量/日")
                    new_delivery = parse_soluble_number(delivery_text, "納品")
                    new_inventory = None if auto_inventory else parse_soluble_number(inventory_text, "在庫")
                    updates = {}
                    if not same_soluble_value(new_usage, usage):
                        updates["usage"] = new_usage
                    if not same_soluble_value(new_delivery, delivery):
                        updates["delivery"] = new_delivery
                    if auto_inventory:
                        if not current_formula:
                            updates["inventory"] = "__AUTO_INVENTORY__"
                    elif current_formula or not same_soluble_value(new_inventory, inventory):
                        updates["inventory"] = new_inventory
                    with st.spinner("元ファイルをバックアップして保存しています…"):
                        changed = save_soluble_changes(
                            row["row"],
                            location,
                            updates,
                        )
                        history_changes = {}
                        if "usage" in updates:
                            history_changes["使用量/日"] = (usage, new_usage)
                        if "delivery" in updates:
                            history_changes["納品"] = (delivery, new_delivery)
                        if "inventory" in updates:
                            before_inventory = "自動計算" if current_formula else inventory
                            after_inventory = "自動計算" if auto_inventory else new_inventory
                            history_changes["在庫"] = (before_inventory, after_inventory)
                        remember_change_history_warning(
                            record_change_history_safely(
                                "顧客",
                                "",
                                SOLUBLE_LOCATION_DISPLAY_NAMES.get(location, location),
                                "変更",
                                history_changes,
                                section=f"ソリュブル在庫 {day.strftime('%Y/%m/%d')}",
                            )
                        )
                    st.success(f"保存しました（{len(changed)}セル更新）。黄色は手入力値です。")
                    st.rerun()
                except Exception as error:
                    st.error(str(error))






# =========================
# WATER it接続（読み取り専用）
# =========================
def resolve_water_it_csv_path():
    """data.csvの場所を、このPythonファイル基準で解決する。"""
    configured = str(WATER_IT_CSV_PATH).strip() or "data.csv"
    path = Path(configured).expanduser()
    if not path.is_absolute():
        path = Path(__file__).resolve().parent / path
    return path


def read_water_it_source_bytes():
    """WATER itのCSVを読み取る。書き込み処理は行わない。"""
    csv_url = str(WATER_IT_CSV_URL).strip()
    if csv_url:
        try:
            response = requests.get(
                csv_url,
                timeout=WATER_IT_REQUEST_TIMEOUT,
                headers={"User-Agent": "Aoyama-WATER-it-readonly-test/1.0"},
            )
            response.raise_for_status()
        except requests.RequestException as exc:
            raise RuntimeError(f"WATER it CSV URLから取得できませんでした：{exc}") from exc

        content_type = str(response.headers.get("Content-Type", "")).lower()
        if "text/html" in content_type:
            raise RuntimeError(
                "WATER_IT_CSV_URLからCSVではなくHTMLが返されました。CSVを直接取得できるURLを設定してください。"
            )
        hostname = urllib.parse.urlparse(csv_url).hostname or "設定URL"
        return response.content, f"WATER_IT_CSV_URL（{hostname}）"

    path = resolve_water_it_csv_path()
    if not path.exists():
        raise RuntimeError(
            f"{path.name} が見つかりません。このPythonファイルと同じフォルダに data.csv を置いてください。"
        )
    if not path.is_file():
        raise RuntimeError(f"WATER_IT_CSV_PATH がファイルではありません：{path.name}")
    return path.read_bytes(), path.name


def normalize_water_it_unit(value):
    text = clean_value(value, blank_text="").strip()
    replacements = {
        "㎏": "kg",
        "ＫＧ": "kg",
        "ｋｇ": "kg",
        "Ｋｇ": "kg",
        "Ｌ": "L",
        "ℓ": "L",
        "㍑": "L",
    }
    return replacements.get(text, text)


def water_it_nonblank(value):
    if value is None or pd.isna(value):
        return False
    text = str(value).strip()
    if not text:
        return False
    return text.lower() not in {
        "nan",
        "none",
        "false",
        "0",
        "0.0",
        "-",
        "なし",
        "正常",
        "異常なし",
    }


def parse_water_it_csv(content):
    """WATER itのCSVを画面表示用に整形する。"""
    dataframe = None
    errors = []
    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            candidate = pd.read_csv(BytesIO(content), encoding=encoding)
            if len(candidate.columns) <= 1:
                raise ValueError("CSVの列を分割できませんでした。")
            dataframe = candidate
            break
        except Exception as exc:
            errors.append(f"{encoding}: {exc}")

    if dataframe is None:
        detail = " / ".join(errors[:2])
        raise RuntimeError(f"data.csvを読み込めませんでした。{detail}")

    dataframe.columns = [str(column).replace("\ufeff", "").strip() for column in dataframe.columns]
    missing = [column for column in WATER_IT_REQUIRED_COLUMNS if column not in dataframe.columns]
    if missing:
        raise RuntimeError(
            "data.csvに必要な列がありません：" + "、".join(missing)
        )

    dataframe = dataframe.copy()
    dataframe["測定日時_解析"] = pd.to_datetime(
        dataframe["測定日時"],
        errors="coerce",
    )

    number_translation = str.maketrans(
        "０１２３４５６７８９．，－＋",
        "0123456789.,-+",
    )
    number_text = (
        dataframe["測定値"]
        .astype(str)
        .str.translate(number_translation)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    dataframe["測定値_数値"] = pd.to_numeric(number_text, errors="coerce")
    dataframe["単位_表示"] = dataframe["単位"].apply(normalize_water_it_unit)
    dataframe["エリア"] = dataframe["エリア"].fillna("").astype(str).str.strip()
    dataframe["ポイント"] = dataframe["ポイント"].fillna("").astype(str).str.strip()
    dataframe["測定項目"] = dataframe["測定項目"].fillna("").astype(str).str.strip()

    dataframe = dataframe[
        dataframe["測定日時_解析"].notna()
        & dataframe["ポイント"].ne("")
        & dataframe["測定項目"].ne("")
    ].copy()
    dataframe.sort_values("測定日時_解析", ascending=False, inplace=True)
    dataframe.reset_index(drop=True, inplace=True)
    return dataframe


@st.cache_data(ttl=60, show_spinner=False)
def load_water_it_data():
    content, source = read_water_it_source_bytes()
    return parse_water_it_csv(content), source


def make_water_it_snapshot_payload(content, filename, dataframe):
    """検証済みCSVを圧縮し、Supabaseへ保存できるJSON文字列にする。"""
    latest_time = dataframe["測定日時_解析"].max()
    oldest_time = dataframe["測定日時_解析"].min()
    return json.dumps(
        {
            "version": WATER_IT_STORAGE_VERSION,
            "filename": str(filename or "data.csv"),
            "sha256": hashlib.sha256(content).hexdigest(),
            "csv_gzip_base64": base64.b64encode(gzip.compress(content)).decode("ascii"),
            "row_count": int(len(dataframe)),
            "point_count": int(dataframe["ポイント"].nunique()),
            "oldest_time": oldest_time.isoformat() if pd.notna(oldest_time) else None,
            "latest_time": latest_time.isoformat() if pd.notna(latest_time) else None,
            "imported_at": get_jst_now().isoformat(),
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )


def decode_water_it_snapshot_payload(payload_text):
    """Supabaseに保存したスナップショットから元のCSVバイト列を復元する。"""
    try:
        payload = json.loads(str(payload_text or ""))
    except Exception as exc:
        raise RuntimeError("保存済みWATER itデータの形式が正しくありません。") from exc
    if not isinstance(payload, dict):
        raise RuntimeError("保存済みWATER itデータの形式が正しくありません。")
    if int(payload.get("version", 0)) != WATER_IT_STORAGE_VERSION:
        raise RuntimeError("保存済みWATER itデータのバージョンが対応外です。")
    encoded = str(payload.get("csv_gzip_base64") or "")
    if not encoded:
        raise RuntimeError("保存済みWATER itデータにCSV本体がありません。")
    try:
        content = gzip.decompress(base64.b64decode(encoded.encode("ascii")))
    except Exception as exc:
        raise RuntimeError("保存済みWATER itデータを復元できませんでした。") from exc
    expected_hash = str(payload.get("sha256") or "")
    if expected_hash and hashlib.sha256(content).hexdigest() != expected_hash:
        raise RuntimeError("保存済みWATER itデータの検証に失敗しました。")
    return content, str(payload.get("filename") or "data.csv"), payload


def save_water_it_snapshot_to_supabase(content, filename, dataframe):
    """選択したCSVを既存Supabaseへ保存する。WATER itやExcelには書き込まない。"""
    if not has_supabase_config():
        raise RuntimeError("Supabase設定がないため、CSVを永続保存できません。")
    now = get_jst_now().isoformat()
    payload = {
        "id": WATER_IT_STORAGE_ID,
        "customer_key": None,
        "customer_name": WATER_IT_STORAGE_CUSTOMER,
        "field_name": WATER_IT_STORAGE_FIELD,
        "content": make_water_it_snapshot_payload(content, filename, dataframe),
        "sort_order": 0,
        "created_at": now,
        "updated_at": now,
    }
    try:
        response = requests.post(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(
                prefer="resolution=merge-duplicates,return=minimal"
            ),
            params={"on_conflict": "id"},
            json=payload,
            timeout=30,
        )
    except Exception as exc:
        raise RuntimeError("WATER itデータをSupabaseへ保存できませんでした。") from exc
    if response.status_code not in (200, 201):
        detail = str(response.text or "").strip()[:500]
        raise RuntimeError(
            f"WATER itデータをSupabaseへ保存できませんでした（{response.status_code}）。"
            + (f" {detail}" if detail else "")
        )
    load_saved_water_it_snapshot.clear()
    # 保存直後は、対象顧客の小さな照合索引も次回表示時に作り直す。
    try:
        load_persisted_water_it_customer_key_index.clear()
    except NameError:
        pass
    st.session_state.pop(WATER_IT_CUSTOMER_INDEX_SESSION_KEY, None)
    st.session_state.pop(WATER_IT_CUSTOMER_INDEX_HASH_SESSION_KEY, None)


@st.cache_data(ttl=30, show_spinner=False)
def load_saved_water_it_snapshot():
    """Supabaseから最後に取り込んだCSVを取得する。未保存ならNoneを返す。"""
    if not has_supabase_config():
        return None
    try:
        response = requests.get(
            get_supabase_customer_information_url(),
            headers=get_supabase_headers(),
            params={
                "select": "content,updated_at",
                "id": f"eq.{WATER_IT_STORAGE_ID}",
                "limit": "1",
            },
            timeout=20,
        )
    except Exception:
        return None
    if response.status_code != 200:
        return None
    try:
        rows = response.json()
    except Exception:
        return None
    if not isinstance(rows, list) or not rows:
        return None
    try:
        content, filename, metadata = decode_water_it_snapshot_payload(rows[0].get("content"))
        dataframe = parse_water_it_csv(content)
    except Exception:
        return None
    return {
        "content": content,
        "filename": filename,
        "metadata": metadata,
        "dataframe": dataframe,
        "updated_at": rows[0].get("updated_at"),
    }


@st.cache_resource(show_spinner=False)
def get_water_it_temporary_store():
    """アプリ再起動まで、選択CSVをサーバーの一時メモリに保持する。"""
    return {"content": None, "name": None, "hash": None}


def get_active_water_it_data():
    """選択中CSV、Supabase保存済みCSV、同梱data.csvの順で読み込む。"""
    uploaded_content = st.session_state.get(WATER_IT_UPLOAD_BYTES_KEY)
    uploaded_name = st.session_state.get(WATER_IT_UPLOAD_NAME_KEY)
    persisted = bool(st.session_state.get(WATER_IT_UPLOAD_PERSISTED_KEY))
    if not uploaded_content:
        temporary_store = get_water_it_temporary_store()
        uploaded_content = temporary_store.get("content")
        uploaded_name = temporary_store.get("name")
        persisted = bool(temporary_store.get("persisted"))
    if uploaded_content:
        uploaded_name = str(uploaded_name or "選択したCSV")
        label = "スマホから選択・Supabase保存済み" if persisted else "スマホから選択（一時）"
        return parse_water_it_csv(uploaded_content), f"{label}：{uploaded_name}"

    saved = load_saved_water_it_snapshot()
    if saved:
        return saved["dataframe"].copy(), f"Supabase保存：{saved['filename']}"

    return load_water_it_data()


def remember_uploaded_water_it_csv(uploaded_file):
    """選択されたCSVを検証し、Supabaseへ自動保存する。"""
    content = uploaded_file.getvalue()
    if not content:
        raise RuntimeError("選択したCSVが空です。")
    dataframe = parse_water_it_csv(content)
    digest = hashlib.sha256(content).hexdigest()
    uploaded_name = uploaded_file.name or "data.csv"

    persisted = False
    st.session_state.pop("water_it_persist_warning_message", None)
    try:
        save_water_it_snapshot_to_supabase(content, uploaded_name, dataframe)
        persisted = True
    except Exception as exc:
        st.session_state["water_it_persist_warning_message"] = str(exc)

    st.session_state[WATER_IT_UPLOAD_BYTES_KEY] = content
    st.session_state[WATER_IT_UPLOAD_NAME_KEY] = uploaded_name
    st.session_state[WATER_IT_UPLOAD_HASH_KEY] = digest
    st.session_state[WATER_IT_UPLOAD_PERSISTED_KEY] = persisted
    # 選択したCSVの対象顧客索引を同じセッション内で再利用する。
    st.session_state[WATER_IT_CUSTOMER_INDEX_HASH_SESSION_KEY] = digest
    st.session_state[WATER_IT_CUSTOMER_INDEX_SESSION_KEY] = (
        build_water_it_customer_key_index(dataframe)
    )
    temporary_store = get_water_it_temporary_store()
    temporary_store.update(
        {
            "content": content,
            "name": uploaded_name,
            "hash": digest,
            "persisted": persisted,
        }
    )
    label = "スマホから選択・Supabase保存済み" if persisted else "スマホから選択（一時）"
    return dataframe, f"{label}：{uploaded_name}"


def clear_uploaded_water_it_csv():
    for key in (
        WATER_IT_UPLOAD_BYTES_KEY,
        WATER_IT_UPLOAD_NAME_KEY,
        WATER_IT_UPLOAD_HASH_KEY,
        WATER_IT_UPLOAD_PERSISTED_KEY,
        "water_it_upload_success_message",
        "water_it_persist_warning_message",
        "water_it_upload_error_message",
    ):
        st.session_state.pop(key, None)
    temporary_store = get_water_it_temporary_store()
    temporary_store.update({"content": None, "name": None, "hash": None, "persisted": False})


def handle_water_it_mobile_upload(widget_key):
    """スマホのファイル選択完了直後にCSVを検証して保持する。"""
    st.session_state.pop("water_it_upload_success_message", None)
    st.session_state.pop("water_it_upload_error_message", None)
    uploaded_file = st.session_state.get(widget_key)
    if uploaded_file is None:
        return
    try:
        dataframe, source = remember_uploaded_water_it_csv(uploaded_file)
        latest_time = dataframe["測定日時_解析"].max()
        saved_text = (
            " Supabaseへ保存しました。"
            if st.session_state.get(WATER_IT_UPLOAD_PERSISTED_KEY)
            else " 一時表示には反映しました。"
        )
        st.session_state["water_it_upload_success_message"] = (
            f"{uploaded_file.name or '選択したファイル'} を受け取りました。"
            f" 最新測定日時：{latest_time.strftime('%Y/%m/%d %H:%M')}。"
            + saved_text
        )
    except Exception as exc:
        st.session_state["water_it_upload_error_message"] = str(exc)


def get_water_it_latest_rows(dataframe):
    if dataframe.empty:
        return dataframe.copy()
    keys = ["エリア", "ポイント", "測定項目"]
    return (
        dataframe.sort_values("測定日時_解析", ascending=False)
        .drop_duplicates(subset=keys, keep="first")
        .reset_index(drop=True)
    )


def normalize_water_it_customer_key(value):
    """WATER itのポイント名と顧客名を安全に照合するための最小正規化。

    表記ゆれを広く吸収すると別顧客を誤って結び付ける可能性があるため、
    Unicode正規化と空白除去だけを行う。
    """
    text = clean_value(value, blank_text="").strip()
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"[\s\u3000]+", "", text)
    return text.casefold()


def water_it_display_name(value):
    """WATER itの元データを変えず、画面上だけ名称を統一する。"""
    text = clean_value(value, blank_text="").strip()
    return WATER_IT_POINT_DISPLAY_NAMES.get(text, text)


def canonical_water_it_customer_key(value):
    """明示した別名だけを同一顧客として扱う。曖昧な部分一致は行わない。"""
    display_name = water_it_display_name(value)
    return normalize_water_it_customer_key(display_name)


def get_water_it_customer_rows(dataframe, customer_name):
    """顧客名と対応するWATER itポイントだけを返す（読み取り専用）。"""
    if dataframe is None or dataframe.empty:
        return dataframe.iloc[0:0].copy() if dataframe is not None else pd.DataFrame()
    target = canonical_water_it_customer_key(customer_name)
    if not target:
        return dataframe.iloc[0:0].copy()
    point_keys = dataframe["ポイント"].map(canonical_water_it_customer_key)
    return dataframe[point_keys == target].copy()

def build_water_it_customer_key_index(dataframe):
    """WATER it対象顧客の照合キーだけを小さなタプルにまとめる。"""
    if dataframe is None or dataframe.empty or "ポイント" not in dataframe.columns:
        return tuple()
    keys = {
        canonical_water_it_customer_key(value)
        for value in dataframe["ポイント"].dropna().tolist()
    }
    return tuple(sorted(key for key in keys if key))

@st.cache_data(ttl=WATER_IT_CUSTOMER_INDEX_TTL_SECONDS, show_spinner=False)
def load_persisted_water_it_customer_key_index():
    """保存済みCSVまたは同梱CSVから、対象顧客の照合キーだけを再利用する。"""
    saved = load_saved_water_it_snapshot()
    if saved:
        return build_water_it_customer_key_index(saved["dataframe"])
    dataframe, _ = load_water_it_data()
    return build_water_it_customer_key_index(dataframe)

def get_active_water_it_customer_key_index():
    """現在有効なWATER itデータの対象顧客キーだけを返す。"""
    uploaded_content = st.session_state.get(WATER_IT_UPLOAD_BYTES_KEY)
    uploaded_hash = st.session_state.get(WATER_IT_UPLOAD_HASH_KEY)
    if not uploaded_content:
        temporary_store = get_water_it_temporary_store()
        uploaded_content = temporary_store.get("content")
        uploaded_hash = temporary_store.get("hash")

    if uploaded_content:
        digest = str(uploaded_hash or hashlib.sha256(uploaded_content).hexdigest())
        cached_digest = st.session_state.get(WATER_IT_CUSTOMER_INDEX_HASH_SESSION_KEY)
        cached_keys = st.session_state.get(WATER_IT_CUSTOMER_INDEX_SESSION_KEY)
        if cached_digest == digest and isinstance(cached_keys, (tuple, list, set)):
            return set(cached_keys)
        keys = build_water_it_customer_key_index(parse_water_it_csv(uploaded_content))
        st.session_state[WATER_IT_CUSTOMER_INDEX_HASH_SESSION_KEY] = digest
        st.session_state[WATER_IT_CUSTOMER_INDEX_SESSION_KEY] = keys
        return set(keys)

    return set(load_persisted_water_it_customer_key_index())

def customer_has_water_it_data(customer_name):
    """対象顧客だけWATER it本体を読み込むための事前判定。"""
    target = canonical_water_it_customer_key(customer_name)
    if not target:
        return False
    try:
        return target in get_active_water_it_customer_key_index()
    except Exception:
        # 索引確認に失敗した場合は従来処理へ戻し、表示機会を失わない。
        return True



def render_customer_water_it_card(customer_name):
    """顧客詳細にWATER itの最新値を読み取り専用で表示する。

    ポイント名と顧客名が一致しない顧客には何も表示しない。
    ExcelやWATER itへの書き込み処理は行わない。
    """
    # 対象外の顧客では、WATER it本体・Supabase保存CSVを読み込まない。
    if not customer_has_water_it_data(customer_name):
        return

    try:
        dataframe, source = get_active_water_it_data()
    except Exception:
        # WATER it側が一時的に読めなくても、既存の顧客詳細は通常どおり表示する。
        return

    customer_rows = get_water_it_customer_rows(dataframe, customer_name)
    if customer_rows.empty:
        return

    latest_rows = get_water_it_latest_rows(customer_rows)
    if latest_rows.empty:
        return

    newest = latest_rows["測定日時_解析"].max()
    point_names = [
        water_it_display_name(value)
        for value in latest_rows["ポイント"].drop_duplicates().tolist()
    ]
    areas = [
        clean_value(value, blank_text="")
        for value in latest_rows["エリア"].drop_duplicates().tolist()
        if clean_value(value, blank_text="")
    ]

    st.markdown("---")
    with st.container(border=True):
        st.subheader("💧 WATER it タンク情報")
        st.caption(
            "読み取り専用表示です。ここからExcelやWATER itへの書き込みは行いません。"
        )
        st.caption(
            f"ポイント：{' / '.join(point_names)}"
            + (f"　｜　エリア：{' / '.join(areas)}" if areas else "")
            + f"　｜　最終受信：{newest.strftime('%Y/%m/%d %H:%M')}"
        )

        rows = list(latest_rows.iterrows())
        for start in range(0, len(rows), 3):
            group = rows[start:start + 3]
            columns = st.columns(len(group))
            for display_column, (_, row) in zip(columns, group):
                with display_column:
                    label = clean_value(row.get("測定項目"))
                    value = format_water_it_value(row.get("測定値_数値"))
                    unit = normalize_water_it_unit(row.get("単位_表示"))
                    st.metric(label, f"{value} {unit}".strip())
                    st.caption(row["測定日時_解析"].strftime("%m/%d %H:%M"))

        alert_messages = []
        for _, row in latest_rows.iterrows():
            item_name = clean_value(row.get("測定項目"))
            for alert in get_water_it_alerts(row):
                alert_messages.append(f"{item_name}｜{alert}")
        if alert_messages:
            st.warning(" / ".join(alert_messages))
        else:
            st.caption(f"状態：異常表示なし　｜　参照：{source}")


def format_water_it_value(value):
    if value is None or pd.isna(value):
        return "未設定"
    try:
        number = float(value)
    except Exception:
        return clean_value(value)
    if not math.isfinite(number):
        return "未設定"
    if number.is_integer():
        return f"{int(number):,}"
    return f"{number:,.2f}".rstrip("0").rstrip(".")


def get_water_it_alerts(row):
    alerts = []
    for column in WATER_IT_ALERT_COLUMNS:
        if column not in row.index:
            continue
        value = row.get(column)
        if water_it_nonblank(value):
            alerts.append(f"{column}: {clean_value(value)}")
    return alerts


def show_water_it_latest_cards(latest_rows):
    for point in latest_rows["ポイント"].drop_duplicates().tolist():
        point_rows = latest_rows[latest_rows["ポイント"] == point].copy()
        if point_rows.empty:
            continue
        first = point_rows.iloc[0]
        area = clean_value(first.get("エリア"), blank_text="未設定")
        newest = point_rows["測定日時_解析"].max()

        with st.container(border=True):
            st.subheader(f"💧 {water_it_display_name(point)}")
            st.caption(f"エリア：{area}　｜　最新：{newest.strftime('%Y/%m/%d %H:%M')}")

            rows = list(point_rows.iterrows())
            for start in range(0, len(rows), 3):
                group = rows[start:start + 3]
                columns = st.columns(len(group))
                for display_column, (_, row) in zip(columns, group):
                    with display_column:
                        label = clean_value(row.get("測定項目"))
                        value = format_water_it_value(row.get("測定値_数値"))
                        unit = normalize_water_it_unit(row.get("単位_表示"))
                        st.metric(label, f"{value} {unit}".strip())
                        st.caption(row["測定日時_解析"].strftime("%m/%d %H:%M"))

            alert_messages = []
            for _, row in point_rows.iterrows():
                item_name = clean_value(row.get("測定項目"))
                for alert in get_water_it_alerts(row):
                    alert_messages.append(f"{item_name}｜{alert}")
            if alert_messages:
                st.warning(" / ".join(alert_messages))
            else:
                st.caption("状態：異常表示なし")


def show_water_it_history(dataframe):
    st.markdown("### 測定履歴")
    points = dataframe["ポイント"].drop_duplicates().tolist()
    selected_point = st.selectbox(
        "ポイント",
        points,
        key="water_it_history_point",
        format_func=water_it_display_name,
    )
    point_rows = dataframe[dataframe["ポイント"] == selected_point].copy()
    items = point_rows["測定項目"].drop_duplicates().tolist()
    item_key_suffix = hashlib.sha1(selected_point.encode("utf-8")).hexdigest()[:12]
    selected_item = st.selectbox(
        "測定項目",
        items,
        key=f"water_it_history_item_{item_key_suffix}",
    )
    history = point_rows[point_rows["測定項目"] == selected_item].copy()
    history.sort_values("測定日時_解析", inplace=True)

    period = st.radio(
        "表示期間",
        ["24時間", "3日間", "7日間", "すべて"],
        horizontal=True,
        key="water_it_history_period",
    )
    if not history.empty and period != "すべて":
        hours = {"24時間": 24, "3日間": 72, "7日間": 168}[period]
        cutoff = history["測定日時_解析"].max() - timedelta(hours=hours)
        history = history[history["測定日時_解析"] >= cutoff].copy()

    chart_data = history.dropna(subset=["測定値_数値"]).set_index("測定日時_解析")[["測定値_数値"]]
    if chart_data.empty:
        st.info("グラフに表示できる数値データがありません。")
    else:
        chart_data = chart_data.rename(columns={"測定値_数値": selected_item})
        st.line_chart(chart_data, use_container_width=True)

    with st.expander("直近の測定値を表示"):
        display = history.sort_values("測定日時_解析", ascending=False).head(100).copy()
        display["測定日時"] = display["測定日時_解析"].dt.strftime("%Y/%m/%d %H:%M")
        display["測定値"] = display["測定値_数値"].apply(format_water_it_value)
        display["単位"] = display["単位_表示"]
        display["ポイント"] = display["ポイント"].map(water_it_display_name)
        st.dataframe(
            display[["測定日時", "エリア", "ポイント", "測定項目", "測定値", "単位"]],
            use_container_width=True,
            hide_index=True,
        )


def show_water_it_test_page():
    st.markdown("---")
    st.header("💧 WATER it CSV取込・保存")
    show_back_home_button("water_it_back_home")
    st.markdown(
        render_page_link("🧪 ソリュブル在庫", page="soluble_inventory"),
        unsafe_allow_html=True,
    )
    st.caption(
        "スマホでWATER itからCSVを手動ダウンロードし、そのCSVを選ぶだけで画面へ反映し、既存のSupabaseへ自動保存します。Excel・WATER it・Dropboxへの書き込みは行いません。"
    )

    st.link_button(
        "🌐 WATER itを開く",
        WATER_IT_LOGIN_URL,
        use_container_width=True,
    )

    with st.container(border=True):
        st.markdown("#### スマホでの手順")
        st.write("1. 上のボタンからWATER itを開いてログインします。")
        st.write("2. リスト画面の『ダウンロード』を押します。")
        st.write("3. このカルテへ戻り、下の『CSVを選ぶ』を押します。")
        st.write("4. ファイル画面で『最近使用したファイル』または『ダウンロード』を開き、一番新しいCSVを選びます。")

    uploader_version = int(st.session_state.get("water_it_uploader_version", 0))
    uploader_key = f"water_it_mobile_csv_uploader_{uploader_version}"

    temporary_store_before_upload = get_water_it_temporary_store()
    has_selected_water_it_csv = bool(
        st.session_state.get(WATER_IT_UPLOAD_BYTES_KEY)
        or temporary_store_before_upload.get("content")
    )
    if not has_selected_water_it_csv:
        if st.button(
            "CSV選択をリセット",
            key=f"water_it_reset_uploader_{uploader_version}",
            use_container_width=True,
        ):
            clear_uploaded_water_it_csv()
            st.session_state["water_it_uploader_version"] = uploader_version + 1
            st.rerun()

    uploaded_file = st.file_uploader(
        "ダウンロードしたファイルを選ぶ",
        type=None,
        accept_multiple_files=False,
        key=uploader_key,
        on_change=handle_water_it_mobile_upload,
        args=(uploader_key,),
        help=(
            "Androidでは『最近使用したファイル』または『ダウンロード』から選びます。"
            "端末によっては、ファイルをタップしたあとに『開く』『選択』『完了』または右上のチェックを押します。"
        ),
    )
    st.caption(
        "CSVだけに絞るとAndroidで選択が戻らない場合があるため、この版ではファイル種類を絞っていません。"
        "選択後に中身を確認し、WATER it形式のCSVだけを反映します。"
    )

    success_message = st.session_state.pop("water_it_upload_success_message", None)
    error_message = st.session_state.pop("water_it_upload_error_message", None)
    persist_warning = st.session_state.pop("water_it_persist_warning_message", None)
    if success_message:
        st.success(success_message)
    if persist_warning:
        st.warning("CSVは画面へ反映しましたが、Supabaseへの保存は完了していません。")
        st.write(persist_warning)
    if error_message:
        st.error("選択したファイルを読み込めませんでした。")
        st.write(error_message)
        st.info("WATER itのリスト画面からダウンロードしたCSVを選んでください。")

    dataframe = None
    source = ""

    if uploaded_file is not None:
        try:
            uploaded_hash = hashlib.sha256(uploaded_file.getvalue()).hexdigest()
            if uploaded_hash != st.session_state.get(WATER_IT_UPLOAD_HASH_KEY):
                with st.spinner("選択したファイルを確認しています…"):
                    dataframe, source = remember_uploaded_water_it_csv(uploaded_file)
            else:
                dataframe, source = get_active_water_it_data()
            st.caption(
                f"選択済み：{uploaded_file.name or '名前なし'}　"
                f"{len(uploaded_file.getvalue()):,} bytes"
            )
        except Exception as exc:
            st.error("選択したファイルを読み込めませんでした。")
            st.write(str(exc))
            st.info("WATER itのリスト画面からダウンロードしたCSVを選んでください。")
            return
    else:
        try:
            dataframe, source = get_active_water_it_data()
        except Exception as exc:
            st.info("まだCSVが選択されていません。まずWATER itからCSVをダウンロードし、上の欄から選んでください。")
            st.caption(str(exc))
            return

    temporary_store = get_water_it_temporary_store()
    if st.session_state.get(WATER_IT_UPLOAD_BYTES_KEY) or temporary_store.get("content"):
        button_col, note_col = st.columns([1, 2])
        with button_col:
            if st.button("選択中のCSVを解除", key="water_it_clear_upload"):
                clear_uploaded_water_it_csv()
                st.session_state["water_it_uploader_version"] = uploader_version + 1
                st.rerun()
        with note_col:
            st.caption("選択したCSVはSupabaseへ保存されます。ここで選択状態を解除しても、最後に保存したデータは残り、顧客詳細から引き続き確認できます。")

    if dataframe is None or dataframe.empty:
        st.warning("CSVに表示できる測定データがありません。")
        return

    latest_rows = get_water_it_latest_rows(dataframe)
    latest_time = dataframe["測定日時_解析"].max()
    oldest_time = dataframe["測定日時_解析"].min()

    st.success(f"読込OK　｜　参照：{source}")
    metric1, metric2, metric3 = st.columns(3)
    with metric1:
        st.metric("最新測定日時", latest_time.strftime("%Y/%m/%d %H:%M"))
    with metric2:
        st.metric("ポイント数", f"{dataframe['ポイント'].nunique()}件")
    with metric3:
        st.metric("読込行数", f"{len(dataframe):,}行")
    st.caption(
        f"データ期間：{oldest_time.strftime('%Y/%m/%d %H:%M')} ～ {latest_time.strftime('%Y/%m/%d %H:%M')}"
    )

    show_water_it_latest_cards(latest_rows)
    show_water_it_history(dataframe)


# =========================
# 仕入先・運送会社（取引先カルテ.xlsx）
# =========================
TRADE_PARTNER_HEADER_ALIASES = {
    "連絡方法レンラクホウホウ": "連絡方法",
    "納品先ノウヒンサキ": "納品先",
    "運賃ウンチン": "運賃",
    "地域チイキ": "地域",
}
TRADE_PARTNER_REQUIRED_SHEETS = (
    TRADE_PARTNER_MASTER_SHEET,
    TRADE_PARTNER_CONTACT_SHEET,
    TRADE_PARTNER_PRODUCT_SHEET,
    TRADE_PARTNER_TRANSPORT_SHEET,
)
TRADE_PARTNER_ID_FIELDS = {
    TRADE_PARTNER_MASTER_SHEET: "取引先ID",
    TRADE_PARTNER_CONTACT_SHEET: "担当者ID",
    TRADE_PARTNER_PRODUCT_SHEET: "仕入商品ID",
    TRADE_PARTNER_TRANSPORT_SHEET: "運送条件ID",
}
TRADE_PARTNER_PRIMARY_FIELDS = {
    TRADE_PARTNER_MASTER_SHEET: "会社名",
    TRADE_PARTNER_CONTACT_SHEET: "担当者名",
    TRADE_PARTNER_PRODUCT_SHEET: "商品名",
    TRADE_PARTNER_TRANSPORT_SHEET: "納品先",
}
TRADE_PARTNER_NOTE_PREFIXES = {
    "supplier": "【仕入先】",
    "carrier": "【運送会社】",
}
TRADE_PARTNER_IMMEDIATE_DATA_KEY = "trade_partner_immediate_data"
TRADE_PARTNER_IMMEDIATE_PARTNER_ID_KEY = "trade_partner_immediate_partner_id"


def normalize_trade_partner_header(value):
    text = clean_value(value, blank_text="").strip()
    return TRADE_PARTNER_HEADER_ALIASES.get(text, text)


def trade_partner_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y/%m/%d")
    text = str(value).strip()
    if text.startswith("="):
        return ""
    return text


def is_trade_partner_marked(value):
    text = trade_partner_text(value).strip().lower()
    return text in {"○", "〇", "1", "true", "yes", "有", "あり"}


def trade_partner_type_label(partner_type):
    return "仕入先" if partner_type == "supplier" else "運送会社"


def trade_partner_home_page(partner_type):
    return "supplier_home" if partner_type == "supplier" else "carrier_home"


def trade_partner_list_page(partner_type):
    return "supplier_list" if partner_type == "supplier" else "carrier_list"


def trade_partner_search_page(partner_type):
    return "supplier_search" if partner_type == "supplier" else "carrier_search"


def trade_partner_category_field(partner_type):
    return "仕入先区分" if partner_type == "supplier" else "運送会社区分"


def get_trade_partner_file_path():
    path = str(TRADE_PARTNER_DROPBOX_FILE_PATH or "").strip()
    return path or TRADE_PARTNER_DROPBOX_DEFAULT_FILE_PATH


XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XLSX_XML_NS = "http://www.w3.org/XML/1998/namespace"

# ElementTreeは、要素名・属性名で直接使われていない名前空間宣言を
# 再シリアライズ時に省略する。Excelはmc:Ignorableに記載された接頭辞の
# xmlns宣言が欠けるとワークシートを破損扱いにするため、宣言一覧を明示的に保持する。
XLSX_NAMESPACE_DECLARATIONS = {
    "": XLSX_MAIN_NS,
    "r": XLSX_REL_NS,
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "xm": "http://schemas.microsoft.com/office/excel/2006/main",
}

# Excelの拡張データ検証やリビジョン情報を壊さないよう、元と同じ名前空間接頭辞を保つ。
for _prefix, _namespace in XLSX_NAMESPACE_DECLARATIONS.items():
    try:
        ET.register_namespace(_prefix, _namespace)
    except Exception:
        pass


def xlsx_tag(local_name):
    return f"{{{XLSX_MAIN_NS}}}{local_name}"


def xlsx_column_name(column_number):
    result = ""
    number = int(column_number)
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def xlsx_column_number(cell_reference):
    match = re.match(r"^([A-Z]+)", str(cell_reference or "").upper())
    if not match:
        return 0
    result = 0
    for char in match.group(1):
        result = result * 26 + (ord(char) - 64)
    return result


def extract_worksheet_namespace_declarations(xml_content):
    """worksheetルートにあるxmlns宣言を接頭辞別に取り出す。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception:
        return {}
    root_match = re.search(r"<worksheet\b[^>]*>", text, re.DOTALL)
    if not root_match:
        return {}
    declarations = {}
    for prefix, uri in re.findall(
        r"\bxmlns(?::([A-Za-z_][\w.-]*))?=[\"']([^\"']+)[\"']",
        root_match.group(0),
    ):
        declarations[prefix or ""] = uri
    return declarations


def ensure_worksheet_namespace_declarations(xml_content, original_declarations=None):
    """
    元のxmlns宣言とmc:Ignorableが要求する宣言をworksheetへ戻す。

    ElementTreeはmc:Ignorableの値にだけ登場するx14ac/xr2/xr3などを
    未使用と判断して削るため、Excelで開く前に必ず宣言を復元する。
    """
    try:
        text = xml_content.decode("utf-8")
    except Exception as error:
        raise ValueError("ワークシートXMLをUTF-8として確認できません。") from error

    root_match = re.search(r"<worksheet\b[^>]*>", text, re.DOTALL)
    if not root_match:
        raise ValueError("ワークシートXMLのルート要素を確認できません。")

    root_tag = root_match.group(0)
    required = dict(original_declarations or {})
    ignorable_match = re.search(
        r"\bmc:Ignorable=[\"']([^\"']*)[\"']",
        root_tag,
    )
    if ignorable_match:
        for prefix in ignorable_match.group(1).split():
            uri = XLSX_NAMESPACE_DECLARATIONS.get(prefix)
            if uri:
                required.setdefault(prefix, uri)

    additions = []
    for prefix, uri in required.items():
        attribute_name = "xmlns" if not prefix else f"xmlns:{prefix}"
        if re.search(
            rf"\b{re.escape(attribute_name)}=[\"']",
            root_tag,
        ):
            continue
        additions.append(f' {attribute_name}="{uri}"')

    if additions:
        root_tag = root_tag[:-1] + "".join(additions) + ">"
        text = text[:root_match.start()] + root_tag + text[root_match.end():]

    return text.encode("utf-8")


def missing_worksheet_ignorable_namespaces(xml_content):
    """mc:Ignorableにあるのにxmlns宣言がない接頭辞を返す。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception:
        return ["XMLをUTF-8として読めません"]
    root_match = re.search(r"<worksheet\b[^>]*>", text, re.DOTALL)
    if not root_match:
        return ["worksheetルートがありません"]
    root_tag = root_match.group(0)
    ignorable_match = re.search(
        r"\bmc:Ignorable=[\"']([^\"']*)[\"']",
        root_tag,
    )
    if not ignorable_match:
        return []
    declared = extract_worksheet_namespace_declarations(xml_content)
    return [
        prefix
        for prefix in ignorable_match.group(1).split()
        if prefix not in declared
    ]


def remove_calc_chain_relationship(xml_content):
    """workbook.xml.relsから古いcalcChain参照だけを取り除く。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception as error:
        raise ValueError("Excelの計算関係情報を読み取れませんでした。") from error

    relationship_pattern = re.compile(
        r"<(?:[A-Za-z_][\w.-]*:)?Relationship\b"
        r"(?=[^>]*(?:"
        r"\bType\s*=\s*[\"'][^\"']*/calcChain[\"']"
        r"|\bTarget\s*=\s*[\"'][^\"']*calcChain\.xml[\"']"
        r"))[^>]*(?:/>|>\s*</(?:[A-Za-z_][\w.-]*:)?Relationship\s*>)",
        re.IGNORECASE | re.DOTALL,
    )
    return relationship_pattern.sub("", text).encode("utf-8")


def remove_calc_chain_content_type(xml_content):
    """[Content_Types].xmlからcalcChainの登録だけを取り除く。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception as error:
        raise ValueError("Excelのコンテンツ種類情報を読み取れませんでした。") from error

    override_pattern = re.compile(
        r"<(?:[A-Za-z_][\w.-]*:)?Override\b"
        r"(?=[^>]*\bPartName\s*=\s*[\"']/xl/calcChain\.xml[\"'])"
        r"[^>]*(?:/>|>\s*</(?:[A-Za-z_][\w.-]*:)?Override\s*>)",
        re.IGNORECASE | re.DOTALL,
    )
    return override_pattern.sub("", text).encode("utf-8")


def set_xml_tag_attribute(tag_text, attribute_name, value):
    """XML開始タグの既存属性を更新し、なければ末尾へ追加する。"""
    pattern = re.compile(
        rf"(\s{re.escape(attribute_name)}\s*=\s*)([\"'])(.*?)(\2)",
        re.IGNORECASE | re.DOTALL,
    )
    if pattern.search(tag_text):
        return pattern.sub(
            lambda match: match.group(1) + match.group(2) + str(value) + match.group(2),
            tag_text,
            count=1,
        )

    closing = "/>" if tag_text.rstrip().endswith("/>") else ">"
    position = tag_text.rfind(closing)
    if position < 0:
        raise ValueError("Excelの再計算設定を更新できませんでした。")
    return tag_text[:position] + f' {attribute_name}="{value}"' + tag_text[position:]


def force_workbook_recalculation(xml_content):
    """Excelを開いた時に数式を自動で全再計算する設定へ更新する。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception as error:
        raise ValueError("Excelのブック設定を読み取れませんでした。") from error

    calc_pattern = re.compile(
        r"<(?:[A-Za-z_][\w.-]*:)?calcPr\b[^>]*>",
        re.IGNORECASE | re.DOTALL,
    )
    match = calc_pattern.search(text)
    if match:
        tag = match.group(0)
        for name, value in (
            ("calcMode", "auto"),
            ("fullCalcOnLoad", "1"),
            ("forceFullCalc", "1"),
        ):
            tag = set_xml_tag_attribute(tag, name, value)
        text = text[:match.start()] + tag + text[match.end():]
    else:
        closing_match = re.search(
            r"</(?:[A-Za-z_][\w.-]*:)?workbook\s*>",
            text,
            re.IGNORECASE,
        )
        if not closing_match:
            raise ValueError("Excelのブック設定にworkbook要素が見つかりません。")
        calc_tag = (
            '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>'
        )
        text = text[:closing_match.start()] + calc_tag + text[closing_match.start():]

    return text.encode("utf-8")


def workbook_recalculation_is_forced(xml_content):
    """保存後のworkbook.xmlに全再計算設定があるか確認する。"""
    try:
        text = xml_content.decode("utf-8")
    except Exception:
        return False
    match = re.search(
        r"<(?:[A-Za-z_][\w.-]*:)?calcPr\b[^>]*>",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return False
    tag = match.group(0)
    required = {
        "calcMode": "auto",
        "fullCalcOnLoad": "1",
        "forceFullCalc": "1",
    }
    for name, expected in required.items():
        attribute = re.search(
            rf"\b{re.escape(name)}\s*=\s*[\"']([^\"']*)[\"']",
            tag,
            re.IGNORECASE,
        )
        if not attribute or attribute.group(1).lower() != expected.lower():
            return False
    return True


class TradePartnerXlsxEditor:
    """セル値だけをXMLで差し替え、入力規則・書式・数式をそのまま保持する。"""

    def __init__(self, content):
        self.original_infos = []
        self.parts = {}
        with zipfile.ZipFile(BytesIO(content), "r") as archive:
            self.original_infos = archive.infolist()
            for info in self.original_infos:
                self.parts[info.filename] = archive.read(info.filename)

        self.shared_strings = self._read_shared_strings()
        self.sheet_paths = self._read_sheet_paths()
        self.sheet_namespace_declarations = {
            path: extract_worksheet_namespace_declarations(self.parts[path])
            for path in self.sheet_paths.values()
            if path in self.parts
        }
        self.sheet_roots = {}
        self.changed_sheet_names = set()

    def _read_shared_strings(self):
        path = "xl/sharedStrings.xml"
        if path not in self.parts:
            return []
        root = ET.fromstring(self.parts[path])
        result = []
        for item in root.findall(xlsx_tag("si")):
            texts = []
            direct_text = item.find(xlsx_tag("t"))
            if direct_text is not None:
                texts.append(direct_text.text or "")
            for run in item.findall(xlsx_tag("r")):
                run_text = run.find(xlsx_tag("t"))
                if run_text is not None:
                    texts.append(run_text.text or "")
            result.append("".join(texts))
        return result

    def _read_sheet_paths(self):
        workbook_root = ET.fromstring(self.parts["xl/workbook.xml"])
        relation_root = ET.fromstring(self.parts["xl/_rels/workbook.xml.rels"])
        relation_map = {
            relation.attrib.get("Id"): relation.attrib.get("Target", "")
            for relation in relation_root.findall(f"{{{XLSX_PACKAGE_REL_NS}}}Relationship")
        }
        result = {}
        sheets = workbook_root.find(xlsx_tag("sheets"))
        if sheets is None:
            return result
        relation_attribute = f"{{{XLSX_REL_NS}}}id"
        for sheet in sheets.findall(xlsx_tag("sheet")):
            name = sheet.attrib.get("name", "")
            target = relation_map.get(sheet.attrib.get(relation_attribute), "")
            if not target:
                continue
            if target.startswith("/"):
                path = target.lstrip("/")
            else:
                path = posixpath.normpath(posixpath.join("xl", target))
            result[name] = path
        return result

    def has_sheet(self, sheet_name):
        return sheet_name in self.sheet_paths

    def get_sheet_root(self, sheet_name):
        if sheet_name not in self.sheet_paths:
            raise ValueError(f"{sheet_name}シートがありません。")
        if sheet_name not in self.sheet_roots:
            self.sheet_roots[sheet_name] = ET.fromstring(self.parts[self.sheet_paths[sheet_name]])
        return self.sheet_roots[sheet_name]

    def get_sheet_data(self, sheet_name):
        root = self.get_sheet_root(sheet_name)
        sheet_data = root.find(xlsx_tag("sheetData"))
        if sheet_data is None:
            sheet_data = ET.SubElement(root, xlsx_tag("sheetData"))
        return sheet_data

    def _display_inline_string(self, container):
        if container is None:
            return ""
        texts = []
        direct_text = container.find(xlsx_tag("t"))
        if direct_text is not None:
            texts.append(direct_text.text or "")
        for run in container.findall(xlsx_tag("r")):
            run_text = run.find(xlsx_tag("t"))
            if run_text is not None:
                texts.append(run_text.text or "")
        return "".join(texts)

    def cell_value_from_element(self, cell):
        if cell is None:
            return None
        formula = cell.find(xlsx_tag("f"))
        if formula is not None:
            return "=" + (formula.text or "")
        cell_type = cell.attrib.get("t", "")
        if cell_type == "inlineStr":
            return self._display_inline_string(cell.find(xlsx_tag("is")))
        value_element = cell.find(xlsx_tag("v"))
        raw_value = value_element.text if value_element is not None else None
        if raw_value is None:
            return None
        if cell_type == "s":
            try:
                return self.shared_strings[int(raw_value)]
            except Exception:
                return ""
        if cell_type == "b":
            return raw_value == "1"
        if cell_type in {"str", "e"}:
            return raw_value
        try:
            number = float(raw_value)
            return int(number) if number.is_integer() else number
        except Exception:
            return raw_value

    def get_row_element(self, sheet_name, row_number, create=False):
        sheet_data = self.get_sheet_data(sheet_name)
        target = int(row_number)
        rows = list(sheet_data.findall(xlsx_tag("row")))
        for row in rows:
            if int(row.attrib.get("r", "0") or 0) == target:
                return row
        if not create:
            return None
        new_row = ET.Element(xlsx_tag("row"), {"r": str(target)})
        inserted = False
        for index, row in enumerate(rows):
            current = int(row.attrib.get("r", "0") or 0)
            if current > target:
                sheet_data.insert(index, new_row)
                inserted = True
                break
        if not inserted:
            sheet_data.append(new_row)
        self.changed_sheet_names.add(sheet_name)
        return new_row

    def get_cell_element(self, sheet_name, row_number, column_number, create=False):
        row = self.get_row_element(sheet_name, row_number, create=create)
        if row is None:
            return None
        reference = f"{xlsx_column_name(column_number)}{int(row_number)}"
        cells = list(row.findall(xlsx_tag("c")))
        for cell in cells:
            if cell.attrib.get("r") == reference:
                return cell
        if not create:
            return None
        new_cell = ET.Element(xlsx_tag("c"), {"r": reference})
        inserted = False
        target_column = int(column_number)
        for index, cell in enumerate(cells):
            if xlsx_column_number(cell.attrib.get("r")) > target_column:
                row.insert(index, new_cell)
                inserted = True
                break
        if not inserted:
            row.append(new_cell)
        self.changed_sheet_names.add(sheet_name)
        return new_cell

    def get_cell_value(self, sheet_name, row_number, column_number):
        return self.cell_value_from_element(
            self.get_cell_element(sheet_name, row_number, column_number, create=False)
        )

    def set_cell_value(self, sheet_name, row_number, column_number, value):
        cell = self.get_cell_element(sheet_name, row_number, column_number, create=True)
        for child_name in ("f", "v", "is"):
            child = cell.find(xlsx_tag(child_name))
            if child is not None:
                cell.remove(child)
        if value is None or value == "":
            cell.attrib.pop("t", None)
        elif isinstance(value, bool):
            cell.attrib["t"] = "b"
            ET.SubElement(cell, xlsx_tag("v")).text = "1" if value else "0"
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.attrib.pop("t", None)
            ET.SubElement(cell, xlsx_tag("v")).text = str(value)
        else:
            text = value.strftime("%Y/%m/%d") if isinstance(value, (datetime, date)) else str(value)
            cell.attrib["t"] = "inlineStr"
            inline = ET.SubElement(cell, xlsx_tag("is"))
            text_element = ET.SubElement(inline, xlsx_tag("t"))
            if text != text.strip() or "\n" in text:
                text_element.attrib[f"{{{XLSX_XML_NS}}}space"] = "preserve"
            text_element.text = text
        self.changed_sheet_names.add(sheet_name)

    def get_max_row(self, sheet_name):
        maximum = 1
        for row in self.get_sheet_data(sheet_name).findall(xlsx_tag("row")):
            try:
                maximum = max(maximum, int(row.attrib.get("r", "0") or 0))
            except Exception:
                pass
        return maximum

    def get_header_map(self, sheet_name):
        result = {}
        row = self.get_row_element(sheet_name, 1, create=False)
        if row is None:
            return result
        for cell in row.findall(xlsx_tag("c")):
            column = xlsx_column_number(cell.attrib.get("r"))
            header = normalize_trade_partner_header(self.cell_value_from_element(cell))
            if header and header not in result:
                result[header] = column
        return result

    def read_sheet(self, sheet_name):
        header_map = self.get_header_map(sheet_name)
        headers = list(header_map.keys())
        rows = []
        for row_number in range(2, self.get_max_row(sheet_name) + 1):
            row = {
                header: self.get_cell_value(sheet_name, row_number, column)
                for header, column in header_map.items()
            }
            row["_row_number"] = row_number
            rows.append(row)
        return {"headers": headers, "rows": rows}

    def validate_worksheet_namespaces(self):
        problems = []
        for sheet_name, path in self.sheet_paths.items():
            if path not in self.parts:
                problems.append(f"{sheet_name}: XMLがありません")
                continue
            missing = missing_worksheet_ignorable_namespaces(self.parts[path])
            if missing:
                problems.append(f"{sheet_name}: " + "、".join(missing))
        if problems:
            raise ValueError(
                "Excel互換性の確認で名前空間宣言が不足しています。保存を中止しました。\n"
                + "\n".join(problems)
            )

    def remove_stale_calculation_chain(self):
        """セル変更前のcalcChainを除去し、Excelへ安全に再計算させる。"""
        self.parts.pop("xl/calcChain.xml", None)

        relationships_path = "xl/_rels/workbook.xml.rels"
        if relationships_path in self.parts:
            self.parts[relationships_path] = remove_calc_chain_relationship(
                self.parts[relationships_path]
            )

        content_types_path = "[Content_Types].xml"
        if content_types_path in self.parts:
            self.parts[content_types_path] = remove_calc_chain_content_type(
                self.parts[content_types_path]
            )

        workbook_path = "xl/workbook.xml"
        if workbook_path not in self.parts:
            raise ValueError("Excelのworkbook.xmlが見つかりません。")
        self.parts[workbook_path] = force_workbook_recalculation(
            self.parts[workbook_path]
        )

    def validate_calculation_state(self):
        """古いcalcChain参照が残らず、全再計算設定が有効か確認する。"""
        if "xl/calcChain.xml" in self.parts:
            raise ValueError("Excelの古い計算順序情報が残っています。保存を中止しました。")

        relationships = self.parts.get("xl/_rels/workbook.xml.rels", b"")
        if re.search(
            rb'(?:relationships/calcChain|Target\s*=\s*["\'][^"\']*calcChain\.xml)',
            relationships,
            re.IGNORECASE,
        ):
            raise ValueError("Excelの計算順序への参照が残っています。保存を中止しました。")

        content_types = self.parts.get("[Content_Types].xml", b"")
        if re.search(rb"/xl/calcChain\.xml", content_types, re.IGNORECASE):
            raise ValueError("Excelの計算順序の種類登録が残っています。保存を中止しました。")

        workbook = self.parts.get("xl/workbook.xml", b"")
        if not workbook_recalculation_is_forced(workbook):
            raise ValueError("Excelの自動再計算設定を確認できません。保存を中止しました。")

    def to_bytes(self):
        for sheet_name in self.changed_sheet_names:
            path = self.sheet_paths[sheet_name]
            serialized = ET.tostring(
                self.get_sheet_root(sheet_name),
                encoding="utf-8",
                xml_declaration=True,
            )
            self.parts[path] = ensure_worksheet_namespace_declarations(
                serialized,
                self.sheet_namespace_declarations.get(path),
            )

        # 以前のアプリ保存で名前空間宣言が欠けたファイルも、次の保存時に
        # 全ワークシートを安全な状態へ戻す。セル値や書式・数式は変更しない。
        for path in set(self.sheet_paths.values()):
            if path not in self.parts:
                continue
            self.parts[path] = ensure_worksheet_namespace_declarations(
                self.parts[path],
                self.sheet_namespace_declarations.get(path),
            )

        self.validate_worksheet_namespaces()

        # セルを書き換えた後に古いcalcChainを残すと、Excelが修復画面を出す。
        # 本体・関連付け・Content Typesの3か所をそろえて除去し、開いた時に
        # Excel自身が数式を全再計算する設定へ更新する。
        self.remove_stale_calculation_chain()
        self.validate_calculation_state()

        output = BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            for info in self.original_infos:
                # calcChain.xmlは意図的に削除しているため、元のZIP一覧にあっても書き戻さない。
                if info.filename not in self.parts:
                    continue
                archive.writestr(info, self.parts[info.filename])
        return output.getvalue()


def read_trade_partner_data_from_editor(editor):
    """検証済みの取引先カルテから、画面表示用データを作る。"""
    missing = [name for name in TRADE_PARTNER_REQUIRED_SHEETS if not editor.has_sheet(name)]
    if missing:
        raise RuntimeError("取引先カルテ.xlsxに必要なシートがありません：" + "、".join(missing))
    return {name: editor.read_sheet(name) for name in TRADE_PARTNER_REQUIRED_SHEETS}


def clear_trade_partner_immediate_data():
    """保存直後だけ使う取引先カルテの最新データを破棄する。"""
    st.session_state.pop(TRADE_PARTNER_IMMEDIATE_DATA_KEY, None)
    st.session_state.pop(TRADE_PARTNER_IMMEDIATE_PARTNER_ID_KEY, None)


def store_trade_partner_immediate_data(data, partner_id):
    """Dropbox反映待ちや古いキャッシュに左右されないよう、保存済みデータを保持する。"""
    target = trade_partner_text(partner_id)
    if not target or not isinstance(data, dict):
        clear_trade_partner_immediate_data()
        return
    st.session_state[TRADE_PARTNER_IMMEDIATE_DATA_KEY] = data
    st.session_state[TRADE_PARTNER_IMMEDIATE_PARTNER_ID_KEY] = target


def get_trade_partner_detail_data(partner_id):
    """同じ取引先の保存直後データがあれば優先し、なければDropboxから読む。"""
    target = trade_partner_text(partner_id)
    immediate_id = trade_partner_text(
        st.session_state.get(TRADE_PARTNER_IMMEDIATE_PARTNER_ID_KEY)
    )
    immediate_data = st.session_state.get(TRADE_PARTNER_IMMEDIATE_DATA_KEY)
    if target and immediate_id == target and isinstance(immediate_data, dict):
        return immediate_data
    if immediate_data is not None or immediate_id:
        clear_trade_partner_immediate_data()
    return load_trade_partner_data()


def verify_trade_partner_saved_result(data, result):
    """保存後データに、対象行と変更値が実在することを確認する。"""
    if not isinstance(result, dict):
        raise RuntimeError("取引先カルテの保存結果を確認できませんでした。")
    sheet_name = result.get("sheet_name")
    record_id = trade_partner_text(result.get("record_id"))
    if sheet_name not in TRADE_PARTNER_ID_FIELDS or not record_id:
        raise RuntimeError("取引先カルテの保存対象を確認できませんでした。")

    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]
    saved_row = next(
        (
            row
            for row in data[sheet_name]["rows"]
            if trade_partner_text(row.get(id_field)) == record_id
        ),
        None,
    )
    if saved_row is None:
        raise RuntimeError(
            f"保存後の確認で{sheet_name}の対象ID（{record_id}）が見つかりません。"
        )

    for header, change in (result.get("changes") or {}).items():
        if not isinstance(change, (tuple, list)) or len(change) < 2:
            continue
        expected = change[1]
        actual = saved_row.get(header)
        if not same_excel_value(actual, expected):
            raise RuntimeError(
                f"保存後の確認で{sheet_name}の「{header}」が更新されていません。"
            )


@st.cache_data(ttl=60, show_spinner=False)
def load_trade_partner_data():
    access_token = get_dropbox_access_token()
    path = get_trade_partner_file_path()
    content, response = download_dropbox_file(path, access_token)
    if content is None:
        raise RuntimeError(
            "取引先カルテ.xlsxをDropboxから取得できませんでした。\n"
            + dropbox_error_text(response)
        )

    editor = TradePartnerXlsxEditor(content)
    return read_trade_partner_data_from_editor(editor)

def get_trade_partner_master_rows(data, partner_type=None):
    rows = []
    for row in data[TRADE_PARTNER_MASTER_SHEET]["rows"]:
        if not trade_partner_text(row.get("会社名")):
            continue
        if partner_type and not is_trade_partner_marked(row.get(trade_partner_category_field(partner_type))):
            continue
        rows.append(row)
    return rows


def get_trade_partner_by_id(data, partner_id):
    target = str(partner_id or "").strip()
    for row in get_trade_partner_master_rows(data):
        if trade_partner_text(row.get("取引先ID")) == target:
            return row
    return None


def get_trade_partner_related_rows(data, sheet_name, partner_id):
    target = str(partner_id or "").strip()
    primary_field = TRADE_PARTNER_PRIMARY_FIELDS[sheet_name]
    result = []
    for row in data[sheet_name]["rows"]:
        if trade_partner_text(row.get("取引先ID")) != target:
            continue
        if not trade_partner_text(row.get(primary_field)):
            continue
        result.append(row)
    return result


def trade_partner_sort_key(row):
    kana = trade_partner_text(row.get("会社名かな"))
    company = trade_partner_text(row.get("会社名"))
    return (kana or company, company)


def make_trade_partner_note_key(partner_type, partner_id, company_name=None):
    """会社名が変わってもメモが外れないよう、区分と取引先IDだけで紐づける。"""
    prefix = TRADE_PARTNER_NOTE_PREFIXES.get(partner_type, "【取引先】")
    return f"{prefix}{partner_id}"


def parse_trade_partner_note_key(value):
    text = trade_partner_text(value)
    for partner_type, prefix in TRADE_PARTNER_NOTE_PREFIXES.items():
        if text.startswith(prefix):
            body = text[len(prefix):]
            partner_id, separator, company = body.partition("|")
            return {
                "partner_type": partner_type,
                "partner_id": partner_id.strip(),
                "company_name": company.strip() if separator else "",
            }
    return None


def ensure_trade_partner_backup_folder(access_token):
    response = call_dropbox_rpc(
        "files/create_folder_v2",
        {"path": TRADE_PARTNER_BACKUP_FOLDER, "autorename": False},
        access_token,
    )
    if response.status_code == 200:
        return
    try:
        summary = str(response.json().get("error_summary", "")).lower()
    except Exception:
        summary = str(getattr(response, "text", "")).lower()
    if "conflict" in summary and "folder" in summary:
        return
    raise RuntimeError(
        "取引先カルテのバックアップフォルダを作成できませんでした。\n"
        + dropbox_error_text(response)
    )


def create_trade_partner_backup(target_path, backup_path, original_content, access_token):
    ensure_trade_partner_backup_folder(access_token)
    copy_response = copy_dropbox_file(target_path, backup_path, access_token)
    if copy_response.status_code == 200:
        metadata = get_dropbox_response_metadata(copy_response)
        if not metadata.get("content_hash") or metadata.get("size") is None:
            metadata = get_dropbox_file_metadata(backup_path, access_token)
        try:
            verify_dropbox_file_metadata(metadata, original_content)
            return
        except Exception:
            call_dropbox_rpc("files/delete_v2", {"path": backup_path}, access_token)
            raise RuntimeError(
                "取引先カルテ.xlsxが別の端末で更新された可能性があります。再読み込みしてやり直してください。"
            )

    backup_response = upload_dropbox_file(
        backup_path,
        original_content,
        access_token,
        mode="add",
    )
    if backup_response.status_code != 200:
        raise RuntimeError(
            "取引先カルテのバックアップを作成できないため、本番ファイルは更新しません。\n"
            + dropbox_error_text(backup_response)
        )
    metadata = get_dropbox_response_metadata(backup_response)
    if not metadata.get("content_hash") or metadata.get("size") is None:
        metadata = get_dropbox_file_metadata(backup_path, access_token)
    verify_dropbox_file_metadata(metadata, original_content)


def trim_trade_partner_backups(access_token, keep=30):
    response = call_dropbox_rpc(
        "files/list_folder",
        {"path": TRADE_PARTNER_BACKUP_FOLDER, "recursive": False, "include_deleted": False},
        access_token,
    )
    if response.status_code != 200:
        return
    try:
        entries = list(response.json().get("entries", []))
    except Exception:
        return
    pattern = re.compile(r"^取引先カルテ_\d{8}_\d{6}_\d+\.xlsx$")
    files = [item for item in entries if pattern.match(str(item.get("name", "")))]
    files.sort(key=lambda item: str(item.get("server_modified", "")), reverse=True)
    for item in files[keep:]:
        path = item.get("path_lower") or item.get("path_display")
        if path:
            call_dropbox_rpc("files/delete_v2", {"path": path}, access_token)


def save_trade_partner_workbook(mutator):
    access_token = get_dropbox_access_token()
    target_path = get_trade_partner_file_path()
    original_content, download_response = download_dropbox_file(target_path, access_token)
    if original_content is None:
        raise RuntimeError(
            "最新の取引先カルテ.xlsxを取得できませんでした。\n"
            + dropbox_error_text(download_response)
        )
    revision = get_download_revision(download_response)
    if not revision:
        raise RuntimeError("Dropboxの更新番号を確認できないため、安全のため保存を中止しました。")

    timestamp = get_jst_now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = f"{TRADE_PARTNER_BACKUP_FOLDER}/取引先カルテ_{timestamp}.xlsx"
    create_trade_partner_backup(
        target_path,
        backup_path,
        original_content,
        access_token,
    )

    editor = TradePartnerXlsxEditor(original_content)
    missing = [name for name in TRADE_PARTNER_REQUIRED_SHEETS if not editor.has_sheet(name)]
    if missing:
        raise ValueError("必要なシートがありません：" + "、".join(missing))
    result = mutator(editor)
    saved_content = editor.to_bytes()

    # XML更新後もブック構造と入力規則の拡張部分が残っていることを確認する。
    verified = TradePartnerXlsxEditor(saved_content)
    missing = [name for name in TRADE_PARTNER_REQUIRED_SHEETS if not verified.has_sheet(name)]
    if missing:
        raise ValueError("保存後の検証で必要なシートがありません：" + "、".join(missing))
    for sheet_name in TRADE_PARTNER_REQUIRED_SHEETS:
        if not verified.get_header_map(sheet_name):
            raise ValueError(f"保存後の検証で{sheet_name}の見出しを確認できません。")
    latest_data = read_trade_partner_data_from_editor(verified)
    verify_trade_partner_saved_result(latest_data, result)

    upload_response = upload_dropbox_file(
        target_path,
        saved_content,
        access_token,
        mode="update",
        rev=revision,
    )
    if upload_response.status_code != 200:
        raise RuntimeError(
            "取引先カルテ.xlsxを更新できませんでした。\n"
            + dropbox_error_text(upload_response)
        )
    metadata = get_dropbox_response_metadata(upload_response)
    if not metadata.get("content_hash") or metadata.get("size") is None:
        metadata = get_dropbox_file_metadata(target_path, access_token)
    verify_dropbox_file_metadata(metadata, saved_content, previous_revision=revision)
    trim_trade_partner_backups(access_token, keep=30)
    load_trade_partner_data.clear()
    store_trade_partner_immediate_data(
        latest_data,
        result.get("partner_id") or result.get("record_id"),
    )
    return result

def trade_partner_input_value(header, value):
    text = str(value or "").strip()
    if not text:
        return None
    if header in {"単価", "運賃"}:
        normalized = text.replace(",", "").translate(
            str.maketrans("０１２３４５６７８９．－", "0123456789.-")
        )
        try:
            number = float(normalized)
            return int(number) if number.is_integer() else number
        except Exception:
            return text
    return text


def update_trade_partner_row(sheet_name, record_id, values):
    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]
    target_id = str(record_id or "").strip()

    def mutator(editor):
        header_map = editor.get_header_map(sheet_name)
        if id_field not in header_map:
            raise ValueError(f"{sheet_name}に{id_field}列がありません。")
        target_row = None
        for row_number in range(2, editor.get_max_row(sheet_name) + 1):
            current_id = trade_partner_text(
                editor.get_cell_value(sheet_name, row_number, header_map[id_field])
            )
            if current_id == target_id:
                target_row = row_number
                break
        if target_row is None:
            raise ValueError(f"{sheet_name}で対象IDが見つかりません。")

        # 取引先マスターの会社名を変更する場合は、区分に関係なく
        # マスター全体で同名が存在しないか確認する。編集中の行は除外する。
        if sheet_name == TRADE_PARTNER_MASTER_SHEET and "会社名" in values:
            if "会社名" not in header_map:
                raise ValueError("取引先マスターに会社名列がありません。")
            old_company = editor.get_cell_value(
                sheet_name, target_row, header_map["会社名"]
            )
            new_company = trade_partner_input_value("会社名", values.get("会社名"))
            if not same_excel_value(old_company, new_company):
                normalized_company = normalize_match_value(new_company)
                if normalized_company:
                    for check_row in range(2, editor.get_max_row(sheet_name) + 1):
                        if check_row == target_row:
                            continue
                        existing = trade_partner_text(
                            editor.get_cell_value(
                                sheet_name, check_row, header_map["会社名"]
                            )
                        )
                        if (
                            existing
                            and normalize_match_value(existing) == normalized_company
                        ):
                            raise ValueError("同じ会社名がすでに登録されています。")

        changes = {}
        for header, value in values.items():
            if header not in header_map or header == id_field or header == "会社名（確認用）":
                continue
            old_value = editor.get_cell_value(sheet_name, target_row, header_map[header])
            new_value = trade_partner_input_value(header, value)
            if not same_excel_value(old_value, new_value):
                editor.set_cell_value(sheet_name, target_row, header_map[header], new_value)
                changes[header] = (old_value, new_value)
        if not changes:
            raise ValueError("変更された項目がありません。")
        partner_id = target_id
        if sheet_name != TRADE_PARTNER_MASTER_SHEET and "取引先ID" in header_map:
            partner_id = trade_partner_text(
                editor.get_cell_value(sheet_name, target_row, header_map["取引先ID"])
            )
        return {
            "sheet_name": sheet_name,
            "record_id": target_id,
            "partner_id": partner_id,
            "changed": len(changes),
            "changes": changes,
        }

    return save_trade_partner_workbook(mutator)


def find_trade_partner_empty_row(editor, sheet_name, header_map, primary_field):
    if primary_field not in header_map:
        raise ValueError(f"{sheet_name}に{primary_field}列がありません。")
    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]
    if id_field not in header_map:
        raise ValueError(f"{sheet_name}に{id_field}列がありません。")
    for row_number in range(2, editor.get_max_row(sheet_name) + 1):
        primary = trade_partner_text(
            editor.get_cell_value(sheet_name, row_number, header_map[primary_field])
        )
        record_id = trade_partner_text(
            editor.get_cell_value(sheet_name, row_number, header_map[id_field])
        )
        if not primary and record_id:
            return row_number, record_id
    raise ValueError(
        f"{sheet_name}に登録用の空き行がありません。ExcelでID付きの空き行を追加してください。"
    )


def create_trade_partner_record(sheet_name, values):
    primary_field = TRADE_PARTNER_PRIMARY_FIELDS[sheet_name]
    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]

    def mutator(editor):
        header_map = editor.get_header_map(sheet_name)
        row_number, record_id = find_trade_partner_empty_row(
            editor,
            sheet_name,
            header_map,
            primary_field,
        )
        if sheet_name == TRADE_PARTNER_MASTER_SHEET:
            company = trade_partner_text(values.get("会社名"))
            if not company:
                raise ValueError("会社名を入力してください。")
            if "会社名" not in header_map:
                raise ValueError("取引先マスターに会社名列がありません。")
            for check_row in range(2, editor.get_max_row(sheet_name) + 1):
                existing = trade_partner_text(
                    editor.get_cell_value(sheet_name, check_row, header_map["会社名"])
                )
                if existing and normalize_match_value(existing) == normalize_match_value(company):
                    raise ValueError("同じ会社名がすでに登録されています。")
        changes = {}
        for header, value in values.items():
            if header not in header_map or header in {id_field, "会社名（確認用）"}:
                continue
            new_value = trade_partner_input_value(header, value)
            editor.set_cell_value(
                sheet_name,
                row_number,
                header_map[header],
                new_value,
            )
            if new_value not in (None, ""):
                changes[header] = ("", new_value)
        partner_id = record_id
        if sheet_name != TRADE_PARTNER_MASTER_SHEET:
            partner_id = trade_partner_text(values.get("取引先ID"))
        return {
            "sheet_name": sheet_name,
            "record_id": record_id,
            "partner_id": partner_id,
            "row_number": row_number,
            "changes": changes,
        }

    return save_trade_partner_workbook(mutator)

def show_top_home_link():
    st.markdown(render_page_link("← トップへ戻る", page="home"), unsafe_allow_html=True)


def show_trade_partner_home_link(partner_type):
    st.markdown(
        render_page_link(
            f"← {trade_partner_type_label(partner_type)}メニューへ戻る",
            page=trade_partner_home_page(partner_type),
        ),
        unsafe_allow_html=True,
    )


def trade_partner_detail_link(row, partner_type, label=None, class_name="dispatch-month-link"):
    partner_id = trade_partner_text(row.get("取引先ID"))
    company = trade_partner_text(row.get("会社名"))
    return render_page_link(
        label or company,
        page="partner_detail",
        partner_id=partner_id,
        partner_type=partner_type,
        class_name=class_name,
    )


def render_trade_partner_directory_cards(rows, partner_type):
    """仕入先一覧・検索結果を、顧客名一覧と同じ押しやすいカードで表示する。"""
    parts = ['<div class="customer-directory">']
    for row in rows:
        partner_id = trade_partner_text(row.get("取引先ID"))
        company = trade_partner_text(row.get("会社名")) or "名称未設定"
        region = trade_partner_text(row.get("地域")) or "未設定"
        url = html.escape(
            make_app_url(
                page="partner_detail",
                partner_id=partner_id,
                partner_type=partner_type,
            ),
            quote=True,
        )
        parts.append(
            (
                f'<a class="customer-directory-item" href="{url}" target="_self">'
                f'<span class="customer-directory-name">{html.escape(company)}</span>'
                f'<span class="customer-directory-meta">地域：{html.escape(region)}</span>'
                '</a>'
            )
        )
    parts.append("</div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


def show_trade_partner_home(partner_type):
    show_top_home_link()
    label = trade_partner_type_label(partner_type)
    icon = "🏢" if partner_type == "supplier" else "🚚"
    st.header(f"{icon} {label}")

    register_page = "supplier_register" if partner_type == "supplier" else "carrier_register"
    st.markdown(
        render_page_link(f"＋ 新しい{label}を登録", page=register_page),
        unsafe_allow_html=True,
    )
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(
            render_page_link(f"📋 {label}一覧", page=trade_partner_list_page(partner_type)),
            unsafe_allow_html=True,
        )
    with col2:
        st.markdown(
            render_page_link(f"🔍 {label}検索", page=trade_partner_search_page(partner_type)),
            unsafe_allow_html=True,
        )

    if partner_type == "supplier":
        st.markdown(
            render_page_link("📦 商品検索", page="supplier_product"),
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            render_page_link("💰 運賃比較", page="carrier_freight_compare"),
            unsafe_allow_html=True,
        )


def show_trade_partner_directory(partner_type):
    show_trade_partner_home_link(partner_type)
    if partner_type == "supplier":
        st.markdown(
            render_page_link("＋ 新しい仕入先を登録", page="supplier_register"),
            unsafe_allow_html=True,
        )
    label = trade_partner_type_label(partner_type)
    st.header(f"📋 {label}一覧")
    data = load_trade_partner_data()
    rows = sorted(get_trade_partner_master_rows(data, partner_type), key=trade_partner_sort_key)
    if not rows:
        st.info(f"登録されている{label}はありません。")
        return
    st.caption(f"{len(rows)}件")
    render_trade_partner_directory_cards(rows, partner_type)


@st.cache_data(ttl=300, show_spinner=False)
def load_trade_partner_search_index(partner_type):
    """仕入先・運送会社検索用の文字列を作り、5分間再利用する。"""
    data = load_trade_partner_data()
    rows = []
    for source in get_trade_partner_master_rows(data, partner_type):
        row = dict(source)
        haystack = " ".join(
            trade_partner_text(row.get(field))
            for field in ("会社名", "会社名かな", "地域")
        )
        row["_検索文字"] = normalize_match_value(haystack).casefold()
        rows.append(row)
    return rows


def show_trade_partner_search(partner_type):
    show_trade_partner_home_link(partner_type)
    label = trade_partner_type_label(partner_type)
    st.header(f"🔍 {label}検索")
    default_keyword = str(get_query_value("partner_search", "")).strip()
    keyword = st.text_input(
        "会社名・会社名かな・地域で検索",
        value=default_keyword,
        placeholder="入力すると候補を表示します",
        key=f"{partner_type}_partner_search_input",
        autocomplete="off",
    ).strip()
    update_query_params(
        page=trade_partner_search_page(partner_type),
        partner_search=keyword or None,
    )
    if not keyword:
        st.info("検索文字を入力してください。")
        return
    target = normalize_match_value(keyword).casefold()
    matches = [
        row for row in load_trade_partner_search_index(partner_type)
        if target in str(row.get("_検索文字") or "")
    ]
    matches.sort(key=trade_partner_sort_key)
    if not matches:
        st.warning("一致する会社が見つかりません。")
        return
    st.caption(f"{len(matches)}件")
    render_trade_partner_directory_cards(matches, partner_type)


def show_supplier_product_search():
    show_trade_partner_home_link("supplier")
    st.header("📦 仕入商品の検索")
    data = load_trade_partner_data()
    supplier_ids = {
        trade_partner_text(row.get("取引先ID"))
        for row in get_trade_partner_master_rows(data, "supplier")
    }
    products = [
        row for row in data[TRADE_PARTNER_PRODUCT_SHEET]["rows"]
        if trade_partner_text(row.get("取引先ID")) in supplier_ids
        and trade_partner_text(row.get("商品名"))
    ]
    keyword = st.text_input(
        "商品名で検索",
        placeholder="例：酒、醤油粕",
        key="supplier_product_search_input",
        autocomplete="off",
    ).strip()
    if not keyword:
        st.info("商品名を入力してください。")
        return
    candidates = sorted({
        trade_partner_text(row.get("商品名"))
        for row in products
        if normalize_match_value(keyword).lower()
        in normalize_match_value(trade_partner_text(row.get("商品名"))).lower()
    })
    if not candidates:
        st.warning("一致する商品名が見つかりません。")
        return
    st.caption("候補の商品名を選んでください。")
    selected = st.session_state.get("selected_supplier_product", "")
    columns = st.columns(min(3, max(1, len(candidates))))
    for index, product_name in enumerate(candidates):
        with columns[index % len(columns)]:
            if st.button(product_name, key=f"supplier_product_candidate_{index}", use_container_width=True):
                st.session_state["selected_supplier_product"] = product_name
                selected = product_name
                st.rerun()
    if selected not in candidates:
        return
    st.markdown(f"### {selected}")
    master_by_id = {
        trade_partner_text(row.get("取引先ID")): row
        for row in get_trade_partner_master_rows(data, "supplier")
    }
    exact_rows = [row for row in products if trade_partner_text(row.get("商品名")) == selected]
    exact_rows.sort(key=lambda row: trade_partner_sort_key(master_by_id.get(trade_partner_text(row.get("取引先ID")), {})))
    for product in exact_rows:
        master = master_by_id.get(trade_partner_text(product.get("取引先ID")))
        if not master:
            continue
        with st.container(border=True):
            st.markdown(trade_partner_detail_link(master, "supplier"), unsafe_allow_html=True)
            for field in data[TRADE_PARTNER_PRODUCT_SHEET]["headers"]:
                if field in {"仕入商品ID", "取引先ID", "会社名（確認用）", "商品名"}:
                    continue
                value = trade_partner_text(product.get(field))
                if value:
                    if field == "備考":
                        st.caption("備考")
                        render_collapsible_record_remarks(value)
                    else:
                        st.write(f"**{field}：** {value}")


def show_carrier_condition_search():
    show_trade_partner_home_link("carrier")
    st.header("🗺 運送条件検索")
    keyword = st.text_input(
        "納品先・地域・運賃などで検索",
        placeholder="例：帯広、釧路",
        key="carrier_condition_search_input",
        autocomplete="off",
    ).strip()
    if not keyword:
        st.info("検索文字を入力してください。")
        return
    data = load_trade_partner_data()
    carrier_ids = {
        trade_partner_text(row.get("取引先ID"))
        for row in get_trade_partner_master_rows(data, "carrier")
    }
    target = normalize_match_value(keyword).lower()
    matches = []
    for row in data[TRADE_PARTNER_TRANSPORT_SHEET]["rows"]:
        partner_id = trade_partner_text(row.get("取引先ID"))
        if partner_id not in carrier_ids:
            continue
        values = [
            trade_partner_text(row.get(header))
            for header in data[TRADE_PARTNER_TRANSPORT_SHEET]["headers"]
            if header not in {"運送条件ID", "取引先ID", "会社名（確認用）"}
        ]
        if target in normalize_match_value(" ".join(values)).lower():
            matches.append(row)
    if not matches:
        st.warning("一致する運送条件が見つかりません。")
        return
    master_by_id = {
        trade_partner_text(row.get("取引先ID")): row
        for row in get_trade_partner_master_rows(data, "carrier")
    }
    for condition in matches:
        master = master_by_id.get(trade_partner_text(condition.get("取引先ID")))
        if not master:
            continue
        with st.container(border=True):
            st.markdown(trade_partner_detail_link(master, "carrier"), unsafe_allow_html=True)
            for field in data[TRADE_PARTNER_TRANSPORT_SHEET]["headers"]:
                if field in {"運送条件ID", "取引先ID", "会社名（確認用）"}:
                    continue
                value = trade_partner_text(condition.get(field))
                if value:
                    if field == "備考":
                        st.caption("備考")
                        render_collapsible_record_remarks(value)
                    else:
                        st.write(f"**{field}：** {value}")


def render_trade_partner_fields(row, headers, excluded=None):
    excluded = set(excluded or [])
    visible = []
    for header in headers:
        if header in excluded or header.startswith("_"):
            continue
        value = trade_partner_text(row.get(header))
        if value:
            visible.append((header, value))
    if not visible:
        st.caption("入力済みの情報はありません。")
        return

    pending = []

    def render_pending_fields():
        nonlocal pending
        if not pending:
            return
        cols = st.columns(2)
        for offset, item in enumerate(pending):
            header, value = item
            with cols[offset]:
                st.caption(header)
                st.markdown(f"**{html.escape(value)}**")
        pending = []

    for header, value in visible:
        if header == "備考":
            render_pending_fields()
            st.caption("備考")
            render_collapsible_record_remarks(value)
            continue
        pending.append((header, value))
        if len(pending) == 2:
            render_pending_fields()
    render_pending_fields()


def render_trade_partner_form_input(header, value="", key=None):
    """備考だけ改行可能な入力欄にし、それ以外の入力方法は変えない。"""
    if header == "備考":
        return st.text_area(
            header,
            value=trade_partner_text(value),
            height=110,
            key=key,
        )
    return st.text_input(
        header,
        value=trade_partner_text(value),
        key=key,
        autocomplete="off",
    )


def trade_partner_history_section(sheet_name):
    return {
        TRADE_PARTNER_MASTER_SHEET: "基本情報",
        TRADE_PARTNER_CONTACT_SHEET: "担当者",
        TRADE_PARTNER_PRODUCT_SHEET: "仕入商品",
        TRADE_PARTNER_TRANSPORT_SHEET: "運送条件",
    }.get(sheet_name, sheet_name)


def render_trade_partner_row_editor(
    sheet_name,
    row,
    headers,
    key_prefix,
    partner_type,
    company_name,
):
    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]
    record_id = trade_partner_text(row.get(id_field))
    excluded_headers = {id_field, "取引先ID", "会社名（確認用）"}
    if sheet_name == TRADE_PARTNER_MASTER_SHEET:
        excluded_headers.update({"仕入先区分", "運送会社区分"})
    if sheet_name == TRADE_PARTNER_CONTACT_SHEET:
        excluded_headers.add("会社名")
    editable_headers = [header for header in headers if header not in excluded_headers]
    with st.expander("編集"):
        with st.form(f"edit_{key_prefix}_{record_id}"):
            inputs = {}
            for header in editable_headers:
                inputs[header] = render_trade_partner_form_input(
                    header,
                    value=row.get(header),
                    key=f"edit_{key_prefix}_{record_id}_{header}",
                )
            submitted = st.form_submit_button("バックアップして保存", use_container_width=True)
        if submitted:
            try:
                with st.spinner("バックアップを作成して保存しています…"):
                    result = update_trade_partner_row(sheet_name, record_id, inputs)
                    new_company_name = (
                        trade_partner_text(inputs.get("会社名"))
                        if sheet_name == TRADE_PARTNER_MASTER_SHEET
                        else company_name
                    ) or company_name
                    remember_change_history_warning(
                        record_change_history_safely(
                            trade_partner_type_label(partner_type),
                            result.get("partner_id") or trade_partner_text(row.get("取引先ID")),
                            new_company_name,
                            "変更",
                            result.get("changes", {}),
                            section=trade_partner_history_section(sheet_name),
                        )
                    )
                st.success("保存しました。")
                st.rerun()
            except Exception as error:
                st.error(str(error))


def render_trade_partner_related_section(
    data,
    sheet_name,
    partner_id,
    title,
    add_label,
    partner_type,
    company_name,
):
    headers = data[sheet_name]["headers"]
    rows = get_trade_partner_related_rows(data, sheet_name, partner_id)
    id_field = TRADE_PARTNER_ID_FIELDS[sheet_name]
    primary_field = TRADE_PARTNER_PRIMARY_FIELDS[sheet_name]
    st.markdown("---")
    st.subheader(title)
    if not rows:
        st.info(f"{title}はまだ登録されていません。")
    for row in rows:
        with st.container(border=True):
            heading = trade_partner_text(row.get(primary_field))
            st.markdown(f"**{html.escape(heading)}**")
            display_excluded = {id_field, "取引先ID", "会社名（確認用）", primary_field}
            if sheet_name == TRADE_PARTNER_CONTACT_SHEET:
                display_excluded.add("会社名")
            render_trade_partner_fields(
                row,
                headers,
                excluded=display_excluded,
            )
            render_trade_partner_row_editor(
                sheet_name,
                row,
                headers,
                key_prefix=f"{sheet_name}_{partner_id}",
                partner_type=partner_type,
                company_name=company_name,
            )

    with st.expander(f"＋ {add_label}"):
        add_excluded_headers = {id_field, "取引先ID", "会社名（確認用）"}
        if sheet_name in {TRADE_PARTNER_CONTACT_SHEET, TRADE_PARTNER_PRODUCT_SHEET}:
            add_excluded_headers.add("会社名")
        editable_headers = [
            header for header in headers
            if header not in add_excluded_headers
        ]
        with st.form(f"add_{sheet_name}_{partner_id}"):
            values = {"取引先ID": partner_id}
            if sheet_name in {TRADE_PARTNER_CONTACT_SHEET, TRADE_PARTNER_PRODUCT_SHEET}:
                values["会社名"] = company_name
            for header in editable_headers:
                values[header] = render_trade_partner_form_input(
                    header,
                    key=f"add_{sheet_name}_{partner_id}_{header}",
                )
            submitted = st.form_submit_button("バックアップして追加", use_container_width=True)
        if submitted:
            try:
                if not trade_partner_text(values.get(primary_field)):
                    raise ValueError(f"{primary_field}を入力してください。")
                with st.spinner("バックアップを作成して保存しています…"):
                    result = create_trade_partner_record(sheet_name, values)
                    remember_change_history_warning(
                        record_change_history_safely(
                            trade_partner_type_label(partner_type),
                            result.get("partner_id") or partner_id,
                            company_name,
                            "追加",
                            result.get("changes", {}),
                            section=trade_partner_history_section(sheet_name),
                        )
                    )
                st.success("追加しました。")
                st.rerun()
            except Exception as error:
                st.error(str(error))


def show_trade_partner_notes(partner_type, partner_id, company_name):
    st.markdown("---")
    st.subheader(f"📝 この{trade_partner_type_label(partner_type)}のメモ")
    show_note_delete_success_message()
    note_key = make_trade_partner_note_key(partner_type, partner_id, company_name)
    input_key = f"trade_partner_note_{partner_type}_{partner_id}"
    clear_key = f"clear_{input_key}"
    if st.session_state.pop(clear_key, False):
        st.session_state[input_key] = ""
    note_text = st.text_area(
        "メモ本文",
        key=input_key,
        height=120,
        help=VOICE_INPUT_HELP,
    )
    if st.button("メモを保存", key=f"save_{input_key}"):
        if add_note(note_key, note_text):
            st.session_state[clear_key] = True
            st.rerun()
    notes = get_notes_for_customer(note_key)
    if not notes:
        st.info("メモはまだありません。")
        return
    st.markdown("#### メモ履歴")
    for note in notes:
        render_note_card(note, show_customer=False)
        render_note_delete_controls(note)


def show_trade_partner_detail(partner_type, partner_id):
    show_trade_partner_home_link(partner_type)
    data = get_trade_partner_detail_data(partner_id)
    master = get_trade_partner_by_id(data, partner_id)
    if not master or not is_trade_partner_marked(master.get(trade_partner_category_field(partner_type))):
        st.warning("選択した会社の情報が見つかりません。")
        return
    label = trade_partner_type_label(partner_type)
    company = trade_partner_text(master.get("会社名"))
    st.title(f"{'🏢' if partner_type == 'supplier' else '🚚'} {company}")
    st.caption(f"{label}ID：{trade_partner_text(master.get('取引先ID'))}")

    map_value = trade_partner_text(master.get("マップ位置")) or trade_partner_text(master.get("住所"))
    if map_value:
        map_url = build_google_maps_url(map_value)
        if map_url:
            show_google_maps_button(map_url)

    st.subheader("基本情報")
    master_headers = data[TRADE_PARTNER_MASTER_SHEET]["headers"]
    render_trade_partner_fields(
        master,
        master_headers,
        excluded={"取引先ID", "仕入先区分", "運送会社区分", "会社名", "会社名かな"},
    )
    render_trade_partner_row_editor(
        TRADE_PARTNER_MASTER_SHEET,
        master,
        master_headers,
        key_prefix=f"master_{partner_type}",
        partner_type=partner_type,
        company_name=company,
    )

    render_trade_partner_related_section(
        data,
        TRADE_PARTNER_CONTACT_SHEET,
        partner_id,
        "担当者",
        "担当者を追加",
        partner_type,
        company,
    )
    if partner_type == "supplier":
        render_trade_partner_related_section(
            data,
            TRADE_PARTNER_PRODUCT_SHEET,
            partner_id,
            "取扱商品",
            "商品を追加",
            partner_type,
            company,
        )
    else:
        render_carrier_freight_section(partner_id, company)
    render_customer_attachments_section(
        company,
        partner_id,
        entity_type=partner_type,
    )
    show_trade_partner_notes(partner_type, partner_id, company)


def show_trade_partner_register(partner_type):
    show_trade_partner_home_link(partner_type)
    label = trade_partner_type_label(partner_type)
    st.header(f"＋ 新しい{label}を登録")
    data = load_trade_partner_data()
    headers = data[TRADE_PARTNER_MASTER_SHEET]["headers"]
    editable_headers = [
        header for header in headers
        if header not in {"取引先ID", "仕入先区分", "運送会社区分"}
    ]
    other_label = "運送会社でもある" if partner_type == "supplier" else "仕入先でもある"
    with st.form(f"register_{partner_type}"):
        values = {}
        for header in editable_headers:
            values[header] = render_trade_partner_form_input(
                header,
                key=f"register_{partner_type}_{header}",
            )
        also_other = st.checkbox(other_label, key=f"register_{partner_type}_also_other")
        submitted = st.form_submit_button("バックアップして登録", use_container_width=True)
    if submitted:
        try:
            if not trade_partner_text(values.get("会社名")):
                raise ValueError("会社名を入力してください。")
            values[trade_partner_category_field(partner_type)] = "○"
            if also_other:
                values[trade_partner_category_field("carrier" if partner_type == "supplier" else "supplier")] = "○"
            with st.spinner("バックアップを作成して登録しています…"):
                result = create_trade_partner_record(TRADE_PARTNER_MASTER_SHEET, values)
                remember_change_history_warning(
                    record_change_history_safely(
                        trade_partner_type_label(partner_type),
                        result.get("partner_id") or result.get("record_id"),
                        trade_partner_text(values.get("会社名")),
                        "登録",
                        result.get("changes", {}),
                        section="基本情報",
                    )
                )
            partner_id = result["record_id"]
            st.session_state["selected_partner_id"] = partner_id
            st.session_state["selected_partner_type"] = partner_type
            st.session_state["page"] = "partner_detail"
            update_query_params(
                page="partner_detail",
                partner_id=partner_id,
                partner_type=partner_type,
            )
            st.rerun()
        except Exception as error:
            st.error(str(error))


def render_trade_note_card(note, category, partner_names=None):
    parsed = parse_trade_partner_note_key(note.get("customer_name"))
    created_at = format_note_datetime(note.get("created_at", ""))
    body = html.escape(clean_value(note.get("body"), blank_text="")).replace("\n", "<br>")
    if parsed:
        partner_names = partner_names or {}
        company_name = (
            partner_names.get((parsed["partner_type"], parsed["partner_id"]))
            or parsed.get("company_name")
            or parsed["partner_id"]
        )
        company_link = render_page_link(
            company_name,
            page="partner_detail",
            partner_id=parsed["partner_id"],
            partner_type=parsed["partner_type"],
            class_name="dispatch-month-link",
        )
        meta = f"{html.escape(created_at)}　{company_link}"
    else:
        customer_name = clean_value(note.get("customer_name"), blank_text="未設定")
        customer_link = build_customer_detail_link(customer_name, class_name="dispatch-month-link")
        meta = f"{html.escape(created_at)}　{customer_link}"
    st.markdown(
        f'<div class="note-card"><div class="note-meta">{meta}</div><div class="note-body">{body}</div></div>',
        unsafe_allow_html=True,
    )


def show_trade_notes_page():
    show_top_home_link()
    st.header("📝 メモ帳")
    show_note_delete_success_message()
    notes = load_notes_from_supabase()
    try:
        partner_data = load_trade_partner_data()
        partner_names = {}
        for partner_type in ("supplier", "carrier"):
            for row in get_trade_partner_master_rows(partner_data, partner_type):
                partner_names[(partner_type, trade_partner_text(row.get("取引先ID")))] = trade_partner_text(row.get("会社名"))
    except Exception:
        partner_names = {}

    # st.tabs は削除確認などで再実行されるたびに先頭の「顧客」へ戻るため、
    # 選択値を session_state に保持できる横並びラジオで区分を切り替える。
    # これにより、仕入先や運送会社のメモを続けて削除しても同じ区分を維持する。
    category_labels = ["すべて", "顧客", "仕入先", "運送会社"]
    category_by_label = {
        "すべて": "all",
        "顧客": "customer",
        "仕入先": "supplier",
        "運送会社": "carrier",
    }
    selected_label = st.radio(
        "表示する区分",
        category_labels,
        horizontal=True,
        key="trade_notes_selected_category",
        label_visibility="collapsed",
    )
    category = category_by_label[selected_label]

    filtered = []
    for note in notes:
        parsed = parse_trade_partner_note_key(note.get("customer_name"))
        if category == "all":
            filtered.append(note)
        elif category == "customer" and parsed is None:
            filtered.append(note)
        elif parsed and parsed["partner_type"] == category:
            filtered.append(note)

    if category in {"all", "customer"}:
        try:
            past_product_items = load_all_past_product_notes_from_supabase()
        except Exception as exc:
            st.warning(f"過去商品メモを読み込めませんでした：{exc}")
            past_product_items = []

        for item in past_product_items:
            product_name = extract_past_product_name(item.get("field_name"))
            customer_name = clean_value(item.get("customer_name"), blank_text="")
            content = clean_value(item.get("content"), blank_text="").strip()
            if not product_name or not customer_name or not content:
                continue
            filtered.append(
                {
                    "id": item.get("id"),
                    "customer_name": customer_name,
                    "body": f"過去に使用した商品：{product_name}\n{content}",
                    "created_at": item.get("updated_at") or item.get("created_at") or "",
                    "_past_product_note": True,
                }
            )

    filtered.sort(
        key=lambda note: str(note.get("created_at") or ""),
        reverse=True,
    )

    if not filtered:
        st.info("メモはまだありません。")
        return

    for note in filtered:
        render_trade_note_card(note, category, partner_names=partner_names)
        if not note.get("_past_product_note"):
            render_note_delete_controls(note)



@st.cache_data(ttl=15, show_spinner=False)
def load_login_browsers_from_supabase():
    if not has_login_audit_config():
        return []
    response = requests.get(
        get_login_browsers_url(),
        headers=get_login_audit_headers(),
        params={
            "select": (
                "id,browser_token_hash,first_seen_at,last_seen_at,first_ip_address,"
                "last_ip_address,browser_name,os_name,device_type,is_active,"
                "is_acknowledged,acknowledged_at"
            ),
            "order": "last_seen_at.desc",
            "limit": "100",
        },
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"ログインブラウザ一覧を取得できませんでした（{response.status_code}）。")
    return response.json() or []


@st.cache_data(ttl=15, show_spinner=False)
def load_login_events_from_supabase(limit=LOGIN_HISTORY_PAGE_SIZE):
    if not has_login_audit_config():
        return []
    safe_limit = max(1, min(int(limit), 300))
    response = requests.get(
        get_login_events_url(),
        headers=get_login_audit_headers(),
        params={
            "select": (
                "id,event_type,occurred_at,account_label,is_new_browser,severity,"
                "ip_address,browser_name,os_name,device_type,details"
            ),
            "order": "occurred_at.desc",
            "limit": str(safe_limit),
        },
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"ログイン履歴を取得できませんでした（{response.status_code}）。")
    return response.json() or []


def get_unacknowledged_login_browser_count():
    try:
        return sum(
            1
            for item in load_login_browsers_from_supabase()
            if bool(item.get("is_active", True)) and not bool(item.get("is_acknowledged"))
        )
    except Exception:
        return 0


def acknowledge_login_browser(browser_id):
    response = requests.patch(
        get_login_browsers_url(),
        headers=get_login_audit_headers(prefer="return=minimal"),
        params={"id": f"eq.{browser_id}"},
        json={
            "is_acknowledged": True,
            "acknowledged_at": utc_now_iso(),
            "updated_at": utc_now_iso(),
        },
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code not in (200, 204):
        raise RuntimeError(f"確認済みに変更できませんでした（{response.status_code}）。")
    clear_login_audit_caches()


def acknowledge_all_login_browsers():
    response = requests.patch(
        get_login_browsers_url(),
        headers=get_login_audit_headers(prefer="return=minimal"),
        params={"is_acknowledged": "eq.false"},
        json={
            "is_acknowledged": True,
            "acknowledged_at": utc_now_iso(),
            "updated_at": utc_now_iso(),
        },
        timeout=LOGIN_AUDIT_REQUEST_TIMEOUT,
    )
    if response.status_code not in (200, 204):
        raise RuntimeError(f"一括確認済みに変更できませんでした（{response.status_code}）。")
    clear_login_audit_caches()


def format_login_event_name(event_type):
    labels = {
        "login_success": "ログイン成功",
        "account_denied": "許可外アカウントを拒否",
        "logout": "ログアウト",
    }
    return labels.get(str(event_type or ""), str(event_type or "不明な操作"))


def format_login_client_summary(item):
    browser = str(item.get("browser_name") or "不明なブラウザ")
    os_name = str(item.get("os_name") or "不明なOS")
    device = str(item.get("device_type") or "不明な端末")
    return f"{device}・{os_name}・{browser}"


def show_login_history_page():
    show_top_home_link()
    st.header("🔐 ログイン履歴")
    st.caption("Microsoft認証後に取引先カルテへ到達したログインを記録します。IPアドレスは参考表示で、本人判定には使いません。")

    if not has_login_audit_config():
        st.error("ログイン履歴にはSupabaseの管理用Secret KeyまたはService Role Keyが必要です。")
        return

    try:
        browsers = load_login_browsers_from_supabase()
        events = load_login_events_from_supabase()
    except Exception as exc:
        st.error(f"ログイン履歴を読み込めませんでした：{exc}")
        st.info("先に app_login_browsers と app_login_events のSQLをSupabaseで実行してください。")
        return

    unacknowledged = [
        item
        for item in browsers
        if bool(item.get("is_active", True)) and not bool(item.get("is_acknowledged"))
    ]
    if unacknowledged:
        st.warning(f"未確認の新しいブラウザが {len(unacknowledged)} 件あります。")
        if st.button("すべて確認済みにする", type="primary", use_container_width=True):
            try:
                acknowledge_all_login_browsers()
                st.success("すべて確認済みにしました。")
                st.rerun()
            except Exception as exc:
                st.error(f"更新できませんでした：{exc}")
    else:
        st.success("未確認の新しいブラウザはありません。")

    current_browser_hash = get_current_browser_token_hash(create_if_missing=False)

    st.markdown("### ブラウザ一覧")
    if not browsers:
        st.info("ブラウザ履歴はまだありません。")
    for item in browsers:
        browser_id = item.get("id")
        is_current = bool(
            current_browser_hash
            and hmac.compare_digest(
                str(item.get("browser_token_hash") or ""),
                current_browser_hash,
            )
        )
        is_acknowledged = bool(item.get("is_acknowledged"))
        title = format_login_client_summary(item)
        if is_current:
            title += "（現在のブラウザ）"
        with st.container(border=True):
            st.markdown(f"**{html.escape(title)}**")
            st.write(f"初回：{format_note_datetime(item.get('first_seen_at'))}")
            st.write(f"最終アクセス：{format_note_datetime(item.get('last_seen_at'))}")
            st.write(f"最終IP：{html.escape(str(item.get('last_ip_address') or '取得できません'))}")
            if is_acknowledged:
                st.caption("確認済み")
            elif st.button(
                "このブラウザを確認済みにする",
                key=f"ack_login_browser_{browser_id}",
                use_container_width=True,
            ):
                try:
                    acknowledge_login_browser(browser_id)
                    st.success("確認済みにしました。")
                    st.rerun()
                except Exception as exc:
                    st.error(f"更新できませんでした：{exc}")

    st.markdown("### 最新のログイン履歴")
    if not events:
        st.info("ログイン履歴はまだありません。")
        return

    for event in events:
        event_name = format_login_event_name(event.get("event_type"))
        occurred_at = format_note_datetime(event.get("occurred_at"))
        client_summary = format_login_client_summary(event)
        new_label = "・新しいブラウザ" if event.get("is_new_browser") else ""
        with st.container(border=True):
            if event.get("event_type") == "account_denied":
                st.error(f"{event_name}{new_label}")
            elif event.get("is_new_browser"):
                st.warning(f"{event_name}{new_label}")
            else:
                st.markdown(f"**{event_name}**")
            st.write(f"日時：{occurred_at}")
            st.write(f"端末：{html.escape(client_summary)}")
            st.write(f"IP：{html.escape(str(event.get('ip_address') or '取得できません'))}")
            account_label = str(event.get("account_label") or "").strip()
            if account_label:
                st.caption(f"Microsoftアカウント表示名：{account_label}")


def show_top_home():
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(render_page_link("👥 顧客", page="customer_home"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_page_link("🏢 仕入先", page="supplier_home"), unsafe_allow_html=True)
    col3, col4 = st.columns(2)
    with col3:
        st.markdown(render_page_link("🚚 運送会社", page="carrier_home"), unsafe_allow_html=True)
    with col4:
        st.markdown(render_page_link("📝 メモ帳", page="trade_notes"), unsafe_allow_html=True)
    col5, col6 = st.columns(2)
    with col5:
        st.markdown(render_page_link("🔎 写真・PDF検索", page="attachment_search"), unsafe_allow_html=True)
    with col6:
        st.markdown(render_page_link("🕘 変更確認", page="change_history"), unsafe_allow_html=True)
    col7, col8 = st.columns(2)
    with col7:
        st.markdown(render_page_link("📄 商品見積り履歴", page="estimates"), unsafe_allow_html=True)
    with col8:
        alert_count = get_unacknowledged_login_browser_count()
        label = "🔐 ログイン履歴"
        if alert_count:
            label += f"（未確認 {alert_count}）"
        st.markdown(render_page_link(label, page="login_history"), unsafe_allow_html=True)

    st.markdown(
        render_page_link(
            "🏨 ホテル・宿泊先情報",
            page="hotel_information",
            class_name="app-nav-link hotel-home-card-link",
        ),
        unsafe_allow_html=True,
    )

    render_home_todo_section()


# =========================
# ホームメニュー
# =========================
def show_home_menu():
    show_top_home_link()
    st.subheader("顧客メニュー")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(render_page_link("👥 顧客名一覧", page="customer_list"), unsafe_allow_html=True)
    with col2:
        st.markdown(render_page_link("🔍 顧客検索", page="customer"), unsafe_allow_html=True)

    col3, col4 = st.columns(2)
    with col3:
        st.markdown(render_page_link("📍 地域検索", page="region"), unsafe_allow_html=True)
    with col4:
        st.markdown(render_page_link("🔎 商品検索", page="product"), unsafe_allow_html=True)

    col5, col6 = st.columns(2)
    with col5:
        st.markdown(render_page_link("🗓 在庫カレンダー", page="calendar"), unsafe_allow_html=True)
    with col6:
        st.markdown(render_page_link("🚚 配車表", page="dispatch_table"), unsafe_allow_html=True)

    col7, col8 = st.columns(2)
    with col7:
        st.markdown(render_page_link("🧪 ソリュブル在庫", page="soluble_inventory"), unsafe_allow_html=True)
    with col8:
        st.markdown(render_page_link("📝 メモ帳", page="notes"), unsafe_allow_html=True)

    col9, _ = st.columns(2)
    with col9:
        st.markdown(render_page_link("💧 WATER it接続", page="water_it_test"), unsafe_allow_html=True)

    st.markdown("---")

# =========================
# 全データバックアップ（読み取り専用）
# =========================
def backup_safe_value(value):
    """CSVへ安全に書き出せる値へ変換する。"""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def backup_csv_bytes(dataframe):
    """Excelで文字化けしにくいUTF-8 BOM付きCSVを返す。"""
    export_df = dataframe.copy()
    for column in export_df.columns:
        export_df[column] = export_df[column].map(backup_safe_value)
    return b"\xef\xbb\xbf" + export_df.to_csv(
        index=False,
        lineterminator="\n",
    ).encode("utf-8")


def backup_dataframe(records, columns=None):
    """空の一覧でも必要な見出しを残す。"""
    dataframe = pd.DataFrame(records or [])
    if columns:
        for column in columns:
            if column not in dataframe.columns:
                dataframe[column] = ""
        ordered = list(columns) + [
            column for column in dataframe.columns if column not in columns
        ]
        dataframe = dataframe[ordered]
    return dataframe


def backup_read_all_supabase_rows(url, label):
    """Supabaseのテーブルを読み取り専用で全件取得する。"""
    if not has_supabase_config():
        raise RuntimeError(f"{label}を取得するためのSupabase設定がありません。")

    rows = []
    page_size = 1000
    offset = 0
    while True:
        try:
            response = requests.get(
                url,
                headers=get_supabase_headers(),
                params={
                    "select": "*",
                    "limit": str(page_size),
                    "offset": str(offset),
                },
                timeout=30,
            )
        except Exception as exc:
            raise RuntimeError(f"{label}の取得中にSupabaseへ接続できませんでした。") from exc

        if response.status_code != 200:
            detail = str(response.text or "").strip()[:500]
            raise RuntimeError(
                f"{label}を取得できませんでした（{response.status_code}）。"
                + (f" {detail}" if detail else "")
            )

        try:
            page = response.json()
        except Exception as exc:
            raise RuntimeError(f"{label}の応答形式が正しくありません。") from exc
        if not isinstance(page, list):
            raise RuntimeError(f"{label}の応答形式が正しくありません。")

        rows.extend(page)
        if len(page) < page_size:
            break
        offset += page_size
    return rows


def backup_get_main_excel_bytes():
    """顧客・在庫の元Excelを読み取り専用で取得する。"""
    if has_dropbox_auth_config():
        content = get_cached_dropbox_excel_content()
        path = get_dropbox_file_path()
        return content, Path(path).name or "配車予定 次郎.xlsm", f"Dropbox: {path}"

    path = Path(EXCEL_FILE)
    if not path.exists():
        raise FileNotFoundError(f"顧客・在庫の元Excelが見つかりません：{path}")
    return path.read_bytes(), path.name, f"ローカル: {path}"


def backup_get_dispatch_excel_bytes():
    """配車表の元Excelを画面と同じ優先順位で取得する。"""
    dropbox_error = None
    if has_dropbox_auth_config():
        try:
            content = get_cached_dispatch_dropbox_content()
            path = str(DISPATCH_DROPBOX_FILE_PATH or DISPATCH_DROPBOX_DEFAULT_FILE_PATH).strip()
            return content, Path(path).name or "配車表1.xlsm", f"Dropbox: {path}"
        except Exception as exc:
            dropbox_error = exc

    path = Path(str(DISPATCH_LOCAL_FILE or "").strip())
    if path.exists():
        return path.read_bytes(), path.name, f"ローカル: {path}"

    message = f"配車表の元Excelが見つかりません：{path}"
    if dropbox_error is not None:
        message += f"\nDropbox取得エラー：{dropbox_error}"
    raise FileNotFoundError(message)


def backup_get_soluble_excel_bytes():
    """ソリュブル在庫の元Excelを読み取り専用で取得する。"""
    content, source = load_soluble_workbook_content()
    return content, SOLUBLE_FILE_NAME, source


def backup_get_trade_partner_excel_bytes():
    """仕入先・運送会社の元Excelを読み取り専用で取得する。"""
    access_token = get_dropbox_access_token()
    path = get_trade_partner_file_path()
    content, response = download_dropbox_file(path, access_token)
    if content is None:
        raise RuntimeError(
            "取引先カルテ.xlsxをDropboxから取得できませんでした。\n"
            + dropbox_error_text(response)
        )
    return content, TRADE_PARTNER_FILE_NAME, f"Dropbox: {path}"


def backup_build_product_usage(customer_df):
    """商品検索と同じ基準で現在使用中・過去使用を一覧化する。"""
    product_rows = get_product_search_rows(customer_df)
    records = []
    if product_rows.empty:
        return backup_dataframe(
            [],
            ["商品名", "利用区分", "顧客名", "地域", "使用数量/日", "使用中行数"],
        )

    for product_name, product_group in product_rows.groupby("_商品名検索", sort=True):
        for customer_name, group in product_group.groupby("_顧客名検索", sort=True):
            current_group = group[
                ~group["使用数量/日"].apply(is_blank_or_zero)
            ].copy()
            regions = [
                clean_value(value, blank_text="").strip()
                for value in group["地域"].tolist()
            ]
            region = next((value for value in regions if value), "")
            if current_group.empty:
                status = "過去に使用"
                usage = ""
                current_count = 0
            else:
                status = "現在使用中"
                current_count = len(current_group)
                usage_values = [
                    clean_value(value, blank_text="").strip()
                    for value in current_group["使用数量/日"].tolist()
                ]
                usage = " / ".join(dict.fromkeys(value for value in usage_values if value))
            records.append(
                {
                    "商品名": product_name,
                    "利用区分": status,
                    "顧客名": customer_name,
                    "地域": region,
                    "使用数量/日": usage,
                    "使用中行数": current_count,
                }
            )
    return pd.DataFrame(records)


def backup_build_calendar_export(customer_df):
    """在庫カレンダーで使う基本項目を元データから書き出す。"""
    candidates = [
        "ID", "顧客名", "地域", "商品名", "メーカー",
        "使用数量/日", "次回配達予定", "残数",
    ]
    columns = [column for column in candidates if column in customer_df.columns]
    if not columns:
        return pd.DataFrame()
    result = customer_df[columns].copy()
    if "次回配達予定" in result.columns:
        result["次回配達予定_解析"] = pd.to_datetime(
            result["次回配達予定"], errors="coerce"
        )
        result.sort_values(
            ["次回配達予定_解析", "顧客名"],
            na_position="last",
            inplace=True,
        )
    return result.reset_index(drop=True)


def backup_build_note_exports(raw_notes):
    """notesテーブルを通常メモとLINE状態へ分ける。"""
    normal_notes = []
    line_statuses = []
    for row in raw_notes:
        row_id = clean_value(row.get("id"), blank_text="")
        if row_id.startswith(LINE_STATUS_NOTE_PREFIX):
            line_statuses.append(
                {
                    "顧客名": row.get("customer_name", ""),
                    "LINE状態": "接続中",
                    "保存ID": row_id,
                    "作成日時": row.get("created_at", ""),
                }
            )
        else:
            normal_notes.append(row)
    return backup_dataframe(normal_notes), backup_dataframe(
        line_statuses,
        ["顧客名", "LINE状態", "保存ID", "作成日時"],
    )


def backup_add_entry(entries, path, content, description, count=""):
    if not isinstance(content, (bytes, bytearray)) or not content:
        raise RuntimeError(f"バックアップ内容が空です：{path}")
    entries.append(
        {
            "path": path,
            "content": bytes(content),
            "description": description,
            "count": count,
        }
    )


def ensure_full_data_backup_dropbox_folder(access_token):
    """全データバックアップ専用フォルダを作る。既存フォルダは成功扱いにする。"""
    folder = str(FULL_DATA_BACKUP_DROPBOX_FOLDER or "").strip().rstrip("/")
    if not folder:
        raise RuntimeError("全データバックアップのDropbox保存先が設定されていません。")

    response = call_dropbox_rpc(
        "files/create_folder_v2",
        {"path": folder, "autorename": False},
        access_token,
    )
    if response.status_code == 200:
        return folder
    if response.status_code == 409:
        try:
            summary = str(response.json().get("error_summary", ""))
            if "conflict" in summary and "folder" in summary:
                return folder
        except Exception:
            pass
    raise RuntimeError(
        "Dropboxに全データバックアップ用フォルダを作成できませんでした。\n"
        + dropbox_error_text(response)
    )


def save_full_data_backup_to_dropbox(filename, content):
    """作成済みZIPを専用Dropboxフォルダへ追加保存し、内容を検証する。"""
    if not filename or not isinstance(content, (bytes, bytearray)) or not content:
        raise RuntimeError("Dropboxへ保存するバックアップZIPが空です。")

    access_token = get_dropbox_access_token()
    folder = ensure_full_data_backup_dropbox_folder(access_token)
    target_path = f"{folder}/{filename}"
    response = upload_dropbox_file(
        target_path,
        bytes(content),
        access_token,
        mode="add",
    )
    if response.status_code == 409:
        raise RuntimeError(
            "同じ名前のバックアップがDropboxに存在します。もう一度作成してください。"
        )
    if response.status_code != 200:
        raise RuntimeError(
            "Dropboxへ全データバックアップを保存できませんでした。\n"
            + dropbox_error_text(response)
        )

    metadata = get_dropbox_response_metadata(response)
    if not metadata.get("content_hash") or metadata.get("size") is None:
        metadata = get_dropbox_file_metadata(target_path, access_token)
    verify_dropbox_file_metadata(metadata, bytes(content))
    return target_path


def create_full_data_backup_zip():
    """現在の保存処理を変更せず、読み取りだけで全データZIPを作る。"""
    created_at = get_jst_now()
    timestamp = created_at.strftime("%Y%m%d_%H%M%S")
    entries = []
    sources = []

    main_excel, main_name, main_source = backup_get_main_excel_bytes()
    dispatch_excel, dispatch_name, dispatch_source = backup_get_dispatch_excel_bytes()
    soluble_excel, soluble_name, soluble_source = backup_get_soluble_excel_bytes()
    trade_excel, trade_name, trade_source = backup_get_trade_partner_excel_bytes()
    sources.extend([main_source, dispatch_source, soluble_source, trade_source])

    customer_df = normalize_excel_table(BytesIO(main_excel))
    dispatch_df = read_dispatch_month_sheets(BytesIO(dispatch_excel))
    soluble_rows_df = backup_dataframe(read_soluble_rows(soluble_excel))
    soluble_summary_df = backup_dataframe(
        list(read_soluble_customer_summaries(soluble_excel).values())
    )

    trade_editor = TradePartnerXlsxEditor(trade_excel)
    missing_sheets = [
        name for name in TRADE_PARTNER_REQUIRED_SHEETS
        if not trade_editor.has_sheet(name)
    ]
    if missing_sheets:
        raise RuntimeError(
            "取引先カルテ.xlsxに必要なシートがありません："
            + "、".join(missing_sheets)
        )
    trade_data = {
        name: trade_editor.read_sheet(name)
        for name in TRADE_PARTNER_REQUIRED_SHEETS
    }

    water_df, water_source = get_active_water_it_data()
    sources.append(f"WATER it表示データ: {water_source}")

    raw_notes = backup_read_all_supabase_rows(
        get_supabase_notes_url(),
        "メモ・LINE状態",
    )
    raw_customer_information = backup_read_all_supabase_rows(
        get_supabase_customer_information_url(),
        "顧客情報",
    )
    change_history_rows = [
        row for row in raw_customer_information
        if clean_value(row.get("customer_name"), blank_text="") == CHANGE_HISTORY_CUSTOMER
    ]
    water_storage_rows = [
        row for row in raw_customer_information
        if clean_value(row.get("id"), blank_text="") == WATER_IT_STORAGE_ID
        or clean_value(row.get("customer_name"), blank_text="") == WATER_IT_STORAGE_CUSTOMER
        or clean_value(row.get("field_name"), blank_text="") == WATER_IT_STORAGE_FIELD
    ]
    customer_information = [
        row for row in raw_customer_information
        if row not in water_storage_rows and row not in change_history_rows
    ]
    water_metadata = []
    for row in water_storage_rows:
        metadata = {
            "id": row.get("id", ""),
            "customer_name": row.get("customer_name", ""),
            "field_name": row.get("field_name", ""),
            "updated_at": row.get("updated_at", ""),
        }
        try:
            payload = json.loads(str(row.get("content") or "{}"))
            metadata.update(
                {
                    "保存ファイル名": payload.get("filename", ""),
                    "SHA256": payload.get("sha256", ""),
                    "行数": payload.get("row_count", ""),
                    "ポイント数": payload.get("point_count", ""),
                    "最古測定日時": payload.get("oldest_time", ""),
                    "最新測定日時": payload.get("latest_time", ""),
                    "取込日時": payload.get("imported_at", ""),
                    "保存形式バージョン": payload.get("version", ""),
                }
            )
        except Exception:
            metadata["解析結果"] = "保存メタデータを解析できませんでした"
        water_metadata.append(metadata)

    notes_df, line_df = backup_build_note_exports(raw_notes)
    product_usage_df = backup_build_product_usage(customer_df)
    calendar_df = backup_build_calendar_export(customer_df)
    estimates_df = estimate_rows_to_dataframe(customer_information)
    carrier_freights_df = carrier_freight_rows_to_dataframe(customer_information)
    onedrive_attachments_df = onedrive_attachment_rows_to_dataframe(customer_information)

    backup_add_entry(entries, f"元Excel/{main_name}", main_excel, "顧客・在庫の元Excel")
    backup_add_entry(entries, f"元Excel/{dispatch_name}", dispatch_excel, "配車表の元Excel")
    backup_add_entry(entries, f"元Excel/{soluble_name}", soluble_excel, "ソリュブル在庫の元Excel")
    backup_add_entry(entries, f"元Excel/{trade_name}", trade_excel, "仕入先・運送会社の元Excel")

    csv_exports = [
        ("CSV/01_顧客商品在庫_全行.csv", customer_df, "アプリが読み取る顧客・商品・在庫の全行"),
        ("CSV/02_商品利用状況.csv", product_usage_df, "現在使用中・過去使用の分類"),
        ("CSV/03_在庫カレンダー.csv", calendar_df, "在庫カレンダーの基本情報"),
        ("CSV/04_配車表.csv", dispatch_df.drop(columns=["_引取日", "_着日"], errors="ignore"), "配車表1月～12月"),
        ("CSV/05_ソリュブル在庫履歴.csv", soluble_rows_df, "ソリュブル在庫の日別情報"),
        ("CSV/06_ソリュブル顧客概要.csv", soluble_summary_df, "ソリュブル顧客の概要"),
        ("CSV/07_WATER_it表示データ.csv", water_df, "アプリで表示できるWATER it情報（元CSVは含まない）"),
        ("CSV/08_WATER_it保存メタデータ.csv", backup_dataframe(water_metadata), "WATER it保存データの概要のみ"),
        ("CSV/09_顧客情報_Supabase生データ.csv", backup_dataframe(customer_information), "顧客情報・過去商品メモ・提案見積り・運送会社運賃・写真資料メタデータの生データ"),
        ("CSV/10_メモ_Supabase生データ.csv", notes_df, "通常メモ・取引先メモの生データ"),
        ("CSV/11_LINE状態.csv", line_df, "LINE接続状態"),
    ]

    next_number = 12
    for sheet_name in TRADE_PARTNER_REQUIRED_SHEETS:
        sheet = trade_data[sheet_name]
        sheet_df = backup_dataframe(
            sheet["rows"],
            list(sheet["headers"]) + ["_row_number"],
        )
        csv_exports.append(
            (
                f"CSV/{next_number:02d}_{sheet_name}.csv",
                sheet_df,
                f"取引先カルテ.xlsxの{sheet_name}シート",
            )
        )
        next_number += 1

    csv_exports.extend(
        [
            (
                f"CSV/{next_number:02d}_変更履歴.csv",
                change_history_rows_to_dataframe(change_history_rows),
                "アプリから保存した変更履歴（メモ帳は対象外）",
            ),
            (
                f"CSV/{next_number + 1:02d}_変更履歴_Supabase生データ.csv",
                backup_dataframe(change_history_rows),
                "変更履歴のSupabase生データ",
            ),
            (
                f"CSV/{next_number + 2:02d}_提案見積り.csv",
                estimates_df,
                "顧客ごとの提案・見積り一覧",
            ),
            (
                f"CSV/{next_number + 3:02d}_運送会社運賃履歴.csv",
                carrier_freights_df,
                "運送会社ごとの運賃履歴",
            ),
            (
                f"CSV/{next_number + 4:02d}_写真資料メタデータ.csv",
                onedrive_attachments_df,
                "OneDriveに保存した顧客の写真・PDFの管理情報（ファイル本体はOneDrive）",
            ),
        ]
    )

    for path, dataframe, description in csv_exports:
        backup_add_entry(
            entries,
            path,
            backup_csv_bytes(dataframe),
            description,
            len(dataframe),
        )

    manifest = pd.DataFrame(
        [
            {
                "ファイル": entry["path"],
                "内容": entry["description"],
                "件数": entry["count"],
                "バイト数": len(entry["content"]),
                "SHA256": hashlib.sha256(entry["content"]).hexdigest(),
            }
            for entry in entries
        ]
    )
    backup_add_entry(
        entries,
        "バックアップ一覧.csv",
        backup_csv_bytes(manifest),
        "ZIP内ファイルの一覧・件数・SHA256",
        len(manifest),
    )

    info_lines = [
        f"作成日時: {created_at.strftime('%Y/%m/%d %H:%M:%S %z')}",
        f"アプリ名: {APP_TITLE}",
        "",
        "取得元:",
        *[f"- {source}" for source in sources],
        "",
        "注意:",
        "- 元のExcel・Supabaseデータは変更しません。作成したZIPだけをDropboxへ追加保存します。",
        "- パスワード、接続キー、secrets.tomlは含みません。",
        "- WATER itの元CSV本体および圧縮保存本文は含みません。",
        "- CSVはUTF-8 BOM付きです。",
        "- 元ExcelとCSVの重複は、データ保全のため意図的に残しています。",
    ]
    backup_add_entry(
        entries,
        "バックアップ情報.txt",
        ("\ufeff" + "\n".join(info_lines) + "\n").encode("utf-8"),
        "バックアップの作成日時・取得元・注意事項",
    )

    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for entry in entries:
            archive.writestr(entry["path"], entry["content"])
    zip_content = output.getvalue()

    with zipfile.ZipFile(BytesIO(zip_content), "r") as archive:
        bad_file = archive.testzip()
        if bad_file:
            raise RuntimeError(f"ZIP内の検証に失敗しました：{bad_file}")
        archived_names = set(archive.namelist())
    expected_names = {entry["path"] for entry in entries}
    if archived_names != expected_names:
        raise RuntimeError("ZIP内のファイル一覧が作成前と一致しません。")

    return {
        "content": zip_content,
        "filename": f"取引先カルテ全データ_{timestamp}.zip",
        "file_count": len(entries),
        "customer_count": int(customer_df["顧客名"].nunique()) if "顧客名" in customer_df.columns else 0,
    }


@st.fragment
def show_full_data_backup_download_button():
    """ダウンロード操作ではアプリ全体を再実行せず、ZIPをブラウザへ渡す。"""
    zip_content = st.session_state.get("full_data_backup_zip_bytes")
    zip_name = st.session_state.get("full_data_backup_zip_name")
    if not zip_content or not zip_name:
        return

    st.download_button(
        "⬇ ZIPをダウンロード",
        data=BytesIO(zip_content),
        file_name=zip_name,
        mime="application/zip",
        key="download_full_data_backup",
        use_container_width=True,
    )


def show_full_data_backup_page():
    """ボタンを押した時だけZIPを作り、Dropboxへ追加保存する。"""
    st.header("📦 全データバックアップ")
    st.write("現在のExcel保存ルールは変更せず、元データを変更せずにバックアップを作成します。")
    st.caption("作成したZIPはDropboxの専用フォルダへ自動保存し、端末にもダウンロードできます。")
    st.caption(f"Dropbox保存先：{FULL_DATA_BACKUP_DROPBOX_FOLDER}")

    bytes_key = "full_data_backup_zip_bytes"
    name_key = "full_data_backup_zip_name"
    summary_key = "full_data_backup_summary"
    dropbox_path_key = "full_data_backup_dropbox_path"
    dropbox_error_key = "full_data_backup_dropbox_error"

    if st.button(
        "📦 全データバックアップを作成",
        key="create_full_data_backup",
        type="primary",
        use_container_width=True,
    ):
        for key in (
            bytes_key,
            name_key,
            summary_key,
            dropbox_path_key,
            dropbox_error_key,
        ):
            st.session_state.pop(key, None)
        try:
            with st.spinner("全データを読み取り、ZIPを作成しています…"):
                result = create_full_data_backup_zip()
            st.session_state[bytes_key] = result["content"]
            st.session_state[name_key] = result["filename"]
            st.session_state[summary_key] = (
                f"{result['file_count']}ファイル・"
                f"{result['customer_count']}顧客のZIPを作成しました。"
            )

            try:
                with st.spinner("作成したZIPをDropboxへ保存しています…"):
                    dropbox_path = save_full_data_backup_to_dropbox(
                        result["filename"],
                        result["content"],
                    )
                st.session_state[dropbox_path_key] = dropbox_path
            except Exception as exc:
                st.session_state[dropbox_error_key] = str(exc)
        except Exception as exc:
            st.error(f"完全バックアップを作成できませんでした：{exc}")

    zip_content = st.session_state.get(bytes_key)
    zip_name = st.session_state.get(name_key)
    if zip_content and zip_name:
        summary = st.session_state.get(summary_key)
        if summary:
            st.success(summary)

        dropbox_path = st.session_state.get(dropbox_path_key)
        dropbox_error = st.session_state.get(dropbox_error_key)
        if dropbox_path:
            st.success(f"Dropboxへ保存しました：{dropbox_path}")
        elif dropbox_error:
            st.warning(
                "Dropboxへの保存に失敗しました。下のボタンから端末へダウンロードしてください。"
            )
            st.error(dropbox_error)

        show_full_data_backup_download_button()


# =========================
# メイン
# =========================
if "page" not in st.session_state:
    st.session_state["page"] = "home"

if "selected_customer" not in st.session_state:
    st.session_state["selected_customer"] = None
if "selected_partner_id" not in st.session_state:
    st.session_state["selected_partner_id"] = None
if "selected_partner_type" not in st.session_state:
    st.session_state["selected_partner_type"] = None

# URLに画面情報がある場合は、ブラウザの戻る・進むに合わせて復元する。
handle_customer_query_param()

current_page = st.session_state.get("page", "home")
customer_pages = {
    "customer_home", "customer_list", "customer", "region", "product", "calendar",
    "dispatch_table", "soluble_inventory", "water_it_test", "notes", "detail",
}
supplier_pages = {
    "supplier_home", "supplier_list", "supplier_search", "supplier_product",
    "supplier_register",
}
carrier_pages = {
    "carrier_home", "carrier_list", "carrier_search",
    "carrier_freight_compare", "carrier_register",
}

with st.sidebar:
    st.title(f"🚚 {APP_TITLE}")
    st.markdown("### メニュー")
    if st.button("🔄 更新", use_container_width=True):
        st.cache_data.clear()
        clear_trade_partner_immediate_data()
        st.rerun()
    st.markdown("---")

    st.markdown(render_page_link("🏠 トップ", page="home"), unsafe_allow_html=True)
    st.markdown(render_page_link("👥 顧客", page="customer_home"), unsafe_allow_html=True)
    st.markdown(render_page_link("🏢 仕入先", page="supplier_home"), unsafe_allow_html=True)
    st.markdown(render_page_link("🚚 運送会社", page="carrier_home"), unsafe_allow_html=True)
    st.markdown(render_page_link("📝 メモ帳", page="trade_notes"), unsafe_allow_html=True)

    if current_page in customer_pages:
        st.markdown("---")
        st.markdown("#### 顧客メニュー")
        st.markdown(render_page_link("👥 顧客名一覧", page="customer_list"), unsafe_allow_html=True)
        st.markdown(render_page_link("🔍 顧客検索", page="customer"), unsafe_allow_html=True)
        st.markdown(render_page_link("📍 地域検索", page="region"), unsafe_allow_html=True)
        st.markdown(render_page_link("🔎 商品検索", page="product"), unsafe_allow_html=True)
        st.markdown(render_page_link("🗓 在庫カレンダー", page="calendar"), unsafe_allow_html=True)
        st.markdown(render_page_link("🚚 配車表", page="dispatch_table"), unsafe_allow_html=True)
        st.markdown(render_page_link("🧪 ソリュブル在庫", page="soluble_inventory"), unsafe_allow_html=True)
        st.markdown(render_page_link("💧 WATER it接続", page="water_it_test"), unsafe_allow_html=True)
        st.markdown(render_page_link("📝 顧客メモ", page="notes"), unsafe_allow_html=True)
    elif current_page in supplier_pages or (
        current_page == "partner_detail" and st.session_state.get("selected_partner_type") == "supplier"
    ):
        st.markdown("---")
        st.markdown("#### 仕入先メニュー")
        st.markdown(render_page_link("📋 仕入先一覧", page="supplier_list"), unsafe_allow_html=True)
        st.markdown(render_page_link("🔍 仕入先検索", page="supplier_search"), unsafe_allow_html=True)
        st.markdown(render_page_link("📦 商品検索", page="supplier_product"), unsafe_allow_html=True)
    elif current_page in carrier_pages or (
        current_page == "partner_detail" and st.session_state.get("selected_partner_type") == "carrier"
    ):
        st.markdown("---")
        st.markdown("#### 運送会社メニュー")
        st.markdown(render_page_link("📋 運送会社一覧", page="carrier_list"), unsafe_allow_html=True)
        st.markdown(render_page_link("🔍 運送会社検索", page="carrier_search"), unsafe_allow_html=True)
        st.markdown(render_page_link("💰 運賃比較", page="carrier_freight_compare"), unsafe_allow_html=True)

    st.markdown("---")
    sidebar_login_alert_count = get_unacknowledged_login_browser_count()
    sidebar_login_label = "🔐 ログイン履歴"
    if sidebar_login_alert_count:
        sidebar_login_label += f"（未確認 {sidebar_login_alert_count}）"
    st.markdown(
        render_page_link(sidebar_login_label, page="login_history"),
        unsafe_allow_html=True,
    )
    st.markdown(
        render_page_link("📦 全データバックアップ", page="data_backup"),
        unsafe_allow_html=True,
    )


col_title, col_logout = st.columns([3, 1])

with col_title:
    st.title(f"🚚 {APP_TITLE}")
    st.caption("顧客・仕入先・運送会社の情報を確認・編集します。")

with col_logout:
    st.write("")
    if st.button("ログアウト"):
        record_logout_event(microsoft_claims, active_login_payload)
        clear_application_login_state(revoke_current=True)
        st.session_state.page = "home"
        st.session_state.selected_customer = None
        st.session_state.selected_partner_id = None
        st.session_state.selected_partner_type = None
        clear_onedrive_auth_state()
        set_query_params_safely({"page": "home", "logout": "1"})
        st.logout()
        st.stop()

show_microsoft_client_secret_expiry_notice()

login_audit_warning = st.session_state.pop("login_audit_warning", None)
if login_audit_warning:
    st.warning(
        "ログイン自体は成功しましたが、ログイン履歴を保存できませんでした。"
        f" 詳細：{login_audit_warning}"
    )

new_browser_notice = bool(st.session_state.pop("new_browser_login_notice", False))
login_security_alert_count = get_unacknowledged_login_browser_count()
if new_browser_notice:
    st.warning("このブラウザからの初回ログインを記録しました。ログイン履歴で確認してください。")
if login_security_alert_count:
    st.markdown(
        render_page_link(
            f"🔐 未確認の新しいブラウザが {login_security_alert_count} 件あります",
            page="login_history",
        ),
        unsafe_allow_html=True,
    )

history_warning = st.session_state.pop("change_history_warning", None)
if history_warning:
    st.warning(history_warning)

global_delete_restore_requested = bool(
    st.session_state.pop(GLOBAL_DELETE_SCROLL_RESTORE_KEY, False)
)

# 専用の顧客検索ページ以外では、すべての画面上部に共通検索欄を表示する。
# 専用ページには同じ検索機能があるため、入力欄が二重にならないよう除外する。
if current_page != "customer":
    render_global_customer_search()

# 各機能ページから、区分メニューを経由せずトップへ直接戻れるようにする。
# 顧客・仕入先・運送会社の各ホームと取引先メモは、従来から同じリンクを
# 表示しているため、二重表示にならないようここでは除外する。
pages_with_existing_top_link = {
    "customer_home",
    "supplier_home",
    "carrier_home",
    "trade_notes",
    "login_history",
}
if current_page != "home" and current_page not in pages_with_existing_top_link:
    show_top_home_link()

try:
    if st.session_state["page"] == "home":
        show_top_home()

    elif st.session_state["page"] == "customer_home":
        show_home_menu()

    elif st.session_state["page"] == "customer":
        show_customer_search(show_home_link=True)

    elif st.session_state["page"] == "customer_list":
        show_customer_directory()

    elif st.session_state["page"] == "region":
        show_region_search()

    elif st.session_state["page"] == "product":
        show_product_search()

    elif st.session_state["page"] == "calendar":
        df = load_data()
        show_dispatch_calendar(df)

    elif st.session_state["page"] == "dispatch_table":
        show_dispatch_board()

    elif st.session_state["page"] == "soluble_inventory":
        show_soluble_inventory_page()

    elif st.session_state["page"] == "water_it_test":
        show_water_it_test_page()

    elif st.session_state["page"] == "notes":
        show_notes_page(None)

    elif st.session_state["page"] == "trade_notes":
        show_trade_notes_page()

    elif st.session_state["page"] == "hotel_information":
        show_hotel_information_page()

    elif st.session_state["page"] == "change_history":
        show_change_history_page()

    elif st.session_state["page"] == "attachment_search":
        show_attachment_search_page()

    elif st.session_state["page"] == "estimates":
        show_estimates_page()

    elif st.session_state["page"] == "data_backup":
        show_full_data_backup_page()

    elif st.session_state["page"] == "login_history":
        show_login_history_page()

    elif st.session_state["page"] == "detail":
        selected = st.session_state.get("selected_customer")
        if selected:
            immediate_df = st.session_state.pop("customer_excel_immediate_df", None)
            if isinstance(immediate_df, pd.DataFrame) and not immediate_df.empty:
                df = immediate_df
            else:
                df = load_data()
            show_customer_detail(df, selected)
        else:
            set_page("customer_home")
            st.rerun()

    elif st.session_state["page"] == "supplier_home":
        show_trade_partner_home("supplier")
    elif st.session_state["page"] == "supplier_list":
        show_trade_partner_directory("supplier")
    elif st.session_state["page"] == "supplier_search":
        show_trade_partner_search("supplier")
    elif st.session_state["page"] == "supplier_product":
        show_supplier_product_search()
    elif st.session_state["page"] == "supplier_register":
        show_trade_partner_register("supplier")

    elif st.session_state["page"] == "carrier_home":
        show_trade_partner_home("carrier")
    elif st.session_state["page"] == "carrier_list":
        show_trade_partner_directory("carrier")
    elif st.session_state["page"] == "carrier_search":
        show_trade_partner_search("carrier")
    elif st.session_state["page"] == "carrier_freight_compare":
        show_carrier_freight_compare()
    elif st.session_state["page"] == "carrier_register":
        show_trade_partner_register("carrier")

    elif st.session_state["page"] == "partner_detail":
        partner_id = st.session_state.get("selected_partner_id")
        partner_type = st.session_state.get("selected_partner_type")
        if partner_id and partner_type in {"supplier", "carrier"}:
            show_trade_partner_detail(partner_type, partner_id)
        else:
            set_page("home")
            st.rerun()
except Exception as e:
    st.error("画面表示中にエラーが発生しました。")
    st.write("原因確認のため、エラー内容を表示しています。")
    st.exception(e)
    st.stop()

render_global_delete_scroll_keeper(restore=global_delete_restore_requested)

st.caption(
    "※ 顧客情報は配車予定 次郎.xlsm、仕入先・運送会社は取引先カルテ.xlsxを読み込んで表示しています。"
)

