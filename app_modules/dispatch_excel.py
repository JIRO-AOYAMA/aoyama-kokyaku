"""配車表1.xlsmの月別シートを読み込み、一覧用データへ整形する。"""

import re
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook


DEFAULT_MONTH_SHEETS = [f"{month}月" for month in range(1, 13)]
DEFAULT_REQUIRED_COLUMNS = [
    "発注番号",
    "引取日",
    "引取先",
    "商品名",
    "数量",
    "運送会社",
    "納品先",
    "着日",
]


def normalize_dispatch_text(value):
    """配車表の表示・絞り込み用に前後空白と連続空白をそろえる。"""
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip()


def read_dispatch_month_sheets(
    excel_source,
    month_sheets=None,
    required_columns=None,
):
    """配車表1.xlsmの1月～12月シートからA～H列だけを結合する。"""
    month_sheets = list(month_sheets or DEFAULT_MONTH_SHEETS)
    required_columns = list(required_columns or DEFAULT_REQUIRED_COLUMNS)

    if isinstance(excel_source, BytesIO):
        source = BytesIO(excel_source.getvalue())
    else:
        source = excel_source

    workbook = load_workbook(source, read_only=True, data_only=True)
    rows = []
    try:
        missing_sheets = [name for name in month_sheets if name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError("月別シートが見つかりません：" + "、".join(missing_sheets))

        for sheet_name in month_sheets:
            ws = workbook[sheet_name]
            headers = [normalize_dispatch_text(ws.cell(1, column).value) for column in range(1, 9)]
            if headers != required_columns:
                raise ValueError(
                    f"{sheet_name}のA～H列の見出しが想定と異なります。\n"
                    f"読み取った見出し：{' / '.join(headers)}"
                )

            for values in ws.iter_rows(min_row=2, max_col=8, values_only=True):
                if not any(value is not None and normalize_dispatch_text(value) for value in values):
                    continue

                record = dict(zip(required_columns, values))
                record["参照シート"] = sheet_name
                rows.append(record)
    finally:
        workbook.close()

    df = pd.DataFrame(rows, columns=required_columns + ["参照シート"])
    if df.empty:
        return df

    for column in ["引取先", "商品名", "数量", "運送会社", "納品先"]:
        df[column] = df[column].map(normalize_dispatch_text)

    pickup_dates = pd.to_datetime(df["引取日"], errors="coerce")
    arrival_dates = pd.to_datetime(df["着日"], errors="coerce")
    df["_引取日"] = pickup_dates.map(lambda value: value.date() if pd.notna(value) else None)
    df["_着日"] = arrival_dates.map(lambda value: value.date() if pd.notna(value) else None)
    return df
