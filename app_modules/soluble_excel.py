"""ソリュブル在庫Excelの読み取り・値変換・安全保存。

Streamlit画面やDropbox接続を含めず、Excelファイルのbytesだけを扱う。
"""

import math
import re
from datetime import date, datetime, timedelta
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


SOLUBLE_SHEET_NAME = "ソリュブル"
SOLUBLE_LOCATIONS = {
    "ノベルズ": {"usage": 3, "delivery": 4, "inventory": 5},
    "コスモアグリ": {"usage": 6, "delivery": 7, "inventory": 8},
}


def soluble_cell_is_manual(cell):
    """Excelで黄色に塗られたセルを手入力値として扱う。"""
    if cell.fill.fill_type != "solid":
        return False
    color = cell.fill.fgColor
    if color.type == "rgb":
        return str(color.rgb or "").upper().endswith("FFFF00")
    return False


def soluble_formula_value(formula_ws, value_ws, row, column, memo=None, visiting=None):
    """計算キャッシュが空でも、対象表の単純な加減式を表示できるようにする。"""
    memo = memo if memo is not None else {}
    visiting = visiting if visiting is not None else set()
    key = (row, column)
    if key in memo:
        return memo[key]
    if key in visiting:
        return None

    raw = formula_ws.cell(row, column).value
    cached = value_ws.cell(row, column).value
    if not (isinstance(raw, str) and raw.startswith("=")):
        memo[key] = raw
        return raw
    if cached is not None:
        memo[key] = cached
        return cached

    expression = raw[1:].replace(" ", "").replace("$", "").upper()
    tokens = re.findall(r"[A-Z]+\d+|\d+(?:\.\d+)?|[+-]", expression)
    if not tokens or "".join(tokens) != expression:
        return None

    visiting.add(key)
    try:
        def token_value(token):
            match = re.fullmatch(r"([A-Z]+)(\d+)", token)
            if match:
                letters, target_row = match.groups()
                target_column = 0
                for letter in letters:
                    target_column = target_column * 26 + (ord(letter) - 64)
                return soluble_formula_value(
                    formula_ws,
                    value_ws,
                    int(target_row),
                    target_column,
                    memo,
                    visiting,
                )
            return float(token)

        result = token_value(tokens[0])
        index = 1
        while index < len(tokens):
            operator = tokens[index]
            right = token_value(tokens[index + 1])
            if result is None:
                result = 0
            if right is None:
                right = 0
            if isinstance(result, (date, datetime)) and isinstance(right, (int, float)):
                result = result + timedelta(days=right if operator == "+" else -right)
            else:
                result = result + right if operator == "+" else result - right
            index += 2
        memo[key] = result
        return result
    except Exception:
        return None
    finally:
        visiting.discard(key)


def soluble_date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def soluble_number_label(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "—"
    if isinstance(value, (int, float)):
        number = float(value)
        if number.is_integer():
            return f"{int(number):,}"
        return f"{number:,.2f}".rstrip("0").rstrip(".")
    return str(value)


def soluble_input_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value)


def parse_soluble_number(text, label):
    cleaned = str(text or "").strip().replace(",", "").replace("，", "")
    if not cleaned:
        return None
    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"{label}は数字で入力してください。") from exc
    return int(number) if number.is_integer() else number


def same_soluble_value(left, right):
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=0, abs_tol=1e-9)
    return left == right


def read_soluble_rows(
    content,
    sheet_name=SOLUBLE_SHEET_NAME,
    locations=None,
):
    """ソリュブルシートの日付・使用量・納品・在庫を読み取る。"""
    locations = locations or SOLUBLE_LOCATIONS
    formula_wb = load_workbook(BytesIO(content), data_only=False, read_only=False)
    value_wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    try:
        if sheet_name not in formula_wb.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        formula_ws = formula_wb[sheet_name]
        value_ws = value_wb[sheet_name]
        memo = {}
        rows = []
        for row_number in range(11, formula_ws.max_row + 1):
            day_value = soluble_formula_value(formula_ws, value_ws, row_number, 2, memo)
            day = soluble_date_value(day_value)
            if day is None:
                continue
            record = {"row": row_number, "date": day}
            for location, columns in locations.items():
                for field, column in columns.items():
                    record[f"{location}_{field}"] = soluble_formula_value(
                        formula_ws,
                        value_ws,
                        row_number,
                        column,
                        memo,
                    )
                    record[f"{location}_{field}_manual"] = soluble_cell_is_manual(
                        formula_ws.cell(row_number, column)
                    )
                    raw_value = formula_ws.cell(row_number, column).value
                    record[f"{location}_{field}_formula"] = (
                        raw_value
                        if isinstance(raw_value, str) and raw_value.startswith("=")
                        else ""
                    )
            rows.append(record)
        return rows
    finally:
        formula_wb.close()
        value_wb.close()


def build_soluble_updated_workbook(
    content,
    row_number,
    location,
    updates,
    sheet_name=SOLUBLE_SHEET_NAME,
    locations=None,
):
    """Excel標準形式のまま対象セルだけを更新し、保存結果を検証する。"""
    locations = locations or SOLUBLE_LOCATIONS
    if location not in locations:
        raise ValueError("対象の会社が正しくありません。")
    if row_number < 11:
        raise ValueError("更新する行が正しくありません。")
    if not updates:
        raise ValueError("変更された項目がありません。")

    workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    original_sheets = list(workbook.sheetnames)
    changed = []
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    clear_fill = PatternFill(fill_type=None)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("ソリュブルシートが見つかりません。")
        ws = workbook[sheet_name]
        if row_number > ws.max_row:
            raise ValueError("更新する日付行が見つかりません。")
        columns = locations[location]

        for field, requested_value in updates.items():
            if field not in columns:
                raise ValueError("更新項目が正しくありません。")
            if requested_value == "__AUTO_INVENTORY__":
                if field != "inventory" or row_number <= 11:
                    raise ValueError("この日は在庫を自動計算にできません。")
                inventory_letter = ws.cell(row_number, columns["inventory"]).column_letter
                usage_letter = ws.cell(row_number, columns["usage"]).column_letter
                delivery_letter = ws.cell(row_number, columns["delivery"]).column_letter
                new_value = (
                    f"={inventory_letter}{row_number - 1}"
                    f"-{usage_letter}{row_number}"
                    f"+{delivery_letter}{row_number}"
                )
                manual = False
            else:
                new_value = requested_value
                manual = True

            cell = ws.cell(row_number, columns[field])
            cell.value = new_value
            cell.fill = yellow_fill if manual else clear_fill
            changed.append((cell.coordinate, new_value, manual))

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    saved_content = output.getvalue()
    verified = load_workbook(BytesIO(saved_content), data_only=False, read_only=False)
    try:
        if list(verified.sheetnames) != original_sheets:
            raise ValueError("保存後にシート構成が変わったため、更新を中止しました。")
        ws = verified[sheet_name]
        for coordinate, expected, expected_manual in changed:
            cell = ws[coordinate]
            if cell.value != expected or soluble_cell_is_manual(cell) != expected_manual:
                raise ValueError(f"保存確認で{sheet_name}!{coordinate}が一致しません。")
    finally:
        verified.close()
    return saved_content, changed
